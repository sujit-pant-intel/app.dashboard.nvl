"""yield_trend.py — Merged trend charts and run comparison.

Consolidates: compare_runs, compareTP, trend_chart, trend_chart_frame

Public API (unchanged):
  CompareFrame  (compareTP.py)
  TrendChartFrame (trend_chart_frame.py)
"""
from __future__ import annotations

# ════════════════════════════════════════════════════════════════
# (formerly compare_runs.py)
# ════════════════════════════════════════════════════════════════
#!/usr/bin/env python3
"""
compare_runs.py
---------------
Reads Dashboard.html, resolves each run's *_out.xlsx, and generates a
standalone comparison HTML report.

Usage:
    python compare_runs.py Dashboard.html
    python compare_runs.py Dashboard.html --out my_comparison.html
    python compare_runs.py Dashboard.html --ref "NCXSDJXP0H51M202611-1"
"""

import sys
import os
import re
import argparse


def _find_wafer_tools(start: 'Path | None' = None, max_levels: int = 6) -> str:
    """Walk up from this file's directory looking for shared/utilities/wafer_tools
    (mirrors scan-dashboard.py / generate_dashboard.py's _find_wafer_tools())."""
    cur = (start or Path(__file__).resolve().parent)
    for _ in range(max_levels):
        for rel in ("shared/utilities/wafer_tools", "utilities/wafer_tools"):
            cand = cur / rel
            if (cand / "wafer_map").is_dir():
                return str(cand)
        parent = cur.parent
        if parent == cur:
            break
        cur = parent
    return ""


def _inject_wafermap_js(html: str) -> str:
    """Inject the shared SVG wafer-map renderer (wmRender) before the main
    dashboard <script> block so it's defined before any JS calls it."""
    try:
        wafer_tools = _find_wafer_tools()
        if not wafer_tools:
            return html
        if wafer_tools not in sys.path:
            sys.path.insert(0, wafer_tools)
        from wafer_map import WAFERMAP_JS
        marker = '\n<script>\n// \u2550'
        if marker not in html:
            marker = '\n<script>\nconst DATA ='
        if marker in html:
            return html.replace(marker, '\n' + WAFERMAP_JS + marker, 1)
        return html.replace('</body>', WAFERMAP_JS + '\n</body>', 1)
    except Exception as e:
        print(f'[yield_trend] WARN: WAFERMAP_JS not injected: {e}')
        return html
import io
import base64
from pathlib import Path


def _wm_inject(html: str) -> str:
    _wm = (
        '<div id="_wm_div" style="position:fixed;top:8px;right:12px;font-size:10px;'
        'font-weight:600;pointer-events:none;z-index:99999;'
        'font-family:Arial,sans-serif;user-select:none;letter-spacing:0.04em;'
        'padding:2px 6px;border-radius:3px;background:transparent;">'
        'Pant, Sujit N \u2014 GEMS FTE</div>'
        '<script>(function(){'
        'function _wm_color(){'
        'var d=document.getElementById("_wm_div");if(!d)return;'
        'var bg=window.getComputedStyle(document.body).backgroundColor;'
        'var m=bg.match(/\\d+/g);'
        'if(m&&m.length>=3){'
        'var r=+m[0],g=+m[1],b=+m[2];'
        'var lum=0.299*r+0.587*g+0.114*b;'
        'd.style.color=lum<128?"rgba(255,255,255,0.9)":"rgba(20,20,20,0.75)";'
        '}else{d.style.color="rgba(255,255,255,0.9)";}'
        '}'
        'if(document.readyState==="loading")'
        '{document.addEventListener("DOMContentLoaded",_wm_color);}'
        'else{_wm_color();}'
        '})();</script>'
    )
    import re as _re_wm
    if '</body>' not in html:
        return html
    html = _re_wm.sub(
        r'<div[^>]*id=["\']_wm_div["\'][^>]*>[\s\S]*?</div>\s*<script[^>]*>[\s\S]*?</script>',
        '', html)
    html = _re_wm.sub(r'<div[^>]*>[^<]*GEMS FTE[^<]*</div>', '', html)
    return html.replace('</body>', _wm + '\n</body>', 1)

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np
    HAVE_MPL = True
except ImportError:
    HAVE_MPL = False

# ---------------------------------------------------------------------------
# 1. Parse Dashboard.html → run records
# ---------------------------------------------------------------------------

def parse_dashboard(dash_path: Path):
    """Return list of dicts: {stem, name, ts, index_href}.
    Only returns blocks from the Yield section (YIELD_START/END or legacy RUNS_START/END)."""
    content = dash_path.read_text(encoding='utf-8')
    # Extract only the Yield section content
    yield_html = ''
    for start_pat, end_pat in [
        (r'<!--\s*YIELD_START\s*-->', r'<!--\s*YIELD_END\s*-->'),
        (r'<!--\s*RUNS_START\s*-->',  r'<!--\s*RUNS_END\s*-->'),   # legacy
    ]:
        m = re.search(start_pat + r'(.*?)' + end_pat, content, re.S)
        if m:
            yield_html = m.group(1)
            break
    # Fall back to full content only if no sentinels found at all
    search_html = yield_html if yield_html.strip() else content
    runs = []
    # Each block: <div class="run-block" data-stem="...">...</div></div>
    block_re = re.compile(
        r'<div class="run-block" data-stem="([^"]+)">([\s\S]*?)</div>\s*</div>',
        re.MULTILINE
    )
    for m in block_re.finditer(search_html):
        stem = m.group(1)
        body = m.group(2)
        # Name and timestamp from run-header
        hdr = re.search(
            r'<span class="arrow">[^<]*</span>\s*([^<]+)<span class="ts">\s*-\s*([^<]*)</span>',
            body
        )
        name = hdr.group(1).strip() if hdr else stem
        ts   = hdr.group(2).strip() if hdr else ''
        # Yield Report href (report-link)
        link_m = re.search(r'class="run-link report-link"[^>]*href="([^"]+)"', body)
        index_href = link_m.group(1) if link_m else None
        runs.append({'stem': stem, 'name': name, 'ts': ts, 'index_href': index_href})
    return runs


# ---------------------------------------------------------------------------
# 2. Resolve *_out.xlsx from an index.html href
# ---------------------------------------------------------------------------

def find_xlsx(dash_dir: Path, index_href: str):
    """Given Dashboard.html dir and the relative/absolute href to index.html,
    return Path to *_out.xlsx in the same folder, or None."""
    if not index_href:
        return None
    # Strip file:// scheme if present
    href = re.sub(r'^file:///', '', index_href).replace('/', os.sep)
    if os.path.isabs(href):
        idx_path = Path(href)
    else:
        idx_path = dash_dir / href
    out_folder = idx_path.parent
    if not out_folder.exists():
        return None
    candidates = sorted(out_folder.glob('*_out.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


# ---------------------------------------------------------------------------
# 2b. Resolve *_BinDistribution.html and parse RDND + Bin Fail tables
# ---------------------------------------------------------------------------

def find_bin_html(output_dir: Path):
    """Return Path to *_BinDistribution.html in output_dir, or None."""
    candidates = sorted(output_dir.glob('*_BinDistribution.html'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def find_group_medians(output_dir: Path):
    """Return Path to Group_Medians.csv in output_dir, or None."""
    candidates = sorted(output_dir.glob('Group_Medians.csv'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def find_raw_csv(output_dir: Path):
    """Return Path to the raw CSV in output_dir's parent (the source CSV), or None."""
    parent = output_dir.parent
    if not parent.exists():
        return None
    # Match CSVs that look like the source data (not bindef or generated)
    candidates = [p for p in sorted(parent.glob('*.csv'))
                  if 'bindef' not in p.name.lower()
                  and '_targets_' not in p.name.lower()]
    # Prefer the one with 'reticle' in the name (most complete), else newest
    reticle = [p for p in candidates if 'reticle' in p.name.lower()]
    if reticle:
        return reticle[0]
    return candidates[0] if candidates else None


def extract_upm_from_csv(csv_path: Path, config_json: str = None):
    """Extract UPM distribution data from raw CSV using config JSON analyses.
    Returns list of dicts compatible with upm_data format, plus a detailed
    list of per-column medians."""
    import pandas as pd
    import json

    if not csv_path or not csv_path.exists():
        return None, None

    # Load config to get analyses + base
    analyses = []
    base_val = None
    upm_prefix = 'UPM_0107'
    filt_value = upm_prefix
    filt_method = 'starts_with'
    if config_json:
        try:
            cfg = json.loads(Path(config_json).read_text(encoding='utf-8'))
            analyses = cfg.get('analyses', [])
            for anl in analyses:
                if anl.get('type') == 'distribution':
                    filt = anl.get('filter', {}).get('match', {})
                    filt_method = filt.get('method', 'starts_with')
                    filt_value = filt.get('value', upm_prefix)
                    agg = anl.get('aggregation', {})
                    if agg.get('mode') == 'percentage':
                        bc = agg.get('base', {})
                        if bc.get('type') == 'fixed':
                            base_val = bc.get('value')
                    break
        except Exception:
            pass

    try:
        df = pd.read_csv(str(csv_path), dtype=object)
    except Exception:
        return None, None

    import re as _re_upm
    import fnmatch as _fnmatch_upm
    _has_wc = ('*' in filt_value or '?' in filt_value)
    if _has_wc:
        upm_cols = [c for c in df.columns if _fnmatch_upm.fnmatch(c, filt_value)]
    elif filt_method == 'regex':
        upm_cols = [c for c in df.columns if _re_upm.search(filt_value, c)]
    elif filt_method == 'contains':
        upm_cols = [c for c in df.columns if filt_value in c]
    elif filt_method == 'starts_with':
        upm_cols = [c for c in df.columns if c.startswith(filt_value)]
    else:
        upm_cols = [c for c in df.columns if c.startswith(filt_value)]
    if not upm_cols:
        return None, None

    if base_val is None:
        base_val = 9154  # default

    # Per-column detail
    col_details = []
    overall_pct = None
    for col in upm_cols:
        vals = pd.to_numeric(df[col], errors='coerce').dropna()
        if vals.empty:
            continue
        med = float(vals.median())
        pct = (med / base_val) * 100
        short = col[len(upm_prefix):].strip('_') if col.startswith(upm_prefix) else col
        col_details.append({
            'col': col, 'short': short, 'median': med,
            'mean': float(vals.mean()), 'count': len(vals), 'pct': pct
        })
        # Use 0950 column as the "main" UPM % if available
        if '0950' in col:
            overall_pct = pct

    if not col_details:
        return None, None

    # If no 0950 column, use the first
    if overall_pct is None:
        overall_pct = col_details[0]['pct']

    # Build legacy-compatible upm_data (single row with upm_pct)
    upm_data = [{'test': 'UPM', 'n_rows': col_details[0]['count'],
                 'sicc_actual': None, 'sicc_target': None,
                 'multiple': None, 'upm_pct': overall_pct}]

    return upm_data, col_details


def parse_group_medians(csv_path: Path) -> list[dict]:
    """Return list of {test, n_rows, sicc_actual, sicc_target, multiple, upm_pct}."""
    rows = []
    try:
        text = csv_path.read_text(encoding='utf-8-sig')
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            return rows
        header = [h.strip() for h in lines[0].split(',')]
        for line in lines[1:]:
            parts = [p.strip() for p in line.split(',')]
            if len(parts) < 2:
                continue
            def _f(i):
                try: return float(parts[i]) if i < len(parts) and parts[i] else None
                except: return None
            rows.append({
                'test':        parts[0] if parts else '',
                'n_rows':      _f(1),
                'sicc_actual': _f(2),
                'sicc_target': _f(3),
                'multiple':    _f(4),
                'upm_pct':     _f(5),
            })
    except Exception as e:
        print(f'Warning: could not read Group_Medians.csv: {e}')
    return rows


def find_cdyn_medians(output_dir: Path):
    """Return Path to cdyn_medians.csv in output_dir, or None."""
    candidates = sorted(output_dir.glob('cdyn_medians.csv'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def parse_cdyn_medians(csv_path: Path) -> list[dict]:
    """Return list of {test, type, actual, expected, ratio}."""
    rows = []
    try:
        text = csv_path.read_text(encoding='utf-8-sig')
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            return rows
        for line in lines[1:]:
            parts = [p.strip() for p in line.split(',')]
            if len(parts) < 2:
                continue
            def _f(i):
                try: return float(parts[i]) if i < len(parts) and parts[i] else None
                except: return None
            rows.append({
                'test':     parts[0] if parts else '',
                'type':     parts[1] if len(parts) > 1 else '',
                'actual':   _f(2),
                'expected': _f(3),
                'ratio':    _f(4),
            })
    except Exception as e:
        print(f'Warning: could not read cdyn_medians.csv: {e}')
    return rows


def _strip_tags(s: str) -> str:
    return re.sub(r'<[^>]+>', '', s).strip()


def parse_bin_html(bin_html_path: Path):
    """Return dict:
        yield_rows:    [{bin, fail_bucket, yield_pct, expected_pct}, ...]
        bin_fail_rows: [{ibin, fail_bucket, fail_pct, fail_count}, ...]
        func_bin_rows: [{ibin, fbin, fail_bucket, fail_pct, fail_count}, ...]
    """
    content = bin_html_path.read_text(encoding='utf-8')

    # --- RDND yield table ---
    yield_rows = []
    yt_m = re.search(r'<table class="yield-table">(.*?)</table>', content, re.DOTALL)
    if yt_m:
        for tr in re.findall(r'<tr>(.*?)</tr>', yt_m.group(1), re.DOTALL):
            tds = re.findall(r'<td(?:[^>]*)>(.*?)</td>', tr, re.DOTALL)
            if len(tds) < 3:
                continue
            # Column order: BIN(0), FAIL BUCKET(1), YIELD(2), EXPECTED(3)
            bin_name     = _strip_tags(tds[0])
            fail_bucket  = _strip_tags(tds[1]) if len(tds) > 1 else ''
            yield_str    = _strip_tags(tds[2]).rstrip('%') if len(tds) > 2 else ''
            expected_str = _strip_tags(tds[3]).rstrip('%') if len(tds) > 3 else ''
            try:    yield_pct    = float(yield_str)
            except: yield_pct    = None
            try:    expected_pct = float(expected_str)
            except: expected_pct = None
            if bin_name:
                yield_rows.append({'bin': bin_name, 'fail_bucket': fail_bucket,
                                   'yield_pct': yield_pct, 'expected_pct': expected_pct})

    # --- Bin Fail Summary table (pareto-tbl with Interface Bin header) ---
    bin_fail_rows = []
    func_bin_rows = []
    bin_summary_rows = []  # new 6-col format: ibin, cat, desc, total, fail_count, fail_pct
    for tbl in re.findall(r'<table class="pareto-tbl"[^>]*>(.*?)</table>', content, re.DOTALL):
        if 'Interface Bin' not in tbl:
            continue
        has_fbin    = 'Functional Bin' in tbl
        has_cat_desc = 'Category' in tbl and 'Description' in tbl
        for tr in re.findall(r'<tr[^>]*>(.*?)</tr>', tbl, re.DOTALL):
            tds = re.findall(r'<td(?:[^>]*)>(.*?)</td>', tr, re.DOTALL)
            if has_cat_desc:
                # new format: Interface Bin(0), Category(1), Description(2),
                #              Total Count(3), Fail Count(4), Yield/Fail %(5)
                if len(tds) < 6:
                    continue
                ibin         = _strip_tags(tds[0]).strip().rstrip('\u26a0').strip()
                cat          = _strip_tags(tds[1])
                desc         = _strip_tags(tds[2])
                fail_cnt_str = _strip_tags(tds[4]).replace(',', '')
                fail_pct_str = _strip_tags(tds[5]).rstrip('%')
                try:    fail_pct   = float(fail_pct_str)
                except: fail_pct   = None
                try:    fail_count = int(fail_cnt_str)
                except: fail_count = None
                if ibin and fail_pct is not None:
                    bin_summary_rows.append({'ibin': ibin, 'cat': cat, 'desc': desc,
                                             'fail_pct': fail_pct, 'fail_count': fail_count})
            elif has_fbin:
                # cols: Interface Bin(0), Functional Bin(1), Fail Bucket(2),
                #       Description(3), Total Count(4), Fail Count(5), Fail Count %(6)
                if len(tds) < 7:
                    continue
                ibin         = _strip_tags(tds[0])
                fbin         = _strip_tags(tds[1])
                fail_bucket  = _strip_tags(tds[2])
                fail_pct_str = _strip_tags(tds[6]).rstrip('%')
                fail_cnt_str = _strip_tags(tds[5]).replace(',', '')
                try:    fail_pct   = float(fail_pct_str)
                except: fail_pct   = None
                try:    fail_count = int(fail_cnt_str)
                except: fail_count = None
                if ibin and fail_pct is not None:
                    func_bin_rows.append({'ibin': ibin, 'fbin': fbin,
                                          'fail_bucket': fail_bucket,
                                          'fail_pct': fail_pct, 'fail_count': fail_count})
            else:
                # old format: Interface Bin(0), Fail Bucket(1), Total Count(2), Fail Count(3), Fail %(4)
                if len(tds) < 5:
                    continue
                ibin         = _strip_tags(tds[0])
                fail_bucket  = _strip_tags(tds[1])
                fail_pct_str = _strip_tags(tds[4]).rstrip('%')
                fail_cnt_str = _strip_tags(tds[3]).replace(',', '')
                try:    fail_pct   = float(fail_pct_str)
                except: fail_pct   = None
                try:    fail_count = int(fail_cnt_str)
                except: fail_count = None
                if ibin and fail_pct is not None:
                    bin_fail_rows.append({'ibin': ibin, 'fail_bucket': fail_bucket,
                                          'fail_pct': fail_pct, 'fail_count': fail_count})

    # Fallback for newer BinDistribution HTML where rows are rendered dynamically
    # from JS arrays/objects (DATA / BFS_DATA / FP_DATA) instead of static <tr> rows.
    if not bin_summary_rows or not func_bin_rows:
        try:
            import json as _json

            def _extract_js_array(var_name: str):
                m = re.search(
                    rf'var\s+{re.escape(var_name)}\s*=\s*(\[[\s\S]*?\]);',
                    content,
                    re.DOTALL,
                )
                if not m:
                    return []
                try:
                    return _json.loads(m.group(1))
                except Exception:
                    return []

            if not yield_rows:
                m_data = re.search(r'var\s+DATA\s*=\s*({[\s\S]*?});', content, re.DOTALL)
                if m_data:
                    try:
                        data_obj = _json.loads(m_data.group(1))
                    except Exception:
                        data_obj = {}
                    rows = data_obj.get('rows', []) if isinstance(data_obj, dict) else []
                    ydefs = data_obj.get('yieldDefs', []) if isinstance(data_obj, dict) else []
                    total_die = data_obj.get('total', 0) if isinstance(data_obj, dict) else 0
                    if not total_die:
                        total_die = sum(
                            int(r.get('total', 0) or 0)
                            for r in rows if isinstance(r, dict)
                        )
                    if total_die and ydefs:
                        for yd in ydefs:
                            if not isinstance(yd, dict):
                                continue
                            bins_list = [str(b) for b in (yd.get('bins_list') or [])]
                            cnt = 0
                            for rr in rows:
                                if not isinstance(rr, dict):
                                    continue
                                bc = rr.get('binCounts') or {}
                                if not isinstance(bc, dict):
                                    continue
                                cnt += sum(int(bc.get(bk, 0) or 0) for bk in bins_list)
                            try:
                                expected_pct = float(yd.get('expected')) if yd.get('expected') not in (None, '') else None
                            except Exception:
                                expected_pct = None
                            yield_rows.append({
                                'bin': str(yd.get('bins', '') or ''),
                                'fail_bucket': str(yd.get('bucket', '') or ''),
                                'yield_pct': (cnt / float(total_die) * 100.0) if total_die else None,
                                'expected_pct': expected_pct,
                            })

            if not bin_summary_rows:
                bfs_rows = _extract_js_array('BFS_DATA')
                for r in bfs_rows:
                    ibin_raw = str(r.get('bin', '')).strip()
                    # Stored values may contain warning symbol (e.g., "15⚠").
                    ibin = ibin_raw.rstrip('\u26a0').strip()
                    try:
                        fail_pct = float(r.get('pct'))
                    except Exception:
                        fail_pct = None
                    try:
                        fail_count = int(r.get('count'))
                    except Exception:
                        fail_count = None
                    if ibin and fail_pct is not None:
                        bin_summary_rows.append({
                            'ibin': ibin,
                            'cat': str(r.get('cat', '') or ''),
                            'desc': str(r.get('desc', '') or ''),
                            'fail_pct': fail_pct,
                            'fail_count': fail_count,
                        })

            if not func_bin_rows:
                fp_rows = _extract_js_array('FP_DATA')
                for r in fp_rows:
                    try:
                        fbin = str(int(float(r.get('fb'))))
                    except Exception:
                        fbin = str(r.get('fb', '') or '')
                    try:
                        fail_pct = float(r.get('pct'))
                    except Exception:
                        fail_pct = None
                    try:
                        fail_count = int(r.get('count'))
                    except Exception:
                        fail_count = None
                    if fbin and fail_pct is not None:
                        # Keep key name "ibin" because compare chart/table logic expects it.
                        func_bin_rows.append({
                            'ibin': fbin,
                            'fbin': fbin,
                            'fail_bucket': str(r.get('bkt', '') or ''),
                            'fail_pct': fail_pct,
                            'fail_count': fail_count,
                        })
        except Exception:
            pass

    return {'yield_rows': yield_rows, 'bin_fail_rows': bin_fail_rows,
            'func_bin_rows': func_bin_rows, 'bin_summary_rows': bin_summary_rows}


# ---------------------------------------------------------------------------
# 3. Read xlsx → structured data
# ---------------------------------------------------------------------------

def read_xlsx(xlsx_path: Path):
    """Return dict with keys:
        num_die, col_headers, groups, totals
        groups: list of {name, rows: [(label, vals...)], sum_vals: (vals...)}
        totals: {label, vals} for 'Yield SUM' row
    """
    if not HAVE_OPENPYXL:
        return None
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    ws = wb[wb.sheetnames[0]]

    # Build raw cell grid
    grid = {}
    fmt_grid = {}
    for ri, row in enumerate(ws.iter_rows(values_only=False)):
        for ci, cell in enumerate(row):
            grid[(ri, ci)] = cell.value
            fmt_grid[(ri, ci)] = cell.number_format or ''

    # Find num_die
    num_die = None
    for ri in range(ws.max_row):
        if grid.get((ri, 0)) == '# Die':
            num_die = grid.get((ri, 1))
            break

    if not num_die:
        return None

    # Memoised formula evaluator
    _ev_cache = {}

    def _ev(ri, ci):
        if (ri, ci) in _ev_cache:
            return _ev_cache[(ri, ci)]
        v = grid.get((ri, ci))
        if v is None:
            _ev_cache[(ri, ci)] = None
            return None
        if not isinstance(v, str) or not v.startswith('='):
            _ev_cache[(ri, ci)] = v
            return v
        result = None
        # =N/B2
        m1 = re.match(r'^=(-?\d+)/B\d+$', v)
        if m1:
            result = int(m1.group(1)) / num_die
        # =SUM(X3:X6)
        elif re.match(r'^=SUM\([A-Z]+\d+:[A-Z]+\d+\)$', v):
            m2 = re.match(r'^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', v)
            if m2:
                c1 = column_index_from_string(m2.group(1)) - 1
                r1 = int(m2.group(2)) - 1
                c2 = column_index_from_string(m2.group(3)) - 1
                r2 = int(m2.group(4)) - 1
                result = sum(
                    _ev(r, c1) for r in range(r1, r2 + 1)
                    if isinstance(_ev(r, c1), (int, float))
                )
        # =SUM(X3,X7,...) — comma separated refs
        elif re.match(r'^=SUM\(([^)]+)\)$', v):
            m3 = re.match(r'^=SUM\(([^)]+)\)$', v)
            if m3:
                total = 0.0
                for ref in m3.group(1).split(','):
                    mr = re.match(r'^([A-Z]+)(\d+)$', ref.strip())
                    if mr:
                        rc = column_index_from_string(mr.group(1)) - 1
                        rr = int(mr.group(2)) - 1
                        sv = _ev(rr, rc)
                        if isinstance(sv, (int, float)):
                            total += sv
                result = total
        _ev_cache[(ri, ci)] = result
        return result

    def _pct(ri, ci):
        v = _ev(ri, ci)
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return round(v * 100, 2)
        return None

    # Find column headers row
    col_hdrs = []
    hdr_ri = None
    for ri in range(ws.max_row):
        if grid.get((ri, 0)) == 'Sub Module':
            hdr_ri = ri
            col_hdrs = [grid.get((ri, ci)) for ci in range(ws.max_column)
                        if grid.get((ri, ci)) is not None]
            break

    if hdr_ri is None:
        return None

    num_cols = len(col_hdrs) - 1  # exclude 'Sub Module' column

    # Which data columns are percentages (header contains '%')
    col_is_pct = ['%' in str(h) for h in col_hdrs[1:]]

    def _read_val(ri, ci):
        """Return pct-scaled value for % columns, raw value otherwise."""
        idx = ci - 1  # 0-based into col_is_pct
        if idx < len(col_is_pct) and col_is_pct[idx]:
            return _pct(ri, ci)
        v = _ev(ri, ci)
        if isinstance(v, (int, float)):
            return v
        return None

    # Parse data rows into groups
    groups = []
    current_rows = []
    totals = None

    for ri in range(hdr_ri + 1, ws.max_row + 1):
        label = grid.get((ri, 0))
        if label is None:
            continue
        label_str = str(label).strip()
        if not label_str or label_str.startswith('#'):
            continue

        vals = [_read_val(ri, ci) for ci in range(1, 1 + num_cols)]

        if label_str.upper() == 'SUM':
            groups.append({
                'rows': current_rows,
                'sum_vals': vals,
            })
            current_rows = []
        elif re.match(r'^Yield\s+SUM', label_str, re.IGNORECASE):
            totals = {'label': label_str, 'vals': vals}
        else:
            current_rows.append({'label': label_str, 'vals': vals})

    # Flush any trailing rows without a SUM (e.g. TPI/Other group)
    if current_rows:
        groups.append({'rows': current_rows, 'sum_vals': [None] * num_cols})

    return {
        'num_die': num_die,
        'col_headers': col_hdrs[1:],
        'col_is_pct': col_is_pct,
        'groups': groups,
        'totals': totals,
        'sheet': wb.sheetnames[0],
    }


# ---------------------------------------------------------------------------
# 4. Chart helpers
# ---------------------------------------------------------------------------

def _fig_b64(fig, dpi=130):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('ascii')


def _esc(s):
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


# Color palette for identifiers (up to 10)
_ID_COLORS = [
    '#2980b9', '#27ae60', '#e74c3c', '#f39c12',
    '#8e44ad', '#16a085', '#d35400', '#2c3e50',
    '#c0392b', '#1abc9c',
]

_GROUP_COLORS = [
    '#eaf4ea', '#e3f0fc', '#fff6e6', '#f3e8fb',
    '#fdecea', '#e0f7fa', '#fdf6e3', '#f9ece8',
]


def build_sum_comparison_chart(runs_data, col_idx=0):
    """Grouped bar chart: each group's SUM value per identifier."""
    if not HAVE_MPL:
        return ''
    labels = [r['name'] for r in runs_data]
    # Collect group names from the run with most groups
    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    groups = best['data']['groups']
    group_names = []
    for g in groups:
        if g['rows']:
            # Use prefix of first row's label as group name
            first = g['rows'][0]['label']
            prefix = re.match(r'^([A-Za-z]+)', first)
            group_names.append(prefix.group(1) if prefix else first[:8])
        else:
            group_names.append('?')

    n_groups = len(group_names)
    n_runs = len(runs_data)
    x = np.arange(n_groups)
    width = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(max(10, n_groups * 1.4), 5))
    for ri, run in enumerate(runs_data):
        if not run['data']:
            continue
        vals = []
        for gi, g in enumerate(run['data']['groups']):
            sv = g['sum_vals'][col_idx] if col_idx < len(g['sum_vals']) else None
            vals.append(sv if sv is not None else 0.0)
        # Pad if fewer groups
        while len(vals) < n_groups:
            vals.append(0.0)
        offset = (ri - n_runs / 2 + 0.5) * width
        bars = ax.bar(x + offset, vals, width, label=run['name'],
                      color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)
        for bar, v in zip(bars, vals):
            if v and v > 0.3:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.1,
                        f'{v:.1f}%', ha='center', va='bottom', fontsize=6.5, rotation=90)

    col_label = runs_data[0]['data']['col_headers'][col_idx] if runs_data[0]['data'] else f'Col {col_idx+1}'
    ax.set_title(f'Group SUM Comparison — {col_label}', fontsize=13, weight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(group_names, rotation=30, ha='right', fontsize=9)
    ax.set_ylabel('Yield / Fallout (%)')
    ax.legend(fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_submodule_chart(runs_data, col_idx=0):
    """Horizontal bar chart: each sub-module row per identifier (top N)."""
    if not HAVE_MPL:
        return ''
    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    all_rows = []
    for g in best['data']['groups']:
        all_rows.extend(g['rows'])

    if not all_rows:
        return ''

    # Sort by first run's value descending (top 20)
    def _first_val(row_label):
        for r in runs_data:
            if not r['data']:
                continue
            for g in r['data']['groups']:
                for row in g['rows']:
                    if row['label'] == row_label:
                        v = row['vals'][col_idx] if col_idx < len(row['vals']) else None
                        return v or 0.0
        return 0.0

    sorted_rows = sorted(all_rows, key=lambda r: _first_val(r['label']), reverse=True)[:20]
    row_labels = [r['label'] for r in sorted_rows]

    n_runs = len(runs_data)
    n_rows = len(row_labels)
    y = np.arange(n_rows)
    height = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(9, max(5, n_rows * 0.45)))
    for ri, run in enumerate(runs_data):
        if not run['data']:
            continue
        vals = []
        for lbl in row_labels:
            v = 0.0
            for g in run['data']['groups']:
                for row in g['rows']:
                    if row['label'] == lbl:
                        v = row['vals'][col_idx] if col_idx < len(row['vals']) else 0.0
                        if v is None: v = 0.0
            vals.append(v)
        offset = (ri - n_runs / 2 + 0.5) * height
        ax.barh(y + offset, vals, height, label=run['name'],
                color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)

    col_label = runs_data[0]['data']['col_headers'][col_idx] if runs_data[0]['data'] else f'Col {col_idx+1}'
    ax.set_title(f'Sub-Module Breakdown — {col_label}', fontsize=13, weight='bold')
    ax.set_yticks(y)
    ax.set_yticklabels(row_labels, fontsize=8)
    ax.set_xlabel('Fallout / Yield (%)')
    ax.legend(fontsize=8)
    ax.grid(axis='x', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_total_yield_chart(runs_data):
    """Single bar chart: Yield SUM (%) per identifier."""
    if not HAVE_MPL:
        return ''
    labels = [r['name'] for r in runs_data]
    vals = []
    for r in runs_data:
        if r['data'] and r['data']['totals']:
            v = r['data']['totals']['vals'][0]
            vals.append(v if v is not None else 0.0)
        else:
            vals.append(0.0)

    if not any(vals):
        return ''

    fig, ax = plt.subplots(figsize=(max(6, len(labels) * 1.2), 4))
    bars = ax.bar(labels, vals, color=[_ID_COLORS[i % len(_ID_COLORS)] for i in range(len(labels))],
                  alpha=0.88, edgecolor='white')
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.2,
                f'{v:.1f}%', ha='center', va='bottom', fontsize=9, weight='bold')
    ax.set_title('Total Yield Loss SUM by Identifier', fontsize=13, weight='bold')
    ax.set_ylabel('Yield Loss (%)')
    ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=9)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_delta_heatmap(runs_data, ref_name, col_idx=0):
    """Heatmap of delta (run - ref) per sub-module."""
    if not HAVE_MPL:
        return ''
    ref_run = next((r for r in runs_data if r['name'] == ref_name), runs_data[0])
    compare_runs = [r for r in runs_data if r['name'] != ref_run['name']]
    if not compare_runs:
        return ''

    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    all_rows = []
    for g in best['data']['groups']:
        all_rows.extend(g['rows'])
    row_labels = [r['label'] for r in all_rows]

    def _get_val(run, lbl):
        if not run['data']:
            return None
        for g in run['data']['groups']:
            for row in g['rows']:
                if row['label'] == lbl:
                    return row['vals'][col_idx] if col_idx < len(row['vals']) else None
        return None

    ref_vals = [_get_val(ref_run, lbl) for lbl in row_labels]
    matrix = []
    col_labels = []
    for r in compare_runs:
        row_vals = [_get_val(r, lbl) for lbl in row_labels]
        deltas = [
            (rv - refv) if (rv is not None and refv is not None) else float('nan')
            for rv, refv in zip(row_vals, ref_vals)
        ]
        matrix.append(deltas)
        col_labels.append(r['name'])

    data_arr = np.array(matrix)  # shape: (n_compare, n_rows)
    if data_arr.size == 0:
        return ''

    fig, ax = plt.subplots(figsize=(max(8, len(row_labels) * 0.55), max(3, len(col_labels) * 0.7 + 1.5)))
    # Diverging colormap: red = higher fallout (worse), green = lower (better)
    vmax = np.nanmax(np.abs(data_arr))
    if vmax == 0:
        vmax = 1
    im = ax.imshow(data_arr, cmap='RdYlGn_r', aspect='auto',
                   vmin=-vmax, vmax=vmax)
    ax.set_xticks(range(len(row_labels)))
    ax.set_xticklabels(row_labels, rotation=60, ha='right', fontsize=7)
    ax.set_yticks(range(len(col_labels)))
    ax.set_yticklabels(col_labels, fontsize=8)
    for yi in range(len(col_labels)):
        for xi in range(len(row_labels)):
            v = data_arr[yi, xi]
            if not np.isnan(v):
                ax.text(xi, yi, f'{v:+.1f}', ha='center', va='center',
                        fontsize=6, color='black')
    plt.colorbar(im, ax=ax, label='Δ % vs reference')
    col_label = runs_data[0]['data']['col_headers'][col_idx] if runs_data[0]['data'] else ''
    ax.set_title(f'Delta vs {ref_name} — {col_label}\n(red = more fallout, green = less)',
                 fontsize=11, weight='bold')
    fig.tight_layout()
    return _fig_b64(fig)


# ---------------------------------------------------------------------------
# 4b. RDND yield-table charts and Bin Fail chart
# ---------------------------------------------------------------------------

# Palette for stacked fail segments (consistent across runs)
_FAIL_COLORS = [
    '#e74c3c', '#e67e22', '#f39c12', '#2ecc71',
    '#1abc9c', '#3498db', '#9b59b6', '#e8177d',
    '#95a5a6', '#34495e', '#c0392b', '#16a085',
]


def build_top10_pareto_chart(runs_data):
    """Horizontal bar chart: top-10 Interface Bins by max fail% — uses Functional Bin table."""
    if not HAVE_MPL:
        return ''
    # Prefer func_bin_rows (has fbin); fall back to bin_fail_rows
    use_fbin = any(r.get('bin_data') and r['bin_data'].get('func_bin_rows')
                   for r in runs_data)
    row_key  = 'func_bin_rows' if use_fbin else 'bin_fail_rows'
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data'].get(row_key)]
    if not valid:
        return ''

    # Collect all ibins; keep max fail% per ibin
    ibin_max    = {}
    ibin_fbin   = {}
    ibin_bucket = {}
    for run in valid:
        for row in run['bin_data'][row_key]:
            k = row['ibin']
            v = row['fail_pct'] or 0.0
            if v > ibin_max.get(k, 0.0):
                ibin_max[k]    = v
                ibin_bucket[k] = row.get('fail_bucket', '')
                ibin_fbin[k]   = row.get('fbin', '')

    top10 = sorted(ibin_max.keys(), key=lambda k: ibin_max[k], reverse=True)[:10]
    if not top10:
        return ''

    n_runs = len(valid)
    n_bins = len(top10)
    y      = np.arange(n_bins)
    height = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(9, max(4, n_bins * 0.65)))
    for ri, run in enumerate(valid):
        vals = [
            next((ro['fail_pct'] for ro in run['bin_data'][row_key] if ro['ibin'] == k), 0.0) or 0.0
            for k in top10
        ]
        offset = (ri - n_runs / 2 + 0.5) * height
        bars = ax.barh(y + offset, vals, height, label=run['name'],
                       color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.87)
        for bar, v in zip(bars, vals):
            if v >= 0.1:
                ax.text(v + 0.05, bar.get_y() + bar.get_height() / 2,
                        f'{v:.2f}%', va='center', ha='left', fontsize=7)

    if use_fbin:
        ylabels = [
            f'iBin {k}  FBin {ibin_fbin.get(k, "—")}  |  {ibin_bucket.get(k, "")}'
            for k in top10
        ]
    else:
        ylabels = [
            f'iBin {k}  |  {ibin_bucket.get(k, "")}'
            for k in top10
        ]
    ax.set_yticks(y)
    ax.set_yticklabels(ylabels, fontsize=7.5)
    ax.invert_yaxis()
    ax.set_xlabel('Fail (%)')
    ax.set_title('Top 10 Interface Bin Fail Pareto', fontsize=13, weight='bold')
    ax.legend(fontsize=8)
    ax.grid(axis='x', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_upm_median_chart(runs_data):
    """Grouped bar chart: SICC Si Actual Median vs Target per test, per run."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('upm_data') and r['upm_data']]
    if not valid:
        return ''

    # Collect test names from run with most rows
    best = max(valid, key=lambda r: len(r['upm_data']))
    # Only tests that have both actual and target
    tests = [row for row in best['upm_data'] if row.get('sicc_actual') is not None]
    if not tests:
        return ''
    test_labels = [row['test'] for row in tests]

    n_tests = len(test_labels)
    n_runs  = len(valid)
    x       = np.arange(n_tests)
    width   = 0.7 / max(n_runs + 1, 2)  # +1 for target bar

    fig, ax = plt.subplots(figsize=(max(7, n_tests * 0.5), 4))

    # Plot target once (from best run)
    targets = [t.get('sicc_target') for t in tests]
    has_target = any(v is not None for v in targets)

    for ri, run in enumerate(valid):
        actuals = []
        for t in tests:
            row = next((r for r in run['upm_data'] if r['test'] == t['test']), None)
            actuals.append(row['sicc_actual'] if row and row.get('sicc_actual') is not None else 0.0)
        offset = ri * width
        ax.bar(x + offset, actuals, width, label=run['name'],
               color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)

    if has_target:
        tgt_vals = [v if v is not None else 0.0 for v in targets]
        offset = n_runs * width
        ax.bar(x + offset, tgt_vals, width, label='Target',
               color='#2c3e50', alpha=0.55, hatch='//')

    # Centre ticks
    tick_offset = (n_runs * width) / 2
    ax.set_xticks(x + tick_offset)
    ax.set_xticklabels(test_labels, rotation=35, ha='right', fontsize=8)
    ax.set_ylabel('SICC Si Actual Median (A)')
    ax.set_title('SICC Median', fontsize=13, weight='bold')
    ax.yaxis.set_major_locator(plt.MaxNLocator(nbins=12, integer=False))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: '0' if v == 0 else f'{v:.3f}'))
    ax.legend(fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def build_upm_pct_chart(runs_data, upm_target_pct=None, upm_target_label='Target'):
    """Bar chart: UPM % per run (single value per run)."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('upm_data') and r['upm_data']]
    if not valid:
        return ''

    labels = []
    vals   = []
    for run in valid:
        pcts = [row['upm_pct'] for row in run['upm_data'] if row.get('upm_pct') is not None]
        if pcts:
            labels.append(run['name'])
            vals.append(pcts[0])  # UPM % is the same for all rows

    if not vals:
        return ''

    fig, ax = plt.subplots(figsize=(max(4, len(labels) * 0.9), 3.5))
    bars = ax.bar(labels, vals,
                  color=[_ID_COLORS[i % len(_ID_COLORS)] for i in range(len(labels))],
                  alpha=0.88, edgecolor='white')
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.1,
                f'{v:.2f}%', ha='center', va='bottom', fontsize=9, weight='bold')
    if upm_target_pct is not None:
        ax.axhline(upm_target_pct, color='red', linewidth=2, linestyle='--',
                   label=f'{upm_target_label}: {upm_target_pct}%', zorder=5)
        ax.legend(fontsize=9)
    ax.set_title('UPM ULVT 950mV (%)', fontsize=13, weight='bold')
    ax.set_ylabel('UPM (%)')
    _ymax = max(vals + ([upm_target_pct] if upm_target_pct is not None else [])) * 1.15 if vals else 100
    ax.set_ylim(0, _ymax)
    ax.yaxis.set_major_locator(plt.MaxNLocator(nbins=10, integer=False))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: '0' if v == 0 else f'{v:.2f}%'))
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=9)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def build_upm_detail_chart(runs_data, upm_target_pct=None, upm_target_label='Target'):
    """Grouped bar chart: UPM median % per column per run."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('upm_detail') and r['upm_detail']]
    if not valid:
        return ''

    import numpy as np

    # Collect all unique short names across runs (preserve order)
    all_shorts = []
    for run in valid:
        for d in run['upm_detail']:
            if d['short'] not in all_shorts:
                all_shorts.append(d['short'])

    if not all_shorts:
        return ''

    n_cols = len(all_shorts)
    n_runs = len(valid)
    bar_w = 0.8 / n_runs
    x = np.arange(n_cols)

    fig, ax = plt.subplots(figsize=(max(6, n_cols * 1.2 + n_runs * 0.3), 4.5))

    for ri, run in enumerate(valid):
        detail_map = {d['short']: d['pct'] for d in run['upm_detail']}
        vals = [detail_map.get(s, 0) for s in all_shorts]
        offset = x + ri * bar_w - (n_runs - 1) * bar_w / 2
        bars = ax.bar(offset, vals, width=bar_w,
                      color=_ID_COLORS[ri % len(_ID_COLORS)],
                      alpha=0.85, edgecolor='white', label=run['name'])
        for bar, v in zip(bars, vals):
            if v > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.15,
                        f'{v:.1f}%', ha='center', va='bottom', fontsize=6, weight='bold')

    if upm_target_pct is not None:
        ax.axhline(upm_target_pct, color='red', linewidth=2, linestyle='--',
                   label=f'{upm_target_label}: {upm_target_pct}%', zorder=5)

    ax.set_title('UPM Distribution Comparison (Median %)', fontsize=12, weight='bold')
    ax.set_ylabel('Median (%)', fontsize=10)
    ax.set_xticks(x)
    ax.set_xticklabels(all_shorts, rotation=45, ha='right', fontsize=7)
    ax.legend(fontsize=8, loc='best')
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def _bin_all_good(bin_str: str) -> bool:
    """Return True if every number in the bin string is <= 4."""
    nums = [int(x) for x in re.findall(r'\d+', str(bin_str))]
    return bool(nums) and all(n <= 4 for n in nums)


# Bin groups to always show in the top subplot — matched by bin string equality
_KEY_BINS = ['1/2/3/4', '1/2']
# Fallback: treat any row whose label exactly matches one of these
_KEY_BIN_TITLES = {
    '1/2/3/4': 'FF+DF  (Bin 1/2/3/4)',
    '1/2':     'FF  (Bin 1/2)',
}


def build_combined_rdnd_chart(runs_data):
    """Single mixed-format chart:
       - Stacked bars (per identifier) for Bins > 4 fail%      [left Y, 0-100%]
       - Line overlay (per key bin) for FF+DF / FF yield%       [right Y, 0-100%]
       - Dashed hline = expected yield per key bin
       - Dashed tick marker = expected fail total per identifier
    """
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]
    if not valid:
        return ''

    from matplotlib.patches import Patch
    from matplotlib.lines import Line2D

    best = max(valid, key=lambda r: len(r['bin_data']['yield_rows']))
    all_yield_rows = best['bin_data']['yield_rows']

    key_rows  = [row for row in all_yield_rows if row['bin'] in _KEY_BINS]
    if not key_rows:
        key_rows = [row for row in all_yield_rows if _bin_all_good(row['bin'])][:2]
    fail_rows = [row for row in all_yield_rows if not _bin_all_good(row['bin'])]

    n_runs    = len(valid)
    id_labels = [r['name'] for r in valid]
    x         = np.arange(n_runs)
    bar_w     = 0.5

    fig, ax_bar = plt.subplots(figsize=(max(10, n_runs * 2.2), 6))
    ax_line = ax_bar.twinx()          # shared x, independent right Y

    legend_elems = []

    # ---- Stacked bars: Bins > 4 fail% ----
    bottoms = np.zeros(n_runs)
    for si, fr in enumerate(fail_rows):
        bin_key = fr['bin']
        seg_lbl = fr.get('fail_bucket') or bin_key
        clr     = _FAIL_COLORS[si % len(_FAIL_COLORS)]
        vals    = np.array([
            (next((ro['yield_pct'] for ro in r['bin_data']['yield_rows']
                   if ro['bin'] == bin_key), None) or 0.0)
            for r in valid
        ])
        ax_bar.bar(x, vals, bar_w, bottom=bottoms, color=clr,
                   alpha=0.80, edgecolor='white', linewidth=0.5, zorder=2)
        for bi, (v, b) in enumerate(zip(vals, bottoms)):
            if v >= 0.8:
                ax_bar.text(bi, b + v / 2, f'{v:.1f}%',
                            ha='center', va='center', fontsize=6.5,
                            color='white', weight='bold')
        legend_elems.append(Patch(color=clr, alpha=0.80, label=seg_lbl))
        bottoms += vals

    # Dashed expected-total tick per identifier
    for xi, run in enumerate(valid):
        exp_total = sum(
            (next((ro['expected_pct'] for ro in run['bin_data']['yield_rows']
                   if ro['bin'] == fr['bin']), None) or 0.0)
            for fr in fail_rows
        )
        if exp_total:
            ax_bar.plot([xi - 0.28, xi + 0.28], [exp_total, exp_total],
                        color='#2c3e50', linestyle='--', linewidth=2.0, zorder=5)

    legend_elems.append(Line2D([0], [0], color='#2c3e50', linestyle='--',
                                linewidth=2.0, label='Expected Fail Total'))

    ax_bar.set_ylabel('Fail (%)',  fontsize=10)
    # Dynamic y-limit: max stacked fail total across all runs + 15% buffer
    max_fail = float(bottoms.max()) if bottoms.max() > 0 else 1.0
    also_exp = []
    for run in valid:
        exp_t = sum(
            (next((ro['expected_pct'] for ro in run['bin_data']['yield_rows']
                   if ro['bin'] == fr['bin']), None) or 0.0)
            for fr in fail_rows
        )
        if exp_t:
            also_exp.append(exp_t)
    if also_exp:
        max_fail = max(max_fail, max(also_exp))
    fail_ylim = min(100.0, max_fail * 1.20)   # 20% headroom, cap at 100
    fail_ylim = max(fail_ylim, 5.0)            # at least 5% so chart is readable
    ax_bar.set_ylim(0, fail_ylim)
    ax_bar.set_xticks(x)
    ax_bar.set_xticklabels(id_labels, rotation=15, ha='right', fontsize=9)
    ax_bar.grid(axis='y', linestyle='--', alpha=0.3, zorder=0)

    # ---- Lines: key-bin yield% (right Y axis) ----
    line_colors = ['#1a73e8', '#e53935', '#2e7d32', '#f57c00']
    for ki, kr in enumerate(key_rows):
        bin_key = kr['bin']
        lbl = _KEY_BIN_TITLES.get(bin_key,
                                  f"{bin_key} {kr.get('fail_bucket', '')}".strip())
        lclr = line_colors[ki % len(line_colors)]
        vals = np.array([
            (next((ro['yield_pct'] for ro in r['bin_data']['yield_rows']
                   if ro['bin'] == bin_key), None) or 0.0)
            for r in valid
        ])
        ax_line.plot(x, vals, marker='o', linewidth=2.4, markersize=8,
                     color=lclr, label=lbl, zorder=6)
        for xi, v in enumerate(vals):
            ax_line.text(xi, v + 1.5, f'{v:.1f}%', ha='center', va='bottom',
                         fontsize=8, color=lclr, weight='bold')
        # Dashed expected hline
        ev = kr.get('expected_pct')
        if ev is not None:
            ax_line.axhline(ev, color=lclr, linestyle=':', linewidth=1.6,
                            alpha=0.65, zorder=4)
            ax_line.text(n_runs - 0.45, ev - 2.5 - ki * 3.0, f'Exp {ev:.1f}%',
                         ha='right', va='top', fontsize=7.5,
                         color=lclr, alpha=0.85)
        legend_elems.append(Line2D([0], [0], color=lclr, marker='o',
                                   linewidth=2.4, markersize=8, label=lbl))

    ax_line.set_ylabel('Yield (%)', fontsize=10, y=0.4)
    ax_line.set_ylim(0, 100)

    ax_bar.set_title(
        'Yield (%) and Fail (%) Chart',
        fontsize=12, weight='bold'
    )
    ax_bar.legend(handles=legend_elems, fontsize=7.5,
                  loc='upper left', bbox_to_anchor=(1.18, 1.0),
                  borderaxespad=0)

    fig.tight_layout(pad=2.0)
    return _fig_b64(fig)


def build_fail_stacked_chart(runs_data):
    """Kept for backward compat — now delegates to combined chart."""
    return build_combined_rdnd_chart(runs_data)


def build_rdnd_delta_chart(runs_data):
    """Bars: (actual yield - expected) per BIN per run. Green = beating target, red = below."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]
    if not valid:
        return ''
    best = max(valid, key=lambda r: len(r['bin_data']['yield_rows']))
    yield_rows    = best['bin_data']['yield_rows']
    bin_labels    = [row['bin'] for row in yield_rows]
    expected_vals = [row['expected_pct'] for row in yield_rows]

    n_bins = len(bin_labels)
    n_runs = len(valid)
    x      = np.arange(n_bins)
    width  = 0.8 / max(n_runs, 1)

    from matplotlib.patches import Patch
    fig, ax = plt.subplots(figsize=(max(12, n_bins * 1.1), 4))
    legend_elems = []
    for ri, run in enumerate(valid):
        actuals = [
            next((r['yield_pct'] for r in run['bin_data']['yield_rows'] if r['bin'] == bl), None)
            for bl in bin_labels
        ]
        deltas = [
            (a - e) if (a is not None and e is not None) else 0.0
            for a, e in zip(actuals, expected_vals)
        ]
        offset = (ri - n_runs / 2 + 0.5) * width
        bar_colors = ['#27ae60' if d > 0 else ('#c0392b' if d < 0 else '#95a5a6')
                      for d in deltas]
        ax.bar(x + offset, deltas, width, color=bar_colors, alpha=0.85)
        legend_elems.append(Patch(color=_ID_COLORS[ri % len(_ID_COLORS)],
                                  alpha=0.85, label=run['name']))
    ax.axhline(0, color='#2c3e50', linewidth=1)
    ax.set_title('RDND Yield Delta vs Expected  (actual - expected)\n'
                 'Green = beating target  |  Red = below target',
                 fontsize=12, weight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(bin_labels, rotation=35, ha='right', fontsize=7)
    ax.set_ylabel('\u0394 Yield (%)')
    ax.legend(handles=legend_elems, fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_bin_fail_chart(runs_data):
    """Horizontal grouped bars: fail% per Interface Bin per run, sorted by max fail."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['bin_fail_rows']]
    if not valid:
        return ''

    # Union of ibins in order from the run with most rows
    best = max(valid, key=lambda r: len(r['bin_data']['bin_fail_rows']))
    ibin_meta = {}
    for row in best['bin_data']['bin_fail_rows']:
        ibin_meta.setdefault(row['ibin'], row['fail_bucket'])

    def _max_fail(ibin):
        return max(
            (next((ro['fail_pct'] for ro in r['bin_data']['bin_fail_rows'] if ro['ibin'] == ibin), 0.0) or 0.0)
            for r in valid
        )

    sorted_ibins = sorted(ibin_meta.keys(), key=_max_fail, reverse=True)
    ibin_labels  = [f"Bin {k} \u2014 {ibin_meta[k]}" for k in sorted_ibins]

    n_rows = len(sorted_ibins)
    n_runs = len(valid)
    y      = np.arange(n_rows)
    height = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(9, max(5, n_rows * 0.42)))
    for ri, run in enumerate(valid):
        vals = [
            next((ro['fail_pct'] for ro in run['bin_data']['bin_fail_rows'] if ro['ibin'] == k), 0.0) or 0.0
            for k in sorted_ibins
        ]
        offset = (ri - n_runs / 2 + 0.5) * height
        ax.barh(y + offset, vals, height, label=run['name'],
                color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)
    ax.set_title('Bin Fail Summary — Fail% per Interface Bin', fontsize=13, weight='bold')
    ax.set_yticks(y)
    ax.set_yticklabels(ibin_labels, fontsize=7)
    ax.set_xlabel('Fail (%)')
    ax.legend(fontsize=8)
    ax.grid(axis='x', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


# ---------------------------------------------------------------------------
# 4c. Run summary table (Program Name, Material Type, Wafer, # Dies)
# ---------------------------------------------------------------------------

def parse_index_meta(dash_dir: Path, index_href: str) -> dict:
    """Parse program, lot(s), wafer(s), material type(s) from a run's BinDistribution HTML
    by reading the embedded DATA JS variable written by bin_distribution_html.py."""
    result = {'program': '', 'lots': [], 'wafers': [], 'material': []}
    if not index_href:
        return result
    try:
        import json as _json_idx
        href = re.sub(r'^file:///', '', index_href).replace('/', os.sep)
        idx_path = dash_dir / href if not os.path.isabs(href) else Path(href)
        out_folder = idx_path.parent
        if not out_folder.exists():
            return result
        # Find the best BinDistribution.html: prefer *_reticle_material_BinDistribution.html
        bin_html = None
        for pat in ('*_reticle_material_BinDistribution.html',
                    '*_material_merged_*BinDistribution.html',
                    '*BinDistribution.html'):
            cands = sorted(out_folder.glob(pat), key=lambda p: p.stat().st_mtime, reverse=True)
            if cands:
                bin_html = cands[0]
                break
        if not bin_html:
            return result
        content = bin_html.read_text(encoding='utf-8', errors='replace')
        # Extract var DATA = {...}; — ends before the next var declaration
        m = re.search(r'var\s+DATA\s*=\s*(\{[\s\S]*?\});\s*(?:var\s|\Z)', content)
        if not m:
            m = re.search(r'var\s+DATA\s*=\s*(\{[\s\S]*?\});', content)
        if m:
            try:
                data = _json_idx.loads(m.group(1))
                rows = data.get('rows', [])
                programs, lots, wafers, mats = set(), set(), set(), set()
                for row in rows:
                    if row.get('program'):
                        programs.add(str(row['program']))
                    if row.get('lot'):
                        lots.add(str(row['lot']))
                    if row.get('wafer') and str(row['wafer']) not in ('', 'all'):
                        wafers.add(str(row['wafer']))
                    if row.get('material'):
                        mats.add(str(row['material']))
                def _sort_nums(s):
                    try:
                        return sorted(s, key=lambda x: int(x) if str(x).isdigit() else x)
                    except Exception:
                        return sorted(s)
                result['program']  = ', '.join(sorted(programs))
                result['lots']     = _sort_nums(lots)
                result['wafers']   = _sort_nums(wafers)
                result['material'] = sorted(mats)
            except Exception:
                pass
    except Exception:
        pass
    return result


def _find_processed_csv(output_dir: Path):
    """Find the best enriched/processed CSV for a run.
    Priority: *_reticle_material.csv > *_material_merged.csv > any CSV in output_dir > parent dir."""
    def _excl(name: str) -> bool:
        n = name.lower()
        return ('bindef' in n or '_targets_' in n or '_bindef' in n)

    def _csvs(d: Path):
        if not d or not d.exists():
            return []
        return [p for p in sorted(d.glob('*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
                if not _excl(p.name)]

    # output_dir is the identifier subfolder — enriched CSVs are written there
    for search_dir in [output_dir,
                       output_dir.parent if output_dir else None,
                       output_dir.parent.parent if output_dir and output_dir.parent else None]:
        if search_dir is None:
            continue
        cands = _csvs(search_dir)
        if not cands:
            continue
        # Strongest preference: fully enriched (reticle + material) file
        for p in cands:
            if '_reticle_material' in p.name.lower():
                return p
        # Second preference: material-merged intermediate
        for p in cands:
            if '_material_merged' in p.name.lower():
                return p
        # Third: any CSV with 'reticle' or 'material' in name
        for p in cands:
            n = p.name.lower()
            if 'reticle' in n or 'material' in n or 'enriched' in n:
                return p
        # Fallback: newest CSV in this dir
        return cands[0]
    return None


def _extract_csv_meta(output_dir: Path) -> dict:
    """Read the processed/enriched CSV to extract Program Name, Material Type, Wafer list."""
    result = {'program': '', 'material': '', 'wafers': ''}
    raw = _find_processed_csv(output_dir)
    if not raw:
        return result
    try:
        import pandas as _pd_meta
        # Read only first 10k rows for speed; use low_memory=False so mixed cols don't truncate
        df = _pd_meta.read_csv(str(raw), dtype=object, nrows=10000)
        prog_col = next((c for c in df.columns if 'program' in c.lower()), None)
        # Material Type column: look for exact match first, then substring
        mat_col = next(
            (c for c in df.columns if c.strip().lower() in
             ('material type', 'materialtype', 'material_type',
              'material type, skew, beol skew')),
            None
        ) or next((c for c in df.columns if 'material type' in c.lower()), None)
        wafer_col = (next((c for c in df.columns if 'sort_wafer' in c.lower()), None)
                     or next((c for c in df.columns if 'wafer' in c.lower()), None))
        if prog_col:
            vals = [str(v) for v in df[prog_col].dropna().unique()]
            result['program'] = ', '.join(vals[:4])
        if mat_col:
            vals = [str(v) for v in df[mat_col].dropna().unique()]
            result['material'] = ', '.join(vals[:6])
        if wafer_col:
            wvals = df[wafer_col].dropna().unique()
            try:
                wvals = sorted(wvals, key=lambda x: int(str(x)) if str(x).isdigit() else str(x))
            except Exception:
                wvals = sorted([str(w) for w in wvals])
            result['wafers'] = ', '.join([str(w) for w in wvals])
    except Exception:
        pass
    return result


def build_run_summary_table_html(runs_data, dash_dir: Path = None) -> str:
    """Compact table: Program Name / Lot(s) / Material Type / Wafer(s) / # Dies — one column per run."""
    metas = []
    for r in runs_data:
        meta = {'program': '', 'lots': '', 'material': '', 'wafers': '', 'num_die': ''}

        # # Dies from xlsx
        if r.get('data') and r['data'].get('num_die') is not None:
            try:
                meta['num_die'] = f"{int(r['data']['num_die']):,}"
            except Exception:
                meta['num_die'] = str(r['data']['num_die'])

        # Program / Lot / Wafer / Material from the run's index.html UDATA
        if dash_dir and r.get('index_href'):
            idx_meta = parse_index_meta(dash_dir, r['index_href'])
            meta['lots']     = _esc(', '.join(idx_meta['lots']))
            meta['wafers']   = _esc(', '.join(idx_meta['wafers']))
            meta['material'] = _esc(', '.join(idx_meta['material']))
            # Program comes from UDATA program field; if empty fall back to run name
            if idx_meta.get('program'):
                meta['program'] = _esc(idx_meta['program'])
            else:
                meta['program'] = _esc(r.get('name', ''))
        else:
            meta['program'] = _esc(r.get('name', ''))

        metas.append(meta)

    ROWS = [
        ('Program Name',  'program'),
        ('Lot(s)',        'lots'),
        ('Material Type', 'material'),
        ('Wafer(s)',      'wafers'),
        ('# Dies',        'num_die'),
    ]

    _th_base = 'white-space:nowrap;padding:6px 10px'
    hdr = f'<th style="background:#1a3a5c;color:#fff;{_th_base}">Metric</th>'
    for ri, r in enumerate(runs_data):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th style="background:{clr};color:#fff;font-weight:bold;'
                f'text-align:center;{_th_base}">'
                f'{_esc(r["name"])}</th>')

    _td_base = 'padding:5px 10px;font-size:20px;white-space:nowrap'
    rows_html = ''
    for label, key in ROWS:
        cells = (f'<td style="font-weight:bold;background:#e8f0fb;color:#1a3a5c;{_td_base}">{label}</td>')
        for mi, m in enumerate(metas):
            bg = '#ffffff' if mi % 2 == 0 else '#f5f8ff'
            val = m[key] or '\u2014'
            cells += (f'<td style="background:{bg};color:#222;{_td_base}">{val}</td>')
        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128204; Run Summary</h2>
  <div style="overflow-x:auto">
  <table style="border-collapse:collapse;table-layout:auto">
    <thead><tr>{hdr}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 4d. RDND and Bin Fail comparison table HTML
# ---------------------------------------------------------------------------

def build_xlsx_comparison_table(runs_data):
    """Side-by-side table of all rows from *_out.xlsx for every identifier."""
    valid = [r for r in runs_data if r.get('data')]
    if not valid:
        return ''

    # All runs (including those with no xlsx) — show '—' columns for missing ones
    all_runs = runs_data

    # Collect all unique sub-module labels in display order from the run with most groups
    best = max(valid, key=lambda r: len(r['data']['groups']))
    ordered_labels = []   # (label, is_sum, group_idx)
    for gi, group in enumerate(best['data']['groups']):
        for row in group['rows']:
            ordered_labels.append((row['label'], False, gi))
        ordered_labels.append(('SUM', True, gi))
    if best['data'] and best['data']['totals']:
        ordered_labels.append((best['data']['totals']['label'], True, -1))

    n_cols = max(len(r['data']['col_headers']) for r in valid)
    col_hdrs = best['data']['col_headers']
    col_is_pct = best['data'].get('col_is_pct', [True] * n_cols)

    def _fmt_val(v, ci):
        """Format cell value: add % for pct columns, plain number for raw columns."""
        if v is None:
            return ''
        if ci < len(col_is_pct) and col_is_pct[ci]:
            return f'{v:.1f}%'
        # Raw column (e.g. Die id count)
        return f'{int(v):,}' if isinstance(v, float) and v == int(v) else f'{v:g}'

    def _get_row_vals(run, lbl, is_sum, gi):
        """Return list of values (one per data column) for a given row."""
        if not run['data']:
            return [None] * n_cols
        if is_sum and lbl.upper() == 'SUM':
            grps = run['data']['groups']
            if gi < len(grps):
                sv = grps[gi]['sum_vals']
                return [sv[ci] if ci < len(sv) else None for ci in range(n_cols)]
            return [None] * n_cols
        if is_sum:  # Yield SUM row
            if run['data']['totals']:
                v = run['data']['totals']['vals']
                return [v[ci] if ci < len(v) else None for ci in range(n_cols)]
            return [None] * n_cols
        for g in run['data']['groups']:
            for row in g['rows']:
                if row['label'] == lbl:
                    return [row['vals'][ci] if ci < len(row['vals']) else None
                            for ci in range(n_cols)]
        return [None] * n_cols

    # Build header — use all_runs so every identifier appears
    hdr = '<th rowspan="2">Sub Module</th>'
    for ri, r in enumerate(all_runs):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        no_data = '' if r.get('data') else ' ⚠ no xlsx'
        hdr += (f'<th colspan="{n_cols}" style="background:{clr};color:#fff;'
                f'font-weight:bold;text-align:center;padding:6px 8px">'
                f'{_esc(r["name"])}{no_data}</th>')
    hdr2 = ''
    for ri, r in enumerate(all_runs):
        for ci in range(n_cols):
            if r.get('data'):
                ch = col_hdrs[ci] if ci < len(col_hdrs) else f'Col {ci+1}'
                hdr2 += f'<th style="font-size:17px;white-space:nowrap">{_esc(ch)}</th>'
            else:
                hdr2 += f'<th style="font-size:17px;white-space:nowrap;color:#aaa">—</th>'

    rows_html = ''
    for lbl, is_sum, gi in ordered_labels:
        grp_clr = _GROUP_COLORS[gi % len(_GROUP_COLORS)] if gi >= 0 else '#d5e8d4'
        is_total = (gi == -1)
        bold = 'font-weight:bold;' if is_sum else ''
        border = 'border-top:2px solid #aaa;' if is_sum and not is_total else ''
        border = 'border-top:3px solid #555;' if is_total else border
        lbl_cell = (f'<td style="{bold}{border}background:{grp_clr};'
                    f'font-size:20px;white-space:nowrap">{_esc(lbl)}</td>')
        cells = lbl_cell
        # Only include valid runs in the highlight calculation
        all_run_vals = [_get_row_vals(rx, lbl, is_sum, gi) for rx in valid]
        for ri, r in enumerate(all_runs):
            if not r.get('data'):
                # No xlsx for this run — show N/A cells
                for ci in range(n_cols):
                    base_st = f'{bold}{border}background:{grp_clr};color:#bbb;'
                    cells += f'<td class="num" style="{base_st}">—</td>'
                continue
            vals = _get_row_vals(r, lbl, is_sum, gi)
            for ci, v in enumerate(vals):
                base_st = f'{bold}{border}background:{grp_clr};'
                if col_is_pct[ci] if ci < len(col_is_pct) else False:
                    col_row_vals = [arv[ci] for arv in all_run_vals]
                    cells += _cell_hl(v, col_row_vals, extra_style=base_st) + _fmt_val(v, ci) + '</td>'
                else:
                    cells += (f'<td class="num" style="{base_st}">'
                              f'{_fmt_val(v, ci)}</td>')
        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128203; Digital Dashboard</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead>
      <tr>{hdr}</tr>
      <tr>{hdr2}</tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


def _cell_hl(v, row_vals, extra_style=''):
    """Return a <td> string. Bold-red if |v - row_mean| > 10 (percentage points)."""
    nums = [x for x in row_vals if x is not None]
    if v is None:
        return f'<td class="num" style="{extra_style}"></td>'
    alert = (len(nums) >= 2 and abs(v - (sum(nums) / len(nums))) > 10)
    st = extra_style + ('color:#c0392b;font-weight:bold;' if alert else '')
    return f'<td class="num" style="{st}">'


def build_rdnd_table_html(runs_data):
    """Table: BIN rows, Expected % + each run's actual yield% + delta vs expected."""
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]
    if not valid:
        return ''
    best = max(valid, key=lambda r: len(r['bin_data']['yield_rows']))
    yield_rows = best['bin_data']['yield_rows']

    hdr = '<th>BIN</th><th>Fail Bucket</th><th>Expected (%)</th>'
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += f'<th style="background:{clr};color:#fff;font-weight:bold;padding:5px 8px">{_esc(r["name"])}</th>'
    # add delta headers between consecutive runs
    if len(valid) >= 2:
        for ri in range(1, len(valid)):
            hdr += f'<th style="background:#34495e;color:#fff;font-weight:bold;padding:5px 8px;font-size:11px">\u0394 {_esc(valid[ri]["name"])}<br>vs {_esc(valid[0]["name"])}</th>'

    rows_html = ''
    for row in yield_rows:
        bl  = row['bin']
        exp = row.get('expected_pct')
        cells = (
            f'<td style="white-space:nowrap;font-size:20px">{_esc(bl)}</td>'
            f'<td style="font-size:20px">{_esc(row.get("fail_bucket", ""))}</td>'
            f'<td class="num" style="color:#555">{f"{exp:.1f}%" if exp is not None else ""}</td>'
        )
        run_vals = []
        for r in valid:
            v = next((ro['yield_pct'] for ro in r['bin_data']['yield_rows'] if ro['bin'] == bl), None)
            run_vals.append(v)
            row_vals = [
                next((ro['yield_pct'] for ro in rx['bin_data']['yield_rows'] if ro['bin'] == bl), None)
                for rx in valid
            ]
            cells += _cell_hl(v, row_vals) + (f'{v:.1f}%' if v is not None else '') + '</td>'
        # delta cells: each run vs first run (baseline)
        if len(valid) >= 2:
            base = run_vals[0]
            for ri in range(1, len(valid)):
                v = run_vals[ri]
                if base is not None and v is not None:
                    delta = v - base
                    sign = '+' if delta > 0 else ''
                    clr = '#27ae60' if delta > 0 else '#c0392b' if delta < 0 else '#555'
                    cells += f'<td class="num" style="color:{clr};font-weight:bold">{sign}{delta:.2f}%</td>'
                else:
                    cells += '<td class="num">\u2014</td>'
        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128203; Yield Table</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead><tr>{hdr}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


def build_bin_fail_table_html(runs_data):
    """Table: Interface Bin rows, each run's Yield/Fail% — uses new bin_summary_rows
    (Category/Description format) when available, else falls back to bin_fail_rows."""
    _CAT_PALETTE = [
        '#dbeeff','#e0f5e0','#fef3cd','#fde0d0','#ece0f8','#d0f4f4',
        '#fce4ec','#e8f5e9','#fff3e0','#e3f2fd','#f3e5f5','#e8eaf6',
    ]

    # Prefer new 6-col format
    use_summary = any(r.get('bin_data') and r['bin_data'].get('bin_summary_rows')
                      for r in runs_data)
    if use_summary:
        valid = [r for r in runs_data if r.get('bin_data') and r['bin_data'].get('bin_summary_rows')]
        if not valid:
            return ''
        best = max(valid, key=lambda r: len(r['bin_data']['bin_summary_rows']))
        all_rows = list({row['ibin']: row for row in best['bin_data']['bin_summary_rows']}.values())

        # Build category→color map in row order
        _cat_color = {}
        for row in all_rows:
            ck = row.get('cat', '').strip().lower()
            if ck and ck not in _cat_color:
                _cat_color[ck] = _CAT_PALETTE[len(_cat_color) % len(_CAT_PALETTE)]

        hdr = '<th>Bin</th><th>Category</th><th>Description</th>'
        for ri, r in enumerate(valid):
            clr = _ID_COLORS[ri % len(_ID_COLORS)]
            hdr += (f'<th style="background:{clr};color:#fff;font-weight:bold;padding:5px 8px">'
                    f'{_esc(r["name"])}<br><span style="font-size:11px;font-weight:normal">Yield/Fail%</span></th>')
        # delta headers
        if len(valid) >= 2:
            for ri in range(1, len(valid)):
                hdr += f'<th style="background:#34495e;color:#fff;font-weight:bold;padding:5px 8px;font-size:11px">\u0394 {_esc(valid[ri]["name"])}<br>vs {_esc(valid[0]["name"])}</th>'

        rows_html = ''
        for row in all_rows:
            key  = row['ibin']
            cat  = row.get('cat', '')
            desc = row.get('desc', '')
            row_bg = _cat_color.get(cat.strip().lower(), '#ffffff')
            cells = (f'<td style="background:{row_bg}">{_esc(key)}</td>'
                     f'<td style="background:{row_bg}">{_esc(cat)}</td>'
                     f'<td style="background:{row_bg}">{_esc(desc)}</td>')
            run_vals = []
            for r in valid:
                v = next((ro['fail_pct'] for ro in r['bin_data']['bin_summary_rows']
                          if ro['ibin'] == key), None)
                run_vals.append(v)
                row_all = [next((ro['fail_pct'] for ro in rx['bin_data']['bin_summary_rows'] if ro['ibin'] == key), None) for rx in valid]
                cells += _cell_hl(v, row_all, extra_style=f'background:{row_bg};') + (f'{v:.2f}%' if v is not None else '\u2014') + '</td>'
            # delta cells
            if len(valid) >= 2:
                base = run_vals[0]
                for ri in range(1, len(valid)):
                    v = run_vals[ri]
                    if base is not None and v is not None:
                        delta = v - base
                        sign = '+' if delta > 0 else ''
                        clr = '#c0392b' if delta > 0 else '#27ae60' if delta < 0 else '#555'
                        cells += f'<td class="num" style="color:{clr};font-weight:bold;background:{row_bg}">{sign}{delta:.2f}%</td>'
                    else:
                        cells += f'<td class="num" style="background:{row_bg}">\u2014</td>'
            rows_html += f'<tr>{cells}</tr>\n'
    else:
        valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['bin_fail_rows']]
        if not valid:
            return ''
        best = max(valid, key=lambda r: len(r['bin_data']['bin_fail_rows']))
        all_rows = list({row['ibin']: row for row in best['bin_data']['bin_fail_rows']}.values())

        hdr = '<th>Interface Bin</th><th>Fail Bucket</th>'
        for ri, r in enumerate(valid):
            clr = _ID_COLORS[ri % len(_ID_COLORS)]
            hdr += f'<th style="background:{clr};color:#fff;font-weight:bold;padding:5px 8px">{_esc(r["name"])}</th>'
        # delta headers
        if len(valid) >= 2:
            for ri in range(1, len(valid)):
                hdr += f'<th style="background:#34495e;color:#fff;font-weight:bold;padding:5px 8px;font-size:11px">\u0394 {_esc(valid[ri]["name"])}<br>vs {_esc(valid[0]["name"])}</th>'

        rows_html = ''
        for row in all_rows:
            key   = row['ibin']
            cells = f'<td>{_esc(key)}</td><td>{_esc(row["fail_bucket"])}</td>'
            run_vals = []
            for r in valid:
                v = next((ro['fail_pct'] for ro in r['bin_data']['bin_fail_rows']
                          if ro['ibin'] == key), None)
                run_vals.append(v)
                row_all = [next((ro['fail_pct'] for ro in rx['bin_data']['bin_fail_rows'] if ro['ibin'] == key), None) for rx in valid]
                cells += _cell_hl(v, row_all) + (f'{v:.2f}%' if v is not None else '\u2014') + '</td>'
            # delta cells
            if len(valid) >= 2:
                base = run_vals[0]
                for ri in range(1, len(valid)):
                    v = run_vals[ri]
                    if base is not None and v is not None:
                        delta = v - base
                        sign = '+' if delta > 0 else ''
                        clr = '#c0392b' if delta > 0 else '#27ae60' if delta < 0 else '#555'
                        cells += f'<td class="num" style="color:{clr};font-weight:bold">{sign}{delta:.2f}%</td>'
                    else:
                        cells += '<td class="num">\u2014</td>'
            rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128196; Bin Fail Summary</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl" style="border-collapse:collapse">
    <thead><tr>{hdr}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 4d. SICC / UPM comparison table HTML
# ---------------------------------------------------------------------------

def build_sicc_table_html(runs_data):
    """Table: one row per test, columns = SICC Actual / SICC Target / Multiple / UPM% per run."""
    valid = [r for r in runs_data if r.get('upm_data') and r['upm_data']]
    if not valid:
        return ''

    # Union of test names, ordered by the run with the most rows
    best  = max(valid, key=lambda r: len(r['upm_data']))
    tests = [row['test'] for row in best['upm_data']]

    def _get(run, test, field):
        row = next((r for r in run['upm_data'] if r['test'] == test), None)
        return row[field] if row else None

    def _fmt(v, decimals=3):
        return f'{v:.{decimals}f}' if v is not None else '—'

    # Build column headers — one group (Actual / Target / Multiple) per run
    hdr = '<th rowspan="2" style="min-width:160px">Test</th>'
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th colspan="3" style="background:{clr};color:#fff;font-weight:bold;'
                f'text-align:left;padding:6px 10px">{_esc(r["name"])}</th>')
    # UPM % — one column per run
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th style="background:{clr};color:#fff;font-weight:bold;text-align:left;'
                f'padding:6px 10px">UPM%<br><span style="font-size:11px">{_esc(r["name"])}</span></th>')

    _th_sub = 'style="font-size:11px;white-space:nowrap;text-align:left!important;padding:4px 10px"'
    hdr2 = ''
    for _ in valid:
        hdr2 += (f'<th {_th_sub}>Actual (A)</th>'
                 f'<th {_th_sub}>Target (A)</th>'
                 f'<th {_th_sub}>Multiple</th>')
    for _ in valid:
        hdr2 += f'<th {_th_sub}>UPM (%)</th>'

    rows_html = ''
    for i, test in enumerate(tests):
        bg = '#f9f9f9' if i % 2 == 0 else '#ffffff'
        cells = f'<td style="white-space:nowrap;font-size:13px;background:{bg}">{_esc(test)}</td>'

        # Collect actuals for highlight comparison
        actuals = [_get(r, test, 'sicc_actual') for r in valid]
        for r in valid:
            act  = _get(r, test, 'sicc_actual')
            tgt  = _get(r, test, 'sicc_target')
            mult = _get(r, test, 'multiple')
            # Red only when actual > target (over spec)
            if act is not None and tgt is not None and act > tgt:
                act_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                act_style = f'background:{bg}'
            # Multiple: red bold only when > 1
            if mult is not None and mult > 1:
                mult_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                mult_style = f'background:{bg}'
            _td = 'font-variant-numeric:tabular-nums;padding:4px 10px'
            cells += (f'<td style="{act_style};{_td}">{_fmt(act)}</td>'
                      f'<td style="background:{bg};{_td}">{_fmt(tgt)}</td>'
                      f'<td style="{mult_style};{_td}">{_fmt(mult, 2)}</td>')

        for r in valid:
            upm = _get(r, test, 'upm_pct')
            cells += f'<td style="background:{bg};font-variant-numeric:tabular-nums;padding:4px 10px">{_fmt(upm, 1)}</td>'

        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#9889; SICC / UPM Table</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead>
      <tr>{hdr}</tr>
      <tr>{hdr2}</tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 4e. CDYN comparison chart + table
# ---------------------------------------------------------------------------

def build_cdyn_median_chart(runs_data):
    """Grouped bar chart: CDYN Actual Median vs Expected per test, per run."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('cdyn_data') and r['cdyn_data']]
    if not valid:
        return ''

    best = max(valid, key=lambda r: len(r['cdyn_data']))
    tests = [row for row in best['cdyn_data'] if row.get('actual') is not None]
    if not tests:
        return ''
    test_labels = [row['test'] for row in tests]

    n_tests = len(test_labels)
    n_runs  = len(valid)
    x       = np.arange(n_tests)
    width   = 0.7 / max(n_runs + 1, 2)

    fig, ax = plt.subplots(figsize=(max(7, n_tests * 0.6), 4))

    expected = [t.get('expected') for t in tests]
    has_expected = any(v is not None for v in expected)

    for ri, run in enumerate(valid):
        actuals = []
        for t in tests:
            row = next((r for r in run['cdyn_data'] if r['test'] == t['test']), None)
            actuals.append(row['actual'] if row and row.get('actual') is not None else 0.0)
        offset = ri * width
        ax.bar(x + offset, actuals, width, label=run['name'],
               color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)

    if has_expected:
        exp_vals = [v if v is not None else 0.0 for v in expected]
        offset = n_runs * width
        ax.bar(x + offset, exp_vals, width, label='Expected',
               color='#2c3e50', alpha=0.55, hatch='//')

    tick_offset = (n_runs * width) / 2
    ax.set_xticks(x + tick_offset)
    ax.set_xticklabels(test_labels, rotation=35, ha='right', fontsize=8)
    ax.set_ylabel('CDYN Actual Median (nF)')
    ax.set_title('CDYN Median', fontsize=13, weight='bold')
    ax.yaxis.set_major_locator(plt.MaxNLocator(nbins=12, integer=False))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: '0' if v == 0 else f'{v:.2f}'))
    ax.legend(fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def build_cdyn_table_html(runs_data):
    """Table: one row per CDYN test, columns = Actual / Expected / Actual/Expected per run."""
    valid = [r for r in runs_data if r.get('cdyn_data') and r['cdyn_data']]
    if not valid:
        return ''

    best  = max(valid, key=lambda r: len(r['cdyn_data']))
    tests = [row['test'] for row in best['cdyn_data']]

    def _get(run, test, field):
        row = next((r for r in run['cdyn_data'] if r['test'] == test), None)
        return row[field] if row else None

    def _fmt(v, decimals=2):
        return f'{v:.{decimals}f}' if v is not None else '\u2014'

    # Header row 1 — test name + one group per run
    hdr = '<th rowspan="2" style="min-width:200px">Test</th>'
    hdr += '<th rowspan="2" style="min-width:80px">Type</th>'
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th colspan="3" style="background:{clr};color:#fff;font-weight:bold;'
                f'text-align:left;padding:6px 10px">{_esc(r["name"])}</th>')

    # Header row 2 — sub-headers
    _th_sub = 'style="font-size:11px;white-space:nowrap;text-align:left!important;padding:4px 10px"'
    hdr2 = ''
    for _ in valid:
        hdr2 += (f'<th {_th_sub}>Actual (nF)</th>'
                 f'<th {_th_sub}>Expected (nF)</th>'
                 f'<th {_th_sub}>Actual/Expected</th>')

    rows_html = ''
    for i, test in enumerate(tests):
        bg = '#f9f9f9' if i % 2 == 0 else '#ffffff'
        test_type = _get(best, test, 'type') or ''
        cells = (f'<td style="white-space:nowrap;font-size:13px;background:{bg}">{_esc(test)}</td>'
                 f'<td style="font-size:12px;background:{bg}">{_esc(test_type)}</td>')

        for r in valid:
            act  = _get(r, test, 'actual')
            exp  = _get(r, test, 'expected')
            ratio = _get(r, test, 'ratio')
            # Red when actual > expected
            if act is not None and exp is not None and act > exp:
                act_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                act_style = f'background:{bg}'
            # Ratio red when > 1
            if ratio is not None and ratio > 1:
                ratio_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                ratio_style = f'background:{bg}'
            _td = 'font-variant-numeric:tabular-nums;padding:4px 10px'
            cells += (f'<td style="{act_style};{_td}">{_fmt(act)}</td>'
                      f'<td style="background:{bg};{_td}">{_fmt(exp)}</td>'
                      f'<td style="{ratio_style};{_td}">{_fmt(ratio)}</td>')

        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#9889; CDYN Median Table</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead>
      <tr>{hdr}</tr>
      <tr>{hdr2}</tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 5. Build comparison table HTML
# ---------------------------------------------------------------------------

def build_comparison_table(runs_data, ref_name, col_idx=0):
    if not runs_data:
        return ''
    ref_run = next((r for r in runs_data if r['name'] == ref_name), runs_data[0])

    def _get_val(run, lbl):
        if not run['data']:
            return None
        for g in run['data']['groups']:
            for row in g['rows']:
                if row['label'] == lbl:
                    return row['vals'][col_idx] if col_idx < len(row['vals']) else None
            if all(row['vals'][col_idx] is None for row in g['rows']) is False:
                if g['sum_vals'] and col_idx < len(g['sum_vals']) and g['rows'] and g['rows'][0]['label'] == lbl:
                    pass
        return None

    def _get_sum(run, gi):
        if not run['data'] or gi >= len(run['data']['groups']):
            return None
        sv = run['data']['groups'][gi]['sum_vals']
        return sv[col_idx] if sv and col_idx < len(sv) else None

    def _get_total(run):
        if not run['data'] or not run['data']['totals']:
            return None
        v = run['data']['totals']['vals']
        return v[col_idx] if v and col_idx < len(v) else None

    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    groups = best['data']['groups']

    # Header row
    run_names = [r['name'] for r in runs_data]
    hdr_cells = ('<th style="background:#1a252f;color:#ecf0f1;font-size:18px">'
                 'Sub Module</th>') + ''.join(
        f'<th style="background:{_ID_COLORS[i % len(_ID_COLORS)]};color:#fff;'
        f'font-size:18px;padding:7px 10px">'
        f'{_esc(n)}</th>'
        for i, n in enumerate(run_names)
    )
    # Add delta columns (vs ref)
    compare_names = [r['name'] for r in runs_data if r['name'] != ref_run['name']]
    for cn in compare_names:
        hdr_cells += (
            f'<th style="background:#4a235a;color:#e8daef;text-align:center;'
            f'font-size:18px;padding:7px 10px">'
            f'&#916; vs <b>{_esc(ref_run["name"])}</b>'
            f'<br><span style="font-size:14px;font-weight:normal">{_esc(cn)}</span></th>'
        )

    rows_html = ''
    for gi, group in enumerate(groups):
        grp_clr = _GROUP_COLORS[gi % len(_GROUP_COLORS)]
        for row in group['rows']:
            lbl = row['label']
            vals = [_get_val(r, lbl) for r in runs_data]
            ref_v = _get_val(ref_run, lbl)
            cells = f'<td style="background:{grp_clr};font-size:20px">{_esc(lbl)}</td>'
            for v in vals:
                cells += f'<td class="num" style="background:{grp_clr}">{f"{v:.1f}%" if v is not None else ""}</td>'
            for r in runs_data:
                if r['name'] == ref_run['name']:
                    continue
                cv = _get_val(r, lbl)
                if cv is not None and ref_v is not None:
                    delta = cv - ref_v
                    clr = '#c0392b' if delta > 0 else ('#27ae60' if delta < 0 else '#555')
                    cells += f'<td class="num delta" style="color:{clr};font-weight:bold">{delta:+.1f}%</td>'
                else:
                    cells += '<td class="num delta">—</td>'
            rows_html += f'<tr>{cells}</tr>\n'

        # SUM row
        sum_vals = [_get_sum(r, gi) for r in runs_data]
        ref_sv = _get_sum(ref_run, gi)
        sum_cells = f'<td style="font-weight:bold;background:{grp_clr};border-top:2px solid #999">SUM</td>'
        for sv in sum_vals:
            sum_cells += (f'<td class="num" style="font-weight:bold;background:{grp_clr};'
                          f'border-top:2px solid #999">{f"{sv:.1f}%" if sv is not None else ""}</td>')
        for r in runs_data:
            if r['name'] == ref_run['name']:
                continue
            sv = _get_sum(r, gi)
            if sv is not None and ref_sv is not None:
                delta = sv - ref_sv
                clr = '#c0392b' if delta > 0 else ('#27ae60' if delta < 0 else '#555')
                sum_cells += (f'<td class="num delta" style="color:{clr};font-weight:bold;'
                              f'border-top:2px solid #999">{delta:+.1f}%</td>')
            else:
                sum_cells += '<td class="num delta" style="border-top:2px solid #999">—</td>'
        rows_html += f'<tr>{sum_cells}</tr>\n'

    # Total yield loss row
    if best['data'] and best['data']['totals']:
        tot_lbl = best['data']['totals']['label']
        tot_vals = [_get_total(r) for r in runs_data]
        ref_tv = _get_total(ref_run)
        tot_cells = f'<td style="font-weight:bold;background:#d5e8d4;border-top:3px solid #555">{_esc(tot_lbl)}</td>'
        for tv in tot_vals:
            tot_cells += (f'<td class="num" style="font-weight:bold;background:#d5e8d4;'
                          f'border-top:3px solid #555">{f"{tv:.1f}%" if tv is not None else ""}</td>')
        for r in runs_data:
            if r['name'] == ref_run['name']:
                continue
            tv = _get_total(r)
            if tv is not None and ref_tv is not None:
                delta = tv - ref_tv
                clr = '#c0392b' if delta > 0 else ('#27ae60' if delta < 0 else '#555')
                tot_cells += (f'<td class="num delta" style="color:{clr};font-weight:bold;'
                              f'border-top:3px solid #555">{delta:+.1f}%</td>')
            else:
                tot_cells += '<td class="num delta" style="border-top:3px solid #555">—</td>'
        rows_html += f'<tr>{tot_cells}</tr>\n'

    return f'''
<div class="section">
  <div style="background:#1a252f;border-radius:6px 6px 0 0;padding:10px 16px;
              display:flex;align-items:center;flex-wrap:wrap;gap:8px">
    <span style="font-size:20px;font-weight:bold;color:#ecf0f1">
      &#128202; Detailed Comparison Table
    </span>
    <span style="font-size:15px;color:#aab7b8;margin-left:4px">
      reference:&nbsp;<b style="color:#3498db">{_esc(ref_run["name"])}</b>
      &nbsp;&nbsp;&#124;&nbsp;&nbsp;
      <span style="color:#e74c3c">&#9650; red = more fallout</span>
      &nbsp;&nbsp;
      <span style="color:#2ecc71">&#9660; green = less</span>
    </span>
  </div>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead><tr>{hdr_cells}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 6. Generate full HTML report
# ---------------------------------------------------------------------------

def _run_sort_key(run):
    """Sort key: file mtime of xlsx (most reliable), then ts string, then name."""
    xlsx = run.get('xlsx_path')
    if xlsx:
        try:
            return Path(xlsx).stat().st_mtime
        except Exception:
            pass
    ts = run.get('ts', '')
    # Try common date patterns: YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY, etc.
    import datetime
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y',
                '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %H:%M', '%m/%d/%Y'):
        try:
            return datetime.datetime.strptime(ts.strip(), fmt).timestamp()
        except Exception:
            pass
    return 0.0


def generate_report(runs_data, output_path: Path, ref_name: str = None, config_json: str = None, dash_dir: Path = None):
    # Preserve the order as given (matches Dashboard.html order)
    # runs_data = sorted(runs_data, key=_run_sort_key)  # removed: use Dashboard order

    valid = [r for r in runs_data if r['data']]
    if not valid:
        print('ERROR: No valid xlsx data found for any run.', file=sys.stderr)
        return

    # Load UPM target from config JSON
    _upm_target_pct = None
    _upm_target_label = 'Target'
    if config_json:
        try:
            import json as _json_cr
            _cfg_cr = _json_cr.loads(open(str(config_json), encoding='utf-8').read())
            for _ut in _cfg_cr.get('upm_target', []):
                if _ut.get('target_%') is not None:
                    _upm_target_pct = float(_ut['target_%'])
                    _upm_target_label = _ut.get('test', 'Target')
                    break
        except Exception:
            pass

    # Resolve the reference run name
    if ref_name is None:
        ref_name = valid[0]['name']
    elif not any(r['name'] == ref_name for r in runs_data):
        print(f'WARNING: --ref "{ref_name}" not found; using first run.')
        ref_name = valid[0]['name']

    # Determine number of data columns (Yield Loss + optional Recovery)
    n_cols = max(len(r['data']['col_headers']) for r in valid)
    col_headers = valid[0]['data']['col_headers']

    charts_html = ''

    # --- Yield Information chart ---
    rdnd_valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]

    # --- Extract UPM from raw CSV if Group_Medians.csv not available ---
    for r in runs_data:
        if not r.get('upm_data'):
            output_dir = None
            if r.get('xlsx_path') and str(r['xlsx_path']):
                output_dir = Path(str(r['xlsx_path'])).parent
            if not output_dir:
                href = r.get('index_href', '')
                if href:
                    import re as _re_tmp
                    href = _re_tmp.sub(r'^file:///', '', href).replace('/', os.sep)
                    idx = Path(href) if os.path.isabs(href) else (output_path.parent / href)
                    output_dir = idx.parent
            if output_dir:
                raw_csv = find_raw_csv(output_dir)
                if raw_csv:
                    upm_data, upm_detail = extract_upm_from_csv(raw_csv, config_json=config_json)
                    if upm_data:
                        r['upm_data'] = upm_data
                    if upm_detail:
                        r['upm_detail'] = upm_detail

    upm_valid  = [r for r in runs_data if r.get('upm_data') and r['upm_data']]

    # --- Run Summary table (always shown) ---
    charts_html += build_run_summary_table_html(runs_data, dash_dir=dash_dir)

    if rdnd_valid:
        combined_b64 = build_combined_rdnd_chart(rdnd_valid)
        if combined_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; Yield Information</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + combined_b64 + '"/></div>')

    # --- Yield Table ---
    if rdnd_valid:
        charts_html += build_rdnd_table_html(rdnd_valid)

    # --- Bin Fail Summary ---
    bf_valid = [r for r in runs_data if r.get('bin_data') and (
        r['bin_data'].get('bin_summary_rows') or r['bin_data'].get('bin_fail_rows'))]
    if bf_valid:
        charts_html += build_bin_fail_table_html(bf_valid)

    # --- Top-10 fail pareto ---
    top10_b64 = build_top10_pareto_chart(runs_data)
    if top10_b64:
        charts_html += ('<div class="section">'
                        '<h2>&#128202; Top 10 Interface Bin Fail Pareto</h2>'
                        '<img class="chart" src="data:image/png;base64,'
                        + top10_b64 + '"/></div>')

    # --- SICC/UPM charts + table ---
    if upm_valid:
        upm_b64 = build_upm_median_chart(upm_valid)
        if upm_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; SICC Median</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + upm_b64 + '"/></div>')
        pct_b64 = build_upm_pct_chart(upm_valid, upm_target_pct=_upm_target_pct,
                                         upm_target_label=_upm_target_label)
        if pct_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; UPM ULVT 950mV (%)</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + pct_b64 + '"/></div>')
        # Detailed per-column UPM comparison
        detail_b64 = build_upm_detail_chart(runs_data, upm_target_pct=_upm_target_pct,
                                            upm_target_label=_upm_target_label)
        if detail_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128200; UPM Distribution Comparison</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + detail_b64 + '"/></div>')
        sicc_tbl = build_sicc_table_html(upm_valid)
        if sicc_tbl:
            charts_html += sicc_tbl

    # --- CDYN charts + table ---
    cdyn_valid = [r for r in runs_data if r.get('cdyn_data') and r['cdyn_data']]
    if cdyn_valid:
        cdyn_b64 = build_cdyn_median_chart(cdyn_valid)
        if cdyn_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; CDYN Median</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + cdyn_b64 + '"/></div>')
        cdyn_tbl = build_cdyn_table_html(cdyn_valid)
        if cdyn_tbl:
            charts_html += cdyn_tbl

    # --- *_out.xlsx digital dashboard (at end) ---
    charts_html += build_xlsx_comparison_table(runs_data)

    # Detailed Comparison Table removed

    # Run summary cards — show all runs, including those with no xlsx
    card_html = ''
    for ri, run in enumerate(runs_data):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        if run.get('data'):
            total = run['data']['totals']['vals'][0] if run['data']['totals'] else None
            total_str = f'{total:.1f}%' if total is not None else '—'
            num_die = run['data']['num_die']
            die_str = f'{int(num_die):,}'
            # FF+DF yield (bin 1/2/3/4) and FF yield (bin 1/2)
            yrows = run['bin_data']['yield_rows'] if run.get('bin_data') else []
            ffdf_row = next((r for r in yrows if r['bin'] == '1/2/3/4'), None)
            ff_row   = next((r for r in yrows if r['bin'] == '1/2'), None)
            ffdf_str = f"{ffdf_row['yield_pct']:.1f}%" if ffdf_row and ffdf_row.get('yield_pct') is not None else '—'
            ff_str   = f"{ff_row['yield_pct']:.1f}%"  if ff_row  and ff_row.get('yield_pct')  is not None else '—'
        else:
            total_str = 'N/A'
            die_str   = 'N/A'
            ffdf_str  = 'N/A'
            ff_str    = 'N/A'
        card_html += f'''
<div class="run-card" style="border-left:4px solid {clr}">
  <div class="run-card-name" style="color:{clr}">{_esc(run["name"])}</div>
  <div class="run-card-ts">{_esc(run["ts"])}</div>
  <div class="run-card-stat">Die: <b>{die_str}</b></div>
  <div class="run-card-stat">FF + DF Yield: <b>{ffdf_str}</b></div>
  <div class="run-card-stat">FF Yield: <b>{ff_str}</b></div>
</div>'''

    html = f'''<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>Test Program Comparison Report</title>
<style>
.dash-link{{font-size:21px;color:#2980b9;margin-bottom:4px;display:block}}
.dash-link a{{color:#2980b9;text-decoration:none}}
.dash-link a:hover{{text-decoration:underline}}
h1{{font-size:33px;color:#2c3e50;margin-bottom:6px}}
h2{{font-size:26px;color:#2c3e50;margin:20px 0 8px;padding-bottom:4px;border-bottom:2px solid #dce1e7}}
h3{{font-size:23px;color:#555;margin:10px 0 4px}}
.subtitle{{font-size:21px;color:#7f8c8d;margin-bottom:18px}}
.ref-note{{font-size:20px;font-weight:normal;color:#7f8c8d;margin-left:8px}}
.section{{background:#fff;border-radius:8px;padding:18px;margin-bottom:20px;
  box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.cards{{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:20px}}
.run-card{{background:#fff;border-radius:6px;padding:12px 16px;min-width:200px;
  box-shadow:0 1px 4px rgba(0,0,0,.08);flex:1}}
.run-card-name{{font-size:23px;font-weight:bold;margin-bottom:4px}}
.run-card-ts{{font-size:18px;color:#95a5a6;margin-bottom:6px}}
.run-card-stat{{font-size:21px;margin-bottom:2px}}
.run-card-src{{font-size:17px;color:#aaa;margin-top:6px;word-break:break-all}}
.chart{{width:100%;height:100%;display:block;object-fit:contain}}
.chart-img-wrap{{display:inline-block;width:100%;min-width:300px;min-height:120px;
  resize:both;overflow:hidden;box-sizing:border-box;
  border:1px solid #dde;border-radius:4px;margin:8px 0;cursor:se-resize}}
.cmp-tbl{{border-collapse:collapse;font-size:20px;width:auto}}
.cmp-tbl th{{background:#34495e;color:#ecf0f1;padding:6px 10px;text-align:left;
  white-space:nowrap;font-size:18px}}
.cmp-tbl thead tr:nth-child(2) th{{text-align:right;font-size:13px;font-weight:normal;
  background:#4a5568;color:#e2e8f0;padding:3px 10px}}
.cmp-tbl td{{padding:4px 10px;border-bottom:1px solid #eee;white-space:nowrap}}
.cmp-tbl tr:hover td{{background:#f9f9f9}}
td.num{{text-align:right;font-variant-numeric:tabular-nums}}
td.delta{{text-align:right;font-variant-numeric:tabular-nums}}
th.delta-hdr{{background:#4a235a;color:#e8daef;text-align:center}}
</style>
</head>
<body>
<h1>&#128200; Test Program Comparison Report</h1>
<div class="dash-link">Source: <a href="Dashboard.html">Dashboard.html</a></div>
<div class="subtitle">
  Runs: <b>{len(runs_data)}</b> &nbsp;|&nbsp; With data: <b>{len(valid)}</b>
  &nbsp;|&nbsp; Reference: <b>{_esc(ref_name)}</b>
</div>
<div class="cards">{card_html}</div>
{charts_html}
<script>
// wrap each static PNG chart in a user-resizable div
document.querySelectorAll('img.chart').forEach(function(img){{
  var w=document.createElement('div');
  w.className='chart-img-wrap';
  w.style.height=(img.naturalHeight||img.height||400)+'px';
  img.parentNode.insertBefore(w,img);
  w.appendChild(img);
}});
</script>
</body>
</html>'''

    if output_path:
        output_path.write_text(_wm_inject(html), encoding='utf-8')
        print(f'Wrote comparison report: {output_path}')
    return html


# ---------------------------------------------------------------------------
# 7. Scan for compare HTML files and update links block in Dashboard.html
# ---------------------------------------------------------------------------

_CMP_PATTERNS = ('compare_*.html', 'compareTP*.html', '*_compare.html',
                 '*_comparison.html', 'compare-*.html')


def find_compare_files(dash_dir: Path) -> list[Path]:
    """Return all compare-report HTML files next to Dashboard.html, sorted by name.
    Case-insensitive: scans all .html files and matches any whose stem contains 'compare'."""
    name_re = re.compile(r'compare', re.IGNORECASE)
    results = []
    for p in sorted(dash_dir.glob('*.html')):
        if p.name.lower() == 'dashboard.html':
            continue
        if name_re.search(p.stem):
            results.append(p)
    return results


def update_dashboard_compare_links(dash_path: Path) -> None:
    """
    Scan the dashboard directory for compare HTML files and rewrite the
    <!-- COMPARISON_REPORT_START/END --> block in Dashboard.html with links.
    Safe to call after any compareTP / compare_runs run.
    """
    dash_dir = dash_path.parent
    files = find_compare_files(dash_dir)

    if not files:
        # Nothing to show — remove the block if present
        dash_html = dash_path.read_text(encoding='utf-8')
        dash_html = re.sub(
            r'\n*<!-- COMPARISON_REPORT_START -->[\s\S]*?<!-- COMPARISON_REPORT_END -->\n*',
            '', dash_html
        )
        dash_path.write_text(dash_html, encoding='utf-8')
        print('No compare files found — removed comparison block from Dashboard.html.')
        return

    link_items = ''.join(
        f'<a class="run-link report-link" href="{p.name}" target="_blank">'
        f'&#128200; {p.stem}</a>\n'
        for p in files
    )

    inject = (
        '\n\n<!-- COMPARISON_REPORT_START -->\n'
        '<div id="_cmp_embed" style="margin-top:18px;padding:10px 14px;background:#2c3e50;'
        'border-radius:6px;display:flex;flex-wrap:wrap;gap:6px;align-items:center">\n'
        '<span style="color:#ecf0f1;font-size:13px;font-weight:bold;margin-right:4px">'
        '&#128200; Comparison Reports:</span>\n'
        + link_items +
        '</div>\n'
        '<!-- COMPARISON_REPORT_END -->\n'
    )

    dash_html = dash_path.read_text(encoding='utf-8')

    # Remove any previous comparison block
    dash_html = re.sub(
        r'\n*<!-- COMPARISON_REPORT_START -->[\s\S]*?<!-- COMPARISON_REPORT_END -->\n*',
        '', dash_html
    )

    # Add ▼ Compare jump link to <h1> once
    jump = (' <a href="#_cmp_embed" style="font-size:15px;color:#2980b9;'
            'text-decoration:none;margin-left:14px;vertical-align:middle">'
            '&#9660; Compare</a>')
    if '#_cmp_embed' not in dash_html:
        dash_html = re.sub(
            r'(<h1[^>]*>)(.*?)(</h1>)',
            lambda m: m.group(1) + m.group(2) + jump + m.group(3),
            dash_html, count=1, flags=re.DOTALL
        )

    # Insert just before <!-- RUNS_END -->
    if '<!-- RUNS_END -->' in dash_html:
        dash_html = dash_html.replace('<!-- RUNS_END -->', inject + '<!-- RUNS_END -->', 1)
    elif '</body>' in dash_html:
        dash_html = dash_html.replace('</body>', inject + '</body>', 1)
    else:
        dash_html += inject

    dash_path.write_text(dash_html, encoding='utf-8')
    print(f'Updated {dash_path.name} with {len(files)} comparison link(s).')


# ---------------------------------------------------------------------------
# 8. Main
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(description='Compare yield runs from Dashboard.html')
    p.add_argument('dashboard', help='Path to Dashboard.html')
    p.add_argument('--out', default='', help='Output HTML path (default: next to Dashboard.html)')
    p.add_argument('--ref', default='', help='Reference run name for comparison (default: first run)')
    args = p.parse_args()

    if not HAVE_OPENPYXL:
        print('ERROR: openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
        sys.exit(1)
    if not HAVE_MPL:
        print('WARNING: matplotlib not installed — charts will be skipped.')

    dash_path = Path(args.dashboard).resolve()
    if not dash_path.exists():
        print(f'ERROR: Dashboard.html not found: {dash_path}', file=sys.stderr)
        sys.exit(1)

    dash_dir = dash_path.parent

    print(f'Parsing {dash_path} ...')
    run_records = parse_dashboard(dash_path)
    if not run_records:
        print('ERROR: No run blocks found in Dashboard.html', file=sys.stderr)
        sys.exit(1)

    print(f'Found {len(run_records)} run(s): {[r["name"] for r in run_records]}')

    # Load xlsx + BinDistribution HTML for each run
    runs_data = []
    for rec in run_records:
        xlsx_p = find_xlsx(dash_dir, rec['index_href'])
        data = None
        if xlsx_p:
            print(f'  [{rec["name"]}] Reading {xlsx_p.name} ...')
            data = read_xlsx(xlsx_p)
            if not data:
                print(f'    WARNING: Could not parse {xlsx_p}')
            output_dir = xlsx_p.parent
        else:
            print(f'  [{rec["name"]}] WARNING: *_out.xlsx not found')
            # Try to resolve output_dir from index_href even without xlsx
            href = re.sub(r'^file:///', '', rec['index_href'] or '').replace('/', os.sep)
            idx_path = dash_dir / href if not os.path.isabs(href) else Path(href)
            output_dir = idx_path.parent if idx_path else None

        bin_data = None
        upm_data = None
        cdyn_data = None
        if output_dir and output_dir.exists():
            bin_p = find_bin_html(output_dir)
            if bin_p:
                print(f'  [{rec["name"]}] Reading {bin_p.name} ...')
                bin_data = parse_bin_html(bin_p)
            else:
                print(f'  [{rec["name"]}] WARNING: *_BinDistribution.html not found')
            gm_p = find_group_medians(output_dir)
            if gm_p:
                print(f'  [{rec["name"]}] Reading {gm_p.name} ...')
                upm_data = parse_group_medians(gm_p)
            cdyn_p = find_cdyn_medians(output_dir)
            if cdyn_p:
                print(f'  [{rec["name"]}] Reading {cdyn_p.name} ...')
                cdyn_data = parse_cdyn_medians(cdyn_p)

        runs_data.append({**rec, 'data': data, 'xlsx_path': xlsx_p or '',
                          'bin_data': bin_data, 'upm_data': upm_data,
                          'cdyn_data': cdyn_data})

    ref_name = args.ref.strip() or None
    out_path = Path(args.out).resolve() if args.out else dash_dir / 'compare_report.html'
    out_path = _safe_html_out_path(out_path, 'compare_report.html')

    print('Generating comparison report ...')
    # Find Product Config JSON in collateral/ folder
    _cfg_json = None
    _collateral = dash_dir / 'collateral'
    if _collateral.exists():
        _cfgs = sorted(_collateral.glob('Product Config*.json'),
                       key=lambda p: p.stat().st_mtime, reverse=True)
        if _cfgs:
            _cfg_json = str(_cfgs[0])
            print(f'  Config: {_cfgs[0].name}')
    generate_report(runs_data, out_path, ref_name=ref_name, config_json=_cfg_json,
                    dash_dir=dash_dir)

    # Update comparison links in Dashboard.html
    update_dashboard_compare_links(dash_path, out_path)

    # Open report in browser
    try:
        os.startfile(str(out_path))
    except Exception:
        pass


def update_dashboard_compare_links(dash_path: Path, compare_report_path: Path):
    """Inject a link to compare_report_path into the <!-- COMPARE_START/END --> section
    of Dashboard.html, using the same run-block style as the Yield section."""
    from datetime import datetime as _dt
    dash_path = Path(dash_path)
    compare_report_path = Path(compare_report_path)
    if not dash_path.exists():
        return

    content = dash_path.read_text(encoding='utf-8')

    COMPARE_START = '<!-- COMPARE_START -->'
    COMPARE_END   = '<!-- COMPARE_END -->'
    YIELD_END     = '<!-- YIELD_END -->'

    # Ensure COMPARE section exists
    if COMPARE_START not in content:
        if YIELD_END in content:
            content = content.replace(
                YIELD_END,
                YIELD_END + '\n<h2 class="section-header">&#128200; Compare TP</h2>\n'
                + COMPARE_START + '\n' + COMPARE_END)
        else:
            # No sentinels at all — append before </body>
            content = content.replace(
                '</body>',
                '<h2 class="section-header">&#128200; Compare TP</h2>\n'
                + COMPARE_START + '\n' + COMPARE_END + '\n</body>')

    # Build relative href from dash_dir to compare report
    try:
        href = os.path.relpath(str(compare_report_path), str(dash_path.parent)).replace('\\', '/')
    except Exception:
        href = compare_report_path.as_uri()

    report_stem = compare_report_path.stem
    ts = _dt.now().strftime('%Y-%m-%d %H:%M')

    # Replace existing block with same stem, or prepend new one
    block_key = report_stem
    new_block = (
        f'<div class="run-block" data-stem="{block_key}">\n'
        f'<div class="run-header" onclick="toggle(this)">'
        f'<span class="arrow">&#9660;</span> {report_stem}'
        f'<span class="ts"> - {ts}</span></div>\n'
        f'<div class="run-body">\n'
        f'<a class="run-link report-link" href="{href}" target="_blank">{report_stem}</a>\n'
        f'</div>\n</div>'
    )

    block_re = re.compile(
        r'<div class="run-block" data-stem="' + re.escape(block_key) +
        r'">\s*<div[^>]*>[\s\S]*?</div>\s*</div>', re.MULTILINE)
    if block_re.search(content):
        content = block_re.sub(new_block, content)
    else:
        content = content.replace(COMPARE_START, COMPARE_START + '\n' + new_block)

    dash_path.write_text(content, encoding='utf-8')


if __name__ == '__main__':
    main()


# ════════════════════════════════════════════════════════════════
# (formerly compareTP.py)
# ════════════════════════════════════════════════════════════════
import os
import sys
import threading
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox

sys.path.insert(0, str(Path(__file__).parent))

# ── Palette (same as dashboard.py) ──────────────────────────────────────────
BG   = '#1a252f'
BG2  = '#2c3e50'
FG   = '#ecf0f1'
FG2  = '#95a5a6'
BLUE = '#2980b9'
ABLU = '#3498db'
GRN  = '#27ae60'
AGRN = '#2ecc71'


def _btn(parent, text, cmd, color=BLUE, acolor=ABLU, width=None):
    kw = dict(text=text, command=cmd, bg=color, fg='white',
              activebackground=acolor, activeforeground='white',
              relief='flat', cursor='hand2', font=('Arial', 9),
              padx=8, pady=3)
    if width:
        kw['width'] = width
    return tk.Button(parent, **kw)


def _lf(parent, text, label_color=FG2):
    f = tk.LabelFrame(parent, text=text, bg=BG, fg=label_color,
                      font=('Arial', 8, 'bold'), padx=6, pady=4,
                      relief='groove', bd=1)
    return f


# ---------------------------------------------------------------------------
# Main window
# ---------------------------------------------------------------------------

class CompareFrame(tk.Frame):
    def __init__(self, parent=None, **kw):
        super().__init__(parent, bg=BG, **kw)

        self._dash_path  = tk.StringVar()
        self._out_var    = tk.StringVar()
        self._run_records = []
        self._check_vars  = []
        self._last_report_path = ''

        self._build_ui()

    # ------------------------------------------------------------------ UI --

    def _build_ui(self):
        P = {'padx': 10, 'pady': 4}

        # ── Title ─────────────────────────────────────────────────────────────
        tk.Label(self, text='TestProgram Compare Tool',
                 bg=BG, fg=ABLU, font=('Arial', 13, 'bold')
                 ).pack(fill='x', padx=10, pady=(8, 2))

        # ── Step 1 ────────────────────────────────────────────────────────────
        frm1 = _lf(self, 'Step 1 — Dashboard.html', ABLU)
        frm1.pack(fill='x', **P)

        entry_row = tk.Frame(frm1, bg=BG)
        entry_row.pack(fill='x')
        tk.Entry(entry_row, textvariable=self._dash_path, width=52,
                 bg=BG2, fg=FG, insertbackground=FG,
                 relief='flat', font=('Consolas', 9)
                 ).pack(side='left', padx=(0, 4), pady=2, expand=True, fill='x')
        _btn(entry_row, 'Browse…', self._browse).pack(side='left', padx=(0, 4))
        _btn(entry_row, 'Load',    self._load,  color='#1f618d').pack(side='left')

        # ── Step 2 ────────────────────────────────────────────────────────────
        frm2 = _lf(self, 'Step 2 — Select identifiers to compare', '#9b59b6')
        frm2.pack(fill='both', expand=True, **P)

        sel_row = tk.Frame(frm2, bg=BG)
        sel_row.pack(fill='x', pady=(2, 4))
        _btn(sel_row, 'Select all',   self._sel_all,  color='#1f618d').pack(side='left', padx=(0, 4))
        _btn(sel_row, 'Deselect all', self._sel_none, color='#6d3b01').pack(side='left')
        tk.Label(sel_row, text='Use ↑↓ to set column order in report',
                 bg=BG, fg=FG2, font=('Arial', 8)).pack(side='left', padx=(12, 0))

        list_outer = tk.Frame(frm2, bg=BG2, relief='flat', bd=1)
        list_outer.pack(fill='both', expand=True)

        self._canvas = tk.Canvas(list_outer, bg=BG2, borderwidth=0,
                                 highlightthickness=0)
        vsb = tk.Scrollbar(list_outer, orient='vertical',
                           command=self._canvas.yview,
                           bg=BG2, troughcolor=BG)
        self._canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        self._canvas.pack(side='left', fill='both', expand=True)

        self._list_inner = tk.Frame(self._canvas, bg=BG2)
        self._canvas_win = self._canvas.create_window(
            (0, 0), window=self._list_inner, anchor='nw')
        self._list_inner.bind('<Configure>', self._on_inner_configure)
        self._canvas.bind('<Configure>', self._on_canvas_configure)

        # ── Step 3 ────────────────────────────────────────────────────────────
        frm3 = _lf(self, 'Step 3 — Output', FG2)
        frm3.pack(fill='x', **P)

        out_row = tk.Frame(frm3, bg=BG)
        out_row.pack(fill='x')
        tk.Label(out_row, text='Output file:', bg=BG, fg=FG,
                 font=('Arial', 9), width=11, anchor='w').pack(side='left')
        tk.Entry(out_row, textvariable=self._out_var, width=46,
                 bg=BG2, fg=FG, insertbackground=FG,
                 relief='flat', font=('Consolas', 9)
                 ).pack(side='left', padx=(0, 4), expand=True, fill='x')
        _btn(out_row, '…', self._browse_out, width=3).pack(side='left')

        # ── Action buttons ────────────────────────────────────────────────────
        btn_row = tk.Frame(self, bg=BG)
        btn_row.pack(pady=(6, 2), padx=10, fill='x')
        self._run_btn = _btn(btn_row, '▶  Generate Report', self._generate,
                             color=GRN, acolor=AGRN)
        self._run_btn.config(font=('Arial', 10, 'bold'), pady=5)
        self._run_btn.pack(side='left', expand=True, fill='x', padx=(0, 4))
        self._open_btn = _btn(btn_row, '  Open Dashboard  ', self._open_dashboard,
                              color='#935116', acolor='#ca6f1e')
        self._open_btn.config(font=('Arial', 10, 'bold'), pady=5, state='disabled')
        self._open_btn.pack(side='left')

        # ── Log ───────────────────────────────────────────────────────────────
        log_frm = _lf(self, 'Log', FG2)
        log_frm.pack(fill='both', expand=False, **P)
        self._log = tk.Text(log_frm, height=6, state='disabled',
                            font=('Consolas', 8), bg='#0d1b26', fg='#a8d8ea',
                            relief='flat', insertbackground=FG)
        self._log.pack(fill='both', expand=True)

    # ---------------------------------------------------------------- events --

    def _on_inner_configure(self, _evt=None):
        self._canvas.configure(scrollregion=self._canvas.bbox('all'))

    def _on_canvas_configure(self, evt):
        self._canvas.itemconfig(self._canvas_win, width=evt.width)

    def _browse(self):
        p = filedialog.askopenfilename(
            title='Select Dashboard.html',
            filetypes=[('HTML files', '*.html'), ('All files', '*.*')])
        if p:
            self._dash_path.set(p)
            self._load()

    def _browse_out(self):
        p = filedialog.asksaveasfilename(
            title='Save report as',
            defaultextension='.html',
            filetypes=[('HTML files', '*.html')])
        if p:
            self._out_var.set(p)

    def _sel_all(self):
        for v in self._check_vars:
            v.set(True)

    def _sel_none(self):
        for v in self._check_vars:
            v.set(False)

    # ------------------------------------------------------------------ load --

    def _load(self):
        path_str = self._dash_path.get().strip()
        if not path_str:
            return
        dash = Path(path_str)
        if not dash.exists():
            messagebox.showerror('Not found', f'File not found:\n{dash}')
            return
        try:
            records = parse_dashboard(dash)
        except Exception as exc:
            messagebox.showerror('Parse error', str(exc))
            return
        if not records:
            messagebox.showwarning('No runs', 'No run blocks found in Dashboard.html')
            return

        self._run_records = records
        self._log_write(f'Loaded {len(records)} identifier(s) from {dash.name}\n')
        self._open_btn.configure(state='normal')

        for w in self._list_inner.winfo_children():
            w.destroy()
        self._check_vars = []

        for i, rec in enumerate(records):
            var = tk.BooleanVar(value=True)
            self._check_vars.append(var)
            row_bg = BG2 if i % 2 == 0 else '#253545'
            row = tk.Frame(self._list_inner, bg=row_bg)
            row.pack(fill='x')
            tk.Label(row, text=f'Col {i+1}', bg=row_bg, fg='#7fb3d3',
                     font=('Arial', 8), width=5).pack(side='left', padx=(4, 0))
            tk.Button(row, text='↑', command=lambda idx=i: self._move_up(idx),
                      bg=BG, fg=FG, relief='flat', font=('Arial', 8),
                      padx=2, pady=0, cursor='hand2').pack(side='left')
            tk.Button(row, text='↓', command=lambda idx=i: self._move_down(idx),
                      bg=BG, fg=FG, relief='flat', font=('Arial', 8),
                      padx=2, pady=0, cursor='hand2').pack(side='left', padx=(0, 4))
            tk.Checkbutton(row, variable=var, bg=row_bg, fg=FG,
                           selectcolor=BG, activebackground=row_bg,
                           activeforeground=FG, relief='flat',
                           font=('Arial', 9),
                           text=rec['name']).pack(side='left', padx=2, pady=2)
            ts = rec.get('ts', '')
            if ts:
                tk.Label(row, text=ts, bg=row_bg, fg=FG2,
                         font=('Arial', 8)).pack(side='right', padx=8)

        self._out_var.set(str(dash.parent / 'compare_report.html'))

    def _move_up(self, idx):
        if idx <= 0 or idx >= len(self._run_records):
            return
        self._run_records[idx-1], self._run_records[idx] = self._run_records[idx], self._run_records[idx-1]
        self._check_vars[idx-1], self._check_vars[idx] = (
            tk.BooleanVar(value=self._check_vars[idx].get()),
            tk.BooleanVar(value=self._check_vars[idx-1].get()))
        self._rebuild_list()

    def _move_down(self, idx):
        if idx < 0 or idx >= len(self._run_records) - 1:
            return
        self._run_records[idx], self._run_records[idx+1] = self._run_records[idx+1], self._run_records[idx]
        self._check_vars[idx], self._check_vars[idx+1] = (
            tk.BooleanVar(value=self._check_vars[idx+1].get()),
            tk.BooleanVar(value=self._check_vars[idx].get()))
        self._rebuild_list()

    def _rebuild_list(self):
        for w in self._list_inner.winfo_children():
            w.destroy()
        new_vars = []
        for i, (rec, var) in enumerate(zip(self._run_records, self._check_vars)):
            new_var = tk.BooleanVar(value=var.get())
            new_vars.append(new_var)
            row_bg = BG2 if i % 2 == 0 else '#253545'
            row = tk.Frame(self._list_inner, bg=row_bg)
            row.pack(fill='x')
            tk.Label(row, text=f'Col {i+1}', bg=row_bg, fg='#7fb3d3',
                     font=('Arial', 8), width=5).pack(side='left', padx=(4, 0))
            tk.Button(row, text='↑', command=lambda idx=i: self._move_up(idx),
                      bg=BG, fg=FG, relief='flat', font=('Arial', 8),
                      padx=2, pady=0, cursor='hand2').pack(side='left')
            tk.Button(row, text='↓', command=lambda idx=i: self._move_down(idx),
                      bg=BG, fg=FG, relief='flat', font=('Arial', 8),
                      padx=2, pady=0, cursor='hand2').pack(side='left', padx=(0, 4))
            tk.Checkbutton(row, variable=new_var, bg=row_bg, fg=FG,
                           selectcolor=BG, activebackground=row_bg,
                           activeforeground=FG, relief='flat',
                           font=('Arial', 9),
                           text=rec['name']).pack(side='left', padx=2, pady=2)
            ts = rec.get('ts', '')
            if ts:
                tk.Label(row, text=ts, bg=row_bg, fg=FG2,
                         font=('Arial', 8)).pack(side='right', padx=8)
        self._check_vars = new_vars

    # ------------------------------------------------------------- generate --

    def _generate(self):
        if not self._run_records:
            messagebox.showwarning('No data', 'Load a Dashboard.html first.')
            return
        selected = [rec for rec, var in zip(self._run_records, self._check_vars)
                    if var.get()]
        if len(selected) < 1:
            messagebox.showwarning('No selection', 'Select at least 1 identifier.')
            return

        out_path  = Path(self._out_var.get().strip() or
                         Path(self._dash_path.get()).parent / 'compare_report.html')
        out_path  = _safe_html_out_path(out_path, 'compare_report.html')
        dash_path = Path(self._dash_path.get())

        self._run_btn.configure(state='disabled', text='Working…', bg=FG2)
        self._log_write('Loading data…\n')

        def _worker():
            try:
                dash_dir  = dash_path.parent
                runs_data = []
                for rec in selected:
                    xlsx_p     = find_xlsx(dash_dir, rec['index_href'])
                    data       = None
                    output_dir = None
                    if xlsx_p:
                        self._log_write(f'  [{rec["name"]}] {xlsx_p.name}\n')
                        data       = read_xlsx(xlsx_p)
                        output_dir = xlsx_p.parent
                    else:
                        self._log_write(f'  [{rec["name"]}] no xlsx found\n')
                        import re as _re
                        href = _re.sub(r'^file:///', '', rec['index_href'] or '').replace('/', os.sep)
                        idx  = dash_dir / href if not os.path.isabs(href) else Path(href)
                        output_dir = idx.parent if idx else None

                    bin_data = None
                    upm_data = None
                    cdyn_data = None
                    if output_dir and output_dir.exists():
                        bin_p = find_bin_html(output_dir)
                        if bin_p:
                            self._log_write(f'  [{rec["name"]}] {bin_p.name}\n')
                            bin_data = parse_bin_html(bin_p)
                        gm_p = find_group_medians(output_dir)
                        if gm_p:
                            self._log_write(f'  [{rec["name"]}] {gm_p.name}\n')
                            upm_data = parse_group_medians(gm_p)
                        cdyn_p = find_cdyn_medians(output_dir)
                        if cdyn_p:
                            self._log_write(f'  [{rec["name"]}] {cdyn_p.name}\n')
                            cdyn_data = parse_cdyn_medians(cdyn_p)

                    # Fallback: extract UPM from raw CSV if Group_Medians not found
                    upm_detail = None
                    if not upm_data and output_dir and output_dir.exists():
                        raw_csv = find_raw_csv(output_dir)
                        if raw_csv:
                            # Find config JSON early for extraction
                            _cfg_tmp = None
                            try:
                                _coll = dash_path.parent / 'collateral'
                                if _coll.exists():
                                    _cfs = sorted(_coll.glob('Product Config*.json'),
                                                  key=lambda p: p.stat().st_mtime, reverse=True)
                                    if _cfs:
                                        _cfg_tmp = str(_cfs[0])
                            except Exception:
                                pass
                            upm_data, upm_detail = extract_upm_from_csv(raw_csv, config_json=_cfg_tmp)
                            if upm_data:
                                self._log_write(f'  [{rec["name"]}] UPM from {raw_csv.name}\n')

                    runs_data.append({**rec, 'data': data,
                                      'xlsx_path': xlsx_p or '',
                                      'bin_data': bin_data,
                                      'upm_data': upm_data,
                                      'upm_detail': upm_detail,
                                      'cdyn_data': cdyn_data})

                self._log_write('Generating report…\n')
                # Find Product Config JSON in collateral/ folder
                _cfg_json = None
                try:
                    _collateral = dash_path.parent / 'collateral'
                    if _collateral.exists():
                        _cfgs = sorted(_collateral.glob('Product Config*.json'),
                                       key=lambda p: p.stat().st_mtime, reverse=True)
                        if _cfgs:
                            _cfg_json = str(_cfgs[0])
                            self._log_write(f'  Config: {_cfgs[0].name}\n')
                except Exception:
                    pass
                generate_report(runs_data, out_path, config_json=_cfg_json,
                                   dash_dir=dash_path.parent)
                self._log_write(f'Done → {out_path}\n')
                # Update comparison links in Dashboard.html
                try:
                    update_dashboard_compare_links(dash_path, out_path)
                    self._log_write(f'Updated {dash_path.name} with compare links.\n')
                except Exception as e:
                    self._log_write(f'Warning: could not update Dashboard.html: {e}\n')
                self._last_report_path = str(out_path)
                self.after(0, lambda: self._open_btn.configure(state='normal'))
            except Exception as exc:
                self._log_write(f'ERROR: {exc}\n')
            finally:
                self.after(0, lambda: self._run_btn.configure(
                    state='normal', text='▶  Generate Report', bg=GRN))

        threading.Thread(target=_worker, daemon=True).start()

    # --------------------------------------------------------------- open --

    def _open_dashboard(self):
        p = self._dash_path.get().strip()
        if p and os.path.isfile(p):
            try:
                os.startfile(p)
            except Exception as exc:
                messagebox.showerror('Error', str(exc))
        else:
            messagebox.showwarning('Not found', 'Dashboard.html not found. Load a Dashboard.html first.')

    # ------------------------------------------------------------------- log --

    def _log_write(self, msg: str):
        def _do():
            self._log.configure(state='normal')
            self._log.insert('end', msg)
            self._log.see('end')
            self._log.configure(state='disabled')
        self.after(0, _do)


# Keep standalone entrypoint
class CompareGUI(tk.Tk):
    """Standalone wrapper — embeds CompareFrame in a Tk root window."""
    def __init__(self):
        super().__init__()
        self.title('TestProgram Compare Tool')
        self.resizable(True, True)
        self.minsize(620, 540)
        frame = CompareFrame(self)
        frame.pack(fill='both', expand=True)


# ---------------------------------------------------------------------------
if __name__ == '__main__':
    CompareGUI().mainloop()


# ════════════════════════════════════════════════════════════════
# (formerly trend_chart.py)
# ════════════════════════════════════════════════════════════════

import sys
import os
import re
import json
import argparse
from pathlib import Path
from datetime import datetime
from collections import OrderedDict
from typing import Any

# Plotly
try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    HAVE_PLOTLY = True
except ImportError:
    HAVE_PLOTLY = False

INTERVALS = ['revision', 'weekly', 'bi-weekly', 'monthly']

_PASS_BINS  = {1, 2, 3, 4}
_FF_BINS    = {1, 2}
_FF_DF_BINS = {1, 2, 3, 4}

def _safe_html_out_path(out_path: Path, default_name: str) -> Path:
    """Coerce out_path to a writable .html file path — if it's an existing
    directory or missing the .html suffix, append/replace with default_name."""
    out_path = Path(out_path)
    if out_path.is_dir():
        return out_path / default_name
    if out_path.suffix.lower() != '.html':
        return out_path.parent / default_name
    return out_path

_FAIL_PALETTE = [
    '#E53935', '#1E88E5', '#43A047', '#FB8C00', '#8E24AA',
    '#00ACC1', '#F4511E', '#3949AB', '#00897B', '#FFB300',
    '#D81B60', '#039BE5', '#7CB342', '#6D4C41', '#546E7A',
    '#C62828', '#283593', '#2E7D32', '#E65100', '#4A148C',
]


# ============================================================================
# 1. Product config helpers
# ============================================================================

def load_product_config(cfg_path: str | Path) -> dict[str, Any]:
    """Load product config JSON, return dict with ibin_name, ibin_target, yield_target."""
    cfg_path = Path(cfg_path)
    try:
        text = cfg_path.read_text(encoding='utf-8')
    except UnicodeDecodeError:
        text = cfg_path.read_text(encoding='cp1252', errors='replace')
    raw = json.loads(text)

    ibin_name: dict[int, str]    = {}
    ibin_target: dict[int, float] = {}
    yield_target: dict[str, float] = {}

    for entry in raw.get('yield_targets', []):
        bin_str  = str(entry.get('bin', ''))
        label    = entry.get('fail_bucket', '') or ''
        yld_pct  = entry.get('yield')

        ibins = []
        for part in re.split(r'[/,\s]+', bin_str):
            part = part.strip()
            if part.isdigit():
                ibins.append(int(part))

        for ib in ibins:
            if label:
                ibin_name[ib] = label
            if yld_pct is not None:
                ibin_target[ib] = float(yld_pct)

        if bin_str in ('1/2', '1/2/3/4'):
            key = 'ff' if bin_str == '1/2' else 'ff_df'
            if yld_pct is not None:
                yield_target[key] = float(yld_pct)

    # ── Enrich ibin_name from bin_map (desc + cat), overrides yield_targets ─
    for bin_str, info in raw.get('bin_map', {}).items():
        try:
            ib = int(bin_str)
        except (ValueError, TypeError):
            continue
        desc = (info.get('desc') or '').strip()
        cat  = (info.get('cat')  or '').strip()
        if cat or desc:
            ibin_name[ib] = f'{cat} \u2014 {desc}' if (cat and desc and cat != desc) else (cat or desc)

    # Extract series names from fail_bucket labels
    ff_name   = 'SDS FF'
    ff_df_name = 'SDS FF+DF'
    for entry in raw.get('yield_targets', []):
        bin_str = str(entry.get('bin', ''))
        lbl = (entry.get('fail_bucket') or '').strip()
        if not lbl:
            continue
        if bin_str == '1/2':
            ff_name = lbl
        elif bin_str == '1/2/3/4':
            ff_df_name = lbl

    return {
        'ibin_name':    ibin_name,
        'ibin_target':  ibin_target,
        'yield_target': yield_target,
        'name':         raw.get('name', ''),
        'ff_name':      ff_name,
        'ff_df_name':   ff_df_name,
        'raw':          raw,
    }


def _find_repo_root(start: Path) -> Path:
    """Walk up from `start` to the ancestor containing a shared/ dir (robust to repo reshuffles)."""
    current = Path(start).resolve()
    for _ in range(12):
        if (current / "shared").is_dir():
            return current
        parent = current.parent
        if parent == current:
            break
        current = parent
    return current


def _find_auto_config(devrevstep: str = '') -> Path | None:
    """Search shared/setup/yield-dashboard/ for a matching .json.
    If devrevstep is given (e.g. '8PF5CV'), prefer a file whose name starts with it.
    Falls back to 'default' file, then first .json found.
    """
    here = Path(__file__).resolve().parent
    d = _find_repo_root(here) / 'shared' / 'setup' / 'config' / 'yield-dashboard'
    if not d.exists():
        return None
    jsons = sorted(d.glob('*.json'))
    if not jsons:
        return None
    if devrevstep:
        key = devrevstep.upper()
        # Exact prefix match: filename starts with devrevstep
        for p in jsons:
            if p.name.upper().startswith(key):
                return p
    # Fallback: prefer 'default' file
    for p in jsons:
        if p.stem.lower().startswith('default'):
            return p
    return jsons[0]


# ============================================================================
# 2. Date / interval helpers
# ============================================================================

_TS_FMTS = (
    '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
    '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y',
    '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %H:%M', '%m/%d/%Y',
    '%Y%m%d',
)


def _parse_date(s: str) -> datetime | None:
    s = (s or '').strip()
    # Strip timezone offset (+HH:MM or -HH:MM) so strptime formats work
    s = re.sub(r'[+-]\d{2}:\d{2}$', '', s).strip()
    for fmt in _TS_FMTS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    m = re.search(r'(\d{8})', s)
    if m:
        try:
            return datetime.strptime(m.group(1), '%Y%m%d')
        except ValueError:
            pass
    return None


def _interval_key(dt: datetime, interval: str, program: str = '') -> str:
    if interval == 'revision':
        return _prog_label(program) if program else dt.strftime('%Y-%m-%d')
    if interval == 'daily':
        return dt.strftime('%Y-%m-%d')
    if interval == 'weekly':
        iso_y, iso_w, _ = dt.isocalendar()
        return f'{iso_y}-W{iso_w:02d}'
    if interval == 'bi-weekly':
        iso_y, iso_w, _ = dt.isocalendar()
        bw = ((iso_w - 1) // 2) * 2 + 1
        return f'{iso_y}-W{bw:02d}/{bw+1:02d}'
    if interval == 'monthly':
        return dt.strftime('%Y-%m')
    return dt.strftime('%Y-%m-%d')


def _interval_sort_key(label: str) -> tuple:
    # Revision label e.g. '61A', '61B', '102C'
    m = re.match(r'^(\d+)([A-Z])$', label)
    if m:
        return (0, int(m.group(1)), ord(m.group(2)), 0)
    m = re.match(r'^(\d{4})-W(\d+)', label)
    if m:
        return (1, int(m.group(1)), int(m.group(2)), 0)
    m = re.match(r'^(\d{4})-(\d{2})$', label)
    if m:
        return (1, int(m.group(1)), int(m.group(2)), 0)
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', label)
    if m:
        return (1, int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return (9999, 0, 0, 0)


# ============================================================================
# 3. Program label
# ============================================================================

def _prog_label(program_name: str, lot: str = '', wafer: str = '') -> str:
    """Extract a short run label, e.g. '61C' from 'NCXSDJXL0H61C002620'."""
    p = str(program_name).strip()
    m = re.search(r'[A-Z](\d{2}[A-Z])', p)
    if m:
        return m.group(1)
    m = re.search(r'(\d{2,3}[A-Z])', p)
    if m:
        return m.group(1)
    return p[-6:] if len(p) > 6 else (p or (lot[-4:] if lot else '?'))


# ============================================================================
# 4. CSV loading
# ============================================================================

_COL_ALIASES = {
    'date':    ['date', 'run date', 'run_date', 'rundate', 'timestamp', 'ts',
                'lots end date time', 'end date time', 'end_date_time'],
    'lot':     ['lot', 'lot_id', 'lotid', 'lot id'],
    'wafer':   ['wafer', 'wafer_num', 'wafer num', 'waferno', 'wafer no', 'wafer_id',
                'sort_wafer', 'sort wafer', 'sort partial wafer id'],
    'program':    ['program name', 'program_name', 'programname', 'program', 'tp', 'test program'],
    'devrevstep': ['devrevstep'],
    'sort_lot': ['sort_lot', 'sortlot', 'sort lot'],
    'ibin':    ['interface bin', 'interface_bin', 'ibin', 'bin', 'bin_num', 'bin num'],
    'count':   ['count', 'die count', 'die_count', 'fail count', 'fail_count'],
    'total':   ['total dies', 'total_dies', 'totaldies', 'total', 'dies',
                'interface_total_bin', 'interface total bin'],
    'material': ['material'],
    'fbin':   ['functional bin', 'functional_bin', 'fbin', 'fb'],
    'bin_desc': ['bin description', 'bindescription', 'bin_description'],
    'x':      ['sort_x', 'sort x', 'x', 'coordx', 'coord_x', 'die_x', 'posx'],
    'y':      ['sort_y', 'sort y', 'y', 'coordy', 'coord_y', 'die_y', 'posy'],
}


def _resolve_cols(header: list[str]) -> dict[str, int]:
    import re as _re
    # Normalize: strip trailing _DIGITS suffix (TRACE raw exports append job number)
    def _norm(h: str) -> str:
        return _re.sub(r'_\d+$', '', h.strip()).strip().lower()
    lower_hdr = [_norm(h) for h in header]
    out = {}
    for canon, aliases in _COL_ALIASES.items():
        for alias in aliases:
            if alias in lower_hdr:
                out[canon] = lower_hdr.index(alias)
                break
    # Detect UPM_*_FULLDIE_*0950* column (any product prefix, e.g. UPM_0107 or UPM_2007)
    import fnmatch as _fnmatch
    _upm950_pat = 'upm_*fulldie*0950*'
    for i, h in enumerate(header):
        if _fnmatch.fnmatch(_re.sub(r'_\d+$', '', h.strip()).strip().lower(), _upm950_pat):
            out['upm_950'] = i
            break
    return out


def _open_csv_source(path: Path):
    """
    Open a CSV, ZIP, or GZ file and return a text-mode file-like object.
    ZIP: uses the first .csv member found.
    GZ: assumes the compressed content is a CSV.
    """
    import io
    suffix = path.suffix.lower()
    if suffix == '.zip':
        import zipfile
        zf = zipfile.ZipFile(path, 'r')
        members = [m for m in zf.namelist() if m.lower().endswith('.csv')]
        if not members:
            members = zf.namelist()  # fallback: first member regardless
        if not members:
            raise ValueError(f'No CSV found inside ZIP: {path.name}')
        raw_bytes = zf.read(members[0])
        zf.close()
        return io.StringIO(raw_bytes.decode('utf-8-sig', errors='replace'))
    elif suffix in ('.gz', '.gzip'):
        import gzip
        with gzip.open(path, 'rb') as gz:
            raw_bytes = gz.read()
        return io.StringIO(raw_bytes.decode('utf-8-sig', errors='replace'))
    else:
        return open(path, newline='', encoding='utf-8-sig')


def load_material_data(product_prefix: str) -> dict[str, str]:
    """
    Load material type data from collateral lot-definition CSV.
    Returns dict: first_7_chars_of_intel_lot7 -> material_type
    E.g., {'Q603S6T': 'NVL816-BLLC-L0 AIO', ...}
    """
    import csv as _csv
    # trend_chart.py is at: .../app.yield.nvl/code/dashboard/yield-dashboard/yld/src/
    # Need to get to: .../app.yield.nvl/shared/material
    script_dir = Path(__file__).resolve().parent
    # Go up to find shared/material directory
    current = script_dir
    for i in range(10):  # Search up to 10 levels
        collateral_dir = current / 'shared' / 'material'
        if collateral_dir.exists():
            break
        current = current.parent
    
    if not collateral_dir.exists():
        return {}
    
    # Find lot-definition CSV starting with product_prefix
    lot_def_files = list(collateral_dir.glob(f'{product_prefix}*.csv'))
    if not lot_def_files:
        return {}
    
    material_map = {}
    for lot_file in lot_def_files:
        try:
            with open(lot_file, newline='', encoding='utf-8-sig') as fh:
                rdr = _csv.DictReader(fh)
                if not rdr.fieldnames:
                    continue
                # Find columns for INTEL_LOT7 and material type
                intel_lot_col = next((c for c in rdr.fieldnames if 'intel_lot' in c.lower()), None)
                wafer_col = next((c for c in rdr.fieldnames if c.strip().lower() == 'waferid'), None)
                mat_col = next((c for c in rdr.fieldnames if 'material' in c.lower()), None)
                
                if not intel_lot_col or not mat_col:
                    continue
                
                for row in rdr:
                    intel_lot_val = row.get(intel_lot_col, '').strip()
                    # Often the lot string has dots like 'Q604SB1.01'
                    clean_lot = intel_lot_val.replace('.', '')
                    mat_val = row.get(mat_col, '').strip()
                    if clean_lot and mat_val:
                        # Index by full lot, first 7 characters, and WaferID
                        lot7 = clean_lot[:7]
                        if wafer_col:
                            raw_w = row.get(wafer_col, '').strip()
                            try:
                                w_key = str(int(float(raw_w)))
                            except (ValueError, TypeError):
                                w_key = raw_w
                            material_map[(clean_lot, w_key)] = mat_val
                            material_map[(lot7, w_key)] = mat_val
                        material_map.setdefault(clean_lot, mat_val)
                        material_map.setdefault(lot7, mat_val)
        except Exception:
            pass
    
    return material_map


def load_csv(path: Path, log=None, grouping_mode: str = 'wafer') -> list[dict]:
    """Parse CSV / ZIP / GZ; return list of per-run dicts.
    
    grouping_mode: 'wafer' (default) = one bar per wafer
                   'lot' = one bar per lot (combines all wafers)
    """
    import csv as _csv
    path = Path(path)
    raw_rows = []
    fh = _open_csv_source(path)
    try:
        rdr = _csv.reader(fh)
        header = next(rdr, [])
        for row in rdr:
            if any(cell.strip() for cell in row):
                raw_rows.append(row)
    finally:
        fh.close()

    col = _resolve_cols(header)
    if 'ibin' not in col:
        raise ValueError(f'Cannot find Interface Bin column.\nHeader: {header}')
    if 'total' not in col and 'count' not in col:
        raise ValueError(f'Need at least a Count or Total Dies column.\nHeader: {header}')
    _per_unit_mode = 'count' not in col  # each row = 1 die
    _has_xy = _per_unit_mode and 'x' in col and 'y' in col  # wafermap needs per-die X/Y

    def _get(row, key, default=''):
        idx = col.get(key)
        if idx is None or idx >= len(row):
            return default
        return row[idx].strip()

    _upm950_divisor = 9154.0  # default MHz → % divisor (NVL816); overridden per product config
    
    # Load material data from collateral (one-time, before processing rows)
    product_prefix = ''
    material_map = {}
    for row in raw_rows[:10]:  # Scan first 10 rows to find product prefix
        devrevstep_sample = _get(row, 'devrevstep', '')
        if devrevstep_sample:
            product_prefix = devrevstep_sample[:6]  # e.g., '8PF5CV' from devrevstep
            break
    if product_prefix:
        material_map = load_material_data(product_prefix)
        # Derive UPM divisor from product config upmInfo base (e.g. 7734 for NVLG-512)
        try:
            _auto_cfg_path = _find_auto_config(product_prefix)
            if _auto_cfg_path:
                import json as _json_lc
                _cfg_lc = _json_lc.loads(_auto_cfg_path.read_text(encoding='utf-8'))
                _upm_info = _cfg_lc.get('upmInfo', [])
                if _upm_info and len(_upm_info[0]) > 2 and _upm_info[0][2]:
                    _upm950_divisor = float(_upm_info[0][2])
                # Re-resolve upm_950 column using the config's exact pattern (beats generic fallback)
                if _upm_info and len(_upm_info[0]) > 1:
                    import fnmatch as _fn_upm
                    _upm_cfg_pat = _upm_info[0][1].strip().lower()
                    for _ci, _ch in enumerate(header):
                        if _fn_upm.fnmatch(_ch.strip().lower(), _upm_cfg_pat):
                            col['upm_950'] = _ci
                            break
        except Exception:
            pass
    
    groups: dict[tuple, dict] = OrderedDict()
    for row in raw_rows:
        lot        = _get(row, 'lot', 'LOT?')
        wafer      = _get(row, 'wafer', '')
        program    = _get(row, 'program', '')
        devrevstep = _get(row, 'devrevstep', '')[:6]  # truncate to 6 chars for grouping
        sort_lot   = _get(row, 'sort_lot', '')
        
        # Material: first check CSV row, then fall back to collateral lookup
        _csv_material = _get(row, 'material', '')
        if not _csv_material:
            lookup_lot = sort_lot if sort_lot else lot
            lot7 = lookup_lot[:7]
            _w_key = ''
            if wafer:
                try:
                    _w_key = str(int(wafer[-2:])) if len(wafer) >= 2 else str(int(float(wafer)))
                except (ValueError, TypeError):
                    _w_key = wafer
            material = (material_map.get((lookup_lot, _w_key))
                        or material_map.get(lookup_lot)
                        or material_map.get((lot7, _w_key))
                        or material_map.get(lot7)
                        or '')
        else:
            material = _csv_material

        date_s  = _get(row, 'date', '')
        ibin_s  = _get(row, 'ibin', '')
        cnt_s   = _get(row, 'count', '1') if not _per_unit_mode else '1'
        tot_s   = '0' if _per_unit_mode else _get(row, 'total', '0')
        upm950_s = _get(row, 'upm_950', '')

        try:
            ibin = int(float(ibin_s))
        except (ValueError, TypeError):
            continue
        try:
            cnt = int(float(cnt_s))
        except (ValueError, TypeError):
            cnt = 0
        try:
            tot = int(float(tot_s))
        except (ValueError, TypeError):
            tot = 0

        # Group by wafer (default) or lot only
        if grouping_mode == 'lot':
            key = (lot[:7], program, material)
        else:
            key = (lot, wafer, program, material)
        
        if key not in groups:
            dt = _parse_date(date_s)
            if dt is None:
                dt = _parse_date(lot) or _parse_date(program)
            groups[key] = {
                'lot': lot[:7] if grouping_mode == 'lot' else lot,
                'wafer': wafer, 'program': program,
                'devrevstep': devrevstep,
                'sort_lot': sort_lot[:7] if grouping_mode == 'lot' else sort_lot,
                'material': material,
                'label': _prog_label(program, lot, wafer if grouping_mode == 'wafer' else ''),
                'date_str': date_s, 'date': dt,
                'total_dies': tot, 'bin_counts': {},
                'upm_950': [],  # per-die [ibin, upm_pct] pairs
                'die_xy': [],   # per-die [x, y, ibin, fbin, upm_pct] for wafermap drilldown
            }
        grp = groups[key]
        if not grp.get('material') and material:
            grp['material'] = material
        wtot = grp.setdefault('_wafer_totals', {})
        if tot > wtot.get(wafer, 0):
            wtot[wafer] = tot
        grp['bin_counts'][ibin] = grp['bin_counts'].get(ibin, 0) + cnt
        # Collect functional-bin breakdown per ibin
        fbin_s = _get(row, 'fbin', '')
        fbin_val = None
        if fbin_s:
            try:
                fbin = int(float(fbin_s))
                fbin_val = fbin
                fb_map = grp.setdefault('fb_counts', {})
                ib_fb  = fb_map.setdefault(ibin, {})
                ib_fb[fbin] = ib_fb.get(fbin, 0) + cnt
                # Collect bin_desc (bin setter string) for fail test module
                bdesc = _get(row, 'bin_desc', '')
                if bdesc:
                    mod_map = grp.setdefault('fb_modules', {})
                    ib_mod  = mod_map.setdefault(ibin, {})
                    ib_mod.setdefault(fbin, {})
                    ib_mod[fbin][bdesc] = ib_mod[fbin].get(bdesc, 0) + cnt
            except (ValueError, TypeError):
                pass
        # Store [ibin, upm_pct] per die for DLCP CDF (same as bin_distribution_html)
        upm_pct_val = None
        if upm950_s:
            try:
                upm_pct_val = round(float(upm950_s) / _upm950_divisor * 100, 2)
                grp['upm_950'].append([ibin, upm_pct_val])
            except (ValueError, TypeError):
                pass
        # Store [x, y, ibin, fbin, upm_pct] per die for wafermap drilldown
        if _has_xy:
            try:
                xy = int(float(_get(row, 'x', ''))), int(float(_get(row, 'y', '')))
                grp['die_xy'].append([xy[0], xy[1], ibin, fbin_val, upm_pct_val])
            except (ValueError, TypeError):
                pass

    result = []
    for grp in groups.values():
        wafer_totals = grp.pop('_wafer_totals', {})
        total     = sum(wafer_totals.values()) or grp['total_dies'] or sum(grp['bin_counts'].values()) or 1
        wafer_count = len(wafer_totals) or 1
        ff_cnt    = sum(c for b, c in grp['bin_counts'].items() if b in _FF_BINS)
        ff_df_cnt = sum(c for b, c in grp['bin_counts'].items() if b in _FF_DF_BINS)
        fail_ibins = {b: c / total * 100
                      for b, c in grp['bin_counts'].items()
                      if b not in _PASS_BINS}
        r = {**grp, 'total_dies': total, 'wafer_count': wafer_count,
             'ff_yield':    ff_cnt    / total * 100,
             'ff_df_yield': ff_df_cnt / total * 100,
             'fail_ibins':  fail_ibins,
             'material':    grp.get('material', ''),
             'upm_950':     grp.get('upm_950', []),
             'fb_counts':   grp.get('fb_counts', {}),
             'fb_modules':  grp.get('fb_modules', {})}
        result.append(r)
        if log:
            n_f = sum(1 for v in fail_ibins.values() if v > 0)
            log(f'  [{grp["label"]}] lot={grp["lot"]} w={grp["wafer"]}  '
                f'FF={r["ff_yield"]:.1f}%  FF+DF={r["ff_df_yield"]:.1f}%  '
                f'fail_ibins={n_f}\n')
    return result


# ============================================================================
# 5. Grouping
# ============================================================================

def group_runs(runs: list[dict], interval: str) -> OrderedDict:
    """Group runs into {interval_label: [run, ...]} ordered chronologically.
    Within each group, runs are sorted by date (earliest to latest)."""
    grouped: dict[str, list] = {}
    for r in runs:
        if interval == 'revision':
            drs = r.get('devrevstep', '')
            key = drs if drs else _prog_label(r.get('program', ''), r.get('lot', ''), r.get('wafer', ''))
        elif r['date']:
            key = _interval_key(r['date'], interval, r.get('program', ''))
        else:
            key = r['lot']
        grouped.setdefault(key, []).append(r)

    # Sort runs within each group by (lot's last test date, run date) so all
    # runs for a lot are contiguous and the latest-tested lot appears last
    for group_runs_list in grouped.values():
        _no_date = datetime.min
        lot_last: dict[str, datetime] = {}
        for r in group_runs_list:
            lot = r['lot']
            dt = r['date'] or _no_date
            if dt > lot_last.get(lot, _no_date):
                lot_last[lot] = dt
        group_runs_list.sort(key=lambda r: (lot_last.get(r['lot'], _no_date),
                                            r['date'] or _no_date))

    def _sk(k):
        sk = _interval_sort_key(k)
        return (*sk, k) if sk != (9999, 0, 0) else (9999, 0, 0, k)

    return OrderedDict(sorted(grouped.items(), key=lambda kv: _sk(kv[0])))


def _aggregate_by_lot(runs: list[dict]) -> list[dict]:
    """Combine per-wafer runs into one pseudo-run per (program, lot[:7], material).

    Mirrors the client-side aggregateByLot() JS used for the 'Program / Lot'
    grouping view — used only to build the initial static trend chart so it
    matches that default sidebar selection; DATA.runs (embedded JSON) keeps
    the original per-wafer entries so the wafer-map drilldown stays accurate.
    """
    groups: dict[tuple, dict] = OrderedDict()
    for r in runs:
        lot7 = (r.get('lot') or '')[:7]
        key = (r.get('program', ''), lot7, r.get('material', ''))
        g = groups.get(key)
        if g is None:
            g = groups[key] = {
                'lot': lot7, 'wafer': '',
                'sort_lot': (r.get('sort_lot') or r.get('lot') or '')[:7],
                'material': r.get('material', ''), 'program': r.get('program', ''),
                'devrevstep': r.get('devrevstep', ''),
                'date': r.get('date'), 'date_str': r.get('date_str', ''),
                'total_dies': 0, 'bin_counts': {}, '_wafers': [],
            }
        g['total_dies'] += r.get('total_dies', 0) or 0
        g['_wafers'].append(r.get('wafer') or '?')
        rd = r.get('date')
        if rd and (not g['date'] or rd > g['date']):
            g['date'] = rd
            g['date_str'] = r.get('date_str', '')
        for ib, cnt in (r.get('bin_counts') or {}).items():
            g['bin_counts'][ib] = g['bin_counts'].get(ib, 0) + cnt

    result = []
    for g in groups.values():
        wafers = g.pop('_wafers')
        total = g['total_dies'] or sum(g['bin_counts'].values()) or 1
        ff_cnt    = sum(c for b, c in g['bin_counts'].items() if b in _FF_BINS)
        ff_df_cnt = sum(c for b, c in g['bin_counts'].items() if b in _FF_DF_BINS)
        fail_ibins = {b: c / total * 100
                      for b, c in g['bin_counts'].items() if b not in _PASS_BINS}
        result.append({
            **g,
            'wafer': wafers[0] if len(wafers) == 1 else f'{len(wafers)}W',
            'wafer_count': len(wafers),
            'total_dies': total,
            'ff_yield': ff_cnt / total * 100,
            'ff_df_yield': ff_df_cnt / total * 100,
            'fail_ibins': fail_ibins,
            'label': _prog_label(g['program'], g['lot'], ''),
        })
    return result


# ============================================================================
# 6. Chart builders (Plotly)
# ============================================================================

def _ibin_display(ibin: int, cfg: dict | None) -> str:
    """Return 'IB N — Category' or just 'IB N'."""
    if cfg and cfg.get('ibin_name', {}).get(ibin):
        return f'IB {ibin} \u2014 {cfg["ibin_name"][ibin]}'
    return f'IB {ibin}'


def build_trend_chart(groups: OrderedDict,
                      top_n_fail_ibins: int = 8,
                      fail_thresh_pct: float = 0.0,
                      interval: str = 'revision',
                      cfg: dict | None = None) -> 'go.Figure':
    """
    Plotly Figure: stacked clustered bars (fail% per iBin) + dual-Y yield lines.
    X-axis = run short labels; period separators shown as vertical dotted lines
    with interval labels as annotations above the chart.
    """
    # --- Top-N fail ibins by cumulative fail%
    global_fail: dict[int, float] = {}
    for runs in groups.values():
        for r in runs:
            for ib, pct in r['fail_ibins'].items():
                global_fail[ib] = global_fail.get(ib, 0) + pct

    top_ibins = sorted(
        [ib for ib, v in global_fail.items() if v >= fail_thresh_pct],
        key=lambda ib: global_fail[ib], reverse=True
    )[:top_n_fail_ibins]

    # --- Flatten runs in order
    all_runs_ordered: list[tuple[str, dict]] = []   # (iv_label, run)
    iv_start_indices: list[int]  = []               # x-index where interval starts
    iv_labels_ordered: list[str] = []

    for iv_label, iv_runs in groups.items():
        iv_start_indices.append(len(all_runs_ordered))
        iv_labels_ordered.append(iv_label)
        for r in iv_runs:
            all_runs_ordered.append((iv_label, r))

    n_runs = len(all_runs_ordered)
    if n_runs == 0:
        fig = go.Figure()
        fig.add_annotation(text='No runs to plot.', xref='paper', yref='paper',
                           x=0.5, y=0.5, showarrow=False, font_size=16)
        return fig

    x_pos      = list(range(n_runs))
    short_lbls = [(r.get('sort_lot') or r['label']) + 
                  (f' ({r.get("material", "")})' if r.get('material') else '')
                  for _, r in all_runs_ordered]

    # Build rich tooltip base per run
    def _run_tip(iv: str, r: dict) -> str:
        d = (r['date_str'] or
             (r['date'].strftime('%Y-%m-%d') if r['date'] else '\u2014'))
        return (f'<b>{r["label"]}</b> | Period: {iv}<br>'
                f'Lot: {r["lot"]}  Wafer: {r["wafer"]}<br>'
                f'Program: {r["program"]}<br>'
                f'Date: {d}  |  Dies: {r["total_dies"]:,}')

    run_tips = [_run_tip(iv, r) for iv, r in all_runs_ordered]

    ff_tgt     = cfg['yield_target'].get('ff')    if cfg else None
    ffdf_tgt   = cfg['yield_target'].get('ff_df') if cfg else None
    ff_name    = (cfg or {}).get('ff_name',   'SDS FF')
    ffdf_name  = (cfg or {}).get('ff_df_name', 'SDS FF+DF')
    chart_name = (cfg or {}).get('name', '')

    fig = make_subplots(specs=[[{'secondary_y': True}]])

    # --- Stacked bar traces (one per fail iBin)
    for bi, ibin in enumerate(top_ibins):
        ibin_lbl = _ibin_display(ibin, cfg)
        tgt      = cfg['ibin_target'].get(ibin) if cfg else None
        bar_y    = []
        hover    = []
        for idx, (iv, r) in enumerate(all_runs_ordered):
            pct = r['fail_ibins'].get(ibin, 0.0)
            bar_y.append(pct)
            htxt = (f'{run_tips[idx]}<br>\u2500\u2500\u2500\u2500\u2500<br>'
                    f'<b>{ibin_lbl}</b><br>Fail: <b>{pct:.2f}%</b>')
            if tgt is not None:
                htxt += f'<br>Target: {tgt:.1f}%'
            hover.append(htxt)

        fig.add_trace(go.Bar(
            x=x_pos, y=bar_y,
            name=ibin_lbl,
            hovertext=hover, hoverinfo='text',
            marker_color=_FAIL_PALETTE[bi % len(_FAIL_PALETTE)],
            marker_line_color='white', marker_line_width=0.4,
            opacity=0.85,
            legendgroup='fail_bins',
        ), secondary_y=False)

    # --- FF yield line
    ff_y  = [r['ff_yield']    for _, r in all_runs_ordered]
    ffdf_y= [r['ff_df_yield'] for _, r in all_runs_ordered]

    ff_hover = [
        f'{run_tips[i]}<br>\u2500\u2500\u2500\u2500\u2500<br>'
        f'<b>{ff_name}</b>: {ff_y[i]:.2f}%'
        + (f'<br>Target: {ff_tgt:.1f}%' if ff_tgt is not None else '')
        for i in range(n_runs)
    ]
    fig.add_trace(go.Scatter(
        x=x_pos, y=ff_y,
        mode='lines+markers+text',
        name=ff_name,
        line=dict(color='#1a73e8', width=2.5),
        marker=dict(size=8),
        text=[f'{v:.1f}%' for v in ff_y],
        textposition='top center', textfont=dict(size=9, color='#1a73e8'),
        hovertext=ff_hover, hoverinfo='text',
        legendgroup='yield_lines',
    ), secondary_y=True)

    ffdf_hover = [
        f'{run_tips[i]}<br>\u2500\u2500\u2500\u2500\u2500<br>'
        f'<b>{ffdf_name}</b>: {ffdf_y[i]:.2f}%'
        + (f'<br>Target: {ffdf_tgt:.1f}%' if ffdf_tgt is not None else '')
        for i in range(n_runs)
    ]
    fig.add_trace(go.Scatter(
        x=x_pos, y=ffdf_y,
        mode='lines+markers+text',
        name=ffdf_name,
        line=dict(color='#2e7d32', width=2.5, dash='dash'),
        marker=dict(size=8, symbol='square'),
        text=[f'{v:.1f}%' for v in ffdf_y],
        textposition='bottom center', textfont=dict(size=9, color='#2e7d32'),
        hovertext=ffdf_hover, hoverinfo='text',
        legendgroup='yield_lines',
    ), secondary_y=True)

    # --- Yield target reference lines
    if ff_tgt is not None:
        fig.add_hline(y=ff_tgt, line_dash='dot', line_color='#1a73e8',
                      line_width=1.5, opacity=0.5, secondary_y=True,
                      annotation_text=f'{ff_name} target {ff_tgt:.1f}%',
                      annotation_position='right',
                      annotation_font_size=10)
    if ffdf_tgt is not None:
        fig.add_hline(y=ffdf_tgt, line_dash='dot', line_color='#2e7d32',
                      line_width=1.5, opacity=0.5, secondary_y=True,
                      annotation_text=f'{ffdf_name} target {ffdf_tgt:.1f}%',
                      annotation_position='right',
                      annotation_font_size=10)

    # --- Period dividers + annotations
    shapes, annots = [], []
    for bi, (bnd, iv_name) in enumerate(zip(iv_start_indices, iv_labels_ordered)):
        end = iv_start_indices[bi + 1] if bi + 1 < len(iv_start_indices) else n_runs
        mid = (bnd + end - 1) / 2

        if bnd > 0:
            shapes.append(dict(
                type='line', x0=bnd - 0.5, x1=bnd - 0.5,
                y0=0, y1=1, yref='paper',
                line=dict(color='#95a5a6', width=1.2, dash='dot'),
            ))
        annots.append(dict(
            x=mid, y=1.06, xref='x', yref='paper',
            text=f'<b>{iv_name}</b>',
            showarrow=False,
            font=dict(size=11, color='#2c3e50'),
            xanchor='center',
        ))

    # --- Y-axis range
    max_stack = max(
        (sum(r['fail_ibins'].get(ib, 0) for ib in top_ibins) for _, r in all_runs_ordered),
        default=0.0,
    )
    fail_ylim = min(100.0, max(max_stack * 1.25, 5.0))

    fig.update_layout(
        barmode='stack',
        plot_bgcolor='#f9f9fb',
        paper_bgcolor='white',
        title=dict(
            text=((f'<b>{chart_name}</b> — ' if chart_name else '')
                  + f'Interface Bin Fail vs. Yield Trend \u2014 <b>{interval}</b> intervals<br>'
                  f'<sup>{n_runs} run{"s" if n_runs != 1 else ""}, '
                  f'{len(groups)} period{"s" if len(groups) != 1 else ""}</sup>'),
            font=dict(size=16),
        ),
        xaxis=dict(
            tickvals=x_pos,
            ticktext=short_lbls,
            tickfont=dict(size=10),
            tickangle=-35,
            showgrid=False,
            title='SORT LOT',
        ),
        yaxis=dict(
            title='Interface Bin Fail (%)',
            range=[0, fail_ylim],
            gridcolor='#e8e8e8',
            zeroline=True, zerolinecolor='#ccc',
        ),
        yaxis2=dict(
            title='Yield (%)',
            range=[0, 105],
            overlaying='y', side='right',
            showgrid=False,
        ),
        legend=dict(
            orientation='v',
            x=1.09, y=1.0,
            bgcolor='rgba(255,255,255,0.85)',
            bordercolor='#ddd', borderwidth=1,
            font=dict(size=11),
        ),
        shapes=shapes,
        annotations=annots,
        margin=dict(l=60, r=200, t=110, b=80),
        hovermode='closest',
        hoverlabel=dict(bgcolor='white', font_size=12, bordercolor='#ccc'),
        autosize=True,
    )
    return fig


def build_pareto_vertical_chart(runs: list[dict],
                                top_n: int = 20,
                                cfg: dict | None = None):
    """
    Fail Pareto Chart (Percentage) — vertical bars.
    X-axis: iBin number labels.
    Left Y-axis: % failure per bin (averaged across runs).
    Right Y-axis: cumulative % reaching 100%.
    Returns (fig, table_rows) where table_rows is a list of dicts.
    """
    # Auto-load product config for ibin descriptions if not supplied
    if cfg is None:
        drs = runs[0].get('devrevstep', '') if runs else ''
        auto = _find_auto_config(drs)
        if auto:
            try:
                cfg = load_product_config(auto)
            except Exception:
                cfg = None
    global_fail: dict[int, float] = {}
    run_count = len(runs)
    for r in runs:
        for ib, pct in r['fail_ibins'].items():
            global_fail[ib] = global_fail.get(ib, 0) + pct

    if not global_fail:
        fig = go.Figure()
        fig.add_annotation(text='No fail bin data.', xref='paper', yref='paper',
                           x=0.5, y=0.5, showarrow=False, font_size=16)
        return fig, []

    avg_fail     = {ib: v / run_count for ib, v in global_fail.items()}
    sorted_ibins = sorted(avg_fail, key=lambda ib: avg_fail[ib], reverse=True)[:top_n]
    total_avg    = sum(avg_fail[ib] for ib in sorted_ibins) or 1.0

    x_labels, bar_vals, hover_txt, colors = [], [], [], []
    table_rows = []
    bin_map = (cfg or {}).get('raw', {}).get('bin_map', {})
    for i, ib in enumerate(sorted_ibins):
        pct  = avg_fail[ib]
        lbl  = _ibin_display(ib, cfg)
        info = bin_map.get(str(ib), {})
        cat  = (info.get('cat') or '').strip()
        desc = (info.get('desc') or '').strip()
        n_fail = sum(r.get('bin_counts', {}).get(ib, 0) for r in runs)
        x_labels.append(lbl)
        bar_vals.append(pct)
        colors.append(_FAIL_PALETTE[i % len(_FAIL_PALETTE)])
        hover_txt.append(
            f'<b>{lbl}</b><br>Avg Fail: <b>{pct:.2f}%</b><br>'
            f'Across {run_count} run{"s" if run_count != 1 else ""}'
        )
        table_rows.append({'ib': ib, 'cat': cat, 'desc': desc,
                           'n_fail': n_fail, 'pct': pct})

    cum_vals: list[float] = []
    running = 0.0
    for v in bar_vals:
        running += v / total_avg * 100
        cum_vals.append(running)

    fig = make_subplots(specs=[[{'secondary_y': True}]])

    fig.add_trace(go.Bar(
        x=x_labels, y=bar_vals,
        name='Avg Fail (%)',
        marker_color=colors,
        marker_line_color='#1a252f', marker_line_width=0.8,
        opacity=0.9,
        hovertext=hover_txt, hoverinfo='text',
        text=[f'{v:.2f}%' for v in bar_vals],
        textposition='outside', textfont=dict(size=10, color='#333', family='Arial'),
    ), secondary_y=False)

    fig.add_trace(go.Scatter(
        x=x_labels, y=cum_vals,
        mode='lines+markers',
        name='Cumulative %',
        line=dict(color='#e67e22', width=2.5),
        marker=dict(size=7, color='#e67e22'),
        hovertemplate='<b>%{x}</b><br>Cumulative: %{y:.1f}%<extra></extra>',
    ), secondary_y=True)

    fig.add_hline(y=80, line_dash='dash', line_color='#e74c3c',
                  line_width=1.8, opacity=0.8,
                  annotation_text='80% cumulative', annotation_position='top right',
                  annotation_font_size=11, annotation_font_color='#e74c3c',
                  secondary_y=True)

    n = len(sorted_ibins)
    chart_name_pv = (cfg or {}).get('name', '')
    title_pv = ((f'<b>{chart_name_pv}</b> \u2014 ' if chart_name_pv else '')
                + f'<b>Fail Pareto Chart (Percentage)</b><br>'
                f'<sup>Top {n} fail bins, averaged across '
                f'{run_count} run{"s" if run_count != 1 else ""}</sup>')
    fig.update_layout(
        plot_bgcolor='#f9f9fb',
        paper_bgcolor='white',
        title=dict(
            text=title_pv,
            font=dict(size=16),
            x=0.5,
            xanchor='center',
        ),
        xaxis=dict(
            title='Interface Bin',
            gridcolor='#e8e8e8',
            tickangle=-35, tickfont=dict(size=10),
        ),
        yaxis=dict(
            title='Fail (%)',
            gridcolor='#e8e8e8', zeroline=True, zerolinecolor='#ccc',
            range=[0, max(bar_vals) * 1.15] if bar_vals else None,
            ticksuffix='%',
        ),
        yaxis2=dict(
            title='Cumulative (%)',
            range=[0, 105],
            showgrid=False,
            ticksuffix='%',
        ),
        legend=dict(x=1.08, y=1.0, bgcolor='rgba(255,255,255,0.85)',
                    bordercolor='#ddd', borderwidth=1),
        margin=dict(l=70, r=120, t=90, b=120),
        hovermode='closest',
        hoverlabel=dict(bgcolor='white', font_size=12, bordercolor='#ccc'),
        autosize=True,
        bargap=0.25,
    )
    return fig, table_rows


def build_pareto_chart(runs: list[dict],
                       top_n: int = 20,
                       cfg: dict | None = None) -> 'go.Figure':
    """
    Overall Interface Bin Pareto -- horizontal bar chart + cumulative % line.
    Sorted by average fail% across all runs.
    """
    global_fail: dict[int, float] = {}
    run_count = len(runs)
    for r in runs:
        for ib, pct in r['fail_ibins'].items():
            global_fail[ib] = global_fail.get(ib, 0) + pct

    if not global_fail:
        fig = go.Figure()
        fig.add_annotation(text='No fail bin data.', xref='paper', yref='paper',
                           x=0.5, y=0.5, showarrow=False, font_size=16)
        return fig

    avg_fail     = {ib: v / run_count for ib, v in global_fail.items()}
    sorted_ibins = sorted(avg_fail, key=lambda ib: avg_fail[ib], reverse=True)[:top_n]
    total_avg    = sum(avg_fail[ib] for ib in sorted_ibins) or 1.0

    y_labels, bar_vals, hover_txt, colors = [], [], [], []
    for i, ib in enumerate(sorted_ibins):
        pct  = avg_fail[ib]
        lbl  = _ibin_display(ib, cfg)
        tgt  = cfg['ibin_target'].get(ib) if cfg else None
        y_labels.append(lbl)
        bar_vals.append(pct)
        colors.append(_FAIL_PALETTE[i % len(_FAIL_PALETTE)])
        htxt = (f'<b>{lbl}</b><br>Avg Fail: <b>{pct:.2f}%</b><br>'
                f'Total fail sum: {global_fail[ib]:.2f}%<br>'
                f'Across {run_count} run{"s" if run_count != 1 else ""}')
        if tgt is not None:
            htxt += f'<br>Target: {tgt:.1f}%'
        hover_txt.append(htxt)

    cum_vals: list[float] = []
    running = 0.0
    for v in bar_vals:
        running += v / total_avg * 100
        cum_vals.append(running)

    fig = make_subplots(specs=[[{'secondary_y': True}]])

    fig.add_trace(go.Bar(
        y=y_labels, x=bar_vals,
        orientation='h',
        name='Avg Fail (%)',
        marker_color=colors,
        marker_line_color='white', marker_line_width=0.5,
        opacity=0.9,
        hovertext=hover_txt, hoverinfo='text',
        text=[f'{v:.2f}%' for v in bar_vals],
        textposition='outside', textfont=dict(size=10, color='#333', family='Arial'),
    ), secondary_y=False)

    fig.add_trace(go.Scatter(
        y=y_labels, x=cum_vals,
        mode='lines+markers',
        name='Cumulative %',
        line=dict(color='#e67e22', width=2.5),
        marker=dict(size=7, color='#e67e22'),
        hovertemplate='<b>%{y}</b><br>Cumulative: %{x:.1f}%<extra></extra>',
    ), secondary_y=True)

    fig.add_vline(x=80, line_dash='dash', line_color='#e74c3c',
                  line_width=1.8, opacity=0.8,
                  annotation_text='80%', annotation_position='top right',
                  annotation_font_size=11, annotation_font_color='#e74c3c')

    n = len(sorted_ibins)
    chart_name_p = (cfg or {}).get('name', '')
    title_p = ((f'<b>{chart_name_p}</b> \u2014 ' if chart_name_p else '')
               + f'Overall Interface Bin Fail Pareto<br>'
               f'<sup>Top {n} fail bins, averaged across '
               f'{run_count} run{"s" if run_count != 1 else ""}</sup>')
    fig.update_layout(
        plot_bgcolor='#f9f9fb',
        paper_bgcolor='white',
        title=dict(
            text=title_p,
            font=dict(size=16),
        ),
        xaxis=dict(
            title='Average Fail (%)',
            gridcolor='#e8e8e8',
            range=[0, max(bar_vals) * 1.25 if bar_vals else 10]),
        xaxis2=dict(title='Cumulative (%)', range=[0, 110],
                    overlaying='x', side='top', showgrid=False),
        yaxis=dict(autorange='reversed', tickfont=dict(size=10),
                   showgrid=False),
        yaxis2=dict(range=[0, 110], overlaying='y', side='right',
                    showgrid=False, visible=False),
        legend=dict(x=1.08, y=1.0, bgcolor='rgba(255,255,255,0.85)',
                    bordercolor='#ddd', borderwidth=1),
        margin=dict(l=220, r=120, t=90, b=60),
        hovermode='closest',
        hoverlabel=dict(bgcolor='white', font_size=12, bordercolor='#ccc'),
        autosize=True,
    )
    return fig


# ============================================================================
# 7. HTML generation
# ============================================================================

def generate_html(csv_path: Path, groups: OrderedDict, runs: list[dict],
                  trend_fig: 'go.Figure', pareto_fig: 'go.Figure',
                  output_path: Path,
                  interval: str = 'revision',
                  top_n: int = 8,
                  cfg_path: str = '',
                  cfg: dict | None = None,
                  pareto_vertical_fig: 'go.Figure | None' = None,
                  pareto_table_rows: list | None = None,
                  grouping_mode: str = 'lot') -> None:
    """Generate a fully interactive self-contained HTML report.

    The report embeds all run data as JSON and uses JavaScript + Plotly.react()
    to refilter/regroup live in the browser — no server needed.
    """
    ts_now = datetime.now().strftime('%Y-%m-%d %H:%M')

    # ── Serialize run data for JS ──────────────────────────────────────────
    ibin_names   = (cfg or {}).get('ibin_name',   {})
    yield_target = (cfg or {}).get('yield_target', {})

    runs_json_list = []
    runs_dlcp_extra = []  # fb_modules only needed for drill-down click; parsed lazily on first DLCP/pareto click
    for r in runs:
        date_s = r['date'].strftime('%Y-%m-%d') if r.get('date') else (
            (r.get('date_str') or '')[:10])
        runs_json_list.append({
            'lot':        r['lot'],
            'wafer':      r.get('wafer', ''),
            'sort_lot':   r.get('sort_lot', ''),
            'material':   r.get('material', ''),
            'program':    r['program'],
            'date':       date_s,
            'total_dies': r['total_dies'],
            'bin_counts': {str(k): v for k, v in r.get('bin_counts', {}).items()},
            'fb_counts':  {str(ib): {str(fb): cnt for fb, cnt in fb_map.items()}
                           for ib, fb_map in r.get('fb_counts', {}).items()},
            'ff_upm':     sorted(up for ib, up in r.get('upm_950', []) if ib in (1, 2)),
            'df_upm':     sorted(up for ib, up in r.get('upm_950', []) if ib in (3, 4)),
        })
        runs_dlcp_extra.append({
            'fb_modules': {str(ib): {str(fb): max(bdesc_map, key=bdesc_map.get)
                           for fb, bdesc_map in fb_bdesc_map.items()}
                           for ib, fb_bdesc_map in r.get('fb_modules', {}).items()},
            'die_xy': r.get('die_xy', []),  # [x, y, ibin, fbin, upm_pct] per die, for wafermap drilldown
        })

    all_progs  = sorted({r['program'] for r in runs})
    all_lots   = sorted({r['lot'] for r in runs})
    all_mats   = sorted({r.get('material', '') for r in runs if r.get('material', '')})
    _has_no_mat = any(not r.get('material') for r in runs)
    if _has_no_mat:
        all_mats.append('')  # sentinel for runs with no material tag
    all_ibins  = sorted({ib for r in runs for ib in r.get('bin_counts', {})})
    fail_ibins = [ib for ib in all_ibins if ib not in _PASS_BINS]

    # lot -> [wafers]
    lot_wafers: dict[str, list[str]] = {}
    for r in runs:
        lot_wafers.setdefault(r['lot'], set()).add(r.get('wafer', ''))
    lot_wafers = {lot: sorted(ws) for lot, ws in lot_wafers.items()}

    # Build fb_map: FB number str -> {cat, desc} from Pass-Bin-Map and fB93xx
    _raw_cfg = (cfg or {}).get('raw', {})
    _fb_map = {}
    for fb_str, info in _raw_cfg.get('Pass-Bin-Map', {}).items():
        _fb_map[str(fb_str)] = {'cat': info.get('cat', ''), 'desc': info.get('desc', '')}
    for entry in _raw_cfg.get('fB93xx', []):
        fb_str = str(entry.get('FB', ''))
        if fb_str:
            _fb_map[fb_str] = {'cat': entry.get('name', ''), 'desc': entry.get('description', '')}

    data_js = json.dumps({
        'runs':         runs_json_list,
        'ibin_names':   {str(k): v for k, v in ibin_names.items()},
        'bin_map':      _raw_cfg.get('bin_map', {}),
        'fb_map':       _fb_map,
        'yield_target': {k: v for k, v in yield_target.items()},
        'pass_bins':    list(_PASS_BINS),
        'ff_bins':      list(_FF_BINS),
        'ff_df_bins':   list(_FF_DF_BINS),
        'palette':      _FAIL_PALETTE,
        'chart_name':   (cfg or {}).get('name', ''),
        'ff_name':      (cfg or {}).get('ff_name', 'SDS FF'),
        'ff_df_name':   (cfg or {}).get('ff_df_name', 'SDS FF+DF'),
    }, ensure_ascii=False, separators=(',', ':'))
    dlcp_extra_json = json.dumps(runs_dlcp_extra, ensure_ascii=False, separators=(',', ':'))

    cfg_note = (f' &nbsp;|&nbsp; Config: <code>{Path(cfg_path).name}</code>'
                if cfg_path else '')

    # ── Build program + ibin checkbox HTML ────────────────────────────────
    prog_checks = ''.join(
        f'<label class="cb-lbl">'
        f'<input type="checkbox" class="prog-cb" value="{p}" checked> '
        f'<span>{p}</span></label>'
        for p in all_progs
    )
    mat_checks = ''.join(
        f'<label class="cb-lbl">'
        f'<input type="checkbox" class="mat-cb" value="{m}" checked> '
        f'<span>{m if m else "(none)"}</span></label>'
        for m in all_mats
    )
    ibin_checks = ''.join(
        f'<label class="cb-lbl" data-fail="{str(ib not in _PASS_BINS).lower()}">'
        f'<input type="checkbox" class="ibin-cb" value="{ib}" checked> '
        f'<span>iBin {ib}'
        + (f' — {ibin_names[ib]}' if ib in ibin_names else '')
        + '</span></label>'
        for ib in all_ibins
    )
    def _wafer_items(lot):
        items = []
        for w in lot_wafers[lot]:
            val   = f'{lot}::{w}'
            label = w or '(no wafer)'
            # Get material for this wafer if available
            wafer_key = f'{lot}::{w}'
            wafer_mat = wafer_materials.get(wafer_key, '')
            wafer_title = f'Wafer: {label}' + (f' - {wafer_mat}' if wafer_mat else '')
            items.append(
                f'<label class="cb-lbl wafer-item" data-lot="{lot}" title="{wafer_title}">'
                f'<input type="checkbox" class="wafer-cb" value="{val}" checked> '
                f'<span>{label}</span></label>'
            )
        return ''.join(items)

    # lot -> material (first non-empty value seen)
    lot_material: dict[str, str] = {}
    # lot::wafer -> material (per-wafer material tracking)
    wafer_materials: dict[str, str] = {}
    # lot -> set of unique materials for this lot
    lot_materials_set: dict[str, set[str]] = {}
    for r in runs:
        lot = r['lot']
        wafer = r.get('wafer', '')
        material = r.get('material', '')
        if lot not in lot_material:
            lot_material[lot] = material
        wafer_key = f'{lot}::{wafer}'
        if wafer_key not in wafer_materials:
            wafer_materials[wafer_key] = material
        if lot not in lot_materials_set:
            lot_materials_set[lot] = set()
        if material:
            lot_materials_set[lot].add(material)

    # Group lots by first 7 chars of sort_lot for cleaner sidebar display
    from collections import OrderedDict as _OD
    lot_groups: dict[str, list[str]] = _OD()
    for lot in all_lots:
        prefix = lot[:7]
        lot_groups.setdefault(prefix, []).append(lot)

    def _material_wafer_items(lot, material):
        """Return wafers for a specific lot+material combination."""
        items = []
        for w in lot_wafers[lot]:
            wafer_key = f'{lot}::{w}'
            wafer_mat = wafer_materials.get(wafer_key, '')
            # Only include wafers with matching material
            if wafer_mat == material:
                val = wafer_key
                label = w or '(no wafer)'
                wafer_title = f'Wafer: {label}' + (f' - {wafer_mat}' if wafer_mat else '')
                items.append(
                    f'<label class="cb-lbl wafer-item" data-lot="{lot}" title="{wafer_title}">'
                    f'<input type="checkbox" class="wafer-cb" value="{val}" checked> '
                    f'<span>{label}</span></label>'
                )
        return ''.join(items)

    def _lot_group_html(prefix: str, lots_in_group: list[str]) -> str:
        # Derive material for the prefix group (first non-empty value among lots)
        prefix_mat = next((lot_material[l] for l in lots_in_group if lot_material.get(l)), '')
        single = len(lots_in_group) == 1 and lots_in_group[0] == prefix
        # Build individual lot rows
        lot_rows = ''
        for lot in lots_in_group:
            mat = lot_material.get(lot, '')
            mat_span = f' <span style="color:#7fb3d3;font-size:10px">({mat})</span>' if mat else ''
            mat_title = f' - {mat}' if mat else ''
            
            # Check if lot has multiple materials (for lot-grouping mode)
            lot_has_multi_mats = len(lot_materials_set.get(lot, set())) > 1
            
            if grouping_mode == 'lot' and lot_has_multi_mats:
                # Build material nesting for lots with multiple materials
                material_rows = ''
                for material in sorted(lot_materials_set.get(lot, [])):
                    if material:  # Skip empty materials
                        material_rows += (
                            f'<div class="material-row" style="margin-left:20px">'
                            f'<input type="checkbox" class="material-cb" id="mat-cb-{lot}-{material}" '
                            f'value="{lot}::{material}" checked>'
                            f'<label for="mat-cb-{lot}-{material}" style="color:#7fb3d3">{material}</label>'
                            f'<span class="wafer-arrow" onclick="toggleMaterialWafers(this)" '
                            f'data-lot="{lot}" data-material="{material}">&#9654;</span>'
                            f'</div>'
                            f'<div class="material-drop" id="mdrop-{lot}-{material}" style="display:none;margin-left:20px">'
                            f'{_material_wafer_items(lot, material)}'
                            f'</div>'
                        )
                lot_rows += (
                    f'<div class="lot-row">'
                    f'<input type="checkbox" class="lot-cb" id="lot-cb-{lot}" value="{lot}" checked onchange="toggleLotWafers(this)">'
                    f'<label for="lot-cb-{lot}" class="lot-label" title="Lot: {lot}{mat_title}">{lot}{mat_span}</label>'
                    f'<span class="wafer-arrow" onclick="toggleWaferDrop(this)" data-lot="{lot}">&#9654;</span>'
                    f'</div>'
                    f'<div class="wafer-drop" id="wdrop-{lot}" style="display:none">'
                    f'{material_rows}'
                    f'</div>'
                )
            else:
                # Standard flat wafer list (wafer mode or single-material lot)
                lot_rows += (
                    f'<div class="lot-row">'
                    f'<input type="checkbox" class="lot-cb" id="lot-cb-{lot}" value="{lot}" checked onchange="toggleLotWafers(this)">'
                    f'<label for="lot-cb-{lot}" class="lot-label" title="Lot: {lot}{mat_title}">{lot}{mat_span}</label>'
                    f'<span class="wafer-arrow" onclick="toggleWaferDrop(this)" data-lot="{lot}">&#9654;</span>'
                    f'</div>'
                    f'<div class="wafer-drop" id="wdrop-{lot}">{_wafer_items(lot)}</div>'
                )
        
        if single:
            # Only one lot in group — no extra nesting; still show material on the row
            return f'<div class="lot-group" data-prefix="{prefix}">{lot_rows}</div>'
        # Multiple lots share this prefix — add group header with material
        grp_id = f'lotgrp-{prefix}'
        mat_tag = f' <span style="color:#7fb3d3;font-size:10px">({prefix_mat})</span>' if prefix_mat else ''
        prefix_title = f'{prefix}... ({len(lots_in_group)} lots)' + (f' - {prefix_mat}' if prefix_mat else '')
        return (
            f'<div class="lot-group" data-prefix="{prefix}">'
            f'<div class="lot-group-hdr">'
            f'<input type="checkbox" class="lot-grp-cb" data-grp="{prefix}" checked onchange="toggleLotGroup(this)">'
            f'<span class="lot-group-lbl" onclick="toggleLotGroupDrop(\'{grp_id}\')" title="{prefix_title}">'
            f'{prefix}&#8230;{mat_tag} ({len(lots_in_group)} lots) &#9654;</span>'
            f'</div>'
            f'<div class="lot-group-drop" id="{grp_id}" style="display:none">'
            + lot_rows +
            f'</div>'
            f'</div>'
        )

    lot_wafer_checks = ''.join(
        _lot_group_html(prefix, lots)
        for prefix, lots in lot_groups.items()
    )

    # ── Pareto table HTML ──────────────────────────────────────────────────
    def _build_pareto_table(rows):
        if not rows:
            return ''
        toolbar = (
            '<div class="pareto-comment-toolbar">'
            '<button class="btn-comment-action" onclick="exportParetoTableCsv()">&#8681; Export Table CSV</button>'
            '<button class="btn-comment-action" onclick="exportComments()">&#8681; Export Comments CSV</button>'
            '<label class="btn-comment-action" style="cursor:pointer">'
            '&#8679; Import Comments CSV'
            '<input type="file" accept=".csv" style="display:none" onchange="importComments(this)">'
            '</label>'
            '</div>'
        )
        hdr = ('<table class="pareto-tbl" id="pareto-summary-tbl"><thead><tr>'
               '<th>IB</th><th>Description</th>'
               '<th>Total Tested (Wafers / Dies)</th><th>N Fail</th><th>Fail (%)</th><th>Comment</th></tr></thead><tbody>')
        grand_dies   = sum(r.get('total_dies', 0) for r in runs) or 1
        grand_wafers = sum(r.get('wafer_count', 1) for r in runs) or 1
        tested_cell  = f'{grand_wafers:,} / {grand_dies:,}'
        body = ''.join(
            f'<tr><td>{r["ib"]}</td>'
            f'<td>{(r["cat"] + " \u2014 " + r["desc"]) if (r["cat"] and r["desc"] and r["cat"] != r["desc"]) else (r["cat"] or r["desc"])}</td>'
            f'<td>{tested_cell}</td>'
            f'<td>{r["n_fail"]}</td><td>{r["pct"]:.2f}%</td>'
            f'<td><textarea class="pareto-comment" data-ib="{r["ib"]}" rows="1" placeholder="Add comment..."></textarea></td></tr>'
            for r in rows
        )
        return toolbar + hdr + body + '</tbody></table>'

    pareto_table_html = _build_pareto_table(pareto_table_rows or [])

    _pv_lots   = len({r.get('lot', '') for r in runs})
    _pv_wafers = sum(r.get('wafer_count', 1) for r in runs)
    _pv_dies   = sum(r.get('total_dies', 0) for r in runs)
    pareto_v_totals_html = f'Lots: <b>{_pv_lots:,}</b> &nbsp;|&nbsp; Wafers: <b>{_pv_wafers:,}</b> &nbsp;|&nbsp; Dies: <b>{_pv_dies:,}</b>'

    # ── Plotly JS: embed inline (no CDN/network dependency) ─────────────────
    _lib_dir = _find_repo_root(Path(__file__).parent) / 'shared' / 'library'
    _plotly_js_files = sorted(_lib_dir.glob('plotly*.min.js')) if _lib_dir.exists() else []
    if _plotly_js_files:
        # Prefer the smallest build present (e.g. cartesian-only) — this report
        # only uses bar/scatter traces, so the full bundle is unneeded bloat.
        _chosen_plotly = min(_plotly_js_files, key=lambda p: p.stat().st_size)
        _plotly_js_content = _chosen_plotly.read_text(encoding='utf-8')
        _plotly_inline = f'<script>{_plotly_js_content}</script>'
        _plotly_inline_at_end = ''
        print(f'Using local Plotly: {_chosen_plotly.name} ({_chosen_plotly.stat().st_size:,} bytes)')
    else:
        _plotly_inline = '<script src="https://cdn.plot.ly/plotly-latest.min.js"></script>'
        _plotly_inline_at_end = ''
        print('Using Plotly CDN fallback')

    # ── Static chart embeds ───────────────────────────────────────────────────
    # Plotly 6.x injects "responsive":true into Plotly.newPlot configs.
    # Plotly.js 3.x handles this correctly, so strip is just a safety measure.
    from plotly.offline import plot as _plotly_plot
    import re as _re
    def _make_div(fig, cfg=None):
        d = _plotly_plot(fig, output_type='div', include_plotlyjs=False,
                         config=cfg or {'displayModeBar': True})
        d = _re.sub(r',?\s*"responsive"\s*:\s*true', '', d)
        m = _re.search(r'<div id="([^"]+)"[^>]*class="plotly-graph-div"', d)
        return d, (m.group(1) if m else '')

    trend_div,      _trend_id    = _make_div(trend_fig,
                                              {'displayModeBar': True, 'scrollZoom': True})
    pareto_div,     _pareto_h_id = _make_div(pareto_fig)
    if pareto_vertical_fig is not None:
        pareto_vert_div, _pareto_v_id = _make_div(pareto_vertical_fig)
    else:
        pareto_vert_div, _pareto_v_id = '<p style="color:#888">Not available</p>', ''

    html = f'''<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>iBin Fail vs Yield Trend — {csv_path.name}</title>
{_plotly_inline}
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:Arial,sans-serif;background:#f0f3f7;display:flex;height:100vh;overflow:hidden}}

/* ── Sidebar ── */
#sidebar{{width:280px;min-width:220px;background:#1a252f;color:#ecf0f1;
  display:flex;flex-direction:column;overflow:hidden;flex-shrink:0}}
#sidebar-header{{padding:12px 14px 8px;border-bottom:1px solid #2c3e50}}
#sidebar-header h1{{font-size:14px;color:#3498db;line-height:1.3}}
#sidebar-header .meta{{font-size:10px;color:#7f8c8d;margin-top:3px}}
#sidebar-watermark{{padding:4px 14px;font-size:10px;font-weight:bold;color:#3498db;border-bottom:1px solid #2c3e50}}
#sidebar-body{{overflow-y:auto;flex:1;padding:0 0 12px}}
#sidebar-body::-webkit-scrollbar{{width:4px}}
#sidebar-body::-webkit-scrollbar-thumb{{background:#2c3e50}}

.ctrl-section{{padding:8px 12px 4px}}
.ctrl-section h3{{font-size:11px;font-weight:bold;color:#95a5a6;
  text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px}}
.sep{{height:1px;background:#2c3e50;margin:4px 10px}}

/* interval radios */
.iv-row{{display:flex;flex-wrap:wrap;gap:4px}}
.iv-row label{{font-size:11px;cursor:pointer;padding:3px 8px;border-radius:4px;
  background:#243342;color:#bdc3c7}}
.date-range-row{{display:flex;flex-wrap:wrap;gap:4px 10px;font-size:11px}}
.date-range-row label{{display:flex;align-items:center;gap:3px;cursor:pointer;color:#bdc3c7}}
.date-range-row{{display:flex;flex-wrap:wrap;gap:4px 10px;font-size:11px}}
.date-range-row label{{display:flex;align-items:center;gap:3px;cursor:pointer;color:#bdc3c7}}
.iv-row input[type=radio]{{display:none}}
.iv-row input[type=radio]:checked + span{{color:#fff;font-weight:bold}}
.iv-row label:has(input:checked){{background:#2980b9}}

/* filter btn row */
.btn-row{{display:flex;gap:4px;margin-bottom:6px}}
.btn-row button{{flex:1;font-size:10px;padding:3px 0;border:none;border-radius:3px;
  cursor:pointer;font-weight:bold}}
.btn-all{{background:#1f618d;color:white}}
.btn-none{{background:#555;color:white}}
.btn-fail{{background:#7d1f1f;color:white}}

/* checkboxes */
.cb-list{{max-height:220px;overflow-y:scroll;background:#243342;border-radius:4px;
  padding:4px;scrollbar-width:thin;scrollbar-color:#3498db #1a252f}}
.cb-list::-webkit-scrollbar{{width:10px}}
.cb-list::-webkit-scrollbar-track{{background:#1a252f;border-radius:4px}}
.cb-list::-webkit-scrollbar-thumb{{background:#3498db;border-radius:4px;min-height:40px}}
.cb-list::-webkit-scrollbar-thumb:hover{{background:#5dade2}}
.cb-lbl{{display:flex;align-items:center;gap:5px;font-size:11px;color:#bdc3c7;
  padding:2px 4px;cursor:pointer;border-radius:3px}}
.cb-lbl:hover{{background:#2c3e50}}
.cb-lbl input{{cursor:pointer;accent-color:#3498db}}

/* FB drilldown table/wafermap toggle */
.fbdd-tab-btn{{font-size:11px;padding:3px 10px;margin-right:4px;cursor:pointer;
  border:1px solid #aaa;border-radius:3px;background:#f5f5f5}}
.fbdd-tab-btn.active{{background:#1f618d;color:white;border-color:#1f618d}}
.fbdd-tab-btn:disabled{{opacity:0.45;cursor:not-allowed}}
/* Wafer-map wafer picker (sidebar-style dropdown-with-checkboxes) */
.wm-dd-panel{{position:absolute;top:100%;left:0;z-index:70;background:#fff;
  border:1px solid #ccd3db;border-radius:5px;box-shadow:0 4px 14px rgba(0,0,0,.18);
  padding:8px;min-width:200px;margin-top:2px}}
.wm-dd-panel .cb-lbl{{color:#333;font-size:11px;white-space:nowrap;cursor:pointer;padding:2px 0}}
.cb-lbl span{{white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
/* lot-wafer nested */
.lot-group{{margin-bottom:2px}}
.lot-group-hdr{{display:flex;align-items:center;gap:5px;padding:2px 4px;border-radius:3px;cursor:pointer}}
.lot-group-hdr:hover{{background:#2c3e50}}
.lot-group-lbl{{font-size:11px;color:#aed6f1;font-weight:bold;flex:1;cursor:pointer;white-space:nowrap}}
.lot-group-drop{{padding-left:12px}}
.lot-row{{display:flex;align-items:center;gap:5px;padding:2px 4px;border-radius:3px}}
.lot-row:hover{{background:#2c3e50}}
.lot-label{{font-size:11px;color:#ecf0f1;cursor:pointer;flex:1;white-space:nowrap;
  overflow:hidden;text-overflow:ellipsis}}
.wafer-arrow{{font-size:9px;color:#7f8c8d;cursor:pointer;user-select:none;
  padding:0 2px;flex-shrink:0}}
.wafer-arrow:hover{{color:#3498db}}
.wafer-drop{{display:none;padding-left:18px;margin-top:2px;background:#1a252f;
  border-left:2px solid #2c3e50;margin-left:0px}}
.wafer-drop.show{{display:block}}
.wafer-drop .cb-lbl{{font-size:10px;color:#95a5a6;padding-left:4px}}

/* options */
.opt-grid{{display:grid;grid-template-columns:auto 1fr;gap:4px 8px;align-items:center}}
.opt-grid label{{font-size:11px;color:#bdc3c7}}
.opt-grid input{{background:#243342;border:1px solid #2c3e50;color:white;
  border-radius:3px;padding:2px 6px;font-size:11px;width:70px}}

/* generate button */
#gen-btn{{margin:10px 12px 4px;padding:8px;background:#27ae60;color:white;
  font-size:13px;font-weight:bold;border:none;border-radius:5px;cursor:pointer;width:calc(100% - 24px)}}
#gen-btn:hover{{background:#2ecc71}}
#gen-btn:active{{background:#1e8449}}

/* stats bar */
#stats-bar{{display:flex;gap:6px;padding:4px 12px;flex-wrap:wrap}}
.stat-chip{{background:#2c3e50;border-radius:4px;padding:2px 8px;font-size:10px;color:#7f8c8d}}
.stat-chip b{{color:#3498db}}

/* ── Main area ── */
#main{{flex:1;display:flex;flex-direction:column;overflow:hidden}}
#tabs{{display:flex;background:#fff;border-bottom:2px solid #dce3eb;padding:0 16px}}
.tab{{padding:10px 18px;cursor:pointer;font-size:13px;color:#7f8c8d;
  border-bottom:2px solid transparent;margin-bottom:-2px}}
.tab.active{{color:#2980b9;border-bottom-color:#2980b9;font-weight:bold}}
#tab-content{{flex:1;overflow:auto;padding:12px 16px;display:flex;flex-direction:column}}

/* chart containers */
.chart-card{{background:white;border-radius:8px;padding:12px;margin-bottom:14px;
  box-shadow:0 1px 4px rgba(0,0,0,.08);display:flex;flex-direction:column}}
.chart-card h2{{font-size:15px;color:#2c3e50;margin-bottom:8px;flex-shrink:0}}
.chart-wrap{{width:100%;height:520px;min-height:200px;min-width:300px;
  position:relative;resize:both;overflow:hidden;box-sizing:border-box;
  border:1px solid #e0e0e0;border-radius:4px}}
.chart-wrap > div{{width:100% !important;height:100% !important}}

/* run table */
#run-table{{width:100%;border-collapse:collapse;font-size:12px}}
#run-table th{{background:#2c3e50;color:white;padding:6px 10px;text-align:left;
  position:sticky;top:0;z-index:2}}
#run-table td{{padding:4px 10px;border-bottom:1px solid #e8eaed}}
#run-table tr:hover td{{background:#f0f4f8}}
.yld-ok{{color:#27ae60;font-weight:bold}}
.yld-mid{{color:#f39c12;font-weight:bold}}
.yld-low{{color:#e74c3c;font-weight:bold}}
code{{background:#eef;padding:1px 4px;border-radius:3px;font-size:11px}}

/* pareto summary table */
.fb-drill-tbl{{width:100%;border-collapse:collapse;font-size:12px;margin-top:4px}}
.fb-drill-tbl th{{background:#1a5276;color:white;padding:6px 10px;text-align:left;font-weight:bold;position:sticky;top:0}}
.fb-drill-tbl th.num{{text-align:right}}
.fb-drill-tbl td{{padding:5px 10px;border-bottom:1px solid #e8eaed;vertical-align:middle}}
.fb-drill-tbl td.num{{text-align:right;font-variant-numeric:tabular-nums}}
.fb-drill-tbl tr:hover td{{background:#eaf4fb}}
.fb-drill-tbl td:first-child,.fb-drill-tbl td:nth-child(2){{font-weight:bold;color:#1a5276}}
.pareto-tbl{{width:100%;border-collapse:collapse;font-size:12px;margin-top:16px}}
.pareto-tbl th{{background:#2c3e50;color:white;padding:6px 10px;text-align:left;font-weight:bold}}
.pareto-tbl td{{padding:4px 10px;border-bottom:1px solid #e8eaed;vertical-align:middle}}
.pareto-tbl tr:hover td{{background:#f0f4f8}}
.pareto-tbl td:first-child{{font-weight:bold;color:#1a5276}}
.pareto-comment-toolbar{{display:flex;gap:8px;margin-bottom:8px;flex-wrap:wrap}}
.btn-comment-action{{background:#2c3e50;color:#ecf0f1;border:1px solid #3d5166;border-radius:4px;
  padding:4px 10px;font-size:11px;cursor:pointer;transition:background 0.15s}}
.btn-comment-action:hover{{background:#3d5166}}
.pareto-comment{{width:100%;min-width:160px;padding:4px 6px;border:1px solid #ddd;border-radius:4px;
  font-size:12px;font-family:inherit;resize:vertical;background:#fffef0;
  transition:border-color 0.2s}}
.pareto-comment:focus{{outline:none;border-color:#1a73e8;background:#fff}}
.pareto-comment.saved{{border-color:#27ae60;background:#f0fff4}}
/* resizable columns */
th.resizable{{position:relative;overflow:visible}}
th.resizable .col-resizer{{position:absolute;right:0;top:0;bottom:0;width:5px;
  cursor:col-resize;user-select:none;z-index:3;background:transparent}}
th.resizable .col-resizer:hover{{background:rgba(255,255,255,0.3)}}

/* ── DLCP (exact match to bin_distribution_html.py) ─────────────────── */
.dlcp-ctrl{{display:flex;align-items:center;gap:12px;flex-wrap:wrap;background:#fff;padding:7px 12px;border-radius:6px;box-shadow:0 1px 4px rgba(0,0,0,.1);flex-shrink:0}}
.dlcp-sumbox{{background:transparent;border-radius:0;padding:0;box-shadow:none;flex-shrink:0;display:flex;flex-direction:column;gap:6px;align-items:stretch}}
.dlcp-sum-panel{{background:#fff;border-radius:6px;padding:8px 14px;box-shadow:0 1px 4px rgba(0,0,0,.1);display:flex;flex-direction:column;gap:4px;min-width:0}}
.dlcp-sum-panel-ttl{{font-size:15px;font-weight:bold;text-transform:uppercase;letter-spacing:.7px;color:#fff;background:#5d6d7e;border-radius:3px;padding:1px 8px;margin-bottom:4px;align-self:flex-start}}
.dlcp-sumrow{{display:flex;gap:10px;flex-wrap:wrap;align-items:center}}
.dlcp-sum-grp{{display:flex;flex-direction:column;padding:4px 14px;border-left:3px solid #dde;min-width:110px}}
.dlcp-sum-grp.pass{{border-color:#2980b9}}.dlcp-sum-grp.marg{{border-color:#d4ac0d}}.dlcp-sum-grp.fail{{border-color:#c0392b}}
.dlcp-sum-lbl{{font-size:17px;color:#000;text-transform:uppercase;letter-spacing:.5px;margin-bottom:2px}}
.dlcp-sum-val{{font-size:26px;font-weight:bold;color:#2c3e50}}.dlcp-sum-pct{{font-size:17px;color:#666;margin-left:4px}}
.dlcp-sum-pct-big{{font-size:33px;font-weight:bold;line-height:1.1}}
.dlcp-sum-sub{{font-size:15px;color:#aaa;margin-top:1px}}
.dlcp-inner{{display:flex;gap:0;flex:1;min-height:0}}
.dlcp-left{{display:flex;flex-direction:column;gap:6px;min-width:0;flex:1;overflow:hidden}}
.dlcp-panel-hdr{{display:flex;align-items:center;gap:5px;flex-shrink:0}}
.dlcp-pbtn{{background:#ecf0f1;border:1px solid #bdc3c7;border-radius:3px;font-size:11px;padding:1px 7px;cursor:pointer;color:#2c3e50;white-space:nowrap}}
.dlcp-pbtn:hover{{background:#d5dbde}}
.dlcp-flt-row input{{width:100%;box-sizing:border-box;font-size:11px;padding:2px 4px;border:1px solid #ccd;border-radius:2px}}
.dlcp-sec-ttl{{font-size:11px;font-weight:bold;color:#5d6d7e;text-transform:uppercase;letter-spacing:.5px;flex-shrink:0}}
.dlcp-tw{{overflow:auto;background:#fff;border-radius:6px;padding:6px;box-shadow:0 1px 4px rgba(0,0,0,.1);flex:1;min-height:0}}
#dlcp-tbl-pane{{display:flex;flex-direction:column;flex:1;min-height:0;overflow:hidden}}
.dlcp-t{{border-collapse:collapse;font-size:12px;white-space:nowrap;width:100%}}
.dlcp-t th{{background:#2c3e50;color:#ecf0f1;padding:5px 10px;text-align:left;position:sticky;top:0;z-index:1}}
.dlcp-t td{{padding:4px 10px;border-bottom:1px solid #eee}}
.dlcp-t tr:nth-child(even) td{{background:#f7f9fc}}.dlcp-t tr:hover td{{background:#eaf3fb}}
.dlcp-t tr.dlcp-rsel td{{background:#d0eaff!important;font-weight:bold}}
.dlcp-t tr.dlcp-runsel td{{opacity:.4}}
.dlcp-t tr{{cursor:pointer}}
.dlcp-ddbtn{{background:none;border:none;color:#aed6f1;cursor:pointer;font-size:10px;padding:0 2px;vertical-align:middle;margin-left:3px}}
.dlcp-ddbtn.on{{color:#f1c40f}}
.dlcp-dd{{position:fixed;background:#fff;border:1px solid #aaa;border-radius:4px;box-shadow:0 4px 16px rgba(0,0,0,.2);z-index:30000;min-width:160px;max-width:260px;font-size:12px;color:#2c3e50}}
.dlcp-dd-srch{{width:100%;box-sizing:border-box;padding:5px 8px;border:none;border-bottom:1px solid #ddd;font-size:12px;outline:none}}
.dlcp-dd-panel{{position:fixed;z-index:9999;background:#fff;border:1px solid #c8cdd5;border-radius:5px;box-shadow:0 4px 16px rgba(0,0,0,.18);min-width:200px;max-width:320px;display:none;flex-direction:column}}
.dlcp-dd-acts{{display:flex;gap:4px;padding:4px 6px;border-bottom:1px solid #eee}}
.dlcp-dd-acts button{{flex:1;padding:2px 6px;font-size:11px;cursor:pointer;border:1px solid #bdc3c7;background:#ecf0f1;border-radius:3px}}
.dlcp-dd-list{{max-height:200px;overflow-y:auto;padding:4px 0}}
.dlcp-dd-item{{display:flex;align-items:center;gap:6px;padding:3px 10px;cursor:pointer}}
.dlcp-dd-item:hover{{background:#eaf0fb}}
.dlcp-dd-foot{{padding:4px 8px;border-top:1px solid #eee;text-align:right}}
.dlcp-dd-foot button{{padding:3px 12px;font-size:11px;cursor:pointer;background:#2c3e50;color:#fff;border:none;border-radius:3px}}
.dlcp-splitter{{width:14px;flex-shrink:0;display:flex;align-items:center;justify-content:center;cursor:pointer;background:#e8ecf0;border-left:1px solid #d0d8e8;border-right:1px solid #d0d8e8;transition:background .15s;user-select:none}}
.dlcp-splitter:hover{{background:#c8d4e8}}
.dlcp-split-arrow{{font-size:12px;color:#5d6d7e}}
.dlcp-cw{{flex:1;background:#fff;border-radius:6px;padding:8px;box-shadow:0 1px 4px rgba(0,0,0,.1);display:flex;flex-direction:column;min-width:0;overflow:hidden;margin-left:8px}}
.dlcp-t .num{{text-align:right}}
.dlcp-note{{font-size:10px;color:#666;background:#f8f9fa;border:1px solid #e4e4e4;border-radius:4px;padding:5px 10px;line-height:1.8;flex-shrink:0}}
.dlcp-note b{{color:#444}}
.upm-hist-overlay{{display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.45);z-index:32000;align-items:center;justify-content:center}}
.upm-hist-overlay.open{{display:flex}}
.upm-hist-box{{background:#f0f2f5;border-radius:8px;box-shadow:0 8px 32px rgba(0,0,0,.4);width:800px;max-width:96vw;height:540px;max-height:92vh;display:flex;flex-direction:column;overflow:hidden;resize:both;min-width:420px;min-height:340px}}
.upm-hist-drag{{cursor:move;background:#1f618d;color:#fff;padding:7px 14px;display:flex;align-items:center;justify-content:space-between;user-select:none;flex-shrink:0}}
.upm-hist-body{{display:flex;flex-direction:column;flex:1;padding:10px;gap:8px;min-height:0;overflow:hidden}}
.upm-hist-stats{{display:flex;flex-wrap:wrap;gap:6px;background:#fff;border-radius:6px;padding:8px 12px;box-shadow:0 1px 4px rgba(0,0,0,.1);flex-shrink:0}}
.upm-hist-stat-grp{{display:flex;flex-direction:column;padding:3px 12px;border-left:3px solid #dde;min-width:100px}}
.upm-hist-stat-lbl{{font-size:10px;color:#888;text-transform:uppercase;letter-spacing:.4px;font-weight:bold}}
.upm-hist-stat-val{{font-size:14px;font-weight:bold;color:#2c3e50}}
.upm-hist-cv-wrap{{flex:1;min-height:0;display:flex}}
</style>
</head>
<body>

<!-- ═══ SIDEBAR ═══ -->
<div id="sidebar">
  <div id="sidebar-header">
    <h1>&#128200; iBin Fail vs. Yield Trend</h1>
    <div class="meta">
      <b>{csv_path.name}</b>{cfg_note}<br>
      Generated: {ts_now}
    </div>
  </div>
  <div id="sidebar-watermark">Pant, Sujit N &mdash; GEMS FTE</div>
  <div id="sidebar-body">

    <!-- Grouping -->
    <div class="ctrl-section">
      <h3>Grouping</h3>
      <div class="iv-row">
        <label><input type="radio" name="groupby" value="lot" checked><span>Program / Lot</span></label>
        <label><input type="radio" name="groupby" value="wafer"><span>Program / Lot / Wafer</span></label>
      </div>
    </div>
    <div class="sep"></div>

    <!-- Interval -->
    <div class="ctrl-section">
      <h3>Interval</h3>
      <div class="iv-row">
        {''.join(f'<label><input type="radio" name="interval" value="{iv}"{"checked" if iv==interval else ""}><span>{iv.capitalize()}</span></label>' for iv in INTERVALS)}
      </div>
    </div>
    <div class="sep"></div>

    <!-- Lot / Wafer filter -->
    <div class="ctrl-section">
      <h3>Lot / Wafer</h3>
      <input type="text" id="lot-search" placeholder="Filter lots..." 
        style="width:100%;box-sizing:border-box;background:#243342;border:1px solid #2c3e50;
        color:#bdc3c7;padding:4px 6px;border-radius:3px;font-size:11px;margin-bottom:6px"
        oninput="filterLots(this.value)">
      <div class="btn-row">
        <button class="btn-all" onclick="lotWaferAll(true)">All</button>
        <button class="btn-none" onclick="lotWaferAll(false)">None</button>
      </div>
      <div class="cb-list" id="lot-wafer-list" style="max-height:350px">{lot_wafer_checks}</div>
    </div>
    <div class="sep"></div>

    <!-- Material filter -->
    <div class="ctrl-section">
      <h3>Material</h3>
      <input type="text" id="mat-search" placeholder="Filter materials..." 
        style="width:100%;box-sizing:border-box;background:#243342;border:1px solid #2c3e50;
        color:#bdc3c7;padding:4px 6px;border-radius:3px;font-size:11px;margin-bottom:6px"
        oninput="filterMats(this.value)">
      <div class="btn-row">
        <button class="btn-all" onclick="selAll('.mat-cb')">All</button>
        <button class="btn-none" onclick="selNone('.mat-cb')">None</button>
      </div>
      <div class="cb-list" id="mat-list">{mat_checks}</div>
    </div>
    <div class="sep"></div>

    <!-- Program filter -->
    <div class="ctrl-section">
      <h3>Test Program</h3>
      <div class="btn-row">
        <button class="btn-all" onclick="selAll('.prog-cb')">All</button>
        <button class="btn-none" onclick="selNone('.prog-cb')">None</button>
      </div>
      <input type="text" placeholder="Search programs…" oninput="sidebarFilterList('prog-list',this.value)" style="font-size:11px;width:100%;box-sizing:border-box;margin-bottom:4px;padding:2px 4px">
      <div class="cb-list" id="prog-list">{prog_checks}</div>
    </div>
    <div class="sep"></div>

    <!-- Interface Bin filter -->
    <div class="ctrl-section">
      <h3>Interface Bin</h3>
      <div class="btn-row">
        <button class="btn-all" onclick="selAll('.ibin-cb')">All</button>
        <button class="btn-none" onclick="selNone('.ibin-cb')">None</button>
        <button class="btn-fail" onclick="selFail()">Fail only</button>
      </div>
      <input type="text" placeholder="Search bins…" oninput="sidebarFilterList('ibin-list',this.value)" style="font-size:11px;width:100%;box-sizing:border-box;margin-bottom:4px;padding:2px 4px">
      <div class="cb-list" id="ibin-list">{ibin_checks}</div>
    </div>
    <div class="sep"></div>

    <!-- Date Range -->
    <div class="ctrl-section">
      <h3>Date Range</h3>
      <div class="date-range-row">
        <label><input type="radio" name="datemode" value="all" checked> All</label>
        <label><input type="radio" name="datemode" value="4w"> 4 wks</label>
        <label><input type="radio" name="datemode" value="6w"> 6 wks</label>
        <label><input type="radio" name="datemode" value="12w"> 12 wks</label>
        <label><input type="radio" name="datemode" value="custom"> Custom</label>
      </div>
      <div id="custom-date-row" style="display:none;margin-top:6px">
        <div style="display:grid;grid-template-columns:auto 1fr;gap:4px 6px;align-items:center;font-size:11px">
          <span style="color:#95a5a6">From</span>
          <input type="date" id="date-from" style="background:#2c3e50;color:#ecf0f1;border:1px solid #3d5166;border-radius:3px;padding:2px 4px;font-size:11px;width:100%">
          <span style="color:#95a5a6">To</span>
          <input type="date" id="date-to" style="background:#2c3e50;color:#ecf0f1;border:1px solid #3d5166;border-radius:3px;padding:2px 4px;font-size:11px;width:100%">
        </div>
      </div>
    </div>
    <div class="sep"></div>

    <!-- Options -->
    <div class="ctrl-section">
      <h3>Options</h3>
      <div class="opt-grid">
        <label>Top N iBins</label>
        <input type="number" id="topn-input" value="{top_n}" min="1" max="30">
        <label>Min fail% thresh</label>
        <input type="number" id="thresh-input" value="0.0" min="0" step="0.5">
      </div>
    </div>

    <button id="gen-btn" onclick="rebuildCharts()">&#9654;&#xFE0E; Apply Filters</button>

    <!-- Stats -->
    <div class="sep"></div>
    <div id="stats-bar"></div>
  </div>
</div>

<!-- ═══ MAIN ═══ -->
<div id="main">
  <div id="tabs">
    <div class="tab active" onclick="showTab('trend',this)">&#128204; Trend</div>
    <div class="tab" onclick="showTab('pareto-h',this)">&#128202; Pareto (horizontal)</div>
    <div class="tab" onclick="showTab('pareto-v',this)">&#128202; Pareto (by bin)</div>
    <div class="tab" onclick="showTab('table',this)">&#128209; Run Table</div>
    <div class="tab" onclick="showTab('dlcp',this)">&#9889; DLCP</div>
  </div>
  <div id="tab-content">
    <div id="tab-trend">
      <div class="chart-card">
        <h2 style="margin-bottom:8px">&#128204; iBin Fail Trend
          <button onclick="exportTrendCsv()" style="font-size:11px;margin-left:10px;padding:2px 8px;cursor:pointer;border:1px solid #aaa;border-radius:3px;background:#f5f5f5">&#8681; CSV</button></h2>
        <div id="trend-totals" style="font-size:12px;color:#555;margin:6px 0">{pareto_v_totals_html}</div>
        <div id="trend-chart" class="chart-wrap">{trend_div}</div>
      </div>
      <div class="chart-card" id="trend-fb-drilldown" style="display:none">
        <h2 id="trend-fb-title">Functional Bin Breakdown — IB <span id="trend-fb-ib"></span>
          <button onclick="exportFbDrilldownCsv('trend-fb-thead','trend-fb-tbody','fb_drilldown')" style="font-size:11px;margin-left:10px;padding:2px 8px;cursor:pointer;border:1px solid #aaa;border-radius:3px;background:#f5f5f5">&#8681; CSV</button></h2>
        <div class="fbdd-tabs" style="margin:4px 0 8px">
          <button class="fbdd-tab-btn active" id="trend-fbview-table-btn" onclick="setFbView('trend','table')">Table</button>
          <button class="fbdd-tab-btn" id="trend-fbview-map-btn" onclick="setFbView('trend','map')">Wafer Map</button>
        </div>
        <div id="trend-fbtable-wrap"><div style="overflow-x:auto"><table class="fb-drill-tbl" id="trend-fb-tbl">
          <thead id="trend-fb-thead"><tr><th>Interface Bin</th><th>Lot (Wafers)</th><th>Functional Bin</th><th>Description</th><th>Fail Test Module</th>
            <th class="num">Total Tested</th><th class="num">Fail Count</th><th class="num">Fail %</th></tr></thead>
          <tbody id="trend-fb-tbody"></tbody>
        </table></div></div>
        <div id="trend-wafermap-wrap" style="display:none">
          <div id="trend-wafermap-selrow" style="display:none;margin-bottom:8px;font-size:12px;position:relative">
            <span class="wm-dd-group" style="position:relative;display:inline-block">
              <span onclick="toggleWmDrop(this,'trend','wafer')" style="cursor:pointer;user-select:none;color:#2c3e50;font-weight:600">Wafers <span class="wm-dd-arrow">&#9654;</span></span>
              <span id="trend-wafermap-summary" style="color:#888;margin-left:6px"></span>
              <div id="trend-wafermap-drop" class="wm-dd-panel" style="display:none">
                <div style="margin-bottom:4px"><button onclick="wmCheckAll('trend',true)" style="font-size:11px">All</button> <button onclick="wmCheckAll('trend',false)" style="font-size:11px">None</button></div>
                <input type="text" placeholder="Search wafers…" oninput="wmFilterChecks('trend',this.value)" style="font-size:11px;width:100%;box-sizing:border-box;margin-bottom:4px;padding:2px 4px">
                <div id="trend-wafermap-checks" style="display:flex;flex-direction:column;max-height:180px;overflow-y:auto"></div>
              </div>
            </span>
            <span class="wm-dd-group" id="trend-wm-fb-wrap" style="position:relative;display:none;margin-left:14px">
              <span onclick="toggleWmDrop(this,'trend','fb')" style="cursor:pointer;user-select:none;color:#2c3e50;font-weight:600">FB <span class="wm-dd-arrow">&#9654;</span></span>
              <span id="trend-wafermap-fbsummary" style="color:#888;margin-left:6px"></span>
              <div id="trend-wafermap-fbdrop" class="wm-dd-panel" style="display:none">
                <div style="margin-bottom:4px"><button onclick="wmFbCheckAll('trend',true)" style="font-size:11px">All</button> <button onclick="wmFbCheckAll('trend',false)" style="font-size:11px">None</button></div>
                <div id="trend-wafermap-fbchecks" style="display:flex;flex-direction:column;max-height:180px;overflow-y:auto"></div>
              </div>
            </span>
          </div>
          <div id="trend-wafermap-grid" style="display:flex;flex-wrap:wrap;gap:14px"></div>
        </div>
      </div>
    </div>
    <div id="tab-pareto-h" style="display:none">
      <div class="chart-card">
        <h2>Interface Bin Fail Pareto (horizontal)</h2>
        <div id="pareto-h-totals" style="font-size:12px;color:#555;margin:6px 0">{pareto_v_totals_html}</div>
        <div id="pareto-h-chart" class="chart-wrap">{pareto_div}</div>
      </div>
      <div class="chart-card" id="pareto-h-fb-drilldown" style="display:none">
        <h2 id="pareto-h-fb-title">Functional Bin Breakdown — IB <span id="pareto-h-fb-ib"></span>
          <button onclick="exportFbDrilldownCsv('pareto-h-fb-thead','pareto-h-fb-tbody','fb_drilldown')" style="font-size:11px;margin-left:10px;padding:2px 8px;cursor:pointer;border:1px solid #aaa;border-radius:3px;background:#f5f5f5">&#8681; CSV</button></h2>
        <div class="fbdd-tabs" style="margin:4px 0 8px">
          <button class="fbdd-tab-btn active" id="pareto-h-fbview-table-btn" onclick="setFbView('pareto-h','table')">Table</button>
          <button class="fbdd-tab-btn" id="pareto-h-fbview-map-btn" onclick="setFbView('pareto-h','map')">Wafer Map</button>
        </div>
        <div id="pareto-h-fbtable-wrap"><div style="overflow-x:auto"><table class="fb-drill-tbl" id="pareto-h-fb-tbl">
          <thead id="pareto-h-fb-thead"><tr><th>Interface Bin</th><th>Lot (Wafers)</th><th>Functional Bin</th><th>Description</th><th>Fail Test Module</th>
            <th class="num">Total Tested</th><th class="num">Fail Count</th><th class="num">Fail %</th></tr></thead>
          <tbody id="pareto-h-fb-tbody"></tbody>
        </table></div></div>
        <div id="pareto-h-wafermap-wrap" style="display:none">
          <div id="pareto-h-wafermap-selrow" style="display:none;margin-bottom:8px;font-size:12px;position:relative">
            <span class="wm-dd-group" style="position:relative;display:inline-block">
              <span onclick="toggleWmDrop(this,'pareto-h','wafer')" style="cursor:pointer;user-select:none;color:#2c3e50;font-weight:600">Wafers <span class="wm-dd-arrow">&#9654;</span></span>
              <span id="pareto-h-wafermap-summary" style="color:#888;margin-left:6px"></span>
              <div id="pareto-h-wafermap-drop" class="wm-dd-panel" style="display:none">
                <div style="margin-bottom:4px"><button onclick="wmCheckAll('pareto-h',true)" style="font-size:11px">All</button> <button onclick="wmCheckAll('pareto-h',false)" style="font-size:11px">None</button></div>
                <input type="text" placeholder="Search wafers…" oninput="wmFilterChecks('pareto-h',this.value)" style="font-size:11px;width:100%;box-sizing:border-box;margin-bottom:4px;padding:2px 4px">
                <div id="pareto-h-wafermap-checks" style="display:flex;flex-direction:column;max-height:180px;overflow-y:auto"></div>
              </div>
            </span>
            <span class="wm-dd-group" id="pareto-h-wm-fb-wrap" style="position:relative;display:none;margin-left:14px">
              <span onclick="toggleWmDrop(this,'pareto-h','fb')" style="cursor:pointer;user-select:none;color:#2c3e50;font-weight:600">FB <span class="wm-dd-arrow">&#9654;</span></span>
              <span id="pareto-h-wafermap-fbsummary" style="color:#888;margin-left:6px"></span>
              <div id="pareto-h-wafermap-fbdrop" class="wm-dd-panel" style="display:none">
                <div style="margin-bottom:4px"><button onclick="wmFbCheckAll('pareto-h',true)" style="font-size:11px">All</button> <button onclick="wmFbCheckAll('pareto-h',false)" style="font-size:11px">None</button></div>
                <div id="pareto-h-wafermap-fbchecks" style="display:flex;flex-direction:column;max-height:180px;overflow-y:auto"></div>
              </div>
            </span>
          </div>
          <div id="pareto-h-wafermap-grid" style="display:flex;flex-wrap:wrap;gap:14px"></div>
        </div>
      </div>
    </div>
    <div id="tab-pareto-v" style="display:none">
      <div class="chart-card">
        <h2>Interface Bin Fail Pareto — by bin</h2>
        <div id="pareto-v-chart" class="chart-wrap">{pareto_vert_div}</div>
        <div id="pareto-v-totals" style="font-size:12px;color:#555;margin:6px 0">{pareto_v_totals_html}</div>
        <label style="display:flex;align-items:center;gap:6px;font-size:12px;color:#555;margin:8px 0;cursor:pointer">
          <input type="checkbox" id="pareto-v-tbl-toggle" onchange="document.getElementById('pareto-v-tbl-wrap').style.display=this.checked?'':'none'">
          Show summary table
        </label>
        <div id="pareto-v-tbl-wrap" style="display:none">{pareto_table_html}</div>
      </div>
      <div class="chart-card" id="pareto-v-fb-drilldown" style="display:none">
        <h2 id="pareto-v-fb-title">Functional Bin Breakdown — IB <span id="pareto-v-fb-ib"></span>
          <button onclick="exportFbDrilldownCsv('pareto-v-fb-thead','pareto-v-fb-tbody','fb_drilldown_v')" style="font-size:11px;margin-left:10px;padding:2px 8px;cursor:pointer;border:1px solid #aaa;border-radius:3px;background:#f5f5f5">&#8681; CSV</button></h2>
        <div class="fbdd-tabs" style="margin:4px 0 8px">
          <button class="fbdd-tab-btn active" id="pareto-v-fbview-table-btn" onclick="setFbView('pareto-v','table')">Table</button>
          <button class="fbdd-tab-btn" id="pareto-v-fbview-map-btn" onclick="setFbView('pareto-v','map')">Wafer Map</button>
        </div>
        <div id="pareto-v-fbtable-wrap"><div style="overflow-x:auto"><table class="fb-drill-tbl" id="pareto-v-fb-tbl">
          <thead id="pareto-v-fb-thead"><tr><th>Interface Bin</th><th>Lot (Wafers)</th><th>Functional Bin</th><th>Description</th><th>Fail Test Module</th>
            <th class="num">Total Tested</th><th class="num">Fail Count</th><th class="num">Fail %</th></tr></thead>
          <tbody id="pareto-v-fb-tbody"></tbody>
        </table></div></div>
        <div id="pareto-v-wafermap-wrap" style="display:none">
          <div id="pareto-v-wafermap-selrow" style="display:none;margin-bottom:8px;font-size:12px;position:relative">
            <span class="wm-dd-group" style="position:relative;display:inline-block">
              <span onclick="toggleWmDrop(this,'pareto-v','wafer')" style="cursor:pointer;user-select:none;color:#2c3e50;font-weight:600">Wafers <span class="wm-dd-arrow">&#9654;</span></span>
              <span id="pareto-v-wafermap-summary" style="color:#888;margin-left:6px"></span>
              <div id="pareto-v-wafermap-drop" class="wm-dd-panel" style="display:none">
                <div style="margin-bottom:4px"><button onclick="wmCheckAll('pareto-v',true)" style="font-size:11px">All</button> <button onclick="wmCheckAll('pareto-v',false)" style="font-size:11px">None</button></div>
                <input type="text" placeholder="Search wafers…" oninput="wmFilterChecks('pareto-v',this.value)" style="font-size:11px;width:100%;box-sizing:border-box;margin-bottom:4px;padding:2px 4px">
                <div id="pareto-v-wafermap-checks" style="display:flex;flex-direction:column;max-height:180px;overflow-y:auto"></div>
              </div>
            </span>
            <span class="wm-dd-group" id="pareto-v-wm-fb-wrap" style="position:relative;display:none;margin-left:14px">
              <span onclick="toggleWmDrop(this,'pareto-v','fb')" style="cursor:pointer;user-select:none;color:#2c3e50;font-weight:600">FB <span class="wm-dd-arrow">&#9654;</span></span>
              <span id="pareto-v-wafermap-fbsummary" style="color:#888;margin-left:6px"></span>
              <div id="pareto-v-wafermap-fbdrop" class="wm-dd-panel" style="display:none">
                <div style="margin-bottom:4px"><button onclick="wmFbCheckAll('pareto-v',true)" style="font-size:11px">All</button> <button onclick="wmFbCheckAll('pareto-v',false)" style="font-size:11px">None</button></div>
                <div id="pareto-v-wafermap-fbchecks" style="display:flex;flex-direction:column;max-height:180px;overflow-y:auto"></div>
              </div>
            </span>
          </div>
          <div id="pareto-v-wafermap-grid" style="display:flex;flex-wrap:wrap;gap:14px"></div>
        </div>
      </div>
    </div>
    <div id="tab-table" style="display:none">
      <div class="chart-card" style="overflow:auto">
        <div id="table-totals" style="font-size:12px;color:#555;margin:0 0 6px">{pareto_v_totals_html}</div>
        <table id="run-table">
          <thead><tr>
            <th>Period</th><th>Date</th><th>Lot</th><th>Wafer</th>
            <th>Program</th><th>FF Yield%</th><th>FF+DF Yield%</th>
            <th>Top Fail Bins</th>
          </tr></thead>
          <tbody id="run-table-body"></tbody>
        </table>
      </div>
    </div>
    <div id="tab-dlcp" style="display:none;flex-direction:column;padding:8px;gap:6px;overflow:hidden">
      <div class="dlcp-ctrl">
        <label style="font-weight:bold">UPM Threshold:</label>
        <input type="range" id="dlcp-sl" min="70" max="100" step="0.5" value="92.5" style="width:180px" oninput="dlcpSliderT()">
        <input type="number" id="dlcp-tv-inp" min="70" max="100" step="0.5" value="92.5" style="width:64px;font-size:13px;padding:2px 4px;border:1px solid #aac;border-radius:3px;text-align:right" oninput="dlcpTxtT(this.value)" onchange="dlcpTxtT(this.value)">
        <span style="color:#1a5276;font-size:13px">%</span>
        <button onclick="dlcpOpenHistT()" style="margin-left:12px;padding:4px 16px;font-size:13px;font-weight:bold;background:linear-gradient(135deg,#1a5276,#2980b9);color:#fff;border:none;border-radius:5px;cursor:pointer;box-shadow:0 2px 6px rgba(0,0,0,.25)">&#128202; UPM Distribution</button>
        <span id="dlcp-no-data-msg" style="color:#c0392b;font-weight:bold;display:none">No UPM die data available.</span>
      </div>
      <div id="dlcp-totals" style="font-size:12px;color:#555;margin:0 0 6px">{pareto_v_totals_html}</div>
      <div class="dlcp-sumbox" id="dlcp-sumbox"></div>
      <div class="dlcp-inner" style="flex:1;min-height:300px">
        <div class="dlcp-left" id="dlcp-left-pane">
          <div class="dlcp-panel-hdr">
            <span class="dlcp-sec-ttl" style="flex:1">Per-Wafer Detail</span>
            <button class="dlcp-pbtn" onclick="dlcpSelAllT()">&#9745; All</button>
            <button class="dlcp-pbtn" onclick="dlcpSelNoneT()">&#9746; None</button>
            <button class="dlcp-pbtn" onclick="dlcpDownloadCsvT()">&#8681; CSV</button>
            <button class="dlcp-pbtn" onclick="dlcpClearFiltersT()">&#10005; Filters</button>
          </div>
          <div id="dlcp-tbl-pane"><div class="dlcp-tw"><table class="dlcp-t"><thead>
            <tr>
              <th rowspan="2">Lot <button class="dlcp-ddbtn" id="dlcp-dd-btn-0" onclick="event.stopPropagation();dlcpDdOpenT(0,this)">&#9660;</button></th>
              <th rowspan="2">Wafer <button class="dlcp-ddbtn" id="dlcp-dd-btn-1" onclick="event.stopPropagation();dlcpDdOpenT(1,this)">&#9660;</button></th>
              <th rowspan="2">Test Program <button class="dlcp-ddbtn" id="dlcp-dd-btn-2" onclick="event.stopPropagation();dlcpDdOpenT(2,this)">&#9660;</button></th>
              <th rowspan="2">Material <button class="dlcp-ddbtn" id="dlcp-dd-btn-3" onclick="event.stopPropagation();dlcpDdOpenT(3,this)">&#9660;</button></th>
              <th class="num" rowspan="2">Total</th><th class="num" rowspan="2">Med UPM%</th>
              <th class="num" colspan="2" style="background:#1a5276">HP (IB1/2, UPM&ge;thr)</th>
              <th class="num" colspan="2" style="background:#7d6608">LP (IB1-4, below thr)</th>
              <th class="num" colspan="2" style="background:#7b241c">Fail (IB&gt;4)</th>
              <th class="num" colspan="2" style="background:#1a7a4a">FF+DF (IB1-4)</th>
              <th class="num" colspan="2" style="background:#1e8449">FF (IB 1,2)</th>
              <th class="num" colspan="2" style="background:#117a65">DF (IB 3-4)</th>
              <th class="num" colspan="2" style="background:#7d3c98">ATOM DF (IB 3)</th>
              <th class="num" colspan="2" style="background:#922b21">CORE DF (IB 4)</th></tr>
            <tr>
              <th class="num" style="background:#1a5276">#</th><th class="num" style="background:#1a5276">% of IB1-4</th>
              <th class="num" style="background:#7d6608">#</th><th class="num" style="background:#7d6608">% of IB1-4</th>
              <th class="num" style="background:#7b241c">#</th><th class="num" style="background:#7b241c">% of total</th>
              <th class="num" style="background:#1a7a4a">#</th><th class="num" style="background:#1a7a4a">% of total</th>
              <th class="num" style="background:#1e8449">#</th><th class="num" style="background:#1e8449">% of IB1-4</th>
              <th class="num" style="background:#117a65">#</th><th class="num" style="background:#117a65">% of IB1-4</th>
              <th class="num" style="background:#7d3c98">#</th><th class="num" style="background:#7d3c98">% of IB1-4</th>
              <th class="num" style="background:#922b21">#</th><th class="num" style="background:#922b21">% of IB1-4</th>
            </tr>
          </thead><tbody id="dlcp-flt-row-t"></tbody><tbody id="dlcp-tb-t"></tbody></table></div></div>
          <div class="dlcp-note" id="dlcp-note-t"></div>
        </div>
        <div class="dlcp-splitter" id="dlcp-splitter-t" onclick="dlcpSplitterToggleT()">
          <span class="dlcp-split-arrow" id="dlcp-split-arrow-t">&#9654;</span>
        </div>
        <div class="dlcp-cw" id="dlcp-right-pane-t">
          <div class="dlcp-panel-hdr" style="margin-bottom:4px">
            <div style="font-size:11px;color:#666;flex:1">CDF of UPM% &mdash; HP/LP (solid) | FF IB1,2 / DF IB3,4 (dashed) | red dashed = threshold</div>
            <button class="dlcp-pbtn" onclick="dlcpSavePngT()">&#128247; PNG</button>
          </div>
          <canvas id="dlcp-cv-t" style="display:block;width:100%;flex:1;border:1px solid #dde;border-radius:4px;min-height:180px"></canvas>
        </div>
      </div>
    </div>

<script>
// ═══════════════════════════════════════ DATA ═══════════════════════════════
const DATA = {data_js};
// fb_modules (functional-bin drill-down text) is large and rarely needed —
// parsed from this separate <script> block on first drill-down click only.
let _dlcpExtraLoaded = false;
function ensureDlcpExtraLoaded() {{
  if (_dlcpExtraLoaded) return;
  const raw = document.getElementById('dlcp-extra-data');
  if (raw) {{
    const extra = JSON.parse(raw.textContent);
    extra.forEach((e, i) => {{ if (DATA.runs[i]) {{ DATA.runs[i].fb_modules = e.fb_modules; DATA.runs[i].die_xy = e.die_xy; }} }});
  }}
  _dlcpExtraLoaded = true;
}}
// Load fb_modules on first user interaction (mousedown fires before the click/
// plotly_click that would need it) rather than blocking the initial page paint.
document.addEventListener('mousedown', ensureDlcpExtraLoaded, {{ once: true, capture: true }});
// Plot element IDs — static embeds use UUIDs; JS targets those same elements
const _TREND_EL   = '{_trend_id}'   || 'trend-chart';
const _PARETO_H_EL= '{_pareto_h_id}'|| 'pareto-h-chart';
const _PARETO_V_EL= '{_pareto_v_id}'|| 'pareto-v-chart';
const PASS_BINS  = new Set(DATA.pass_bins);
const FF_BINS    = new Set(DATA.ff_bins);
const FF_DF_BINS = new Set(DATA.ff_df_bins);
const PALETTE    = DATA.palette;

function ibinLabel(ib) {{
  const n = DATA.ibin_names[String(ib)];
  return n ? `iBin ${{ib}} \u2014 ${{n}}` : `iBin ${{ib}}`;
}}

// ═══════════════════════════════════════ TABS ══════════════════════════════
let _activeTab = 'trend';
let _firstLoad = true;  // gates lazy pareto/table rendering only; trend chart always fully rebuilds
const _PLOTLY_CFG     = {{displayModeBar:true, scrollZoom:true}};
const _PLOTLY_CFG_STD = {{displayModeBar:true}};
function resizeActiveChart() {{
  if (typeof Plotly === 'undefined') return;
  const ids = {{'trend':_TREND_EL,'pareto-h':_PARETO_H_EL,'pareto-v':_PARETO_V_EL}};
  const el = document.getElementById(ids[_activeTab]);
  if (el) Plotly.Plots.resize(el);
}}
window.addEventListener('resize', function() {{ resizeActiveChart(); if(_activeTab==='dlcp') requestAnimationFrame(_dlcpRenderCdfT); }});
// ResizeObserver: fires when user drags the chart-wrap handle
const _ro = new ResizeObserver(entries => {{
  for (const e of entries) {{
    const wrap = e.target;
    const plotDiv = wrap.querySelector('.js-plotly-plot');
    if (typeof Plotly !== 'undefined' && plotDiv) Plotly.Plots.resize(plotDiv);
  }}
}});
document.querySelectorAll('.chart-wrap').forEach(el => _ro.observe(el));
function showTab(name, btn) {{
  document.querySelectorAll('#tab-content > div').forEach(d => d.style.display = 'none');
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  var tabEl = document.getElementById('tab-' + name);
  if (!tabEl) return;
  tabEl.style.display = 'flex';
  tabEl.style.flexDirection = 'column';
  // Use passed element (this) — avoids reliance on global event object (broken in Firefox)
  if (btn) btn.classList.add('active');
  _activeTab = name;
  resizeActiveChart();
  // Lazy render — wrapped in try/catch so errors never block tab switching
  try {{
    const runs = window._pendingRuns;
    if (runs && (name === 'pareto-h' || name === 'pareto-v') && !_paretoRendered) {{
      const pareto = buildParetoTraces(runs, 20);
      Plotly.react(_PARETO_H_EL, pareto.traces, pareto.layout, _PLOTLY_CFG_STD).then(() => {{
        document.getElementById(_PARETO_H_EL).on('plotly_click', function(d) {{
          const pt = d.points[0];
          const ibNum = parseInt((pt.x || '').toString().match(/\\d+/)?.[0]);
          if (!isNaN(ibNum)) showFbDrilldown(ibNum, window._lastFilteredRuns, 'pareto-h', window._lastFilteredRuns);
        }});
      }});
      const paretoV = buildParetoVertTraces(runs, 20);
      Plotly.react(_PARETO_V_EL, paretoV.traces, paretoV.layout, _PLOTLY_CFG_STD).then(() => {{
        document.getElementById(_PARETO_V_EL).on('plotly_click', function(d) {{
          const pt = d.points[0];
          const ibNum = parseInt((pt.x || '').toString().match(/\\d+/)?.[0]);
          if (!isNaN(ibNum)) showFbDrilldown(ibNum, window._lastFilteredRuns, 'pareto-v', window._lastFilteredRuns);
        }});
      }});
      updateParetoTable(paretoV.tableRows);
      _paretoRendered = true;
    }}
    if (runs && name === 'dlcp') {{
      updateDlcp(runs);
      // Trigger CDF render when canvas actually becomes visible in the viewport.
      // IntersectionObserver is the most reliable: fires AFTER layout completes,
      // exactly when the hidden tab → visible transition is done.
      (function() {{
        var _cv = document.getElementById('dlcp-cv-t');
        if (!_cv) return;
        // IntersectionObserver: fires when canvas enters viewport
        if (window.IntersectionObserver && !_cv._dlcpIO) {{
          _cv._dlcpIO = true;
          new IntersectionObserver(function(entries) {{
            if (entries[0].isIntersecting && _activeTab === 'dlcp') {{
              requestAnimationFrame(_dlcpRenderCdfT);
            }}
          }}, {{threshold: 0.01}}).observe(_cv);
        }}
        // ResizeObserver: re-renders on resize (e.g. splitter drag)
        if (window.ResizeObserver && !_cv._dlcpRO2) {{
          _cv._dlcpRO2 = true;
          new ResizeObserver(function() {{
            if (_activeTab === 'dlcp' && _cv.clientWidth > 0) _dlcpRenderCdfT();
          }}).observe(_cv);
        }}
        // Fallback: direct call after two frames in case observers don't fire
        requestAnimationFrame(function() {{
          requestAnimationFrame(function() {{
            if (_activeTab === 'dlcp' && _cv.clientWidth > 0) _dlcpRenderCdfT();
          }});
        }});
      }})();
    }}
  }} catch(e) {{ console.warn('showTab lazy render error:', e); }}
}}

// ═══════════════════════════════════════ FILTER HELPERS ════════════════════
function selAll(sel)  {{ document.querySelectorAll(sel).forEach(c => c.checked = true);  }}
function selNone(sel) {{ document.querySelectorAll(sel).forEach(c => c.checked = false); }}

// Filter a sidebar checkbox list (e.g. Test Program, Interface Bin) by label text
function sidebarFilterList(listId, text) {{
  const q = (text || '').trim().toLowerCase();
  const list = document.getElementById(listId);
  if (!list) return;
  list.querySelectorAll('label').forEach(lbl => {{
    lbl.style.display = !q || lbl.textContent.toLowerCase().includes(q) ? '' : 'none';
  }});
}}
function toggleLotWafers(cb) {{
  const lot  = cb.value;
  const drop = document.getElementById('wdrop-' + lot);
  const checked = cb.checked;
  if (drop) {{
    drop.querySelectorAll('.wafer-cb').forEach(c => c.checked = checked);
    if (!checked) drop.style.display = 'none';
  }}
}}
function toggleLotGroup(cb) {{
  const prefix = cb.dataset.grp;
  const checked = cb.checked;
  // Toggle all lot-cb inside this group
  const grp = cb.closest('.lot-group');
  if (!grp) return;
  grp.querySelectorAll('.lot-cb').forEach(lotCb => {{
    lotCb.checked = checked;
    toggleLotWafers(lotCb);
  }});
}}
function toggleLotGroupDrop(id) {{
  const el = document.getElementById(id);
  if (!el) return;
  const open = el.style.display !== 'none';
  el.style.display = open ? 'none' : '';
  // flip arrow in the label
  const lbl = el.previousElementSibling && el.previousElementSibling.querySelector('.lot-group-lbl');
  if (lbl) lbl.innerHTML = lbl.innerHTML.replace(open ? '&#9660;' : '&#9654;', open ? '&#9654;' : '&#9660;');
}}
function toggleWaferDrop(span) {{
  const lot  = span.dataset.lot;
  const drop = document.getElementById('wdrop-' + lot);
  if (!drop) return;
  const open = drop.classList.contains('show');
  if (open) {{
    drop.classList.remove('show');
    span.innerHTML = '&#9654;';
  }} else {{
    drop.classList.add('show');
    span.innerHTML = '&#9660;';
  }}
}}

function toggleMaterialWafers(span) {{
  const lot = span.dataset.lot;
  const material = span.dataset.material;
  const drop = document.getElementById('mdrop-' + lot + '-' + material);
  if (!drop) return;
  const open = drop.classList.contains('show');
  if (open) {{
    drop.classList.remove('show');
    span.innerHTML = '&#9654;';
  }} else {{
    drop.classList.add('show');
    span.innerHTML = '&#9660;';
  }}
}}
function filterLots(query) {{
  const lowerQuery = query.toLowerCase();
  document.querySelectorAll('.lot-group').forEach(group => {{
    // Match against any lot-label or lot-group-lbl inside the group
    const labels = [...group.querySelectorAll('.lot-label, .lot-group-lbl')];
    const matches = labels.some(el => el.textContent.toLowerCase().includes(lowerQuery));
    group.style.display = matches ? '' : 'none';
  }});
}}
function filterMats(query) {{
  const lowerQuery = query.toLowerCase();
  document.querySelectorAll('#mat-list .cb-lbl').forEach(lbl => {{
    lbl.style.display = lbl.textContent.toLowerCase().includes(lowerQuery) ? '' : 'none';
  }});
}}
function lotWaferAll(checked) {{
  document.querySelectorAll('.lot-group').forEach(grp => {{
    if (grp.style.display === 'none') return;
    grp.querySelectorAll('.lot-cb').forEach(cb => {{
      cb.checked = checked;
      const drop = document.getElementById('wdrop-' + cb.value);
      if (drop) drop.querySelectorAll('.wafer-cb').forEach(c => c.checked = checked);
    }});
    const grpCb = grp.querySelector('.lot-grp-cb');
    if (grpCb) grpCb.checked = checked;
  }});
}}
function selFail() {{
  document.querySelectorAll('.ibin-cb').forEach(c => {{
    c.checked = (c.closest('.cb-lbl').dataset.fail === 'true');
  }});
}}

// ═══════════════════════════════════════ GROUPING ══════════════════════════
function getMondayOf(d) {{
  const day = d.getDay();
  const diff = (day === 0) ? -6 : 1 - day;
  const m = new Date(d); m.setDate(d.getDate() + diff);
  return m.toISOString().slice(0, 10);
}}
function getWorkWeek(dateStr) {{
  const d = new Date(dateStr + 'T00:00:00');
  const jan4 = new Date(d.getFullYear(), 0, 4);
  const startOfW1 = new Date(jan4);
  startOfW1.setDate(jan4.getDate() - ((jan4.getDay() + 6) % 7));
  const ww = Math.floor((d - startOfW1) / (7 * 24 * 3600 * 1000)) + 1;
  return 'WW' + String(ww).padStart(2, '0');
}}
function getRevisionKey(program) {{
  if (!program) return 'Unknown';
  let m = program.match(/[A-Z](\\d{{2}}[A-Z])/);
  if (m) return m[1];
  m = program.match(/(\\d{{2,3}}[A-Z])/);
  if (m) return m[1];
  return program.length > 6 ? program.slice(-6) : program;
}}
function revSortKey(rev) {{
  const m = rev.match(/^(\\d+)([A-Z])$/);
  if (m) return [parseInt(m[1]), m[2].charCodeAt(0)];
  return [9999, 0];
}}
function getPeriodKey(run, interval) {{
  if (interval === 'revision') return getRevisionKey(run.program);
  const dateStr = run.date || '';
  if (!dateStr) return 'Unknown';
  const d = new Date(dateStr + 'T00:00:00');
  if (isNaN(d)) return 'Unknown';
  if (interval === 'weekly')    return getWorkWeek(dateStr);
  if (interval === 'bi-weekly') {{
    const mon = getMondayOf(d);
    const ms  = new Date(mon + 'T00:00:00').getTime();
    const wk  = Math.floor(ms / (7 * 24 * 3600 * 1000));
    const bwk = Math.floor(wk / 2) * 2;
    const bwkDate = new Date(bwk * 7 * 24 * 3600 * 1000).toISOString().slice(0, 10);
    return getWorkWeek(bwkDate);
  }}
  if (interval === 'monthly')   return dateStr.slice(0, 7);
  return dateStr.slice(0, 10);
}}
function groupRuns(runs, interval) {{
  const g = {{}};
  for (const r of runs) {{
    const k = getPeriodKey(r, interval);
    (g[k] = g[k] || []).push(r);
  }}
  // Within each group, sort by (lot's last date, run date) so all runs for a
  // lot are contiguous and the latest-tested lot appears at the end
  for (const grpRuns of Object.values(g)) {{
    const lotLast = {{}};
    for (const r of grpRuns) {{
      if (r.date && (!lotLast[r.lot] || r.date > lotLast[r.lot])) lotLast[r.lot] = r.date;
    }}
    grpRuns.sort((a, b) => {{
      const la = lotLast[a.lot] || '', lb = lotLast[b.lot] || '';
      if (la !== lb) return la < lb ? -1 : 1;
      const da = a.date || '', db = b.date || '';
      return da < db ? -1 : da > db ? 1 : 0;
    }});
  }}
  const entries = Object.entries(g);
  if (interval === 'revision') {{
    entries.sort((a, b) => {{
      const ka = revSortKey(a[0]), kb = revSortKey(b[0]);
      return ka[0] !== kb[0] ? ka[0] - kb[0] : ka[1] - kb[1];
    }});
  }} else {{
    entries.sort((a, b) => a[0] < b[0] ? -1 : 1);
  }}
  return Object.fromEntries(entries);
}}

// ═══════════════════════════════════════ PER-RUN STATS ═════════════════════
function upmMedian(run) {{
  const vals = (run.ff_upm || []).concat(run.df_upm || []);
  if (!vals.length) return null;
  const sorted = vals.slice().sort((a, b) => a - b);
  const m = Math.floor(sorted.length / 2);
  return sorted.length % 2 === 0 ? (sorted[m - 1] + sorted[m]) / 2 : sorted[m];
}}
function runStats(run) {{
  let total = run.total_dies || 0;
  if (!total) total = Object.values(run.bin_counts).reduce((s, v) => s + v, 0) || 1;
  const failIbins = {{}};
  let ff = 0, ffdf = 0;
  for (const [ibStr, cnt] of Object.entries(run.bin_counts)) {{
    const ib = parseInt(ibStr);
    if (!PASS_BINS.has(ib)) failIbins[ib] = cnt / total * 100;
    if (FF_BINS.has(ib))    ff   += cnt;
    if (FF_DF_BINS.has(ib)) ffdf += cnt;
  }}
  return {{ failIbins, ffYield: ff / total * 100, ffDfYield: ffdf / total * 100, upmMed: upmMedian(run) }};
}}
function runLabel(r) {{
  const p = r.program.length > 18 ? r.program.slice(-18) : r.program;
  return `${{r.wafer || '?'}}-${{p}}`;
}}

// ═══════════════════════════════════════ AGGREGATE BY LOT ════════════════
// Merge multiple wafers belonging to the same (program, lot) into one run.
function aggregateByLot(runs) {{
  const map = new Map();
  for (const r of runs) {{
    const lot7 = (r.lot || '').substring(0, 7);
    const key = r.program + '\x00' + lot7 + '\x00' + (r.material || '');
    if (!map.has(key)) {{
      map.set(key, {{
        lot: lot7, wafer: '', sort_lot: ((r.sort_lot || r.lot) || '').substring(0, 7),
        material: r.material, program: r.program,
        date: r.date, total_dies: 0, bin_counts: {{}}, fb_counts: {{}}, fb_modules: {{}}, ff_upm: [], df_upm: [],
        _wafers: [], _sourceRuns: [], _n: 0,
      }});
    }}
    const agg = map.get(key);
    agg.total_dies += (r.total_dies || 0);
    agg._n++;
    agg._wafers.push(r.wafer || '?');
    agg._sourceRuns.push(r);
    if (r.date && (!agg.date || r.date > agg.date)) agg.date = r.date;
    for (const [ib, cnt] of Object.entries(r.bin_counts))
      agg.bin_counts[ib] = (agg.bin_counts[ib] || 0) + cnt;
    for (const [ib, fbMap] of Object.entries(r.fb_counts || {{}})) {{
      if (!agg.fb_counts[ib]) agg.fb_counts[ib] = {{}};
      for (const [fb, cnt] of Object.entries(fbMap))
        agg.fb_counts[ib][fb] = (agg.fb_counts[ib][fb] || 0) + cnt;
    }}
    for (const [ib, fbMap] of Object.entries(r.fb_modules || {{}})) {{
      if (!agg.fb_modules[ib]) agg.fb_modules[ib] = {{}};
      Object.assign(agg.fb_modules[ib], fbMap);
    }}
    if (r.ff_upm && r.ff_upm.length) agg.ff_upm.push(...r.ff_upm);
    if (r.df_upm && r.df_upm.length) agg.df_upm.push(...r.df_upm);
  }}
  return [...map.values()].map(agg => ({{
    ...agg,
    wafer: agg._wafers.length === 1 ? agg._wafers[0] : `${{agg._n}}W`,
  }}));
}}

// ═══════════════════════════════════════ BUILD TREND CHART ═════════════════
function buildTrendTraces(groups, topN, thresh, groupMode, skipTraces) {{
  const flat = [];  // {{period, run, stats}}
  const globalFail = {{}};
  for (const [period, runs] of Object.entries(groups)) {{
    for (const run of runs) {{
      const stats = runStats(run);
      flat.push({{ period, run, stats }});
      for (const [ib, pct] of Object.entries(stats.failIbins))
        globalFail[ib] = (globalFail[ib] || 0) + pct;
    }}
  }}
  if (!flat.length) return {{ traces: [], layout: {{}}, flat }};
  // First load: static embed already shows a chart — skip building the (potentially
  // large) trace/hovertext arrays entirely since they'd be discarded unused.
  if (skipTraces) return {{ traces: [], layout: {{}}, flat }};

  const topIbins = Object.entries(globalFail)
    .filter(([, v]) => v >= thresh)
    .sort((a, b) => b[1] - a[1]).slice(0, topN)
    .map(([ib]) => parseInt(ib));

  const xPos    = flat.map((_, i) => i);
  const xLabels = flat.map(({{ run }}) => {{
    const base = run.sort_lot || run.lot;
    const wfr  = groupMode === 'wafer' ? ` W${{run.wafer || '?'}}` : '';
    const mat  = run.material ? `(${{run.material}})` : '';
    return (base + wfr + (mat ? `\n${{mat}}` : ''));
  }});
  const traces  = [];

  topIbins.forEach((ib, bi) => {{
    const y     = flat.map(({{ stats }}) => stats.failIbins[ib] || 0);
    const hover = flat.map(({{ period, run, stats }}) => {{
      const pct = (stats.failIbins[ib] || 0).toFixed(2);
      const totalDies = run.total_dies || Object.values(run.bin_counts || {{}}).reduce((s,v)=>s+v,0) || 1;
      const nFail = Math.round((pct / 100) * totalDies);
      return `<b>${{runLabel(run)}}</b><br>Period: ${{period}}<br>Lot: ${{run.lot}}(${{run.material || '?'}})&nbsp;&nbsp;Wafer: ${{run.wafer}}<br>Program: ${{run.program}}<br>Date: ${{run.date}}<br>\u2500\u2500\u2500\u2500<br><b>${{ibinLabel(ib)}}</b><br>Fail: <b>${{nFail}} (${{pct}}%)</b>`;
    }});
    traces.push({{ type:'bar', x:xPos, y, name:ibinLabel(ib),
      hovertext:hover, hoverinfo:'text',
      marker:{{color:PALETTE[bi%PALETTE.length],line:{{color:'white',width:0.4}}}},
      opacity:0.85, yaxis:'y' }});
  }});

  const ffY    = flat.map(({{ stats }}) => stats.ffYield);
  const ffdfY  = flat.map(({{ stats }}) => stats.ffDfYield);
  const ffTgt  = DATA.yield_target.ff   ?? null;
  const ffdfTgt= DATA.yield_target.ff_df ?? null;

  const ffName   = DATA.ff_name   || 'SDS FF';
  const ffdfName = DATA.ff_df_name || 'SDS FF+DF';

  // Build IB/FB breakdown for hover - filter to specific IBs
  const buildIbFbBreakdown = (run, includeIBs) => {{
    const breakdown = [];
    const fbCounts = run.fb_counts || {{}};
    const totalDies = run.total_dies || Object.values(run.bin_counts || {{}}).reduce((s,v)=>s+v,0) || 1;
    
    // Filter IBs: sort and include only those in includeIBs array
    const sortedIBs = Object.keys(fbCounts).map(Number).sort((a,b)=>a-b);
    for (const ib of sortedIBs) {{
      if (includeIBs && !includeIBs.includes(ib)) continue;  // Skip if not in filter list
      const fbData = fbCounts[ib] || {{}};
      const ibCount = Object.values(fbData).reduce((s,v)=>s+v,0);
      const ibPct = ((ibCount / totalDies) * 100).toFixed(2);
      
      const fbList = [];
      const sortedFBs = Object.keys(fbData).map(Number).sort((a,b)=>a-b);
      for (const fb of sortedFBs) {{
        const fbCnt = fbData[fb] || 0;
        const fbPct = ((fbCnt / totalDies) * 100).toFixed(2);
        fbList.push(`FB${{fb}}: ${{fbCnt}}(${{fbPct}}%)`);
      }}
      breakdown.push(`IB${{ib}}: ${{ibCount}}(${{ibPct}}%)<br>&nbsp;&nbsp;${{fbList.join(', ')}}`);
    }}
    
    // Add summary for specific FBs
    const targetFBs = [126, 226, 326, 426];
    const fbSummary = [];
    for (const targetFB of targetFBs) {{
      let fbTotal = 0;
      for (const fbData of Object.values(fbCounts)) {{
        fbTotal += fbData[targetFB] || 0;
      }}
      if (fbTotal > 0) {{
        const fbPct = ((fbTotal / totalDies) * 100).toFixed(2);
        fbSummary.push(`FB${{targetFB}}: ${{fbTotal}}(${{fbPct}}%)`);
      }}
    }}
    
    if (breakdown.length) {{
      let result = '<br>\u2500\u2500\u2500\u2500<br>' + breakdown.join('<br>');
      if (fbSummary.length) result += '<br>\u2500\u2500\u2500\u2500<br>' + fbSummary.join(', ');
      return result;
    }}
    return '';
  }};

  traces.push({{ type:'scatter', x:xPos, y:ffY,
    mode:'lines+markers+text', name:ffName,
    line:{{color:'#1a73e8',width:2.5}}, marker:{{size:8}},
    text:ffY.map(v => v.toFixed(1)+'%'), textposition:'top center',
    textfont:{{size:9,color:'#1a73e8'}},
    hovertext:flat.map(({{ run, stats }}) => `<b>${{run.sort_lot || runLabel(run)}}</b><br>${{ffName}}: <b>${{stats.ffYield.toFixed(2)}}%</b>${{buildIbFbBreakdown(run, [1, 2])}}`),
    hoverinfo:'text', legendgroup:'yield_lines', yaxis:'y2' }});

  traces.push({{ type:'scatter', x:xPos, y:ffdfY,
    mode:'lines+markers+text', name:ffdfName,
    line:{{color:'#2e7d32',width:2.5,dash:'dash'}}, marker:{{size:8,symbol:'square'}},
    text:ffdfY.map(v => v.toFixed(1)+'%'), textposition:'bottom center',
    textfont:{{size:9,color:'#2e7d32'}},
    hovertext:flat.map(({{ run, stats }}) => `<b>${{run.sort_lot || runLabel(run)}}</b><br>${{ffdfName}}: <b>${{stats.ffDfYield.toFixed(2)}}%</b>${{buildIbFbBreakdown(run, [1, 2, 3, 4])}}`),
    hoverinfo:'text', legendgroup:'yield_lines', yaxis:'y2' }});

  const upmY = flat.map(({{ stats }}) => stats.upmMed);
  if (upmY.some(v => v != null)) {{
    traces.push({{ type:'scatter', x:xPos, y:upmY,
      mode:'lines+markers+text', name:'UPM Median (%)', connectgaps:true,
      line:{{color:'#ff0000',width:2.8,dash:'dot'}}, marker:{{size:7,symbol:'diamond',color:'#ff0000'}},
      text:upmY.map(v => v != null ? v.toFixed(1)+'%' : ''), textposition:'top center',
      textfont:{{size:9,color:'#ff0000'}},
      hovertext:flat.map(({{ run, stats }}) => stats.upmMed != null
        ? `<b>${{run.sort_lot || runLabel(run)}}</b><br>UPM Median: <b>${{stats.upmMed.toFixed(2)}}%</b>`
        : `<b>${{run.sort_lot || runLabel(run)}}</b><br>UPM Median: \u2014`),
      hoverinfo:'text', legendgroup:'yield_lines', yaxis:'y2' }});
  }}

  // Period dividers
  const shapes = [], annots = [];
  let idx = 0;
  for (const [period, runs] of Object.entries(groups)) {{
    const start = idx, end = idx + runs.length;
    const mid   = (start + end - 1) / 2;
    if (start > 0)
      shapes.push({{ type:'line', x0:start-0.5, x1:start-0.5, y0:0, y1:1,
        yref:'paper', line:{{color:'#95a5a6',width:1.2,dash:'dot'}} }});
    annots.push({{ x:mid, y:1.06, xref:'x', yref:'paper', text:`<b>${{period}}</b>`,
      showarrow:false, font:{{size:11,color:'#2c3e50'}}, xanchor:'center' }});
    idx = end;
  }}

  // Target lines + annotations
  const hlines = [], tgtAnnots = [];
  if (ffTgt != null) {{
    hlines.push({{ type:'line', x0:0, x1:1, xref:'paper',
      y0:ffTgt, y1:ffTgt, yref:'y2',
      line:{{color:'#1a73e8',width:2.5,dash:'dot'}}, opacity:0.85 }});
    tgtAnnots.push({{ x:1, xref:'paper', y:ffTgt, yref:'y2',
      text:`${{ffName}} target ${{ffTgt.toFixed(1)}}%`,
      showarrow:false, xanchor:'left', font:{{size:10,color:'#1a73e8'}},
      bgcolor:'rgba(255,255,255,0.7)' }});
  }}
  if (ffdfTgt != null) {{
    hlines.push({{ type:'line', x0:0, x1:1, xref:'paper',
      y0:ffdfTgt, y1:ffdfTgt, yref:'y2',
      line:{{color:'#2e7d32',width:2.5,dash:'dot'}}, opacity:0.85 }});
    tgtAnnots.push({{ x:1, xref:'paper', y:ffdfTgt, yref:'y2',
      text:`${{ffdfName}} target ${{ffdfTgt.toFixed(1)}}%`,
      showarrow:false, xanchor:'left', font:{{size:10,color:'#2e7d32'}},
      bgcolor:'rgba(255,255,255,0.7)' }});
  }}

  const maxStack = flat.reduce((mx, {{ stats }}) => {{
    const s = topIbins.reduce((t, ib) => t + (stats.failIbins[ib]||0), 0);
    return Math.max(mx, s);
  }}, 0);
  const failYlim = Math.min(100, Math.max(maxStack * 1.25, 5));

  const layout = {{
    barmode:'stack', plot_bgcolor:'#f9f9fb', paper_bgcolor:'white',
    title: {{ text: `${{DATA.chart_name ? '<b>' + DATA.chart_name + '</b> \u2014 ' : ''}}Interface Bin Fail vs. Yield Trend`, font:{{size:16}}, y:0.97, yanchor:'top' }},
    xaxis: {{ tickvals:xPos, ticktext:xLabels, tickfont:{{size:11}},
      tickangle:-45, showgrid:false, title:'SORT LOT',
      automargin:true }},
    yaxis: {{ title:'Interface Bin Fail (%)', range:[0,failYlim],
      gridcolor:'#e8e8e8', zeroline:true, zerolinecolor:'#ccc' }},
    yaxis2: {{ title:'Yield (%)', range:[0,105], overlaying:'y', side:'right',
      showgrid:false }},
    legend: {{ orientation:'v', x:1.01, y:0.0, xanchor:'left', yanchor:'bottom',
      bgcolor:'rgba(255,255,255,0.85)', bordercolor:'#ddd', borderwidth:1 }},
    shapes: [...shapes, ...hlines], annotations: [...annots, ...tgtAnnots],
    margin: {{ l:60, r:180, t:100, b:220 }},
    hovermode:'closest', autosize:true,
  }};

  return {{ traces, layout, flat }};
}}

// ═══════════════════════════════════════ BUILD PARETO (HORIZONTAL) ═══════
function buildParetoTraces(runs, topN) {{
  const totals = {{}};
  const n = runs.length || 1;
  for (const run of runs) {{
    for (const [ib, pct] of Object.entries(runStats(run).failIbins))
      totals[ib] = (totals[ib] || 0) + pct;
  }}
  const sorted = Object.entries(totals)
    .map(([ib, t]) => ({{ ib:parseInt(ib), avg:t/n }}))
    .sort((a, b) => b.avg - a.avg).slice(0, topN);

  const x = sorted.map(e => ibinLabel(e.ib));
  const y = sorted.map(e => e.avg);
  return {{
    traces: [{{ type:'bar', x, y,
      marker:{{color:sorted.map((_, i) => PALETTE[i%PALETTE.length])}},
      hovertemplate:'%{{x}}<br>Avg Fail: %{{y:.2f}}%<extra></extra>',
      name:'Avg Fail %' }}],
    layout: {{
      plot_bgcolor:'#f9f9fb', paper_bgcolor:'white',
      title: {{ text: (DATA.chart_name ? '<b>' + DATA.chart_name + '</b> \u2014 ' : '') + 'Overall Interface Bin Fail Pareto', font:{{size:16}} }},
      xaxis:{{tickangle:-45, tickfont:{{size:11}}, automargin:true}},
      yaxis:{{title:'Avg Fail (%)', gridcolor:'#e8e8e8'}},
      margin:{{l:60,r:40,t:60,b:180}}, showlegend:false, autosize:true,
    }},
  }};
}}

// ═══════════════════════════════════════ BUILD PARETO (VERTICAL / BY BIN) ═
function buildParetoVertTraces(runs, topN) {{
  const totals = {{}};
  const n = runs.length || 1;
  for (const run of runs) {{
    for (const [ib, pct] of Object.entries(runStats(run).failIbins))
      totals[ib] = (totals[ib] || 0) + pct;
  }}
  const sorted = Object.entries(totals)
    .map(([ib, t]) => ({{ ib:parseInt(ib), avg:t/n }}))
    .sort((a, b) => b.avg - a.avg).slice(0, topN);

  if (!sorted.length) return {{ traces:[], layout:{{}}, tableRows:[] }};

  const totalAvg = sorted.reduce((s, e) => s + e.avg, 0) || 1;
  const x = sorted.map(e => ibinLabel(e.ib));
  const y = sorted.map(e => e.avg);
  const cum = [];
  let running = 0;
  for (const v of y) {{ running += v / totalAvg * 100; cum.push(running); }}

  const binMap = DATA.bin_map || {{}};
  const tableRows = sorted.map(e => {{
    const info = binMap[String(e.ib)] || {{}};
    const nFail = runs.reduce((s, r) => s + ((r.bin_counts || {{}})[String(e.ib)] || 0), 0);
    return {{ ib: e.ib, cat: info.cat || '', desc: info.desc || '', nFail, pct: e.avg }};
  }});

  const maxY = y[0] || 1;
  return {{
    traces: [
      {{ type:'bar', x, y,
        name:'Avg Fail (%)',
        marker:{{color:sorted.map((_, i) => PALETTE[i%PALETTE.length]),
                line:{{color:'#1a252f',width:0.8}}}},
        opacity:0.9,
        text:y.map(v => v.toFixed(2)+'%'), textposition:'outside',
        textfont:{{size:10,color:'#333'}},
        hovertemplate:'%{{x}}<br>Avg Fail: %{{y:.2f}}%<extra></extra>' }},
      {{ type:'scatter', x, y:cum,
        name:'Cumulative %', yaxis:'y2',
        mode:'lines+markers',
        line:{{color:'#e67e22',width:2.5}},
        marker:{{size:7,color:'#e67e22'}},
        hovertemplate:'%{{x}}<br>Cumulative: %{{y:.1f}}%<extra></extra>' }},
    ],
    layout: {{
      plot_bgcolor:'#f9f9fb', paper_bgcolor:'white',
      title: {{ text: (DATA.chart_name ? '<b>' + DATA.chart_name + '</b> \u2014 ' : '') + '<b>Fail Pareto Chart (Percentage)</b>', font:{{size:16}} }},
      xaxis:{{tickangle:-45, tickfont:{{size:10}}, automargin:true}},
      yaxis:{{title:'Fail (%)', gridcolor:'#e8e8e8', range:[0,maxY*1.2], ticksuffix:'%'}},
      yaxis2:{{title:'Cumulative (%)', overlaying:'y', side:'right',
               range:[0,105], showgrid:false, ticksuffix:'%'}},
      legend:{{x:1.08,y:1.0,bgcolor:'rgba(255,255,255,0.85)',bordercolor:'#ddd',borderwidth:1}},
      margin:{{l:70,r:120,t:60,b:180}}, autosize:true,
    }},
    tableRows,
  }};
}}

// ═══════════════════════════════════════ UPDATE PARETO TABLE ══════════════
function _updateParetoVTotals(runs) {{
  const lots   = new Set(runs.map(r => r.lot));
  const wafers = runs.reduce((s, r) => s + (r.wafer_count || 1), 0);
  const dies   = runs.reduce((s, r) => s + (r.total_dies || 0), 0);
  const html   = `Lots: <b>${{lots.size.toLocaleString()}}</b> &nbsp;|&nbsp; Wafers: <b>${{wafers.toLocaleString()}}</b> &nbsp;|&nbsp; Dies: <b>${{dies.toLocaleString()}}</b>`;
  // Same totals summary shown on every tab (Trend, Pareto h/v, Run Table, DLCP)
  ['pareto-v-totals', 'trend-totals', 'pareto-h-totals', 'table-totals', 'dlcp-totals'].forEach(id => {{
    const div = document.getElementById(id);
    if (div) div.innerHTML = html;
  }});
}}

function updateParetoTable(tableRows) {{
  const saved = loadComments();
  const wrapper = document.getElementById('pareto-v-tbl-wrap');
  if (!wrapper) return;
  const oldTbl = wrapper.querySelector('.pareto-tbl');
  if (oldTbl) oldTbl.remove();
  if (!tableRows || !tableRows.length) return;

  const filteredRuns = window._lastFilteredRuns || [];
  const grandTotal = filteredRuns.reduce((s, r) => s + (r.total_dies || Object.values(r.bin_counts || {{}}).reduce((a,v)=>a+v,0) || 0), 0) || 1;
  const grandWafers = filteredRuns.reduce((s, r) => s + (r.wafer_count || 1), 0) || 1;

  let html = '<table class="pareto-tbl" id="pareto-summary-tbl"><thead><tr>'
    + '<th>Interface Bin</th><th>Description</th>'
    + '<th class="num">Total Tested (Wafers / Dies)</th><th class="num">Fail Count</th>'
    + '<th class="num">Fail %</th><th>Comment</th></tr></thead><tbody>';

  for (const r of tableRows) {{
    const ibStr = String(r.ib);
    const desc = (r.cat && r.desc && r.cat !== r.desc) ? r.cat + ' \u2014 ' + r.desc : (r.cat || r.desc || '');
    const pct = r.pct.toFixed(2);
    const savedCmt = (saved[ibStr] || '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
    html += `<tr><td>${{r.ib}}</td><td>${{desc}}</td>`
          + `<td class="num">${{grandWafers.toLocaleString()}} / ${{grandTotal.toLocaleString()}}</td>`
          + `<td class="num">${{r.nFail.toLocaleString()}}</td><td class="num">${{pct}}%</td>`
          + `<td><textarea class="pareto-comment" data-ib="${{ibStr}}" rows="1" placeholder="Add comment...">${{savedCmt}}</textarea></td></tr>`;
  }}

  html += '</tbody></table>';
  wrapper.insertAdjacentHTML('beforeend', html);
  initParetoComments();
  const tbl = wrapper.querySelector('.pareto-tbl');
  if (tbl) resizableCols(tbl);
}}

// ═══════════════════════════════════════ TABLE + STATS ═════════════════════
function updateTable(flat) {{
  const tbody = document.getElementById('run-table-body');
  if (!flat || !flat.length) {{ tbody.innerHTML = ''; return; }}
  // Build all rows at once (single innerHTML = much faster than 579 insertAdjacentHTML)
  const rows = [];
  for (const {{ period, run, stats }} of flat) {{
    const top5 = Object.entries(stats.failIbins)
      .sort((a, b) => b[1] - a[1]).slice(0, 5)
      .map(([ib, p]) => `iBin ${{ib}}: ${{parseFloat(p).toFixed(1)}}%`).join(' | ');
    const ffCls = stats.ffYield >= 80 ? 'yld-ok' : (stats.ffYield >= 50 ? 'yld-mid' : 'yld-low');
    rows.push(`<tr><td>${{period}}</td><td>${{run.date}}</td><td>${{run.lot}}</td><td>${{run.wafer}}</td>`
      + `<td>${{run.program}}</td>`
      + `<td class="${{ffCls}}" style="text-align:right">${{stats.ffYield.toFixed(1)}}%</td>`
      + `<td style="text-align:right">${{stats.ffDfYield.toFixed(1)}}%</td>`
      + `<td style="font-size:11px;color:#555">${{top5}}</td></tr>`);
  }}
  tbody.innerHTML = rows.join('');
  const runTbl = document.getElementById('run-table');
  if (runTbl) resizableCols(runTbl);
}}
function updateStats(runs, flat) {{
  const f = flat || [];
  const progs = new Set(runs.map(r => r.program));
  const n = f.length || 1;
  const avgFF   = f.reduce((s, e) => s + e.stats.ffYield,   0) / n;
  const avgFFDF = f.reduce((s, e) => s + e.stats.ffDfYield, 0) / n;
  const ffLbl   = DATA.ff_name   || 'SDS FF';
  const ffdfLbl = DATA.ff_df_name || 'SDS FF+DF';
  document.getElementById('stats-bar').innerHTML =
    `<div class="stat-chip">Runs: <b>${{flat.length}}</b></div>
     <div class="stat-chip">Programs: <b>${{progs.size}}</b></div>
     <div class="stat-chip">Avg ${{ffLbl}}: <b>${{avgFF.toFixed(1)}}%</b></div>
     <div class="stat-chip">Avg ${{ffdfLbl}}: <b>${{avgFFDF.toFixed(1)}}%</b></div>`;
}}

// ═══════════════════════════════════════ MAIN REBUILD ══════════════════════
function rebuildCharts() {{
  const interval = document.querySelector('input[name="interval"]:checked')?.value || 'revision';
  const topN   = parseInt(document.getElementById('topn-input').value)   || 8;
  const thresh = parseFloat(document.getElementById('thresh-input').value) || 0;

  const selProgs  = new Set([...document.querySelectorAll('.prog-cb:checked')].map(c => c.value));
  const selMats   = new Set([...document.querySelectorAll('.mat-cb:checked')].map(c => c.value));
  const selMatsHasNone = selMats.has('');
  const selIbins  = new Set([...document.querySelectorAll('.ibin-cb:checked')].map(c => parseInt(c.value)));
  const selLots   = new Set([...document.querySelectorAll('.lot-cb:checked')].map(c => c.value));
  const selWafers = new Set([...document.querySelectorAll('.wafer-cb:checked')].map(c => c.value));
  const dateMode  = document.querySelector('input[name="datemode"]:checked')?.value || 'all';
  const dateFrom  = document.getElementById('date-from').value;  // YYYY-MM-DD or ''
  const dateTo    = document.getElementById('date-to').value;
  function dateInRange(d) {{
    if (!d) return true;
    if (dateMode === 'all') return true;
    const wks = dateMode === '4w' ? 4 : dateMode === '6w' ? 6 : dateMode === '12w' ? 12 : 0;
    if (wks) {{
      const cutoff = new Date(); cutoff.setDate(cutoff.getDate() - wks * 7);
      return new Date(d + 'T00:00:00') >= cutoff;
    }}
    if (dateMode === 'custom') {{
      if (dateFrom && d < dateFrom) return false;
      if (dateTo   && d > dateTo)   return false;
      return true;
    }}
    return true;
  }}

  const filteredRuns = DATA.runs
    .filter(r => dateInRange(r.date))
    .filter(r => selProgs.has(r.program))
    .filter(r => r.material ? selMats.has(r.material) : selMatsHasNone)
    .filter(r => selLots.has(r.lot))
    .filter(r => selWafers.has(r.lot + '::' + (r.wafer || '')))
    .map(r => ({{
      ...r,
      bin_counts: Object.fromEntries(
        Object.entries(r.bin_counts).filter(([ib]) => selIbins.has(parseInt(ib)))
      ),
    }}));

  // Store globally so click handlers can access filtered runs
  window._lastFilteredRuns = filteredRuns;
  _updateParetoVTotals(filteredRuns);

  const groupMode = document.querySelector('input[name="groupby"]:checked')?.value || 'lot';
  const runsForChart = groupMode === 'lot' ? aggregateByLot(filteredRuns) : filteredRuns;
  const groups = groupRuns(runsForChart, interval);
  // Always build full traces (UPM line + click targets) — the static server-rendered
  // embed lacks the UPM trace and has no click handler bound, so first load needs this too.
  const {{ traces, layout, flat }} = buildTrendTraces(groups, topN, thresh, groupMode, false);
  window._lastFlat = flat;

  // Guard all Plotly calls — if CDN fails, tables/DLCP/tabs still work
  if (typeof Plotly !== 'undefined') {{
    Plotly.react(_TREND_EL, traces, layout, _PLOTLY_CFG).then(() => {{
      document.getElementById(_TREND_EL).on('plotly_click', function(d) {{
        const pt = d.points[0];
        if (pt.data.type === 'bar') {{
          const ibNum = parseInt((pt.data.name || '').match(/\\d+/)?.[0]);
          if (!isNaN(ibNum)) {{
            const entry = (window._lastFlat || [])[pt.pointIndex];
            const barRuns = entry ? (entry.run._sourceRuns || [entry.run]) : window._lastFilteredRuns;
            showFbDrilldown(ibNum, barRuns, 'trend', window._lastFilteredRuns);
          }}
        }} else if (pt.data.type === 'scatter' && pt.data.hoverinfo === 'text') {{
          const htArr = pt.data.hovertext;
          const hoverText = Array.isArray(htArr) ? (htArr[pt.pointIndex] || '') : (htArr || '');
          const ex = d.event ? d.event.pageX : (pt.xaxis ? pt.xaxis._offset : 200);
          const ey = d.event ? d.event.pageY : 200;
          if (hoverText) showStickyTooltip(hoverText, ex, ey);
        }}
      }});
    }});
  }} else {{
    // Plotly not loaded — show inline warning in the trend chart area
    var tc = document.getElementById('trend-chart');
    if (tc) tc.innerHTML = '<div style="padding:20px;color:#e74c3c;font-family:Arial;font-size:13px">&#9888; Plotly chart library failed to load.<br>Check shared/library/ path.</div>';
  }}

  // Lazy: only build/render pareto charts if their tab is currently visible (skip entirely on
  // first load — buildParetoTraces' hover-text construction is expensive and unused otherwise)
  if ((typeof Plotly !== 'undefined') && !_firstLoad && (_activeTab === 'pareto-h' || _activeTab === 'pareto-v' || !_paretoRendered)) {{
    const pareto = buildParetoTraces(filteredRuns, 20);
    Plotly.react(_PARETO_H_EL, pareto.traces, pareto.layout, _PLOTLY_CFG_STD).then(() => {{
      document.getElementById(_PARETO_H_EL).on('plotly_click', function(d) {{
        const pt = d.points[0];
        const ibNum = parseInt((pt.x || '').toString().match(/\\d+/)?.[0]);
        if (!isNaN(ibNum)) showFbDrilldown(ibNum, window._lastFilteredRuns, 'pareto-h', window._lastFilteredRuns);
      }});
    }});
    const paretoV = buildParetoVertTraces(filteredRuns, 20);
    Plotly.react(_PARETO_V_EL, paretoV.traces, paretoV.layout, _PLOTLY_CFG_STD).then(() => {{
      document.getElementById(_PARETO_V_EL).on('plotly_click', function(d) {{
        const pt = d.points[0];
        const ibNum = parseInt((pt.x || '').toString().match(/\\d+/)?.[0]);
        if (!isNaN(ibNum)) showFbDrilldown(ibNum, window._lastFilteredRuns, 'pareto-v', window._lastFilteredRuns);
      }});
    }});
    updateParetoTable(paretoV.tableRows);
    if (_activeTab === 'pareto-h' || _activeTab === 'pareto-v') _paretoRendered = true;
  }}
  if (_activeTab === 'dlcp') updateDlcp(filteredRuns);

  // On first load: skip heavy DOM/Plotly ops — static embed already shows chart.
  // Only update stats (fast). Table + pareto charts render lazily on user action.
  if (_firstLoad) {{
    updateStats(filteredRuns, flat);
    window._lastFilteredRuns = filteredRuns;
    window._lastFlat = flat;
    window._pendingRuns = filteredRuns;
    _firstLoad = false;
    return;
  }}

  // Always update table/stats (cheap DOM ops)
  updateTable(flat);
  updateStats(filteredRuns, flat);
  // Stash filtered runs so lazy tabs can render on first show
  window._pendingRuns = filteredRuns;
  _firstLoad = false;
}}

// ═══════════════════════════════════════ STICKY TOOLTIP (FF/FF+DF HOVER) ═
function showStickyTooltip(text, x, y) {{
  if (!text) return;
  
  // Remove existing tooltip if any
  const existing = document.getElementById('sticky-tooltip');
  if (existing) existing.remove();
  
  const tooltip = document.createElement('div');
  tooltip.id = 'sticky-tooltip';
  tooltip.style.position = 'fixed';
  tooltip.style.top = (y + 10) + 'px';
  tooltip.style.left = (x + 10) + 'px';
  tooltip.style.background = '#fff';
  tooltip.style.border = '2px solid #2c3e50';
  tooltip.style.borderRadius = '6px';
  tooltip.style.padding = '12px';
  tooltip.style.maxWidth = '400px';
  tooltip.style.maxHeight = '300px';
  tooltip.style.overflowY = 'auto';
  tooltip.style.boxShadow = '0 4px 12px rgba(0,0,0,0.15)';
  tooltip.style.zIndex = '10000';
  tooltip.style.fontSize = '12px';
  tooltip.style.fontFamily = 'monospace';
  tooltip.style.whiteSpace = 'pre-wrap';
  tooltip.style.wordWrap = 'break-word';
  
  tooltip.innerHTML = (text || '').replace(/\\n/g, '<br>');
  
  // Add close button
  const closeBtn = document.createElement('button');
  closeBtn.innerHTML = '✕';
  closeBtn.style.position = 'absolute';
  closeBtn.style.top = '4px';
  closeBtn.style.right = '4px';
  closeBtn.style.background = '#e74c3c';
  closeBtn.style.color = 'white';
  closeBtn.style.border = 'none';
  closeBtn.style.borderRadius = '50%';
  closeBtn.style.width = '24px';
  closeBtn.style.height = '24px';
  closeBtn.style.cursor = 'pointer';
  closeBtn.style.fontSize = '14px';
  closeBtn.style.padding = '0';
  closeBtn.onclick = () => tooltip.remove();
  tooltip.appendChild(closeBtn);
  
  document.body.appendChild(tooltip);
  
  // Close on escape key
  const closeOnEsc = (e) => {{
    if (e.key === 'Escape') {{
      tooltip.remove();
      document.removeEventListener('keydown', closeOnEsc);
    }}
  }};
  document.addEventListener('keydown', closeOnEsc);
  
  // Close when clicking outside
  setTimeout(() => {{
    document.addEventListener('click', (e) => {{
      if (e.target !== tooltip && !tooltip.contains(e.target)) {{
        tooltip.remove();
      }}
    }}, {{ once: true }});
  }}, 0);
}}

// ═══════════════════════════════════════ FB DRILLDOWN TABLE ═
// ═══════════════════════════════════════ COLUMN RESIZE ═══════════════════
function resizableCols(table) {{
  const ths = table.querySelectorAll('thead th');
  ths.forEach(th => {{
    th.classList.add('resizable');
    const hdl = document.createElement('div');
    hdl.className = 'col-resizer';
    th.appendChild(hdl);
    let startX, startW;
    hdl.addEventListener('mousedown', e => {{
      startX = e.pageX;
      startW = th.offsetWidth;
      const onMove = ev => {{ th.style.minWidth = Math.max(40, startW + ev.pageX - startX) + 'px'; }};
      const onUp   = () => {{ document.removeEventListener('mousemove', onMove); document.removeEventListener('mouseup', onUp); }};
      document.addEventListener('mousemove', onMove);
      document.addEventListener('mouseup', onUp);
      e.preventDefault();
    }});
  }});
}}
function resizeAllTables() {{
  document.querySelectorAll('table').forEach(t => resizableCols(t));
}}

function showFbDrilldown(ibNum, runs, tabPrefix, selectedRuns) {{
  ensureDlcpExtraLoaded();  // fb_modules is loaded lazily on first drill-down click
  const ibStr   = String(ibNum);
  const binMap  = DATA.bin_map || {{}};
  const ibInfo  = binMap[ibStr] || {{}};

  // Calculate total dies from selected wafers (sidebar selections)
  let barTotal = 0;
  const selectedList = selectedRuns || runs;
  for (const run of selectedList) {{
    barTotal += run.total_dies || Object.values(run.bin_counts).reduce((s,v)=>s+v,0) || 1;
  }}
  
  // Aggregate fb_counts from clicked bar runs, but use barTotal from selected wafers
  const fbTotals = {{}};    // fb -> {{cnt, lotWafers: Map<lot, Set<wafer>>}}
  const fbModules = {{}};   // fb -> {{bdesc -> count}}
  // run objects here may be snapshots taken before fb_modules was lazy-loaded —
  // fall back to the live DATA.runs entry (now populated) matched by identity.
  function _liveFbModules(run) {{
    if (run.fb_modules) return run.fb_modules;
    const live = DATA.runs.find(dr => dr.lot === run.lot && dr.wafer === run.wafer && dr.program === run.program);
    return (live && live.fb_modules) || {{}};
  }}
  for (const run of runs) {{
    const lot = run.sort_lot || run.lot || '';
    const wafer = run.wafer || '';
    const ibFb = (run.fb_counts || {{}})[ibStr] || {{}};
    for (const [fb, cnt] of Object.entries(ibFb)) {{
      if (!fbTotals[fb]) fbTotals[fb] = {{cnt: 0, lotWafers: new Map()}};
      fbTotals[fb].cnt += cnt;
      if (lot) {{
        if (!fbTotals[fb].lotWafers.has(lot)) fbTotals[fb].lotWafers.set(lot, new Set());
        if (wafer) fbTotals[fb].lotWafers.get(lot).add(wafer);
      }}
    }}
    const ibMod = _liveFbModules(run)[ibStr] || {{}};
    for (const [fb, bdesc] of Object.entries(ibMod)) {{
      if (!fbModules[fb]) fbModules[fb] = {{}};
      fbModules[fb][bdesc] = (fbModules[fb][bdesc] || 0) + 1;
    }}
  }}

  // Return sorted list of unique bin description strings for an FB (by frequency)
  function _fbBdescs(fb) {{
    const bmap = fbModules[String(fb)];
    if (!bmap) return [];
    return Object.entries(bmap).sort((a,b)=>b[1]-a[1]).map(e=>e[0]);
  }}

  // Build FB description from fb_map (Pass-Bin-Map + fB93xx) with bin_map fallback
  function fbDesc(fb) {{
    const fbInfo = (DATA.fb_map || {{}})[String(fb)];
    if (fbInfo) {{
      const cat  = fbInfo.cat  || '';
      const desc = fbInfo.desc || '';
      if (cat && desc && cat !== desc) return `${{cat}} \u2014 ${{desc}}`;
      return cat || desc;
    }}
    // Fallback: try bin_map (interface-bin level)
    const ibInfo = (DATA.bin_map || {{}})[String(fb)] || {{}};
    const cat  = ibInfo.cat  || '';
    const desc = ibInfo.desc || ibInfo.description || '';
    if (cat && desc && cat !== desc) return `${{cat}} \u2014 ${{desc}}`;
    if (cat || desc) return cat || desc;
    // Last fallback: use the raw bin description string from the CSV
    const bdescs = _fbBdescs(fb);
    return bdescs.length > 0 ? bdescs[0] : '';
  }}

  // Sort by fail count descending
  const rows = Object.entries(fbTotals)
    .map(([fb, d]) => {{
      const lotWaferStr = [...d.lotWafers.entries()]
        .sort((a, b) => a[0].localeCompare(b[0]))
        .map(([lot, wafers]) => wafers.size ? `${{lot}}(${{[...wafers].sort().join(',')}})` : lot)
        .join(', ');
      return {{fb: parseInt(fb), cnt: d.cnt, lotWaferStr}};
    }})
    .sort((a, b) => b.cnt - a.cnt);

  const hasFbData = rows.length > 0;
  const ibLabel = ibinLabel(ibNum) || `IB ${{ibNum}}`;
  const ibCat  = ibInfo.cat  || '';
  const ibDsc  = ibInfo.desc || ibInfo.description || '';
  const ibDescTxt = (ibCat && ibDsc && ibCat !== ibDsc) ? ` \u2014 ${{ibCat}} / ${{ibDsc}}`
                  : (ibCat || ibDsc) ? ` \u2014 ${{ibCat || ibDsc}}` : '';

  let html = '';
  if (!hasFbData) {{
    html = `<tr><td colspan="8" style="text-align:center;color:#888;padding:12px">No functional bin data available for IB ${{ibNum}}</td></tr>`;
  }} else {{
    for (const {{lotWaferStr, fb, cnt}} of rows) {{
      const pct = barTotal > 0 ? (cnt / barTotal * 100).toFixed(2) : '—';
      const desc = fbDesc(fb) || ibInfo.desc || '';
      const mods = _fbBdescs(fb);
      let modCell = '';
      if (mods.length > 0) {{
        const first = mods[0];
        const disp = first.length > 45 ? first.substring(0, 43) + '..' : first;
        const tip = mods.join('&#10;').replace(/"/g, '&quot;');
        modCell = `<span title="${{tip}}">${{disp}}</span>` + (mods.length > 1 ? ` <span style="color:#aaa;font-size:10px">(+${{mods.length-1}} more)</span>` : '');
      }}
      html += `<tr>
        <td>${{ibNum}}</td>
        <td>${{lotWaferStr || '—'}}</td>
        <td>FB${{fb}}</td>
        <td>${{desc}}</td>
        <td>${{modCell}}</td>
        <td class="num">${{barTotal}}</td>
        <td class="num">${{cnt.toLocaleString()}}</td>
        <td class="num">${{pct}}%</td>
      </tr>`;
    }}
  }}

  document.getElementById(`${{tabPrefix}}-fb-ib`).textContent = `${{ibNum}}${{ibDescTxt}}`;
  document.getElementById(`${{tabPrefix}}-fb-tbody`).innerHTML = html;
  const drillCard = document.getElementById(`${{tabPrefix}}-fb-drilldown`);
  drillCard.style.display = '';
  drillCard.scrollIntoView({{behavior:'smooth', block:'nearest'}});
  const tbl = document.getElementById(`${{tabPrefix}}-fb-tbl`);
  if (tbl) resizableCols(tbl);

  // Wafer map: collect every run in this drilldown that actually has per-die X/Y data.
  // A bar can aggregate multiple wafers (e.g. a lot-level bar) — show all as thumbnails,
  // with checkboxes to narrow down which ones are displayed.
  const mapBtn = document.getElementById(`${{tabPrefix}}-fbview-map-btn`);
  const eligible = runs.filter(r => {{
    const dxy = _liveDieXy(r);
    if (!dxy.length) return false;
    // Only show wafers where the selected IB actually failed a die on that wafer.
    return dxy.some(d => d[2] === ibNum);
  }});
  window._fbddRuns = window._fbddRuns || {{}};
  window._fbddRuns[tabPrefix] = eligible;
  window._fbddIb = window._fbddIb || {{}};
  window._fbddIb[tabPrefix] = ibNum;
  const selRow = document.getElementById(`${{tabPrefix}}-wafermap-selrow`);
  const checksEl = document.getElementById(`${{tabPrefix}}-wafermap-checks`);
  if (checksEl) {{
    checksEl.innerHTML = eligible.map((r, i) => `<label class="cb-lbl">
        <input type="checkbox" class="wm-chk-${{tabPrefix}}" value="${{i}}" checked onchange="renderWafermapGrid('${{tabPrefix}}');_wmUpdateSummary('${{tabPrefix}}')"> ${{r.lot || ''}} ${{r.wafer || ''}}${{r.material ? ` (${{r.material}})` : ''}}
      </label>`).join('');
  }}
  if (selRow) selRow.style.display = eligible.length > 1 ? '' : 'none';
  _wmUpdateSummary(tabPrefix);

  // FB picker: list every functional bin seen under this IB across the eligible wafers.
  const fbSet = new Set();
  eligible.forEach(r => {{
    const fbCounts = (r.fb_counts && r.fb_counts[ibNum]) || {{}};
    Object.keys(fbCounts).forEach(fb => fbSet.add(parseInt(fb)));
  }});
  const fbList = [...fbSet].sort((a, b) => a - b);
  window._fbddFb = window._fbddFb || {{}};
  window._fbddFb[tabPrefix] = new Set(fbList);
  const fbWrap   = document.getElementById(`${{tabPrefix}}-wm-fb-wrap`);
  const fbChecks = document.getElementById(`${{tabPrefix}}-wafermap-fbchecks`);
  if (fbChecks) {{
    fbChecks.innerHTML = fbList.map(fb => `<label class="cb-lbl">
        <input type="checkbox" class="wmfb-chk-${{tabPrefix}}" value="${{fb}}" checked onchange="_wmSyncFbSelection('${{tabPrefix}}');renderWafermapGrid('${{tabPrefix}}')"> FB${{fb}}
      </label>`).join('');
  }}
  if (fbWrap) fbWrap.style.display = fbList.length ? '' : 'none';
  _wmUpdateFbSummary(tabPrefix);
  if (mapBtn) {{
    const hasDies = eligible.length > 0;
    mapBtn.disabled = !hasDies;
    mapBtn.title = hasDies ? '' : 'No wafers with per-die X/Y data in this selection';
  }}
  setFbView(tabPrefix, 'table');
}}

// Fetch per-die [x,y,ibin,fbin,upm_pct] for a run, falling back to a live DATA.runs lookup
// since the run object passed to showFbDrilldown may be a stale pre-lazy-load snapshot.
function _liveDieXy(run) {{
  if (!run) return [];
  if (run.die_xy) return run.die_xy;
  const live = DATA.runs.find(dr => dr.lot === run.lot && dr.wafer === run.wafer && dr.program === run.program);
  return (live && live.die_xy) || [];
}}

function setFbView(tabPrefix, view) {{
  const tableBtn = document.getElementById(`${{tabPrefix}}-fbview-table-btn`);
  const mapBtn   = document.getElementById(`${{tabPrefix}}-fbview-map-btn`);
  const tableWrap = document.getElementById(`${{tabPrefix}}-fbtable-wrap`);
  const mapWrap   = document.getElementById(`${{tabPrefix}}-wafermap-wrap`);
  if (view === 'map' && mapBtn && mapBtn.disabled) view = 'table';
  tableBtn.classList.toggle('active', view === 'table');
  mapBtn.classList.toggle('active', view === 'map');
  tableWrap.style.display = view === 'table' ? '' : 'none';
  mapWrap.style.display   = view === 'map' ? '' : 'none';
  if (view === 'map') renderWafermapGrid(tabPrefix);
}}

function wmCheckAll(tabPrefix, on) {{
  document.querySelectorAll(`.wm-chk-${{tabPrefix}}`).forEach(cb => {{
    if (cb.closest('.cb-lbl').style.display !== 'none') cb.checked = on;
  }});
  renderWafermapGrid(tabPrefix);
  _wmUpdateSummary(tabPrefix);
}}

// Filter the wafer checkbox list by lot/wafer/material text
function wmFilterChecks(tabPrefix, text) {{
  const q = (text || '').trim().toLowerCase();
  const checksEl = document.getElementById(`${{tabPrefix}}-wafermap-checks`);
  if (!checksEl) return;
  checksEl.querySelectorAll('.cb-lbl').forEach(lbl => {{
    lbl.style.display = !q || lbl.textContent.toLowerCase().includes(q) ? '' : 'none';
  }});
}}

function toggleWmDrop(span, tabPrefix, kind) {{
  const suffix = kind === 'fb' ? '-wafermap-fbdrop' : '-wafermap-drop';
  const drop  = document.getElementById(`${{tabPrefix}}${{suffix}}`);
  const arrow = span.querySelector('.wm-dd-arrow');
  if (!drop) return;
  const open = drop.style.display !== 'none';
  drop.style.display = open ? 'none' : 'block';
  if (arrow) arrow.innerHTML = open ? '&#9654;' : '&#9660;';
}}
// Close any open wafer/FB-picker dropdown when clicking outside its group
document.addEventListener('click', (e) => {{
  document.querySelectorAll('.wm-dd-panel').forEach(panel => {{
    if (panel.style.display === 'none') return;
    const grp = panel.closest('.wm-dd-group');
    if (grp && !grp.contains(e.target)) {{
      panel.style.display = 'none';
      const arrow = grp.querySelector('.wm-dd-arrow');
      if (arrow) arrow.innerHTML = '&#9654;';
    }}
  }});
}});
function _wmUpdateSummary(tabPrefix) {{
  const summary = document.getElementById(`${{tabPrefix}}-wafermap-summary`);
  if (!summary) return;
  // Denominator = total wafers in the current sidebar-filtered dataset, not just
  // the wafers impacted by this bin, so users can see e.g. "75/200".
  const totalWafers = (window._lastFilteredRuns || []).reduce((s, r) => s + (r.wafer_count || 1), 0);
  const checked = document.querySelectorAll(`.wm-chk-${{tabPrefix}}:checked`).length;
  summary.textContent = totalWafers ? `(${{checked}}/${{totalWafers}} shown)` : '';
}}

function wmFbCheckAll(tabPrefix, on) {{
  document.querySelectorAll(`.wmfb-chk-${{tabPrefix}}`).forEach(cb => cb.checked = on);
  _wmSyncFbSelection(tabPrefix);
  renderWafermapGrid(tabPrefix);
}}

function _wmSyncFbSelection(tabPrefix) {{
  const checked = [...document.querySelectorAll(`.wmfb-chk-${{tabPrefix}}:checked`)].map(cb => parseInt(cb.value));
  window._fbddFb = window._fbddFb || {{}};
  window._fbddFb[tabPrefix] = new Set(checked);
  _wmUpdateFbSummary(tabPrefix);
}}

function _wmUpdateFbSummary(tabPrefix) {{
  const summary = document.getElementById(`${{tabPrefix}}-wafermap-fbsummary`);
  if (!summary) return;
  const total = document.querySelectorAll(`.wmfb-chk-${{tabPrefix}}`).length;
  const checked = document.querySelectorAll(`.wmfb-chk-${{tabPrefix}}:checked`).length;
  summary.textContent = total ? `(${{checked}}/${{total}} shown)` : '';
}}

// Render one small thumbnail wafer map per checked wafer; click a thumbnail to zoom in.
function renderWafermapGrid(tabPrefix) {{
  const grid = document.getElementById(`${{tabPrefix}}-wafermap-grid`);
  if (!grid) return;
  const eligible = (window._fbddRuns || {{}})[tabPrefix] || [];
  const selIb = (window._fbddIb || {{}})[tabPrefix];
  const selFb = (window._fbddFb || {{}})[tabPrefix];
  const checked = [...document.querySelectorAll(`.wm-chk-${{tabPrefix}}:checked`)].map(cb => parseInt(cb.value));
  const shown = checked.length ? checked.map(i => eligible[i]).filter(Boolean) : eligible;
  grid.innerHTML = '';
  const legend = document.createElement('div');
  legend.style.width = '100%';
  legend.innerHTML = _wmLegendHtml(selIb, selFb, tabPrefix);
  grid.appendChild(legend);
  if (!shown.length) {{
    const none = document.createElement('div');
    none.style.cssText = 'color:#888;font-size:13px;padding:8px 0';
    none.textContent = 'No wafers selected';
    grid.appendChild(none);
    return;
  }}
  shown.forEach(run => {{
    const tile = document.createElement('div');
    tile.style.cssText = 'cursor:pointer;text-align:center;width:150px';
    tile.title = 'Click to zoom in';
    const label = document.createElement('div');
    label.style.cssText = 'font-size:11px;color:#444;margin-bottom:3px;font-weight:600';
    label.textContent = `${{run.lot || ''}} ${{run.wafer || ''}}${{run.material ? ` (${{run.material}})` : ''}}`;
    const canvasDiv = document.createElement('div');
    tile.appendChild(label);
    tile.appendChild(canvasDiv);
    tile.onclick = () => openWaferZoom(run, selIb, selFb);
    grid.appendChild(tile);
    _drawWafermap(canvasDiv, run, 150, selIb, selFb);
  }});
}}

function openWaferZoom(run, selIb, selFb) {{
  const overlay = document.getElementById('wm-zoom-overlay');
  const title    = document.getElementById('wm-zoom-title');
  const legend   = document.getElementById('wm-zoom-legend');
  const canvas   = document.getElementById('wm-zoom-canvas');
  if (!overlay || !canvas) return;
  title.textContent = `${{run.lot || ''}} ${{run.wafer || ''}}${{run.material ? ` (${{run.material}})` : ''}}`;
  if (legend) legend.innerHTML = _wmLegendHtml(selIb, selFb);
  _drawWafermap(canvas, run, Math.min(640, window.innerWidth * 0.8), selIb, selFb);
  overlay.style.display = 'flex';
}}

function closeWaferZoom() {{
  const overlay = document.getElementById('wm-zoom-overlay');
  if (overlay) overlay.style.display = 'none';
}}

function _drawWafermap(container, run, width, selIb, selFb) {{
  if (!container) return;
  const raw = _liveDieXy(run);
  if (!raw.length) {{
    container.innerHTML = '<div style="color:#888;font-size:12px;padding:8px 0">No per-die X/Y data available for this wafer</div>';
    return;
  }}
  const dies = raw.map(d => ({{x: d[0], y: d[1], ib: d[2], fb: d[3], upm: d[4]}}));
  function colorFor(ib, fb) {{
    if (selIb != null && ib === selIb) {{
      // FB deselected in the FB picker — show blank/muted instead of highlighted
      if (selFb && selFb.size && !selFb.has(fb)) return '#f4f6f6';
      return '#e74c3c';  // selected IB — bright red, highlighted
    }}
    if (ib === 1 || ib === 2) return '#1e8449';   // FF pass
    if (ib === 3 || ib === 4) return '#117a65';   // DF pass
    return '#2471a3';                              // other fail — blue
  }}
  if (typeof wmRender !== 'function') {{
    container.innerHTML = '<div style="color:#c0392b;font-size:12px">Wafer map renderer unavailable</div>';
    return;
  }}
  wmRender(container, {{
    dies: dies,
    colorFn: d => colorFor(d.ib, d.fb),
    tooltipFn: d => {{
      const fbTxt  = (d.fb === null || d.fb === undefined) ? '—' : `FB${{d.fb}}`;
      const upmTxt = (d.upm === null || d.upm === undefined) ? '—' : `${{d.upm.toFixed(2)}}%`;
      return `Die (${{d.x}}, ${{d.y}})<br>IB ${{d.ib}} &middot; ${{fbTxt}} &middot; UPM ${{upmTxt}}`;
    }},
    width: width,
  }});
}}

// Color-code legend for the wafer map (matches colorFor() in _drawWafermap)
function _wmLegendHtml(selIb, selFb) {{
  const items = [['#1e8449', 'FF Pass (IB1/2)'], ['#117a65', 'DF Pass (IB3/4)']];
  if (selIb != null) items.push(['#e74c3c', `Selected IB ${{selIb}} (fail)`]);
  items.push(['#2471a3', selIb != null ? 'Other Fail' : 'Fail']);
  if (selIb != null && selFb) items.push(['#f4f6f6', 'Deselected FB (hidden)']);
  return '<div style="display:flex;flex-wrap:wrap;gap:10px;font-size:11px;color:#444;margin-bottom:8px">' +
    items.map(([c, l]) => `<span style="display:inline-flex;align-items:center;gap:4px"><span style="width:11px;height:11px;background:${{c}};display:inline-block;border-radius:2px;border:1px solid rgba(0,0,0,0.15)"></span>${{l}}</span>`).join('') +
    '</div>';
}}

// ═══════════════════════════════════════ DLCP SPLIT ANALYSIS (bin_distribution layout) ═══════

// ── State ──────────────────────────────────────────────────────────────────
var _dlcpT = 92.5;
var _dlcpDeselT = new Set();
var _dlcpFltValsT = {{}};
var _dlcpDdFltT = {{}};
var _dlcpDdCurColT = -1;
var _dlcpDdPendingT = null;

function updateDlcp(runs) {{
  window._dlcpRuns = runs;
  _dlcpRenderT();
}}

function _dlcpRowKeyT(lot, wafer) {{ return String(lot) + '|' + String(wafer); }}
function _dlcpIsRowSelT(key) {{ return !_dlcpDeselT.has(key); }}

function _dlcpComputeRowsT() {{
  var res = [], hasDie = false;
  var runs = window._dlcpRuns || [];
  runs.forEach(function(run) {{
    if (!run) return;
    var ffUpm = run.ff_upm, dfUpm = run.df_upm;
    if (!ffUpm && !dfUpm && run.dies && run.dies.length) {{ ffUpm=[]; dfUpm=[]; run.dies.forEach(function(d){{ var ib=d[0],up=d[1]; if(up==null)return; if(ib===1||ib===2)ffUpm.push(up); else if(ib===3||ib===4)dfUpm.push(up); }}); }}
    ffUpm=ffUpm||[]; dfUpm=dfUpm||[];
    if (ffUpm.length || dfUpm.length) hasDie = true;
    // Counts from bin_counts (authoritative, includes dies without UPM data)
    var bc = run.bin_counts || {{}};
    var nFF  = (parseInt(bc['1'])||0) + (parseInt(bc['2'])||0);
    var nDF3 = parseInt(bc['3'])||0;
    var nDF4 = parseInt(bc['4'])||0;
    var ib14 = nFF + nDF3 + nDF4;
    var nC   = Math.max(0, (run.total_dies||0) - ib14);
    // HP/LP: binary search on pre-sorted ff_upm (ascending) at threshold
    var lo=0, hi=ffUpm.length;
    while(lo<hi){{var mid=(lo+hi)>>1;if(ffUpm[mid]<_dlcpT)lo=mid+1;else hi=mid;}}
    var nA = ffUpm.length - lo;  // HP: ff_upm items >= _dlcpT
    var nB = (ffUpm.length - nA) + dfUpm.length;  // LP
    // Median across all IB1-4 UPM values
    var uv = ffUpm.concat(dfUpm);
    uv.sort(function(a,b){{return a-b;}});
    var med=null;
    if (uv.length) {{ var m=Math.floor(uv.length/2); med=uv.length%2===0?(uv[m-1]+uv[m])/2:uv[m]; }}
    res.push({{
      lot:run.lot||'', wafer:run.wafer||'', prog:run.program||'', mat:run.material||'',
      tot:run.total_dies||ib14, med:med,
      nA:nA, nB:nB, nC:nC, nFF:nFF, nDF34:nDF3+nDF4, nDF3:nDF3, nDF4:nDF4
    }});
  }});
  return {{rows:res, noDies:!hasDie}};
}}

function _dlcpRenderSummaryT(tA,tB,tC,tN,medAll,tFF,tDF34,tDF3,tDF4) {{
  var sb=document.getElementById('dlcp-sumbox'); if(!sb)return;
  if(!tN){{sb.innerHTML='<span style="color:#999;font-size:12px">No data</span>';return;}}
  var mTxt=medAll!=null?medAll.toFixed(2)+'%':'\u2014';
  var tIB14=tFF+(tDF3||0)+(tDF4||0);
  var row1='<div class="dlcp-sum-panel">'
    +'<div class="dlcp-sum-panel-ttl" style="background:#34495e">Overview</div>'
    +'<div class="dlcp-sumrow">'
    +'<div class="dlcp-sum-grp"><div class="dlcp-sum-lbl">Total Die</div><div class="dlcp-sum-val">'+tN+'</div></div>'
    +'<div class="dlcp-sum-grp"><div class="dlcp-sum-lbl">Med UPM%</div><div class="dlcp-sum-val">'+mTxt+'</div></div>'
    +'<div class="dlcp-sum-grp" style="border-color:#1a7a4a"><div class="dlcp-sum-lbl">FF+DF Yield</div><div class="dlcp-sum-pct-big" style="color:#1a7a4a">'+(tN>0?((tIB14||0)/tN*100).toFixed(1):0)+'%</div><div class="dlcp-sum-sub">N='+(tIB14||0)+' \u00b7 of total</div></div>'
    +'<div class="dlcp-sum-grp" style="border-color:#1e8449"><div class="dlcp-sum-lbl">FF (IB 1,2) Yield</div><div class="dlcp-sum-pct-big" style="color:#1e8449">'+(tN>0?((tFF||0)/tN*100).toFixed(1):0)+'%</div><div class="dlcp-sum-sub">N='+(tFF||0)+' \u00b7 of total</div></div>'
    +'<div class="dlcp-sum-grp fail"><div class="dlcp-sum-lbl">Fail (IB&gt;4)</div><div class="dlcp-sum-pct-big" style="color:#c0392b">'+(tN>0?(tC/tN*100).toFixed(1):0)+'%</div><div class="dlcp-sum-sub">N='+tC+' \u00b7 of total</div></div>'
    +'</div></div>';
  var row2='<div class="dlcp-sum-panel">'
    +'<div class="dlcp-sum-panel-ttl" style="background:#1a5276">DLCP Split</div>'
    +'<div class="dlcp-sumrow">'
    +'<div class="dlcp-sum-grp pass"><div class="dlcp-sum-lbl">HP (IB1/2, UPM\u2265thr)</div><div class="dlcp-sum-pct-big" style="color:#1a5276">'+(tA+tB>0?(tA/(tA+tB)*100).toFixed(1):0)+'%</div><div class="dlcp-sum-sub">N='+tA+' \u00b7 of IB1-4</div></div>'
    +'<div class="dlcp-sum-grp marg"><div class="dlcp-sum-lbl">LP (IB1-4, below thr)</div><div class="dlcp-sum-pct-big" style="color:#ba6b0a">'+(tA+tB>0?(tB/(tA+tB)*100).toFixed(1):0)+'%</div><div class="dlcp-sum-sub">N='+tB+' \u00b7 of IB1-4</div></div>'
    +'</div>'
    +'<div style="font-size:13px;font-weight:bold;color:#555;text-transform:uppercase;letter-spacing:.5px;margin:8px 0 3px 4px;border-bottom:1px solid #e0e0e0;padding-bottom:3px">FF/DF Breakdown</div>'
    +'<div class="dlcp-sumrow">'
    +'<div class="dlcp-sum-grp" style="border-color:#1e8449"><div class="dlcp-sum-lbl" style="font-size:13px">FF (IB 1,2)</div><div class="dlcp-sum-pct-big" style="color:#1e8449;font-size:26px">'+(tIB14>0?((tFF||0)/tIB14*100).toFixed(1):0)+'%</div><div class="dlcp-sum-sub" style="font-size:12px">N='+(tFF||0)+' \u00b7 of IB1-4</div></div>'
    +'<div class="dlcp-sum-grp" style="border-color:#117a65"><div class="dlcp-sum-lbl" style="font-size:13px">DF (IB 3-4)</div><div class="dlcp-sum-pct-big" style="color:#117a65;font-size:26px">'+(tIB14>0?((tDF34||0)/tIB14*100).toFixed(1):0)+'%</div><div class="dlcp-sum-sub" style="font-size:12px">N='+(tDF34||0)+' \u00b7 of IB1-4</div></div>'
    +'<div class="dlcp-sum-grp" style="border-color:#7d3c98"><div class="dlcp-sum-lbl" style="font-size:13px">ATOM DF (IB 3)</div><div class="dlcp-sum-pct-big" style="color:#7d3c98;font-size:26px">'+(tIB14>0?((tDF3||0)/tIB14*100).toFixed(1):0)+'%</div><div class="dlcp-sum-sub" style="font-size:12px">N='+(tDF3||0)+' \u00b7 of IB1-4</div></div>'
    +'<div class="dlcp-sum-grp" style="border-color:#a04000"><div class="dlcp-sum-lbl" style="font-size:13px">CORE DF (IB 4)</div><div class="dlcp-sum-pct-big" style="color:#a04000;font-size:26px">'+(tIB14>0?((tDF4||0)/tIB14*100).toFixed(1):0)+'%</div><div class="dlcp-sum-sub" style="font-size:12px">N='+(tDF4||0)+' \u00b7 of IB1-4</div></div>'
    +'</div></div>';
  sb.innerHTML=row1+row2;
}}

function _dlcpBuildFilterRowT() {{
  var fr=document.getElementById('dlcp-flt-row-t'); if(!fr)return;
  var nc=22, html='<tr style="background:#f0f4ff">';
  for(var ci=0;ci<nc;ci++) {{
    if(ci<4){{html+='<td></td>';continue;}}
    html+='<td style="padding:1px 3px"><input data-ci="'+ci+'" placeholder="e.g. &gt;50" style="width:100%;box-sizing:border-box;font-size:10px;padding:1px 3px;border:1px solid #ccd;border-radius:2px" oninput="dlcpFltInputT('+ci+',this.value)"></td>';
  }}
  fr.innerHTML=html+'</tr>';
}}
function _dlcpNumValT(txt){{var s=txt.replace(/%/g,'').replace(/\u2014/g,'').trim();var n=parseFloat(s);return isNaN(n)?null:n;}}
function _dlcpNumTestT(fv,cellTxt){{
  var m=fv.match(/^(>=|<=|!=|>|<|=)?\\s*([\\d.]+)$/);if(!m)return true;
  var op=m[1]||'=',thr=parseFloat(m[2]);var val=_dlcpNumValT(cellTxt);if(val===null)return false;
  if(op==='>') return val>thr;if(op==='<') return val<thr;if(op==='>=') return val>=thr;
  if(op==='<=') return val<=thr;if(op==='!=') return val!==thr;return val===thr;
}}
function dlcpFltInputT(ci,val){{_dlcpFltValsT[ci]=(val||'').trim();_dlcpApplyFilterT();}}
function _dlcpApplyFilterT(){{
  var tb=document.getElementById('dlcp-tb-t');if(!tb)return;
  var rows=tb.getElementsByTagName('tr');
  for(var i=0;i<rows.length;i++){{
    var cells=rows[i].getElementsByTagName('td');var show=true;
    var colVals=[cells[0]?cells[0].textContent:'',cells[1]?cells[1].textContent:'',cells[2]?cells[2].textContent:'',cells[3]?cells[3].textContent:''];
    [0,1,2,3].forEach(function(ci){{var fs=_dlcpDdFltT[ci];if(fs&&fs.size>0&&!fs.has(colVals[ci]))show=false;}});
    Object.keys(_dlcpFltValsT).forEach(function(ci){{
      var fv=_dlcpFltValsT[ci];if(!fv)return;
      var cellTxt=cells[parseInt(ci)]?cells[parseInt(ci)].textContent:'';
      if(!_dlcpNumTestT(fv,cellTxt))show=false;
    }});
    rows[i].style.display=show?'':'none';
  }}
}}
function _dlcpEscT(s){{return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}}

function _dlcpRenderTableT(){{
  var r=_dlcpComputeRowsT();
  var tb=document.getElementById('dlcp-tb-t');if(!tb)return;
  var nd=document.getElementById('dlcp-no-data-msg');
  if(nd)nd.style.display=r.noDies?'':'none';
  if(r.noDies){{tb.innerHTML='<tr><td colspan="22" style="padding:14px;color:#7f8c8d;text-align:center">No die-level UPM data available.</td></tr>';_dlcpRenderSummaryT(0,0,0,0,null,0,0,0,0);return;}}
  var tA=0,tB=0,tC=0,tN=0,tFF=0,tDF34=0,tDF3=0,tDF4=0,html='';
  r.rows.forEach(function(x){{
    var t=x.nA+x.nB+x.nC;if(!t)return;
    var key=_dlcpRowKeyT(x.lot,x.wafer);
    var isSel=_dlcpIsRowSelT(key);
    var ddOk=true;
    var ddVals=[x.lot,x.wafer,x.prog,x.mat];
    [0,1,2,3].forEach(function(ci){{var fs=_dlcpDdFltT[ci];if(fs&&fs.size>0&&!fs.has(ddVals[ci]))ddOk=false;}});
    var visStyle=ddOk?'':'display:none';
    var f12=x.nA+x.nB, f14=x.nFF+x.nDF3+x.nDF4;
    if(ddOk){{tA+=isSel?x.nA:0;tB+=isSel?x.nB:0;tC+=isSel?x.nC:0;tN+=isSel?t:0;
              tFF+=isSel?x.nFF:0;tDF34+=isSel?x.nDF34:0;tDF3+=isSel?x.nDF3:0;tDF4+=isSel?x.nDF4:0;}}
    html+='<tr data-key="'+_dlcpEscT(key)+'" class="'+(isSel?'dlcp-rsel':'dlcp-runsel')+'" style="'+visStyle+'" onclick="dlcpRowClickT(this.getAttribute(\\'data-key\\'))">'+
      '<td>'+_dlcpEscT(x.lot)+'</td>'+'<td>'+_dlcpEscT(x.wafer)+'</td>'+'<td style="color:#555;font-size:11px">'+_dlcpEscT(x.prog)+'</td>'+'<td style="color:#555;font-size:11px">'+_dlcpEscT(x.mat)+'</td>'+
      '<td class="num">'+t+'</td>'+'<td class="num">'+(x.med!=null?x.med.toFixed(2)+'%':'\u2014')+'</td>'+
      '<td class="num" style="color:#1a5276;font-weight:bold">'+x.nA+'</td>'+'<td class="num" style="color:#1a5276">'+(f12>0?(x.nA/f12*100).toFixed(1):'\u2014')+'%</td>'+
      '<td class="num" style="color:#ba6b0a">'+x.nB+'</td>'+'<td class="num" style="color:#ba6b0a">'+(f12>0?(x.nB/f12*100).toFixed(1):'\u2014')+'%</td>'+
      '<td class="num" style="color:#922b21">'+x.nC+'</td>'+'<td class="num" style="color:#922b21">'+(t>0?(x.nC/t*100).toFixed(1):'\u2014')+'%</td>'+
      '<td class="num" style="color:#1a7a4a;font-weight:bold">'+f14+'</td>'+'<td class="num" style="color:#1a7a4a">'+(t>0?(f14/t*100).toFixed(1):'\u2014')+'%</td>'+
      '<td class="num" style="color:#1e8449;font-weight:bold">'+x.nFF+'</td>'+'<td class="num" style="color:#1e8449">'+(f14>0?(x.nFF/f14*100).toFixed(1):'\u2014')+'%</td>'+
      '<td class="num" style="color:#117a65;font-weight:bold">'+x.nDF34+'</td>'+'<td class="num" style="color:#117a65">'+(f14>0?(x.nDF34/f14*100).toFixed(1):'\u2014')+'%</td>'+
      '<td class="num" style="color:#7d3c98">'+x.nDF3+'</td>'+'<td class="num" style="color:#7d3c98">'+(f14>0?(x.nDF3/f14*100).toFixed(1):'\u2014')+'%</td>'+
      '<td class="num" style="color:#a04000">'+x.nDF4+'</td>'+'<td class="num" style="color:#a04000">'+(f14>0?(x.nDF4/f14*100).toFixed(1):'\u2014')+'%</td>'+
      '</tr>';
  }});
  tb.innerHTML=html;
  _dlcpBuildFilterRowT();
  /* Global median: weighted median of per-run medians already computed in r.rows
     — avoids sorting hundreds of thousands of raw die values on every render */
  var selMeds=[];
  r.rows.forEach(function(x){{
    var key=_dlcpRowKeyT(x.lot,x.wafer);
    var isSel=_dlcpIsRowSelT(key);
    var ddVals=[x.lot,x.wafer,x.prog,x.mat],ddOk=true;
    [0,1,2,3].forEach(function(ci){{var fs=_dlcpDdFltT[ci];if(fs&&fs.size>0&&!fs.has(ddVals[ci]))ddOk=false;}});
    if(isSel&&ddOk&&x.med!=null)selMeds.push(x.med);
  }});
  selMeds.sort(function(a,b){{return a-b;}});
  var medAll=null;
  if(selMeds.length){{var m2=Math.floor(selMeds.length/2);medAll=selMeds.length%2===0?(selMeds[m2-1]+selMeds[m2])/2:selMeds[m2];}}
  _dlcpRenderSummaryT(tA,tB,tC,tN,medAll,tFF,tDF34,tDF3,tDF4);
  var noteEl=document.getElementById('dlcp-note-t');
  if(noteEl)noteEl.innerHTML='<b>HP%</b> = HP / (HP+LP) &nbsp;|&nbsp; <b>LP%</b> = LP / (HP+LP) &nbsp;|&nbsp; <b>Fail%</b> = Fail / Total &nbsp;|&nbsp; <b>FF/DF%</b> = count / IB1-4 total &nbsp;|&nbsp; Threshold: <b>'+_dlcpT.toFixed(1)+'%</b>';
}}

function _dlcpRenderCdfT(){{
  var cv=document.getElementById('dlcp-cv-t');if(!cv)return;
  var W=cv.clientWidth||560,H=cv.clientHeight||280;
  cv.width=W;cv.height=H;
  var ctx=cv.getContext('2d');ctx.clearRect(0,0,W,H);
  var hp=[],lp=[],ff=[],df=[];
  /* Collect all IB1-4 dies from pre-sorted ff_upm/df_upm — no per-die ibin lookup needed */
  var runs=window._dlcpRuns||[];
  runs.forEach(function(run){{
    if(!run)return;
    var k=_dlcpRowKeyT(run.lot||'',run.wafer||'');if(!_dlcpIsRowSelT(k))return;
    // Also respect dropdown column filters (lot/wafer/material)
    var ddVals=[run.lot||'',run.wafer||'',run.program||'',run.material||''];
    var ddOk=true;
    [0,1,2,3].forEach(function(ci){{var fs=_dlcpDdFltT[ci];if(fs&&fs.size>0&&!fs.has(ddVals[ci]))ddOk=false;}});
    if(!ddOk)return;
    var ffU=run.ff_upm,dfU=run.df_upm;
    if(!ffU&&!dfU&&run.dies&&run.dies.length){{ffU=[];dfU=[];run.dies.forEach(function(d){{var ib=d[0],up=d[1];if(up==null)return;if(ib===1||ib===2)ffU.push(up);else if(ib===3||ib===4)dfU.push(up);}});}}
    ffU=ffU||[];dfU=dfU||[];
    ffU.forEach(function(up){{ff.push(up);if(up>=_dlcpT)hp.push(up);else lp.push(up);}});
    dfU.forEach(function(up){{df.push(up);lp.push(up);}});
  }});
  hp.sort(function(a,b){{return a-b;}});lp.sort(function(a,b){{return a-b;}});
  ff.sort(function(a,b){{return a-b;}});df.sort(function(a,b){{return a-b;}});
  if(!hp.length&&!lp.length){{
    ctx.fillStyle='#999';ctx.font='13px Arial';ctx.textAlign='center';
    ctx.fillText('No UPM die data in selected wafers',W/2,H/2);return;
  }}
  var ML=52,MR=16,MT=22,MB=42,PW=W-ML-MR,PH=H-MT-MB;
  /* Use loop-based min/max — Math.min.apply blows the call stack with large arrays */
  var xMn=Infinity,xMx=-Infinity;
  ff.forEach(function(v){{if(v<xMn)xMn=v;if(v>xMx)xMx=v;}});
  df.forEach(function(v){{if(v<xMn)xMn=v;if(v>xMx)xMx=v;}});
  if(!isFinite(xMn)){{xMn=0;xMx=100;}}
  xMn=Math.floor(xMn*2)/2-1;xMx=Math.ceil(xMx*2)/2+1;
  if(xMx-xMn<4){{xMn-=2;xMx+=2;}}
  function xp(v){{return ML+(v-xMn)/(xMx-xMn)*PW;}}
  function yp(v){{return MT+PH-v/100*PH;}}
  ctx.strokeStyle='#e8e8e8';ctx.lineWidth=1;
  for(var yi=0;yi<=4;yi++){{ctx.beginPath();ctx.moveTo(ML,yp(yi*25));ctx.lineTo(ML+PW,yp(yi*25));ctx.stroke();}}
  if(_dlcpT>=xMn&&_dlcpT<=xMx){{
    var tx=xp(_dlcpT);ctx.save();ctx.strokeStyle='#e74c3c';ctx.lineWidth=1.5;ctx.setLineDash([5,4]);
    ctx.beginPath();ctx.moveTo(tx,MT);ctx.lineTo(tx,MT+PH);ctx.stroke();
    ctx.setLineDash([]);ctx.fillStyle='#e74c3c';ctx.font='11px Arial';ctx.textAlign='center';
    ctx.fillText(_dlcpT.toFixed(1)+'%',tx,MT-5);ctx.restore();
  }}
  function drawCdf(arr,col,dash){{if(!arr.length)return;
    ctx.save();ctx.strokeStyle=col;ctx.lineWidth=2;if(dash)ctx.setLineDash([6,3]);
    var n=arr.length;
    ctx.beginPath();ctx.moveTo(xp(arr[0]),yp(0));
    var lastPx=-1;
    for(var i=0;i<n;i++){{
      var px=Math.round(xp(arr[i]));
      if(px===lastPx)continue;
      lastPx=px;
      ctx.lineTo(xp(arr[i]),yp(i/n*100));
      if(i<n-1)ctx.lineTo(xp(arr[i+1]),yp(i/n*100));
    }}
    ctx.lineTo(ML+PW,yp(100));ctx.stroke();ctx.restore();
  }}
  drawCdf(df,'#8e44ad',true);drawCdf(ff,'#27ae60',true);
  drawCdf(lp,'#e67e22',false);drawCdf(hp,'#2980b9',false);
  ctx.strokeStyle='#555';ctx.lineWidth=1;ctx.beginPath();ctx.moveTo(ML,MT);ctx.lineTo(ML,MT+PH);ctx.lineTo(ML+PW,MT+PH);ctx.stroke();
  ctx.fillStyle='#555';ctx.font='11px Arial';ctx.textAlign='right';
  for(var yi2=0;yi2<=4;yi2++)ctx.fillText(yi2*25+'%',ML-4,yp(yi2*25)+4);
  ctx.textAlign='center';var rng=xMx-xMn,stp=rng>20?5:rng>10?2:1,xs=Math.ceil(xMn/stp)*stp;
  for(var xv=xs;xv<=xMx;xv+=stp)ctx.fillText(xv.toFixed(0)+'%',xp(xv),MT+PH+14);
  ctx.fillStyle='#2c3e50';ctx.font='bold 11px Arial';ctx.textAlign='center';
  ctx.fillText('UPM %',ML+PW/2,H-4);
  ctx.save();ctx.translate(13,MT+PH/2);ctx.rotate(-Math.PI/2);ctx.fillText('Cumulative %',0,0);ctx.restore();
  var ib14=ff.length+df.length;
  var selCount=0;runs.forEach(function(r){{
    if(!r)return;
    var ddVals=[r.lot||'',r.wafer||'',r.material||''],ddOk=true;
    [0,1,2].forEach(function(ci){{var fs=_dlcpDdFltT[ci];if(fs&&fs.size>0&&!fs.has(ddVals[ci]))ddOk=false;}});
    if(ddOk&&_dlcpIsRowSelT(_dlcpRowKeyT(r.lot||'',r.wafer||'')))selCount++;
  }});
  ctx.fillStyle='#888';ctx.font='10px Arial';ctx.textAlign='right';
  ctx.fillText(selCount+' wafer'+(selCount!==1?'s':'')+' \u00b7 '+(ff.length+df.length)+' dies',ML+PW,MT-4);
  function lgPct(n,d){{return d>0?(n/d*100).toFixed(1)+'%':'0%';}}
  function drawLgEntry(x,y,lineCol,lineDash,pctTxt,nTxt,pctCol){{
    if(lineDash){{ctx.save();ctx.strokeStyle=lineCol;ctx.lineWidth=2;ctx.setLineDash([6,3]);ctx.beginPath();ctx.moveTo(x,y+2);ctx.lineTo(x+22,y+2);ctx.stroke();ctx.restore();}}
    else{{ctx.fillStyle=lineCol;ctx.fillRect(x,y+1,22,3);}}
    ctx.font='bold 12px Arial';ctx.fillStyle=pctCol||lineCol;ctx.textAlign='left';ctx.fillText(pctTxt,x+26,y+8);
    var pw=ctx.measureText(pctTxt).width;ctx.font='10px Arial';ctx.fillStyle='#aaa';ctx.fillText(' '+nTxt,x+26+pw,y+8);
  }}
  var ly=MT+8;
  drawLgEntry(ML,ly,'#2980b9',false,lgPct(hp.length,ib14),'HP  n='+hp.length,'#2980b9');
  drawLgEntry(ML+210,ly,'#e67e22',false,lgPct(lp.length,ib14),'LP  n='+lp.length,'#e67e22');
  var ly2=ly+16;
  drawLgEntry(ML,ly2,'#27ae60',true,lgPct(ff.length,ib14),'FF IB1,2  n='+ff.length,'#27ae60');
  drawLgEntry(ML+210,ly2,'#8e44ad',true,lgPct(df.length,ib14),'DF IB3,4  n='+df.length,'#8e44ad');
}}

function _dlcpRenderT(){{
  _dlcpRenderTableT();
  /* Render CDF immediately (synchronous, uses current canvas dims) and also
     after two rAF frames so any flex/reflow triggered by table rebuild settles */
  _dlcpRenderCdfT();
  requestAnimationFrame(function(){{ requestAnimationFrame(_dlcpRenderCdfT); }});
}}

function dlcpSliderT(){{
  var sl=document.getElementById('dlcp-sl');if(!sl)return;
  _dlcpT=parseFloat(sl.value);
  var inp=document.getElementById('dlcp-tv-inp');if(inp)inp.value=_dlcpT.toFixed(1);
  _dlcpRenderT();rebuildCharts();
}}
function dlcpTxtT(val){{
  var v=parseFloat(val);if(isNaN(v))return;
  v=Math.max(70,Math.min(100,v));_dlcpT=v;
  var sl=document.getElementById('dlcp-sl');if(sl)sl.value=v;
  _dlcpRenderT();rebuildCharts();
}}

function dlcpRowClickT(key){{
  if(_dlcpDeselT.has(key))_dlcpDeselT.delete(key);else _dlcpDeselT.add(key);
  var tb=document.getElementById('dlcp-tb-t');if(!tb)return;
  var rows=tb.getElementsByTagName('tr');
  for(var i=0;i<rows.length;i++){{
    var k=rows[i].getAttribute('data-key');
    if(k===key){{rows[i].classList.toggle('dlcp-runsel',_dlcpDeselT.has(key));rows[i].classList.toggle('dlcp-rsel',!_dlcpDeselT.has(key));}}
  }}
  _dlcpRenderT();
}}
function dlcpSelAllT(){{_dlcpDeselT.clear();_dlcpRenderTableT();_dlcpRenderT();}}
function dlcpSelNoneT(){{
  var r=_dlcpComputeRowsT();r.rows.forEach(function(x){{_dlcpDeselT.add(_dlcpRowKeyT(x.lot,x.wafer));}});
  _dlcpRenderTableT();_dlcpRenderT();
}}

function dlcpDdOpenT(col,btn){{
  if(_dlcpDdCurColT===col){{_dlcpDdCloseT();return;}}
  _dlcpDdCurColT=col;
  var panel=document.getElementById('dlcp-dd-panel-t');if(!panel)return;
  var srch=document.getElementById('dlcp-dd-srch-t');if(srch)srch.value='';
  var r=_dlcpComputeRowsT(),vals=[],seen=new Set();
  r.rows.forEach(function(x){{
    // Only include values for rows that will actually render (t > 0)
    var t=x.nA+x.nB+x.nC;if(!t)return;
    var v=col===0?x.lot:col===1?x.wafer:col===2?x.prog:x.mat;if(!seen.has(v)){{seen.add(v);vals.push(v);}}}});
  vals.sort();
  var cur=_dlcpDdFltT[col]||null;
  _dlcpDdPendingT=cur?new Set(cur):null;
  var lst=document.getElementById('dlcp-dd-list-t');if(!lst)return;
  lst.innerHTML=vals.map(function(v){{
    var chk=(!_dlcpDdPendingT||_dlcpDdPendingT.has(v))?'checked':'';
    return '<label class="dlcp-dd-item"><input type="checkbox" value="'+_dlcpEscT(v)+'" '+chk+' onchange="dlcpDdChkT(this)"> <span>'+_dlcpEscT(v)+'</span></label>';
  }}).join('');
  var r2=btn.getBoundingClientRect();
  panel.style.display='block';panel.style.left=r2.left+'px';panel.style.top=(r2.bottom+2)+'px';
  document.querySelectorAll('.dlcp-ddbtn').forEach(function(b){{b.classList.remove('on');}});
  btn.classList.add('on');
  setTimeout(function(){{document.addEventListener('click',_dlcpDdOutsideT,{{once:true}});}},0);
}}
function _dlcpDdOutsideT(e){{
  var panel=document.getElementById('dlcp-dd-panel-t');
  if(panel&&panel.contains(e.target)){{
    document.addEventListener('click',_dlcpDdOutsideT,{{once:true}});return;
  }}
  dlcpDdApplyT();
}}
function _dlcpDdCloseT(){{
  _dlcpDdCurColT=-1;
  var panel=document.getElementById('dlcp-dd-panel-t');if(panel)panel.style.display='none';
  document.querySelectorAll('.dlcp-ddbtn').forEach(function(b){{b.classList.remove('on');}});
}}
function dlcpDdChkT(inp){{
  var v=inp.value,chk=inp.checked;
  if(!_dlcpDdPendingT){{
    var r=_dlcpComputeRowsT();_dlcpDdPendingT=new Set();
    r.rows.forEach(function(x){{_dlcpDdPendingT.add(_dlcpDdCurColT===0?x.lot:_dlcpDdCurColT===1?x.wafer:_dlcpDdCurColT===2?x.prog:x.mat);}});
  }}
  if(chk)_dlcpDdPendingT.add(v);else _dlcpDdPendingT.delete(v);
}}
function dlcpDdSelAllT(){{
  _dlcpDdPendingT=null;
  var lst=document.getElementById('dlcp-dd-list-t');if(!lst)return;
  lst.querySelectorAll('input[type=checkbox]').forEach(function(cb){{cb.checked=true;}});
}}
function dlcpDdSelNoneT(){{
  _dlcpDdPendingT=new Set();
  var lst=document.getElementById('dlcp-dd-list-t');if(!lst)return;
  lst.querySelectorAll('input[type=checkbox]').forEach(function(cb){{cb.checked=false;}});
}}
function dlcpDdSearchT(q){{
  q=(q||'').toLowerCase();
  var lst=document.getElementById('dlcp-dd-list-t');if(!lst)return;
  lst.querySelectorAll('.dlcp-dd-item').forEach(function(el){{el.style.display=(!q||el.textContent.toLowerCase().indexOf(q)>=0)?'':'none';}});
}}
function dlcpDdApplyT(){{
  if(_dlcpDdCurColT>=0){{
    _dlcpDdFltT[_dlcpDdCurColT]=(_dlcpDdPendingT&&_dlcpDdPendingT.size>0)?_dlcpDdPendingT:null;
    var btn2=document.getElementById('dlcp-dd-btn-'+_dlcpDdCurColT);
    if(btn2)btn2.classList.toggle('on',!!_dlcpDdFltT[_dlcpDdCurColT]);
  }}
  _dlcpDdCloseT();_dlcpRenderTableT();_dlcpRenderT();
}}
function dlcpClearFiltersT(){{
  _dlcpFltValsT={{}};_dlcpDdFltT={{}};
  document.querySelectorAll('.dlcp-ddbtn').forEach(function(b){{b.classList.remove('on');}});
  var fr=document.getElementById('dlcp-flt-row-t');if(fr){{var inps=fr.getElementsByTagName('input');for(var i=0;i<inps.length;i++)inps[i].value='';}}
  var tb=document.getElementById('dlcp-tb-t');if(tb){{var rows=tb.getElementsByTagName('tr');for(var i=0;i<rows.length;i++)rows[i].style.display='';}}
  _dlcpRenderT();
}}
function dlcpSplitterToggleT(){{
  var rp=document.getElementById('dlcp-right-pane-t');
  var arr=document.getElementById('dlcp-split-arrow-t');
  if(!rp||!arr)return;
  var hidden=rp.style.display==='none';
  rp.style.display=hidden?'':'none';
  arr.innerHTML=hidden?'&#9654;':'&#9664;';
  if(!hidden)requestAnimationFrame(_dlcpRenderCdfT);
}}
function dlcpSavePngT(){{
  var cv=document.getElementById('dlcp-cv-t');if(!cv)return;
  var a=document.createElement('a');a.href=cv.toDataURL('image/png');a.download='dlcp_cdf.png';a.click();
}}
function dlcpDownloadCsv(){{dlcpDownloadCsvT();}}
function dlcpDownloadCsvT(){{
  var r=_dlcpComputeRowsT();
  var hdr=['Lot','Wafer','Test Program','Material','Total','Med UPM%','HP#','HP%','LP#','LP%','Fail#','Fail%','FF+DF#','FF+DF%','FF#','FF%','DF#','DF%','ATOM DF#','ATOM DF%','CORE DF#','CORE DF%'];
  var tb=document.getElementById('dlcp-tb-t'),visKeys=new Set();
  if(tb){{var trs=tb.getElementsByTagName('tr');for(var vi=0;vi<trs.length;vi++){{if(trs[vi].style.display!=='none'){{var vk=trs[vi].getAttribute('data-key');if(vk)visKeys.add(vk);}}}};}}
  function q(s){{var v=String(s==null?'':s);return(v.indexOf(',')>=0||v.indexOf('"')>=0)?'"'+v.replace(/"/g,'""')+'"':v;}}
  var lines=[hdr.join(',')];
  r.rows.forEach(function(x){{
    var k=_dlcpRowKeyT(x.lot,x.wafer);
    if(visKeys.size>0&&!visKeys.has(k))return;
    var t=x.nA+x.nB+x.nC;if(!t)return;
    var f12=x.nA+x.nB,f14=x.nFF+x.nDF3+x.nDF4;
    lines.push([x.lot,x.wafer,x.prog,x.mat,t,x.med!=null?x.med.toFixed(2):'',
      x.nA,f12>0?(x.nA/f12*100).toFixed(1):'',x.nB,f12>0?(x.nB/f12*100).toFixed(1):'',
      x.nC,t>0?(x.nC/t*100).toFixed(1):'',f14,t>0?(f14/t*100).toFixed(1):'',
      x.nFF,f14>0?(x.nFF/f14*100).toFixed(1):'',x.nDF34,f14>0?(x.nDF34/f14*100).toFixed(1):'',
      x.nDF3,f14>0?(x.nDF3/f14*100).toFixed(1):'',x.nDF4,f14>0?(x.nDF4/f14*100).toFixed(1):''].map(q).join(','));
  }});
  var blob=new Blob([lines.join('\\n')],{{type:'text/csv'}});
  var a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='dlcp_table.csv';a.click();
}}

function dlcpOpenHistT(){{
  var ov=document.getElementById('upm-hist-overlay-t');if(!ov)return;
  ov.classList.add('open');
  var box=document.getElementById('upm-hist-box-t'),drag=document.getElementById('upm-hist-drag-t');
  if(box&&drag&&!drag._histDrag){{
    drag._histDrag=true;
    drag.addEventListener('mousedown',function(e){{
      e.preventDefault();var startX=e.clientX,startY=e.clientY,r=box.getBoundingClientRect();
      var startL=r.left,startT2=r.top;box.style.position='fixed';box.style.margin='0';
      function mm(e2){{box.style.left=(startL+e2.clientX-startX)+'px';box.style.top=(startT2+e2.clientY-startY)+'px';}}
      function mu(){{document.removeEventListener('mousemove',mm);document.removeEventListener('mouseup',mu);}}
      document.addEventListener('mousemove',mm);document.addEventListener('mouseup',mu);
    }});
  }}
  requestAnimationFrame(_dlcpRenderHistT);
}}
function dlcpCloseHistT(){{
  var ov=document.getElementById('upm-hist-overlay-t');if(ov)ov.classList.remove('open');
}}
function _dlcpHistStatsT(arr){{
  if(!arr||!arr.length)return null;
  var n=arr.length,sorted=arr.slice().sort(function(a,b){{return a-b;}});
  var mean=sorted.reduce(function(s,v){{return s+v;}},0)/n;
  var sigma=Math.sqrt(sorted.reduce(function(s,v){{return s+(v-mean)*(v-mean);}},0)/n);
  var med=n%2===0?(sorted[n/2-1]+sorted[n/2])/2:sorted[Math.floor(n/2)];
  function pct(p){{var i=(p/100)*(n-1),lo=Math.floor(i),hi=Math.ceil(i);return lo===hi?sorted[lo]:sorted[lo]+(sorted[hi]-sorted[lo])*(i-lo);}}
  return{{n:n,mean:mean,med:med,sigma:sigma,min:sorted[0],max:sorted[n-1],p5:pct(5),p25:pct(25),p75:pct(75),p95:pct(95)}};
}}
function _dlcpRenderHistT(){{
  var cv=document.getElementById('upm-hist-cv-t');if(!cv)return;
  var W=cv.clientWidth||740,H=cv.clientHeight||260;cv.width=W;cv.height=H;
  var ctx=cv.getContext('2d');ctx.clearRect(0,0,W,H);
  /* Use pre-sorted ff_upm/df_upm — no downsampling needed */
  var runs=window._dlcpRuns||[];
  var hp2=[],lp2=[];
  runs.forEach(function(run){{
    if(!run)return;
    var k=_dlcpRowKeyT(run.lot||'',run.wafer||'');if(!_dlcpIsRowSelT(k))return;
    var ddVals=[run.lot||'',run.wafer||'',run.material||''],ddOk=true;
    [0,1,2].forEach(function(ci){{var fs=_dlcpDdFltT[ci];if(fs&&fs.size>0&&!fs.has(ddVals[ci]))ddOk=false;}});
    if(!ddOk)return;
    var ffU=run.ff_upm,dfU=run.df_upm;
    if(!ffU&&!dfU&&run.dies&&run.dies.length){{ffU=[];dfU=[];run.dies.forEach(function(d){{var ib=d[0],up=d[1];if(up==null)return;if(ib===1||ib===2)ffU.push(up);else if(ib===3||ib===4)dfU.push(up);}});}}
    ffU=ffU||[];dfU=dfU||[];
    ffU.forEach(function(up){{if(up>=_dlcpT)hp2.push(up);else lp2.push(up);}});
    dfU.forEach(function(up){{lp2.push(up);}});
  }});
  var all2=hp2.concat(lp2);
  if(!all2.length){{
    ctx.fillStyle='#999';ctx.font='13px Arial';ctx.textAlign='center';ctx.fillText('No UPM data in selected wafers',W/2,H/2);
    var sd=document.getElementById('upm-hist-stats-t');if(sd)sd.innerHTML='<span style="color:#999">No data</span>';return;
  }}
  var asP=_dlcpHistStatsT(all2),hsP=_dlcpHistStatsT(hp2),lsP=_dlcpHistStatsT(lp2);
  function fmt(v){{return v==null?'\u2014':v.toFixed(2);}}
  function sCard(lbl,s,col){{
    if(!s)return '<div class="upm-hist-stat-grp" style="border-color:'+col+'"><div class="upm-hist-stat-lbl" style="color:'+col+'">'+lbl+'</div><div class="upm-hist-stat-val">N=0</div></div>';
    return'<div class="upm-hist-stat-grp" style="border-color:'+col+'"><div class="upm-hist-stat-lbl" style="color:'+col+'">'+lbl+'</div><div class="upm-hist-stat-val" style="color:'+col+'">N='+s.n+'</div>'
      +'<div style="font-size:11px;color:#444">Median: <b>'+fmt(s.med)+'%</b> &nbsp; Mean: <b>'+fmt(s.mean)+'%</b></div>'
      +'<div style="font-size:11px;color:#444">\u03c3: <b>'+fmt(s.sigma)+'%</b> &nbsp; Min: '+fmt(s.min)+'% &nbsp; Max: '+fmt(s.max)+'%</div>'
      +'<div style="font-size:10px;color:#777">P5: '+fmt(s.p5)+'% &nbsp; P25: '+fmt(s.p25)+'% &nbsp; P75: '+fmt(s.p75)+'% &nbsp; P95: '+fmt(s.p95)+'%</div></div>';
  }}
  var sd2=document.getElementById('upm-hist-stats-t');
  if(sd2)sd2.innerHTML=sCard('All IB1-4',asP,'#2c3e50')+sCard('HP (IB1/2 \u2265thr)',hsP,'#2980b9')+sCard('LP (IB1-4 <thr)',lsP,'#e67e22');
  var xMn2=Infinity,xMx2=-Infinity;
  all2.forEach(function(v){{if(v<xMn2)xMn2=v;if(v>xMx2)xMx2=v;}});
  xMn2=Math.floor(xMn2);xMx2=Math.ceil(xMx2);
  if(xMx2-xMn2<4){{xMn2-=2;xMx2+=2;}}
  var bins=Math.min(80,Math.max(20,Math.round((xMx2-xMn2)*2))),bw2=(xMx2-xMn2)/bins;
  function mBins2(arr2){{var b=new Array(bins).fill(0);arr2.forEach(function(v){{var i2=Math.min(bins-1,Math.floor((v-xMn2)/bw2));if(i2>=0)b[i2]++;}});return b;}}
  var hpB2=mBins2(hp2),lpB2=mBins2(lp2),maxC2=0;
  for(var i=0;i<bins;i++){{var s2=hpB2[i]+lpB2[i];if(s2>maxC2)maxC2=s2;}}
  if(!maxC2)return;
  var ML2=46,MR2=14,MT2=20,MB2=38,PW2=W-ML2-MR2,PH2=H-MT2-MB2;
  function xp2(v){{return ML2+(v-xMn2)/(xMx2-xMn2)*PW2;}}
  ctx.strokeStyle='#ececec';ctx.lineWidth=1;
  for(var gi=0;gi<=4;gi++){{var gy=MT2+PH2*gi/4;ctx.beginPath();ctx.moveTo(ML2,gy);ctx.lineTo(ML2+PW2,gy);ctx.stroke();}}
  for(var bi=0;bi<bins;bi++){{
    var bx0=xp2(xMn2+bi*bw2)+0.5,bx1=xp2(xMn2+(bi+1)*bw2)-0.5,bW2=Math.max(1,bx1-bx0);
    var lpH=lpB2[bi]/maxC2*PH2,hpH=hpB2[bi]/maxC2*PH2;
    if(lpB2[bi]>0){{ctx.fillStyle='rgba(230,126,34,0.78)';ctx.fillRect(bx0,MT2+PH2-lpH,bW2,lpH);}}
    if(hpB2[bi]>0){{ctx.fillStyle='rgba(41,128,185,0.82)';ctx.fillRect(bx0,MT2+PH2-lpH-hpH,bW2,hpH);}}
  }}
  if(_dlcpT>=xMn2&&_dlcpT<=xMx2){{
    var tx2=xp2(_dlcpT);ctx.save();ctx.strokeStyle='#e74c3c';ctx.lineWidth=2;ctx.setLineDash([5,4]);
    ctx.beginPath();ctx.moveTo(tx2,MT2);ctx.lineTo(tx2,MT2+PH2);ctx.stroke();
    ctx.setLineDash([]);ctx.fillStyle='#e74c3c';ctx.font='11px Arial';ctx.textAlign='center';
    ctx.fillText(_dlcpT.toFixed(1)+'%',tx2,MT2-5);ctx.restore();
  }}
  ctx.strokeStyle='#555';ctx.lineWidth=1;ctx.beginPath();ctx.moveTo(ML2,MT2);ctx.lineTo(ML2,MT2+PH2);ctx.lineTo(ML2+PW2,MT2+PH2);ctx.stroke();
  ctx.fillStyle='#555';ctx.font='10px Arial';ctx.textAlign='right';
  for(var yi3=0;yi3<=4;yi3++)ctx.fillText(Math.round(maxC2*yi3/4),ML2-4,MT2+PH2-PH2*yi3/4+4);
  ctx.textAlign='center';
  var rng2=xMx2-xMn2,xstp2=rng2>40?5:rng2>20?2:1,xs2=Math.ceil(xMn2/xstp2)*xstp2;
  for(var xv2=xs2;xv2<=xMx2;xv2+=xstp2)ctx.fillText(xv2+'%',xp2(xv2),MT2+PH2+13);
  ctx.fillStyle='#2c3e50';ctx.font='bold 11px Arial';ctx.textAlign='center';
  ctx.fillText('UPM %',ML2+PW2/2,H-4);
  ctx.save();ctx.translate(12,MT2+PH2/2);ctx.rotate(-Math.PI/2);ctx.fillText('Count',0,0);ctx.restore();
}}

// ResizeObserver for CDF canvas is now set up in showTab('dlcp') after the tab
// is visible, so it fires with the real canvas dimensions (see showTab).

// ═══════════════════════════════════════ PARETO COMMENTS ══════════════════
const COMMENT_KEY = 'pareto_comments';
function loadComments() {{
  try {{ return JSON.parse(localStorage.getItem(COMMENT_KEY) || '{{}}'); }} catch(e) {{ return {{}}; }}
}}
function saveComment(ib, text) {{
  const all = loadComments();
  if (text) all[ib] = text; else delete all[ib];
  localStorage.setItem(COMMENT_KEY, JSON.stringify(all));
}}
function initParetoComments() {{
  const saved = loadComments();
  document.querySelectorAll('.pareto-comment').forEach(ta => {{
    const ib = ta.dataset.ib;
    if (saved[ib]) {{ ta.value = saved[ib]; ta.classList.add('saved'); }}
    ta.addEventListener('input', () => ta.classList.remove('saved'));
    ta.addEventListener('blur', () => {{
      saveComment(ib, ta.value.trim());
      ta.classList.toggle('saved', !!ta.value.trim());
    }});
  }});
}}

function exportParetoTableCsv() {{
  const saved = loadComments();
  const rows = [['Interface Bin', 'Description', 'Total Tested', 'Fail Count', 'Fail %', 'Comment']];
  document.querySelectorAll('#pareto-summary-tbl tbody tr').forEach(tr => {{
    const cells = tr.querySelectorAll('td');
    if (cells.length < 5) return;
    const ib    = cells[0].textContent.trim();
    const desc  = cells[1].textContent.trim();
    const total = cells[2].textContent.trim();
    const nf    = cells[3].textContent.trim();
    const pct   = cells[4].textContent.trim();
    const cmt   = saved[ib] || '';
    rows.push([ib, desc, total, nf, pct, cmt]);
  }});
  const csv = rows.map(r => r.map(v => '"' + String(v).replace(/"/g, '""') + '"').join(',')).join('\\n');
  const a = document.createElement('a');
  a.href = 'data:text/csv;charset=utf-8,' + encodeURIComponent(csv);
  a.download = 'pareto_summary.csv';
  a.click();
}}

function exportComments() {{
  const saved = loadComments();
  const rows = [['IB', 'Comment']];
  document.querySelectorAll('.pareto-comment').forEach(ta => {{
    rows.push([ta.dataset.ib, saved[ta.dataset.ib] || '']);
  }});
  const csv = rows.map(r => r.map(v => '"' + String(v).replace(/"/g, '""') + '"').join(',')).join('\\n');
  const a = document.createElement('a');
  a.href = 'data:text/csv;charset=utf-8,' + encodeURIComponent(csv);
  a.download = 'pareto_comments.csv';
  a.click();
}}

function importComments(input) {{
  const file = input.files[0];
  if (!file) return;
  const reader = new FileReader();
  reader.onload = e => {{
    const lines = e.target.result.split(/\\r?\\n/).slice(1); // skip header
    lines.forEach(line => {{
      if (!line.trim()) return;
      const m = line.match(/^"(\\d+)","((?:[^"]|"")*)"/);
      if (!m) return;
      const ib = m[1], text = m[2].replace(/""/g, '"');
      saveComment(ib, text);
    }});
    initParetoComments();  // refresh visible textareas
    input.value = '';       // reset file picker
  }};
  reader.readAsText(file);
}}

var _paretoRendered = false;
window.addEventListener('load', () => {{ rebuildCharts(); initParetoComments(); resizeAllTables(); }});

// Listen to interval radio changes
document.querySelectorAll('input[name="interval"]').forEach(rb =>
  rb.addEventListener('change', rebuildCharts));
document.querySelectorAll('input[name="groupby"]').forEach(
  rb => rb.addEventListener('change', rebuildCharts));

// Listen to date mode radio changes
document.querySelectorAll('input[name="datemode"]').forEach(rb => rb.addEventListener('change', () => {{
  document.getElementById('custom-date-row').style.display =
    document.querySelector('input[name="datemode"]:checked')?.value === 'custom' ? 'block' : 'none';
  rebuildCharts();
}}));
document.getElementById('date-from').addEventListener('change', rebuildCharts);
document.getElementById('date-to').addEventListener('change', rebuildCharts);
</script>
<!-- DLCP column-filter dropdown panel -->
<div id="dlcp-dd-panel-t" class="dlcp-dd-panel">
  <input id="dlcp-dd-srch-t" class="dlcp-dd-srch" type="text" placeholder="Search…" oninput="dlcpDdSearchT(this.value)">
  <div class="dlcp-dd-acts">
    <button onclick="dlcpDdSelAllT()">All</button>
    <button onclick="dlcpDdSelNoneT()">None</button>
  </div>
  <div id="dlcp-dd-list-t" class="dlcp-dd-list"></div>
  <div class="dlcp-dd-foot"><button onclick="dlcpDdApplyT()">Apply</button></div>
</div>
<!-- Wafer-map zoom overlay (shared by all 3 FB-drilldown tabs) -->
<!-- z-index kept below the shared wafer-map tooltip (z-index:99990 in WAFERMAP_JS) so hover tips render above the modal -->
<div id="wm-zoom-overlay" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.65);z-index:99980;align-items:center;justify-content:center" onclick="if(event.target===this)closeWaferZoom()">
  <div style="background:#fff;border-radius:8px;padding:14px;max-width:90vw;max-height:90vh;overflow:auto;position:relative">
    <button onclick="closeWaferZoom()" style="position:absolute;top:6px;right:8px;border:none;background:#e74c3c;color:#fff;border-radius:50%;width:26px;height:26px;cursor:pointer;font-size:14px">&#10005;</button>
    <div id="wm-zoom-title" style="font-size:13px;font-weight:600;margin-bottom:4px;color:#2c3e50"></div>
    <div id="wm-zoom-legend" style="margin-bottom:8px"></div>
    <div id="wm-zoom-canvas"></div>
  </div>
</div>
<script type="application/json" id="dlcp-extra-data">{dlcp_extra_json}</script>
</body>
</html>
'''
    html = _inject_wafermap_js(html)
    output_path.write_text(html, encoding='utf-8')
    print(f'Wrote interactive report: {output_path}')


# ============================================================================
# 8. CLI
# ============================================================================

def main():
    os.umask(0o002)  # ensure generated files are group-writable on NFS/Samba
    ap = argparse.ArgumentParser(description='Interactive iBin Fail vs. Yield Trend')
    ap.add_argument('csv', help='Input CSV file')
    ap.add_argument('--cfg', default='',
                    help='Product config JSON (auto-detected if omitted)')
    ap.add_argument('--interval', choices=INTERVALS, default='revision')
    ap.add_argument('--topn',   type=int,   default=8)
    ap.add_argument('--thresh', type=float, default=0.0)
    ap.add_argument('--group',  choices=['wafer', 'lot'], default='wafer',
                    help='Histogram grouping: lot (default) = one bar per lot, wafer = one bar per wafer')
    ap.add_argument('--out',    default='',
                    help='Output HTML path (default: <csv>_trend.html)')
    args = ap.parse_args()

    if not HAVE_PLOTLY:
        print('ERROR: plotly not installed.  Run: pip install plotly', file=sys.stderr)
        sys.exit(1)

    csv_path = Path(args.csv).resolve()
    if not csv_path.exists():
        print(f'ERROR: file not found: {csv_path}', file=sys.stderr)
        sys.exit(1)

    # Load product config
    cfg, cfg_path = None, ''
    cfg_src = args.cfg or ''
    if not cfg_src:
        runs_preview = load_csv(csv_path, grouping_mode=args.group)
        drs = runs_preview[0].get('devrevstep', '') if runs_preview else ''
        auto = _find_auto_config(drs)
        if auto:
            cfg_src = str(auto)
            print(f'Auto-detected product config: {auto.name}')
    if cfg_src and Path(cfg_src).exists():
        cfg      = load_product_config(cfg_src)
        cfg_path = cfg_src
        print(f'Loaded product config: {Path(cfg_src).name}')
    else:
        print('No product config - ibin names and targets not shown.')

    print(f'Loading {csv_path} ... (grouping mode: {args.group})')
    runs = load_csv(csv_path, log=lambda s: print(s, end=''), grouping_mode=args.group)
    print(f'Loaded {len(runs)} run(s).')

    groups = group_runs(_aggregate_by_lot(runs), args.interval)
    print(f'Grouped into {len(groups)} {args.interval} period(s).')

    print('Building charts ...')
    trend_fig  = build_trend_chart(groups, top_n_fail_ibins=args.topn,
                                    fail_thresh_pct=args.thresh,
                                    interval=args.interval, cfg=cfg)
    pareto_fig = build_pareto_chart(runs, top_n=20, cfg=cfg)
    pareto_vertical_fig, pareto_table_rows = build_pareto_vertical_chart(runs, top_n=20, cfg=cfg)

    out_path = (Path(args.out).resolve() if args.out
                else csv_path.parent / (csv_path.stem + '_trend.html'))
    out_path = _safe_html_out_path(out_path, csv_path.stem + '_trend.html')
    generate_html(csv_path, groups, runs, trend_fig, pareto_fig, out_path,
                  interval=args.interval, top_n=args.topn, cfg_path=cfg_path, cfg=cfg,
                  pareto_vertical_fig=pareto_vertical_fig,
                  pareto_table_rows=pareto_table_rows,
                  grouping_mode=args.group)


if __name__ == '__main__':
    main()


# ════════════════════════════════════════════════════════════════
# (formerly trend_chart_frame.py)
# ════════════════════════════════════════════════════════════════
import os
import sys
import threading
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

sys.path.insert(0, str(Path(__file__).parent))

# ── Palette (matches dashboard theme) ────────────────────────────────────────
BG    = '#1a252f'
BG2   = '#2c3e50'
BG3   = '#243342'   # slightly lighter — right-panel card bg
FG    = '#ecf0f1'
FG2   = '#95a5a6'
BLUE  = '#2980b9'
ABLU  = '#3498db'
GRN   = '#27ae60'
AGRN  = '#2ecc71'
WARN  = '#f39c12'
RED   = '#c0392b'


def _btn(parent, text, cmd, color=BLUE, acolor=ABLU, width=None):
    kw = dict(text=text, command=cmd, bg=color, fg='white',
              activebackground=acolor, activeforeground='white',
              relief='flat', cursor='hand2', font=('Arial', 9),
              padx=8, pady=3)
    if width:
        kw['width'] = width
    return tk.Button(parent, **kw)


def _lf(parent, text, label_color=FG2):
    return tk.LabelFrame(parent, text=text, bg=BG, fg=label_color,
                         font=('Arial', 8, 'bold'), padx=6, pady=4,
                         relief='groove', bd=1)


def _sep(parent):
    """Thin horizontal divider."""
    tk.Frame(parent, bg=BG2, height=1).pack(fill='x', padx=6, pady=4)


# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
class TrendChartFrame(tk.Frame):
    """Simple GUI: pick CSV + config, set interval/top-n, generate interactive HTML."""

    def __init__(self, parent=None, **kw):
        super().__init__(parent, bg=BG, **kw)
        self._csv_var     = tk.StringVar()
        self._cfg_var     = tk.StringVar()
        self._out_var     = tk.StringVar()
        self._last_report = ''
        self._build_ui()

    def _build_ui(self):
        tk.Label(self, text='iBin Fail vs. Yield Trend',
                 bg=BG, fg=ABLU, font=('Arial', 13, 'bold')
                 ).pack(fill='x', padx=12, pady=(10, 2))
        tk.Label(self,
                 text='Generates a self-contained interactive HTML report.\n'
                      'Filter by program, bin, and interval directly in the browser.',
                 bg=BG, fg=FG2, font=('Arial', 9), justify='left'
                 ).pack(fill='x', padx=12)
        _sep(self)

        # CSV
        tk.Label(self, text='Input File (.csv / .zip / .gz)',
                 bg=BG, fg=FG, font=('Arial', 9, 'bold')
                 ).pack(fill='x', padx=12)
        fr = tk.Frame(self, bg=BG)
        fr.pack(fill='x', padx=12, pady=(2, 6))
        tk.Entry(fr, textvariable=self._csv_var,
                 bg=BG2, fg=FG, insertbackground=FG, relief='flat',
                 font=('Consolas', 9)
                 ).pack(side='left', fill='x', expand=True, padx=(0, 4))
        _btn(fr, 'Browse...', self._browse_csv, width=8).pack(side='left')

        # Product config
        tk.Label(self, text='Product Config JSON (optional)',
                 bg=BG, fg=FG, font=('Arial', 9, 'bold')
                 ).pack(fill='x', padx=12)
        fr2 = tk.Frame(self, bg=BG)
        fr2.pack(fill='x', padx=12, pady=(2, 6))
        tk.Entry(fr2, textvariable=self._cfg_var,
                 bg=BG2, fg=FG, insertbackground=FG, relief='flat',
                 font=('Consolas', 9)
                 ).pack(side='left', fill='x', expand=True, padx=(0, 4))
        _btn(fr2, 'Browse...', self._browse_cfg, color='#1f618d', width=8).pack(side='left', padx=(0, 3))
        _btn(fr2, 'X', lambda: self._cfg_var.set(''), color='#555', width=2).pack(side='left')

        _sep(self)

        # Output path
        tk.Label(self, text='Output HTML', bg=BG, fg=FG,
                 font=('Arial', 9, 'bold')).pack(fill='x', padx=12, pady=(8, 0))
        fr3 = tk.Frame(self, bg=BG)
        fr3.pack(fill='x', padx=12, pady=(2, 6))
        tk.Entry(fr3, textvariable=self._out_var,
                 bg=BG2, fg=FG, insertbackground=FG, relief='flat',
                 font=('Consolas', 9)
                 ).pack(side='left', fill='x', expand=True, padx=(0, 4))
        _btn(fr3, '...', self._browse_out, width=3).pack(side='left')

        _sep(self)

        _sep(self)

        # Buttons
        btn_row = tk.Frame(self, bg=BG)
        btn_row.pack(fill='x', padx=12)
        self._run_btn = _btn(btn_row, 'Generate Interactive HTML',
                             self._generate, color=GRN, acolor=AGRN)
        self._run_btn.config(font=('Arial', 10, 'bold'), pady=6)
        self._run_btn.pack(side='left', fill='x', expand=True, padx=(0, 4))
        self._open_btn = _btn(btn_row, 'Open', self._open_report,
                              color='#935116', acolor='#ca6f1e')
        self._open_btn.config(state='disabled', pady=6)
        self._open_btn.pack(side='left')

        # Log
        _sep(self)
        tk.Label(self, text='Log', bg=BG, fg=FG2,
                 font=('Arial', 8, 'bold')).pack(fill='x', padx=12)
        log_frm = tk.Frame(self, bg='#0d1b26')
        log_frm.pack(fill='both', expand=True, padx=12, pady=(2, 10))
        self._log = tk.Text(log_frm, state='disabled',
                            font=('Consolas', 8), bg='#0d1b26', fg='#a8d8ea',
                            relief='flat', wrap='word')
        sb = tk.Scrollbar(log_frm, command=self._log.yview, bg=BG)
        self._log.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        self._log.pack(fill='both', expand=True)

    def _browse_csv(self):
        p = filedialog.askopenfilename(
            title='Select input CSV / ZIP / GZ',
            filetypes=[('Supported files', '*.csv *.zip *.gz *.gzip'),
                       ('All files', '*.*')])
        if not p:
            return
        self._csv_var.set(p)
        self._out_var.set(str(Path(p).parent / (Path(p).stem + '_trend.html')))
        if not self._cfg_var.get().strip():
            try:
                preview = load_csv(Path(p))
                drs = preview[0].get('devrevstep', '') if preview else ''
            except Exception:
                drs = ''
            auto = _find_auto_config(drs)
            if auto:
                self._cfg_var.set(str(auto))
                self._log_write(f'Auto-detected config: {auto.name}\n')

    def _browse_cfg(self):
        p = filedialog.askopenfilename(
            title='Select product config JSON',
            filetypes=[('JSON files', '*.json'), ('All files', '*.*')])
        if p:
            self._cfg_var.set(p)

    def _browse_out(self):
        p = filedialog.asksaveasfilename(
            title='Save report as', defaultextension='.html',
            filetypes=[('HTML files', '*.html')])
        if p:
            self._out_var.set(p)

    def _open_report(self):
        if self._last_report and os.path.isfile(self._last_report):
            try:
                os.startfile(self._last_report)
            except Exception as e:
                messagebox.showerror('Error', str(e))

    def _log_write(self, msg: str):
        def _do():
            self._log.configure(state='normal')
            self._log.insert('end', msg)
            self._log.see('end')
            self._log.configure(state='disabled')
        try:
            self.after(0, _do)
        except Exception:
            pass

    def _generate(self):
        csv_str = self._csv_var.get().strip()
        if not csv_str or not os.path.isfile(csv_str):
            messagebox.showwarning('No CSV', 'Browse and select a CSV file first.')
            return

        csv_path = Path(csv_str)
        out_str  = self._out_var.get().strip()
        out_path = Path(out_str) if out_str else csv_path.parent / (csv_path.stem + '_trend.html')
        out_path = _safe_html_out_path(out_path, csv_path.stem + '_trend.html')
        interval = 'revision'
        top_n    = 8
        thresh   = 0.0
        cfg_path = self._cfg_var.get().strip()

        self._run_btn.configure(state='disabled', text='Working...', bg=FG2)
        self._log_write(f'Loading {csv_path.name}...\n')

        def _worker():
            try:
                runs = load_csv(csv_path, log=self._log_write, grouping_mode='wafer')
                self._log_write(f'Loaded {len(runs)} run(s). Building charts...\n')

                cfg = None
                if cfg_path and Path(cfg_path).exists():
                    cfg = load_product_config(cfg_path)
                    self._log_write(f'Config: {Path(cfg_path).name}\n')
                else:
                    drs = runs[0].get('devrevstep', '') if runs else ''
                    auto = _find_auto_config(drs)
                    if auto:
                        cfg = load_product_config(auto)
                        self._log_write(f'Config (auto): {auto.name}\n')

                groups    = group_runs(_aggregate_by_lot(runs), interval)
                trend_fig = build_trend_chart(
                    groups, top_n_fail_ibins=top_n,
                    fail_thresh_pct=thresh, interval=interval, cfg=cfg)
                pareto_fig     = build_pareto_chart(runs, top_n=20, cfg=cfg)
                pareto_vert_fig, pareto_tbl = build_pareto_vertical_chart(runs, top_n=20, cfg=cfg)

                generate_html(csv_path, groups, runs, trend_fig, pareto_fig,
                                 out_path, interval=interval, top_n=top_n,
                                 cfg_path=cfg_path, cfg=cfg,
                                 pareto_vertical_fig=pareto_vert_fig,
                                 pareto_table_rows=pareto_tbl)
                self._last_report = str(out_path)
                self._log_write(f'Done -> {out_path}\n')
                try:
                    os.startfile(str(out_path))
                except Exception:
                    pass

                def _done():
                    self._open_btn.configure(state='normal')
                self.after(0, _done)

            except Exception as exc:
                import traceback
                self._log_write(f'ERROR: {exc}\n{traceback.format_exc()}\n')
            finally:
                def _re():
                    self._run_btn.configure(state='normal',
                                            text='Generate Interactive HTML',
                                            bg=GRN)
                self.after(0, _re)

        threading.Thread(target=_worker, daemon=True).start()


if __name__ == '__main__':
    root = tk.Tk()
    root.title('Trend Chart - Debug')
    root.configure(bg='#1a252f')
    root.geometry('820x560')
    frame = TrendChartFrame(root)
    frame.pack(fill='both', expand=True)
    root.mainloop()

