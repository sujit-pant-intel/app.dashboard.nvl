"""_pipeline_html.py - HTML generation mixins."""
import os
from _pipeline_constants import _SRC_DIR, _ROOT_DIR, _FROZEN, _LOADER, SICC_UPM_SCRIPT, SICC_CDYN_UPM_SCRIPT, _wm_inject
from _pipeline_constants import _SRC_DIR, _ROOT_DIR, _FROZEN, _LOADER, SICC_UPM_SCRIPT, SICC_CDYN_UPM_SCRIPT, _wm_inject
import sys


class PipelineHtmlMixin:
    def _build_pareto_html(self, resolved_csv, bindef_csv, out_dir, tag=None, dashboard_html=None, bucket_json=None):
        """Top-10 fail pareto for FUNCTIONAL_BIN_* where INTERFACE_BIN_* > 4.
        - Appends the pareto table into the BinDistribution HTML.
        - Generates a standalone pareto heatmap bar-chart HTML (pareto_heatmap.html).
        Returns path to the heatmap HTML, or None on failure."""
        try:
            import pandas as _pd
            from pathlib import Path as _P
            import base64 as _b64
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as _plt
            import matplotlib.colors as _mcolors
            import matplotlib.patches as _mpatches
            import numpy as _np
            import io as _io
            import re as _re2
            import matplotlib.tri as _mtri
            _pareto_entries = []  # list of (rank_int, short_label_str, html_path)
            _od = _P(out_dir)
            # Detect IBIN wafer maps so BFS rows can link to wafermap.html
            _wm_url = ''
            try:
                _wm_url = 'wafermap.html' if any((_od / 'heatmap').glob('*_IBIN_WaferMap_*.html')) else ''
            except Exception:
                pass

            # Build per-lot wafermap URL map so fbTileClick can navigate to the
            # correct per-lot file with the correct wafer hash (#wafer-W).
            _wm_files_dict: dict = {}
            try:
                if _wm_url:
                    import re as _re_wmf
                    import pandas as _pd_wmf
                    _hd_wmf = _od / 'heatmap'
                    _stem_wmf = _P(resolved_csv).stem
                    def _lot_safe_wmf(s):
                        return _re_wmf.sub(r'[^0-9A-Za-z_-]', '_', str(s))
                    _df_wmf = _pd_wmf.read_csv(str(resolved_csv), dtype=object, low_memory=False)
                    # Use same lot-column priority as generate_heatmap and generate_bin_png:
                    # prefer SORT_LOT (present in all merged CSVs), then exact 'lot',
                    # then any column whose name contains 'lot' but not 'slot'.
                    _lot_col_wmf = next(
                        (c for c in _df_wmf.columns if c.lower() == 'sort_lot'), None)
                    if _lot_col_wmf is None:
                        _lot_col_wmf = next(
                            (c for c in _df_wmf.columns if c.lower() == 'lot'), None)
                    if _lot_col_wmf is None:
                        _lot_col_wmf = next(
                            (c for c in _df_wmf.columns if 'lot' in c.lower() and 'slot' not in c.lower()), None)
                    if _lot_col_wmf:
                        for _lv_wmf in _df_wmf[_lot_col_wmf].dropna().unique():
                            _ls_wmf = _lot_safe_wmf(str(_lv_wmf))
                            _wf_wmf = _hd_wmf / f'{_stem_wmf}_IBIN_WaferMap_{_ls_wmf}.html'
                            if _wf_wmf.exists():
                                _wm_files_dict[str(_lv_wmf)] = f'heatmap/{_wf_wmf.name}'
            except Exception:
                pass

            # ── Digital Dashboard HTML — generated FIRST, independent of fail bins ──────
            # Moved here so early-returns below still include the DD html path.
            _xlsx_tbl_html = ''
            _xlsx_tbl_dyn = ''
            _xlsx_p = None
            _dd_html_path = None
            _dd_js_rows = []
            _dd_js_hdrs = []
            try:
                # ── Digital Dashboard — bin-category breakdown via get_dd_update.py ──
                import sys as _sys_dd0, pandas as _pd_dd0, logging as _log_dd0
                # Suppress info-level chatter from get_dd_update during import/use
                _log_dd0.getLogger('get_dd_update').setLevel(_log_dd0.WARNING)
                _gdd_dir = str(_SRC_DIR)
                if _gdd_dir not in _sys_dd0.path:
                    _sys_dd0.path.insert(0, _gdd_dir)
                from get_dd_update import (
                    moduleMap             as _dd_modmap,
                    _buildBinDefsFromDF   as _dd_buildBinDefs,
                    getYieldByModule      as _dd_getYldByMod,
                    getYieldFromModYield  as _dd_gfy,
                    updateDefeatureModCnts as _dd_updateDfMC,
                )

                try:
                    _dfd = _df_wmf.copy()
                except (NameError, AttributeError):
                    _dfd = _pd_dd0.read_csv(str(resolved_csv), dtype=object, low_memory=False)

                # Ensure DATA_BIN and INTERFACE_BIN are numeric
                _db_col_dd = next(
                    (c for c in _dfd.columns if 'DATA_BIN' in c.upper() and 'TOTAL' not in c.upper()), None)
                # Match same patterns as getIBinCol: INTERFACE_BIN, IB@, or IB DIEBIN
                _ib_col_dd = next(
                    (c for c in _dfd.columns if (
                        ('INTERFACE_BIN' in c.upper() and 'TOTAL' not in c.upper())
                        or 'IB@' in c or 'IB DIEBIN' in c
                    )), None)
                if not _db_col_dd:
                    raise ValueError(
                        f'No DATA_BIN column found — columns: {list(_dfd.columns[:30])}')
                _dfd = _dfd.copy()
                _dfd[_db_col_dd] = _pd_dd0.to_numeric(_dfd[_db_col_dd], errors='coerce')
                # Normalise column names so getBinCol/getIBinCol (case-sensitive) find them
                if _db_col_dd != 'DATA_BIN':
                    _dfd = _dfd.rename(columns={_db_col_dd: 'DATA_BIN'})
                    _db_col_dd = 'DATA_BIN'
                if _ib_col_dd:
                    _dfd[_ib_col_dd] = _pd_dd0.to_numeric(_dfd[_ib_col_dd], errors='coerce')
                    if _ib_col_dd != 'INTERFACE_BIN':
                        _dfd = _dfd.rename(columns={_ib_col_dd: 'INTERFACE_BIN'})
                        _ib_col_dd = 'INTERFACE_BIN'
                    _ib_real_dd = True
                else:
                    # getYieldByModule calls getIBinCol which returns False if no IB column →
                    # df[False] would KeyError.  Provide a synthetic all-zero column instead.
                    _dfd['INTERFACE_BIN'] = 0
                    _ib_col_dd = 'INTERFACE_BIN'
                    _ib_real_dd = False

                if _ib_real_dd:
                    # Actual die = rows where IB is a valid number after to_numeric.
                    # AQUA CSVs may contain footer/summary rows with no IB value (→ NaN).
                    # Filtering them out gives the true die count (matches yield summary).
                    # getYieldByModule handles IB=93 inking die via its numB93 path even
                    # when DATA_BIN is NaN, so passing the full die set is correct.
                    _dfd_full_dd = _dfd[_dfd['INTERFACE_BIN'].notna()].reset_index(drop=True)
                    _numDie_dd   = len(_dfd_full_dd)
                    _dfd         = _dfd_full_dd.dropna(subset=[_db_col_dd]).reset_index(drop=True)
                else:
                    # No real IB → can only use die with valid DATA_BIN; fall back to
                    # yesterday's post-dropna count so Yield SUM stays 100%.
                    _dfd = _dfd.dropna(subset=[_db_col_dd]).reset_index(drop=True)
                    _numDie_dd   = len(_dfd)
                    _dfd_full_dd = _dfd

                if len(_dfd) == 0:
                    raise ValueError('Empty dataframe after dropping NA bins')

                _binDefs_dd  = _dd_buildBinDefs(_dfd)
                # Pass _dfd_full_dd (all die including IB=93) so getYieldByModule's numB93
                # path counts inking/unsort die — same as getDD passing the raw yieldDF.
                _modYield_dd = _dd_getYldByMod(yldDF=_dfd_full_dd, binDefs=_binDefs_dd, modMap=_dd_modmap)

                # Compute Recovery Bins (3-4) per module via LOGTRACKER decoding ─────
                # productInfo varies by product; default to NVL816 values (dfBins=[3,4],
                # numAtoms=16, numCores=8).  Works for NVL816-BLLC too.  For other products
                # with different atom/core counts the LOGTRACKER regex won't match and
                # defeatureModCnts will be empty (safe fallback).
                _df_prod_info_dd = {'dfBins': [3, 4], 'numAtoms': 16, 'numCores': 8}
                _defeature_dd = {}
                try:
                    _dd_updateDfMC(dfModCnts=_defeature_dd, yieldDF=_dfd_full_dd,
                                   productInfo=_df_prod_info_dd, waferLvl=False,
                                   binDefs=_binDefs_dd, modMap=_dd_modmap)
                except Exception:
                    _defeature_dd = {}

                if _modYield_dd:
                    def _esc0(s):
                        return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    def _fmt_dd0(n, nd):
                        return '0.00%' if n == 0 else f'{n / nd * 100:.2f}% ({n:,})'
                    _gfy0     = lambda cat: _dd_gfy(_modYield_dd, 'dd', cat)
                    _tag_str0 = f' &mdash; {_esc0(str(tag))}' if tag else ''
                    _SUM_BG   = '#e0ede0'

                    # Derive wafer count and TP string for column headers (matching makeOutXl)
                    _nw_dd = 0
                    try:
                        _wfr_col_dd = next((c for c in _dfd.columns
                                            if c.upper() == 'SORT_WAFER'), None)
                        _lot_col_dd = next((c for c in _dfd.columns
                                            if c.upper() == 'SORT_LOT'), None)
                        if _wfr_col_dd and _lot_col_dd:
                            _nw_dd = int(_dfd.groupby([_lot_col_dd, _wfr_col_dd]).ngroups)
                        elif _wfr_col_dd:
                            _nw_dd = int(_dfd[_wfr_col_dd].nunique())
                    except Exception:
                        pass
                    _tp_dd = ''
                    try:
                        import re as _re_tp_dd
                        _prog_col_dd = next((c for c in _dfd.columns
                                             if 'PROGRAM' in c.upper()), None)
                        if _prog_col_dd:
                            _pname_dd = str(_dfd[_prog_col_dd].dropna().iloc[0])
                            _m_tp = _re_tp_dd.search(r'([5-7]\d[A-Za-z0-9_]+)', _pname_dd)
                            if _m_tp:
                                _tp_dd = _m_tp.group(1)
                    except Exception:
                        pass
                    _nw_str = f'{_nw_dd}W ' if _nw_dd else ''
                    _tp_str = f'{_tp_dd} ' if _tp_dd else ''

                    # Rows: list of (cells_list, is_sum, bg_color)
                    # cells = [label, col2_loss%, col3_recovery%]
                    # col3 (Recovery Bins 3-4) comes from _defeature_dd computed via
                    # updateDefeatureModCnts / LOGTRACKER decode.  Blank for rows that
                    # makeOutXl leaves empty (CCF, NONCCF, RESET, Bin, Repair, Others).
                    _dd_rows_raw = []
                    _all_grp3_sum = 0.0  # cumulative col3 for Bins SUM row

                    def _fmt_c3(mod_key):
                        """Return formatted col3 string for a given module key."""
                        _n3 = _defeature_dd.get(mod_key, 0)
                        return f'{_n3 / _numDie_dd * 100:.1f}%' if _n3 else '0.0%'

                    # ARR / FUN / SCN groups
                    _GRP_BG = {'ARR': '#fef9f0', 'FUN': '#f0f8ff', 'SCN': '#f0fff4'}
                    for _grp0 in ['ARR', 'FUN', 'SCN']:
                        _gbg0  = _GRP_BG[_grp0]
                        _gn0   = 0
                        _grp3  = 0.0
                        for _sub0 in ['ATOM', 'CCF', 'CORE', 'NONCCF']:
                            _n0 = _gfy0(f'{_grp0}_{_sub0}')
                            _gn0 += _n0
                            if _sub0 in ('ATOM', 'CORE'):
                                _mk3 = f'{_grp0}_{_sub0}'
                                _n3 = _defeature_dd.get(_mk3, 0)
                                _grp3 += _n3 / _numDie_dd
                                _c3 = _fmt_c3(_mk3)
                            else:
                                _c3 = ''
                            _dd_rows_raw.append(
                                ([f'{_grp0}_{_sub0}', _fmt_dd0(_n0, _numDie_dd), _c3], False, _gbg0))
                        _grp3_str = f'{_grp3 * 100:.1f}%' if _grp3 else '0.0%'
                        _dd_rows_raw.append(
                            (['SUM', _fmt_dd0(_gn0, _numDie_dd), _grp3_str], True, _SUM_BG))
                        _all_grp3_sum += _grp3

                    # RESET
                    _n0 = _gfy0('Reset')
                    _dd_rows_raw.append((['RESET (19,35)', _fmt_dd0(_n0, _numDie_dd), ''], False, '#fff'))

                    # Pass bins
                    _bins_n0 = 0
                    for _bi0 in [1, 2, 3, 4]:
                        if _bi0 == 1:
                            _n0 = _gfy0('Bin 1') + _gfy0('Bin 198 (Vmin Repair)')
                        elif _bi0 == 2:
                            _n0 = _gfy0('Bin 2 (Hard Repair)') + _gfy0('Bin 202 (Vmax Repair)')
                        else:
                            _n0 = _gfy0(f'Bin {_bi0}')
                        _bins_n0 += _n0
                        _dd_rows_raw.append(([f'Bin {_bi0}', _fmt_dd0(_n0, _numDie_dd), ''], False, '#eaf4ff'))
                    _bins_sum_c3 = f'{_all_grp3_sum * 100:.1f}%' if _all_grp3_sum else '0.0%'
                    _dd_rows_raw.append((['SUM', _fmt_dd0(_bins_n0, _numDie_dd), _bins_sum_c3], True, _SUM_BG))

                    # Repair rows
                    _rep_n0 = 0
                    for _rname0 in ['Bin 198 (Vmin Repair)', 'Bin 202 (Vmax Repair)', 'Bin 2 (Hard Repair)']:
                        _n0 = _gfy0(_rname0)
                        _rep_n0 += _n0
                        _dd_rows_raw.append(
                            ([f'Repair {_rname0}', _fmt_dd0(_n0, _numDie_dd), ''], False, '#fffbe6'))
                    _dd_rows_raw.append((['SUM', _fmt_dd0(_rep_n0, _numDie_dd), ''], True, _SUM_BG))

                    # Other categories (only show if non-zero)
                    for _oname0 in ['Analog', 'TPI Foundry', 'TPI Other', 'TPI Other - B93',
                                    'TPI Other - B98', 'TPI Other - B99', 'HVQK (B26)']:
                        _n0 = _gfy0(_oname0)
                        if _n0 > 0:
                            _dd_rows_raw.append(([_oname0, _fmt_dd0(_n0, _numDie_dd), ''], False, '#fff'))

                    # Yield SUM
                    _yield_sum_n0 = sum(_gfy0(b) for b in [
                        'Bin 1', 'Bin 198 (Vmin Repair)', 'Bin 2 (Hard Repair)',
                        'Bin 202 (Vmax Repair)', 'Bin 3', 'Bin 4',
                        'Reset', 'ARR_ATOM', 'ARR_CCF', 'ARR_CORE', 'ARR_NONCCF',
                        'FUN_ATOM', 'FUN_CCF', 'FUN_CORE', 'FUN_NONCCF',
                        'SCN_ATOM', 'SCN_CCF', 'SCN_CORE', 'SCN_NONCCF',
                        'Analog', 'TPI Foundry', 'TPI Other', 'TPI Other - B93',
                        'TPI Other - B98', 'TPI Other - B99', 'HVQK (B26)',
                    ])
                    _dd_rows_raw.append(
                        (['Yield SUM (%)', _fmt_dd0(_yield_sum_n0, _numDie_dd), ''], True, '#d4edda'))

                    # Build HTML — 3-column format matching makeOutXl output
                    _hdrs_dd = [
                        'Sub Module',
                        f'{_nw_str}{_tp_str}Yield Loss (Fail Bins) (%)',
                        f'{_nw_str}{_tp_str}Recovery Bins (3-4) (%)',
                    ]
                    _dd_js_hdrs = _hdrs_dd
                    _hdr_html0  = ''.join(f'<th>{_esc0(h)}</th>' for h in _hdrs_dd)
                    _body_html0 = ''
                    for _row0, _is_sum0, _clr0 in _dd_rows_raw:
                        _cells0 = [_esc0(str(v)) for v in _row0]
                        _dd_js_rows.append({'cells': _cells0, 'bg': _clr0, 'bold': bool(_is_sum0)})
                        _sty0 = (f'font-weight:bold;border-top:1px solid #bbb;background:{_clr0}'
                                 if _is_sum0 else f'background:{_clr0}')
                        _body_html0 += (
                            f'<tr style="{_sty0}">'
                            f'<td>{_cells0[0]}</td>'
                            f'<td class="num">{_cells0[1]}</td>'
                            f'<td class="num">{_cells0[2]}</td>'
                            f'</tr>\n'
                        )
                    _xlsx_tbl_html = (
                        f'<h3 style="font-size:13px;margin:14px 0 6px;color:#2c3e50">'
                        f'&#128196; Digital Dashboard Summary{_tag_str0}</h3>'
                        f'<div style="overflow-x:auto;margin-bottom:14px">'
                        f'<table class="pareto-tbl">'
                        f'<thead><tr>{_hdr_html0}</tr></thead>'
                        f'<tbody>{_body_html0}</tbody></table></div>'
                    )
                    _xlsx_tbl_dyn = (
                        f'<h3 style="font-size:13px;margin:14px 0 6px;color:#2c3e50">'
                        f'&#128196; Digital Dashboard Summary{_tag_str0}</h3>'
                        f'<div style="overflow-x:auto;margin-bottom:14px">'
                        f'<table class="pareto-tbl">'
                        f'<thead id="dd-thead"><tr>{_hdr_html0}</tr></thead>'
                        f'<tbody id="dd-tbody"></tbody></table></div>'
                    )
            except Exception as _dd_exc0:
                # Show the error in the DD section rather than silently leaving it blank
                _xlsx_tbl_html = (
                    f'<p style="color:#c0392b;font-size:11px;border:1px solid #e74c3c;'
                    f'background:#fdf0ef;padding:4px 8px;margin:8px 0;border-radius:3px">'
                    f'<b>DD build error:</b> {str(_dd_exc0)[:500]}</p>'
                )

            if False:  # digital_dashboard.html removed — DD table embedded in BinDistribution.html
                try:
                    _opener_dd0 = None
                    _tbl_body0 = (_xlsx_tbl_dyn if _xlsx_tbl_dyn else
                                  _xlsx_tbl_html if _xlsx_tbl_html else '')
                    import json as _json_dd0
                    _dd_script0 = (
                        '<script>\nvar DD_ROWS=' + _json_dd0.dumps(_dd_js_rows, ensure_ascii=False) + ';\n'
                        + r'''function esc(s){return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
function _ddFmtCell(val){
  var raw=String(val);var bm=raw.match(/^(.+?)\s*\((\d[\d,]*)\)$/);
  if(!bm)return esc(raw);
  var ns=bm[2].replace(/,/g,'').replace(/\B(?=(\d{3})+(?!\d))/g,',');
  return esc(bm[1])+' <span style="font-size:10px;color:#aaa;font-weight:normal">('+ns+')</span>';
}
function rDDTbl(){var tbody=document.getElementById('dd-tbody');if(!tbody)return;var rows=DD_ROWS.slice();var html='';rows.forEach(function(row){var ts=row.bold?'font-weight:bold;border-top:2px solid #aaa;background:'+row.bg:'background:'+row.bg;html+='<tr style="'+ts+'">';row.cells.forEach(function(c,i){if(i<1)html+='<td>'+esc(c)+'</td>';else html+='<td class="num">'+_ddFmtCell(c)+'</td>';});html+='</tr>';});tbody.innerHTML=html;}
if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',rDDTbl);}else{rDDTbl();}
</script>'''
                    ) if _dd_js_rows else ''
                    _dd_page0 = (
                        f'<!doctype html>\n<html>\n<head><meta charset="utf-8">'
                        f'<title>Digital Dashboard{_tag_str_dd0}</title>\n'
                        f'{_CSS0}\n'
                        f'<style>\nbody{{font-family:Arial,sans-serif;background:#f4f6f8;padding:20px}}\n'
                        f'.dd-open-btn{{display:inline-block;padding:6px 14px;background:#2980b9;'
                        f'color:#fff;border-radius:4px;text-decoration:none;font-size:12px;margin-bottom:16px}}\n'
                        f'.dd-open-btn:hover{{background:#3498db}}\n'
                        f'.sort-btn{{cursor:pointer;user-select:none;padding-right:8px!important}}\n'
                        f'.sort-btn:hover{{background:#3d5166!important}}\n'
                        f'.sort-arr{{font-size:9px;margin-left:2px;opacity:.7}}\n'
                        f'</style>\n</head>\n<body>\n'
                        f'<a class="dd-open-btn" href="{_xlsx_href_dd0}" target="_blank">'
                        f'&#128196; Open {os.path.basename(_xlsx_p)} in Excel</a>\n'
                        f'{_tbl_body0}\n{_dd_script0}\n</body>\n</html>'
                    )
                    _dd_html_path = _od / 'digital_dashboard.html'
                    _dd_html_path.write_text(_wm_inject(_dd_page0), encoding='utf-8')
                    # Only copy *_out.xlsx into the run folder (not the cumulative dashboard)
                    try:
                        import shutil as _shutil0
                        if _P(_xlsx_p).name.endswith('_out.xlsx'):
                            _xlsx_dest0 = _od / _P(_xlsx_p).name
                            if _xlsx_dest0.resolve() != _P(_xlsx_p).resolve():
                                _shutil0.copy2(_xlsx_p, str(_xlsx_dest0))
                    except Exception:
                        pass
                except Exception:
                    _dd_html_path = None

            # ── PARETO ANALYSIS (requires fail bins) — early returns now include DD path ──
            df = _pd.read_csv(resolved_csv, dtype=object)
            ib_col = next((c for c in df.columns if 'INTERFACE_BIN' in c.upper() and 'TOTAL' not in c.upper()), None)
            fb_col = next((c for c in df.columns if 'FUNCTIONAL_BIN' in c.upper() and 'TOTAL' not in c.upper()), None)
            if not ib_col or not fb_col:
                return None, [], _dd_html_path, _xlsx_p
            df[ib_col] = _pd.to_numeric(df[ib_col], errors='coerce')
            df[fb_col] = _pd.to_numeric(df[fb_col], errors='coerce')
            # ── deduplicate: one row per physical die (lot + wafer + X + Y) ──
            # CSVs merged across programs can have multiple rows per die, inflating counts.
            try:
                _lot_c = (next((c for c in df.columns if c.lower() == 'sort_lot'), None)
                          or next((c for c in df.columns if c.lower() == 'lot'), None)
                          or next((c for c in df.columns if 'lot' in c.lower() and 'slot' not in c.lower()), None))
                _wfr_c = (next((c for c in df.columns if c.lower() == 'sort_wafer'), None)
                          or next((c for c in df.columns if 'wafer' in c.lower()), None))
                _x_c  = (next((c for c in df.columns if c.lower() == 'sort_x'), None)
                          or next((c for c in df.columns if 'sort_x' in c.lower()), None))
                _y_c  = (next((c for c in df.columns if c.lower() == 'sort_y'), None)
                          or next((c for c in df.columns if 'sort_y' in c.lower()), None))
                _dedup_key = [c for c in [_lot_c, _wfr_c, _x_c, _y_c] if c]
                if _dedup_key:
                    df = df.drop_duplicates(subset=_dedup_key, keep='last').reset_index(drop=True)
            except Exception:
                pass
            total_all = len(df)
            if total_all == 0:
                return None, [], _dd_html_path, _xlsx_p
            filtered = df[df[ib_col] > 4].copy()
            if filtered.empty:
                return None, [], _dd_html_path, _xlsx_p
            total_fail = len(filtered)

            counts = (filtered.groupby([fb_col])
                      .size().reset_index(name='FailCount'))
            counts = counts.sort_values('FailCount', ascending=False).reset_index(drop=True)
            if counts.empty:
                return None, [], _dd_html_path, _xlsx_p
            counts['FailPct'] = counts['FailCount'] / total_all * 100
            counts = counts.sort_values(['FailPct', 'FailCount', fb_col], ascending=[False, False, True]).reset_index(drop=True)

            # bindef lookup:
            #   Primary : 'Bin Description_{op}' column from AQUA CSV — the raw test-program
            #             leaf-bin name (e.g. B62500000_FAIL_ARR_CCF_SSA_…).  Often contains
            #             8-digit leaf-bin names that are too verbose to display as-is.
            #   Secondary: bindef_csv file — typically has shorter shared-bin names
            #             (e.g. B6250_FAIL_ARR_CCF_SHARED_BIN) that are more readable.
            #   When the CSV column value matches the 8-digit leaf-bin pattern (B\d{7,}_),
            #   the secondary dict is tried before giving up.
            bindef_dict = {}
            _bindef_file_dict = {}
            _bd_col = next((c for c in df.columns if c.startswith('Bin Description_')), None)
            if _bd_col:
                try:
                    _bd_sub = df[[fb_col, _bd_col]].dropna(subset=[fb_col]).copy()
                    _bd_sub = _bd_sub.drop_duplicates(subset=[fb_col])
                    bindef_dict = {
                        f'FB{int(float(k))}': str(v)
                        for k, v in zip(_bd_sub[fb_col], _bd_sub[_bd_col])
                        if _pd.notna(v) and str(v).strip()
                    }
                except Exception:
                    pass
            if bindef_csv and os.path.isfile(str(bindef_csv)):
                try:
                    bd = _pd.read_csv(bindef_csv, header=0)
                    cols = bd.columns.tolist()
                    _bindef_file_dict = dict(zip(bd[cols[0]].astype(str), bd[cols[1]].astype(str)))
                except Exception:
                    pass
            if not bindef_dict:
                bindef_dict = _bindef_file_dict

            # fail bucket lookup: load from yieldtarget_info JSON if available, else built-ins
            import json as _json_yt
            _BUILTIN_BUCKETS = [
                ('1/2',   'SDS FF yield'),
                ('1/2/3/4', 'SDS FF+DF yield'),
                ('1',    'SDS FF (No Repair) yield'),
                ('2',    'MBIST Repair'),
                ('3/4',  'Recovery (Defeatured)'),
                ('3',    'Recovery (Atom Defeatured)'),
                ('4',    'Recovery (Core Defeatured)'),
                ('41/42/47/76/77/81/82', 'SCAN (post-recovery)'),
                ('20/21/33/60/61/62/63/65', 'ARRAY MBIST (post-recovery)'),
                ('11/13/16/25/27/28/32/36/39/46/48/51/64/71/74/75', 'ANALOG (post-recovery)'),
                ('7/8/9/10/15/18/43', 'TPI (Foundry)'),
                ('31/88/91/94/97/98/99/93', 'TPI (Bump/DiePrep/Test)'),
                ('19/35', 'RESET'),
                ('12/44/45/70/80/85/86', 'Functional'),
                ('26',   'HVQK'),
            ]
            _bin_map = {}  # bin_number_str -> {"cat":..., "desc":...}
            _pass_bin_map = {}  # pass fb_number_str -> {"cat":..., "desc":...}
            _yt_buckets = _BUILTIN_BUCKETS  # list of (bin_str, label)
            _yt_file = dashboard_html  # start with dashboard path hint
            # Search for yieldtarget JSON next to the CSV, or next to the dashboard
            for _yt_search_dir in filter(None, [_od, _P(resolved_csv).parent,
                                                _P(dashboard_html).parent if dashboard_html else None]):
                for _yt_cand in _P(_yt_search_dir).glob('yieldtarget_input*.json'):
                    _yt_file = str(_yt_cand)
                    break
                if _yt_file and str(_yt_file) != str(dashboard_html):
                    break
            # Also accept the bucket_json arg (yieldtarget_info field)
            if dashboard_html and not (_yt_file and _yt_file != dashboard_html):
                pass  # keep searching
            _yt_json_path = bucket_json or _yt_file
            _fb93xx_map = {}  # {fb_str: description} for 93xx handler/skip bins
            if _yt_json_path and os.path.isfile(str(_yt_json_path)):
                try:
                    _yt_data = _json_yt.loads(open(str(_yt_json_path), encoding='utf-8').read())
                    # Support both formats:
                    #   flat-array: [{"bin":"1/2", "fail_bucket":"...", ...}, ...]
                    #   new dict:   {"yield_targets": [...], "bin_map": {...}, ...}
                    if isinstance(_yt_data, list):
                        _yt_entries = _yt_data
                    else:
                        _yt_entries = _yt_data.get('yield_targets', [])
                        _bin_map = {str(k): v for k, v in _yt_data.get('bin_map', {}).items()}
                        _pass_bin_map = {str(k): v for k, v in _yt_data.get('Pass-Bin-Map', {}).items()}
                        for _e93 in _yt_data.get('fB93xx', []):
                            if isinstance(_e93, dict) and 'FB' in _e93 and 'description' in _e93:
                                _fb93xx_map[str(_e93['FB'])] = _e93['description']
                    if _yt_entries:
                        _yt_buckets = [(e['bin'], e['fail_bucket']) for e in _yt_entries
                                       if 'bin' in e and 'fail_bucket' in e]
                except Exception:
                    pass
                # fB93xx may be in a malformed outer block — extract via regex as fallback
                if not _fb93xx_map:
                    try:
                        import re as _re_93
                        _raw93 = open(str(_yt_json_path), encoding='utf-8').read()
                        _m93 = _re_93.search(r'"fB93xx"\s*:\s*(\[.*?\])', _raw93, _re_93.DOTALL)
                        if _m93:
                            for _e93r in _json_yt.loads(_m93.group(1)):
                                if isinstance(_e93r, dict) and 'FB' in _e93r and 'description' in _e93r:
                                    _fb93xx_map[str(_e93r['FB'])] = _e93r['description']
                    except Exception:
                        pass
            _ib_to_bucket = {}
            for _tok_field, _blabel in _yt_buckets:
                for _t in _re2.findall(r'\d+', _tok_field):
                    if _t not in _ib_to_bucket:
                        _ib_to_bucket[_t] = _blabel

            def _bucket(ib_val):
                return _ib_to_bucket.get(str(int(ib_val)), '')

            def _bin_map_cat(n_str):
                """Return (cat, desc): cat from bin_map (IB prefix fallback),
                desc from bindef_dict with _bindef_file_dict as fallback when the
                primary value is an 8-digit leaf-bin name.
                1-digit prefix fallback is only tried for ≤3-digit FBs (e.g. 301→IB 3);
                4-digit FBs (e.g. 3350→IB 33) stop at 2-digit so bin_map["3"] never
                hijacks bins whose real IB is 33 (ARRAY MBIST)."""
                _entry = _bin_map.get(n_str, {})
                _cat = _entry.get('cat', '') if _entry else ''
                if not _cat:
                    _fb_s = str(n_str)
                    _prefixes = [2] if len(_fb_s) >= 4 else [2, 1]
                    for _n in _prefixes:
                        if len(_fb_s) >= _n:
                            _ib_e = _bin_map.get(_fb_s[:_n], {})
                            if _ib_e:
                                _cat = _ib_e.get('cat', '')
                                break
                _bd = bindef_dict.get(f'FB{n_str}', '')
                if not _bd or _re2.search(r'B\d{7,}_', _bd):
                    _bd = _bindef_file_dict.get(f'FB{n_str}', '') or _bd
                _dsc = _bd if _bd and not _re2.search(r'B\d{7,}_', _bd) else ''
                return _cat, _dsc

            def _desc(fb_val):
                try:
                    v = int(fb_val)
                    _bd = bindef_dict.get(f'FB{v}', '')
                    if not _bd or _re2.search(r'B\d{7,}_', _bd):
                        _bd = _bindef_file_dict.get(f'FB{v}', '') or _bd
                    if _bd and not _re2.search(r'B\d{7,}_', _bd):
                        return _bd
                except Exception:
                    pass
                return ''

            def _fb_bucket_desc(fb_val):
                """Return (fail_bucket, description) for a functional bin.
                Prefer bin_map category/description, then fallback to legacy bucket/desc lookups.
                When FB has no direct match, derive IB from the first 2 digits of the FB number
                (primary: Yield Summary (filtered) via _ib_to_bucket; fallback: Bin Fail Summary
                lookup using first 2 — then first 1 — digits of the FB as the IB key).
                For 93xx bins, description is sourced from the fB93xx array in the product JSON.
                """
                try:
                    _v = int(fb_val)
                except Exception:
                    return '', ''
                # For 93xx handler/skip bins, use fB93xx descriptions from product JSON
                if 9300 <= _v <= 9399 and _fb93xx_map:
                    _93desc = _fb93xx_map.get(str(_v), '')
                    return _ib_to_bucket.get('93', 'Handler/Skip'), _93desc
                _cat, _dsc = _bin_map_cat(str(_v))
                if not _cat:
                    try:
                        _cat = _bucket(_v)
                    except Exception:
                        _cat = ''
                # If still no category, derive IB from FB digits:
                # "first 2 digits of FB bin = IB bin" (Yield Summary / Bin Fail Summary fallback)
                if not _cat:
                    _fb_str = str(_v)
                    for _n in (2, 1):
                        if len(_fb_str) >= _n:
                            _cat = _ib_to_bucket.get(_fb_str[:_n], '')
                            if _cat:
                                break
                if not _dsc:
                    _dsc = _desc(_v)
                return _cat or '', _dsc or ''

            def _esc(s):
                return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

            # ── shared table HTML fragment ─────────────────────────────────
            CSS = """
<style>
.pareto-wrap{font-family:Arial,sans-serif;padding:12px 0;font-size:12px}
.pareto-wrap h3{font-size:13px;margin-bottom:6px;color:#2c3e50}
.pareto-wrap p.sub{font-size:11px;color:#7f8c8d;margin-bottom:10px}
.pareto-tbl{border-collapse:collapse;white-space:nowrap;font-size:12px}
.pareto-tbl th{background:#2c3e50;color:#ecf0f1;padding:5px 12px;text-align:left;font-weight:bold;white-space:nowrap}
.pareto-tbl td{padding:4px 12px;border-bottom:1px solid #dde}
.pareto-tbl tr:nth-child(even) td{background:#eaf0fb}
.pareto-tbl tr:hover td{background:#d6eaff}
.pareto-tbl .num{text-align:left}
.sort-btn{cursor:pointer;user-select:none;padding-right:8px!important}
.sort-btn:hover{background:#3d5166!important}
.sort-arr{font-size:9px;margin-left:2px;opacity:.7}
.flt-btn{background:none;border:none;color:#aed6f1;cursor:pointer;font-size:11px;padding:0 0 0 4px;vertical-align:middle;opacity:.85}
.flt-btn:hover{opacity:1;color:#fff}
.flt-btn.active{color:#f1c40f!important;opacity:1}
.dd-panel{position:fixed;background:#fff;border:1px solid #aaa;border-radius:4px;box-shadow:0 4px 16px rgba(0,0,0,.18);z-index:9999;min-width:180px;max-width:280px;font-family:Arial,sans-serif;font-size:12px;color:#2c3e50}
.dd-panel .dd-search{width:100%;box-sizing:border-box;padding:5px 8px;border:none;border-bottom:1px solid #ddd;font-size:12px;outline:none}
.dd-panel .dd-acts{display:flex;gap:4px;padding:4px 6px;border-bottom:1px solid #eee}
.dd-panel .dd-acts button{flex:1;padding:2px 6px;font-size:11px;cursor:pointer;border:1px solid #bdc3c7;background:#ecf0f1;border-radius:3px}
.dd-panel .dd-acts button:hover{background:#d5dbde}
.dd-panel .dd-list{max-height:200px;overflow-y:auto;padding:4px 0}
.dd-panel .dd-item{display:flex;align-items:center;gap:6px;padding:3px 10px;cursor:pointer}
.dd-panel .dd-item:hover{background:#eaf0fb}
.dd-panel .dd-item input{margin:0;cursor:pointer}
.dd-panel .dd-footer{padding:4px 8px;border-top:1px solid #eee;text-align:right}
.dd-panel .dd-footer button{padding:3px 12px;font-size:11px;cursor:pointer;background:#2c3e50;color:#fff;border:none;border-radius:3px}
.dd-panel .dd-footer button:hover{background:#3d5166}
/* collapsible panels */
.yp{background:#fff;border:1px solid #ccd;border-radius:5px;margin-bottom:12px;box-shadow:0 1px 4px rgba(0,0,0,.1)}
.yp-bar{display:flex;justify-content:space-between;align-items:center;background:#2c3e50;color:#ecf0f1;padding:6px 12px;border-radius:4px 4px 0 0;cursor:pointer;user-select:none}
.yp-ttl{font-size:12px;font-weight:bold;flex:1}
.yp-btns{display:flex;gap:4px;flex-shrink:0}
.yp-btn{background:none;border:1px solid rgba(255,255,255,.35);color:#ecf0f1;cursor:pointer;font-size:12px;padding:1px 7px;border-radius:3px;line-height:1.4}
.yp-btn:hover{background:rgba(255,255,255,.2)}
.yp-body{padding:12px 14px;overflow:auto;resize:vertical;min-height:30px;box-sizing:border-box}
.yp-body.yp-col{display:none}
.yp.yp-max{position:fixed;top:8px;left:8px;right:8px;bottom:8px;z-index:10000;margin:0;box-shadow:0 8px 32px rgba(0,0,0,.35)}
.yp.yp-max .yp-body{height:calc(100% - 36px);resize:none}
.yp.yp-max .yp-bar{border-radius:4px 4px 0 0}
</style>"""

            import json as _json_par
            # Build per-wafer FP data for interactive reactivity
            _lot_col_fp = (next((c for c in df.columns if c.lower() == 'sort_lot'), None)
                           or next((c for c in df.columns if c.lower() == 'lot'), None)
                           or next((c for c in df.columns if 'lot' in c.lower() and 'slot' not in c.lower()), None))
            _wfr_col_fp = next((c for c in df.columns if 'sort_wafer' in c.lower() or ('wafer' in c.lower() and 'sort_wafer' not in c.lower())), None)
            _fp_wafer_json = '{}'
            if _lot_col_fp and _wfr_col_fp:
                try:
                    _fp_wafer_data = {}
                    _filtered2 = df[df[ib_col] > 4].copy()
                    for (_lot_v2, _wfr_v2), _grp2 in _filtered2.groupby([_lot_col_fp, _wfr_col_fp]):
                        _key2 = str(_lot_v2) + '|' + str(_wfr_v2)
                        _counts2 = _grp2.groupby([fb_col]).size()
                        _fp_wafer_data[_key2] = {str(int(fb2)): int(ct2) for fb2, ct2 in _counts2.items()}
                    import json as _json_fpw2
                    _fp_wafer_json = _json_fpw2.dumps(_fp_wafer_data, ensure_ascii=False)
                except Exception:
                    _fp_wafer_json = '{}'
            # Per-FB unique Bin Description values → "Fail Test Module" column
            _fp_bd_map = {}  # {fb_int: [sorted list of unique non-empty descriptions]}
            if _bd_col:
                try:
                    _filt_bd = filtered[[fb_col, _bd_col]].copy()
                    _filt_bd[fb_col] = _pd.to_numeric(_filt_bd[fb_col], errors='coerce')
                    for _fb_bd, _grp_bd in _filt_bd.dropna(subset=[fb_col]).groupby(fb_col):
                        _uniq_bd = sorted(set(
                            str(v).strip() for v in _grp_bd[_bd_col].dropna()
                            if str(v).strip()
                        ))
                        if _uniq_bd:
                            _fp_bd_map[int(_fb_bd)] = _uniq_bd
                except Exception:
                    pass
            _fp_rows_data = []
            _fp_data_json = '[]'
            _dd_wafer_json = '{}'
            rows_html = ''
            for _ri2, r in enumerate(counts.itertuples(), start=1):
                bkt, desc = _fb_bucket_desc(getattr(r, fb_col))
                _fbv0 = int(getattr(r, fb_col))
                _entry_url = f'heatmap/pareto/pareto_{_ri2:02d}_FB{_fbv0}.html'
                _fp_rows_data.append({'fb': _fbv0, 'bkt': bkt, 'desc': desc,
                    'total': total_all, 'count': int(r.FailCount), 'pct': round(float(r.FailPct), 1), 'url': _entry_url,
                    'mods': _fp_bd_map.get(_fbv0, [])})
                rows_html += (
                    f'<tr class="pareto-row" onclick="paretoNav(\'{ _entry_url}\')" '
                    f'style="cursor:pointer" title="Click to view wafer map">'
                    f'<td>{_fbv0}</td>'
                    f'<td>{_esc(bkt)}</td>'
                    f'<td>{_esc(desc)}</td>'
                    f'<td class="num">{total_all:,}</td>'
                    f'<td class="num">{int(r.FailCount):,}</td>'
                    f'<td class="num">{r.FailPct:.1f}%</td>'
                    f'</tr>\n'
                )
            _fp_data_json = _json_par.dumps(_fp_rows_data, ensure_ascii=False)
            # ── Pass Pareto (IB <= 4) ──────────────────────────────────────────────
            _pp_rows_data = []
            _pp_data_json = '[]'
            _pp_wafer_json = '{}'
            _total_pass = total_all - total_fail
            try:
                _pass_filt = df[df[ib_col] <= 4].copy()
                _total_pass = len(_pass_filt)
                if _total_pass > 0:
                    _pc = (_pass_filt.groupby([fb_col])
                           .size().reset_index(name='PassCount')
                           .sort_values('PassCount', ascending=False)
                           .reset_index(drop=True))
                    _pc['PassPct'] = _pc['PassCount'] / total_all * 100
                    if _lot_col_fp and _wfr_col_fp:
                        _ppwd = {}
                        for (_lpp, _wpp), _gpp in _pass_filt.groupby([_lot_col_fp, _wfr_col_fp]):
                            _kpp = str(_lpp) + '|' + str(_wpp)
                            _ctspp = _gpp.groupby([fb_col]).size()
                            _ppwd[_kpp] = {str(int(_fbb)): int(_ct) for _fbb, _ct in _ctspp.items()}
                        import json as _json_ppw
                        _pp_wafer_json = _json_ppw.dumps(_ppwd, ensure_ascii=False)
                    _pp_bd_map = {}
                    if _bd_col:
                        _ppbd = _pass_filt[[fb_col, _bd_col]].copy()
                        _ppbd[fb_col] = _pd.to_numeric(_ppbd[fb_col], errors='coerce')
                        for _fb_ppb, _g_ppb in _ppbd.dropna(subset=[fb_col]).groupby(fb_col):
                            _u = sorted(set(
                                str(v).strip() for v in _g_ppb[_bd_col].dropna() if str(v).strip()
                            ))
                            if _u:
                                _pp_bd_map[int(_fb_ppb)] = _u
                    for _, _rpp in _pc.iterrows():
                        _fbvpp = int(_rpp[fb_col])
                        _dscpp = (_pass_bin_map.get(str(_fbvpp), {}).get('desc', '')
                                  or _fb_bucket_desc(_fbvpp)[1])
                        _pp_rows_data.append({'fb': _fbvpp, 'desc': _dscpp,
                            'total': total_all, 'count': int(_rpp.PassCount),
                            'pct': round(float(_rpp.PassPct), 1), 'mods': _pp_bd_map.get(_fbvpp, [])})
                    _pp_data_json = _json_par.dumps(_pp_rows_data, ensure_ascii=False)
            except Exception:
                pass
            # ── _out.xlsx / digital_dashboard.html — removed; DD table built from CSV above ──
            if False:  # xlsx-based DD fallback removed
                try:
                    _csvpar_xl = _P(resolved_csv).parent
                    _dd_parent = (_P(dashboard_html).parent
                                  if dashboard_html and os.path.isfile(str(dashboard_html))
                                  else None)
                    # Only use *_out.xlsx (per-run data)
                    for _sdir in filter(None, [_od, _csvpar_xl, _dd_parent,
                                               _csvpar_xl / 'output']):
                        _cands = list(_P(_sdir).glob('*_out.xlsx'))
                        if _cands:
                            _xlsx_p = str(sorted(_cands, key=lambda p: p.stat().st_mtime, reverse=True)[0])
                            break
                    if _xlsx_p:
                        import openpyxl as _opxl, re as _re_xl
                        _xwb = _opxl.load_workbook(_xlsx_p, data_only=False)
                        _xws = _xwb[_xwb.sheetnames[0]]
                        _num_die = None
                        _col_hdrs = []
                        _grid = {}
                        _fmt_grid = {}
                        for _ri, _xrow in enumerate(_xws.iter_rows(values_only=False)):
                            for _ci, _cell in enumerate(_xrow):
                                _grid[(_ri, _ci)] = _cell.value
                                _fmt_grid[(_ri, _ci)] = _cell.number_format or ''
                        for _ri in range(_xws.max_row):
                            if _grid.get((_ri, 0)) == '# Die':
                                _num_die = _grid.get((_ri, 1))
                                break
                        _ev_cache = {}
                        def _ev(ri, ci):
                            if (ri, ci) in _ev_cache:
                                return _ev_cache[(ri, ci)]
                            v = _grid.get((ri, ci))
                            if v is None:
                                _ev_cache[(ri, ci)] = None
                                return None
                            if not isinstance(v, str) or not v.startswith('='):
                                _ev_cache[(ri, ci)] = v
                                return v
                            result = None
                            _m1 = _re_xl.match(r'^=(-?\d+)/B\d+$', v)
                            if _m1 and _num_die:
                                result = int(_m1.group(1)) / _num_die
                            elif _re_xl.match(r'^=SUM\([A-Z]+\d+:[A-Z]+\d+\)$', v):
                                _m2 = _re_xl.match(r'^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', v)
                                if _m2:
                                    _c1 = _opxl0.utils.column_index_from_string(_m2.group(1)) - 1
                                    _r1 = int(_m2.group(2)) - 1
                                    _c2 = _opxl0.utils.column_index_from_string(_m2.group(3)) - 1
                                    _r2 = int(_m2.group(4)) - 1
                                    result = sum(
                                        _ev(r, _c1) for r in range(_r1, _r2 + 1)
                                        if isinstance(_ev(r, _c1), (int, float))
                                    )
                            elif _re_xl.match(r'^=SUM\(([^)]+)\)$', v):
                                _m3 = _re_xl.match(r'^=SUM\(([^)]+)\)$', v)
                                if _m3:
                                    total = 0.0
                                    for _ref in _m3.group(1).split(','):
                                        _mr = _re_xl.match(r'^([A-Z]+)(\d+)$', _ref.strip())
                                        if _mr:
                                            _rc = _opxl.utils.column_index_from_string(_mr.group(1)) - 1
                                            _rr = int(_mr.group(2)) - 1
                                            sv = _ev(_rr, _rc)
                                            if isinstance(sv, (int, float)):
                                                total += sv
                                    result = total
                            _ev_cache[(ri, ci)] = result
                            return result
                        def _fmt_v(ri, ci):
                            v = _ev(ri, ci)
                            if v is None:
                                return ''
                            fmt = _fmt_grid.get((ri, ci), '')
                            if isinstance(v, (int, float)) and '%' in fmt:
                                _ps = f'{v * 100:.1f}%'
                                if ci >= 1 and _num_die and _num_die > 0:
                                    return f'{_ps} ({int(round(v * _num_die)):,})'
                                return _ps
                            if isinstance(v, float):
                                return f'{int(v):,}' if v == int(v) else f'{v:,.2f}'
                            if isinstance(v, int):
                                return f'{v:,}'
                            return str(v) if v is not None else ''
                        _xl_rows = []
                        _col_hdrs = []
                        for _ri in range(_xws.max_row):
                            _label = _grid.get((_ri, 0))
                            if _label == 'Sub Module':
                                _col_hdrs = [str(_grid.get((_ri, _ci)) or '')
                                             for _ci in range(_xws.max_column)]
                                continue
                            if _label == '# Die':
                                continue
                            if all(_grid.get((_ri, _ci)) is None for _ci in range(_xws.max_column)):
                                continue
                            _row_disp = [_fmt_v(_ri, _ci) for _ci in range(len(_col_hdrs) or _xws.max_column)]
                            while _row_disp and _row_disp[-1] == '':
                                _row_disp.pop()
                            if any(v for v in _row_disp):
                                _xl_rows.append(_row_disp)
                        _GROUP_COLORS = ['#eaf4ea','#e3f0fc','#fff6e6','#f3e8fb',
                                         '#fdecea','#e0f7fa','#fdf6e3','#f9ece8']
                        _group_idx = 0
                        _row_groups = []
                        for _xlr in _xl_rows:
                            _l = str(_xlr[0]).upper() if _xlr[0] else ''
                            _is_sum = _l == 'SUM' or _l.startswith('TOTAL')
                            _row_groups.append((_xlr, _is_sum, _group_idx))
                            if _is_sum:
                                _group_idx += 1
                        if _xl_rows and _col_hdrs:
                            _col_hdrs_v = [str(h) for h in _col_hdrs if h is not None]
                            _hdr_html = ''.join(f'<th>{_esc(h)}</th>' for h in _col_hdrs_v)
                            _hdr_html_dyn = ''.join(
                                f'<th>{_esc(h)}</th>'
                                for h in _col_hdrs_v
                            )
                            _dd_js_hdrs = _col_hdrs_v
                            _body_html = ''
                            for _xlr, _is_sum, _gi in _row_groups:
                                _grp_clr = _GROUP_COLORS[_gi % len(_GROUP_COLORS)]
                                _cells_b = [str(v) if v is not None else '' for i, v in enumerate(_xlr) if i < len(_col_hdrs)]
                                if _cells_b and _cells_b[0] == 'Repair Bin 2 (Hard Repair)':
                                    _cells_b[0] = 'Repair Bin 201 (Vnom Repair)'
                                _dd_js_rows.append({'cells': _cells_b, 'bg': _grp_clr, 'bold': bool(_is_sum)})
                                _rw_style = (f' style="font-weight:bold;border-top:2px solid #aaa;'
                                             f'background:{_grp_clr}"'
                                             if _is_sum else
                                             f' style="background:{_grp_clr}"')
                                _body_html += f'<tr{_rw_style}>' + ''.join(
                                    f'<td class="num">{_esc(str(v)) if v is not None else ""}</td>'
                                    if i > 0 else
                                    f'<td>{_esc(str(v)) if v is not None else ""}</td>'
                                    for i, v in enumerate(_xlr)
                                    if i < len(_col_hdrs)
                                ) + '</tr>\n'
                            _tag_str = f' &mdash; {_esc(tag)}' if tag else ''
                            _die_str = (f'<span style="font-size:11px;color:#7f8c8d">&nbsp;|&nbsp;'
                                        f'# Die: <b>{int(_num_die):,}</b></span>') if _num_die else ''
                            _xlsx_tbl_html = (
                                f'<h3 style="font-size:13px;margin:14px 0 6px;color:#2c3e50">'
                                f'&#128196; Digital Dashboard Summary{_tag_str}</h3>'
                                f'{_die_str}'
                                f'<div style="overflow-x:auto;margin-bottom:14px">'
                                f'<table class="pareto-tbl">'
                                f'<thead><tr>{_hdr_html}</tr></thead>'
                                f'<tbody>{_body_html}</tbody></table></div>'
                            )
                            _xlsx_tbl_dyn = (
                                f'<h3 style="font-size:13px;margin:14px 0 6px;color:#2c3e50">'
                                f'&#128196; Digital Dashboard Summary{_tag_str}</h3>'
                                f'{_die_str}'
                                f'<div style="overflow-x:auto;margin-bottom:14px">'
                                f'<table class="pareto-tbl">'
                                f'<thead id="dd-thead"><tr>{_hdr_html_dyn}</tr></thead>'
                                f'<tbody id="dd-tbody"></tbody></table></div>'
                            )
                        # write digital_dashboard.html — always when xlsx found
                        try:
                            _opener_dd = getattr(self, '_opener_port', None)
                            if _opener_dd:
                                import urllib.parse as _updd
                                _xlsx_href_dd = (f'http://127.0.0.1:{_opener_dd}/open?path='
                                                 + _updd.quote(str(_xlsx_p), safe=''))
                            else:
                                _xlsx_href_dd = _P(_xlsx_p).as_uri()
                            _tag_str_dd = f' &#8212; {_esc(tag)}' if tag else ''
                            _tbl_body = (_xlsx_tbl_dyn if _xlsx_tbl_dyn else
                                         _xlsx_tbl_html if _xlsx_tbl_html else
                                         '<p style="color:#7f8c8d;font-size:12px;margin-top:16px">'
                                         'Table preview not available — open in Excel to view.</p>')
                            import json as _json_dd
                            _dd_script = (
                                '<script>\nvar DD_ROWS=' + _json_dd.dumps(_dd_js_rows, ensure_ascii=False) + ';\n'
                                + r'''function esc(s){return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
function _ddFmtCell(val){
  var raw=String(val);var bm=raw.match(/^(.+?)\s*\((\d[\d,]*)\)$/);
  if(!bm)return esc(raw);
  var ns=bm[2].replace(/,/g,'').replace(/\B(?=(\d{3})+(?!\d))/g,',');
  return esc(bm[1])+' <span style="font-size:10px;color:#aaa;font-weight:normal">('+ns+')</span>';
}
function rDDTbl(){var tbody=document.getElementById('dd-tbody');if(!tbody)return;var rows=DD_ROWS.slice();var html='';rows.forEach(function(row){var ts=row.bold?'font-weight:bold;border-top:2px solid #aaa;background:'+row.bg:'background:'+row.bg;html+='<tr style="'+ts+'">';row.cells.forEach(function(c,i){if(i<1)html+='<td>'+esc(c)+'</td>';else html+='<td class="num">'+_ddFmtCell(c)+'</td>';});html+='</tr>';});tbody.innerHTML=html;}
if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',rDDTbl);}else{rDDTbl();}
</script>'''
                            ) if _dd_js_rows else ''
                            _dd_page = (
                                f'<!doctype html>\n<html>\n<head><meta charset="utf-8">'
                                f'<title>Digital Dashboard{_tag_str_dd}</title>\n'
                                f'{CSS}\n'
                                f'<style>\n'
                                f'body{{font-family:Arial,sans-serif;background:#f4f6f8;padding:20px}}\n'
                                f'.dd-open-btn{{display:inline-block;padding:6px 14px;background:#2980b9;'
                                f'color:#fff;border-radius:4px;text-decoration:none;font-size:12px;margin-bottom:16px}}\n'
                                f'.dd-open-btn:hover{{background:#3498db}}\n'
                                f'</style>\n</head>\n<body>\n'
                                f'<a class="dd-open-btn" href="{_xlsx_href_dd}" target="_blank">'
                                f'&#128196; Open {_esc(os.path.basename(_xlsx_p))} in Excel</a>\n'
                                f'{_tbl_body}\n{_dd_script}\n</body>\n</html>'
                            )
                            _dd_html_path = _od / 'digital_dashboard.html'
                            _dd_html_path.write_text(_wm_inject(_dd_page), encoding='utf-8')
                            # Only copy *_out.xlsx into the run folder (not the cumulative dashboard)
                            try:
                                import shutil as _shutil
                                if _P(_xlsx_p).name.endswith('_out.xlsx'):
                                    _xlsx_dest = _od / _P(_xlsx_p).name
                                    if _xlsx_dest.resolve() != _P(_xlsx_p).resolve():
                                        _shutil.copy2(_xlsx_p, str(_xlsx_dest))
                            except Exception:
                                pass
                        except Exception:
                            _dd_html_path = None
                except Exception:
                    pass

            # ── Bin Fail Summary table ────────────────────────────────────
            _bf_data_json = '[]'
            hist_tbl_html = ''
            try:
                import io as _io2
                # aggregate ALL die by IB (no IB > 4 filter)
                ib_all = (df.groupby(ib_col)
                          .size().reset_index(name='FailCount'))
                ib_all['FailPct'] = ib_all['FailCount'] / total_all * 100
                ib_all = ib_all.sort_values(ib_col)
                _labels = [str(int(v)) for v in ib_all[ib_col]]
                _vals_h = ib_all['FailPct'].tolist()
                _fc_h   = [int(v) for v in ib_all['FailCount'].tolist()]

                # ── bin fail summary table — bin_map order (or _yt_buckets order if no bin_map) ──
                # Build a lookup: ib_str -> (fail_count, fail_pct)
                _ib_fc_map = {str(int(v)): (int(fc), float(fp))
                              for v, fc, fp in zip(ib_all[ib_col], _fc_h, _vals_h)}
                # Color palette — one distinct color per category (assigned in first-seen order)
                _CAT_PALETTE = [
                    '#dbeeff',  # soft blue
                    '#e0f5e0',  # soft green
                    '#fef3cd',  # soft yellow
                    '#fde0d0',  # soft orange
                    '#ece0f8',  # soft purple
                    '#d0f4f4',  # soft teal
                    '#fce4ec',  # soft pink
                    '#e8f5e9',  # mint
                    '#fff3e0',  # peach
                    '#e3f2fd',  # pale blue
                    '#f3e5f5',  # lavender
                    '#e8eaf6',  # indigo pale
                ]

                # Build the ordered list of (bin_tok, cat, desc) to render
                _ordered_rows = []   # [(bin_tok_str, cat, desc), ...]
                _yt_only_bins = []  # bins in yield_targets but NOT in bin_map — shown at end if fc>0
                _seen_order = set()
                import re as _re_bf2
                if _bin_map:
                    # Primary: bin_map key insertion order (JSON bin_map order)
                    for _bk, _bme in _bin_map.items():
                        _seen_order.add(_bk)
                        _ordered_rows.append((_bk, _bme.get('cat',''), _bme.get('desc','')))
                # Collect yield_targets bins NOT in bin_map → tail (only if fc>0)
                for _tok_fld2, _blbl2 in (_yt_buckets or []):
                    for _bt2 in _re_bf2.findall(r'\d+', _tok_fld2):
                        if _bt2 not in _seen_order:
                            _seen_order.add(_bt2)
                            _yt_only_bins.append((_bt2, _blbl2))

                # Assign palette colors in row order
                _cat_color_map = {}
                for _, _c, _ in _ordered_rows:
                    _ck = _c.strip().lower()
                    if _ck and _ck not in _cat_color_map:
                        _cat_color_map[_ck] = _CAT_PALETTE[len(_cat_color_map) % len(_CAT_PALETTE)]

                # Main rows
                import re as _re_bf
                _bin_map_keys = set(_bin_map.keys())
                _tbl_rows = ''
                for _bin_tok, _cat2, _dsc2 in _ordered_rows:
                    _fc2, _fp2 = _ib_fc_map.get(_bin_tok, (0, 0.0))
                    _ckey2 = _cat2.strip().lower()
                    _row_bg = _cat_color_map.get(_ckey2, '#ffffff')
                    _tbl_rows += (
                        f'<tr style="background:{_row_bg}">'
                        f'<td>{_bin_tok}</td>'
                        f'<td>{_esc(_cat2)}</td>'
                        f'<td>{_esc(_dsc2)}</td>'
                        f'<td class="num">{total_all:,}</td>'
                        f'<td class="num">{_fc2:,}</td>'
                        f'<td class="num">{_fp2:.2f}%</td>'
                        f'</tr>\n'
                    )

                # Tail rows: bins in yield_targets but not in bin_map — only if fc > 0
                _tail_rows = ''
                for _lbl_t, _bkt_t in _yt_only_bins:
                    _fc_t, _fp_t = _ib_fc_map.get(_lbl_t, (0, 0.0))
                    if _fc_t > 0:
                        _tail_rows += (
                            f'<tr style="background:#fff9e6;font-style:italic" title="In yield_targets but not in bin_map">'
                            f'<td>{_lbl_t}</td>'
                            f'<td></td>'
                            f'<td>{_esc(_bkt_t)}</td>'
                            f'<td class="num">{total_all:,}</td>'
                            f'<td class="num">{_fc_t:,}</td>'
                            f'<td class="num">{_fp_t:.2f}%</td>'
                            f'</tr>\n'
                        )

                # Extra rows: bins with failures NOT in the ordered list
                _extra_rows = ''
                for _lbl_x, _fc_x, _fp_x in sorted(
                        [(_l, _fc, _fp) for _l, (_fc, _fp) in _ib_fc_map.items()
                         if _l not in _seen_order and _fc > 0],
                        key=lambda t: int(t[0])):
                    _bkt_x = _ib_to_bucket.get(_lbl_x, '')
                    _extra_rows += (
                        f'<tr style="background:#ffe0e0;font-style:italic" title="Not in bin_map — add to JSON">'
                        f'<td>{_lbl_x} &#9888;</td>'
                        f'<td></td>'
                        f'<td>{_esc(_bkt_x)}</td>'
                        f'<td class="num">{total_all:,}</td>'
                        f'<td class="num">{_fc_x:,}</td>'
                        f'<td class="num">{_fp_x:.2f}%</td>'
                        f'</tr>\n'
                    )
                if _extra_rows:
                    _extra_rows = (
                        f'<tr style="background:#c0392b;color:#fff"><td colspan="6" '
                        f'style="padding:4px 12px;font-size:11px;font-weight:bold">'
                        f'&#9888; Bins below have failures but are not in bin_map — add to JSON</td></tr>\n'
                    ) + _extra_rows

                import json as _json_bf
                _bf_rows_data = []
                for _btn2, _ct2, _ds2 in _ordered_rows:
                    _fc2b, _fp2b = _ib_fc_map.get(_btn2, (0, 0.0))
                    _ckey2b = _ct2.strip().lower()
                    _row_bgb = _cat_color_map.get(_ckey2b, '#ffffff')
                    _bf_rows_data.append({'bin': _btn2, 'cat': _ct2, 'desc': _ds2,
                        'total': total_all, 'count': _fc2b, 'pct': round(_fp2b, 2), 'bg': _row_bgb})
                for _lbl_t2, _bkt_t2 in _yt_only_bins:
                    _fc_t2, _fp_t2 = _ib_fc_map.get(_lbl_t2, (0, 0.0))
                    if _fc_t2 > 0:
                        _bf_rows_data.append({'bin': _lbl_t2, 'cat': '', 'desc': _bkt_t2,
                            'total': total_all, 'count': _fc_t2, 'pct': round(_fp_t2, 2), 'bg': '#fff9e6'})
                for _lx, _fcx, _fpx in sorted(
                        [(_l, _fc, _fp) for _l, (_fc, _fp) in _ib_fc_map.items()
                         if _l not in _seen_order and _fc > 0],
                        key=lambda t: int(t[0])):
                    _bktx = _ib_to_bucket.get(_lx, '')
                    _bf_rows_data.append({'bin': _lx + '\u26a0', 'cat': '', 'desc': _bktx,
                        'total': total_all, 'count': _fcx, 'pct': round(_fpx, 2), 'bg': '#ffe0e0'})
                _bf_data_json = _json_bf.dumps(_bf_rows_data, ensure_ascii=False)
                hist_tbl_html = (
                    f'<h3 style="font-size:13px;margin:14px 0 6px;color:#2c3e50">'
                    f'&#128202; Bin Fail Summary'
                    + (f' &mdash; {_esc(tag)}' if tag else '') +
                    (' <span style="font-size:10px;font-weight:normal;color:#666">(click row to view IBIN wafer map)</span>' if _wm_url else '') +
                    ' <button onclick="exportTblCsv(\'bfs-thead\',\'bfs-tbody\',\'bin_fail_summary\')" style="font-size:11px;margin-left:8px;padding:2px 7px;cursor:pointer">&#8681; Export CSV</button>'
                    '</h3>'
                    '<table class="pareto-tbl" id="bfs-tbl">'
                    '<thead id="bfs-thead">'
                    '<tr>'
                    '<th class="sort-btn" onclick="bfsClickHdr(0)">Interface Bin <span class="sort-arr"></span><button class="flt-btn" id="bfs-fb-0" onclick="event.stopPropagation();ddOpen(\'bfs\',0,this)" title="Filter">&#9660;</button></th>'
                    '<th class="sort-btn" onclick="bfsClickHdr(1)">Category <span class="sort-arr"></span><button class="flt-btn" id="bfs-fb-1" onclick="event.stopPropagation();ddOpen(\'bfs\',1,this)" title="Filter">&#9660;</button></th>'
                    '<th class="sort-btn" onclick="bfsClickHdr(2)">Description <span class="sort-arr"></span><button class="flt-btn" id="bfs-fb-2" onclick="event.stopPropagation();ddOpen(\'bfs\',2,this)" title="Filter">&#9660;</button></th>'
                    '<th class="sort-btn num" onclick="bfsClickHdr(3)">Total Count <span class="sort-arr"></span></th>'
                    '<th class="sort-btn num" onclick="bfsClickHdr(4)">Count <span class="sort-arr"></span></th>'
                    '<th class="sort-btn num" onclick="bfsClickHdr(5)">Yield/Fail (%) <span class="sort-arr"></span></th>'
                    '</tr>'
                    '</thead>'
                    '<tbody id="bfs-tbody"></tbody></table>'
                )
            except Exception:
                pass

            # xlsx open-link for end of page
            _xlsx_endlink_html = ''
            if _xlsx_p:
                try:
                    _opener_lk = getattr(self, '_opener_port', None)
                    if _opener_lk:
                        import urllib.parse as _up_lk
                        _xlsx_href_lk = (f'http://127.0.0.1:{_opener_lk}/open?path='
                                         + _up_lk.quote(str(_xlsx_p), safe=''))
                    else:
                        _xlsx_href_lk = _P(_xlsx_p).as_uri()
                    _xlsx_endlink_html = (
                        f'<div style="margin-top:24px;border-top:1px solid #dde;padding-top:12px">'
                        f'<a style="display:inline-block;padding:7px 16px;background:#2980b9;'
                        f'color:#fff;border-radius:4px;text-decoration:none;font-size:12px" '
                        f'href="{_xlsx_href_lk}" target="_blank">'
                        f'&#128196; Open {_esc(os.path.basename(_xlsx_p))} in Excel</a></div>'
                    )
                except Exception:
                    pass

            # ── Per-wafer DD module breakdown (for interactive wafer filtering) ──────────
            _dd_wafer_data = {}   # {wafer_key: {total: N, mods: {module: count}}}
            try:
                if bindef_dict and fb_col and _wfr_col_fp and _wfr_col_fp in df.columns:
                    import re as _re_mm
                    # Extract MODULE_MAP from get_dd_update.py source using brace counting
                    _mm_map = {}
                    try:
                        _src_dir_mm = _P(__file__).parent
                        _gdd_src = (_src_dir_mm / 'get_dd_update.py').read_text(encoding='utf-8', errors='ignore')
                        # Find first occurrence of 'moduleMap = {' inside getDD()
                        _mm_start_tok = '    moduleMap = {'
                        _mm_idx = _gdd_src.find(_mm_start_tok)
                        if _mm_idx >= 0:
                            _depth, _mm_end = 0, _mm_idx
                            for _ci, _ch in enumerate(_gdd_src[_mm_idx:]):
                                if _ch == '{': _depth += 1
                                elif _ch == '}':
                                    _depth -= 1
                                    if _depth == 0:
                                        _mm_end = _mm_idx + _ci + 1
                                        break
                            _mm_src = 'MODULE_MAP = ' + _gdd_src[_mm_idx:_mm_end].lstrip().replace('moduleMap = ', '', 1)
                            _exec_ns = {}
                            exec(compile(_mm_src, '<moduleMap>', 'exec'), _exec_ns)
                            _mm_map = _exec_ns.get('MODULE_MAP', {})
                    except Exception:
                        pass

                    # Functional-Bin-Map: authoritative pass-bin lookup from product config JSON.
                    # Maps FB integer → DD row: FF Yield (No Repair / Vmin) → 'Bin 1',
                    # FF Yield (Vmax / Defect) → 'Bin 2', DF Atom → 'Bin 3', DF Core → 'Bin 4'.
                    _fbm = {}
                    try:
                        if bucket_json and os.path.isfile(str(bucket_json)):
                            import json as _json_fbm
                            _fbm_raw = _json_fbm.loads(
                                open(str(bucket_json), encoding='utf-8').read()
                            ).get('Functional-Bin-Map', {})
                            for _fbm_k, _fbm_v in _fbm_raw.items():
                                _fbm_cat = _fbm_v.get('cat', '')
                                _fbm_desc = _fbm_v.get('desc', '')
                                _d = _fbm_desc.lower()
                                if 'FF Yield' in _fbm_cat:
                                    # Primary bin: all FF die → IB 1 (Bin 1) or IB 2 (Bin 2)
                                    if 'vmax' in _d or 'defect' in _d:
                                        _fbm[int(_fbm_k)] = 'Bin 2'
                                    else:  # No Repair, Vmin, or unknown FF
                                        _fbm[int(_fbm_k)] = 'Bin 1'
                                elif 'DF Yield' in _fbm_cat:
                                    _fbm[int(_fbm_k)] = 'Bin 3' if 'Atom' in _fbm_desc else 'Bin 4'
                        # HVQK_REBIN variants: bindef says HVQK but IB is 1 or 2 — force primary bin
                        _fbm[126] = 'Bin 1'  # B126_PASS_HVQK_REBIN → IB1
                        _fbm[226] = 'Bin 2'  # B226_PASS_HVQK_REBIN → IB2
                    except Exception:
                        pass
                    # Repair sub-rows: informational decomposition of Bin1/Bin2 (double-counted)
                    _fbm_repair = {
                        198: 'Repair Bin 198 (Vmin Repair)',   # subset of Bin 1
                        201: 'Repair Bin 201 (Vnom Repair)',  # subset of Bin 2
                        202: 'Repair Bin 202 (Vmax Repair)',   # subset of Bin 2
                        226: 'Repair Bin 201 (Vnom Repair)',  # HVQK_REBIN IB2 → Vnom Repair
                    }

                    # NVL bindef description fallback: ordered to handle more-specific before less-specific
                    _NVL_BINDEF_FALLBACK = [
                        ('ARR_CCF',   'ARR_CCF'),
                        ('ARR_ATOM',  'ARR_ATOM'),
                        ('ARR_CORE',  'ARR_CORE'),
                        ('ARR_',      'ARR_NONCCF'),
                        ('SCN_CCF',   'SCN_CCF'),
                        ('SCN_ATOM',  'SCN_ATOM'),
                        ('SCN_CORE',  'SCN_CORE'),
                        ('SCN_',      'SCN_NONCCF'),   # SCN_UNCORE → NONCCF
                        ('FUN_CCF',   'FUN_CCF'),
                        ('FUN_ATOM',  'FUN_ATOM'),
                        ('FUN_CORE',  'FUN_CORE'),
                        ('FUN_',      'FUN_NONCCF'),
                        ('CLK_',      'FUN_NONCCF'),   # clock domain failures → functional NONCCF
                        ('DRV_RESET', 'RESET (19,35)'),
                        ('PTH_',      'Analog'),
                        ('EIO_',      'Analog'),
                        ('TPI_VCC',   'TPI Foundry'),
                        ('TPI_SIU',   'TPI Foundry'),
                        ('TPI_',      'TPI Foundry'),
                    ]

                    def _fb_to_dd_module(fb_val):
                        """Return dd module name for a functional bin value."""
                        try:
                            _fb_int = int(float(fb_val))
                            # Authoritative pass-bin lookup: Functional-Bin-Map wins over bindef regex
                            if _fb_int in _fbm:
                                return _fbm[_fb_int]
                            _bdesc = bindef_dict.get(f'FB{_fb_int}', '')
                            if not _bdesc:
                                return None
                            # B26x or HVQK_REBIN check (matches HVQK pass/fail rebins)
                            if _re_mm.search(r'B26\d', _bdesc) or 'HVQK_REBIN' in _bdesc:
                                return 'HVQK (B26)'
                            for _rgx, _mod in _mm_map.items():
                                if _re_mm.search(_rgx, _bdesc):
                                    return _mod.get('dd', '')
                            # Fallback: extract module directly from NVL bindef description
                            if '_FAIL_' in _bdesc:
                                for _pat, _dd_name in _NVL_BINDEF_FALLBACK:
                                    if _pat in _bdesc:
                                        return _dd_name
                        except Exception:
                            pass
                        return None

                    # Display name normalisation: map MODULE_MAP dd names → DD table row labels.
                    # FBs covered by Functional-Bin-Map already return 'Bin 1/2/3/4' directly.
                    # These entries handle pass-bin FB variants that fall through to MODULE_MAP
                    # (e.g. B226_PASS → 'Bin 2 (Hard Repair)') and legacy name aliases.
                    _dd_disp = {
                        'Bin 198 (Vmin Repair)': 'Bin 1',   # MODULE_MAP B198_PASS fallback
                        'Bin 1 (No Repair)': 'Bin 1',       # MODULE_MAP B1xx_PASS label
                        'Bin 2 (Hard Repair)': 'Bin 2',         # MODULE_MAP B201/B226_PASS fallback (legacy alias)
                        'FB 201 (Vnom Repair)': 'Repair Bin 201 (Vnom Repair)',  # old name alias
                        'Bin 202 (Vmax Repair)': 'Bin 2',   # MODULE_MAP B202_PASS fallback
                        'Reset': 'RESET (19,35)',
                        'HVQK (26)': 'HVQK (B26)',
                    }
                    _repair_sub_keys = set(_fbm_repair.values())

                    def _build_wafer_entry(_slice_df):
                        _tot = len(_slice_df)
                        _mods = {}
                        for _fb_v, _n in _slice_df[fb_col].value_counts().items():
                            _mod = _fb_to_dd_module(_fb_v)
                            if _mod:
                                _disp = _dd_disp.get(_mod, _mod)
                                _mods[_disp] = _mods.get(_disp, 0) + int(_n)
                            # Repair sub-row: informational double-count (matches XLSX structure)
                            try:
                                _rsub = _fbm_repair.get(int(float(_fb_v)))
                                if _rsub:
                                    _mods[_rsub] = _mods.get(_rsub, 0) + int(_n)
                            except Exception:
                                pass
                        # Catch-all: die with no primary classification → TPI Other
                        # Exclude repair sub-rows (double-counted) from the primary sum
                        _primary_sum = sum(v for k, v in _mods.items() if k not in _repair_sub_keys)
                        _unc = _tot - _primary_sum
                        if _unc > 0:
                            _mods['TPI Other'] = _mods.get('TPI Other', 0) + _unc
                        # Pass count: die with assigned IB = pass (consistent with XLSX Yield SUM)
                        _pass = int(_slice_df[ib_col].notna().sum())
                        return {'total': _tot, 'mods': _mods, 'pass': _pass}

                    # All-wafers entry
                    _dd_wafer_data['all'] = _build_wafer_entry(df)
                    # Per-wafer entries keyed by lot|wafer (same composite key as FP_WAFER_DATA)
                    if _lot_col_fp and _lot_col_fp in df.columns:
                        for (_lv, _wv), _wgrp in df.groupby([_lot_col_fp, _wfr_col_fp]):
                            _dd_wafer_data[str(_lv) + '|' + str(_wv)] = _build_wafer_entry(_wgrp)
                    else:
                        for _wv, _wgrp in df.groupby(_wfr_col_fp):
                            _dd_wafer_data[str(_wv)] = _build_wafer_entry(_wgrp)
            except Exception:
                pass
            _dd_wafer_json = _json_par.dumps(_dd_wafer_data, ensure_ascii=False)

            # ── Build dynamic DD table (sort-enabled) from JS data, outside any try/except ──
            _xlsx_tbl_dyn = ''
            if _dd_js_rows and _dd_js_hdrs:
                _tag_str_dyn = f' &mdash; {_esc(str(tag))}' if tag else ''
                import re as _re_ddhdr
                _hdr_dyn = ''.join(
                    f'<th class="sort-btn" onclick="ddTblClickHdr({_di})">'
                    f'{_esc(_re_ddhdr.sub(r"^\d+W\s+", "", str(_dh)))} <span class="sort-arr"></span></th>'
                    for _di, _dh in enumerate(_dd_js_hdrs)
                )
                _xlsx_tbl_dyn = (
                    f'<h3 style="font-size:13px;margin:14px 0 6px;color:#2c3e50">'
                    f'&#128196; Digital Dashboard Summary{_tag_str_dyn}'
                    f' <button onclick="exportTblCsv(\'dd-thead\',\'dd-tbody\',\'digital_dashboard\')"'
                    f' style="font-size:11px;margin-left:8px;padding:2px 7px;cursor:pointer">&#8681; CSV</button>'
                    f'</h3>'
                    f'<div style="overflow-x:auto;margin-bottom:14px">'
                    f'<table class="pareto-tbl">'
                    f'<thead id="dd-thead"><tr>{_hdr_dyn}</tr></thead>'
                    f'<tbody id="dd-tbody"></tbody></table></div>'
                )

            _par_script = (
                '<script>\n'
                'var BFS_DATA=' + _bf_data_json + ';\n'
                'var FP_DATA=' + _fp_data_json + ';\n'
                'var FP_WAFER_DATA=' + _fp_wafer_json + ';\n'
                + 'var PP_DATA=' + _pp_data_json + ';\n'
                + 'var PP_WAFER_DATA=' + _pp_wafer_json + ';\n'
                'var WM_URL=' + _json_par.dumps(_wm_url) + ';\n'
                'var WM_FILES=' + _json_par.dumps(_wm_files_dict) + ';\n'
                + 'var DD_TAG=' + _json_par.dumps(str(tag) if tag else '') + ';\n'
                + 'var DD_HDRS=' + _json_par.dumps(_dd_js_hdrs) + ';\n'
                + 'var DD_ROWS=' + _json_par.dumps(_dd_js_rows) + ';\n'
                + 'var DD_WAFER_DATA=' + _dd_wafer_json + ';\n'
                + r'''function esc(s){return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
function paretoNav(url){try{var f=window.parent.document.getElementById('frame');if(f){f.src=url;return;}else{window.open(url,'_blank');return;}}catch(e){}try{window.parent.postMessage({navFrame:url},'*');}catch(e2){window.open(url,'_blank');}}
function _ddGetSelWafers(){
  if(!window.IC||!IC.sR||!IC.DATA)return null;
  if(IC.sR.size===IC.DATA.rows.length)return null; // all selected = no filter
  var s=new Set();
  IC.sR.forEach(function(i){var r=IC.DATA.rows[i];if(r)s.add((r.lot||'')+'|'+(r.wafer||''));});
  return s.size?s:null;
}
function rDD(){
  var sec=document.getElementById('dd-section');if(!sec)return;
  sec.style.opacity='1';sec.style.filter='';
  var notice=document.getElementById('dd-lot-notice');if(!notice)return;
  var selWafers=_ddGetSelWafers();
  var _allWfrs=window.IC&&IC.DATA?IC.DATA.rows.length:Object.keys(DD_WAFER_DATA).filter(function(k){return k!=='all';}).length;
  var _selWfrN=selWafers?selWafers.size:_allWfrs;
  var _wfrInfoEl=document.getElementById('dd-wfr-info');
  if(_wfrInfoEl)_wfrInfoEl.textContent=_selWfrN+(_selWfrN!==_allWfrs?' / '+_allWfrs:'')+' Wafer'+(_allWfrs!==1?'s':'');
  if(!selWafers||!DD_TAG){notice.style.display='none';rDDTbl();return;}
  var wArr=Array.from(selWafers).sort();
  var wList=wArr.slice(0,8).join(', ')+(wArr.length>8?'\u2026':'');
  var hasWaferData=Object.keys(DD_WAFER_DATA).length>1;
  if(hasWaferData){
    notice.textContent='\u2139 Showing wafer-level data for: '+wList;
  }else{
    notice.textContent='\u2139 Table data: lot '+DD_TAG+' \u2014 active wafer filter: '+wList;
  }
  notice.style.cssText='display:block;font-size:11px;color:#5d6b7a;background:#eaf4fb;border:1px solid #b8d4e8;border-radius:3px;padding:3px 8px;margin-bottom:6px';
  rDDTbl();
}
function _ddFmtCell(val,cnt){
  /* Render a DD cell value. If cnt is provided (filtered case), append it.
     Otherwise parse "(N)" already embedded in the string from the static case. */
  var raw=String(val);
  var n=(cnt!==undefined&&cnt!==null)?cnt:null;
  var base=raw;
  if(n===null){
    var bm=raw.match(/^(.+?)\s*\((\d[\d,]*)\)$/);
    if(bm){base=bm[1];n=bm[2].replace(/,/g,'');}
  }
  var s=esc(base);
  if(n!==null&&n!==''&&n!=='0'){
    var ns=typeof n==='number'?n.toLocaleString():String(parseInt(n,10)).replace(/\B(?=(\d{3})+(?!\d))/g,',');
    s+=' <span style="font-size:10px;color:#aaa;font-weight:normal">('+ns+')</span>';
  }
  return s;
}
function rDDTbl(){
  var tbody=document.getElementById('dd-tbody');if(!tbody)return;
  var rows=DD_ROWS.slice();
  var selWafers=_ddGetSelWafers();
  // Per-wafer override: aggregate selected wafers from DD_WAFER_DATA
  if(selWafers&&Object.keys(DD_WAFER_DATA).length>1){
    var aggTotal=0,aggMods={};
    selWafers.forEach(function(wk){
      var wd=DD_WAFER_DATA[wk];if(!wd)return;
      aggTotal+=wd.total;
      Object.keys(wd.mods).forEach(function(m){aggMods[m]=(aggMods[m]||0)+wd.mods[m];});
    });
    if(aggTotal>0){
      var groupSum=0,groupRec=0,yieldRec=0,groupSumCnt=0;
      rows=DD_ROWS.map(function(row){
        var nr={cells:row.cells.slice(),bg:row.bg,bold:row.bold};
        var name=nr.cells[0];
        if(name==='SUM'){
          nr.cells[1]=groupSum.toFixed(2)+'%';nr._cnt=groupSumCnt;
          nr.cells[2]=groupRec>0?groupRec.toFixed(2)+'%':'';
          if(groupRec>0)nr._cnt2=Math.round(groupRec*aggTotal/100);
          yieldRec+=groupRec;groupSum=0;groupRec=0;groupSumCnt=0;
        }else if(name==='Yield SUM (%)'){
          var _passTotal=0;
          selWafers.forEach(function(wk){
            var wd=DD_WAFER_DATA[wk];if(!wd)return;
            _passTotal+=(wd.pass!==undefined?wd.pass:wd.total);
          });
          nr.cells[1]=(_passTotal/aggTotal*100).toFixed(2)+'%';nr._cnt=_passTotal;
          nr.cells[2]=yieldRec>0?yieldRec.toFixed(2)+'%':'';
          if(yieldRec>0)nr._cnt2=Math.round(yieldRec*aggTotal/100);
        }else{
          var cnt=aggMods[name]||0,pct=cnt/aggTotal*100;
          nr.cells[1]=pct.toFixed(2)+'%';groupSum+=pct;nr._cnt=cnt;groupSumCnt+=cnt;
          // Scale col3 proportionally: wafer_recovery = lot_recovery * (wafer_fail / lot_fail)
          var lotP2=parseFloat(row.cells[1])||0,lotP3=parseFloat(row.cells[2])||0;
          if(lotP3>0&&lotP2>0){var sc=lotP3*pct/lotP2;nr.cells[2]=sc.toFixed(2)+'%';nr._cnt2=Math.round(sc*aggTotal/100);groupRec+=sc;}
          else{nr.cells[2]='';}
        }
        return nr;
      });
      // Normalize recovery counts: proportional scaling can exceed actual Bin3+4 total
      // on outlier wafers. Cap so total attributed recovery <= actual Bin3+4 die.
      // Only count leaf module rows (not SUM/bold rows, not Yield SUM which double-counts).
      var _recCap=(aggMods['Bin 3']||0)+(aggMods['Bin 4']||0);
      if(_recCap>0){
        var _recSum=0;
        rows.forEach(function(r){if(r._cnt2&&!r.bold&&r.cells[0]!=='Yield SUM (%)')_recSum+=r._cnt2;});
        if(_recSum>_recCap){
          var _sf=_recCap/_recSum;
          // Re-scale leaf module rows
          rows.forEach(function(r){
            if(r._cnt2&&!r.bold&&r.cells[0]!=='Yield SUM (%)'){
              r._cnt2=Math.round(r._cnt2*_sf);
              r.cells[2]=r._cnt2>0?(r._cnt2/aggTotal*100).toFixed(2)+'%':'';
            }
          });
          // Re-accumulate SUM and Yield SUM rows from scaled leaf values
          var _gRec2=0,_yRec2=0;
          rows.forEach(function(r){
            if(r.bold&&r.cells[0]==='SUM'){
              r._cnt2=_gRec2>0?_gRec2:undefined;
              r.cells[2]=_gRec2>0?(_gRec2/aggTotal*100).toFixed(2)+'%':'';
              _yRec2+=_gRec2;_gRec2=0;
            }else if(!r.bold&&r.cells[0]==='Yield SUM (%)'){
              r._cnt2=_yRec2>0?_yRec2:undefined;
              r.cells[2]=_yRec2>0?(_yRec2/aggTotal*100).toFixed(2)+'%':'';
            }else if(!r.bold&&r._cnt2&&r.cells[0]!=='Yield SUM (%)'){_gRec2+=r._cnt2;}
          });
        }
      }
    }
  }
  var html='';
  rows.forEach(function(row){
    var ts=row.bold?'font-weight:bold;border-top:2px solid #aaa;background:'+row.bg:'background:'+row.bg;
    html+='<tr style="'+ts+'">';
    row.cells.forEach(function(c,i){
      if(i===0)html+='<td>'+esc(c)+'</td>';
      else html+='<td class="num">'+_ddFmtCell(c,i===1?row._cnt:row._cnt2)+'</td>';
    });
    html+='</tr>';
  });
  tbody.innerHTML=html;
}
window._updatePareto=function(){rBFS();rFP();rPP();rDD();};
function ypTgl(id){
  var b=document.getElementById('ypb-'+id);
  if(!b)return;
  var col=b.classList.toggle('yp-col');
  var btn=document.getElementById('ypmin-'+id);
  if(btn)btn.innerHTML=col?'&#43;':'&#8722;';
}
function ypMax(id){
  var p=document.getElementById('yp-'+id);
  if(!p)return;
  var on=p.classList.toggle('yp-max');
  var btn=document.getElementById('ypmax-'+id);
  if(btn)btn.innerHTML=on?'&#10066;':'&#10064;';
  var b=document.getElementById('ypb-'+id);
  if(b)b.classList.remove('yp-col');
  if(on){document.body.style.overflow='hidden';}else{document.body.style.overflow='';}
}

/* ── Excel-style dropdown filter ── */
var _ddState={};  // {tbl: [{col, allowed: Set|null}]}
var _ddOpen=null; // {panel, tbl, col, btnEl}
function _ddVals(tbl,col){
  var src=tbl==='bfs'?BFS_DATA:tbl==='pp'?PP_DATA:FP_DATA;
  var keys=tbl==='bfs'?['bin','cat','desc','total','count']:tbl==='pp'?['fb','desc','total','count']:['fb','bkt','desc','total','count'];
  var k=keys[col]; var seen=[];var set=new Set();
  src.forEach(function(r){var v=String(r[k]);if(!set.has(v)){set.add(v);seen.push(v);}});
  return seen.sort(function(a,b){var na=parseFloat(a),nb=parseFloat(b);return(!isNaN(na)&&!isNaN(nb))?(na-nb):a.localeCompare(b);});
}
function _ddAllowed(tbl,col){
  if(!_ddState[tbl])return null;
  var e=_ddState[tbl][col];
  return(e&&e.size>0)?e:null;
}
function ddOpen(tbl,col,btn){
  if(_ddOpen){_ddClose(false);}
  var vals=_ddVals(tbl,col);
  var allowed=_ddAllowed(tbl,col);
  var panel=document.createElement('div');
  panel.className='dd-panel';
  panel.innerHTML='<input class="dd-search" placeholder="Search\u2026" oninput="ddSearch(this)">'
    +'<div class="dd-acts"><button onclick="ddSelAll()">Select All</button><button onclick="ddClearAll()">Clear</button></div>'
    +'<div class="dd-list" id="dd-list"></div>'
    +'<div class="dd-footer"><button onclick="ddApply()">OK</button></div>';
  document.body.appendChild(panel);
  var r=btn.getBoundingClientRect();
  panel.style.top=(r.bottom+2)+'px';
  var left=r.left;
  if(left+parseInt(panel.style.minWidth||180)>window.innerWidth)left=Math.max(0,r.right-200);
  panel.style.left=left+'px';
  _ddOpen={panel:panel,tbl:tbl,col:col,btnEl:btn,vals:vals,checked:allowed?new Set(allowed):new Set(vals)};
  ddRenderList(_ddOpen.vals);
  setTimeout(function(){document.addEventListener('mousedown',_ddOutside);},0);
}
function ddRenderList(vals){
  if(!_ddOpen)return;
  var list=document.getElementById('dd-list');
  var html='';
  vals.forEach(function(v){
    var chk=_ddOpen.checked.has(v)?' checked':'';
    html+='<label class="dd-item"><input type="checkbox"'+chk+' data-val="'+v.replace(/&/g,'&amp;').replace(/"/g,'&quot;')+'">'+esc(v)+'</label>';
  });
  list.innerHTML=html;
  list.querySelectorAll('input').forEach(function(inp){inp.onchange=function(){ddToggle(inp,inp.dataset.val);};});
}
function ddSearch(inp){
  if(!_ddOpen)return;
  var q=(inp.value||'').toLowerCase();
  var filtered=q?_ddOpen.vals.filter(function(v){return v.toLowerCase().indexOf(q)>=0;}):_ddOpen.vals;
  ddRenderList(filtered);
}
function ddToggle(cb,val){if(!_ddOpen)return;if(cb.checked)_ddOpen.checked.add(val);else _ddOpen.checked.delete(val);}
function ddSelAll(){if(!_ddOpen)return;_ddOpen.vals.forEach(function(v){_ddOpen.checked.add(v);});ddRenderList(_ddOpen.vals);}
function ddClearAll(){if(!_ddOpen)return;_ddOpen.checked.clear();ddRenderList(_ddOpen.vals);}
function ddApply(){
  if(!_ddOpen)return;
  var t=_ddOpen.tbl,c=_ddOpen.col,chk=_ddOpen.checked,vals=_ddOpen.vals;
  if(!_ddState[t])_ddState[t]={};
  var allSel=(chk.size===vals.length);
  _ddState[t][c]=allSel?null:new Set(chk);
  var btn=document.getElementById(t+'-fb-'+c);
  if(btn)btn.classList.toggle('active',!allSel);
  _ddClose(false);
  if(t==='bfs')rBFS();else if(t==='pp')rPP();else rFP();
}
function _ddClose(rerender){
  if(!_ddOpen)return;
  document.removeEventListener('mousedown',_ddOutside);
  if(_ddOpen.panel.parentNode)_ddOpen.panel.parentNode.removeChild(_ddOpen.panel);
  _ddOpen=null;
}
function _ddOutside(e){if(_ddOpen&&!_ddOpen.panel.contains(e.target)){_ddApplyOnOutside();}}
function _ddApplyOnOutside(){ddApply();}

/* ── BFS category colours (computed from data so they always differ per category) ── */
var _bfsCatClr=(function(){
  var pal=['#dbeeff','#e0f5e0','#fef3cd','#fde0d0','#ece0f8','#d0f4f4','#fce4ec','#e8f5e9','#fff3e0','#e3f2fd','#f3e5f5','#e8eaf6'];
  var map={};var idx=0;
  BFS_DATA.forEach(function(r){var k=(r.cat||'').trim().toLowerCase();if(k&&!map[k])map[k]=pal[idx++%pal.length];});
  return map;
})();
function _bfsBg(cat){return _bfsCatClr[(cat||'').trim().toLowerCase()]||'#ffffff';}

/* ── BFS ── */
var bfsSort={col:4,dir:'desc'};
function bfsGetFiltered(){
  return BFS_DATA.filter(function(r){
    var vals=[r.bin,r.cat,r.desc,String(r.total),String(r.count)];
    if(!_ddState.bfs)return true;
    return Object.keys(_ddState.bfs).every(function(ci){
      var allowed=_ddState.bfs[ci];if(!allowed)return true;
      return allowed.has(vals[parseInt(ci)]);
    });
  });
}
function bfsFilterLive(r){
  var vals=[r.bin,r.cat,r.desc,String(r.total),String(r.count)];
  if(!_ddState.bfs)return true;
  return Object.keys(_ddState.bfs).every(function(ci){var allowed=_ddState.bfs[ci];if(!allowed)return true;return allowed.has(vals[parseInt(ci)]);});
}
function rBFS(){
  var keys=['bin','cat','desc','total','count','pct'];
  var rows;
  if(window.IC&&IC.gFC){
    var fc=IC.gFC();var cn=fc.counts,tot=fc.total;
    rows=BFS_DATA.map(function(r){
      var binKey=r.bin.replace(/[^\d]/g,'');
      var cnt=cn[binKey]||0;
      var pct=tot>0?cnt/tot*100:0;
      return {bin:r.bin,cat:r.cat,desc:r.desc,total:tot,count:cnt,pct:parseFloat(pct.toFixed(2)),bg:r.bg};
    }).filter(bfsFilterLive);
  }else{
    rows=bfsGetFiltered();
  }
  if(bfsSort.col>=0){var k=keys[bfsSort.col];rows=rows.slice().sort(function(a,b){var av=a[k],bv=b[k];var cmp=(typeof av==='number'&&typeof bv==='number')?(av-bv):String(av).localeCompare(String(bv));return bfsSort.dir==='asc'?cmp:-cmp;});}
  // Hide good bins (1–4) from display
  rows=rows.filter(function(row){var n=+row.bin.replace(/[^\d]/g,'');return !(n>=1&&n<=4);});
  var tbody=document.getElementById('bfs-tbody');if(!tbody)return;
  var sarr=document.querySelectorAll('#bfs-thead .sort-arr');Array.prototype.forEach.call(sarr,function(el,i){el.textContent=i===bfsSort.col?(bfsSort.dir==='asc'?' \u25b2':' \u25bc'):'';});
  var html='';rows.forEach(function(row){var _bg=_bfsBg(row.cat);var _trAttr=WM_URL?'style="background:'+_bg+';cursor:pointer" onclick="paretoNav(\''+WM_URL+'\')" title="Click to view IBIN wafer map"':'style="background:'+_bg+'"';html+='<tr '+_trAttr+'>';html+='<td>'+esc(row.bin)+'</td><td>'+esc(row.cat)+'</td><td>'+esc(row.desc)+'</td>';html+='<td class="num">'+row.total.toLocaleString()+'</td><td class="num">'+row.count.toLocaleString()+'</td><td class="num">'+row.pct.toFixed(2)+'%</td>';html+='</tr>';});
  tbody.innerHTML=html;
}
function bfsClickHdr(col){
  if(bfsSort.col===col){bfsSort.dir=bfsSort.dir==='asc'?'desc':'asc';}else{bfsSort.col=col;bfsSort.dir='asc';}
  var arr=document.querySelectorAll('#bfs-thead .sort-arr');Array.prototype.forEach.call(arr,function(el,i){el.textContent=i===col?(bfsSort.dir==='asc'?' \u25b2':' \u25bc'):'';});rBFS();
}

/* ── PP (Pass Pareto, IB\u22644) ── */
var ppSort={col:0,dir:'asc'};
function ppGetFiltered(){
  return PP_DATA.filter(function(r){
    if((r.count||0)<=0)return false;
    var vals=[String(r.fb),r.desc,String(r.total),String(r.count)];
    if(!_ddState.pp)return true;
    return Object.keys(_ddState.pp).every(function(ci){var allowed=_ddState.pp[ci];if(!allowed)return true;return allowed.has(vals[parseInt(ci)]);});
  });
}
function rPP(){
  var keys=['fb','desc','total','count','pct','mods'];
  var pp_rows;
  if(window.IC&&IC.sR&&IC.DATA&&Object.keys(PP_WAFER_DATA).length>0){
    var totals={};
    IC.sR.forEach(function(i){var row=IC.DATA.rows[i];if(!row)return;var key=(row.lot||'')+'|'+(row.wafer||'');var wd=PP_WAFER_DATA[key];if(wd)Object.keys(wd).forEach(function(fbkey){totals[fbkey]=(totals[fbkey]||0)+wd[fbkey];});});
    var totDie=IC.gFC?IC.gFC().total:0;
    var descMap={},modsMap={};
    PP_DATA.forEach(function(r){descMap[String(r.fb)]=r.desc||'';modsMap[String(r.fb)]=r.mods||[];});
    pp_rows=Object.keys(totals).map(function(fbkey){
      var fb=parseInt(fbkey);var cnt=totals[fbkey];var pct=totDie>0?cnt/totDie*100:0;
      return {fb:fb,desc:descMap[fbkey]||'',total:totDie,count:cnt,pct:parseFloat(pct.toFixed(1)),mods:modsMap[fbkey]||[]};
    });
    pp_rows.sort(function(a,b){return b.count-a.count||a.fb-b.fb;});
    pp_rows=pp_rows.filter(function(r){
      if(r.count<=0)return false;
      var vals=[String(r.fb),r.desc,String(r.total),String(r.count)];
      if(!_ddState.pp)return true;
      return Object.keys(_ddState.pp).every(function(ci){var allowed=_ddState.pp[ci];if(!allowed)return true;return allowed.has(vals[parseInt(ci)]);});
    });
  }else{pp_rows=ppGetFiltered();}
  if(ppSort.col>=0){var k=keys[ppSort.col];pp_rows=pp_rows.slice().sort(function(a,b){var av=a[k],bv=b[k];var cmp=(typeof av==='number'&&typeof bv==='number')?(av-bv):String(av).localeCompare(String(bv));return ppSort.dir==='asc'?cmp:-cmp;});}
  var tbody=document.getElementById('pp-tbody');if(!tbody)return;
  var html='';pp_rows.forEach(function(row){html+='<tr>';html+='<td>'+row.fb+'</td><td>'+esc(row.desc)+'</td>';html+='<td class="num">'+row.total.toLocaleString()+'</td><td class="num">'+row.count.toLocaleString()+'</td><td class="num">'+row.pct.toFixed(1)+'%</td>';var mods=row.mods||[];if(mods.length>0){var d=esc(mods[0].length>45?mods[0].substring(0,43)+'..':mods[0]);if(mods.length>1)d+=' <span style="color:#aaa;font-size:10px">(+'+(mods.length-1)+' more)</span>';html+='<td title="'+mods.join('&#10;').replace(/"/g,'&quot;')+'">'+d+'</td>';}else{html+='<td></td>';}html+='</tr>';});
  tbody.innerHTML=html;
}
function ppClickHdr(col){
  if(ppSort.col===col){ppSort.dir=ppSort.dir==='asc'?'desc':'asc';}else{ppSort.col=col;ppSort.dir='asc';}
  var arr=document.querySelectorAll('#pp-thead .sort-arr');Array.prototype.forEach.call(arr,function(el,i){el.textContent=i===col?(ppSort.dir==='asc'?' \u25b2':' \u25bc'):'';});rPP();
}

/* ── FP ── */
var fpSort={col:5,dir:'desc'};
function fpGetFiltered(){
  return FP_DATA.filter(function(r){
    if((r.count||0)<=0)return false;
    var vals=[String(r.fb),r.bkt,r.desc,String(r.total),String(r.count)];
    if(!_ddState.fp)return true;
    return Object.keys(_ddState.fp).every(function(ci){
      var allowed=_ddState.fp[ci];if(!allowed)return true;
      return allowed.has(vals[parseInt(ci)]);
    });
  });
}
function rFP(){
  var keys=['fb','bkt','desc','total','count','pct','mods'];
  var fp_rows;
  if(window.IC&&IC.sR&&IC.DATA&&Object.keys(FP_WAFER_DATA).length>0){
    var totals={};
    IC.sR.forEach(function(i){
      var row=IC.DATA.rows[i];if(!row)return;
      var key=(row.lot||'')+'|'+(row.wafer||'');
      var wd=FP_WAFER_DATA[key];
      if(wd)Object.keys(wd).forEach(function(fbkey){totals[fbkey]=(totals[fbkey]||0)+wd[fbkey];});
    });
    var totDie=IC.gFC?IC.gFC().total:0;
    var urlMap={},descMap={},modsMap={};
    FP_DATA.forEach(function(r){urlMap[String(r.fb)]=r.url;descMap[String(r.fb)]={bkt:r.bkt,desc:r.desc};modsMap[String(r.fb)]=r.mods||[];});
    fp_rows=Object.keys(totals).map(function(fbkey){
      var fb=parseInt(fbkey);var cnt=totals[fbkey];
      var pct=totDie>0?cnt/totDie*100:0;var info=descMap[fbkey]||{bkt:'',desc:''};
      return {fb:fb,bkt:info.bkt,desc:info.desc,total:totDie,count:cnt,pct:parseFloat(pct.toFixed(1)),url:urlMap[fbkey]||'',mods:modsMap[fbkey]||[]};
    });
    fp_rows.sort(function(a,b){return b.pct-a.pct||b.count-a.count||a.fb-b.fb;});
    fp_rows=fp_rows.filter(function(r){
      if(r.count<=0)return false;
      var vals=[String(r.fb),r.bkt,r.desc,String(r.total),String(r.count)];
      if(!_ddState.fp)return true;
      return Object.keys(_ddState.fp).every(function(ci){var allowed=_ddState.fp[ci];if(!allowed)return true;return allowed.has(vals[parseInt(ci)]);});
    });
  }else{
    fp_rows=fpGetFiltered();
  }
  if(fpSort.col>=0){var k=keys[fpSort.col];fp_rows=fp_rows.slice().sort(function(a,b){var av=a[k],bv=b[k];var cmp=(typeof av==='number'&&typeof bv==='number')?(av-bv):String(av).localeCompare(String(bv));return fpSort.dir==='asc'?cmp:-cmp;});}
  var tbody=document.getElementById('fp-tbody');if(!tbody)return;
  var html='';fp_rows.forEach(function(row){var url=row.url.replace(/'/g,"\\'");html+='<tr class="pareto-row" onclick="paretoNav(\''+url+'\')" style="cursor:pointer" title="Click to view wafer map">';html+='<td>'+row.fb+'</td><td>'+esc(row.bkt)+'</td><td>'+esc(row.desc)+'</td>';html+='<td class="num">'+row.total.toLocaleString()+'</td><td class="num">'+row.count.toLocaleString()+'</td><td class="num">'+row.pct.toFixed(1)+'%</td>';var mods=row.mods||[];if(mods.length>0){var d=esc(mods[0].length>45?mods[0].substring(0,43)+'..':mods[0]);if(mods.length>1)d+=' <span style="color:#aaa;font-size:10px">(+'+(mods.length-1)+' more)</span>';html+='<td title="'+mods.join('&#10;').replace(/"/g,'&quot;')+'">'+d+'</td>';}else{html+='<td></td>';}html+='</tr>';});
  tbody.innerHTML=html;
}
function fpClickHdr(col){
  if(fpSort.col===col){fpSort.dir=fpSort.dir==='asc'?'desc':'asc';}else{fpSort.col=col;fpSort.dir='asc';}
  var arr=document.querySelectorAll('#fp-thead .sort-arr');Array.prototype.forEach.call(arr,function(el,i){el.textContent=i===col?(fpSort.dir==='asc'?' \u25b2':' \u25bc'):'';});rFP();
}

if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',function(){rBFS();rFP();rPP();rDD();rDDTbl();});}else{rBFS();rFP();rPP();rDD();rDDTbl();}
function exportTblCsv(headId,bodyId,fname){
  function cellText(td){return td.textContent.replace(/\s+/g,' ').trim();}
  function q(s){return(s.indexOf(',')>=0||s.indexOf('"')>=0||s.indexOf('\n')>=0)?'"'+s.replace(/"/g,'""')+'"':s;}
  var head=document.getElementById(headId);
  var body=document.getElementById(bodyId);
  if(!head||!body)return;
  var lines=[];
  Array.from(head.querySelectorAll('tr')).forEach(function(tr){
    lines.push(Array.from(tr.querySelectorAll('th,td')).map(function(c){return q(cellText(c));}).join(','));
  });
  Array.from(body.querySelectorAll('tr')).forEach(function(tr){
    lines.push(Array.from(tr.querySelectorAll('th,td')).map(function(c){return q(cellText(c));}).join(','));
  });
  if(!lines.length)return;
  var blob=new Blob([lines.join('\r\n')],{type:'text/csv'});
  var a=document.createElement('a');a.href=URL.createObjectURL(blob);
  a.download=(fname||'export')+'.csv';document.body.appendChild(a);a.click();
  setTimeout(function(){document.body.removeChild(a);URL.revokeObjectURL(a.href);},100);
}
</script>
'''
            )
            _dd_section_html = ''
            _dd_tbl_embed = _xlsx_tbl_dyn if _xlsx_tbl_dyn else _xlsx_tbl_html
            if _dd_tbl_embed:
                _dd_section_html = (
                    '<div id="dd-section" style="margin-top:14px;transition:opacity 0.25s">\n'
                    '<div id="dd-lot-notice"></div>\n'
                    '<div id="dd-wfr-info" style="font-size:11px;color:#5d6b7a;margin-bottom:4px"></div>\n'
                    + _dd_tbl_embed
                    + '</div>\n'
                )

            def _yp(pid, title_html, body_html):
                return (
                    f'<div class="yp" id="yp-{pid}">\n'
                    f'<div class="yp-bar" onclick="ypTgl(\'{pid}\')">'
                    f'<span class="yp-ttl">{title_html}</span>'
                    f'<span class="yp-btns" onclick="event.stopPropagation()">'
                    f'<button class="yp-btn" id="ypmin-{pid}" onclick="ypTgl(\'{pid}\'" title="Collapse / Expand">&#8722;</button>'
                    f'<button class="yp-btn" id="ypmax-{pid}" onclick="ypMax(\'{pid}\'" title="Full screen">&#10064;</button>'
                    f'</span></div>\n'
                    f'<div class="yp-body" id="ypb-{pid}">{body_html}</div>\n'
                    f'</div>\n'
                )

            _pp_body = (
                f'<h3 style="font-size:13px;margin:4px 0 4px;color:#2c3e50">&#9989; Pass Pareto &mdash; Functional Bin (IB &le; 4) <span style="font-size:10px;font-weight:normal;color:#666">(click header to sort)</span> <button onclick="exportTblCsv(\'pp-thead\',\'pp-tbody\',\'pass_pareto_fbin\')" style="font-size:11px;margin-left:8px;padding:2px 7px;cursor:pointer">&#8681; CSV</button></h3>\n'
                + f'<p class="sub">Total: <b>{total_all:,}</b> die &nbsp;|&nbsp; IB&le;4: <b>{_total_pass:,}</b> &nbsp;|&nbsp; ranked highest first</p>\n'
                + '<table class="pareto-tbl">\n'
                + '<thead id="pp-thead"><tr>\n'
                + '  <th class="sort-btn" onclick="ppClickHdr(0)">Functional Bin <span class="sort-arr"></span><button class="flt-btn" id="pp-fb-0" onclick="event.stopPropagation();ddOpen(\'pp\',0,this)" title="Filter">&#9660;</button></th>\n'
                + '  <th class="sort-btn" onclick="ppClickHdr(1)">Description <span class="sort-arr"></span><button class="flt-btn" id="pp-fb-1" onclick="event.stopPropagation();ddOpen(\'pp\',1,this)" title="Filter">&#9660;</button></th>\n'
                + '  <th class="sort-btn num" onclick="ppClickHdr(2)">Total <span class="sort-arr"></span></th>\n'
                + '  <th class="sort-btn num" onclick="ppClickHdr(3)">Count <span class="sort-arr"></span></th>\n'
                + '  <th class="sort-btn num" onclick="ppClickHdr(4)">Pass % <span class="sort-arr"></span></th>\n'
                + '  <th class="sort-btn" onclick="ppClickHdr(5)">Module <span class="sort-arr"></span></th>\n'
                + '</tr></thead>\n'
                + '<tbody id="pp-tbody"></tbody>\n'
                + '</table>\n'
            )
            _fp_body = (
                f'<h3 style="font-size:13px;margin:4px 0 4px;color:#2c3e50">&#128200; Fail Pareto &mdash; Functional Bin (IB &gt; 4) <span style="font-size:10px;font-weight:normal;color:#666">(click row to view wafer map)</span> <button onclick="exportTblCsv(\'fp-thead\',\'fp-tbody\',\'fail_pareto_fbin\')" style="font-size:11px;margin-left:8px;padding:2px 7px;cursor:pointer">&#8681; CSV</button></h3>\n'
                + f'<p class="sub">Total: <b>{total_all:,}</b> die &nbsp;|&nbsp; IB&gt;4: <b>{total_fail:,}</b> &nbsp;|&nbsp; ranked highest first</p>\n'
                + '<table class="pareto-tbl">\n'
                + '<thead id="fp-thead"><tr>\n'
                + '  <th class="sort-btn" onclick="fpClickHdr(0)">Functional Bin <span class="sort-arr"></span><button class="flt-btn" id="fp-fb-0" onclick="event.stopPropagation();ddOpen(\'fp\',0,this)" title="Filter">&#9660;</button></th>\n'
                + '  <th class="sort-btn" onclick="fpClickHdr(1)">Fail Bucket <span class="sort-arr"></span><button class="flt-btn" id="fp-fb-1" onclick="event.stopPropagation();ddOpen(\'fp\',1,this)" title="Filter">&#9660;</button></th>\n'
                + '  <th class="sort-btn" onclick="fpClickHdr(2)">Description <span class="sort-arr"></span><button class="flt-btn" id="fp-fb-2" onclick="event.stopPropagation();ddOpen(\'fp\',2,this)" title="Filter">&#9660;</button></th>\n'
                + '  <th class="sort-btn num" onclick="fpClickHdr(3)">Total <span class="sort-arr"></span></th>\n'
                + '  <th class="sort-btn num" onclick="fpClickHdr(4)">Count <span class="sort-arr"></span></th>\n'
                + '  <th class="sort-btn num" onclick="fpClickHdr(5)">Fail % <span class="sort-arr"></span></th>\n'
                + '  <th class="sort-btn" onclick="fpClickHdr(6)">Fail Test Module <span class="sort-arr"></span></th>\n'
                + '</tr></thead>\n'
                + '<tbody id="fp-tbody"></tbody>\n'
                + '</table>\n'
            )

            table_fragment = (
                CSS + _par_script
                + '<div class="pareto-wrap">\n'
                + '<div style="display:flex;gap:12px;align-items:flex-start">\n'
                + '<div style="flex:1;min-width:0">' + _yp('bfs', f'&#128202; Bin Fail Summary &mdash; {_esc(tag) if tag else ""}', hist_tbl_html) + '</div>\n'
                + '<div style="flex:1;min-width:0">' + _yp('dd',  f'&#128196; Digital Dashboard Summary &mdash; {_esc(tag) if tag else ""}', _dd_section_html) + '</div>\n'
                + '</div>\n'
                + _yp('pp',   f'&#9989; Pass Pareto &mdash; Functional Bin (IB &le; 4)', _pp_body)
                + _yp('fp',   f'&#128200; Fail Pareto &mdash; Functional Bin (IB &gt; 4)', _fp_body)
                + '</div>\n'
            )

            # ── inject into BinDistribution HTML ──────────────────────────
            csv_stem = _P(resolved_csv).stem
            bin_html = _od / f'{csv_stem}_BinDistribution.html'
            if not bin_html.exists():
                _bh_cands = sorted(_od.glob('*_BinDistribution.html'),
                                   key=lambda p: p.stat().st_mtime, reverse=True)
                if _bh_cands:
                    bin_html = _bh_cands[0]
            if bin_html.exists():
                try:
                    import re as _re_inj
                    _INJ_S = '<!-- PARETO_INJECT_START -->'
                    _INJ_E = '<!-- PARETO_INJECT_END -->'
                    _inject_block = f'{_INJ_S}\n{table_fragment}\n{_INJ_E}'
                    txt = bin_html.read_text(encoding='utf-8')
                    if _INJ_S in txt:
                        _ib_ref = _inject_block  # capture for lambda (avoids re escape processing)
                        txt = _re_inj.sub(
                            _re_inj.escape(_INJ_S) + r'.*?' + _re_inj.escape(_INJ_E),
                            lambda _m: _ib_ref, txt, count=1, flags=_re_inj.DOTALL)
                    else:
                        txt = txt.replace('</body>', _inject_block + '\n</body>', 1)
                    bin_html.write_text(_wm_inject(txt), encoding='utf-8')
                except Exception:
                    pass

            # ── per-entry wafer heatmaps (one per top-10 row) ────────────
            def _find_coord_cols(_df):
                _xc = next((c for c in _df.columns if 'sort_x' in c.lower() or c.lower() == 'x' or 'coordx' in c.lower()), None)
                _yc = next((c for c in _df.columns if 'sort_y' in c.lower() or c.lower() == 'y' or 'coordy' in c.lower()), None)
                if not _xc:
                    _xc = next((c for c in _df.columns if c.lower().endswith('_x') or 'posx' in c.lower()), None)
                if not _yc:
                    _yc = next((c for c in _df.columns if c.lower().endswith('_y') or 'posy' in c.lower()), None)
                return _xc, _yc

            def _to_coord_ser(_df, _col):
                if _col not in _df.columns:
                    return None
                _s = _pd.to_numeric(_df[_col], errors='coerce')
                if _s.isnull().all() or (_s.nunique() == 1 and (_s.dropna().empty or _s.dropna().iloc[0] == 0)):
                    _ex = _df[_col].astype(str).str.extract(r'(-?\d+)', expand=False)
                    if _ex.notnull().any():
                        try:
                            _s = _pd.to_numeric(_ex, errors='coerce').fillna(_s)
                        except Exception:
                            pass
                if _s.isnull().all():
                    return None
                if _s.nunique(dropna=True) <= 1 or (_s != 0).sum() / max(1, len(_s)) < 0.05:
                    return None
                return _s.fillna(0).astype(int)

            try:
                _xc, _yc = _find_coord_cols(df)
                _hx = _to_coord_ser(df, _xc) if _xc else None
                _hy = _to_coord_ser(df, _yc) if _yc else None
                if _hx is not None and _hy is not None:
                    # Global wafer geometry for all pareto heatmaps
                    _g_x_all = _hx.values.astype(float)
                    _g_y_all = _hy.values.astype(float)
                    _g_wcx = (_g_x_all.min() + _g_x_all.max()) / 2.0
                    _g_wcy = (_g_y_all.min() + _g_y_all.max()) / 2.0
                    _g_xrange = _g_x_all.max() - _g_x_all.min()
                    _g_yrange = _g_y_all.max() - _g_y_all.min()
                    _g_die_dx = 1.0
                    _g_die_dy = (_g_xrange / _g_yrange) if _g_yrange > 0 else 1.0
                    _g_gap = 0.9
                    # Axis limits: max abs centered coord + half die + 1%
                    _g_xext = (abs(_g_x_all - _g_wcx).max() + 0.5) * 1.025
                    _g_yext = ((abs(_g_y_all - _g_wcy).max() + 0.5) * _g_die_dy) * 1.025
                    _coord_df = _pd.DataFrame({'_hx': _hx, '_hy': _hy})
                    _lot_col = (next((c for c in df.columns if c.lower() == 'sort_lot'), None)
                                or next((c for c in df.columns if c.lower() == 'lot'), None)
                                or next((c for c in df.columns if 'lot' in c.lower() and 'slot' not in c.lower()), None))
                    _wfr_col = next((c for c in df.columns if 'wafer' in c.lower() or 'sort_wafer' in c.lower()), None)
                    _mat_col = next((c for c in df.columns if 'material' in c.lower()), None)

                    # ── detect LayoutX / LayoutY / Reticle for reticle overlay ──
                    _lx_col_p = next((c for c in df.columns if c.lower() in ('layoutx', 'layout_x')), None)
                    _ly_col_p = next((c for c in df.columns if c.lower() in ('layouty', 'layout_y')), None)
                    _ret_col_p = next((c for c in df.columns if c.lower() in ('reticle', 'reticle_number', 'reticlenumber')), None)
                    _has_reticle_p = bool(_lx_col_p and _ly_col_p)
                    _rlookup_p = {}
                    if _has_reticle_p:
                        _rdf_p = _pd.DataFrame({
                            'sx': _hx.values,
                            'sy': _hy.values,
                            'lx': _pd.to_numeric(df[_lx_col_p], errors='coerce').values,
                            'ly': _pd.to_numeric(df[_ly_col_p], errors='coerce').values,
                        }).dropna()
                        if not _rdf_p.empty:
                            for _, _rr_p in _rdf_p.iterrows():
                                _rlookup_p[(int(_rr_p['sx']), int(_rr_p['sy']))] = (_rr_p['lx'], _rr_p['ly'])

                    # Pre-compute reticle number labels once; shared by composite + per-wafer
                    _ret_labels_p = []
                    if _rlookup_p and _ret_col_p:
                        _ret_seen_p0 = set()
                        for _rk_p0, _ in _rlookup_p.items():
                            if _rk_p0 in _ret_seen_p0:
                                continue
                            _ret_seen_p0.add(_rk_p0)
                            _rv_p0 = df.loc[((_hx == _rk_p0[0]) & (_hy == _rk_p0[1])), _ret_col_p]
                            if _rv_p0.empty:
                                continue
                            try:
                                _ret_labels_p.append((
                                    (_rk_p0[0] - _g_wcx) * _g_die_dx,
                                    (_rk_p0[1] - _g_wcy) * _g_die_dy,
                                    str(int(_rv_p0.iloc[0]))
                                ))
                            except (ValueError, TypeError):
                                pass

                    def _draw_reticle_p(ax):
                        """Draw reticle boundaries + numbers on any pareto axis."""
                        if not _rlookup_p:
                            return
                        _xs_up = sorted(set(int(k[0]) for k in _rlookup_p))
                        _ys_up = sorted(set(int(k[1]) for k in _rlookup_p))
                        for _yip in _ys_up:
                            for _xii in range(len(_xs_up) - 1):
                                _k1p = (_xs_up[_xii], _yip)
                                _k2p = (_xs_up[_xii + 1], _yip)
                                _lx1p = _rlookup_p.get(_k1p, (None,))[0]
                                _lx2p = _rlookup_p.get(_k2p, (None,))[0]
                                if _lx1p is not None and _lx2p is not None and _lx1p != _lx2p:
                                    _bxp = ((_xs_up[_xii] + _xs_up[_xii + 1]) / 2 - _g_wcx) * _g_die_dx
                                    _byp = (_yip - _g_wcy) * _g_die_dy
                                    ax.plot([_bxp, _bxp],
                                            [_byp - _g_die_dy * 0.5, _byp + _g_die_dy * 0.5],
                                            color='blue', linewidth=1.5, alpha=0.8, zorder=5)
                        for _xip in _xs_up:
                            for _yii in range(len(_ys_up) - 1):
                                _k1p = (_xip, _ys_up[_yii])
                                _k2p = (_xip, _ys_up[_yii + 1])
                                _ly1p = _rlookup_p.get(_k1p, (None, None))[1]
                                _ly2p = _rlookup_p.get(_k2p, (None, None))[1]
                                if _ly1p is not None and _ly2p is not None and _ly1p != _ly2p:
                                    _bxp2 = (_xip - _g_wcx) * _g_die_dx
                                    _byp2 = ((_ys_up[_yii] + _ys_up[_yii + 1]) / 2 - _g_wcy) * _g_die_dy
                                    ax.plot([_bxp2 - 0.5, _bxp2 + 0.5], [_byp2, _byp2],
                                            color='blue', linewidth=1.5, alpha=0.8, zorder=5)
                        for _rxp, _ryp, _rlp in _ret_labels_p:
                            ax.text(_rxp, _ryp, _rlp,
                                    ha='center', va='center', fontsize=4,
                                    color='blue', fontweight='bold', alpha=0.7, zorder=6)

                    _pareto_dir = _od / 'heatmap' / 'pareto'
                    _pareto_dir.mkdir(parents=True, exist_ok=True)

                    counts = counts.iloc[:0]  # pareto entries removed from sidebar; skip per-bin wafermap rendering
                    for _ri, _rr in counts.iterrows():
                        try:
                            _fbv = int(_rr[fb_col])
                            _fc = int(_rr['FailCount'])
                            _fp = float(_rr['FailPct'])
                            _bkt_e, _de = _fb_bucket_desc(_fbv)
                            _mask = (df[ib_col] > 4) & (df[fb_col] == _fbv)

                            _e_coords, _e_vals = [], []
                            for (_xv, _yv), _grp in _coord_df.groupby(['_hx', '_hy']):
                                _tot = len(_grp)
                                _hits = _mask[_grp.index].sum()
                                _e_coords.append((int(_xv), int(_yv)))
                                _e_vals.append((_hits / _tot) * 100 if _tot else 0.0)
                            if not _e_coords:
                                continue

                            _xs_e = _np.array([c[0] for c in _e_coords])
                            _ys_e = _np.array([c[1] for c in _e_coords])
                            _vs_e = _np.rint(_np.array(_e_vals)).astype(float)

                            # lot/wafer breakdown
                            _g_cols = [c for c in [_lot_col, _wfr_col] if c]
                            if _g_cols:
                                _lw_e = []
                                for _gk, _grp2 in df.groupby(_g_cols):
                                    _t2 = len(_grp2)
                                    _h2 = int(_mask[_grp2.index].sum())
                                    _p2 = int(round(_h2 / _t2 * 100)) if _t2 else 0
                                    if not isinstance(_gk, tuple):
                                        _gk = (_gk,)
                                    _mat2 = str(_grp2[_mat_col].iloc[0]) if _mat_col and not _grp2[_mat_col].dropna().empty else ''
                                    _lw_e.append((str(_gk[0] if len(_gk) >= 1 else ''),
                                                  str(_gk[1] if len(_gk) >= 2 else ''),
                                                  _mat2, f'{_t2:,}', f'{_h2:,}', f'{_p2}%'))
                            else:
                                _h2 = int(_mask.sum())
                                _t2 = len(df)
                                _p2 = int(round(_h2 / _t2 * 100)) if _t2 else 0
                                _lw_e = [('ALL', 'ALL', '', f'{_t2:,}', f'{_h2:,}', f'{_p2}%')]

                            # figure — round wafer map (wafer_map_simple.py style)
                            _fig_e, _ax_e = _plt.subplots(figsize=(7, 7))
                            # Fixed 0-100% scale across all bins for easy comparison
                            _vmax_e = 100
                            _norm_e = _mcolors.Normalize(vmin=0, vmax=_vmax_e)
                            _cmap_e = _plt.cm.RdYlGn_r
                            # build per-position value lookup
                            _pos_val_e = {(_xs_e[_i], _ys_e[_i]): _vs_e[_i] for _i in range(len(_xs_e))}
                            for (_xv, _yv), _pv in _pos_val_e.items():
                                _px_e = (_xv - _g_wcx) * _g_die_dx
                                _py_e = (_yv - _g_wcy) * _g_die_dy
                                _clr_e = _cmap_e(_norm_e(_pv))
                                _rect_e = _mpatches.Rectangle(
                                    (_px_e - _g_die_dx * _g_gap / 2, _py_e - _g_die_dy * _g_gap / 2),
                                    _g_die_dx * _g_gap, _g_die_dy * _g_gap,
                                    linewidth=0.3, edgecolor='gray', facecolor=_clr_e,
                                    rasterized=True
                                )
                                _ax_e.add_patch(_rect_e)
                            # annotate top 5 hotspots
                            try:
                                for _ti in _np.argsort(_vs_e)[-5:][::-1]:
                                    _ax_e.text((_xs_e[_ti] - _g_wcx) * _g_die_dx,
                                               (_ys_e[_ti] - _g_wcy) * _g_die_dy,
                                               f'{int(_vs_e[_ti])}%',
                                               color='black', fontsize=7, ha='center', va='center', fontweight='bold')
                            except Exception:
                                pass
                            _sdesc = _de[:40] if _de else f'FB{_fbv}'
                            _ax_e.set_title(
                                f'FB{_fbv}  \u2014  {_sdesc}\nFail: {_fc:,}  ({_fp:.1f}%)',
                                fontsize=10, weight='bold')
                            _ax_e.set_xlabel('Sort X', fontsize=10)
                            _ax_e.set_ylabel('Sort Y', fontsize=10)
                            _ax_e.set_aspect('equal')
                            _ax_e.set_xlim(-_g_xext, _g_xext)
                            _ax_e.set_ylim(-_g_yext, _g_yext)
                            # remap ticks to original coords
                            _yt_e = [t for t in _ax_e.get_yticks() if -_g_yext <= t <= _g_yext]
                            _ax_e.set_yticks(_yt_e)
                            _ax_e.set_yticklabels([f'{v / _g_die_dy + _g_wcy:.0f}' for v in _yt_e], fontsize=8)
                            _xt_e = [t for t in _ax_e.get_xticks() if -_g_xext <= t <= _g_xext]
                            _ax_e.set_xticks(_xt_e)
                            _ax_e.set_xticklabels([f'{v + _g_wcx:.0f}' for v in _xt_e], fontsize=8)
                            _ax_e.set_xlim(-_g_xext, _g_xext)
                            _ax_e.set_ylim(-_g_yext, _g_yext)
                            _ax_e.axhline(0, color='black', linewidth=0.5, linestyle='--', alpha=0.3)
                            _ax_e.axvline(0, color='black', linewidth=0.5, linestyle='--', alpha=0.3)
                            _ax_e.grid(True, alpha=0.2)
                            # ── reticle overlay on FBIN pareto composite ──
                            _draw_reticle_p(_ax_e)
                            _sm_e = _plt.cm.ScalarMappable(cmap=_cmap_e, norm=_norm_e)
                            _sm_e.set_array([])
                            _cb_e = _fig_e.colorbar(_sm_e, ax=_ax_e, fraction=0.046, pad=0.04)
                            _cb_e.set_label('% fail', fontsize=9)
                            _fig_e.tight_layout()
                            # Save composite as inline SVG (fallback to PNG)
                            _comp_svg_e = None
                            try:
                                from xml.etree import ElementTree as _ET_ce
                                _ET_ce.register_namespace('', 'http://www.w3.org/2000/svg')
                                _ET_ce.register_namespace('xlink', 'http://www.w3.org/1999/xlink')
                                _comp_svgid_e = f'pcomp_FB{_fbv}'
                                _comp_buf_e = _io.StringIO()
                                _fig_e.savefig(_comp_buf_e, format='svg', bbox_inches='tight')
                                _comp_s_e = _re2.sub(r'<\?xml[^>]*\?>', '', _comp_buf_e.getvalue()).strip()
                                _comp_rt_e = _ET_ce.fromstring(_comp_s_e)
                                if _comp_rt_e.get('viewBox') is None:
                                    _crw_e = _comp_rt_e.get('width', '800pt'); _crh_e = _comp_rt_e.get('height', '600pt')
                                    def _cpt_e(s):
                                        try: return float(_re2.sub(r'[^\d.]', '', s))
                                        except: return 0
                                    _comp_rt_e.set('viewBox', f'0 0 {_cpt_e(_crw_e):.1f} {_cpt_e(_crh_e):.1f}')
                                _comp_rt_e.attrib.pop('width', None); _comp_rt_e.attrib.pop('height', None)
                                _comp_rt_e.set('width', '100%'); _comp_rt_e.set('id', _comp_svgid_e)
                                _cidmap_e = {}
                                for _cel_e in _comp_rt_e.iter():
                                    _coid_e = _cel_e.get('id')
                                    if _coid_e and _coid_e != _comp_svgid_e:
                                        _cnid_e = f'{_comp_svgid_e}_{_coid_e}'; _cidmap_e[_coid_e] = _cnid_e; _cel_e.set('id', _cnid_e)
                                if _cidmap_e:
                                    def _cfu_e(v):
                                        return _re2.sub(r'url\(#([^)]+)\)',
                                            lambda m: f'url(#{_cidmap_e.get(m.group(1), m.group(1))})', v)
                                    _cxlh_e = '{http://www.w3.org/1999/xlink}href'
                                    for _cel_e in _comp_rt_e.iter():
                                        for _ca_e, _cav_e in list(_cel_e.attrib.items()):
                                            if 'url(#' in _cav_e: _cel_e.set(_ca_e, _cfu_e(_cav_e))
                                            if _ca_e in (_cxlh_e, 'href') and _cav_e.startswith('#'):
                                                _cref_e = _cav_e[1:]
                                                if _cref_e in _cidmap_e: _cel_e.set(_ca_e, f'#{_cidmap_e[_cref_e]}')
                                _comp_svg_e = _ET_ce.tostring(_comp_rt_e, encoding='unicode')
                            except Exception:
                                _comp_svg_e = None
                            _plt.close(_fig_e)
                            if not _comp_svg_e:
                                _buf_e = _io.BytesIO()
                                # re-render if SVG failed (shouldn't happen)
                                _buf_e.seek(0)
                                _b64_e = ''  # will not be used if _comp_svg_e is set

                            # ── per-wafer SVGs for interactive lot/wafer table ──
                            def _san(s):
                                return _re2.sub(r'[^0-9A-Za-z_-]', '_', str(s))
                            from xml.etree import ElementTree as _ET
                            _lw_svgs_e = []
                            _g_cols2 = [c for c in [_lot_col, _wfr_col] if c]
                            if _g_cols2 and _hx is not None and _hy is not None:
                                for _lw_row_e in _lw_e:
                                    try:
                                        _lv_e, _wv_e = _lw_row_e[0], _lw_row_e[1]
                                        _lw_sel_e = _pd.Series([True] * len(df), index=df.index)
                                        if _lot_col:
                                            _lw_sel_e &= df[_lot_col].fillna('').astype(str).eq(_lv_e)
                                        if _wfr_col:
                                            _lw_sel_e &= df[_wfr_col].fillna('').astype(str).eq(_wv_e)
                                        _lw_idx_e = _lw_sel_e[_lw_sel_e].index
                                        if len(_lw_idx_e) == 0:
                                            _lw_svgs_e.append((_lv_e, _wv_e, None, None)); continue
                                        _lx_e2 = _hx[_lw_idx_e].values.astype(float)
                                        _ly_e2 = _hy[_lw_idx_e].values.astype(float)
                                        _lmv_e = _mask[_lw_idx_e].values
                                        _cd_e = {}
                                        for _ci_e, (_cx_e, _cy_e) in enumerate(zip(_lx_e2.astype(int), _ly_e2.astype(int))):
                                            _cd_e.setdefault((_cx_e, _cy_e), [0, 0])
                                            _cd_e[(_cx_e, _cy_e)][0] += 1
                                            if _lmv_e[_ci_e]:
                                                _cd_e[(_cx_e, _cy_e)][1] += 1
                                        _lw_xs_e = _np.array([k[0] for k in _cd_e], dtype=float)
                                        _lw_ys_e = _np.array([k[1] for k in _cd_e], dtype=float)
                                        _lw_vs_e = _np.array([v[1] / v[0] * 100 for v in _cd_e.values()])
                                        _lw_fig_e, _lw_ax_e = _plt.subplots(figsize=(7, 7))
                                        _lw_xp_e = (_lw_xs_e - _g_wcx) * _g_die_dx
                                        _lw_yp_e = (_lw_ys_e - _g_wcy) * _g_die_dy
                                        for _di_e in range(len(_lw_xp_e)):
                                            _clr_e2 = _cmap_e(_norm_e(float(_np.rint(_lw_vs_e[_di_e]))))
                                            _lw_ax_e.add_patch(_mpatches.Rectangle(
                                                (_lw_xp_e[_di_e] - _g_die_dx * _g_gap / 2,
                                                 _lw_yp_e[_di_e] - _g_die_dy * _g_gap / 2),
                                                _g_die_dx * _g_gap, _g_die_dy * _g_gap,
                                                linewidth=0.3, edgecolor='gray', facecolor=_clr_e2,
                                                rasterized=True))
                                        _lw_ax_e.set_title(f'Lot {_lv_e}  Wafer {_wv_e}\nFB{_fbv}', fontsize=9)
                                        _lw_ax_e.set_aspect('equal')
                                        _lw_ax_e.set_xlim(-_g_xext, _g_xext); _lw_ax_e.set_ylim(-_g_yext, _g_yext)
                                        _lw_ax_e.set_xlabel('Sort X', fontsize=8); _lw_ax_e.set_ylabel('Sort Y', fontsize=8)
                                        _yt_lw = [t for t in _lw_ax_e.get_yticks() if -_g_yext <= t <= _g_yext]
                                        _lw_ax_e.set_yticks(_yt_lw)
                                        _lw_ax_e.set_yticklabels([f'{v / _g_die_dy + _g_wcy:.0f}' for v in _yt_lw], fontsize=7)
                                        _xt_lw = [t for t in _lw_ax_e.get_xticks() if -_g_xext <= t <= _g_xext]
                                        _lw_ax_e.set_xticks(_xt_lw)
                                        _lw_ax_e.set_xticklabels([f'{v + _g_wcx:.0f}' for v in _xt_lw], fontsize=7)
                                        _lw_ax_e.axhline(0, color='black', linewidth=0.5, linestyle='--', alpha=0.3)
                                        _lw_ax_e.axvline(0, color='black', linewidth=0.5, linestyle='--', alpha=0.3)
                                        _lw_ax_e.grid(True, alpha=0.2)
                                        _draw_reticle_p(_lw_ax_e)
                                        try:
                                            _sm_lw = _plt.cm.ScalarMappable(cmap=_cmap_e, norm=_norm_e)
                                            _sm_lw.set_array([])
                                            _lw_fig_e.colorbar(_sm_lw, ax=_lw_ax_e, fraction=0.046, pad=0.04, label='% fail')
                                        except Exception:
                                            pass
                                        _lw_fig_e.tight_layout()
                                        _svgbuf_e = _io.StringIO()
                                        _lw_fig_e.savefig(_svgbuf_e, format='svg', bbox_inches='tight')
                                        _plt.close(_lw_fig_e)
                                        _svgs_e = _re2.sub(r'<\?xml[^>]*\?>', '', _svgbuf_e.getvalue()).strip()
                                        _svgid_e = f'psvg_{_san(_lv_e)}_{_san(_wv_e)}_FB{_fbv}'
                                        try:
                                            _ET.register_namespace('', 'http://www.w3.org/2000/svg')
                                            _ET.register_namespace('xlink', 'http://www.w3.org/1999/xlink')
                                            _rt_e = _ET.fromstring(_svgs_e)
                                            if _rt_e.get('viewBox') is None:
                                                _rw_e = _rt_e.get('width', '800pt'); _rh_e = _rt_e.get('height', '600pt')
                                                def _pt_e(s):
                                                    try: return float(_re2.sub(r'[^\d.]', '', s))
                                                    except: return 0
                                                _rt_e.set('viewBox', f'0 0 {_pt_e(_rw_e):.1f} {_pt_e(_rh_e):.1f}')
                                            _rt_e.attrib.pop('width', None); _rt_e.attrib.pop('height', None)
                                            _rt_e.set('width', '100%'); _rt_e.set('id', _svgid_e)
                                            _idmap_e = {}
                                            for _el_e in _rt_e.iter():
                                                _oid_e = _el_e.get('id')
                                                if _oid_e and _oid_e != _svgid_e:
                                                    _nid_e = f'{_svgid_e}_{_oid_e}'; _idmap_e[_oid_e] = _nid_e; _el_e.set('id', _nid_e)
                                            if _idmap_e:
                                                def _fu_e(v):
                                                    return _re2.sub(r'url\(#([^)]+)\)',
                                                        lambda m: f'url(#{_idmap_e.get(m.group(1), m.group(1))})', v)
                                                _xlh_e = '{http://www.w3.org/1999/xlink}href'
                                                for _el_e in _rt_e.iter():
                                                    for _a_e, _av_e in list(_el_e.attrib.items()):
                                                        if 'url(#' in _av_e: _el_e.set(_a_e, _fu_e(_av_e))
                                                        if _a_e in (_xlh_e, 'href') and _av_e.startswith('#'):
                                                            _ref_e = _av_e[1:]
                                                            if _ref_e in _idmap_e: _el_e.set(_a_e, f'#{_idmap_e[_ref_e]}')
                                            _svgs_e = _ET.tostring(_rt_e, encoding='unicode')
                                        except Exception:
                                            pass
                                        _lw_svgs_e.append((_lv_e, _wv_e, _svgs_e, _svgid_e))
                                    except Exception as _lw_exc_e:
                                        _lw_svgs_e.append((_lv_e, _wv_e, None, None))

                            # build SVG blocks + row onclick attrs
                            _irows_e = [
                                ('Rank', str(_ri + 1)),
                                ('Functional Bin', str(_fbv)),
                                ('Description', _de),
                                ('Count', f'{_fc:,}'),
                                ('Fail %', f'{_fp:.1f}%'),
                            ]
                            _ihtml = ''.join(
                                f'<tr><td>{_esc(k)}</td><td>{_esc(v)}</td></tr>\n'
                                for k, v in _irows_e)
                            _svg_blocks_e = ''
                            _svg_id_map_e = {}
                            for _lv2_e, _wv2_e, _svgstr_e, _sid_e in _lw_svgs_e:
                                if _svgstr_e and _sid_e:
                                    _svg_id_map_e[(_lv2_e, _wv2_e)] = _sid_e
                                    _svg_blocks_e += (
                                        f'<div id="wrap_{_sid_e}" class="hm-wafer-view" '
                                        f'style="display:none;width:570px">{_svgstr_e}</div>\n'
                                    )
                            _has_svgs_e = bool(_svg_blocks_e)
                            _lwhtml = ''
                            for rr in _lw_e:
                                _sid2_e = _svg_id_map_e.get((rr[0], rr[1]))
                                _oc_e = f' onclick="hmShowWafer(\'{_sid2_e}\',this)"' if _sid2_e else ''
                                _cs_e = ' cursor:pointer;' if _sid2_e else ''
                                _lwhtml += (
                                    f'<tr class="lw-row"{_oc_e} style="{_cs_e}">'
                                    f'<td>{_esc(rr[0])}</td><td>{_esc(rr[1])}</td>'
                                    + (f'<td>{_esc(rr[2])}</td>' if _mat_col else '')
                                    + f'<td class="num">{_esc(rr[3] if _mat_col else rr[2])}</td>'
                                    f'<td class="num">{_esc(rr[4] if _mat_col else rr[3])}</td>'
                                    f'<td class="num">{_esc(rr[5] if _mat_col else rr[4])}</td></tr>\n'
                                )
                            _tot_units = sum(int(rr[3 if _mat_col else 2].replace(',','')) for rr in _lw_e)
                            _tot_fails = sum(int(rr[4 if _mat_col else 3].replace(',','')) for rr in _lw_e)
                            _tot_pct = int(round(_tot_fails / _tot_units * 100)) if _tot_units else 0
                            _lwhtml += (
                                f'<tr style="font-weight:bold;background:#dde8f7">'
                                + (f'<td colspan="3">TOTAL</td>' if _mat_col else '<td colspan="2">TOTAL</td>')
                                + f'<td class="num">{_tot_units:,}</td>'
                                f'<td class="num">{_tot_fails:,}</td>'
                                f'<td class="num">{_tot_pct}%</td></tr>\n'
                            )
                            _js_e = ''
                            if _has_svgs_e:
                                _js_e = (
                                    '<script>\n'
                                    'function hmShowWafer(svgId,row){\n'
                                    '  var allViews=document.querySelectorAll(".hm-wafer-view");\n'
                                    '  var allRows=document.querySelectorAll(".lw-row");\n'
                                    '  var composite=document.getElementById("hm-composite");\n'
                                    '  var alreadyActive=row.classList.contains("lw-active");\n'
                                    '  allViews.forEach(function(d){d.style.display="none";});\n'
                                    '  allRows.forEach(function(r){r.classList.remove("lw-active");r.style.background="";});\n'
                                    '  if(alreadyActive){composite.style.display="block";}\n'
                                    '  else{composite.style.display="none";\n'
                                    '    var el=document.getElementById("wrap_"+svgId);\n'
                                    '    if(el)el.style.display="block";\n'
                                    '    row.classList.add("lw-active");row.style.background="#ddeeff";}\n'
                                    '}\n</script>\n'
                                    '<style>.lw-row:hover td{background:#f0f4ff !important;cursor:pointer;}'
                                    '.lw-row.lw-active td{background:#ddeeff !important;}</style>\n'
                                )
                            _hint_e = ' <span style="font-size:10px;font-weight:normal;color:#666">(click row to show wafer map)</span>' if _has_svgs_e else ''
                            _comp_inner_e = (_comp_svg_e if _comp_svg_e
                                             else f'<img src="data:image/png;base64,{_b64_e}" style="max-width:100%;height:auto"/>')
                            _entry_html_str = (
                                '<!doctype html>\n<html><head><meta charset="utf-8">\n'
                                '<meta name="viewport" content="width=device-width,initial-scale=1">\n'
                                '<style>\n'
                                'html,body{margin:0;padding:8px;background:#fff;font-family:Arial,sans-serif;font-size:12px}\n'
                                'img{max-width:100%;height:auto;display:block;margin-bottom:14px}\n'
                                '.info-table,.lw-table{border-collapse:collapse;font-size:11px;margin-bottom:14px}\n'
                                '.info-table td,.lw-table th,.lw-table td{padding:3px 10px;border:1px solid #ccc;'
                                'white-space:nowrap;text-align:left}\n'
                                '.info-table td:first-child,.lw-table th{font-weight:bold;background:#f0f0f0}\n'
                                '.lw-table tr:hover td{background:#f5f5f5}\n'
                                '.num{text-align:right}\n'
                                '#hm-composite{max-width:700px}\n'
                                f'</style>{_js_e}</head><body>\n'
                                f'<div id="hm-composite">{_comp_inner_e}</div>\n'
                                f'{_svg_blocks_e}'
                                '<table class="info-table"><tbody>' + _ihtml + '</tbody></table>\n'
                                f'<h3 style="font-size:12px;margin:8px 0 4px"><svg width="11" height="11" viewBox="0 0 24 24" fill="currentColor" style="vertical-align:middle;margin-right:3px"><path d="M10 18h4v-2h-4v2zM3 6v2h18V6H3zm3 7h12v-2H6v2z"/></svg> Lot / Wafer breakdown{_hint_e}</h3>\n'
                                '<table class="lw-table">\n'
                                + ('<thead><tr><th>LOT</th><th>WAFER</th><th>MATERIAL TYPE</th><th>Total Count</th><th>Count</th><th>% FAIL</th></tr></thead>\n' if _mat_col else '<thead><tr><th>LOT</th><th>WAFER</th><th>Total Count</th><th>Count</th><th>% FAIL</th></tr></thead>\n')
                                + '<tbody>' + _lwhtml + '</tbody></table>\n'
                                '</body></html>'
                            )
                            _fname = f'pareto_{_ri + 1:02d}_FB{_fbv}.html'
                            _ep = _pareto_dir / _fname
                            _ep.write_text(_wm_inject(_entry_html_str), encoding='utf-8')
                            _short_lbl = _de[:25] if _de else f'FB{_fbv}'
                            _pareto_entries.append((_ri + 1, f'FB{_fbv} {_short_lbl}', _ep))
                        except Exception:
                            pass
            except Exception:
                pass

            return None, _pareto_entries, _dd_html_path, _xlsx_p, None, []
        except Exception:
            return None, [], None, None, None, []

    def _build_master_html(self, resolved_csv, dashboard_path=None, plot_html=None, plot_tag_files=None, sicc_links=None, output_dir=None, opener_port=None, bindef_csv=None, bucket_json=None, tag=None):
        """Build output/index.html with a sidebar linking BinDistribution, heatmap HTMLs, plots, and Excel files.
        sicc_links: list of (abs_path, label, css_class) from SICC/UPM headless run.
        output_dir: explicit output folder; defaults to <csv_parent>/output."""
        try:
            from pathlib import Path as _P
            import os as _os2
            csv_p = _P(resolved_csv)
            if output_dir and _os2.path.isdir(output_dir):
                out_dir = _P(output_dir)
            else:
                out_dir = csv_p.parent / 'output'
            out_dir.mkdir(parents=True, exist_ok=True)
            stem = csv_p.stem
            bin_html = out_dir / f'{stem}_BinDistribution.html'
            # Fallback 1: any *_BinDistribution.html already written in this folder
            # (e.g. from the pre-reticle CSV whose stem differs from the merged CSV)
            if not bin_html.exists():
                _bin_cands = sorted(
                    out_dir.glob('*_BinDistribution.html'),
                    key=lambda p: p.stat().st_mtime, reverse=True)
                if _bin_cands:
                    bin_html = _bin_cands[0]
            heat_dir = out_dir / 'heatmap'
            heat_htmls = sorted(
                f for f in (heat_dir.glob('*.html') if heat_dir.exists() else [])
                if '_IBIN_WaferMap_' not in f.name
                and '_IBIN_ALL_WaferMap' not in f.name
            )
            # Detect per-lot IBIN wafer maps (generated by generate_all_ibin_wafer_map)
            _wm_files = sorted(heat_dir.glob(f'{stem}_IBIN_WaferMap_*.html')) if heat_dir.exists() else []
            # plots HTML
            plots_html_path = _P(plot_html) if plot_html and os.path.isfile(str(plot_html)) else None

            # Resolve SICC links: use what was passed in, then scan multiple candidate dirs
            _sicc_resolved = list(sicc_links) if sicc_links else []
            if not _sicc_resolved:
                _sicc_scan_dirs = [
                    out_dir,
                    csv_p.parent,
                    csv_p.parent.parent / 'output',
                    csv_p.parent.parent,
                ]
                if output_dir:
                    _sicc_scan_dirs.insert(0, _P(output_dir))
                for _scan_dir in _sicc_scan_dirs:
                    _all = list(_scan_dir.glob('*.html')) if _scan_dir.exists() else []
                    _sicc_cands = [p for p in _all
                                   if 'sicc_upm_dashboard' in p.name.lower()]
                    if _sicc_cands:
                        _sicc_cands.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                        _sicc_resolved.append((str(_sicc_cands[0]), 'Dashboard SICC_CDYN_UPM', 'sicc-link'))
                        break

            if not bin_html.exists() and not heat_htmls and not _wm_files:
                return None

            # Generate pareto HTML
            pareto_html_path = None
            pareto_entries = []
            _dd_html_path = None   # safe default — set inside try block below
            _dd_xlsx_path = None
            ibin_pareto_html_path = None
            ibin_pareto_entries = []
            try:
                _pareto_result = self._build_pareto_html(resolved_csv, bindef_csv, out_dir, tag=tag, dashboard_html=dashboard_path, bucket_json=bucket_json)
                if isinstance(_pareto_result, tuple) and len(_pareto_result) >= 6:
                    _pareto, pareto_entries, _dd_html_path, _dd_xlsx_path, _ibin_pareto, ibin_pareto_entries = _pareto_result
                    if _ibin_pareto and os.path.isfile(_ibin_pareto):
                        ibin_pareto_html_path = _P(_ibin_pareto)
                    if not isinstance(ibin_pareto_entries, list):
                        ibin_pareto_entries = []
                elif isinstance(_pareto_result, tuple) and len(_pareto_result) >= 4:
                    _pareto, pareto_entries, _dd_html_path, _dd_xlsx_path = _pareto_result[:4]
                elif isinstance(_pareto_result, tuple) and len(_pareto_result) == 2:
                    _pareto, pareto_entries = _pareto_result
                    _dd_html_path, _dd_xlsx_path = None, None
                else:
                    _pareto, pareto_entries, _dd_html_path, _dd_xlsx_path = _pareto_result, [], None, None
                if not isinstance(pareto_entries, list):
                    pareto_entries = []
                if _pareto and os.path.isfile(_pareto):
                    pareto_html_path = _P(_pareto)
            except Exception:
                pass

            def _nav(label, href):
                return (f'<a class="nav-link" href="#" '
                        f'onclick="load(\'{href}\',this);return false;">'
                        f'{label}</a>\n')

            def _sub_nav(label, href):
                return (f'<a class="nav-link sub-link" href="#" '
                        f'onclick="load(\'{href}\',this);return false;">'
                        f'{label}</a>\n')

            def _subsub_nav(label, href):
                return (f'<a class="nav-link subsub-link" href="#" '
                        f'onclick="load(\'{href}\',this);return false;">'
                        f'{label}</a>\n')

            # ── Yield section (was "Bin Distribution") ─────────────────────────
            yield_section = _nav('&#128229; Yield Dashboard', bin_html.name) if bin_html.exists() else ''


            heat_section = ''
            if heat_htmls:
                import re as _re3, json as _json3

                _BUCKET_ORDER = [
                    '1/2/3/4','1/2','3/4','1','2','3','4',
                    '41/42/47/76/77/81/82',
                    '20/21/33/60/61/62/63/65',
                    '11/13/16/25/27/28/32/36/39/46/48/51/64/71/74/75',
                    '7/8/9/10/15/18/43',
                    '31/88/91/94/97/98/99 + 93',
                    '19/35','12/44/45/70/80/85/86','26',
                ]
                if bucket_json and os.path.isfile(str(bucket_json)):
                    try:
                        _bj = _json3.loads(open(bucket_json, encoding='utf-8').read())
                        if isinstance(_bj, dict):
                            # new format: {"yield_targets": [...], "bin_map": {...}, ...}
                            _yt_list = _bj.get('yield_targets',
                                               _bj.get('value', _bj.get('Value', [])))
                        else:
                            _yt_list = _bj
                        _BUCKET_ORDER = [str(e.get('bin', '')) for e in _yt_list if e.get('bin')]
                    except Exception:
                        pass

                def _digkey(s):
                    return '_'.join(_re3.findall(r'\d+', str(s)))

                _bin_rank = {}
                for _bi, _bf in enumerate(_BUCKET_ORDER):
                    _k = _digkey(_bf)
                    if _k not in _bin_rank:
                        _bin_rank[_k] = _bi

                def _heat_sort_key(h):
                    _label_part = h.stem.replace(f'{stem}_Heatmap_bin_', '')
                    _k = _digkey(_label_part)
                    return _bin_rank.get(_k, 9999)

                for h in sorted(heat_htmls, key=_heat_sort_key):
                    label = h.stem.replace(f'{stem}_Heatmap_bin_', 'Bin ').replace('_', '/')
                    heat_section += _sub_nav(label, f'heatmap/{h.name}')
            if not heat_section:
                heat_section = '<span class="nav-link sub-link" style="opacity:0.45;cursor:default">Bin Heatmaps (skipped)</span>\n'

            # Custom Plots: non-UPM tags only
            plots_section = ''
            if plot_tag_files:
                for _ptag, _p in plot_tag_files.items():
                    _pp = _P(_p)
                    if _pp.exists() and 'UPM' not in _ptag.upper():
                        plots_section += _nav(_ptag, _pp.name)
            elif plots_html_path and plots_html_path.exists():
                plots_section = _nav('&#128202; Plots', plots_html_path.name)

            first_src = (bin_html.name if bin_html.exists()
                         else (f'heatmap/{heat_htmls[0].name}' if heat_htmls else ''))

            # SICC/UPM sidebar — plot.html + SICC/CDYN Report HTML + UPM-tagged plots
            import re as _re4
            sicc_section = ''
            _plot_html_candidate = out_dir / 'plot.html'
            if _plot_html_candidate.exists():
                sicc_section += _nav('&#128202; SICC / CDYN / UPM Dashboard', 'plot.html')
            # Python SICC/CDYN Report (most recent *_sicc_analysis.html in run output folder)
            _sa_cands = sorted(out_dir.glob('*_sicc_analysis.html'),
                               key=lambda p: p.stat().st_mtime, reverse=True)
            if _sa_cands:
                sicc_section += _nav('&#128202; UPM/SICC/CDYN Dashboard', _sa_cands[0].name)
            if plot_tag_files:
                for _ptag, _p in plot_tag_files.items():
                    _pp = _P(_p)
                    if _pp.exists() and 'UPM' in _ptag.upper():
                        # Normalise display: uppercase, _DIST → _DISTRIBUTION
                        _display = _re4.sub(r'(?i)_dist$', '_DISTRIBUTION', _ptag).upper()
                        sicc_section += _nav(f'&#128202; {_display}', _pp.name)

            plots_sidebar = (
                f'  <div class="sec">Custom Plots</div>\n  {plots_section}'
                if plots_section else ''
            )
            sicc_sidebar = (
                f'  <div class="sec">SICC / CDYN / UPM</div>\n  {sicc_section}'
                if sicc_section else ''
            )

            # ── UPM Distribution analysis from config JSON ─────────────────────
            _upm_dist_section = ''
            try:
                if bucket_json and os.path.isfile(str(bucket_json)):
                    import json as _json_upm
                    import pandas as _pd_upm
                    import matplotlib
                    matplotlib.use('Agg')
                    import matplotlib.pyplot as _plt_upm
                    import numpy as _np_upm
                    import io as _io_upm
                    import base64 as _b64_upm

                    _cfg_upm = _json_upm.loads(open(str(bucket_json), encoding='utf-8').read())
                    _analyses = _cfg_upm.get('analyses', []) if isinstance(_cfg_upm, dict) else []
                    _upm_targets = _cfg_upm.get('upm_target', []) if isinstance(_cfg_upm, dict) else []
                    _upm_tgt_pct = None
                    _upm_tgt_label = 'Target'
                    for _ut in _upm_targets:
                        if _ut.get('target_%') is not None:
                            _upm_tgt_pct = float(_ut['target_%'])
                            _upm_tgt_label = _ut.get('test', 'Target')
                            break

                    _csv_df_upm = _pd_upm.read_csv(str(resolved_csv), dtype=object)

                    for _anl in _analyses:
                        if _anl.get('type') != 'distribution':
                            continue
                        _atag = _anl.get('tag', 'UPM_dist')
                        _filt = _anl.get('filter', {})
                        _filt_col = _filt.get('column', '')
                        _filt_match = _filt.get('match', {})
                        _method = _filt_match.get('method', '')
                        _filt_val = _filt_match.get('value', '')
                        _agg = _anl.get('aggregation', {})
                        _agg_mode = _agg.get('mode', 'count')
                        _base_cfg = _agg.get('base', {})
                        _base_val = _base_cfg.get('value', 1) if _base_cfg.get('type') == 'fixed' else None

                        # Find matching columns
                        import re as _re_upm
                        import fnmatch as _fnmatch_upm
                        _has_wc = ('*' in _filt_val or '?' in _filt_val)
                        if _has_wc:
                            _match_cols = [c for c in _csv_df_upm.columns if _fnmatch_upm.fnmatch(c, _filt_val)]
                        elif _method == 'starts_with' and _filt_val:
                            _match_cols = [c for c in _csv_df_upm.columns if c.startswith(_filt_val)]
                        elif _method == 'contains' and _filt_val:
                            _match_cols = [c for c in _csv_df_upm.columns if _filt_val in c]
                        elif _method == 'exact' and _filt_val:
                            _match_cols = [c for c in _csv_df_upm.columns if c == _filt_val]
                        elif _method == 'regex' and _filt_val:
                            _match_cols = [c for c in _csv_df_upm.columns if _re_upm.search(_filt_val, c)]
                        else:
                            _match_cols = [c for c in _csv_df_upm.columns if _filt_col.lower() in c.lower()]

                        if not _match_cols:
                            continue

                        # Compute values per column
                        _dist_data = []
                        for _mc in _match_cols:
                            _vals = _pd_upm.to_numeric(_csv_df_upm[_mc], errors='coerce').dropna()
                            if _vals.empty:
                                continue
                            _median = float(_vals.median())
                            _mean = float(_vals.mean())
                            _count = len(_vals)
                            if _agg_mode == 'percentage' and _base_val:
                                _pct = (_median / _base_val) * 100
                            else:
                                _pct = _median
                            # Short label: strip common prefix
                            _short = _mc[len(_filt_val):].strip('_') if _filt_val and _mc.startswith(_filt_val) else _mc
                            _dist_data.append({
                                'col': _mc, 'short': _short,
                                'median': _median, 'mean': _mean,
                                'count': _count, 'pct': _pct
                            })

                        if not _dist_data:
                            continue

                        _ylabel = _anl.get('output', {}).get('percentage_label', '%')

                        # ── Collect per-wafer per-column medians for interactive JS ──
                        import json as _json_upm_ic
                        _lot_col_upm = (next((c for c in _csv_df_upm.columns if c.lower() == 'sort_lot'), None)
                                        or next((c for c in _csv_df_upm.columns if c.lower() == 'lot'), None)
                                        or next((c for c in _csv_df_upm.columns if 'lot' in c.lower() and 'slot' not in c.lower()), None))
                        _wfr_col_upm = next((c for c in _csv_df_upm.columns
                                             if 'sort_wafer' in c.lower()
                                             or ('wafer' in c.lower() and 'sort_wafer' not in c.lower())), None)
                        _prg_col_upm = next((c for c in _csv_df_upm.columns if 'program' in c.lower()), None)
                        _mat_col_upm = next((c for c in _csv_df_upm.columns if 'material' in c.lower()), None)
                        _grp_cols_upm = [c for c in [_prg_col_upm, _lot_col_upm, _wfr_col_upm] if c]
                        _short_labels = [d['short'] for d in _dist_data]
                        _full_cols    = [d['col']   for d in _dist_data]
                        _upm_wfr_rows = []
                        try:
                            if _grp_cols_upm:
                                for _ukeys, _ugdf in _csv_df_upm.groupby(_grp_cols_upm):
                                    if not isinstance(_ukeys, tuple):
                                        _ukeys = (_ukeys,)
                                    _ukd = dict(zip(_grp_cols_upm, _ukeys))
                                    _mat_val = str(_ugdf[_mat_col_upm].iloc[0]) if _mat_col_upm and not _ugdf[_mat_col_upm].dropna().empty else ''
                                    _wrow = {
                                        'program':  str(_ukd.get(_prg_col_upm, '')),
                                        'lot':      str(_ukd.get(_lot_col_upm, '')),
                                        'wafer':    str(_ukd.get(_wfr_col_upm, '')),
                                        'material': _mat_val,
                                        'total':    int(len(_ugdf)),
                                        'medians':  {},
                                        'vals':     {},
                                    }
                                    for _mc2, _sh2 in zip(_full_cols, _short_labels):
                                        _vs2 = _pd_upm.to_numeric(_ugdf[_mc2], errors='coerce').dropna()
                                        if not _vs2.empty:
                                            _v2 = (_vs2 / _base_val * 100) if (_agg_mode == 'percentage' and _base_val) else _vs2
                                            _wrow['medians'][_sh2] = round(float(_v2.median()), 4)
                                            _wrow['vals'][_sh2] = [round(float(x), 3) for x in _v2.values]
                                    _upm_wfr_rows.append(_wrow)
                            else:
                                _wrow = {'program': '', 'lot': '', 'wafer': 'all',
                                         'material': '', 'total': len(_csv_df_upm), 'medians': {}}
                                for _d2 in _dist_data:
                                    _wrow['medians'][_d2['short']] = round(_d2['pct'], 4)
                                _upm_wfr_rows.append(_wrow)
                        except Exception:
                            for _d2 in _dist_data:
                                _upm_wfr_rows = [{'program': '', 'lot': '', 'wafer': 'all',
                                                  'material': '', 'total': _d2['count'],
                                                  'medians': {_d2['short']: round(_d2['pct'], 4)}}]

                        # Per-column pre-binned histogram data (all rows — for detail view)
                        _upm_hist_data = {}
                        for _d2 in _dist_data:
                            try:
                                _hv = _pd_upm.to_numeric(_csv_df_upm[_d2['col']], errors='coerce').dropna()
                                if _agg_mode == 'percentage' and _base_val:
                                    _hv = (_hv / _base_val) * 100
                                _hv = _hv.values
                                _hn = min(60, max(15, len(_hv) // 20))
                                import numpy as _np_upm2
                                _hcounts, _hedges = _np_upm2.histogram(_hv, bins=_hn)
                                _upm_hist_data[_d2['short']] = {
                                    'edges': [round(float(e), 4) for e in _hedges],
                                    'counts': [int(c) for c in _hcounts],
                                }
                            except Exception:
                                pass

                        _upm_ic_data = _json_upm_ic.dumps({
                            'tag': _atag,
                            'columns': _short_labels,
                            'target': _upm_tgt_pct,
                            'targetLabel': _upm_tgt_label,
                            'ylabel': _ylabel,
                            'mode': _agg_mode,
                            'hasMaterial': bool(_mat_col_upm),
                            'rows': _upm_wfr_rows,
                            'hist': _upm_hist_data,
                        }, ensure_ascii=False)

                        _upm_html = (
                            '<!doctype html>\n<html>\n<head>\n<meta charset="utf-8">\n'
                            '<title>UPM Distribution</title>\n'
                            '<style>\n'
                            '*{box-sizing:border-box;margin:0;padding:0}\n'
                            'body{font-family:Arial,sans-serif;font-size:14px;background:#f0f2f5;color:#2c3e50}\n'
                            '.pw{padding:10px 14px;margin:0}\n'
                            '.ib{display:flex;flex-wrap:wrap;gap:8px;padding:8px 12px;background:#2c3e50;'
                            'color:#ecf0f1;border-radius:6px;margin-bottom:10px;font-size:13px}\n'
                            '.ib b{color:#f1c40f}\n'
                            '.main-layout{display:flex;gap:0;align-items:flex-start}\n'
                            '.filter-col{flex:0 0 auto;min-width:280px;margin-right:14px}\n'
                            '.content-col{flex:1;min-width:0}\n'
                            '.cp{background:#fff;border-radius:6px;padding:10px;'
                            'box-shadow:0 1px 4px rgba(0,0,0,.1);margin-bottom:10px}\n'
                            '.ctr{display:flex;align-items:baseline;gap:8px;margin-bottom:6px}\n'
                            '.ct{font-size:15px;font-weight:bold;color:#2c3e50}\n'
                            '.si{font-size:13px;color:#7f8c8d}\n'
                            '.hs{width:50%;display:block}\n'
                            '.fs{background:#fff;border-radius:6px;padding:10px;'
                            'box-shadow:0 1px 4px rgba(0,0,0,.1);margin-bottom:10px}\n'
                            '.fh{display:flex;align-items:center;gap:8px;margin-bottom:8px;flex-wrap:wrap}\n'
                            '.ft{font-size:15px;font-weight:bold;color:#2c3e50}\n'
                            '.ri{font-size:13px;color:#7f8c8d}\n'
                            '.ftw{overflow-x:auto;max-height:calc(100vh - 120px);overflow-y:auto}\n'
                            '.ftbl{border-collapse:collapse;font-size:13px;white-space:nowrap}\n'
                            '.ftbl th{background:#2c3e50;color:#ecf0f1;padding:6px 12px;text-align:left;'
                            'position:sticky;top:0;z-index:1}\n'
                            '.ftbl td{padding:5px 12px;border-bottom:1px solid #eee}\n'
                            '.ftbl .fi th{padding:2px 4px!important;background:#3d5166;top:37px}\n'
                            '.ftbl .fi input{width:100%;box-sizing:border-box;font-size:11px;padding:2px 4px;border:1px solid #5d7384;border-radius:2px;background:#fff;color:#2c3e50}\n'
                            '.fr{cursor:pointer;transition:background .1s}\n'
                            '.fr:hover td{background:#f0f4ff}\n'
                            '.frs td{background:#d6eaff!important;font-weight:bold}\n'
                            '.num{text-align:right}\n'
                            '.flt-btn{background:none;border:none;color:#aed6f1;cursor:pointer;font-size:11px;padding:0 0 0 4px;vertical-align:middle;opacity:.85}\n'
                            '.flt-btn:hover{opacity:1;color:#fff}\n'
                            '.flt-btn.active{color:#f1c40f!important;opacity:1}\n'
                            '.dd-panel{position:fixed;background:#fff;border:1px solid #aaa;border-radius:4px;box-shadow:0 4px 16px rgba(0,0,0,.18);z-index:9999;min-width:180px;max-width:280px;font-family:Arial,sans-serif;font-size:12px;color:#2c3e50}\n'
                            '.dd-panel .dd-search{width:100%;box-sizing:border-box;padding:5px 8px;border:none;border-bottom:1px solid #ddd;font-size:12px;outline:none}\n'
                            '.dd-panel .dd-acts{display:flex;gap:4px;padding:4px 6px;border-bottom:1px solid #eee}\n'
                            '.dd-panel .dd-acts button{flex:1;padding:2px 6px;font-size:11px;cursor:pointer;border:1px solid #bdc3c7;background:#ecf0f1;border-radius:3px}\n'
                            '.dd-panel .dd-list{max-height:200px;overflow-y:auto;padding:4px 0}\n'
                            '.dd-panel .dd-item{display:flex;align-items:center;gap:6px;padding:3px 10px;cursor:pointer}\n'
                            '.dd-panel .dd-item:hover{background:#eaf0fb}\n'
                            '.dd-panel .dd-item input{margin:0;cursor:pointer}\n'
                            '.dd-panel .dd-footer{padding:4px 8px;border-top:1px solid #eee;text-align:right}\n'
                            '.dd-panel .dd-footer button{padding:3px 12px;font-size:11px;cursor:pointer;background:#2c3e50;color:#fff;border:none;border-radius:3px}\n'
                            '.dp{background:#fff;border-radius:6px;padding:10px;'
                            'box-shadow:0 1px 4px rgba(0,0,0,.1);margin-bottom:10px;display:none}\n'
                            '.dpt{font-size:15px;font-weight:bold;color:#2c3e50;margin-bottom:8px}\n'
                            '.cb{padding:3px 10px;font-size:12px;cursor:pointer;border:1px solid #bdc3c7;'
                            'background:#ecf0f1;border-radius:3px;color:#2c3e50}\n'
                            '.cb:hover{background:#d5dbde}\n'
                            '.stbl{border-collapse:collapse;font-size:13px}\n'
                            '.stbl th{background:#2c3e50;color:#ecf0f1;padding:6px 12px;text-align:left}\n'
                            '.stbl td{padding:5px 12px;border-bottom:1px solid #dde}\n'
                            '.stbl tr:nth-child(even) td{background:#eaf0fb}\n'
                            '.stbl tr:hover td{background:#d6eaff}\n'
                            '</style>\n</head>\n<body>\n<div class="pw">\n'
                            '<div class="ib"><span>&#128202; UPM Distribution</span></div>\n'
                            '<div class="main-layout">\n'
                            '<div class="filter-col">\n'
                            '<div class="fs">\n'
                            '  <div class="fh">\n'
                            '    <span class="ft"><svg width="13" height="13" viewBox="0 0 24 24" fill="currentColor" style="vertical-align:middle;margin-right:4px"><path d="M10 18h4v-2h-4v2zM3 6v2h18V6H3zm3 7h12v-2H6v2z"/></svg> Filter by Lot / Wafer</span>\n'
                            '    <button class="cb" onclick="U.selAll()">Select All</button>\n'
                            '    <button class="cb" onclick="U.clrAll()">Clear</button>\n'
                            '    <span id="row-info" class="ri"></span>\n'
                            '  </div>\n'
                            '  <div class="ftw">\n'
                            '    <table class="ftbl">\n'
                            '      <thead><tr>\n'
                            '        <th>TestProgram <button class="flt-btn" id="uft-fb-0" onclick="event.stopPropagation();U.ftDdOpen(0,this)" title="Filter">&#9660;</button></th>'
                            '<th>Lot <button class="flt-btn" id="uft-fb-1" onclick="event.stopPropagation();U.ftDdOpen(1,this)" title="Filter">&#9660;</button></th>'
                            '<th>Wafer <button class="flt-btn" id="uft-fb-2" onclick="event.stopPropagation();U.ftDdOpen(2,this)" title="Filter">&#9660;</button></th>'
                            +('<th>MaterialType <button class="flt-btn" id="uft-fb-3" onclick="event.stopPropagation();U.ftDdOpen(3,this)" title="Filter">&#9660;</button></th>' if _mat_col_upm else '')
                            +f'<th class="num">Total</th>\n'
                            '      </tr>\n'
                            '      </thead>\n'
                            '      <tbody id="ftbody"></tbody>\n'
                            '    </table>\n'
                            '  </div>\n'
                            '</div>\n'
                            '</div>\n'
                            '<div class="content-col">\n'
                            '<div class="cp">\n'
                            '  <div class="ctr"><span class="ct"><svg width="13" height="13" viewBox="0 0 24 24" fill="currentColor" style="vertical-align:middle;margin-right:4px"><path d="M10 18h4v-2h-4v2zM3 6v2h18V6H3zm3 7h12v-2H6v2z"/></svg> Distribution per Test</span></div>\n'
                            '  <div style="overflow-x:auto">\n'
                            '    <table class="dtbl">\n'
                            '      <thead><tr>\n'
                            '        <th>#</th><th>Test</th><th class="num">Median</th><th class="num">Wafers</th>'
                            '      </tr></thead>\n'
                            '      <tbody id="dist-tbody"></tbody>\n'
                            '    </table>\n'
                            '  </div>\n'
                            '</div>\n'
                            '<div class="dp" id="detail-panel">\n'
                            '  <div class="dpt" id="detail-title"></div>\n'
                            '  <svg id="hist-svg" class="hs"></svg>\n'
                            '  <div style="margin-top:10px;overflow-x:auto">\n'
                            '    <table class="stbl">\n'
                            '      <thead><tr>\n'
                            '        <th>Program</th><th>Lot</th><th>Wafer</th>'+('<th>MaterialType</th>' if _mat_col_upm else '')+f'\n'
                            '        <th class="num">N Die</th><th class="num">Min</th><th class="num">Median</th><th class="num">Max</th>\n'
                            '      </tr></thead>\n'
                            '      <tbody id="detail-tbody"></tbody>\n'
                            '    </table>\n'
                            '  </div>\n'
                            '</div>\n'
                            '</div>\n'
                            '</div>\n'
                            '</div>\n'
                            '<script>\nvar UDATA=' + _upm_ic_data + ';\n'
                            + r'''var U=(function(){
'use strict';
var sR=new Set(UDATA.rows.map(function(_,i){return i;}));
var lR=-1;
var selCol=UDATA.columns.length?UDATA.columns[0]:null;
function esc(s){return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
function fmt(v){return (v===null||v===undefined)?'—':v.toFixed(2);}
function gMed(){
  var agg={};
  UDATA.columns.forEach(function(c){agg[c]=[];});
  sR.forEach(function(i){
    var r=UDATA.rows[i];
    UDATA.columns.forEach(function(c){if(r.medians[c]!==undefined)agg[c].push(r.medians[c]);});
  });
  var out={};
  UDATA.columns.forEach(function(c){
    var arr=agg[c].slice().sort(function(a,b){return a-b;});
    if(!arr.length){out[c]=null;return;}
    var m=arr.length%2===0?(arr[arr.length/2-1]+arr[arr.length/2])/2:arr[Math.floor(arr.length/2)];
    out[c]=m;
  });
  return out;
}
var _ftDdState={};var _ftDdOpen_=null;
function ftDdOpen(col,btn){
  if(_ftDdOpen_){_ftDdClose();}
  var allVals=[];var seen=new Set();
  UDATA.rows.forEach(function(row){
    var cols=[row.program,row.lot,row.wafer].concat(UDATA.hasMaterial?[row.material||'']:[]);
    var v=String(cols[col]||'');if(!seen.has(v)){seen.add(v);allVals.push(v);}
  });
  allVals.sort(function(a,b){return a.localeCompare(b);});
  var allowed=_ftDdState[col];
  var checked=allowed?new Set(allowed):new Set(allVals);
  var panel=document.createElement('div');panel.className='dd-panel';
  panel.innerHTML='<input class="dd-search" placeholder="Search\u2026">'
    +'<div class="dd-acts"><button class="uft-sel">Select All</button><button class="uft-clr">Clear</button></div>'
    +'<div class="dd-list" id="uft-dd-list"></div>'
    +'<div class="dd-footer"><button class="uft-ok">OK</button></div>';
  document.body.appendChild(panel);
  var r=btn.getBoundingClientRect();
  panel.style.top=(r.bottom+2)+'px';
  panel.style.left=Math.min(r.left,window.innerWidth-200)+'px';
  _ftDdOpen_={panel:panel,col:col,btn:btn,allVals:allVals,checked:checked};
  _ftRenderList(allVals);
  panel.querySelector('.dd-search').oninput=function(){var q=(this.value||'').toLowerCase();var fl=q?_ftDdOpen_.allVals.filter(function(v){return v.toLowerCase().indexOf(q)>=0;}):_ftDdOpen_.allVals;_ftRenderList(fl);};
  panel.querySelector('.uft-sel').onclick=function(){_ftSelAll();};
  panel.querySelector('.uft-clr').onclick=function(){_ftClearAll();};
  panel.querySelector('.uft-ok').onclick=function(){_ftApply();};
  setTimeout(function(){document.addEventListener('mousedown',_ftOutside);},0);
}
function _ftRenderList(vals){
  var list=document.getElementById('uft-dd-list');if(!list)return;
  var h='';vals.forEach(function(v){var c=_ftDdOpen_&&_ftDdOpen_.checked.has(v)?' checked':'';
    h+='<label class="dd-item"><input type="checkbox"'+c+' data-val="'+v.replace(/&/g,'&amp;').replace(/"/g,'&quot;')+'">'+esc(v)+'</label>';
  });list.innerHTML=h;
  list.querySelectorAll('input').forEach(function(inp){inp.onchange=function(){_ftToggle(inp,inp.dataset.val);};});
}
function _ftToggle(cb,v){if(_ftDdOpen_){if(cb.checked)_ftDdOpen_.checked.add(v);else _ftDdOpen_.checked.delete(v);}}
function _ftSelAll(){if(_ftDdOpen_){_ftDdOpen_.allVals.forEach(function(v){_ftDdOpen_.checked.add(v);});_ftRenderList(_ftDdOpen_.allVals);}}
function _ftClearAll(){if(_ftDdOpen_){_ftDdOpen_.checked.clear();_ftRenderList(_ftDdOpen_.allVals);}}
function _ftApply(){
  if(!_ftDdOpen_)return;
  var c=_ftDdOpen_.col,chk=_ftDdOpen_.checked,all=_ftDdOpen_.allVals;
  _ftDdState[c]=(chk.size===all.length)?null:new Set(chk);
  var b=document.getElementById('uft-fb-'+c);if(b)b.classList.toggle('active',!!_ftDdState[c]);
  _ftDdClose();rFilter();
}
function _ftDdClose(){
  if(!_ftDdOpen_)return;
  document.removeEventListener('mousedown',_ftOutside);
  if(_ftDdOpen_.panel.parentNode)_ftDdOpen_.panel.parentNode.removeChild(_ftDdOpen_.panel);
  _ftDdOpen_=null;
}
function _ftOutside(e){if(_ftDdOpen_&&!_ftDdOpen_.panel.contains(e.target)){_ftApply();}}
function rFilter(){
  var tb=document.getElementById('ftbody');
  var html='';
  UDATA.rows.forEach(function(row,i){
    var cols=[row.program,row.lot,row.wafer].concat(UDATA.hasMaterial?[row.material||'']:[]);
    var show=Object.keys(_ftDdState).every(function(ci){var s=_ftDdState[ci];return !s||s.has(String(cols[parseInt(ci)]||''));});
    if(!show)return;
    var s=sR.has(i);
    html+='<tr class="fr'+(s?' frs':'')+'" onclick="U.toggleRow('+i+',event)">';
    html+='<td>'+esc(row.program)+'</td><td>'+esc(row.lot)+'</td><td>'+esc(row.wafer)+'</td>';
    if(UDATA.hasMaterial)html+='<td>'+esc(row.material||'')+'</td>';
    html+='<td class="num">'+row.total.toLocaleString()+'</td></tr>';
  });
  tb.innerHTML=html;
  document.getElementById('row-info').textContent=
    sR.size<UDATA.rows.length?'('+sR.size+'/'+UDATA.rows.length+' selected)':'';
}
function rDetail(){
  var dp=document.getElementById('detail-panel');
  if(!selCol){dp.style.display='none';return;}
  dp.style.display='block';
  document.getElementById('detail-title').textContent='Distribution \u2014 '+selCol;
  // Collect raw die values from all selected wafers
  var allRaw=[];
  var perWafer=[];
  sR.forEach(function(i){
    var r=UDATA.rows[i];
    var rv=(r.vals&&r.vals[selCol])||[];
    if(rv.length){
      allRaw=allRaw.concat(rv);
      perWafer.push({prog:r.program,lot:r.lot,wafer:r.wafer,material:r.material||'',vals:rv,median:r.medians[selCol]});
    } else if(r.medians[selCol]!==undefined){
      perWafer.push({prog:r.program,lot:r.lot,wafer:r.wafer,material:r.material||'',vals:[],median:r.medians[selCol]});
    }
  });
  allRaw.sort(function(a,b){return a-b;});
  function medOfArr(arr){
    if(!arr.length)return null;
    var n=arr.length;
    return n%2===0?(arr[n/2-1]+arr[n/2])/2:arr[Math.floor(n/2)];
  }
  function compHist(arr,nb){
    if(!arr.length)return null;
    var lo=arr[0],hi=arr[arr.length-1];
    if(lo===hi){lo-=0.5;hi+=0.5;}
    var n=Math.min(nb,Math.max(3,arr.length));
    var step=(hi-lo)/n,counts=new Array(n).fill(0),edges=[];
    for(var k=0;k<=n;k++)edges.push(lo+k*step);
    arr.forEach(function(v){var idx=Math.min(n-1,Math.floor((v-lo)/step));counts[idx]++;});
    return{edges:edges,counts:counts};
  }
  var useRaw=allRaw.length>0;
  var hd;
  if(useRaw){
    var nb=Math.min(60,Math.max(10,Math.floor(allRaw.length/30)+10));
    hd=compHist(allRaw,nb);
  } else {
    hd=UDATA.hist[selCol];
  }
  var svg=document.getElementById('hist-svg');
  var W=svg.clientWidth||280,H=240,pl=60,pr=16,pt=32,pb=62,cW=W-pl-pr,cH=H-pt-pb;
  var p=[];
  p.push('<rect width="'+W+'" height="'+H+'" fill="#f8f9fa"/>');
  if(hd){
    var maxC=Math.max.apply(null,hd.counts);
    var n=hd.counts.length,bw=cW/n;
    var tgt=UDATA.target;
    for(var i=0;i<n;i++){
      var clr='#3498db';
      var bh=maxC>0?hd.counts[i]/maxC*cH:0;
      var x=pl+i*bw,y=pt+cH-bh;
      p.push('<rect x="'+x.toFixed(1)+'" y="'+y.toFixed(1)+'" width="'+(bw-.5).toFixed(1)+'" height="'+bh.toFixed(1)+'" fill="'+clr+'" opacity="0.8"/>');
    }
    var lo=hd.edges[0],hi=hd.edges[hd.edges.length-1];
    // Median line (of raw die; fallback to wafer medians)
    var medV=useRaw?medOfArr(allRaw):null;
    if(medV===null&&perWafer.length){
      var meds=perWafer.map(function(w){return w.median;}).filter(function(v){return v!==undefined&&v!==null;}).sort(function(a,b){return a-b;});
      medV=medOfArr(meds);
    }
    if(medV!==null&&medV!==undefined){
      var mx=pl+(medV-lo)/(hi-lo)*cW;
      if(mx>=pl&&mx<=pl+cW){
        p.push('<line x1="'+mx.toFixed(1)+'" x2="'+mx.toFixed(1)+'" y1="'+pt+'" y2="'+(pt+cH)+'" stroke="#8B4513" stroke-width="2.5" stroke-dasharray="5,3"/>');
        p.push('<text x="'+(mx+4).toFixed(1)+'" y="'+(pt+22)+'" font-family="Arial" font-size="22" font-weight="bold" fill="#8B4513">Med:'+medV.toFixed(2)+'</text>');
      }
    }
    p.push('<text x="'+pl+'" y="'+(H-6)+'" text-anchor="middle" font-family="Arial" font-size="18" fill="#777">'+lo.toFixed(1)+'</text>');
    p.push('<text x="'+(pl+cW)+'" y="'+(H-6)+'" text-anchor="middle" font-family="Arial" font-size="18" fill="#777">'+hi.toFixed(1)+'</text>');
    p.push('<text x="'+(pl+cW/2).toFixed(1)+'" y="'+(H-6)+'" text-anchor="middle" font-family="Arial" font-size="18" fill="#777">'+esc(UDATA.ylabel)+'</text>');
    var nLbl=useRaw?'('+allRaw.length+' raw die, '+perWafer.length+' wafer'+(perWafer.length>1?'s':'')+')':'(wafer medians only)';
    p.push('<text x="'+(pl+cW/2).toFixed(1)+'" y="'+(pt-6)+'" text-anchor="middle" font-family="Arial" font-size="18" fill="#999">'+nLbl+'</text>');
  }
  p.push('<line x1="'+pl+'" x2="'+pl+'" y1="'+pt+'" y2="'+(pt+cH)+'" stroke="#aaa"/>');
  p.push('<line x1="'+pl+'" x2="'+(W-pr)+'" y1="'+(pt+cH)+'" y2="'+(pt+cH)+'" stroke="#aaa"/>');
  svg.setAttribute('viewBox','0 0 '+W+' '+H);
  svg.setAttribute('height',H);
  svg.innerHTML=p.join('');
  // Per-wafer stats table
  perWafer.sort(function(a,b){return (a.median||0)-(b.median||0);});
  var tb=document.getElementById('detail-tbody');
  var html='';
  perWafer.slice().reverse().forEach(function(r){
    var sv=r.vals.slice().sort(function(a,b){return a-b;});
    var mn=sv.length?sv[0].toFixed(2):'\u2014';
    var mx2=sv.length?sv[sv.length-1].toFixed(2):'\u2014';
    var med2=sv.length?medOfArr(sv).toFixed(2):(r.median!==undefined?r.median.toFixed(2):'\u2014');
    var n2=sv.length||'\u2014';
    html+='<tr><td>'+esc(r.prog)+'</td><td>'+esc(r.lot)+'</td><td>'+esc(r.wafer)+'</td>';
    if(UDATA.hasMaterial)html+='<td>'+esc(r.material)+'</td>';
    html+='<td class="num">'+n2+'</td><td class="num">'+mn+'</td><td class="num">'+med2+'</td><td class="num">'+mx2+'</td></tr>';
  });
  tb.innerHTML=html;
}
function rDistTable(){
  var tb=document.getElementById('dist-tbody');
  if(!tb)return;
  var med=gMed();
  var html='';
  UDATA.columns.forEach(function(col,idx){
    var v=med[col];
    var hd=UDATA.hist[col];
    var isSel=(col===selCol);
    var cnt=0;
    sR.forEach(function(i){if(UDATA.rows[i].medians[col]!==undefined)cnt++;});
    html+='<tr class="'+(isSel?'sel-row':'')+'" onclick="U.clickCol(\''+esc(col)+'\')">';
    html+='<td>'+(idx+1)+'</td>';
    html+='<td>'+esc(col)+'</td>';
    html+='<td class="num">'+(v!==null?v.toFixed(2):'\u2014')+' '+esc(UDATA.ylabel)+'</td>';
    html+='<td class="num">'+cnt+'</td>';
    html+='</tr>';
  });
  tb.innerHTML=html;
}
function upd(){
  rFilter();rDetail();rDistTable();
}
function clickCol(col){
  selCol=(selCol===col)?null:col;
  upd();
}
function toggleRow(idx,ev){
  if(ev&&ev.shiftKey&&lR>=0){
    var lo=Math.min(idx,lR),hi=Math.max(idx,lR);
    for(var i=lo;i<=hi;i++)sR.add(i);
  }else if(ev&&(ev.ctrlKey||ev.metaKey)){
    if(sR.has(idx)){if(sR.size>1)sR.delete(idx);}else sR.add(idx);
  }else{
    if(sR.size===UDATA.rows.length){sR.clear();sR.add(idx);}
    else if(sR.size===1&&sR.has(idx)){UDATA.rows.forEach(function(_,i){sR.add(i);});}
    else if(sR.has(idx)){sR.delete(idx);}
    else{sR.add(idx);}
  }
  lR=idx;upd();
}
function selAll(){
  var visible=[];
  UDATA.rows.forEach(function(row,i){
    var cols=[row.program,row.lot,row.wafer].concat(UDATA.hasMaterial?[row.material||'']:[]);
    var show=Object.keys(_ftDdState).every(function(ci){var s=_ftDdState[ci];return !s||s.has(String(cols[parseInt(ci)]||''));});
    if(show)visible.push(i);
  });
  visible.forEach(function(i){sR.add(i);});lR=-1;upd();
}
function clrAll(){sR.clear();if(UDATA.rows.length)sR.add(0);lR=-1;upd();}
function colFilter(inp,ci){}
if(document.readyState==='loading'){
  document.addEventListener('DOMContentLoaded',function(){upd();window.addEventListener('resize',function(){rDetail();rDistTable();});});
}else{upd();window.addEventListener('resize',function(){rDetail();rDistTable();});}
return{clickCol:clickCol,toggleRow:toggleRow,selAll:selAll,clrAll:clrAll,
  ftDdOpen:ftDdOpen};
})();
</script>
</body></html>'''
                        )
                        _upm_fname = f'upm_distribution_{_atag}.html'
                        _upm_path = out_dir / _upm_fname
                        _upm_path.write_text(_wm_inject(_upm_html), encoding='utf-8')
                        _upm_dist_section += _nav('&#128202; UPM Distribution', _upm_fname)
            except Exception:
                pass

            _tag_display = tag if tag else 'Yield Analysis'
            # Build combined Wafer Map page — single entry in sidebar
            _wm_nav = ''
            if _wm_files:
                try:
                    import pandas as _pd_wm
                    import re as _re_wm
                    def _wm_sanitize(s):
                        return _re_wm.sub(r'[^0-9A-Za-z_-]', '_', str(s))
                    _df_wm = _pd_wm.read_csv(str(resolved_csv), dtype=object)
                    _lot_wm = (next((c for c in _df_wm.columns if c.lower() == 'sort_lot'), None)
                               or next((c for c in _df_wm.columns if c.lower() == 'lot'), None)
                               or next((c for c in _df_wm.columns if 'lot' in c.lower() and 'slot' not in c.lower()), None))
                    _wfr_wm = next((c for c in _df_wm.columns if 'sort_wafer' in c.lower() or ('wafer' in c.lower() and 'sort_wafer' not in c.lower())), None)
                    _prg_wm = next((c for c in _df_wm.columns if 'program' in c.lower()), None)
                    # Mirror bin_distribution_html.py: drop rows with blank/nan lot or wafer
                    # (bin_distribution groups by [prog,lot,wafer] so NaN keys are auto-dropped)
                    _wm_null = {'', 'nan', 'none', 'na'}
                    # Mirror bin_distribution_html.py: group by [prog,lot,wafer] drops any row
                    # where ANY of those three keys is NaN (pandas groupby dropna=True default).
                    for _wm_filt_col in [_lot_wm, _wfr_wm, _prg_wm]:
                        if _wm_filt_col:
                            _df_wm = _df_wm[_df_wm[_wm_filt_col].notna() & ~_df_wm[_wm_filt_col].astype(str).str.strip().str.lower().isin(_wm_null)]
                    _mat_wm = next((c for c in _df_wm.columns if 'material' in c.lower()), None)
                    # map sanitized lot → href
                    _lot_href_map = {_wf.stem[len(f'{stem}_IBIN_WaferMap_'):]: f'heatmap/{_wf.name}' for _wf in _wm_files}
                    def _find_lot_href(lot_v):
                        _k = _wm_sanitize(str(lot_v))
                        if _k in _lot_href_map: return _lot_href_map[_k]
                        if _k:  # guard: empty string matches everything via 'in'
                            for k, v in _lot_href_map.items():
                                if str(lot_v) in k or k in str(lot_v): return v
                        return ''
                    _wm_mat_th = '<th>Material</th>' if _mat_wm else ''
                    _wm_nav_rows = ''
                    _first_lot_href = next(iter(_lot_href_map.values()), '') if _lot_href_map else ''
                    _grp_wm_cols = [c for c in [_lot_wm, _wfr_wm] if c]
                    if _lot_wm:
                        for _lot_v, _lot_df in _df_wm.groupby(_lot_wm):
                            _ls = str(_lot_v)
                            if not _ls or _ls.lower() in ('nan', 'none', 'na'):
                                continue  # skip blank/nan lot values
                            _lhref = _find_lot_href(_ls)
                            if not _lhref:
                                continue  # skip lots with no heatmap file
                            _lprog = str(_lot_df[_prg_wm].iloc[0]) if _prg_wm else ''
                            _lprog = '' if _lprog.lower() in ('nan', 'none', 'na') else _lprog
                            _ltot = len(_lot_df)
                            _lhref_js = _lhref.replace("'", "\\'")
                            _wm_nav_rows += (
                                f'<tr class="wm-lot-row" data-lot="{_ls}" onclick="wmLoad(\'{_lhref_js}\',this,event)">'
                                f'<td><b>{_lprog}</b></td><td><b>{_ls}</b></td>'
                                f'<td style="color:#888;font-style:italic">&#8213; ALL &#8213;</td>'
                                + (f'<td></td>' if _mat_wm else '')
                                + f'<td style="text-align:right"><b>{_ltot:,}</b></td></tr>\n'
                            )
                            if _wfr_wm:
                                _wfr_grp_cols = [_wfr_wm] + ([_prg_wm] if _prg_wm else [])
                                for _wfr_grp_key, _wdf in _lot_df.groupby(_wfr_grp_cols):
                                    if _prg_wm:
                                        _wfr_v, _prg_v = _wfr_grp_key
                                    else:
                                        _wfr_v = _wfr_grp_key; _prg_v = ''
                                    _ws = str(_wfr_v)
                                    _whref = _lhref + f'#wafer-{_ws}' if _lhref else ''
                                    _whref_js = _whref.replace("'", "\\'")
                                    _wmat = str(_wdf[_mat_wm].iloc[0]) if _mat_wm and not _wdf[_mat_wm].dropna().empty else ''
                                    _wmat = '' if _wmat.lower() in ('nan', 'none', 'na') else _wmat
                                    _wprog = str(_prg_v) if _prg_v else (str(_wdf[_prg_wm].iloc[0]) if _prg_wm else '')
                                    _wprog = '' if _wprog.lower() in ('nan', 'none', 'na') else _wprog
                                    _wtot = len(_wdf)
                                    _patkey = f'{_ls}::{_ws}::{_wprog}'
                                    _patkey_js = _patkey.replace("'", "\\'")
                                    _wm_nav_rows += (
                                        f'<tr class="wm-wafer-row" data-patkey="{_patkey}" onclick="wmLoad(\'{_whref_js}\',this,event)">'
                                        f'<td style="padding-left:20px;color:#555">{_wprog}</td>'
                                        f'<td style="padding-left:20px;color:#555">{_ls}</td>'
                                        f'<td style="padding-left:20px">{_ws}</td>'
                                        + (f'<td>{_wmat}</td>' if _mat_wm else '')
                                        + f'<td style="text-align:right">{_wtot:,}</td></tr>\n'
                                    )

                    # ── Per-wafer die data + reticle for pattern analysis ──────────
                    import json as _json_wm_pat
                    _wm_pat_wafers = {}
                    _x_wm = next((c for c in _df_wm.columns if c == 'SORT_X'), None)
                    _y_wm = next((c for c in _df_wm.columns if c == 'SORT_Y'), None)
                    _ib_wm = next((c for c in _df_wm.columns if 'INTERFACE_BIN' in c.upper() and 'TOTAL' not in c.upper()), None)
                    _fb_wm = next((c for c in _df_wm.columns if 'FUNCTIONAL_BIN' in c.upper() and 'TOTAL' not in c.upper()), None)
                    _upm_cols_wm = [c for c in _df_wm.columns if c.upper().startswith('UPM_')]
                    import re as _re_upm, fnmatch as _fnmatch_upm
                    _upm_labels_wm = []
                    _upm_cols_matched = []
                    _upm_info_raw = []
                    if bucket_json and os.path.isfile(str(bucket_json)):
                        try:
                            import json as _json_upm_info
                            _bj_upm = _json_upm_info.loads(open(str(bucket_json), encoding='utf-8').read())
                            _upm_info_raw = _bj_upm.get('upmInfo', [])
                        except Exception:
                            pass
                    if _upm_info_raw:
                        for _ui_entry in _upm_info_raw:
                            _ui_label = _ui_entry[0] if len(_ui_entry) > 0 else ''
                            _ui_pat = _ui_entry[1] if len(_ui_entry) > 1 else ''
                            _ui_target = _ui_entry[2] if len(_ui_entry) > 2 else None
                            _matched = next((c for c in _upm_cols_wm if _fnmatch_upm.fnmatch(c, _ui_pat)), None)
                            if _matched:
                                _upm_cols_matched.append(_matched)
                                # Strip trailing "(%) - SDS" style suffixes for compact display
                                import re as _re_upm2
                                _clean_lbl = _re_upm2.sub(r'\s*\(%\)\s*-\s*\S+\s*$', '', _ui_label).strip()
                                _upm_labels_wm.append([_clean_lbl, _ui_target])
                    else:
                        def _upm_short(c):
                            parts = c.split('_')
                            freq = parts[1] if len(parts) > 1 else ''
                            test_name = parts[2] if len(parts) > 2 else ''
                            volt_raw = parts[4] if len(parts) > 4 else ''
                            stat = parts[5] if len(parts) > 5 else ''
                            cell = _re_upm.search(r'(ULVT|HVT|SVT|LVT|HP)', test_name, _re_upm.I)
                            cell_lbl = cell.group(1).upper() if cell else test_name[:4]
                            freq_lbl = str(int(freq)) if freq.isdigit() else freq
                            volt_lbl = (str(int(volt_raw)) + 'mV') if volt_raw.isdigit() else volt_raw
                            return f'UPM {cell_lbl} {freq_lbl}/{volt_lbl} {stat}'
                        _upm_cols_matched = _upm_cols_wm
                        _upm_labels_wm = [[_upm_short(c), None] for c in _upm_cols_wm]
                    _hw_cols_wm = []
                    for _hw_nm_wm in ['sort partial wafer id', 'cell id', 'unit tester id', 'unit tester site id', 'unit tiu', 'thermal head id']:
                        _hw_c_wm = next((c for c in _df_wm.columns if c.lower().startswith(_hw_nm_wm) or c.lower().replace('_',' ').startswith(_hw_nm_wm)), None)
                        if _hw_c_wm:
                            _hw_cols_wm.append(_hw_c_wm)
                    _hw_fields_wm = [c.replace('_',' ').title() for c in _hw_cols_wm]
                    if _x_wm and _y_wm and _ib_wm and _lot_wm and _wfr_wm:
                        try:
                            import pandas as _pd_pat
                            _grp_wm_cols_p = [_lot_wm, _wfr_wm] + ([_prg_wm] if _prg_wm else [])
                            _grp_wm = _df_wm.groupby(_grp_wm_cols_p)
                            for _grp_key_p, _wdf_p in _grp_wm:
                                if _prg_wm:
                                    _lv_p, _wv_p, _pv_p = _grp_key_p
                                else:
                                    _lv_p, _wv_p = _grp_key_p; _pv_p = ''
                                _pk = f'{_lv_p}::{_wv_p}::{_pv_p}'
                                _dies_p = []
                                _die_cols = [_x_wm, _y_wm, _ib_wm] + ([_fb_wm] if _fb_wm else []) + _upm_cols_matched
                                _arr = _wdf_p[_die_cols].values
                                _upm_start = (4 if _fb_wm else 3)
                                for _r in _arr:
                                    try:
                                        _dx_p = int(float(_r[0])) if _r[0] is not None and str(_r[0]) not in ('nan','None','') else None
                                        _dy_p = int(float(_r[1])) if _r[1] is not None and str(_r[1]) not in ('nan','None','') else None
                                        _ib_p = int(float(_r[2])) if _r[2] is not None and str(_r[2]) not in ('nan','None','') else None
                                        _fb_p = int(float(_r[3])) if _fb_wm and _r[3] is not None and str(_r[3]) not in ('nan','None','') else None
                                        _upm_p = [round(float(_r[_upm_start+_ui]), 1) if _r[_upm_start+_ui] is not None and str(_r[_upm_start+_ui]) not in ('nan','None','') else None for _ui in range(len(_upm_cols_matched))]
                                        if _dx_p is not None and _dy_p is not None:
                                            _dies_p.append([_dx_p, _dy_p, _ib_p, _fb_p] + _upm_p)
                                    except Exception:
                                        pass
                                _ib_fb_p = {}
                                if _fb_wm:
                                    try:
                                        _fb_arr_p = _wdf_p[[_ib_wm, _fb_wm]].values
                                        for _fr in _fb_arr_p:
                                            try:
                                                _fib_p = int(float(_fr[0])) if _fr[0] is not None and str(_fr[0]) not in ('nan','None','') else None
                                                _ffb_p = int(float(_fr[1])) if _fr[1] is not None and str(_fr[1]) not in ('nan','None','') else None
                                                if _fib_p is not None and _ffb_p is not None:
                                                    _fibk = str(_fib_p); _ffbk = str(_ffb_p)
                                                    if _fibk not in _ib_fb_p: _ib_fb_p[_fibk] = {}
                                                    _ib_fb_p[_fibk][_ffbk] = _ib_fb_p[_fibk].get(_ffbk, 0) + 1
                                            except Exception:
                                                pass
                                    except Exception:
                                        pass
                                _ib_hw_p = {}
                                if _hw_cols_wm:
                                    try:
                                        _hw_arr_p = _wdf_p[[_ib_wm] + _hw_cols_wm].values
                                        for _hr in _hw_arr_p:
                                            try:
                                                _hib_p = int(float(_hr[0])) if _hr[0] is not None and str(_hr[0]) not in ('nan','None','') else None
                                                if _hib_p is None: continue
                                                _hvals_p = [str(_hr[i+1]).strip() if _hr[i+1] is not None and str(_hr[i+1]) not in ('nan','None','') else '' for i in range(len(_hw_cols_wm))]
                                                _hkey_p = '|'.join(_hvals_p)
                                                _hibk_p = str(_hib_p)
                                                if _hibk_p not in _ib_hw_p: _ib_hw_p[_hibk_p] = {}
                                                _ib_hw_p[_hibk_p][_hkey_p] = _ib_hw_p[_hibk_p].get(_hkey_p, 0) + 1
                                            except Exception:
                                                pass
                                    except Exception:
                                        pass
                                if _dies_p:
                                    _mat_p = str(_wdf_p[_mat_wm].iloc[0]) if _mat_wm and not _wdf_p[_mat_wm].dropna().empty else ''
                                    _dvs_wm = next((c for c in _df_wm.columns if 'devrewstep' in c.lower().replace('v','') or c.lower().startswith('devrewstep') or c.lower().startswith('devrevstep')), None)
                                    _dev_wm2 = next((c for c in _df_wm.columns if c.lower() == 'device'), None)
                                    _pfx_p = ''
                                    if _dev_wm2 and not _wdf_p[_dev_wm2].dropna().empty:
                                        _pfx_p = str(_wdf_p[_dev_wm2].iloc[0]).strip()[:6]
                                    elif _dvs_wm and not _wdf_p[_dvs_wm].dropna().empty:
                                        _pfx_p = str(_wdf_p[_dvs_wm].iloc[0]).strip()[:6]
                                    _wm_pat_wafers[_pk] = {'dies': _dies_p, 'lot': str(_lv_p), 'wafer': str(_wv_p), 'program': str(_pv_p), 'material': _mat_p, 'pfx': _pfx_p, 'ibToFb': _ib_fb_p, 'ibToHw': _ib_hw_p}
                        except Exception:
                            pass
                    _wm_ret_map = {}
                    _wm_ret_shots = []
                    _wm_ret_site_totals = {}
                    _wm_ret_site_labels = {}
                    try:
                        import glob as _glob_wm2
                        _ret_cands_wm = []
                        # Walk up from script location until shared/ is found, then use shared/reticle/
                        def _find_repo_root_wm(start):
                            d = _P(start)
                            for _ in range(10):
                                if (d / 'shared').is_dir():
                                    return d
                                if d.parent == d:
                                    break
                                d = d.parent
                            return None
                        _repo_root_wm = _find_repo_root_wm(_P(__file__).resolve().parent)
                        if _repo_root_wm:
                            _shared_ret_wm = _repo_root_wm / 'shared' / 'reticle'
                            _ret_cands_wm.extend(_glob_wm2.glob(str(_shared_ret_wm / '*.csv')))
                        _ret_cands_wm = list({p for p in _ret_cands_wm if _P(p).is_file() and 'reticle' in _P(p).name.lower()})
                        _ret_cands_wm.sort()
                        def _wm_build_ret_info(csv_path):
                            import pandas as _pd_ret_wm2
                            _rdf_wm = _pd_ret_wm2.read_csv(csv_path)
                            _rc = {c.lower().replace(' ', '').replace('_', ''): c for c in _rdf_wm.columns}
                            _rdx2 = _rc.get('diex'); _rdy2 = _rc.get('diey')
                            _rrx2 = _rc.get('reticlediex'); _rry2 = _rc.get('reticlediey')
                            _rrs2 = _rc.get('reticleshot')
                            _rret2 = _rc.get('reticle')
                            if not (_rdx2 and _rdy2 and _rrx2 and _rry2 and _rrs2):
                                return {}, [], {}, {}
                            _rdf2_cols = [_rdx2, _rdy2, _rrx2, _rry2, _rrs2] + ([_rret2] if _rret2 else [])
                            _rdf2 = _rdf_wm[_rdf2_cols].dropna().copy()
                            for _rc3 in [_rdx2, _rdy2, _rrx2, _rry2]:
                                _rdf2[_rc3] = _rdf2[_rc3].astype(int)
                            _off_x = round((_rdf2[_rdx2].min() + _rdf2[_rdx2].max()) / 2)
                            _off_y = round((_rdf2[_rdy2].min() + _rdf2[_rdy2].max()) / 2)
                            _rdf2['_SX'] = (_rdf2[_rdx2] - _off_x).astype(int)
                            _rdf2['_SY'] = (_rdf2[_rdy2] - _off_y).astype(int)
                            _snames = list(_rdf2[_rrs2].unique())
                            _sidx = {s: i for i, s in enumerate(_snames)}
                            _sbbox = {}
                            for _, _rv in _rdf2.iterrows():
                                _s = _rv[_rrs2]; _dx = int(_rv['_SX']); _dy = int(_rv['_SY'])
                                if _s not in _sbbox:
                                    _sbbox[_s] = [_dx, _dy, _dx, _dy]
                                else:
                                    _b = _sbbox[_s]
                                    _b[0]=min(_b[0],_dx); _b[1]=min(_b[1],_dy)
                                    _b[2]=max(_b[2],_dx); _b[3]=max(_b[3],_dy)
                            import re as _re_sg
                            def _parse_sg(nm):
                                _m = _re_sg.search(r'X(-?\d+)Y(-?\d+)', nm)
                                return (int(_m.group(1)), int(_m.group(2))) if _m else (0, 0)
                            _shots = [_sbbox[s] + list(_parse_sg(s)) for s in _snames]
                            _rm = {}
                            _sst = {}
                            _slbl = {}
                            for _, _rv in _rdf2.iterrows():
                                _kk = f"{int(_rv['_SX'])},{int(_rv['_SY'])}"
                                _rm[_kk] = [int(_rv[_rrx2]), int(_rv[_rry2]), _sidx[_rv[_rrs2]]]
                                _sk3 = f"{int(_rv[_rrx2])},{int(_rv[_rry2])}"
                                _sst.setdefault(_sk3, set()).add(_rv[_rrs2])
                                if _rret2 and _sk3 not in _slbl:
                                    try: _slbl[_sk3] = int(_rv[_rret2])
                                    except Exception: pass
                            _st = {k: len(v) for k, v in _sst.items()}
                            return _rm, _shots, _st, _slbl
                        # Gather unique prefixes from wafer data
                        _all_pfxs = list({w.get('pfx','') for w in _wm_pat_wafers.values() if w.get('pfx','')})
                        _wm_all_ret_maps = {}
                        for _pfx_r in _all_pfxs:
                            _pfx_csv = next((p for p in _ret_cands_wm if _pfx_r.lower() in _P(p).name.lower()), None)
                            if not _pfx_csv and _ret_cands_wm:
                                _pfx_csv = _ret_cands_wm[0]
                            if _pfx_csv:
                                _rm_r, _sh_r, _st_r, _slbl_r = _wm_build_ret_info(_pfx_csv)
                                if _rm_r:
                                    _wm_all_ret_maps[_pfx_r] = {'retMap': _rm_r, 'retShots': _sh_r, 'retSiteTotals': _st_r, 'retSiteLabels': _slbl_r}
                        # Also load any remaining reticle CSVs not matched by prefix
                        for _cand in _ret_cands_wm:
                            _cand_matched = any(_pfx_r.lower() in _P(_cand).name.lower() for _pfx_r in _all_pfxs)
                            if not _cand_matched:
                                _rm_r, _sh_r, _st_r, _slbl_r = _wm_build_ret_info(_cand)
                                if _rm_r:
                                    _fn_key = _P(_cand).stem[:8]
                                    _wm_all_ret_maps[_fn_key] = {'retMap': _rm_r, 'retShots': _sh_r, 'retSiteTotals': _st_r, 'retSiteLabels': _slbl_r}
                        # Set defaults from the map that matches actual wafer data,
                        # preferring the one with the most reticle sites (6-die over 4-die)
                        if _wm_all_ret_maps:
                            _matched_infos = [_wm_all_ret_maps[_k] for _k in _all_pfxs if _k in _wm_all_ret_maps]
                            _first_info = max(_matched_infos, key=lambda _v: len(_v.get('retSiteLabels', {}))) if _matched_infos else next(iter(_wm_all_ret_maps.values()))
                            _wm_ret_map = _first_info['retMap']
                            _wm_ret_shots = _first_info['retShots']
                            _wm_ret_site_totals = _first_info['retSiteTotals']
                            _wm_ret_site_labels = _first_info.get('retSiteLabels', {})
                    except Exception:
                        pass
                    # ── IB color map: same palette/MD5 as generate_heatmap_from_csv ──
                    import hashlib as _hashlib_wm
                    _WM_PASS_CLR = {'1': '#00ff44', '2': '#7ddb8a', '3': '#3d3d3d', '4': '#b0b0b0'}
                    _WM_FAIL_PAL = [
                        ('#ff0000',''),('#ff6600',''),('#ff8800',''),('#ffcc00',''),
                        ('#0055ff',''),('#00aaff',''),('#aa00ff',''),('#cc00ff',''),
                        ('#ff0066',''),('#ff33aa',''),('#00bbee',''),('#ff3333',''),
                        ('#6699ff',''),('#cc0099',''),('#ffaa00',''),('#336bff',''),
                        ('#cc0000',''),('#cc4400',''),('#cc9900',''),('#0033cc',''),
                        ('#6600cc',''),('#dd4499',''),('#dd2288',''),('#0099cc',''),
                        ('#ff6666',''),('#ffdd55',''),('#5500cc',''),('#ff5500',''),
                        ('#990000','///'),('#994400','///'),('#cc7700','xxx'),('#003399','xxx'),
                        ('#660099','xxx'),('#005580','+++'),('#990066','+++'),('#003d5c','---'),
                        ('#660000','////'),('#cc3300','////'),('#e6b800','xxxx'),('#000099','xxxx'),
                        ('#330066','++++'),('#7700aa','++++'),('#550000','||||'),('#1a0066','||||'),
                    ]
                    _WM_PAL_N = len(_WM_FAIL_PAL)
                    def _wm_md5f(n_int, salt):
                        h = _hashlib_wm.md5(f'{salt}:{n_int}'.encode()).hexdigest()
                        return int(h[:8], 16) / 0xFFFFFFFF
                    _wm_ib_color_map = dict(_WM_PASS_CLR)
                    _wm_all_fail = sorted(
                        {str(_d[2]) for _wpk, _wdata in _wm_pat_wafers.items()
                         for _d in _wdata['dies'] if _d[2] is not None and str(_d[2]) not in _WM_PASS_CLR},
                        key=lambda s: int(s) if s.isdigit() else 0
                    )
                    _wm_assigned_pal = set()
                    for _wfn in _wm_all_fail:
                        try: _wni = int(_wfn)
                        except ValueError: _wni = hash(_wfn) & 0xFFFF
                        _widx = int(_wm_md5f(_wni, 'color') * _WM_PAL_N) % _WM_PAL_N
                        for _woff in range(_WM_PAL_N):
                            _wcand = (_widx + _woff) % _WM_PAL_N
                            if _wcand not in _wm_assigned_pal:
                                _widx = _wcand; break
                        _wm_assigned_pal.add(_widx)
                        _wm_ib_color_map[_wfn] = _WM_FAIL_PAL[_widx][0]

                    # Build IB-number → short description map from builtin bucket table
                    import re as _re_wm_desc
                    _WM_IB_BUCKETS = [
                        ('1',   'Pass (No-Repair)'), ('2', 'Pass (Repair)'),
                        ('3',   'Good (Atom Defeat)'), ('4', 'Good (Core Defeat)'),
                        ('41/42/47/76/77/81/82', 'Scan'),
                        ('20/21/33/60/61/62/63/65', 'Array MBIST'),
                        ('11/13/16/25/27/28/32/36/39/46/48/51/64/71/74/75', 'Analog'),
                        ('7/8/9/10/15/18/43', 'TPI (Foundry)'),
                        ('31/88/91/94/97/98/99/93', 'TPI (Bump/Test)'),
                        ('19/35', 'Reset'), ('12/44/45/70/80/85/86', 'Functional'),
                        ('26', 'HVQK'),
                    ]
                    _wm_ib_desc = {}
                    for _wm_tok, _wm_lbl in _WM_IB_BUCKETS:
                        for _wm_t in _re_wm_desc.findall(r'\d+', _wm_tok):
                            if _wm_t not in _wm_ib_desc:
                                _wm_ib_desc[_wm_t] = _wm_lbl

                    # Build yieldDefs for WM_PAT — load directly from bucket_json
                    import re as _re_wm_yd, json as _json_wm_yd
                    _wm_yield_defs = []
                    _wm_yt_src = []
                    try:
                        _wm_yt_src = _yt_entries  # available if called from same run
                    except NameError:
                        # load directly from the spec file
                        _wm_yt_path = bucket_json
                        if not _wm_yt_path:
                            # search same dirs as bin-distribution loader
                            import glob as _glob_wm
                            for _sd in [data_csv.parent if hasattr(data_csv,'parent') else None,
                                        Path(out_dir) if out_dir else None]:
                                if not _sd: continue
                                for _pat in ['**/*spec*.json','**/*yield*.json','**/*bucket*.json']:
                                    _cands = list(Path(_sd).glob(_pat))
                                    if _cands: _wm_yt_path = str(_cands[0]); break
                                if _wm_yt_path: break
                        if _wm_yt_path and os.path.isfile(str(_wm_yt_path)):
                            try:
                                _wm_jd = _json_wm_yd.loads(open(str(_wm_yt_path), encoding='utf-8').read())
                                _wm_yt_src = _wm_jd.get('yield_targets', _wm_jd) if isinstance(_wm_jd, dict) else _wm_jd
                            except Exception:
                                pass
                    for _wm_yte in (_wm_yt_src or []):
                        if not isinstance(_wm_yte, dict): continue
                        _bins_f = str(_wm_yte.get('bin',''))
                        _exp_f = str(_wm_yte.get('yield', _wm_yte.get('expected_yield_percent','')))
                        _bucket_f = str(_wm_yte.get('fail_bucket',''))
                        _bins_list = _re_wm_yd.findall(r'\d+', _bins_f)
                        if _bins_f and _exp_f:
                            _wm_yield_defs.append({'bins': _bins_f, 'bucket': _bucket_f, 'expected': _exp_f, 'bins_list': _bins_list})
                    # Load FB descriptions for the inline FB modal
                    _wm_fb_desc = {}
                    try:
                        import json as _json_wm_fbd, re as _re_wm_fbd, csv as _csv_wm_fbd
                        if bucket_json and os.path.isfile(str(bucket_json)):
                            _jd_fbd = _json_wm_fbd.loads(open(str(bucket_json), encoding='utf-8').read())
                            for _e_fbd in (_jd_fbd.get('fB93xx', []) if isinstance(_jd_fbd, dict) else []):
                                if isinstance(_e_fbd, dict) and 'FB' in _e_fbd and 'description' in _e_fbd:
                                    _wm_fb_desc[str(_e_fbd['FB'])] = {'desc': str(_e_fbd['description'])}
                        if not _wm_fb_desc and resolved_csv:
                            for _bm_d2 in [Path(str(resolved_csv)).parent]:
                                for _bm_c2 in list(_bm_d2.glob('*bindef*.csv')) + list(_bm_d2.glob('*_bindef.csv')):
                                    try:
                                        with open(str(_bm_c2), encoding='utf-8') as _bm_f2:
                                            for _bm_r2 in _csv_wm_fbd.reader(_bm_f2):
                                                if len(_bm_r2) >= 2:
                                                    _m_fbd2 = _re_wm_fbd.match(r'^FB(\d+)$', str(_bm_r2[0]).strip(), _re_wm_fbd.IGNORECASE)
                                                    if _m_fbd2: _wm_fb_desc[_m_fbd2.group(1)] = {'desc': str(_bm_r2[1]).strip()}
                                        if _wm_fb_desc: break
                                    except Exception:
                                        pass
                                if _wm_fb_desc: break
                    except Exception:
                        pass
                    _wm_pat_js_data = _json_wm_pat.dumps({
                        'wafers': _wm_pat_wafers,
                        'retMap': _wm_ret_map,
                        'retShots': _wm_ret_shots,
                        'retSiteTotals': _wm_ret_site_totals,
                        'retSiteLabels': _wm_ret_site_labels,
                        'retMaps': _wm_all_ret_maps,
                        'hasReticle': bool(_wm_ret_map),
                        'ibColors': _wm_ib_color_map,
                        'ibDesc': _wm_ib_desc,
                        'yieldDefs': _wm_yield_defs,
                        'fbDescriptions': _wm_fb_desc,
                        'hwFields': _hw_fields_wm,
                        'upmCols': _upm_labels_wm,
                    }, separators=(',', ':'))

                    pk = "'+pk+'"  # JS forEach variable name (used in string concat below)
                    _wm_fb_analyze_js = (
                        'var _wmFbChecked=new Set(),_wmFbKeys=[],_wmFbTotals={},_wmFbIb=null,_wmFbIbTotal=0,_wmFbTotalDies=0,_wmFbPk=null;\n'
                        'function _wmEscH(s){return String(s).replace(/&/g,\'&amp;\').replace(/</g,\'&lt;\').replace(/>/g,\'&gt;\');}\n'
                        'function _wmCloseAnalyze(){var e=document.getElementById(\'wm-analyze-overlay\');if(e)e.remove();}\n'
                        'function _wmAnalyzeBins(bins){\n'
                        '  if(!bins||!bins.length)return;\n'
                        '  var ibCounts={};\n'
                        '  Object.keys(WM_PAT.wafers||{}).forEach(function(pk){\n'
                        '    (WM_PAT.wafers[pk].dies||[]).forEach(function(d){\n'
                        '      if(d[2]!==null&&d[2]!==undefined&&bins.indexOf(d[2])>=0){\n'
                        '        var k=String(d[2]);ibCounts[k]=(ibCounts[k]||0)+1;\n'
                        '      }\n'
                        '    });\n'
                        '  });\n'
                        '  var ibKeys=Object.keys(ibCounts).sort(function(a,b){return ibCounts[b]-ibCounts[a];});\n'
                        '  if(!ibKeys.length){alert(\'No die data for the selected bins\');return;}\n'
                        '  if(ibKeys.length===1){_wmOpenFbModal(+ibKeys[0]);return;}\n'
                        '  var ex=document.getElementById(\'wm-analyze-overlay\');if(ex)ex.remove();\n'
                        '  var ov=document.createElement(\'div\');ov.id=\'wm-analyze-overlay\';\n'
                        '  ov.style.cssText=\'position:fixed;top:0;left:0;right:0;bottom:0;z-index:99999;background:rgba(0,0,0,.55);display:flex;align-items:center;justify-content:center\';\n'
                        '  var box=document.createElement(\'div\');\n'
                        '  box.style.cssText=\'background:#fff;border-radius:8px;box-shadow:0 8px 32px rgba(0,0,0,.3);padding:20px 24px;min-width:260px;max-width:340px;font-family:Arial,sans-serif\';\n'
                        '  var ttl=document.createElement(\'div\');\n'
                        '  ttl.style.cssText=\'font-weight:bold;font-size:14px;color:#2c3e50;margin-bottom:14px\';\n'
                        '  ttl.textContent=\'Select Interface Bin to analyze\';\n'
                        '  box.appendChild(ttl);\n'
                        '  ibKeys.forEach(function(ibk){\n'
                        '    var col=(WM_PAT.ibColors&&WM_PAT.ibColors[ibk])||\'#555\';\n'
                        '    var btn=document.createElement(\'button\');\n'
                        '    btn.style.cssText=\'display:block;width:100%;text-align:left;padding:8px 12px;margin-bottom:6px;border:1px solid \'+col+\';border-radius:4px;background:#fff;cursor:pointer;font-size:13px;font-family:Arial,sans-serif\';\n'
                        '    btn.innerHTML=\'<b style="color:\'+col+\'">IB\'+ibk+\'</b> <span style="color:#888;font-size:12px">\u2014 \'+ibCounts[ibk].toLocaleString()+\' die</span>\';\n'
                        '    btn.onmouseover=function(){this.style.background=\'#f0f4ff\';};btn.onmouseout=function(){this.style.background=\'#fff\';};\n'
                        '    var _ibkN=+ibk;btn.onclick=function(){ov.remove();_wmOpenFbModal(_ibkN);};\n'
                        '    box.appendChild(btn);\n'
                        '  });\n'
                        '  var cancelBtn=document.createElement(\'button\');\n'
                        '  cancelBtn.textContent=\'Cancel\';\n'
                        '  cancelBtn.style.cssText=\'margin-top:4px;padding:5px 12px;border:1px solid #ccc;border-radius:4px;background:#f8f8f8;cursor:pointer;font-size:12px;width:100%;font-family:Arial,sans-serif\';\n'
                        '  cancelBtn.onclick=function(){ov.remove();};\n'
                        '  box.appendChild(cancelBtn);\n'
                        '  ov.appendChild(box);\n'
                        '  ov.addEventListener(\'click\',function(e){if(e.target===ov)ov.remove();});\n'
                        '  document.body.appendChild(ov);\n'
                        '}\n'
                        'function _wmOpenFbModal(ib){\n'
                        '  var fbTotals={};var ibTotal=0;var totalDies=0;\n'
                        '  Object.keys(WM_PAT.wafers||{}).forEach(function(pk){\n'
                        '    var wdata=WM_PAT.wafers[pk];\n'
                        '    totalDies+=(wdata.dies||[]).length;\n'
                        '    var fbMap=(wdata.ibToFb&&wdata.ibToFb[String(ib)])||{};\n'
                        '    Object.keys(fbMap).forEach(function(fb){fbTotals[fb]=(fbTotals[fb]||0)+fbMap[fb];ibTotal+=fbMap[fb];});\n'
                        '  });\n'
                        '  var fbKeys=Object.keys(fbTotals).sort(function(a,b){return fbTotals[b]-fbTotals[a];});\n'
                        '  if(!fbKeys.length){alert(\'No functional bin data for IB\'+ib+\'. Add FUNCTIONAL_BIN column to your data.\');return;}\n'
                        '  _wmFbIb=ib;_wmFbKeys=fbKeys.slice();_wmFbTotals=fbTotals;\n'
                        '  _wmFbChecked=new Set(fbKeys);_wmFbIbTotal=ibTotal;_wmFbTotalDies=totalDies;\n'
                        '  var ex=document.getElementById(\'wm-analyze-overlay\');if(ex)ex.remove();\n'
                        '  var ov=document.createElement(\'div\');ov.id=\'wm-analyze-overlay\';\n'
                        '  ov.style.cssText=\'position:fixed;top:0;left:0;right:0;bottom:0;z-index:99999;background:rgba(0,0,0,.55);display:flex;align-items:center;justify-content:center\';\n'
                        '  var modal=document.createElement(\'div\');\n'
                        '  modal.style.cssText=\'background:#fff;border-radius:8px;box-shadow:0 8px 40px rgba(0,0,0,.35);display:flex;flex-direction:column;width:min(92vw,860px);max-height:90vh;font-family:Arial,sans-serif;overflow:hidden\';\n'
                        '  var hdr=document.createElement(\'div\');\n'
                        '  hdr.style.cssText=\'background:#2c3e50;color:#fff;padding:10px 16px;display:flex;justify-content:space-between;align-items:center;flex-shrink:0\';\n'
                        '  var htitle=document.createElement(\'b\');htitle.id=\'wm-fb-title\';\n'
                        '  htitle.textContent=\'IB\'+ib+\' \u2014 Functional Bin Breakdown\';\n'
                        '  var hclose=document.createElement(\'button\');hclose.innerHTML=\'&times;\';\n'
                        '  hclose.style.cssText=\'background:none;border:none;color:#fff;font-size:22px;cursor:pointer;line-height:1;padding:0 4px\';\n'
                        '  hclose.onclick=function(){ov.remove();};\n'
                        '  hdr.appendChild(htitle);hdr.appendChild(hclose);modal.appendChild(hdr);\n'
                        '  var cw=document.createElement(\'div\');\n'
                        '  cw.style.cssText=\'padding:12px 16px 4px;flex-shrink:0;background:#f8f9fa\';\n'
                        '  var svg=document.createElementNS(\'http://www.w3.org/2000/svg\',\'svg\');\n'
                        '  svg.id=\'wm-fb-svg\';svg.style.cssText=\'width:100%;display:block;overflow:visible\';\n'
                        '  cw.appendChild(svg);modal.appendChild(cw);\n'
                        '  var frow=document.createElement(\'div\');frow.id=\'wm-fb-frow\';\n'
                        '  frow.style.cssText=\'padding:6px 16px;display:flex;flex-wrap:wrap;align-items:center;gap:8px;border-top:1px solid #e8ecf0;flex-shrink:0;background:#fafbfc\';\n'
                        '  var flbl=document.createElement(\'span\');\n'
                        '  flbl.style.cssText=\'font-size:11px;font-weight:bold;color:#555;flex-shrink:0\';\n'
                        '  flbl.textContent=\'Filter by FB:\';\n'
                        '  var allBtn=document.createElement(\'button\');allBtn.textContent=\'All\';\n'
                        '  allBtn.style.cssText=\'font-size:11px;padding:2px 8px;border:1px solid #bbb;border-radius:3px;cursor:pointer;background:#fff\';\n'
                        '  allBtn.onclick=function(){_wmFbChecked=new Set(_wmFbKeys);_wmFbRender();};\n'
                        '  var noneBtn=document.createElement(\'button\');noneBtn.textContent=\'None\';\n'
                        '  noneBtn.style.cssText=allBtn.style.cssText;\n'
                        '  noneBtn.onclick=function(){_wmFbChecked.clear();_wmFbRender();};\n'
                        '  frow.appendChild(flbl);frow.appendChild(allBtn);frow.appendChild(noneBtn);\n'
                        '  modal.appendChild(frow);\n'
                        '  var tw=document.createElement(\'div\');tw.id=\'wm-fb-tblwrap\';\n'
                        '  tw.style.cssText=\'flex:1;overflow-y:auto;min-height:0\';\n'
                        '  modal.appendChild(tw);\n'
                        '  ov.appendChild(modal);\n'
                        '  ov.addEventListener(\'click\',function(e){if(e.target===ov)ov.remove();});\n'
                        '  document.body.appendChild(ov);\n'
                        '  _wmFbRender();\n'
                        '}\n'
                        'function _wmFbRender(){\n'
                        '  var ib=_wmFbIb,fbTotals=_wmFbTotals,fbKeys=_wmFbKeys;\n'
                        '  var ibTotal=_wmFbIbTotal,totalDies=_wmFbTotalDies;\n'
                        '  var fbDesc=WM_PAT.fbDescriptions||{};\n'
                        '  var PAL=[\'#e74c3c\',\'#e67e22\',\'#f39c12\',\'#2ecc71\',\'#1abc9c\',\'#3498db\',\'#9b59b6\',\'#e91e63\',\'#00bcd4\',\'#8bc34a\',\'#ff9800\',\'#795548\'];\n'
                        '  var n=Math.min(fbKeys.length,30);\n'
                        '  var barH=13,gap=3,pl=54,pr=10,pt=8,pb=8,labelW=260,cW=360;\n'
                        '  var H=n*(barH+gap)+pt+pb+20;\n'
                        '  var maxCnt=0;\n'
                        '  fbKeys.slice(0,n).forEach(function(fb){if(_wmFbChecked.has(fb)&&(fbTotals[fb]||0)>maxCnt)maxCnt=fbTotals[fb];});\n'
                        '  if(!maxCnt)fbKeys.slice(0,n).forEach(function(fb){if((fbTotals[fb]||0)>maxCnt)maxCnt=fbTotals[fb];});\n'
                        '  if(!maxCnt)maxCnt=1;\n'
                        '  var p=[];\n'
                        '  p.push(\'<rect width="\'+(pl+cW+labelW+pr)+\'" height="\'+H+\'" fill="#f8f9fa" rx="3"/>\');\n'
                        '  for(var i=0;i<n;i++){\n'
                        '    var fb=fbKeys[i],cnt=fbTotals[fb]||0,pct=ibTotal>0?cnt/ibTotal*100:0;\n'
                        '    var fbd=(fbDesc[fb]&&fbDesc[fb].desc)?fbDesc[fb].desc:\'\';\n'
                        '    var y=pt+gap+(barH+gap)*i;\n'
                        '    var sel=_wmFbChecked.has(fb);\n'
                        '    var clr=sel?PAL[i%PAL.length]:\'#ddd\';\n'
                        '    var bw=sel?Math.max(2,(cnt/maxCnt)*cW):2;\n'
                        '    var fbFailPct=totalDies>0?cnt/totalDies*100:0;\n'
                        '    var txtClr=sel?\'#333\':\'#bbb\';\n'
                        '    var tip=\'FB\'+fb+(fbd?\' \u2014 \'+_wmEscH(fbd.substring(0,40)):\'\')+\': \'+cnt.toLocaleString()+\' (\'+pct.toFixed(1)+\'% IB | \'+fbFailPct.toFixed(2)+\'% fail)\';\n'
                        '    p.push(\'<rect x="\'+pl+\'" y="\'+y.toFixed(1)+\'" width="\'+bw.toFixed(1)+\'" height="\'+barH.toFixed(1)+\'" fill="\'+clr+\'" rx="2"><title>\'+tip+\'</title></rect>\');\n'
                        '    p.push(\'<text x="\'+( pl-4)+\'" y="\'+( y+barH/2+4).toFixed(1)+\'" text-anchor="end" font-family="Arial" font-size="11" fill="\'+txtClr+\'">FB\'+fb+\'</text>\');\n'
                        '    if(sel){\n'
                        '      var lbl=cnt.toLocaleString()+\' (\'+pct.toFixed(1)+\'% IB | \'+fbFailPct.toFixed(2)+\'% fail)\'+(fbd?\' \u2014 \'+_wmEscH(fbd.substring(0,22)):\'\');\n'
                        '      p.push(\'<text x="\'+( pl+bw+5).toFixed(1)+\'" y="\'+( y+barH/2+4).toFixed(1)+\'" font-family="Arial" font-size="10" fill="#555">\'+lbl+\'</text>\');\n'
                        '    }\n'
                        '  }\n'
                        '  if(fbKeys.length>30)p.push(\'<text x="\'+( pl+10)+\'" y="\'+( H-4)+\'" font-family="Arial" font-size="11" fill="#888">\u2026 and \'+(fbKeys.length-30)+\' more</text>\');\n'
                        '  var svg=document.getElementById(\'wm-fb-svg\');\n'
                        '  if(svg){svg.setAttribute(\'viewBox\',\'0 0 \'+(pl+cW+labelW+pr)+\' \'+H);svg.setAttribute(\'height\',H);svg.innerHTML=p.join(\'\');}\n'
                        '  var chkTot=0;fbKeys.forEach(function(fb){if(_wmFbChecked.has(fb))chkTot+=fbTotals[fb]||0;});\n'
                        '  var sfx=(_wmFbChecked.size<fbKeys.length)?\' \u2014 \'+_wmFbChecked.size+\'/\'+fbKeys.length+\' selected\':\'\';\n'
                        '  var tm=document.getElementById(\'wm-fb-title\');\n'
                        '  if(tm)tm.textContent=\'IB\'+ib+\' \u2014 Functional Bin Breakdown (\'+chkTot.toLocaleString()+\' / \'+ibTotal.toLocaleString()+\' die)\'+sfx;\n'
                        '  var frow=document.getElementById(\'wm-fb-frow\');\n'
                        '  if(frow){\n'
                        '    while(frow.children.length>3)frow.removeChild(frow.lastChild);\n'
                        '    fbKeys.forEach(function(fb){\n'
                        '      var lbl=document.createElement(\'label\');\n'
                        '      lbl.style.cssText=\'font-size:11px;display:flex;align-items:center;gap:3px;cursor:pointer;user-select:none\';\n'
                        '      var cb=document.createElement(\'input\');cb.type=\'checkbox\';\n'
                        '      cb.checked=_wmFbChecked.has(fb);cb.style.cursor=\'pointer\';\n'
                        '      (function(fbk,cbi){cbi.onchange=function(){if(cbi.checked)_wmFbChecked.add(fbk);else _wmFbChecked.delete(fbk);_wmFbRender();};})(fb,cb);\n'
                        '      var sp=document.createElement(\'span\');sp.textContent=\'FB\'+fb;\n'
                        '      lbl.appendChild(cb);lbl.appendChild(sp);frow.appendChild(lbl);\n'
                        '    });\n'
                        '  }\n'
                        '  var wrap=document.getElementById(\'wm-fb-tblwrap\');\n'
                        '  if(wrap){\n'
                        '    var tbl=document.createElement(\'table\');\n'
                        '    tbl.style.cssText=\'border-collapse:collapse;width:100%;font-size:12px\';\n'
                        '    var thead=document.createElement(\'thead\');\n'
                        '    var hr=document.createElement(\'tr\');\n'
                        '    hr.style.cssText=\'background:#2c3e50;color:#fff;position:sticky;top:0;z-index:1\';\n'
                        '    [\'Functional Bin\',\'Description\',\'Count\',\'% of IB\',\'Fail %\'].forEach(function(h,hi){\n'
                        '      var th=document.createElement(\'th\');th.textContent=h;\n'
                        '      th.style.cssText=\'padding:6px 10px;text-align:\'+(hi>=2?\'right\':\'left\')+\';white-space:nowrap\';\n'
                        '      hr.appendChild(th);\n'
                        '    });\n'
                        '    thead.appendChild(hr);tbl.appendChild(thead);\n'
                        '    var tbody=document.createElement(\'tbody\');\n'
                        '    fbKeys.forEach(function(fb,idx){\n'
                        '      var cnt=fbTotals[fb]||0,pct=ibTotal>0?cnt/ibTotal*100:0;\n'
                        '      var fbFailPct=totalDies>0?cnt/totalDies*100:0;\n'
                        '      var fbd=(fbDesc[fb]&&fbDesc[fb].desc)?fbDesc[fb].desc:\'\';\n'
                        '      var sel=_wmFbChecked.has(fb);\n'
                        '      var tr=document.createElement(\'tr\');\n'
                        '      tr.style.cssText=\'background:\'+(idx%2?\'#f8f9fa\':\'#fff\')+\';\'+(sel?\'\':\'opacity:0.4\');\n'
                        '      [[\'FB\'+fb,\'bold\',\'#2c3e50\'],[fbd,\'\',\'\'],[cnt.toLocaleString(),\'\',\'\'],[pct.toFixed(1)+\'%\',\'bold\',\'#c0392b\'],[fbFailPct.toFixed(2)+\'%\',\'\',\'#888\']].forEach(function(c,ci){\n'
                        '        var td=document.createElement(\'td\');\n'
                        '        td.style.cssText=\'padding:4px 10px;border-bottom:1px solid #f0f0f0;text-align:\'+(ci>=2?\'right\':\'left\');\n'
                        '        if(c[1])td.style.fontWeight=c[1];\n'
                        '        if(c[2])td.style.color=c[2];\n'
                        '        td.textContent=c[0];tr.appendChild(td);\n'
                        '      });\n'
                        '      tbody.appendChild(tr);\n'
                        '    });\n'
                        '    tbl.appendChild(tbody);wrap.innerHTML=\'\';wrap.appendChild(tbl);\n'
                        '  }\n'
                        '}\n'
                    )
                    _wm_combined_html = (
                        '<!doctype html>\n<html><head><meta charset="utf-8">\n'
                        '<meta name="viewport" content="width=device-width,initial-scale=1">\n'
                        '<style>\n'
                        'html,body{display:flex;flex-direction:column;height:100%;overflow:hidden;margin:0;padding:0;font-family:Arial,sans-serif;font-size:12px}\n'
                        '.wm-hdr{padding:6px 10px;background:#2c3e50;color:#ecf0f1;font-size:13px;font-weight:bold;flex-shrink:0}\n'
                        '.wm-nav{height:160px;overflow-y:auto;flex-shrink:0;border-bottom:2px solid #2c3e50}\n'
                        '.wm-tbl{border-collapse:collapse;width:auto;font-size:12px}\n'
                        '.wm-tbl th{background:#2c3e50;color:#ecf0f1;padding:5px 10px;text-align:left;position:sticky;top:0;z-index:2;white-space:nowrap}\n'
                        '.flt-btn{background:none;border:none;color:#aed6f1;cursor:pointer;font-size:11px;padding:0 0 0 4px;vertical-align:middle;opacity:.85}\n'
                        '.flt-btn:hover{opacity:1;color:#fff}\n'
                        '.flt-btn.active{color:#f1c40f!important;opacity:1}\n'
                        '.dd-panel{position:fixed;background:#fff;border:1px solid #aaa;border-radius:4px;box-shadow:0 4px 16px rgba(0,0,0,.18);z-index:9999;min-width:180px;max-width:280px;font-family:Arial,sans-serif;font-size:12px;color:#2c3e50}\n'
                        '.dd-panel .dd-search{width:100%;box-sizing:border-box;padding:5px 8px;border:none;border-bottom:1px solid #ddd;font-size:12px;outline:none}\n'
                        '.dd-panel .dd-acts{display:flex;gap:4px;padding:4px 6px;border-bottom:1px solid #eee}\n'
                        '.dd-panel .dd-acts button{flex:1;padding:2px 6px;font-size:11px;cursor:pointer;border:1px solid #bdc3c7;background:#ecf0f1;border-radius:3px}\n'
                        '.dd-panel .dd-list{max-height:200px;overflow-y:auto;padding:4px 0}\n'
                        '.dd-panel .dd-item{display:flex;align-items:center;gap:6px;padding:3px 10px;cursor:pointer}\n'
                        '.dd-panel .dd-item:hover{background:#eaf0fb}\n'
                        '.dd-panel .dd-item input{margin:0;cursor:pointer}\n'
                        '.dd-panel .dd-footer{padding:4px 8px;border-top:1px solid #eee;text-align:right}\n'
                        '.dd-panel .dd-footer button{padding:3px 12px;font-size:11px;cursor:pointer;background:#2c3e50;color:#fff;border:none;border-radius:3px}\n'
                        '.wm-tbl td{padding:4px 10px;border-bottom:1px solid #eee;white-space:nowrap}\n'
                        '.wm-lot-row{cursor:pointer;background:#eaf0fb}\n'
                        '.wm-lot-row:hover td{background:#c8dcf5!important}\n'
                        '.wm-lot-row.wm-active td{background:#aaccff!important}\n'
                        '.wm-wafer-row{cursor:pointer}\n'
                        '.wm-wafer-row:hover td{background:#f0f4ff!important}\n'
                        '.wm-wafer-row.wm-active td{background:#ddeeff!important;font-weight:bold}\n'
                        '#wm-frames{flex:1;display:flex;min-height:0;overflow:hidden}\n'
                        '#wm-frames iframe{flex:1;border:none;min-width:0;border-left:2px solid #2c3e50}\n'
                        '#wm-frames iframe:first-child{border-left:none}\n'
                        '.wm-resize{height:6px;background:#e2e8f0;cursor:ns-resize;display:flex;align-items:center;justify-content:center;user-select:none;flex-shrink:0}\n'
                        '.wm-resize::after{content:\'\u2014\';color:#aaa;font-size:10px}\n'
                        '.wm-resize:hover,.wm-resize.dragging{background:#2980b9}\n'
                        '.wm-resize:hover::after,.wm-resize.dragging::after{color:#fff}\n'
                        '#wm-map-bar{display:none;padding:4px 10px;background:#6c3483;flex-shrink:0;align-items:center;justify-content:flex-start;gap:10px;border-bottom:1px solid #512e5f}\n'
                        '#wm-map-bar-info{font-size:11px;color:#aed6f1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}\n'
                        '#wm-pat-btn{font-size:11px;padding:3px 10px;background:#27ae60;color:#fff;border:none;border-radius:3px;cursor:pointer;white-space:nowrap;flex-shrink:0}\n'
                        '#wm-pat-btn:hover{background:#1e8449}\n'
                        '.wm-pat-overlay{display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.45);z-index:23000;pointer-events:none}\n'
                        '.wm-pat-overlay.open{display:block;pointer-events:none}\n'
                        '.wm-pat-box{position:absolute;left:3vw;top:36px;background:#f0f2f5;border-radius:8px;box-shadow:0 8px 32px rgba(0,0,0,.35);width:94vw;max-width:1400px;height:78vh;min-width:640px;min-height:360px;max-height:95vh;display:flex;flex-direction:column;pointer-events:auto;resize:both;overflow:hidden}\n'
                        '.wm-pat-drag{cursor:move;background:#145a32;color:#fff;padding:8px 14px;border-radius:8px 8px 0 0;display:flex;align-items:center;gap:10px;user-select:none;flex-shrink:0}\n'
                        '.wm-pat-body2{display:flex;flex-direction:column;flex:1;padding:8px;gap:6px;min-height:0;overflow:hidden}\n'
                        '.wm-pat-inner2{display:flex;gap:0;flex:1;min-height:0;overflow:auto}\n'
                        '.wm-pat-left{display:flex;flex-direction:column;gap:6px;flex:none;width:55%;min-width:180px;min-height:0;overflow:hidden}\n'
                        '.wm-pat-vsplit{width:6px;cursor:ew-resize;background:#e0e0e0;flex-shrink:0;display:flex;align-items:center;justify-content:center;border-radius:3px;margin:0 1px}\n'
                        '.wm-pat-vsplit:hover{background:#bbb}\n'
                        '.wm-pat-vsplit::after{content:"";width:2px;height:30px;background:#999;border-radius:1px}\n'
                        '.wm-pat-maps-wrap{overflow:auto;background:#fff;border-radius:6px;padding:8px;box-shadow:0 1px 4px rgba(0,0,0,.1);flex:1;min-height:0}\n'
                        '.wm-pat-maps{display:flex;flex-wrap:wrap;gap:10px}\n'
                        '.wm-pat-ltab-bar{display:flex;gap:0;border-bottom:2px solid #d5d8dc;flex-shrink:0;margin-bottom:4px}\n'
                        '.wm-pat-ltab{font-size:11px;padding:3px 10px;cursor:pointer;border:1px solid transparent;border-bottom:none;border-radius:4px 4px 0 0;color:#666;background:none;white-space:nowrap}\n'
                        '.wm-pat-ltab.on{border-color:#d5d8dc;background:#fff;color:#145a32;font-weight:bold;margin-bottom:-2px}\n'
                        '.wm-pat-lpane{display:none;flex:1;min-height:0;overflow:auto}\n'
                        '.wm-pat-lpane.on{display:flex;flex-direction:column}\n'
                        '.wm-pat-right{display:flex;flex-direction:column;gap:4px;flex:1;min-width:200px;min-height:0;overflow:hidden}\n'
                        '.wm-pat-scores{flex:0 0 auto;background:#fff;border-radius:6px;box-shadow:0 1px 4px rgba(0,0,0,.1);display:flex;flex-direction:column;overflow:hidden}\n'
                        '.wm-pat-scores-resize{height:5px;background:#e0e0e0;cursor:ns-resize;flex-shrink:0;border-radius:2px;margin:1px 0;display:flex;align-items:center;justify-content:center}\n'
                        '.wm-pat-scores-resize:hover{background:#bbb}\n'
                        '.wm-pat-scores-resize::after{content:"";width:30px;height:2px;background:#999;border-radius:1px}\n'
                        '.wm-pat-tbl-wrap{overflow:auto;flex:1;min-height:0;padding:4px}\n'
                        '.wm-t{border-collapse:collapse;font-size:11px;width:100%}\n'
                        '.wm-t th{background:#145a32;color:#fff;padding:4px 8px;text-align:left;position:sticky;top:0;z-index:1;white-space:nowrap}\n'
                        '.wm-t td{padding:3px 8px;border-bottom:1px solid #eee}\n'
                        '.wm-t tr:nth-child(even) td{background:#f7f9fc}\n'
                        '.wm-bar-bg{background:#e8e8e8;border-radius:3px;height:8px;width:90px;display:inline-block;vertical-align:middle}\n'
                        '.wm-bar-fg{height:8px;border-radius:3px;display:block}\n'
                        '.wm-impact-row{display:flex;align-items:center;gap:6px;font-size:11px;margin-bottom:2px}\n'
                        '.wm-impact-lbl{width:80px;min-width:80px;text-align:right;font-weight:bold;white-space:nowrap;flex-shrink:0}\n'
                        '.wm-impact-bar{flex:1;background:#e8e8e8;border-radius:3px;height:8px;position:relative}\n'
                        '.wm-impact-fill{height:8px;border-radius:3px;position:absolute;left:0;top:0}\n'
                        '.wm-pat-impact{background:#fff;border-radius:5px;padding:5px 8px;box-shadow:0 1px 3px rgba(0,0,0,.1);flex:1;overflow:auto;min-height:0}\n'
                        '.wm-pat-ctrl{display:flex;align-items:flex-start;gap:8px;flex-shrink:0;padding:2px 0;flex-wrap:wrap}\n'
                        '.wm-pat-filtbar{background:#fff;border:1px solid #d5d8dc;border-radius:5px;padding:3px 8px;display:flex;flex-wrap:wrap;gap:3px 8px;align-items:center;flex:1;max-height:58px;overflow-y:auto}\n'
                        '.wm-pat-wcb{font-size:11px;cursor:pointer;display:flex;align-items:center;gap:2px;padding:1px 4px;border-radius:3px}\n'
                        '.wm-pat-wcb:hover{background:#f0f4fa}\n'
                        '.wm-pat-wcb input{cursor:pointer;margin:0}\n'
                        '.wm-pat-tabs{display:flex;gap:0;border-bottom:2px solid #d5d8dc;flex-shrink:0}\n'
                        '.wm-pat-tab{font-size:11px;padding:4px 12px;cursor:pointer;border:1px solid transparent;border-bottom:none;border-radius:5px 5px 0 0;color:#666;background:none;white-space:nowrap}\n'
                        '.wm-pat-tab.on{border-color:#d5d8dc;background:#fff;color:#145a32;font-weight:bold;margin-bottom:-2px}\n'
                        '.wm-pat-tabpane{display:none;flex:1;min-height:0;overflow:auto}\n'
                        '.wm-pat-tabpane.on{display:flex;flex-direction:column}\n'
                        '.wm-pat-binrow{display:flex;flex-wrap:wrap;gap:3px 6px;font-size:11px;padding:3px 6px;background:#fff;border-radius:5px;box-shadow:0 1px 3px rgba(0,0,0,.1);flex-shrink:0}\n'
                        '.wm-pat-bincb{display:flex;align-items:center;gap:3px;cursor:pointer;padding:1px 3px;border-radius:3px;white-space:nowrap}\n'
                        '.wm-pat-bincb:hover{background:#f0f4fa}\n'
                        '.wm-pat-bincb input{cursor:pointer;margin:0}\n'
                        '.wm-pat-binsw{width:10px;height:10px;border-radius:2px;flex-shrink:0;display:inline-block}\n'
                        '.wm-wlbl{font-size:12px;font-weight:bold;color:#2c3e50;text-align:center;margin-bottom:3px}\n'
                        '.wm-pat-close{background:none;border:none;color:#a9dfbf;cursor:pointer;font-size:20px;line-height:1;padding:0}\n'
                        '.wm-pat-close:hover{color:#fff}\n'
                        '.wm-dd-btn{font-size:11px;padding:2px 9px;background:rgba(255,255,255,0.15);color:#fff;border:1px solid rgba(255,255,255,0.3);border-radius:4px;cursor:pointer;white-space:nowrap;flex-shrink:0}\n'
                        '.wm-dd-btn:hover,.wm-dd-btn.open{background:rgba(255,255,255,0.28)}\n'
                        '#wm-dd-panel{position:fixed;z-index:30000;background:#fff;border:1px solid #ccc;border-radius:6px;box-shadow:0 4px 18px rgba(0,0,0,.28);min-width:220px;max-width:320px;max-height:360px;display:flex;flex-direction:column;overflow:hidden;font-family:Arial,sans-serif}\n'
                        '.wm-dd-search{padding:6px 10px;border:none;border-bottom:1px solid #e0e0e0;font-size:12px;outline:none;flex-shrink:0;width:100%;box-sizing:border-box}\n'
                        '.wm-dd-acts{display:flex;gap:10px;padding:3px 10px;border-bottom:1px solid #eee;background:#f7f9fc;flex-shrink:0}\n'
                        '.wm-dd-acts span{font-size:11px;color:#2471a3;cursor:pointer;text-decoration:underline}\n'
                        '.wm-dd-scroll{overflow-y:auto;flex:1;min-height:0}\n'
                        '.wm-dd-item{display:flex;align-items:center;gap:6px;padding:4px 10px;font-size:12px;cursor:pointer;user-select:none}\n'
                        '.wm-dd-item:hover{background:#f0f4fa}\n'
                        '.wm-dd-item input{cursor:pointer;margin:0;flex-shrink:0}\n'
                        '.wm-dd-lg{border-bottom:1px solid #f0f0f0}\n'
                        '.wm-dd-lhdr{display:flex;align-items:center;gap:4px;padding:5px 8px;cursor:pointer;font-size:12px;background:#f7f9fc;user-select:none}\n'
                        '.wm-dd-lhdr:hover{background:#e8f0fe}\n'
                        '.wm-dd-larr{font-size:10px;color:#666;width:12px;flex-shrink:0}\n'
                        '.wm-dd-lc{padding-left:16px}\n'
                        '.wm-dd-wi{display:flex;align-items:center;gap:5px;padding:3px 8px;font-size:11px;cursor:pointer;user-select:none}\n'
                        '.wm-dd-wi:hover{background:#f0f4fa}\n'
                        '.wm-dd-wi input{cursor:pointer;margin:0;flex-shrink:0}\n'
                        '</style></head><body>\n'
                        '<div class="wm-hdr">'
                        '<svg width="13" height="13" viewBox="0 0 24 24" fill="currentColor" style="vertical-align:middle;margin-right:6px"><path d="M10 18h4v-2h-4v2zM3 6v2h18V6H3zm3 7h12v-2H6v2z"/></svg>'
                        'IBIN Wafer Map'
                        ' <span style="font-size:11px;font-weight:normal;opacity:.75">(click to view \u2014 Ctrl+click to compare side-by-side)</span>'
                        '</div>\n'
                        '<div class="wm-nav">\n'
                        '<table class="wm-tbl"><thead><tr>'
                        '<th>Test Program <button class="flt-btn" id="wmft-fb-0" onclick="event.stopPropagation();wmFtDdOpen(0,this)" title="Filter">&#9660;</button></th>'
                        '<th>Lot <button class="flt-btn" id="wmft-fb-1" onclick="event.stopPropagation();wmFtDdOpen(1,this)" title="Filter">&#9660;</button></th>'
                        '<th>Wafer <button class="flt-btn" id="wmft-fb-2" onclick="event.stopPropagation();wmFtDdOpen(2,this)" title="Filter">&#9660;</button></th>'
                        + ('<th>Material <button class="flt-btn" id="wmft-fb-3" onclick="event.stopPropagation();wmFtDdOpen(3,this)" title="Filter">&#9660;</button></th>' if _mat_wm else '')
                        + '<th>Die Count</th></tr>'
                        + '</thead>\n'
                        + '<tbody id="wm-tbody">' + _wm_nav_rows + '</tbody></table>\n'
                        '</div>\n'
                        '<div class="wm-resize" id="wm-nav-resize"></div>\n'
                        '<div id="wm-map-bar">'
                        '<button id="wm-pat-btn" onclick="wmOpenPat()">&#128202; Wafer Pattern Analysis</button>'
                        '<span id="wm-map-bar-info"></span>'
                        '</div>\n'
                        '<div id="wm-frames"></div>\n'
                        '<div class="wm-pat-overlay" id="wm-pat-overlay">\n'
                        '  <div class="wm-pat-box" id="wm-pat-box" style="position:relative">\n'
                        '    <div class="wm-pat-drag" id="wm-pat-drag">\n'
                        '      <div id="wm-dd-btn-wrap" style="display:flex;gap:5px;align-items:center;flex-shrink:0">\n'
                        '        <button class="wm-dd-btn" id="wm-dd-prog-btn" style="display:none" onclick="_wmDdOpen(\'prog\',this)">Programs &#9660;</button>\n'
                        '        <button class="wm-dd-btn" id="wm-dd-lotwafer-btn" onclick="_wmDdOpen(\'lotwafer\',this)">Lots/Wafers &#9660;</button>\n'
                        '        <label id="wm-criteria-miss-lbl" style="display:inline-flex;align-items:center;gap:4px;font-size:11px;cursor:pointer;white-space:nowrap;border:1px solid rgba(255,255,255,0.4);border-radius:12px;padding:2px 9px;color:rgba(255,255,255,0.85);background:rgba(255,255,255,0.12);margin-left:4px" title="Show only wafers where at least one bin does not meet its expected yield target"><input type="checkbox" id="wm-criteria-miss-chk" onchange="_wmToggleCriteriaMiss(this.checked)" style="cursor:pointer;margin:0">&#9888; Wafers missing yield target</label>\n'
                        '        <button id="wm-criteria-cfg-btn" onclick="_wmShowCriteriaCfg()" title="Configure which yield targets to check" style="font-size:11px;padding:2px 8px;border-radius:10px;border:1px solid rgba(255,255,255,0.4);background:rgba(255,255,255,0.12);color:rgba(255,255,255,0.85);cursor:pointer;white-space:nowrap">&#9881; Criteria</button>\n'
                        +(
                        '        <label id="wm-upm-overlay-lbl" style="display:inline-flex;align-items:center;gap:4px;font-size:11px;cursor:pointer;white-space:nowrap;border:1px solid rgba(255,255,255,0.4);border-radius:12px;padding:2px 9px;color:rgba(255,255,255,0.85);background:rgba(255,255,255,0.12);margin-left:4px" title="Overlay UPM heatmap: fill = UPM%, opacity = pass/fail"><input type="checkbox" id="wm-upm-overlay-chk" onchange="_wmToggleUpmOverlay(this.checked)" style="cursor:pointer;margin:0">&#127777; Overlay UPM</label>\n'
                        if _upm_labels_wm else '')
                        +'      </div>\n'
                        '      <span style="flex:1"></span>\n'
                        '      <button id="wm-pat-mode-btn" onclick="_wmPatToggleCanvasMode()" title="Switch to Canvas for fast interactive debug" style="background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.4);color:#fff;font-size:11px;cursor:pointer;padding:2px 9px;border-radius:4px;margin-right:8px">&#128247; SVG mode</button>\n'
                        '      <div style="font-size:9px;color:rgba(255,255,255,0.75);font-family:Arial,sans-serif;pointer-events:none;user-select:none;letter-spacing:0.03em;white-space:nowrap">Pant, Sujit N \u2014 GEMS FTE</div>\n'
                        '      <button class="wm-pat-close" onclick="wmHidePat()">&times;</button>\n'
                        '    </div>\n'
                        '    <div id="wm-pat-prog-picker" style="display:none"></div>\n'
                        '    <div id="wm-pat-lot-picker" style="display:none"></div>\n'
                        '    <div id="wm-pat-wafer-picker" style="display:none"></div>\n'
                        '    <div class="wm-pat-body2">\n'
                        '      <div class="wm-pat-ctrl" id="wm-pat-ctrl"></div>\n'
                        '      <div class="wm-pat-inner2">\n'
                        '        <div class="wm-pat-left">\n'
                        '          <div class="wm-pat-ltab-bar">\n'
                        '            <button class="wm-pat-ltab on" data-ltab="wafers" onclick="wmPatLTab(this.dataset.ltab)">&#128447; Wafer Maps</button>\n'
                        '            <button class="wm-pat-ltab" data-ltab="composite" onclick="wmPatLTab(this.dataset.ltab)">&#128300; Composite Map</button>\n'
                        '          </div>\n'
                        '          <div class="wm-pat-lpane on" id="wm-pat-lpane-wafers">\n'
                        '            <div id="wm-upm-legend" style="display:none;flex-direction:column;padding:6px 10px 4px;background:#1e2733;border-bottom:1px solid #34495e;color:#ecf0f1;user-select:none"></div>\n'
                        '            <div class="wm-pat-maps-wrap"><div class="wm-pat-maps" id="wm-pat-maps"></div>\n'
                        '            </div>\n'
                        '          </div>\n'
                        '          <div class="wm-pat-lpane" id="wm-pat-lpane-composite" style="overflow:auto;padding:4px">\n'
                        '            <div id="wm-pat-modemap-body"><span style="color:#aaa;font-size:11px">Select a lot to view composite map</span></div>\n'
                        '          </div>\n'
                        '        </div>\n'
                        '        <div class="wm-pat-vsplit" id="wm-pat-vsplit"></div>\n'
                        '        <div class="wm-pat-right">\n'
                        '          <div class="wm-pat-tabs" id="wm-pat-tabs">\n'
                        '            <button class="wm-pat-tab on" id="wm-pat-tab-impact" data-tab="impact" onclick="wmPatTab(this.dataset.tab)">&#128269; Bin Impact</button>\n'
                        '            <button class="wm-pat-tab" id="wm-pat-tab-composite2" data-tab="composite2" onclick="wmPatTab(this.dataset.tab)">&#128300; Composite Map</button>\n'
                        +('            <button class="wm-pat-tab" id="wm-pat-tab-reticle" data-tab="reticle" onclick="wmPatTab(this.dataset.tab)">&#127760; Reticle</button>\n' if _wm_ret_map else '')
                        +'            <button class="wm-pat-tab" id="wm-pat-tab-guide" data-tab="guide" onclick="wmPatTab(this.dataset.tab)">&#8505; Guide</button>\n'
                        +'          </div>\n'
                        '          <div class="wm-pat-tabpane on" id="wm-pat-pane-impact">\n'
                        '            <div class="wm-pat-impact" id="wm-pat-impact-body"><span style="color:#aaa;font-size:11px">No data yet</span></div>\n'
                        '          </div>\n'
                        '          <div class="wm-pat-tabpane" id="wm-pat-pane-composite2" style="overflow:auto;padding:4px">\n'
                        '            <div id="wm-pat-modemap-body2"><span style="color:#aaa;font-size:11px">Select a lot to view composite map</span></div>\n'
                        '          </div>\n'
                        '          <div class="wm-pat-tabpane" id="wm-pat-pane-guide">\n'
                        '            <div style="padding:8px;overflow-y:auto;font-size:11px;flex:1"><table style="border-collapse:collapse;width:100%;font-size:11px"><thead><tr><th style="background:#145a32;color:#fff;padding:4px 8px;text-align:left;white-space:nowrap">Pattern</th><th style="background:#145a32;color:#fff;padding:4px 8px;text-align:left;">What it means</th><th style="background:#145a32;color:#fff;padding:4px 8px;text-align:left;">Typical process suspects</th></tr></thead><tbody><tr><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-weight:bold;color:#c0392b;white-space:nowrap">&#11044; CENTER</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;">Failures at wafer center. Center-hot or center-cold non-uniformity.</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-size:10px;line-height:1.6"><b>CMP:</b> center dishing/over-polish, slurry<br><b>Dep:</b> showerhead center zone drift<br><b>Etch:</b> center etch-rate bias, plasma peak<br><b>Implant:</b> beam centering drift</td></tr><tr style="background:#f7f9fc"><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-weight:bold;color:#e67e22;white-space:nowrap">&#11044; EDGE</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;">Failures at the wafer periphery (&gt;75% radius). Edge-boundary effects.</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-size:10px;line-height:1.6"><b>CMP:</b> edge fast-polish, retaining ring wear<br><b>Dep:</b> edge gas-flow boundary, film rolloff<br><b>Etch:</b> edge non-uniformity, chuck gap<br><b>Wafer:</b> edge chips, dicing contamination</td></tr><tr><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-weight:bold;color:#8e44ad;white-space:nowrap">&#11044; DONUT</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;">Ring of failures at mid-radius (40&ndash;70%). Annular non-uniformity.</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-size:10px;line-height:1.6"><b>CMP:</b> multi-zone carrier head ring<br><b>Dep:</b> showerhead mid-radius flow ring<br><b>Etch:</b> plasma standing wave at mid-radius<br><b>Spin:</b> solvent evaporation ring (Marangoni)</td></tr><tr style="background:#f7f9fc"><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-weight:bold;color:#2471a3;white-space:nowrap">&#11044; SYSTEMATIC</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;">Failures repeat at same die XY across wafers, tiling with reticle pitch.</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-size:10px;line-height:1.6"><b>Litho/mask:</b> mask particle, OPC error<br><b>Design:</b> antenna rule, density DRC near limit<br><b>CMP:</b> dishing at dense array, pad groove artifact<br><b>Etch:</b> micro-loading at pattern density</td></tr><tr><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-weight:bold;color:#1f618d;white-space:nowrap">&#11044; RETICLE</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;">Same die within reticle field fails across majority of shots. Mask-born defect.</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-size:10px;line-height:1.6"><b>Litho/mask:</b> reticle particle, chrome/phase defect<br><b>OPC:</b> insufficient correction on critical edge<br><b>Scanner:</b> lens aberration at fixed field position<br><b>Haze:</b> crystalline growth on absorber</td></tr><tr style="background:#f7f9fc"><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-weight:bold;color:#27ae60;white-space:nowrap">&#11044; RANDOM</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;">No spatial pattern. Failures scattered across the wafer.</td><td style="padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;font-size:10px;line-height:1.6"><b>Particles:</b> tool-born particles, ambient contamination<br><b>Crystal:</b> dislocations, stacking faults, EPI hillocks<br><b>Equipment:</b> arcing, ESD, pressure spike<br><b>RDF:</b> intrinsic statistical variability at advanced nodes</td></tr></tbody></table><div style="margin-top:8px;padding:4px 8px;background:#fafafa;border:1px solid #e4e4e4;border-radius:4px;font-size:10px;color:#666;line-height:1.7"><b>Score:</b> 0&ndash;100% measures spatial enrichment over random expectation using 6 radial bands: inner-core (&lt;0.15r), center (0.15&ndash;0.40r), mid (0.40&ndash;0.60r), outer-mid (0.60&ndash;0.75r), edge (0.75&ndash;0.90r), edge-ring (&gt;0.90r). Scores are not mutually exclusive. <b>Confidence:</b> HIGH (n&ge;50) / MEDIUM (20&ndash;49) / LOW (&lt;20 dies &mdash; interpret with caution). <b>Primary</b>&thinsp;=&thinsp;highest score. <b>Driver IB</b>&thinsp;=&thinsp;bin with most fail dies. <b>Lot Trend</b>&thinsp;=&thinsp;avg scores across wafers per lot.</div></div>\n'
                        '          </div>\n'
                        +('          <div class="wm-pat-tabpane" id="wm-pat-pane-reticle">\n'
                          '            <div id="wm-pat-reticle-body" style="padding:8px;overflow:auto;font-size:11px;flex:1"><span style="color:#aaa;font-size:11px">Select a lot to view reticle analysis</span></div>\n'
                          '          </div>\n' if _wm_ret_map else '')
                        +'          <div class="wm-pat-binrow" id="wm-pat-binrow"></div>\n'
                        '          <div id="wm-pat-fbrow" style="display:none;flex-wrap:wrap;gap:3px 6px;font-size:11px;padding:3px 6px 3px 18px;background:#f0f4fa;border-radius:5px;box-shadow:0 1px 3px rgba(0,0,0,.08);flex-shrink:0;border-left:3px solid #2471a3"></div>\n'
                        '          <div class="wm-pat-binrow" id="wm-pat-retrow" style="display:none"></div>\n'
                        '          <div class="wm-pat-binrow" id="wm-pat-shotrow" style="display:none"></div>\n'
                        '        </div>\n'
                        '      </div>\n'
                        '      <div class="wm-pat-scores-resize" id="wm-pat-scores-resize"></div>\n'
                        '      <div class="wm-pat-scores" id="wm-pat-scores-panel" style="height:240px">\n'
                        '        <div style="background:#145a32;color:#fff;font-size:11px;font-weight:bold;padding:4px 10px;flex-shrink:0">&#128202; Pattern Scores</div>\n'
                        '        <div style="flex-shrink:0;border-bottom:2px solid #1a5276;background:#eaf4fb"><div id="wm-pat-lot-trend" style="overflow-x:auto;max-height:72px;padding:2px 4px"></div></div>\n'
                        '        <div class="wm-pat-tbl-wrap" style="flex:1;min-height:0"><table class="wm-t"><thead><tr>\n'
                        '          <th>Lot</th><th>Wafer</th><th>Material</th><th>Primary</th><th>Conf.</th><th>Fail%</th><th>Driver IB</th>\n'
                        '          <th>Center</th><th>Edge</th><th>Donut</th><th>Systematic</th>'
                        +('<th>Reticle</th>' if _wm_ret_map else '')+\
                        ('<th>Top Die Loc</th>' if _wm_ret_map else '')+'\n'
                        '          <th>Random</th>\n'
                        '        </tr></thead><tbody id="wm-pat-tbody"></tbody></table></div>\n'
                        '      </div>\n'
                        '    </div>\n'
                        '  </div>\n'
                        '<div id="wm-dd-panel" style="display:none;position:fixed;z-index:30000;background:#fff;border:1px solid #ccc;border-radius:6px;box-shadow:0 4px 18px rgba(0,0,0,.28);min-width:220px;max-width:320px;max-height:360px;flex-direction:column;overflow:hidden;font-family:Arial,sans-serif">\n'
                        '  <input class="wm-dd-search" placeholder="Search\u2026" oninput="_wmDdSearch(this.value)">\n'
                        '  <div class="wm-dd-scroll" id="wm-dd-body"></div>\n'
                        '</div>\n'
                        '</div>\n'
                        '<script>\n'
                        f'</script>\n'
                        f'<script type="application/json" id="wm-pat-json">{_wm_pat_js_data}</script>\n'
                        '<script>\n'
                        'var WM_PAT=JSON.parse(document.getElementById(\'wm-pat-json\').textContent);\n'
                        f'var BIN_DIST_HTML={repr(bin_html.name if bin_html.exists() else "")};\n'
                        'var _wmSel=new Map();\n'
                        'var _wmFtDdState={};\n'
                        'var _wmFtDdOpen_=null;\n'
                        'var _wmCurPatkey=null;\n'
                        'var _wmCurLot=null;\n'
                        'function wmFtDdOpen(col,btn){\n'
                        '  if(_wmFtDdOpen_){_wmFtDdClose();}\n'
                        '  var rows=document.querySelectorAll("#wm-tbody tr");\n'
                        '  var allVals=[];var seen=new Set();\n'
                        '  rows.forEach(function(tr){\n'
                        '    var tds=tr.querySelectorAll("td");\n'
                        '    var v=tds[col]?tds[col].textContent.trim():"";\n'
                        '    if(!seen.has(v)){seen.add(v);allVals.push(v);}\n'
                        '  });\n'
                        '  allVals.sort(function(a,b){return a.localeCompare(b);});\n'
                        '  var allowed=_wmFtDdState[col];\n'
                        '  var checked=allowed?new Set(allowed):new Set(allVals);\n'
                        '  var panel=document.createElement("div");panel.className="dd-panel";\n'
                        '  panel.innerHTML=\'<input class="dd-search" placeholder="Search\u2026">\'+'
                        '\'<div class="dd-acts"><button onclick="wmFtSelAll()">Select All</button><button onclick="wmFtClearAll()">Clear</button></div>\'+'
                        '\'<div class="dd-list" id="wmft-dd-list"></div>\'+'
                        '\'<div class="dd-footer"><button onclick="wmFtApply()">OK</button></div>\';\n'
                        '  document.body.appendChild(panel);\n'
                        '  panel.querySelector(".dd-search").oninput=function(){var q=(this.value||"").toLowerCase();var fl=q?_wmFtDdOpen_.allVals.filter(function(v){return v.toLowerCase().indexOf(q)>=0;}):_wmFtDdOpen_.allVals;wmFtRenderList(fl);};\n'
                        '  var r=btn.getBoundingClientRect();\n'
                        '  panel.style.top=(r.bottom+2)+"px";\n'
                        '  panel.style.left=Math.min(r.left,window.innerWidth-200)+"px";\n'
                        '  _wmFtDdOpen_={panel:panel,col:col,btn:btn,allVals:allVals,checked:checked};\n'
                        '  wmFtRenderList(allVals);\n'
                        '  setTimeout(function(){document.addEventListener("mousedown",wmFtOutside);},0);\n'
                        '}\n'
                        'function wmFtRenderList(vals){\n'
                        '  var list=document.getElementById("wmft-dd-list");if(!list)return;\n'
                        '  var h="";vals.forEach(function(v){var c=_wmFtDdOpen_&&_wmFtDdOpen_.checked.has(v)?" checked":"";\n'
                        '    h+=\'<label class="dd-item"><input type="checkbox"\'+c+\' data-val="\'+v.replace(/&/g,"&amp;").replace(/"/g,"&quot;")+\'"> \'+v.replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;")+\'</label>\';\n'
                        '  });list.innerHTML=h;\n'
                        '  list.querySelectorAll("input").forEach(function(inp){inp.onchange=function(){wmFtToggle(inp,inp.dataset.val);};});\n'
                        '}\n'
                        'function wmFtToggle(cb,v){if(_wmFtDdOpen_){if(cb.checked)_wmFtDdOpen_.checked.add(v);else _wmFtDdOpen_.checked.delete(v);}}\n'
                        'function wmFtSelAll(){if(_wmFtDdOpen_){_wmFtDdOpen_.allVals.forEach(function(v){_wmFtDdOpen_.checked.add(v);});wmFtRenderList(_wmFtDdOpen_.allVals);}}\n'
                        'function wmFtClearAll(){if(_wmFtDdOpen_){_wmFtDdOpen_.checked.clear();wmFtRenderList(_wmFtDdOpen_.allVals);}}\n'
                        'function wmFtApply(){\n'
                        '  if(!_wmFtDdOpen_)return;\n'
                        '  var c=_wmFtDdOpen_.col,chk=_wmFtDdOpen_.checked,all=_wmFtDdOpen_.allVals;\n'
                        '  _wmFtDdState[c]=(chk.size===all.length)?null:new Set(chk);\n'
                        '  var b=document.getElementById("wmft-fb-"+c);if(b)b.classList.toggle("active",!!_wmFtDdState[c]);\n'
                        '  _wmFtDdClose();\n'
                        '  wmApplyFilter();\n'
                        '}\n'
                        'function _wmFtDdClose(){\n'
                        '  if(!_wmFtDdOpen_)return;\n'
                        '  document.removeEventListener("mousedown",wmFtOutside);\n'
                        '  if(_wmFtDdOpen_.panel.parentNode)_wmFtDdOpen_.panel.parentNode.removeChild(_wmFtDdOpen_.panel);\n'
                        '  _wmFtDdOpen_=null;\n'
                        '}\n'
                        'function wmFtOutside(e){if(_wmFtDdOpen_&&!_wmFtDdOpen_.panel.contains(e.target)){wmFtApply();}}\n'
                        'function wmApplyFilter(){\n'
                        '  var rows=document.querySelectorAll("#wm-tbody tr");\n'
                        '  var firstVisible=null;\n'
                        '  rows.forEach(function(tr){\n'
                        '    var tds=tr.querySelectorAll("td");\n'
                        '    var cols=Array.prototype.map.call(tds,function(td){return td.textContent.trim();});\n'
                        '    var show=Object.keys(_wmFtDdState).every(function(ci){\n'
                        '      var s=_wmFtDdState[ci];return !s||s.has(cols[parseInt(ci)]||"");\n'
                        '    });\n'
                        '    tr.style.display=show?"":"none";\n'
                        '    if(show&&!firstVisible)firstVisible=tr;\n'
                        '  });\n'
                        '  // Auto-navigate to first visible row so filter immediately changes the map\n'
                        '  if(firstVisible)firstVisible.click();\n'
                        '}\n'
                        '/* ── Pattern analysis functions ── */\n'
                        'function _wmIbColor(ib){\n'
                        '  if(ib===null||ib===undefined)return"#e0e0e0";\n'
                        '  var c=WM_PAT.ibColors&&WM_PAT.ibColors[String(parseInt(ib))];\n'
                        '  return c||"#aaaaaa";\n'
                        '}\n'
                        'var _wmFailThr=3;var _wmEdgeExcRows=0;var _wmCriteriaMissOnly=false;var _wmCriteriaDisabled=new Set();\n'
                        'var _wmUpmOverlay=false;var _wmUpmDistOpen=false;\n'
                        'function _wmToggleUpmOverlay(on){_wmUpmOverlay=on;if(!on){_wmUpmDistOpen=false;}  _wmPatRender();}\n'
                        'function _wmToggleUpmDist(){\n'
                        '  _wmUpmDistOpen=!_wmUpmDistOpen;\n'
                        '  var btn=document.getElementById("wm-upm-dist-btn");\n'
                        '  if(btn)btn.style.background=_wmUpmDistOpen?"rgba(255,255,255,0.25)":"rgba(255,255,255,0.1)";\n'
                        '  _wmRenderUpmDist();\n'
                        '}\n'
                        'function _wmRenderUpmDist(){\n'
                        '  var el=document.getElementById("wm-upm-dist-chart");if(!el)return;\n'
                        '  if(!_wmUpmDistOpen){el.style.display="none";return;}\n'
                        '  var _upmEntry=WM_PAT.upmCols&&WM_PAT.upmCols[0];\n'
                        '  var _upmTgt=Array.isArray(_upmEntry)?_upmEntry[1]:null;\n'
                        '  if(!_upmTgt){el.style.display="none";return;}\n'
                        '  var bins=new Array(30).fill(0);var pctMin=80,pctMax=115;\n'
                        '  var vals=[];\n'
                        '  Object.keys(WM_PAT.wafers||{}).forEach(function(pk){\n'
                        '    var w=WM_PAT.wafers[pk];if(!w||!w.dies)return;\n'
                        '    w.dies.forEach(function(d){if(d[0]===null)return;var v=d[4];if(v===null||v===undefined)return;var pct=v/_upmTgt*100;vals.push(pct);var bi=Math.floor((pct-pctMin)/(pctMax-pctMin)*bins.length);if(bi>=0&&bi<bins.length)bins[bi]++;});\n'
                        '  });\n'
                        '  if(!vals.length){el.style.display="none";return;}\n'
                        '  var maxBin=Math.max.apply(null,bins)||1;\n'
                        '  var bw=8,bGap=1,chartW=bins.length*(bw+bGap),chartH=60;\n'
                        '  var bars=bins.map(function(cnt,i){\n'
                        '    var pct=pctMin+(i+0.5)*(pctMax-pctMin)/bins.length;\n'
                        '    var h=Math.round(cnt/maxBin*chartH);if(h<1&&cnt>0)h=1;\n'
                        '    var col=_wmUpmColor(pct);\n'
                        '    var x=i*(bw+bGap),y=chartH-h;\n'
                        '    return\'<rect x="\'+x+\'" y="\'+y+\'" width="\'+bw+\'" height="\'+h+\'" fill="\'+col+\'" opacity="0.9"><title>\'+pct.toFixed(1)+\'% (n=\'+cnt+\')</title></rect>\';\n'
                        '  }).join("");\n'
                        '  var tickPcts=[90,93,97,100];var tickSvg=tickPcts.map(function(tp){\n'
                        '    var tx=Math.round((tp-pctMin)/(pctMax-pctMin)*chartW);\n'
                        '    return\'<line x1="\'+tx+\'" y1="0" x2="\'+tx+\'" y2="\'+chartH+\'" stroke="rgba(255,255,255,0.3)" stroke-width="1" stroke-dasharray="2,2"/>\'\n'
                        '      +\'<text x="\'+tx+\'" y="\'+( chartH+10)+\'" text-anchor="middle" font-size="8" fill="rgba(255,255,255,0.65)">\'+tp+\'%</text>\';\n'
                        '  }).join("");\n'
                        '  var med=vals.slice().sort(function(a,b){return a-b;})[Math.floor(vals.length/2)];\n'
                        '  var avg=vals.reduce(function(s,v){return s+v;},0)/vals.length;\n'
                        '  var mx=Math.round((med-pctMin)/(pctMax-pctMin)*chartW);\n'
                        '  var ax=Math.round((avg-pctMin)/(pctMax-pctMin)*chartW);\n'
                        '  var markers=\'<line x1="\'+mx+\'" y1="0" x2="\'+mx+\'" y2="\'+chartH+\'" stroke="#fff" stroke-width="1.5"/>\'\n'
                        '    +\'<line x1="\'+ax+\'" y1="0" x2="\'+ax+\'" y2="\'+chartH+\'" stroke="#ffee00" stroke-width="1.5" stroke-dasharray="3,2"/>\';\n'
                        '  var svgH=chartH+16;\n'
                        '  el.style.display="block";\n'
                        '  el.innerHTML=\'<div style="font-size:10px;color:rgba(255,255,255,0.7);margin-bottom:3px">Distribution (n=\'+vals.length+\' dies) &nbsp;<span style="color:#fff">&#9474; median=\'+med.toFixed(1)+\'%</span> &nbsp;<span style="color:#ffee00">&#9135;&#9135; avg=\'+avg.toFixed(1)+\'%</span></div>\'\n'
                        '    +\'<svg width="\'+chartW+\'" height="\'+svgH+\'" style="display:block;overflow:visible">\'+bars+tickSvg+markers+\'</svg>\';\n'
                        '}\n'
                        'function _wmUpmColor(pct){\n'
                        '  // pct range 90-100%; green->blue->orange->brown\n'
                        '  var t=Math.max(0,Math.min(1,(pct-90)/(100-90)));\n'
                        '  if(t<0.333){var s=t/0.333;return"rgb(0,"+Math.round(220-s*140)+","+Math.round(s*255)+")";} // green->blue\n'
                        '  if(t<0.667){var s=(t-0.333)/0.334;return"rgb("+Math.round(s*255)+","+Math.round(80+s*80)+","+Math.round(255-s*255)+")";} // blue->orange\n'
                        '  var s=(t-0.667)/0.333;return"rgb("+Math.round(255-s*85)+","+Math.round(160-s*110)+",0)"; // orange->brown\n'
                        '}\n'
                        'function _wmIsFail(ib){if(ib===null||ib===undefined)return false;var n=parseInt(ib);return n>=_wmFailThr;}\n'
                        'function _wmCritRebuildRows(){\n'
                        '  var defs=WM_PAT.yieldDefs||[];\n'
                        '  var tbody=document.getElementById("wm-cc-tbody");if(!tbody)return;\n'
                        '  tbody.innerHTML=defs.map(function(def,i){\n'
                        '    var dis=_wmCriteriaDisabled.has(i);\n'
                        '    return "<tr style=\'border-bottom:1px solid #eee\'><td style=\'padding:4px 8px\'><input type=\'checkbox\' id=\'wm-cc-"+i+"\' "+(dis?"":"checked")+" onchange=\'_wmCriteriaCfgToggle("+i+",this.checked)\' style=\'cursor:pointer\'></td>"\n'
                        '      +"<td style=\'padding:4px 8px;font-size:12px\'>"+(def.bucket||"\u2014")+"</td>"\n'
                        '      +"<td style=\'padding:4px 8px;font-size:11px;color:#555\'>IB"+(def.bins||def.bins_list.join("/"))+"</td>"\n'
                        '      +"<td style=\'padding:4px 8px;font-size:12px;font-weight:bold;color:#2471a3\'>"+(def.expected||"\u2014")+"%</td></tr>";\n'
                        '  }).join("")||\'<tr><td colspan="4" style="padding:10px;color:#888;text-align:center">No criteria loaded</td></tr>\';\n'
                        '  _wmCritCfgUpdateCount();\n'
                        '}\n'
                        'function _wmCritLoadJson(file){\n'
                        '  if(!file)return;\n'
                        '  var reader=new FileReader();\n'
                        '  reader.onload=function(e){\n'
                        '    try{\n'
                        '      var obj=JSON.parse(e.target.result);\n'
                        '      var targets=obj.yield_targets||obj.yieldTargets||obj;\n'
                        '      if(!Array.isArray(targets))throw new Error("Expected array under yield_targets");\n'
                        '      var newDefs=targets.map(function(t){\n'
                        '        var bins=String(t.bin||t.bins||"");\n'
                        '        var bins_list=bins.split("/").map(function(b){return b.trim();}).filter(Boolean);\n'
                        '        return{bins:bins,bin:bins,bucket:t.fail_bucket||t.bucket||"",expected:String(t.yield||t.expected||""),bins_list:bins_list};\n'
                        '      });\n'
                        '      WM_PAT.yieldDefs=newDefs;\n'
                        '      _wmCriteriaDisabled=new Set();\n'
                        '      _wmCritRebuildRows();\n'
                        '      _wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow&&_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '      var inf=document.getElementById("wm-crit-json-info");\n'
                        '      if(inf){inf.style.color="#27ae60";inf.textContent="\u2713 Loaded "+newDefs.length+" criteria from "+file.name;}\n'
                        '    }catch(err){\n'
                        '      var inf=document.getElementById("wm-crit-json-info");\n'
                        '      if(inf){inf.style.color="#c0392b";inf.textContent="\u2716 "+err.message;}\n'
                        '    }\n'
                        '  };\n'
                        '  reader.readAsText(file);\n'
                        '}\n'
                        'function _wmShowCriteriaCfg(){\n'
                        '  var ex=document.getElementById("wm-crit-cfg-modal");if(ex){ex.remove();return;}\n'
                        '  var defs=WM_PAT.yieldDefs||[];\n'
                        '  var tableRows=defs.length?defs.map(function(def,i){\n'
                        '    var dis=_wmCriteriaDisabled.has(i);\n'
                        '    return "<tr style=\'border-bottom:1px solid #eee\'><td style=\'padding:4px 8px\'><input type=\'checkbox\' id=\'wm-cc-"+i+"\' "+(dis?"":"checked")+" onchange=\'_wmCriteriaCfgToggle("+i+",this.checked)\' style=\'cursor:pointer\'></td>"\n'
                        '      +"<td style=\'padding:4px 8px;font-size:12px\'>"+(def.bucket||"\u2014")+"</td>"\n'
                        '      +"<td style=\'padding:4px 8px;font-size:11px;color:#555\'>IB"+(def.bins||def.bins_list.join("/"))+"</td>"\n'
                        '      +"<td style=\'padding:4px 8px;font-size:12px;font-weight:bold;color:#2471a3\'>"+(def.expected||"\u2014")+"%</td></tr>";\n'
                        '  }).join(""):\'<tr><td colspan="4" style="padding:10px;color:#888;text-align:center">No criteria loaded \u2014 load a JSON file below</td></tr>\';\n'
                        '  var m=document.createElement("div");m.id="wm-crit-cfg-modal";\n'
                        '  m.style.cssText="position:fixed;z-index:99999;top:50%;left:50%;transform:translate(-50%,-50%);background:#fff;border:2px solid #1f618d;border-radius:8px;box-shadow:0 8px 32px rgba(0,0,0,.4);min-width:400px;max-width:580px;max-height:82vh;display:flex;flex-direction:column;font-family:Arial,sans-serif;font-size:12px";\n'
                        '  m.innerHTML="<div style=\'background:#1f618d;color:#fff;padding:8px 14px;border-radius:6px 6px 0 0;display:flex;justify-content:space-between;align-items:center;cursor:move;user-select:none\' id=\'wm-crit-cfg-hdr\'>"\n'
                        '    +"<b>&#9881; Yield Criteria Configuration</b>"\n'
                        '    +"<button onclick=\'document.getElementById(&quot;wm-crit-cfg-modal&quot;).remove()\' style=\'background:none;border:none;color:#fff;font-size:16px;cursor:pointer;padding:0 4px\'>&times;</button></div>"\n'
                        '    +"<div style=\'padding:6px 10px;background:#eaf0fb;font-size:11px;color:#555;border-bottom:1px solid #d0d8e8\'>Check items to include in the &quot;Wafers missing yield target&quot; filter. Load a custom JSON to override targets.</div>"\n'
                        '    +"<div style=\'padding:5px 10px;border-bottom:1px solid #e8eef6;background:#f7f9fc;display:flex;align-items:center;gap:8px;flex-shrink:0\'>"\n'
                        '    +"<span style=\'font-size:11px;font-weight:bold;color:#1f618d;white-space:nowrap\'>&#128193; Load JSON:</span>"\n'
                        '    +"<input type=\'file\' id=\'wm-crit-json-inp\' accept=\'.json\' style=\'font-size:11px;flex:1;min-width:0\' onchange=\'_wmCritLoadJson(this.files[0])\'>"\n'
                        '    +"<span id=\'wm-crit-json-info\' style=\'font-size:11px;white-space:nowrap\'></span></div>"\n'
                        '    +"<div style=\'display:flex;gap:6px;padding:6px 10px;border-bottom:1px solid #eee;flex-shrink:0\'>"\n'
                        '    +"<button onclick=\'_wmCriteriaCfgAll(true)\' style=\'font-size:11px;padding:2px 10px;cursor:pointer;border:1px solid #bbb;border-radius:3px;background:#f5f5f5\'>Select All</button>"\n'
                        '    +"<button onclick=\'_wmCriteriaCfgAll(false)\' style=\'font-size:11px;padding:2px 10px;cursor:pointer;border:1px solid #bbb;border-radius:3px;background:#f5f5f5\'>Clear All</button>"\n'
                        '    +"<span id=\'wm-crit-cfg-count\' style=\'font-size:11px;color:#555;margin-left:auto;line-height:2\'></span></div>"\n'
                        '    +"<div style=\'overflow-y:auto;flex:1\'>"\n'
                        '    +"<table style=\'border-collapse:collapse;width:100%\'><thead><tr style=\'background:#2c3e50;color:#fff\'>"\n'
                        '    +"<th style=\'padding:4px 8px;text-align:left\'>Enable</th><th style=\'padding:4px 8px;text-align:left\'>Bucket</th>"\n'
                        '    +"<th style=\'padding:4px 8px;text-align:left\'>Bins</th><th style=\'padding:4px 8px;text-align:left\'>Target</th></tr></thead>"\n'
                        '    +"<tbody id=\'wm-cc-tbody\'>"+tableRows+"</tbody></table></div>"\n'
                        '    +"<div style=\'padding:6px 10px;border-top:1px solid #eee;text-align:right;flex-shrink:0\'>"\n'
                        '    +"<button onclick=\'document.getElementById(&quot;wm-crit-cfg-modal&quot;).remove()\' style=\'font-size:11px;padding:3px 14px;cursor:pointer;background:#1f618d;color:#fff;border:none;border-radius:4px\'>Done</button></div>";\n'
                        '  document.body.appendChild(m);\n'
                        '  _wmCritCfgUpdateCount();\n'
                        '  (function(){var dx=0,dy=0,drag=false;var hd=document.getElementById("wm-crit-cfg-hdr");if(hd){hd.addEventListener("mousedown",function(e){if(e.button!==0)return;drag=true;dx=e.clientX-m.offsetLeft;dy=e.clientY-m.offsetTop;e.preventDefault();});document.addEventListener("mousemove",function(e){if(!drag)return;m.style.transform="none";m.style.left=(e.clientX-dx)+"px";m.style.top=(e.clientY-dy)+"px";});document.addEventListener("mouseup",function(){drag=false;});}})();\n'
                        '}\n'
                        'function _wmCriteriaCfgToggle(i,on){if(on){_wmCriteriaDisabled.delete(i);}else{_wmCriteriaDisabled.add(i);}\n'
                        '  _wmCritCfgUpdateCount();\n'
                        '  _wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow&&_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function _wmCriteriaCfgAll(on){var defs=WM_PAT.yieldDefs||[];\n'
                        '  if(on){_wmCriteriaDisabled=new Set();}else{defs.forEach(function(_,i){_wmCriteriaDisabled.add(i);});}\n'
                        '  defs.forEach(function(_,i){var cb=document.getElementById("wm-cc-"+i);if(cb)cb.checked=on;});\n'
                        '  _wmCritCfgUpdateCount();\n'
                        '  _wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow&&_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function _wmCritCfgUpdateCount(){var defs=WM_PAT.yieldDefs||[];var n=defs.length-_wmCriteriaDisabled.size;var el=document.getElementById("wm-crit-cfg-count");if(el)el.textContent=n+" / "+defs.length+" active";var btn=document.getElementById("wm-criteria-cfg-btn");if(btn){btn.style.background=_wmCriteriaDisabled.size>0?"rgba(243,156,18,0.5)":"rgba(255,255,255,0.12)";btn.style.borderColor=_wmCriteriaDisabled.size>0?"#f39c12":"rgba(255,255,255,0.4)";}}\n'
                        'function _wmGetCriteriaMissBins(pk){\n'
                        '  var wdata=WM_PAT.wafers[pk];if(!wdata||!wdata.dies)return[];\n'
                        '  var bc={},validDies=0;\n'
                        '  wdata.dies.forEach(function(d){var ib=d[2];if(ib!==null&&ib!==undefined){var k=String(ib);bc[k]=(bc[k]||0)+1;validDies++;}});\n'
                        '  var total=validDies||wdata.dies.length;\n'
                        '  var miss=[];\n'
                        '  (WM_PAT.yieldDefs||[]).forEach(function(def,_di){\n'
                        '    if(_wmCriteriaDisabled.has(_di))return;\n'
                        '    if(!def.expected)return;var exp=parseFloat(def.expected);if(isNaN(exp))return;\n'
                        '    var cnt=def.bins_list.reduce(function(s,b){return s+(bc[b]||0);},0);\n'
                        '    var pct=total>0?cnt/total*100:0;\n'
                        '    var hasBin1=def.bins_list.indexOf(\'1\')>=0;\n'
                        '    var fails=hasBin1?(pct<exp):(pct>exp);\n'
                        '    if(fails)def.bins_list.forEach(function(b){if(miss.indexOf(b)<0)miss.push(b);});\n'
                        '  });\n'
                        '  return miss;\n'
                        '}\n'
                        'function _wmGetCriteriaMissInfo(pk){\n'
                        '  var wdata=WM_PAT.wafers[pk];if(!wdata||!wdata.dies)return[];\n'
                        '  var bc={},total=wdata.dies.length;\n'
                        '  wdata.dies.forEach(function(d){var ib=d[2];if(ib!==null&&ib!==undefined){var k=String(ib);bc[k]=(bc[k]||0)+1;}});\n'
                        '  var info=[];\n'
                        '  (WM_PAT.yieldDefs||[]).forEach(function(def,_di){\n'
                        '    if(_wmCriteriaDisabled.has(_di))return;\n'
                        '    if(!def.expected)return;var exp=parseFloat(def.expected);if(isNaN(exp))return;\n'
                        '    var cnt=def.bins_list.reduce(function(s,b){return s+(bc[b]||0);},0);\n'
                        '    var pct=total>0?cnt/total*100:0;\n'
                        '    var hasBin1=def.bins_list.indexOf(\'1\')>=0;\n'
                        '    var fails=hasBin1?(pct<exp):(pct>exp);\n'
                        '    if(fails)info.push(\'Bin \'+def.bins+\' (\'+def.bucket+\'): \'+pct.toFixed(1)+\'% vs \'+exp+\'% exp\');\n'
                        '  });\n'
                        '  return info;\n'
                        '}\n'
                        'function _wmShowCriteriaTable(pk,evt){\n'
                        '  if(evt)evt.stopPropagation();\n'
                        '  var ci=_wmGetCriteriaMissInfo(pk);if(!ci.length){_wmPatSoloWafer(pk);return;}\n'
                        '  var existing=document.getElementById("wm-crit-modal");if(existing)existing.remove();\n'
                        '  var rows=ci.map(function(r){return"<tr><td style=\'padding:4px 10px;border-bottom:1px solid #eee\'>"+r+"</td></tr>";}).join("");\n'
                        '  var m=document.createElement("div");m.id="wm-crit-modal";\n'
                        '  m.style.cssText="position:fixed;z-index:99999;top:50%;left:50%;transform:translate(-50%,-50%);background:#fff;border:2px solid #c0392b;border-radius:8px;box-shadow:0 8px 32px rgba(0,0,0,.35);min-width:300px;max-width:500px;font-family:Arial,sans-serif;font-size:12px";\n'
                        '  m.innerHTML="<div style=\'background:#c0392b;color:#fff;padding:8px 14px;border-radius:6px 6px 0 0;display:flex;justify-content:space-between;align-items:center\'><b>&#9888; Wafers missing yield target</b><button onclick=\'document.getElementById(&quot;wm-crit-modal&quot;).remove()\'style=\'background:none;border:none;color:#fff;font-size:16px;cursor:pointer;padding:0 4px\'>&times;</button></div>"+\n'
                        '    "<div style=\'padding:6px 4px\'><table style=\'border-collapse:collapse;width:100%\'>"+rows+"</table></div>"+\n'
                        '    "<div style=\'padding:4px 14px 10px;text-align:right\'><button onclick=\'_wmPatSoloWafer(&quot;"+pk+"&quot;);document.getElementById(&quot;wm-crit-modal&quot;).remove()\' style=\'font-size:11px;padding:3px 12px;cursor:pointer;background:#2471a3;color:#fff;border:none;border-radius:4px\'>Focus wafer</button></div>";\n'
                        '  document.body.appendChild(m);\n'
                        '  setTimeout(function(){document.addEventListener("click",function _cl(e){var md=document.getElementById("wm-crit-modal");if(md&&!md.contains(e.target)){md.remove();document.removeEventListener("click",_cl,true);}},true);},100);\n'
                        '}\n'
                        'var _wmIconTipEl=null,_wmTipShownBins=new Set();\n'
                        'function _wmApplyBinFilter(){\n'
                        '  var checked=_wmTipShownBins;\n'
                        '  document.querySelectorAll("#wm-pat-maps rect[data-ib]").forEach(function(r){\n'
                        '    var ib=r.getAttribute("data-ib");\n'
                        '    var baseOp=r.getAttribute("data-op")||"1";\n'
                        '    r.setAttribute("opacity",(checked.size===0||checked.has(ib))?baseOp:"0");\n'
                        '  });\n'
                        '  document.querySelectorAll("#wm-pat-maps text[data-ib]").forEach(function(r){\n'
                        '    var ib=r.getAttribute("data-ib");\n'
                        '    var baseOp=r.getAttribute("data-op")||"1";\n'
                        '    r.setAttribute("opacity",(checked.size===0||checked.has(ib))?baseOp:"0");\n'
                        '  });\n'
                        '}\n'
                        'function _wmResetTipBins(){\n'
                        '  _wmTipShownBins=new Set();\n'
                        '  _wmPatBinChecked=null;\n'
                        '  _wmPatFbFilter={};\n'
                        '  _wmCloseFbSubRow();\n'
                        '  _wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow&&_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '  var t=document.getElementById(\'wm-icon-tip\');if(t){var cbs=t.querySelectorAll(\'input[type=checkbox]\');cbs.forEach(function(c){c.checked=false;});}\n'
                        '}\n'
                        'function _wmTipToggleBins(binsArr,on){\n'
                        '  binsArr.forEach(function(b){if(on)_wmTipShownBins.add(String(b));else _wmTipShownBins.delete(String(b));});\n'
                        '  _wmPatBinChecked=_wmTipShownBins.size>0?new Set(_wmTipShownBins):null;\n'
                        '  _wmPatRender();\n'
                        '  var hdr=document.getElementById(\'wm-tip-hdr\');if(hdr)hdr.style.background=_wmTipShownBins.size>0?\'#922b21\':\'#c0392b\';\n'
                        '}\n'
                        'function _wmShowIconTip(el){\n'
                        '  var existing=document.getElementById(\'wm-icon-tip\');\n'
                        '  if(existing){\n'
                        '    if(_wmIconTipEl===el){_wmResetTipBins();existing.remove();_wmIconTipEl=null;return;}\n'
                        '    existing.remove();\n'
                        '  }\n'
                        '  var pk=el.dataset.pk;if(!pk)return;\n'
                        '  _wmFbPk=pk;\n'
                        '  var wdata=WM_PAT.wafers[pk];if(!wdata||!wdata.dies)return;\n'
                        '  var bc={},total=wdata.dies.length;\n'
                        '  wdata.dies.forEach(function(d){var ib=d[2];if(ib!==null&&ib!==undefined){var k=String(ib);bc[k]=(bc[k]||0)+1;}});\n'
                        '  var rows=[],anyFail=false;\n'
                        '  (WM_PAT.yieldDefs||[]).forEach(function(def,_di){\n'
                        '    if(_wmCriteriaDisabled.has(_di))return;\n'
                        '    if(!def.expected)return;var exp=parseFloat(def.expected);if(isNaN(exp))return;\n'
                        '    var cnt=def.bins_list.reduce(function(s,b){return s+(bc[b]||0);},0);\n'
                        '    var pct=total>0?cnt/total*100:0;\n'
                        '    var hasBin1=def.bins_list.indexOf(\'1\')>=0;\n'
                        '    var fails=hasBin1?(pct<exp):(pct>exp);\n'
                        '    if(!fails)return;\n'
                        '    anyFail=true;\n'
                        '    var bj=JSON.stringify(def.bins_list.map(Number));\n'
                        '    var chk=_wmTipShownBins.size>0&&def.bins_list.every(function(b){return _wmTipShownBins.has(String(b));});\n'
                        '    rows.push(\'<tr style="border-bottom:1px solid #f0e0e0"><td style="padding:3px 8px 3px 4px;font-size:11px;color:#333">\'+( def.bucket||\'-\')+\'</td><td style="padding:3px 8px;font-size:11px;color:#555">IB\'+def.bins+\'</td><td style="padding:3px 8px;font-size:11px;font-weight:bold;color:#c0392b">\'+pct.toFixed(1)+\'%</td><td style="padding:3px 8px;font-size:11px;color:#777">\'+exp+\'%</td><td style="padding:3px 4px;text-align:center"><input type="checkbox"\'+(chk?\' checked\':\'\')+\' onchange="_wmTipToggleBins(\'+bj+\',this.checked)" style="cursor:pointer;width:14px;height:14px"></td><td style="padding:3px 4px;text-align:center"><button title="Analyze" onclick="_wmAnalyzeBins(\'+bj+\')" style="background:none;border:1px solid #c8d4e0;border-radius:3px;cursor:pointer;font-size:11px;padding:0 4px;line-height:16px;color:#1a5276">&#128300;</button></td></tr>\');\n'
                        '  });\n'
                        '  if(!anyFail)return;\n'
                        '  var t=document.createElement(\'div\');t.id=\'wm-icon-tip\';\n'
                        '  t.style.cssText=\'position:fixed;z-index:999999;background:#fff;border:2px solid #c0392b;border-radius:6px;box-shadow:0 4px 16px rgba(0,0,0,.35);padding:0;pointer-events:auto;min-width:280px\';\n'
                        '  t.innerHTML=\'<div id="wm-tip-hdr" style="background:#c0392b;color:#fff;padding:5px 8px;border-radius:4px 4px 0 0;display:flex;justify-content:space-between;align-items:center"><span style="font-weight:bold;font-size:11px">&#9888; Yield miss \u2014 check to highlight</span><button onclick="_wmResetTipBins();document.getElementById(&quot;wm-icon-tip&quot;).remove();_wmIconTipEl=null" style="background:none;border:none;color:#fff;font-size:15px;cursor:pointer;padding:0 2px;line-height:1">&times;</button></div><div style="padding:6px 8px"><table style="border-collapse:collapse;width:100%"><thead><tr style="background:#fdf2f2"><th style="padding:2px 8px 2px 4px;font-size:10px;text-align:left;color:#922b21">Bucket</th><th style="padding:2px 8px;font-size:10px;text-align:left;color:#922b21">Bin</th><th style="padding:2px 8px;font-size:10px;text-align:left;color:#922b21">Actual</th><th style="padding:2px 8px;font-size:10px;text-align:left;color:#922b21">Expected</th><th style="padding:2px 4px;font-size:10px;text-align:center;color:#922b21">Show Only</th></tr></thead><tbody>\'+rows.join(\'\')+\'</tbody></table><div style="text-align:right;margin-top:4px"><button onclick="_wmResetTipBins()" style="font-size:10px;padding:2px 8px;cursor:pointer;border:1px solid #ccc;border-radius:3px;background:#f9f9f9">Reset</button></div></div>\';\n'
                        '  document.body.appendChild(t);\n'
                        '  var r=el.getBoundingClientRect();\n'
                        '  var tx=r.left,ty=r.bottom+4;\n'
                        '  if(tx+t.offsetWidth>window.innerWidth-8)tx=window.innerWidth-t.offsetWidth-8;\n'
                        '  if(ty+t.offsetHeight>window.innerHeight-8)ty=r.top-t.offsetHeight-4;\n'
                        '  t.style.left=tx+\'px\';t.style.top=ty+\'px\';\n'
                        '  _wmIconTipEl=el;\n'

                        '  (function(){var tdx=0,tdy=0,tdrag=false;var th=document.getElementById(\'wm-tip-hdr\');if(th){th.style.cursor=\'move\';th.addEventListener(\'mousedown\',function(e){if(e.button!==0)return;tdrag=true;var tr=t.getBoundingClientRect();tdx=e.clientX-tr.left;tdy=e.clientY-tr.top;e.stopPropagation();e.preventDefault();});document.addEventListener(\'mousemove\',function _tm(e){if(!tdrag)return;t.style.left=(e.clientX-tdx)+\'px\';t.style.top=(e.clientY-tdy)+\'px\';});document.addEventListener(\'mouseup\',function(){tdrag=false;});}})();\n'
                        '  setTimeout(function(){\n'
                        '    document.addEventListener(\'click\',function _tc(e){\n'
                        '      var ti=document.getElementById(\'wm-icon-tip\');\n'
                        '      if(ti&&!ti.contains(e.target)&&e.target!==el&&!document.getElementById(\'wm-analyze-overlay\')&&!document.getElementById(\'wm-hw-overlay\')&&!document.getElementById(\'wm-ba-overlay\')){_wmResetTipBins();ti.remove();_wmIconTipEl=null;document.removeEventListener(\'click\',_tc,true);}\n'
                        '    },true);\n'
                        '  },50);\n'
                        '}\n'
                        'function _wmCloseAnalyze(){var e=document.getElementById(\'wm-analyze-overlay\');if(e)e.remove();}\n'
                        f'{_wm_fb_analyze_js}'
                        'var _wmDieTt=null;\n'
                        'function _wmDieHover(e,svg){\n'
                        '  var t=e.target;if(!t||t.tagName!=="rect"||!t.dataset.x)return;\n'
                        '  var ib=t.dataset.ib,fb=t.dataset.fb,dx=t.dataset.x,dy=t.dataset.y;\n'
                        '  var upmRaw=t.dataset.upm||"";var upmCols=WM_PAT.upmCols||[];\n'
                        '  if(!_wmDieTt){_wmDieTt=document.createElement("div");_wmDieTt.id="wm-die-tt";\n'
                        '    _wmDieTt.style.cssText="position:fixed;z-index:99998;background:rgba(30,30,30,0.92);color:#fff;font-size:11px;font-family:Arial,sans-serif;padding:5px 9px;border-radius:5px;pointer-events:none;white-space:nowrap;box-shadow:0 2px 8px rgba(0,0,0,.4);line-height:1.6";\n'
                        '    document.body.appendChild(_wmDieTt);}\n'
                        '  var ibDesc=(WM_PAT.ibColors&&WM_PAT.ibColors[ib])?\'<span style="display:inline-block;width:8px;height:8px;background:\'+WM_PAT.ibColors[ib]+\';border-radius:2px;margin-right:3px;vertical-align:middle"></span>\':\'\';\n'
                        '  var fbLine=fb!==null&&fb!==\'\'?\'<br><b>FB:</b> \'+fb:\'\';\n'
                        '  var fbDesc=(WM_PAT.fbDescriptions&&WM_PAT.fbDescriptions[fb]&&WM_PAT.fbDescriptions[fb].desc)?\'<span style="color:#aaa;font-size:10px"> \u2014 \'+WM_PAT.fbDescriptions[fb].desc.substring(0,30)+\'</span>\':\'\';\n'
                        '  var upmLines=\'\';\n'
                        '  if(upmCols.length&&upmRaw){\n'
                        '    var upmVals=upmRaw.split(\'|\');\n'
                        '    var entry=upmCols[0];var tgt=Array.isArray(entry)?entry[1]:null;\n'
                        '    var v=upmVals[0]!==undefined&&upmVals[0]!==\'\'?parseFloat(upmVals[0]):null;\n'
                        '    if(v!==null&&!isNaN(v)){var disp=tgt?((v/tgt*100).toFixed(1)+\'%\'):v;upmLines=\'<br><b>UPM:</b> \'+disp;}\n'
                        '  }\n'
                        '  _wmDieTt.innerHTML=\'<b>X:</b> \'+dx+\' &nbsp; <b>Y:</b> \'+dy+\'<br>\'+ibDesc+\'<b>IB:</b> \'+ib+fbLine+fbDesc+upmLines;\n'
                        '  var margin=12;\n'
                        '  var ttW=_wmDieTt.offsetWidth||140,ttH=_wmDieTt.offsetHeight||52;\n'
                        '  var left=e.clientX+margin,top=e.clientY-ttH-margin;\n'
                        '  if(left+ttW>window.innerWidth)left=e.clientX-ttW-margin;\n'
                        '  if(top<4)top=e.clientY+margin;\n'
                        '  _wmDieTt.style.left=left+"px";_wmDieTt.style.top=top+"px";\n'
                        '  _wmDieTt.style.display="";\n'
                        '}\n'
                        'function _wmDieHoverOut(){\n'
                        '  if(_wmDieTt)_wmDieTt.style.display="none";\n'
                        '}\n'
                        'function _wmZoomWafer(pk,triggerEl){\n'
                        '  var ex=document.getElementById(\'wm-zoom-modal\');if(ex)ex.remove();\n'
                        '  _wmFbPk=pk;\n'
                        '  var wdata=WM_PAT.wafers[pk];if(!wdata||!wdata.dies)return;\n'
                        '  var _pks=pk.split("::");var mLot=wdata.lot||_pks[0]||pk,mWfr=wdata.wafer||_pks[1]||"";\n'
                        '  var origSvg=triggerEl?triggerEl.querySelector(\'svg\'):null;\n'
                        '  if(!origSvg&&_wmPatCanvasMode){\n'
                        '    var _tmpC=document.createElement(\'div\');var _prevM=_wmPatCanvasMode;\n'
                        '    _wmPatCanvasMode=false;_wmPatRenderTile(pk,_tmpC);_wmPatCanvasMode=_prevM;\n'
                        '    origSvg=_tmpC.querySelector(\'svg\');\n'
                        '  }\n'
                        '  if(!origSvg)return;\n'
                        '  var origW=parseFloat(origSvg.getAttribute(\'width\')),origH=parseFloat(origSvg.getAttribute(\'height\'));\n'
                        '  var maxDim=Math.min(window.innerWidth*0.45,window.innerHeight*0.45,420);\n'
                        '  var scale=Math.min(maxDim/origW,maxDim/origH);\n'
                        '  var newW=Math.round(origW*scale),newH=Math.round(origH*scale);\n'
                        '  var clonedSvg=origSvg.cloneNode(true);\n'
                        '  clonedSvg.id=\'wm-zoom-svg\';\n'
                        '  clonedSvg.setAttribute(\'viewBox\',\'0 0 \'+origW+\' \'+origH);\n'
                        '  clonedSvg.setAttribute(\'width\',newW);clonedSvg.setAttribute(\'height\',newH);\n'
                        '  clonedSvg.style.cssText=\'display:block;margin:0 auto;flex-shrink:0;\';\n'
                        '  var clipId=\'wzmC\'+pk.replace(/[^a-z0-9]/gi,\'_\');\n'
                        '  var cpEl=clonedSvg.querySelector(\'clipPath\');\n'
                        '  if(cpEl){cpEl.id=clipId;var gEl=clonedSvg.querySelector(\'[clip-path]\');if(gEl)gEl.setAttribute(\'clip-path\',\'url(#\'+clipId+\')\');}\n'
                        '  clonedSvg.querySelectorAll(\'rect[fill="none"]\').forEach(function(r){var sw=parseFloat(r.getAttribute(\'stroke-width\')||1);r.setAttribute(\'stroke-width\',(sw*0.5).toFixed(2));r.setAttribute(\'stroke\',\'#2471a3\');});\n'
                        '  var _zmShown=new Set();\n'
                        '  window._wmZmToggle=function(binsArr,on){\n'
                        '    binsArr.forEach(function(b){if(on)_zmShown.add(String(b));else _zmShown.delete(String(b));});\n'
                        '    clonedSvg.querySelectorAll(\'rect[data-ib]\').forEach(function(r){\n'
                        '      var ib=r.getAttribute(\'data-ib\');\n'
                        '      var baseOp=r.getAttribute(\'data-op\')||r.getAttribute(\'opacity\')||\'1\';\n'
                        '      r.setAttribute(\'opacity\',(_zmShown.size===0||_zmShown.has(ib))?baseOp:\'0\');\n'
                        '    });\n'
                        '  };\n'
                        '  var dies=wdata.dies;\n'
                        '  var bc2={},total2=dies.length;\n'
                        '  dies.forEach(function(d){var ib=d[2];if(ib!==null&&ib!==undefined){var k=String(ib);bc2[k]=(bc2[k]||0)+1;}});\n'
                        '  var missRows=[];\n'
                        '  (WM_PAT.yieldDefs||[]).forEach(function(def){\n'
                        '    if(!def.expected)return;var exp=parseFloat(def.expected);if(isNaN(exp))return;\n'
                        '    var cnt=def.bins_list.reduce(function(s,b){return s+(bc2[b]||0);},0);\n'
                        '    var pct=total2>0?cnt/total2*100:0;\n'
                        '    var hasBin1=def.bins_list.indexOf(\'1\')>=0;\n'
                        '    var fails=hasBin1?(pct<exp):(pct>exp);\n'
                        '    var failCol=fails?\'#c0392b\':\'#27ae60\',sym=fails?\'&#9888;\':\'&#10003;\';\n'
                        '    var swCol=_wmIbColor(def.bins_list[0]||null);\n'
                        '    var bj=JSON.stringify(def.bins_list.map(Number));\n'
                        '    missRows.push(\'<tr style="border-bottom:1px solid #f0f0f0"><td style="padding:4px;font-size:12px;white-space:nowrap"><span style="display:inline-block;width:11px;height:11px;border-radius:2px;background:\'+swCol+\';vertical-align:middle;margin-right:4px;border:1px solid rgba(0,0,0,.15)"></span><span style="color:\'+failCol+\'">\'+sym+\'</span>&nbsp;\'+(def.bucket||\'—\')+\'</td><td style="padding:4px 8px;font-size:12px;color:#555">IB\'+def.bins+\'</td><td style="padding:4px 8px;font-size:12px;font-weight:bold;color:\'+failCol+\'">\'+pct.toFixed(1)+\'%</td><td style="padding:4px 8px;font-size:12px;color:#777">\'+exp+\'%</td><td style="padding:4px 6px;text-align:center"><input type="checkbox" onchange="_wmZmToggle(\'+bj+\',this.checked)" style="cursor:pointer;width:13px;height:13px"></td><td style="padding:4px 6px;text-align:center"><button title="Analyze" onclick="_wmAnalyzeBins(\'+bj+\')" style="background:none;border:1px solid #c8d4e0;border-radius:3px;cursor:pointer;font-size:11px;padding:0 4px;line-height:16px;color:#1a5276">&#128300;</button></td></tr>\');\n'
                        '  });\n'
                        '  var tableHtml=missRows.length?\'<table style="border-collapse:collapse;width:100%"><thead><tr style="background:#f4f6f9"><th style="padding:3px 4px;font-size:10px;text-align:left;color:#555">Bucket</th><th style="padding:3px 8px;font-size:10px;text-align:left;color:#555">Bin</th><th style="padding:3px 8px;font-size:10px;text-align:left;color:#555">Actual</th><th style="padding:3px 8px;font-size:10px;text-align:left;color:#555">Expected</th><th style="padding:3px 4px;font-size:10px;text-align:center;color:#555">Show Only</th><th style="padding:3px 4px;font-size:10px;text-align:center;color:#555"></th></tr></thead><tbody>\'+missRows.join(\'\')+ \'</tbody></table>\':\'\';\n'
                        '  var m2=document.createElement(\'div\');m2.id=\'wm-zoom-modal\';\n'
                        '  var _mw=Math.max(newW+48,460);\n'
                        '  var _mh=Math.min(newH+340,Math.round(window.innerHeight*0.92));\n'
                        '  var _ml=Math.round((window.innerWidth-_mw)/2),_mt=Math.round((window.innerHeight-_mh)/2);\n'
                        '  m2.style.cssText=\'position:fixed;z-index:99999;top:\'+ _mt+\'px;left:\'+_ml+\'px;width:\'+_mw+\'px;height:\'+_mh+\'px;resize:both;overflow:hidden;background:#fff;border:1px solid #bdc3c7;border-radius:10px;box-shadow:0 12px 40px rgba(0,0,0,.4);min-width:260px;min-height:180px;font-family:Arial,sans-serif\';\n'
                        '  var hdr=document.createElement(\'div\');\n'
                        '  hdr.style.cssText=\'background:#2c3e50;color:#fff;padding:8px 14px;border-radius:10px 10px 0 0;display:flex;justify-content:space-between;align-items:center;cursor:move;user-select:none;flex-shrink:0\';\n'
                        '  hdr.innerHTML=\'<b>\'+mLot+\' — Wafer \'+mWfr+\'</b><button onclick="document.getElementById(\\&quot;wm-zoom-modal\\&quot;).remove();delete window._wmZmToggle" style="background:none;border:none;color:#fff;font-size:18px;cursor:pointer;padding:0 4px">&times;</button>\';\n'
                        '  var body2=document.createElement(\'div\');body2.style.cssText=\'padding:12px;overflow-y:auto;height:calc(100% - 42px);box-sizing:border-box;display:flex;flex-direction:column;align-items:center;gap:10px\';\n'
                        '  body2.appendChild(clonedSvg);\n'
                        '  if(tableHtml){var td2=document.createElement(\'div\');td2.style.marginTop=\'10px\';td2.innerHTML=tableHtml;body2.appendChild(td2);}\n'
                        '  m2.style.display=\'flex\';m2.style.flexDirection=\'column\';\n'
                        '  m2.appendChild(hdr);m2.appendChild(body2);\n'
                        '  document.body.appendChild(m2);\n'
                        '  requestAnimationFrame(function(){\n'
                        '    body2.style.height=\'auto\';body2.style.overflow=\'visible\';\n'
                        '    var needed=body2.scrollHeight+hdr.offsetHeight+20;\n'
                        '    body2.style.height=\'\';body2.style.overflow=\'\';\n'
                        '    var maxH=Math.round(window.innerHeight*0.92);\n'
                        '    var fitH=Math.min(needed,maxH);\n'
                        '    m2.style.height=fitH+\'px\';\n'
                        '    var fw=Math.round(Math.max(newW+48,body2.scrollWidth+32,460)*1.15);\n'
                        '    fw=Math.min(fw,Math.round(window.innerWidth*0.92));\n'
                        '    m2.style.width=fw+\'px\';\n'
                        '    m2.style.left=Math.round((window.innerWidth-fw)/2)+\'px\';\n'
                        '    m2.style.top=Math.round((window.innerHeight-fitH)/2)+\'px\';\n'
                        '  });\n'
                        '  (function(){var dx=0,dy=0,drag=false;hdr.addEventListener(\'mousedown\',function(e){if(e.button!==0)return;drag=true;dx=e.clientX-m2.offsetLeft;dy=e.clientY-m2.offsetTop;e.preventDefault();});document.addEventListener(\'mousemove\',function(e){if(!drag)return;m2.style.left=(e.clientX-dx)+\'px\';m2.style.top=(e.clientY-dy)+\'px\';});document.addEventListener(\'mouseup\',function(){drag=false;});})();\n'
                        '  setTimeout(function(){document.addEventListener(\'click\',function _zc(e){var md=document.getElementById(\'wm-zoom-modal\');if(md&&!md.contains(e.target)&&!document.getElementById(\'wm-analyze-overlay\')&&!document.getElementById(\'wm-hw-overlay\')&&!document.getElementById(\'wm-ba-overlay\')){md.remove();delete window._wmZmToggle;document.removeEventListener(\'click\',_zc,true);}},true);},100);\n'
                        '}\n'
'function _wmToggleCriteriaMiss(on){\n'
'  _wmCriteriaMissOnly=on;\n'
'  var lbl=document.getElementById("wm-criteria-miss-lbl");\n'
'  if(lbl){lbl.style.background=on?"rgba(192,57,43,0.6)":"rgba(255,255,255,0.12)";lbl.style.borderColor=on?"#fff":"rgba(255,255,255,0.4)";}\n'
'  var chk=document.getElementById("wm-criteria-miss-chk");if(chk)chk.checked=on;\n'
'  if(on){\n'
'    var _allK=Object.keys(WM_PAT.wafers).filter(function(k){return _wmPatMatchLots(k)&&_wmPatMatchProgs(k);});\n'
'    var _failK=_allK.filter(function(k){return _wmGetCriteriaMissBins(k).length>0;});\n'
'    _wmPatSelWafers=_failK.length>0?new Set(_failK):null;\n'
'  } else {\n'
'    _wmPatSelWafers=null;\n'
'  }\n'
'  _wmDdUpdateBtns();_wmDdRefresh();\n'
'  _wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
'}\n'
                        'function _wmSetFailThr(v){_wmFailThr=+v;var sel=document.getElementById("wm-fail-thr");if(sel)sel.value=v;document.querySelectorAll("input[name=\'wm-thr-rb\']").forEach(function(rb){rb.checked=(+rb.value===_wmFailThr);});if(typeof _wmPatRender==="function")_wmPatRender();if(typeof wmPatRenderReticle==="function")wmPatRenderReticle();}\n'
                        'function _wmSetEdgeRows(n){_wmEdgeExcRows=+n;_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);document.querySelectorAll("select.wm-edge-sel").forEach(function(s){s.value=String(_wmEdgeExcRows);});}\n'
                        'function _wmPatSetZoom(delta){_wmPatZoom=Math.min(4,Math.max(0.25,Math.round((_wmPatZoom+delta)*100)/100));_wmPatRender();}\n'
                        'function _wmPatSoloWafer(pk){\n'
                        '  if(_wmPatSelWafers&&_wmPatSelWafers.size===1&&_wmPatSelWafers.has(pk)){_wmPatSelWafers=null;}\n'
                        '  else{_wmPatSelWafers=new Set([pk]);}\n'
                        '  _wmDdUpdateBtns();_wmDdSetIndeterminate();_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function _wmScorePattern(failXn,failYn){\n'
                        '  var N=failXn.length;\n'
                        '  if(!N)return{center:0,edge:0,donut:0,systematic:0,random:1,confidence:"LOW"};\n'
                        '  // 6 radial bands; expected fractions: uniform circular P(r<R)=R^2\n'
                        '  var B1=0,B2=0,B3=0,B4=0,B5=0,B6=0,q=[0,0,0,0];\n'
                        '  for(var i=0;i<N;i++){\n'
                        '    var r=Math.sqrt(failXn[i]*failXn[i]+failYn[i]*failYn[i]);\n'
                        '    if(r<0.15)B1++;else if(r<0.40)B2++;else if(r<0.60)B3++;else if(r<0.75)B4++;else if(r<0.90)B5++;else B6++;\n'
                        '    var xi=failXn[i],yi=failYn[i];\n'
                        '    if(xi>=0&&yi>=0)q[0]++;else if(xi<0&&yi>=0)q[1]++;else if(xi<0&&yi<0)q[2]++;else q[3]++;\n'
                        '  }\n'
                        '  var fC=(B1+B2)/N,eC=0.16;\n'
                        '  var fE=(B5+B6)/N,eE=0.4375;\n'
                        '  var fM=(B3+B4)/N,eM=0.4025;\n'
                        '  var centerScore=Math.max(0,Math.min(1,(fC-eC)/(1-eC)));\n'
                        '  var edgeScore  =Math.max(0,Math.min(1,(fE-eE)/(1-eE)));\n'
                        '  var midEnrich  =Math.max(0,(fM-eM)/(1-eM));\n'
                        '  var donutScore =Math.min(1,midEnrich*2*(1-Math.max(centerScore,edgeScore)*0.7));\n'
                        '  var sampleConf =Math.min(1,N/20);\n'
                        '  var qImbal=(Math.max.apply(null,q)-Math.min.apply(null,q))/N;\n'
                        '  var systematicScore=Math.min(1,qImbal*2.5)*sampleConf;\n'
                        '  var dominated=Math.max(centerScore,edgeScore,donutScore,systematicScore);\n'
                        '  var randomScore=Math.max(0,Math.min(1,1-dominated));\n'
                        '  var conf=N<20?"LOW":N<50?"MEDIUM":"HIGH";\n'
                        '  return{center:+centerScore.toFixed(2),edge:+edgeScore.toFixed(2),donut:+donutScore.toFixed(2),systematic:+systematicScore.toFixed(2),random:+randomScore.toFixed(2),confidence:conf};\n'
                        '}\n'
                        'function _wmScoreReticle(actX,actY,rm,st){\n'
                        '  if(!rm)rm=WM_PAT.retMap;if(!st)st=WM_PAT.retSiteTotals;\n'
                        '  if(!rm||!st||!actX||!actX.length)return 0;\n'
                        '  var siteShots={},siteCnt={},N=actX.length;\n'
                        '  for(var i=0;i<N;i++){var info=rm[actX[i]+","+actY[i]];if(!info)continue;var sk=info[0]+","+info[1];var si=String(info[2]);if(!siteShots[sk]){siteShots[sk]={};siteCnt[sk]=0;}siteShots[sk][si]=true;siteCnt[sk]++;}\n'
                        '  var sites=Object.keys(siteShots);if(!sites.length)return 0;\n'
                        '  var maxSiteScore=0,weightedSum=0,totalMapped=0;\n'
                        '  sites.forEach(function(sk){var totShots=st[sk]||1;var failShots=Object.keys(siteShots[sk]).length;var score=failShots/totShots;var cnt=siteCnt[sk];totalMapped+=cnt;weightedSum+=score*cnt;if(score>maxSiteScore)maxSiteScore=score;});\n'
                        '  if(!totalMapped)return 0;\n'
                        '  var raw=(weightedSum/totalMapped)*0.4+maxSiteScore*0.6;\n'
                        '  var sampleConf=Math.min(1,N/15);\n'
                        '  return Math.min(1,raw*sampleConf);\n'
                        '}\n'
                        'function _wmPrimary(sc){\n'
                        '  var best="random",bv=sc.random;\n'
                        '  ["center","edge","donut","systematic","reticle"].forEach(function(k){if(sc[k]!==undefined&&sc[k]>bv){bv=sc[k];best=k;}});\n'
                        '  return{center:"CENTER",edge:"EDGE",donut:"DONUT",systematic:"SYSTEMATIC",reticle:"RETICLE",random:"RANDOM"}[best]||best.toUpperCase();\n'
                        '}\n'
                        'var _pColors={CENTER:"#c0392b",EDGE:"#e67e22",DONUT:"#8e44ad",SYSTEMATIC:"#2471a3",RETICLE:"#1f618d",RANDOM:"#27ae60"};\n'
                        'var _wmPatBinChecked=null;\n'
                        'var _wmmHlIb=null;\n'
                        'var _wmmHeatMode=false;\n'
                        'var _wmPatSelWafers=null;\n'
                        'var _wmPatZoom=1.0;\n'
                        'var _wmPatCurLots=null;\n'
                        'var _wmPatCurProgs=null;\n'
                        'var _wmPatRetUnchecked=null;\n'
                        'var _wmPatSiteToShots=null;\n'
                        'function _wmPatGetLot(k){return k.split("::")[0]||k;}\n'
                        'function _wmPatGetWfr(k){return k.split("::")[1]||k;}\n'
                        'function _wmPatGetProg(k){return k.split("::")[2]||"";}\n'
                        'function _wmPatAllLots(){var s=new Set();Object.keys(WM_PAT.wafers).forEach(function(k){s.add(_wmPatGetLot(k));});return Array.from(s);}\n'
                        'function _wmPatAllProgs(){var s=new Set();Object.keys(WM_PAT.wafers).forEach(function(k){var p=_wmPatGetProg(k);if(p)s.add(p);});return Array.from(s).sort();}\n'
                        'function _wmPatMatchLots(k){if(!_wmPatCurLots)return true;var lot=_wmPatGetLot(k);for(var i=0;i<_wmPatCurLots.length;i++){if(_wmPatCurLots[i]===lot)return true;}return false;}\n'
                        'function _wmPatMatchProgs(k){if(!_wmPatCurProgs)return true;var p=_wmPatGetProg(k);if(!p)return true;for(var i=0;i<_wmPatCurProgs.length;i++){if(_wmPatCurProgs[i]===p)return true;}return false;}\n'
                        'function _wmRetInfoFor(pk){var pfx=(WM_PAT.wafers[pk]||{}).pfx||"";var m=WM_PAT.retMaps&&WM_PAT.retMaps[pfx];return m||{retMap:WM_PAT.retMap,retShots:WM_PAT.retShots,retSiteTotals:WM_PAT.retSiteTotals};}\n'
                        'function _wmRetInfoForLot(lot){var keys=Object.keys(WM_PAT.wafers).filter(function(k){return k.split("::")[0]===lot;});var _ri=keys.length?_wmRetInfoFor(keys[0]):{retMap:WM_PAT.retMap,retSiteTotals:WM_PAT.retSiteTotals};if(!_ri.retSiteLabels)_ri.retSiteLabels=WM_PAT.retSiteLabels||{};return _ri;}\n'
                        'var _wmPatIsPopup=false;\n'
                        'var _wmPatPopupWin=null;\n'
                        'var _wmPatCanvasMode=true,_wmPatObserver=null,_wmPatRenderedKeys=new Set(),_wmPatLastKeys=[];\n'
                        'function wmOpenPat(){\n'
                        '  var base=location.href.split("#")[0];\n'
                        '  var url=base+"#wpa";\n'
                        '  if(_wmPatPopupWin&&!_wmPatPopupWin.closed){_wmPatPopupWin.focus();return;}\n'
                        '  _wmPatPopupWin=window.open(url,"WaferPatternAnalysis","width=1400,height=850,resizable=yes,scrollbars=yes");\n'
                        '  if(_wmPatPopupWin)_wmPatPopupWin.focus();\n'
                        '}\n'
                        'function wmHidePat(){\n'
                        '  if(_wmPatIsPopup){window.close();return;}\n'
                        '  var ov=document.getElementById("wm-pat-overlay");\n'
                        '  if(ov)ov.classList.remove("open");\n'
                        '}\n'
                        'function wmPatTab(t){\n'
                        '  ["impact","composite2","reticle","guide"].forEach(function(n){\n'
                        '    var btn=document.getElementById("wm-pat-tab-"+n);\n'
                        '    var pane=document.getElementById("wm-pat-pane-"+n);\n'
                        '    if(btn)btn.classList.toggle("on",n===t);\n'
                        '    if(pane)pane.classList.toggle("on",n===t);\n'
                        '  });\n'
                        '  if(t==="reticle")wmPatRenderReticle();\n'
                        '}\n'
                        'function _wmPatGetSiteShots(){\n'
                        '  if(_wmPatSiteToShots)return _wmPatSiteToShots;\n'
                        '  _wmPatSiteToShots={};\n'
                        '  if(WM_PAT.hasReticle){\n'
                        '    var _allRms=WM_PAT.retMaps?Object.values(WM_PAT.retMaps).map(function(x){return x.retMap;}):[];\n'
                        '    if(WM_PAT.retMap)_allRms.push(WM_PAT.retMap);\n'
                        '    _allRms.forEach(function(rm){if(!rm)return;Object.keys(rm).forEach(function(k){\n'
                        '      var info=rm[k];var sk=info[0]+","+info[1];\n'
                        '      if(!_wmPatSiteToShots[sk])_wmPatSiteToShots[sk]=new Set();\n'
                        '      _wmPatSiteToShots[sk].add(info[2]);\n'
                        '    });});\n'
                        '  }\n'
                        '  return _wmPatSiteToShots;\n'
                        '}\n'
                        'function wmPatRetSiteToggle(sk,on){\n'
                        '  if(!on){if(!_wmPatRetUnchecked)_wmPatRetUnchecked=new Set();_wmPatRetUnchecked.add(sk);}\n'
                        '  else{if(_wmPatRetUnchecked){_wmPatRetUnchecked.delete(sk);if(_wmPatRetUnchecked.size===0)_wmPatRetUnchecked=null;}}\n'
                        '  _wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function wmPatRetClear(){_wmPatRetUnchecked=null;_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);}\n'
                        'function wmPatRenderReticle(){\n'
                        '  var el=document.getElementById("wm-pat-reticle-body");\n'
                        '  if(!el)return;\n'
                        '  if(!WM_PAT.hasReticle||(!WM_PAT.retMap&&(!WM_PAT.retMaps||!Object.keys(WM_PAT.retMaps).length))){el.innerHTML=\'<span style="color:#aaa;font-size:11px">No reticle mapping loaded.</span>\';return;}\n'
                        '  var allKeys=Object.keys(WM_PAT.wafers).filter(function(k){return _wmPatMatchLots(k)&&_wmPatMatchProgs(k);});\n'
                        '  var keys=_wmPatSelWafers===null?allKeys:allKeys.filter(function(k){return _wmPatSelWafers.has(k);});\n'
                        '  var nWafers=keys.filter(function(pk){var w=WM_PAT.wafers[pk];return w&&w.dies&&w.dies.length;}).length;\n'
                        '  var _retXMin=Infinity,_retXMax=-Infinity,_retYMin=Infinity,_retYMax=-Infinity;\n  keys.forEach(function(pk){var _w=WM_PAT.wafers[pk];var _d=_w&&_w.dies?_w.dies:_w;if(_d)_d.forEach(function(d){if(d[0]!==null){if(d[0]<_retXMin)_retXMin=d[0];if(d[0]>_retXMax)_retXMax=d[0];if(d[1]<_retYMin)_retYMin=d[1];if(d[1]>_retYMax)_retYMax=d[1];}});});\n  if(_retXMin===Infinity){_retXMin=0;_retXMax=0;_retYMin=0;_retYMax=0;}\n'
                        '  var siteFailShots={},siteFailCount={},grandTotalFail=0;\n'
                        '  var shotFailData={},shotWaferHits={};\n'
                        '  keys.forEach(function(pk){\n'
                        '    var wdata=WM_PAT.wafers[pk];\n'
                        '    var dies=wdata&&wdata.dies?wdata.dies:wdata;\n'
                        '    if(!dies||!dies.length)return;\n'
                        '    var _shotsSeen={};\n'
                        '    dies.forEach(function(d){\n'
                        '      var x=d[0],y=d[1],ib=d[2];if(x===null||x===undefined)return;\n'
                        '      var _ib2=typeof ib==="number"?ib:parseInt(ib);if(isNaN(_ib2))return;\n'
                        '      var binOn=(_wmPatBinChecked===null||_wmPatBinChecked.has(String(ib)));\n'
                        '      if(binOn&&ib!==null&&ib!==undefined){var _fbFr=_wmPatFbFilter[String(ib)];if(_fbFr!==undefined&&_fbFr!==null){var _fbValR=d[3]!==undefined&&d[3]!==null?String(d[3]):null;if(_fbValR===null||!_fbFr.has(_fbValR))binOn=false;}}\n'
                        '      if(!binOn)return;\n'
                        '      if(_wmPatBinChecked===null&&!_wmIsFail(_ib2))return;\n'
                        '      var _lri=_wmRetInfoFor(pk);var info=_lri.retMap&&_lri.retMap[x+","+y];if(!info)return;\n'
                        '      var sk=info[0]+","+info[1];var shotIdx=String(info[2]);\n'
                        '      if(_wmEdgeExcRows>0&&(x<_retXMin+_wmEdgeExcRows||x>_retXMax-_wmEdgeExcRows||y<_retYMin+_wmEdgeExcRows||y>_retYMax-_wmEdgeExcRows))return;\n'
                        '      if(!siteFailShots[sk])siteFailShots[sk]={};\n'
                        '      if(!siteFailShots[sk][pk])siteFailShots[sk][pk]=new Set();\n'
                        '      siteFailShots[sk][pk].add(shotIdx);\n'
                        '      siteFailCount[sk]=(siteFailCount[sk]||0)+1;\n'
                        '      grandTotalFail++;\n'
                        '      if(!shotFailData[shotIdx])shotFailData[shotIdx]={cnt:0,sites:{}};\n'
                        '      shotFailData[shotIdx].cnt++;\n'
                        '      shotFailData[shotIdx].sites[sk]=true;\n'
                        '      if(!_shotsSeen[shotIdx]){_shotsSeen[shotIdx]=true;shotWaferHits[shotIdx]=(shotWaferHits[shotIdx]||0)+1;}\n'
                        '    });\n'
                        '  });\n'
                        '  var _lrsl=(keys.length?_wmRetInfoFor(keys[0]):{}).retSiteLabels||WM_PAT.retSiteLabels||{};\n'
                        '  var _lrst=(keys.length?_wmRetInfoFor(keys[0]):{}).retSiteTotals||WM_PAT.retSiteTotals||{};\n'
                        '  var _siteNum=WM_PAT._retSiteNum||{};\n'
                        '  var sites=Object.keys(siteFailCount);\n'
                        '  if(!sites.length){el.innerHTML=\'<span style="color:#7f8c8d;font-size:11px">No fail dies mapped to reticle sites for selected wafers/bins.</span>\';return;}\n'
                        '  sites.sort(function(a,b){return siteFailCount[b]-siteFailCount[a];});\n'
                        '  var _clrLeg=\'<div style="margin-top:5px;font-size:10px;color:#888"><b>Color:</b> <span style="background:#fde8e8;padding:1px 4px;border-radius:2px">Red \u226570% hit</span> &nbsp; <span style="background:#fef3cd;padding:1px 4px;border-radius:2px">Yellow 40\u201369%</span></div>\';\n'
                        '  var _clrLink=(_wmPatRetUnchecked&&_wmPatRetUnchecked.size>0?\'<div style="margin-bottom:4px"><a href="#" onclick="wmPatRetClear();return false" style="color:#c0392b;font-weight:bold;font-size:10px">\u00d7 Clear highlights</a></div>\':\'\');\n'
                        '  // TABLE A: By Reticle Die Loc\n'
                        '  var h=\'<style>.wmret th{border-right:1px solid rgba(255,255,255,0.35);border-bottom:1px solid rgba(255,255,255,0.2)}.wmret td{border-right:1px solid #c8d8e8;border-bottom:1px solid #e8eef4}.wmret th:last-child,.wmret td:last-child{border-right:none}</style><div style="font-weight:bold;font-size:11px;color:#1f618d;margin:4px 0 2px;padding-bottom:2px;border-bottom:2px solid #1f618d">\u25a3 Table A \u2014 By Reticle Die Loc</div>\';\n'
                        '  h+=_clrLink;\n'
                        '  h+=\'<table class=\"wmret\" style="border-collapse:collapse;font-size:11px;width:auto;white-space:nowrap;display:block;margin-left:0"><thead><tr>\';\n'
                        '  h+=\'<th style="background:#1f618d;color:#fff;padding:2px 4px;text-align:center" title="Highlight on map">\u2611</th>\';\n'
                        '  h+=\'<th style="background:#1f618d;color:#fff;padding:2px 4px">Loc #</th>\';\n'
                        '  h+=\'<th style="background:#1f618d;color:#fff;padding:2px 4px">RX</th>\';\n'
                        '  h+=\'<th style="background:#1f618d;color:#fff;padding:2px 4px">RY</th>\';\n'
                        '  h+=\'<th style="background:#1f618d;color:#fff;padding:2px 4px">Fail Dies</th>\';\n'
                        '  h+=\'<th style="background:#1f618d;color:#fff;padding:2px 4px">%</th>\';\n'
                        '  h+=\'<th style="background:#1f618d;color:#fff;padding:2px 4px">Wafer Hits</th>\';\n'
                        '  h+=\'<th style="background:#1f618d;color:#fff;padding:2px 4px">Hit%</th>\';\n'
                        '  h+=\'<th style="background:#1f618d;color:#fff;padding:2px 4px">Shots/wfr</th></tr></thead><tbody>\';\n'
                        '  var altRow=false;\n'
                        '  sites.forEach(function(sk){\n'
                        '    var parts=sk.split(",");var rx=parts[0],ry=parts[1];\n'
                        '    var locNum=_siteNum[sk]||(_lrsl[sk]!=null?_lrsl[sk]:"-");\n'
                        '    var fc=siteFailCount[sk];\n'
                        '    var pctF=grandTotalFail>0?(fc/grandTotalFail*100).toFixed(1):"0.0";\n'
                        '    var waferHits=Object.keys(siteFailShots[sk]).length;\n'
                        '    var hitPct=nWafers>0?(waferHits/nWafers*100).toFixed(0):0;\n'
                        '    var heatPct=nWafers>0?waferHits/nWafers:0;\n'
                        '    var totShots=(_lrst[sk])||1;\n'
                        '    var bg=heatPct>=0.7?"#fde8e8":heatPct>=0.4?"#fef3cd":altRow?"#f0f4fb":"#fff";\n'
                        '    var isChk=!(_wmPatRetUnchecked&&_wmPatRetUnchecked.has(sk));\n'
                        '    var dimRow=_wmPatRetUnchecked&&_wmPatRetUnchecked.has(sk);\n'
                        '    h+=\'<tr style="background:\'+bg+\';\'+( dimRow?"opacity:0.3":"")+\'">\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:center"><input type="checkbox" data-sk="\'+sk+\'" \'+( isChk?\'checked \':\' \')+\'onchange="wmPatRetSiteToggle(this.dataset.sk,this.checked)"></td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:center;font-weight:bold;color:#1a5276">\'+locNum+\'</td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:center">\'+rx+\'</td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:center">\'+ry+\'</td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:right">\'+fc+\'</td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:right">\'+pctF+\'%</td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:right">\'+waferHits+\'/\'+nWafers+\'</td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:right;font-weight:\'+( heatPct>=0.7?"bold":"normal")+\';color:\'+( heatPct>=0.7?"#c0392b":heatPct>=0.4?"#e67e22":"#27ae60")+\'">\'+( +hitPct)+\'%</td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:right;color:#888">\'+totShots+\'</td></tr>\';\n'
                        '    altRow=!altRow;\n'
                        '  });\n'
                        '  h+=\'</tbody></table>\'+_clrLeg;\n'
                        '  // TABLE B: By Shot #\n'
                        '  var shots=Object.keys(shotFailData).sort(function(a,b){return shotFailData[b].cnt-shotFailData[a].cnt;});\n'
                        '  var _clrShotLink=(_wmPatShotUnchecked&&_wmPatShotUnchecked.size>0?\'<div style="margin-bottom:4px"><a href="#" onclick="_wmPatToggleShotAll(true);return false" style="color:#c0392b;font-weight:bold;font-size:10px">\\u00d7 Clear shot filter</a></div>\':\'\');\n'
                        '  h+=\'<div style="font-weight:bold;font-size:11px;color:#6c3483;margin:10px 0 2px;padding-bottom:2px;border-bottom:2px solid #6c3483">\\u25a3 Table B \\u2014 By Shot # (stage/scanner systematic)</div>\';\n'
                        '  h+=_clrShotLink;\n'
                        '  h+=\'<table class="wmret" style="border-collapse:collapse;font-size:11px;width:auto;white-space:nowrap;display:block;margin-left:0"><thead><tr>\';\n'
                        '  h+=\'<th style="background:#6c3483;color:#fff;padding:2px 4px;text-align:center" title="Show/hide on map">\\u2611 <a href="#" onclick="_wmPatToggleShotAll(true);return false" style="color:#dcc6f0;font-size:9px;text-decoration:none">All</a> <a href="#" onclick="_wmPatToggleShotAll(false);return false" style="color:#dcc6f0;font-size:9px;text-decoration:none">None</a></th>\';\n'
                        '  h+=\'<th style="background:#6c3483;color:#fff;padding:2px 4px">Shot #</th>\';\n'
                        '  h+=\'<th style="background:#6c3483;color:#fff;padding:2px 4px">Fail Dies</th>\';\n'
                        '  h+=\'<th style="background:#6c3483;color:#fff;padding:2px 4px">Die Locs</th>\';\n'
                        '  h+=\'<th style="background:#6c3483;color:#fff;padding:2px 4px">Wafer Hits</th>\';\n'
                        '  h+=\'<th style="background:#6c3483;color:#fff;padding:2px 4px">Hit%</th></tr></thead><tbody>\';\n'
                        '  altRow=false;\n'
                        '  shots.forEach(function(si){\n'
                        '    var sd=shotFailData[si];\n'
                        '    var wh=shotWaferHits[si]||0;\n'
                        '    var hp=nWafers>0?(wh/nWafers*100).toFixed(0):0;\n'
                        '    var heatP=nWafers>0?wh/nWafers:0;\n'
                        '    var bg2=heatP>=0.7?"#f3e5f5":heatP>=0.4?"#ede7f6":altRow?"#f9f4fc":"#fff";\n'
                        '    var locNums=Object.keys(sd.sites).map(function(sk){return _siteNum[sk]||sk;}).sort(function(a,b){return(+a)-(+b);}).join(", ");\n'
                        '    var isShotOn=!(_wmPatShotUnchecked&&_wmPatShotUnchecked.has(+si));\n'
                        '    h+=\'<tr style="background:\'+bg2+\';\'+(!isShotOn?"opacity:0.3":"")+\'">\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:center"><input type="checkbox" data-si="\'+si+\'" \'+(isShotOn?"checked ":"")+\'onchange="_wmPatShotToggle(+this.dataset.si,this.checked)"></td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:center;font-weight:bold;color:#6c3483">Shot \'+si+\'</td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:right">\'+sd.cnt+\'</td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;font-size:10px;color:#555">Loc \'+locNums+\'</td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:right">\'+wh+\'/\'+nWafers+\'</td>\';\n'
                        '    h+=\'<td style="padding:1px 4px;text-align:right;font-weight:\'+(heatP>=0.7?"bold":"normal")+\';color:\'+(heatP>=0.7?"#c0392b":heatP>=0.4?"#9b59b6":"#27ae60")+\'">\'+( +hp)+\'%</td></tr>\';\n'
                        '    altRow=!altRow;\n'
                        '  });\n'
                        '  h+=\'</tbody></table><div style="margin-top:5px;font-size:10px;color:#888"><b>Color:</b> <span style="background:#f3e5f5;padding:1px 4px;border-radius:2px">Purple \u226570% hit</span> &nbsp; <span style="background:#ede7f6;padding:1px 4px;border-radius:2px">Light 40\u201369%</span></div><div style=\"margin-top:4px;font-size:10px;color:#888;font-style:italic\">\u26a0\ufe0f Edge-ring note: Reticle die locs that map to wafer edge positions may show elevated fail rates due to normal edge yield loss, not a reticle defect \u2014 compare vs. edge exclusion zone before concluding reticle fault.</div>\';\n'
                        '  el.innerHTML=\'<div style="float:left;text-align:left">\'+h+\'</div>\';\n'
                        '}\n'
                        'function wmPatBinToggle(ibk,on){\n'
                        '  if(_wmPatBinChecked===null){\n'
                        '    _wmPatBinChecked=new Set();\n'
                        '    document.querySelectorAll("#wm-pat-binrow input[data-ib]").forEach(function(inp){if(inp.checked)_wmPatBinChecked.add(inp.dataset.ib);});\n'
                        '  }\n'
                        '  if(on){_wmPatBinChecked.add(String(ibk));}else{_wmPatBinChecked.delete(String(ibk));}\n'
                        '  _wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function _wmPatToggleBinAll(on){_wmPatBinChecked=on?null:new Set();if(on){_wmPatFbFilter={};_wmCloseFbSubRow();}_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);}\n'
                        'function wmPatWaferToggle(pk,on){\n'
                        '  if(_wmPatSelWafers===null){\n'
                        '    var keys=Object.keys(WM_PAT.wafers).filter(function(k){return _wmPatMatchLots(k)&&_wmPatMatchProgs(k);});\n'
                        '    _wmPatSelWafers=new Set(keys);\n'
                        '  }\n'
                        '  if(on){_wmPatSelWafers.add(pk);if(_wmPatCurLots!==null&&_wmPatCurLots.indexOf(_wmPatGetLot(pk))<0)_wmPatCurLots.push(_wmPatGetLot(pk));}\n  else{_wmPatSelWafers.delete(pk);}\n'
                        '  _wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function _wmPatToggleCanvasMode(){\n'
                        '  _wmPatCanvasMode=!_wmPatCanvasMode;\n'
                        '  var btn=document.getElementById("wm-pat-mode-btn");\n'
                        '  if(btn)btn.innerHTML=_wmPatCanvasMode?"&#128247; SVG mode":"&#9889; Fast mode";\n'
                        '  var maps=document.getElementById("wm-pat-maps");if(!maps)return;\n'
                        '  maps.querySelectorAll(".wm-pat-tile-ph[data-idx]").forEach(function(el){\n'
                        '    var idx=parseInt(el.dataset.idx,10),pk=_wmPatLastKeys[idx];\n'
                        '    if(pk&&_wmPatRenderedKeys.has(pk)){var tc=el.querySelector(".wm-pat-tile-content");if(tc)_wmPatRenderTile(pk,tc);}\n'
                        '  });\n'
                        '}\n'
                        'function _wmPatRenderTile(pk,container){\n'
                        '  var wdata=WM_PAT.wafers[pk];if(!wdata)return;\n'
                        '  var dies=wdata.dies?wdata.dies:wdata;if(!dies||!dies.length)return;\n'
                        '  var FIXED_W=Math.round(190*_wmPatZoom),pad=2;\n'
                        '  var xs=[],ys=[];\n'
                        '  dies.forEach(function(d){if(d[0]!==null){xs.push(d[0]);ys.push(d[1]);}});\n'
                        '  if(!xs.length)return;\n'
                        '  var xMin=Math.min.apply(null,xs),xMax=Math.max.apply(null,xs);\n'
                        '  var yMin=Math.min.apply(null,ys),yMax=Math.max.apply(null,ys);\n'
                        '  var cs=Math.max(2,(FIXED_W-pad*2)/(xMax-xMin+1));\n'
                        '  var xSpan=xMax-xMin,ySpan=yMax-yMin;\n'
                        '  var csy=(xSpan>0&&ySpan>0)?(cs*xSpan/ySpan):cs;\n'
                        '  var W=FIXED_W,H=Math.round((yMax-yMin+1)*csy+pad*2);\n'
                        '  var xCtr=(xMin+xMax)/2,yCtr=(yMin+yMax)/2;\n'
                        '  var xRad=xSpan/2||1,yRad=ySpan/2||1;\n'
                        '  var cx=(pad+(xCtr-xMin)*cs+cs*0.45).toFixed(1);\n'
                        '  var cy=(pad+(yMax-yCtr)*csy+csy*0.45).toFixed(1);\n'
                        '  var rx=(xRad*cs+cs*0.5).toFixed(1);\n'
                        '  var ry=(yRad*csy+csy*0.5).toFixed(1);\n'
                        '  var clipId="wmpc_"+pk.replace(/[^a-z0-9]/gi,"_");\n'
                        '  var _stripePatId="wmsp_"+pk.replace(/[^a-z0-9]/gi,"_");\n'
                        '  var _pkRetInfo=_wmRetInfoFor(pk);\n'
                        '  var _pkShots=(_pkRetInfo.retShots&&_pkRetInfo.retShots.length)?_pkRetInfo.retShots:WM_PAT.retShots;\n'
                        '  var failShotIdx=new Set();\n'
                        '  container.style.width=W+"px";container.style.height=H+"px";container.style.background="";container.style.borderRadius="";\n'
                        '  if(_wmPatCanvasMode){\n'
                        '    var cv=document.createElement("canvas");\n'
                        '    cv.width=W;cv.height=H;cv.style.display="block";\n'
                        '    var ctx=cv.getContext("2d");\n'
                        '    dies.forEach(function(d){\n'
                        '      var x=d[0],y=d[1],ib=d[2];if(x===null)return;\n'
                        '      var ibKey=(ib!==null&&ib!==undefined)?ib:null;\n'
                        '      var binOn=(_wmPatBinChecked===null||_wmPatBinChecked.has(String(ibKey)));\n'
                        '      if(binOn&&ibKey!==null){var _fbF=_wmPatFbFilter[String(ibKey)];if(_fbF!==undefined&&_fbF!==null){var _fv=d[3]!==undefined&&d[3]!==null?String(d[3]):null;if(_fv===null||!_fbF.has(_fv))binOn=false;}}\n'
                        '      ctx.globalAlpha=binOn?1:0.08;\n'
                        '      ctx.fillStyle=_wmIbColor(ibKey);\n'
                        '      ctx.fillRect(pad+(x-xMin)*cs,pad+(yMax-y)*csy,cs*0.9,csy*0.9);\n'
                        '      if(_wmIsFail(ibKey)&&ibKey!==null&&binOn&&WM_PAT.hasReticle){var _ri2=_pkRetInfo.retMap&&_pkRetInfo.retMap[x+","+y];if(_ri2)failShotIdx.add(_ri2[2]);}\n'
                        '    });\n'
                        '    ctx.globalAlpha=1;\n'
                        '    if(WM_PAT.hasReticle&&_pkShots&&_pkShots.length){\n'
                        '      ctx.lineWidth=0.7;\n'
                        '      _pkShots.forEach(function(s,si){\n'
                        '        var sx=pad+(s[0]-xMin)*cs,sy=pad+(yMax-s[3])*csy,sw=(s[2]-s[0]+1)*cs,sh=(s[3]-s[1]+1)*csy;\n'
                        '        ctx.strokeStyle=failShotIdx.has(si)?"#c0392b":"#2471a3";\n'
                        '        ctx.lineWidth=failShotIdx.has(si)?1.5:0.7;\n'
                        '        ctx.globalAlpha=failShotIdx.has(si)?0.9:0.35;\n'
                        '        ctx.strokeRect(sx,sy,sw,sh);ctx.globalAlpha=1;\n'
                        '      });\n'
                        '    }\n'
                        '    ctx.strokeStyle="#bdc3c7";ctx.lineWidth=1.5;\n'
                        '    ctx.beginPath();ctx.ellipse(parseFloat(cx),parseFloat(cy),parseFloat(rx),parseFloat(ry),0,0,2*Math.PI);ctx.stroke();\n'
                        '    /* Die-loc numbers */\n'
                        '    if(WM_PAT.hasReticle&&WM_PAT._retSiteNum&&cs>=4){\n'
                        '      var _dlFs=Math.max(4,Math.min(7,Math.round(cs*0.55)));\n'
                        '      ctx.font="bold "+_dlFs+"px Arial";ctx.textAlign="right";ctx.textBaseline="top";\n'
                        '      dies.forEach(function(d){\n'
                        '        var x=d[0],y=d[1];if(x===null)return;\n'
                        '        var dtInf=_pkRetInfo.retMap&&_pkRetInfo.retMap[x+","+y];if(!dtInf)return;\n'
                        '        var dtTag=String(WM_PAT._retSiteNum[dtInf[0]+","+dtInf[1]]||"");if(!dtTag)return;\n'
                        '        ctx.fillStyle="#000";ctx.globalAlpha=1;\n'
                        '        ctx.fillText(dtTag,pad+(x-xMin)*cs+cs-0.5,pad+(yMax-y)*csy+0.5);\n'
                        '      });\n'
                        '    }\n'
                        '    /* Hover: build die lookup then attach listener once */\n'
                        '    cv._dl={};cv._xMn=xMin;cv._yMx=yMax;cv._cs=cs;cv._csy=csy;cv._pad=pad;\n'
                        '    dies.forEach(function(d){\n'
                        '      var x=d[0],y=d[1];if(x===null)return;\n'
                        '      var ib=(d[2]!==null&&d[2]!==undefined)?d[2]:null;\n'
                        '      var fb=(d[3]!==undefined&&d[3]!==null)?d[3]:"";\n'
                        '      var uv=(WM_PAT.upmCols&&WM_PAT.upmCols.length)?WM_PAT.upmCols.map(function(_,i){var v=d[4+i];return(v!==undefined&&v!==null)?v:"";}).join("|"):"";\n'
                        '      cv._dl[x+","+y]={ib:ib,fb:fb,uv:uv,x:x,y:y};\n'
                        '    });\n'
                        '    if(!cv._hvBound){cv._hvBound=true;\n'
                        '      cv.addEventListener("mousemove",function(e){\n'
                        '        var r=cv.getBoundingClientRect(),sx=cv.width/r.width,sy=cv.height/r.height;\n'
                        '        var cx2=(e.clientX-r.left)*sx,cy2=(e.clientY-r.top)*sy;\n'
                        '        var dx=Math.round(cv._xMn+(cx2-cv._pad)/cv._cs),dy=Math.round(cv._yMx-(cy2-cv._pad)/cv._csy);\n'
                        '        var dd=cv._dl&&cv._dl[dx+","+dy];\n'
                        '        if(!dd){_wmDieHoverOut();return;}\n'
                        '        var tt=document.getElementById("wm-die-tt");\n'
                        '        if(!tt){tt=document.createElement("div");tt.id="wm-die-tt";tt.style.cssText="position:fixed;z-index:99998;background:rgba(30,30,30,0.92);color:#fff;font-size:11px;font-family:Arial,sans-serif;padding:5px 9px;border-radius:5px;pointer-events:none;white-space:nowrap;box-shadow:0 2px 8px rgba(0,0,0,.4);line-height:1.6";document.body.appendChild(tt);}\n'
                        '        var ibSw=(WM_PAT.ibColors&&WM_PAT.ibColors[String(dd.ib)])?"<span style=\'display:inline-block;width:8px;height:8px;background:"+WM_PAT.ibColors[String(dd.ib)]+";border-radius:2px;margin-right:3px;vertical-align:middle\'></span>":"";\n'
                        '        var fbLine=dd.fb!==""?"<br><b>FB:</b> "+dd.fb:"";\n'
                        '        var upmLine="";if(WM_PAT.upmCols&&WM_PAT.upmCols.length&&dd.uv){var _uE=WM_PAT.upmCols[0],_uT=Array.isArray(_uE)?_uE[1]:null,_uV=parseFloat(dd.uv.split("|")[0]);if(!isNaN(_uV))upmLine="<br><b>UPM:</b> "+(_uT?(_uV/_uT*100).toFixed(1)+"%":_uV);}\n'
                        '        tt.innerHTML="<b>X:</b> "+dd.x+" &nbsp; <b>Y:</b> "+dd.y+"<br>"+ibSw+"<b>IB:</b> "+dd.ib+fbLine+upmLine;\n'
                        '        var m=12,tW=tt.offsetWidth||140,tH=tt.offsetHeight||52;\n'
                        '        var lft=e.clientX+m,top=e.clientY-tH-m;\n'
                        '        if(lft+tW>window.innerWidth)lft=e.clientX-tW-m;if(top<4)top=e.clientY+m;\n'
                        '        tt.style.left=lft+"px";tt.style.top=top+"px";tt.style.display="";\n'
                        '      });\n'
                        '      cv.addEventListener("mouseleave",function(){_wmDieHoverOut();});\n'
                        '    }\n'
                        '    container.innerHTML="";container.appendChild(cv);\n'
                        '  } else {\n'
                        '    var rects=[],retOut="";\n'
                        '    var _hasFbFilter=Object.keys(_wmPatFbFilter).some(function(k){return _wmPatFbFilter[k]!==null;});\n'
                        '    dies.forEach(function(d){\n'
                        '      var x=d[0],y=d[1],ib=d[2];if(x===null)return;\n'
                        '      var px=(pad+(x-xMin)*cs).toFixed(1),py=(pad+(yMax-y)*csy).toFixed(1);\n'
                        '      var ibKey=(ib!==null&&ib!==undefined)?ib:null;\n'
                        '      var fill=_wmIbColor(ibKey);\n'
                        '      var binOn=(_wmPatBinChecked===null||_wmPatBinChecked.has(String(ibKey)));\n'
                        '      if(binOn&&ibKey!==null){var _fbF=_wmPatFbFilter[String(ibKey)];if(_fbF!==undefined&&_fbF!==null){var _fv2=d[3]!==undefined&&d[3]!==null?String(d[3]):null;if(_fv2===null||!_fbF.has(_fv2))binOn=false;}}\n'
                        '      var opacity=binOn?"1":"0";\n'
                        '      if(_wmUpmOverlay){\n'
                        '        var _ue=WM_PAT.upmCols&&WM_PAT.upmCols[0];var _ut=Array.isArray(_ue)?_ue[1]:null;\n'
                        '        var _ur=d[4]!==undefined&&d[4]!==null?d[4]:null;\n'
                        '        if(_wmIsFail(ibKey)){fill=(_ur!==null&&_ut)?_wmUpmColor(_ur/_ut*100):"#c8c8c8";opacity="1";}\n'
                        '        else if(_ur!==null&&_ut){fill=_wmUpmColor(_ur/_ut*100);opacity="1";}\n'
                        '        else{fill="#c8c8c8";opacity="0.5";}\n'
                        '      }\n'
                        '      if(binOn&&_wmPatRetUnchecked&&_wmPatRetUnchecked.size>0&&WM_PAT.hasReticle){var _dri=_pkRetInfo.retMap&&_pkRetInfo.retMap[x+","+y];if(_dri&&_wmPatRetUnchecked.has(_dri[0]+","+_dri[1]))opacity="0.12";}\n'
                        '      if(binOn&&_wmPatShotUnchecked&&_wmPatShotUnchecked.size>0&&WM_PAT.hasReticle){var _sri=_pkRetInfo.retMap&&_pkRetInfo.retMap[x+","+y];if(_sri&&_wmPatShotUnchecked.has(_sri[2]))opacity="0.08";}\n'
                        '      if(_wmIsFail(ibKey)&&ibKey!==null&&binOn&&WM_PAT.hasReticle){var _ri3=_pkRetInfo.retMap&&_pkRetInfo.retMap[x+","+y];if(_ri3)failShotIdx.add(_ri3[2]);}\n'
                        '      var _isEdge=(_wmEdgeExcRows>0&&(x<xMin+_wmEdgeExcRows||x>xMax-_wmEdgeExcRows||y<yMin+_wmEdgeExcRows||y>yMax-_wmEdgeExcRows));if(_isEdge)opacity="0.15";\n'
                        '      var _fbVal2=d[3]!==undefined&&d[3]!==null?d[3]:"";\n'
                        '      var _upmVals2=(WM_PAT.upmCols&&WM_PAT.upmCols.length)?WM_PAT.upmCols.map(function(_,_ui){var _v=d[4+_ui];return(_v!==undefined&&_v!==null)?_v:"";}).join("|"):"";\n'
                        '      rects.push(\'<rect x="\'+px+\'" y="\'+py+\'" width="\'+(cs*0.9).toFixed(1)+\'" height="\'+(csy*0.9).toFixed(1)+\'" fill="\'+fill+\'" opacity="\'+opacity+\'" data-ib="\'+ibKey+\'" data-op="\'+opacity+\'" data-x="\'+x+\'" data-y="\'+y+\'" data-fb="\'+_fbVal2+\'" data-upm="\'+_upmVals2+\'"/>\');\n'
                        '      if(_wmUpmOverlay&&_wmPatBinChecked!==null&&_wmPatBinChecked.has(String(ibKey))){rects.push(\'<rect x="\'+px+\'" y="\'+py+\'" width="\'+(cs*0.9).toFixed(1)+\'" height="\'+(csy*0.9).toFixed(1)+\'" fill="rgba(220,0,0,0.85)" pointer-events="none"/>\');}\n'
                        '      var _dieTag="";\n'
                        '      if(WM_PAT.hasReticle&&WM_PAT._retSiteNum){var _dtMap=_pkRetInfo.retMap;var _dtInf=_dtMap&&_dtMap[x+","+y];if(_dtInf){var _dtSk=_dtInf[0]+","+_dtInf[1];_dieTag=String(WM_PAT._retSiteNum[_dtSk]||"");}}\n'
                        '      var _tagFs=Math.max(4,Math.min(7,Math.round(cs*0.55)));\n'
                        '      if(_dieTag&&cs>=4){rects.push(\'<text x="\'+(parseFloat(px)+cs-0.5).toFixed(1)+\'" y="\'+(parseFloat(py)+_tagFs+0.5).toFixed(1)+\'" text-anchor="end" font-size="\'+_tagFs+\'" fill="#000" font-weight="bold" opacity="1" data-ib="\'+ibKey+\'" data-op="\'+opacity+\'" pointer-events="none">\'+_dieTag+\'</text>\');}\n'
                        '    });\n'
                        '    if(WM_PAT.hasReticle&&_pkShots&&_pkShots.length){\n'
                        '      _pkShots.forEach(function(s,si){\n'
                        '        var sx=(pad+(s[0]-xMin)*cs).toFixed(1),sy=(pad+(yMax-s[3])*csy).toFixed(1),sw=((s[2]-s[0]+1)*cs).toFixed(1),sh=((s[3]-s[1]+1)*csy).toFixed(1);\n'
                        '        retOut+=\'<rect x="\'+sx+\'" y="\'+sy+\'" width="\'+sw+\'" height="\'+sh+\'" fill="none" stroke="#2471a3" stroke-width="0.7" opacity="0.35"/>\';\n'
                        '        if(cs>=6){var tx=(+sx+(+sw)/2).toFixed(1),ty=(+sy+7).toFixed(1);retOut+=\'<text x="\'+tx+\'" y="\'+ty+\'" text-anchor="middle" font-size="5" fill="#2471a3" opacity="0.7" pointer-events="none">\'+si+\'</text>\';}\n'
                        '      });\n'
                        '      _pkShots.forEach(function(s,si){\n'
                        '        if(!failShotIdx.has(si))return;\n'
                        '        var sx=(pad+(s[0]-xMin)*cs).toFixed(1),sy=(pad+(yMax-s[3])*csy).toFixed(1),sw=((s[2]-s[0]+1)*cs).toFixed(1),sh=((s[3]-s[1]+1)*csy).toFixed(1);\n'
                        '        retOut+=\'<rect x="\'+sx+\'" y="\'+sy+\'" width="\'+sw+\'" height="\'+sh+\'" fill="none" stroke="#c0392b" stroke-width="1.5" opacity="0.9"/>\';\n'
                        '      });\n'
                        '    }\n'
                        '    container.innerHTML=\'<svg width="\'+W+\'" height="\'+H+\'" style="display:block;margin:0 auto;cursor:crosshair" onmousemove="_wmDieHover(event,this)" onmouseleave="_wmDieHoverOut()">\'\n'
                        '      +\'<defs><clipPath id="\'+clipId+\'"><ellipse cx="\'+cx+\'" cy="\'+cy+\'" rx="\'+rx+\'" ry="\'+ry+\'"/></clipPath>\'\n'
                        '      +\'<pattern id="\'+_stripePatId+\'" patternUnits="userSpaceOnUse" width="8" height="8" patternTransform="rotate(45)"><rect width="4" height="8" fill="rgba(255,255,255,0.9)"/><rect x="4" width="4" height="8" fill="rgba(0,0,0,0.9)"/></pattern></defs>\'\n'
                        '      +\'<g clip-path="url(#\'+clipId+\')">\'+rects.join("")+retOut+\'</g>\'\n'
                        '      +\'<ellipse cx="\'+cx+\'" cy="\'+cy+\'" rx="\'+rx+\'" ry="\'+ry+\'" fill="none" stroke="#bdc3c7" stroke-width="1.5"/></svg>\';\n'
                        '  }\n'
                        '  _wmPatRenderedKeys.add(pk);\n'
                        '}\n'
                        'function wmShowPatLot(lots){\n'
                        '  if(typeof lots==="string")lots=[lots];\n'
                        '  _wmPatCurLots=lots;\n'
                        '  _wmPatCurProgs=null;\n'
                        '  _wmPatSelWafers=null;\n'
                        '  _wmPatBinChecked=null;\n'
                        '  _wmPatFbFilter={};\n'
                        '  _wmCloseFbSubRow();\n'
                        '  _wmPatRetUnchecked=null;\n'
                        '  _wmDdUpdateBtns();\n'
                        '  _wmPatBuildLotPicker();\n'
                        '  var ov=document.getElementById("wm-pat-overlay");\n'
                        '  if(ov)ov.classList.add("open");\n'
                        '  _wmPatRender();\n'
                        '  _wmPatInitDrag();\n'
                        '}\n'
                        'function _wmPatBuildProgPicker(){_wmDdUpdateBtns();_wmDdRefresh();}\n'
                        'function _wmPatProgToggle(p,on){\n'
                        '  if(!_wmPatCurProgs)_wmPatCurProgs=_wmPatAllProgs().slice();\n'
                        '  if(on){if(_wmPatCurProgs.indexOf(p)<0)_wmPatCurProgs.push(p);}else{_wmPatCurProgs=_wmPatCurProgs.filter(function(x){return x!==p;});}\n'
                        '  if(!on&&_wmPatSelWafers){var rem=[];_wmPatSelWafers.forEach(function(pk){if(_wmPatGetProg(pk)===p)rem.push(pk);});rem.forEach(function(pk){_wmPatSelWafers.delete(pk);});}\n'
                        '  _wmDdUpdateBtns();_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function _wmPatProgAll(){_wmPatCurProgs=null;_wmPatSelWafers=null;_wmDdUpdateBtns();_wmDdRefresh();_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);}\n'
                        'function _wmPatProgNone(){_wmPatCurProgs=[];_wmPatSelWafers=new Set();_wmDdUpdateBtns();_wmDdRefresh();_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);}\n'
                        'function _wmPatBuildLotPicker(){_wmDdUpdateBtns();_wmDdRefresh();}\n'
                        'function _wmPatBuildWaferPicker(){_wmDdUpdateBtns();_wmDdRefresh();}\n'
                        'function _wmDdLotCbChange(cb){\n'
                        '  _wmPatLotToggle(cb.dataset.lot,cb.checked);\n'
                                                '}\n'
                        'function _wmPatLotToggle(lt,on){\n'
                        '  if(!_wmPatCurLots)_wmPatCurLots=_wmPatAllLots().slice();\n'
                        '  if(on){\n'
                        '    if(_wmPatCurLots.indexOf(lt)<0)_wmPatCurLots.push(lt);\n'
                        '    /* If in specific-wafer mode, add all wafers of this lot so they appear */\n'
                        '    if(_wmPatSelWafers!==null){Object.keys(WM_PAT.wafers).forEach(function(pk){if(_wmPatGetLot(pk)===lt)_wmPatSelWafers.add(pk);});}\n'
                        '  }else{\n'
                        '    _wmPatCurLots=_wmPatCurLots.filter(function(x){return x!==lt;});\n'
                        '    if(_wmPatSelWafers){\n'
                        '      var remove=[];_wmPatSelWafers.forEach(function(pk){if(_wmPatGetLot(pk)===lt)remove.push(pk);});\n'
                        '      remove.forEach(function(pk){_wmPatSelWafers.delete(pk);});\n'
                        '    }\n'
                        '  }\n'
                        '  _wmDdUpdateBtns();_wmDdSetIndeterminate();_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function _wmPatLotAll(){_wmPatCurLots=null;_wmPatSelWafers=null;_wmDdUpdateBtns();_wmDdRefresh();_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);}\n'
                        'function _wmPatLotNone(){_wmPatCurLots=[];_wmPatSelWafers=new Set();_wmDdUpdateBtns();_wmDdRefresh();_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);}\n'
                        'function _wmPatWaferAll(){_wmPatSelWafers=null;_wmDdUpdateBtns();_wmDdRefresh();_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);}\n'
                        'function _wmPatWaferNone(){_wmPatSelWafers=new Set();_wmDdUpdateBtns();_wmDdRefresh();_wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);}\n'
                        'var _wmDdCur=null;\n'
                        'function _wmDdOpen(type,btn){\n'
                        '  var panel=document.getElementById("wm-dd-panel");\n'
                        '  if(!panel)return;\n'
                        '  if(_wmDdCur===type){_wmDdClose();return;}\n'
                        '  _wmDdClose();\n'
                        '  _wmDdCur=type;\n'
                        '  var body=document.getElementById("wm-dd-body");\n'
                        '  if(!body)return;\n'
                        '  if(type==="prog")body.innerHTML=_wmDdBuildProg();\n'
                        '  else if(type==="lotwafer"){body.innerHTML=_wmDdBuildLotWafer();_wmDdSetIndeterminate();}\n'
                        '  var r=btn.getBoundingClientRect();\n'
                        '  panel.style.left=Math.max(4,Math.min(r.left,window.innerWidth-240))+"px";\n'
                        '  panel.style.top=(r.bottom+4)+"px";\n'
                        '  panel.style.display="flex";\n'
                        '  var s=panel.querySelector(".wm-dd-search");if(s){s.value="";s.focus();}\n'
                        '  document.querySelectorAll(".wm-dd-btn").forEach(function(b){b.classList.remove("open");});\n'
                        '  btn.classList.add("open");\n'
                        '  setTimeout(function(){document.addEventListener("mousedown",_wmDdOutside);},0);\n'
                        '}\n'
                        'function _wmDdClose(){\n'
                        '  var panel=document.getElementById("wm-dd-panel");\n'
                        '  if(panel)panel.style.display="none";\n'
                        '  document.removeEventListener("mousedown",_wmDdOutside);\n'
                        '  document.querySelectorAll(".wm-dd-btn").forEach(function(b){b.classList.remove("open");});\n'
                        '  _wmDdCur=null;\n'
                        '}\n'
                        'function _wmDdOutside(e){\n'
                        '  var panel=document.getElementById("wm-dd-panel");\n'
                        '  if(!panel||panel.contains(e.target))return;\n'
                        '  if(_wmDdCur){var b=document.getElementById("wm-dd-"+_wmDdCur+"-btn");if(b&&b.contains(e.target))return;}\n'
                        '  _wmDdClose();\n'
                        '}\n'
                        'function _wmDdSearch(q){\n'
                        '  var body=document.getElementById("wm-dd-body");if(!body)return;\n'
                        '  var ql=q.toLowerCase();\n'
                        '  if(_wmDdCur==="lotwafer"||_wmDdCur==="wafer"){\n'
                        '    body.querySelectorAll(".wm-dd-lg").forEach(function(grp){\n'
                        '      var lot=(grp.dataset.lot||"").toLowerCase();\n'
                        '      var lotMatch=!ql||lot.indexOf(ql)>=0;\n'
                        '      var vis=0;\n'
                        '      grp.querySelectorAll(".wm-dd-wi").forEach(function(wi){\n'
                        '        var wn=(wi.dataset.wn||"").toLowerCase();\n'
                        '        var show=!ql||lotMatch||wn.indexOf(ql)>=0;\n'
                        '        wi.style.display=show?"":"none";\n'
                        '        if(show)vis++;\n'
                        '      });\n'
                        '      grp.style.display=(vis||!ql||lotMatch)?"":"none";\n'
                        '      if(ql&&vis){var lc=grp.querySelector(".wm-dd-lc");if(lc)lc.style.display="block";}\n'
                        '    });\n'
                        '  }else{\n'
                        '    body.querySelectorAll(".wm-dd-item").forEach(function(item){\n'
                        '      var show=!ql||(item.dataset.val||"").toLowerCase().indexOf(ql)>=0;\n'
                        '      item.style.display=show?"":"none";\n'
                        '    });\n'
                        '  }\n'
                        '}\n'
                        'function _wmDdBuildProg(){\n'
                        '  var all=_wmPatAllProgs();\n'
                        '  var h=\'<div class="wm-dd-acts"><span onclick="_wmPatProgAll()">All</span><span onclick="_wmPatProgNone()">None</span></div>\';\n'
                        '  all.forEach(function(p){\n'
                        '    var on=!_wmPatCurProgs||_wmPatCurProgs.indexOf(p)>=0;\n'
                        '    h+=\'<label class="wm-dd-item" data-val="\'+p+\'"><input type="checkbox" data-prog="\'+p+\'"\'+(on?\' checked\':\'\')+\' onchange="_wmPatProgToggle(this.dataset.prog,this.checked)"> \'+p+\'</label>\';\n'
                        '  });\n'
                        '  return h;\n'
                        '}\n'
                        'function _wmDdBuildLotWafer(){\n'
                        '  var allLots=_wmPatAllLots();\n'
                        '  var activeProgs=_wmPatCurProgs;\n'
                        '  var multiProg=_wmPatAllProgs().length>1;\n'
                        '  var wKeys=Object.keys(WM_PAT.wafers).sort(function(a,b){\n'
                        '    var la=_wmPatGetLot(a),lb=_wmPatGetLot(b);if(la!==lb)return la<lb?-1:1;\n'
                        '    var wa=parseInt(_wmPatGetWfr(a))||0,wb=parseInt(_wmPatGetWfr(b))||0;if(wa!==wb)return wa-wb;\n'
                        '    return(_wmPatGetProg(a)<_wmPatGetProg(b)?-1:1);\n'
                        '  });\n'
                        '  var lotWafers={};\n'
                        '  wKeys.forEach(function(pk){\n'
                        '    var lt=_wmPatGetLot(pk);\n'
                        '    if(activeProgs&&activeProgs.length){var pg=_wmPatGetProg(pk);if(pg&&activeProgs.indexOf(pg)<0)return;}\n'
                        '    if(!lotWafers[lt])lotWafers[lt]=[];\n'
                        '    lotWafers[lt].push(pk);\n'
                        '  });\n'
                        '  var h=\'<div class="wm-dd-acts"><span onclick="_wmPatLotAll();_wmPatWaferAll()">All</span><span onclick="_wmPatLotNone();_wmPatWaferNone()">None</span></div>\';\n'
                        '  allLots.forEach(function(lt){\n'
                        '    var pks=lotWafers[lt]||[];\n'
                        '    var lotOn=!_wmPatCurLots||_wmPatCurLots.indexOf(lt)>=0;\n'
                        '    var _lotMats=[];pks.forEach(function(pk){var m=(WM_PAT.wafers[pk]&&WM_PAT.wafers[pk].material)||"";if(m&&_lotMats.indexOf(m)<0)_lotMats.push(m);});\n'
                        '    var _mFirst=_lotMats[0]||"",_mAll=_lotMats.join(", ");\n'
                        '    var _matLbl=_mFirst?\'<span style="font-size:9px;color:#8e6a2a;margin-left:4px;font-weight:normal" title="\'+_mAll+\'">[\'+(  _mFirst.length>20?_mFirst.slice(0,19)+"\u2026":_mFirst)+(_lotMats.length>1?"+"+(_lotMats.length-1):"")+\']</span>\':\'\';\n'
                        '    h+=\'<div class="wm-dd-lg" data-lot="\'+lt+\'">\'+\n'
                        '       \'<div class="wm-dd-lhdr" style="display:flex;align-items:center">\'+\n'
                        '       \'<input type="checkbox" data-lot="\'+lt+\'"\'+(lotOn?\' checked\':\'\')+\' onclick="event.stopPropagation();_wmDdLotCbChange(this)" style="margin:0 6px 0 2px;flex-shrink:0">\'+\n'
                        '       lt+_matLbl+\'<span style="font-size:10px;color:#888;margin-left:4px">(\'+pks.length+\')</span>\'+\n'
                        '       (pks.length?\'<span class="wm-dd-exp" onclick="event.stopPropagation();_wmDdLotToggle(this.parentElement)" title="Show/hide wafers" style="margin-left:auto;padding:0 6px;cursor:pointer;color:#2471a3;font-weight:bold;font-size:13px">&#9654;</span>\':\'\')+\n'
                        '       \'</div>\'+\n'
                        '       \'<div class="wm-dd-lc" style="display:none">\';\n'
                        '    pks.forEach(function(pk){\n'
                        '      var wn=_wmPatGetWfr(pk),pg=_wmPatGetProg(pk);\n'
                        '      var on=_wmPatSelWafers===null||_wmPatSelWafers.has(pk);\n'
                        '      var mat=(WM_PAT.wafers[pk]&&WM_PAT.wafers[pk].material)||"";\n'
                        '      var lbl=\'W\'+wn+(multiProg&&pg?\'<sup style="font-size:9px">\'+pg+\'</sup>\':\'\')+(mat?\'<span style="font-size:9px;color:#8e6a2a;margin-left:3px">[\'+mat+\']</span>\':\'\')\n'
                        '      h+=\'<label class="wm-dd-wi" data-pk="\'+pk+\'" data-wn="W\'+wn+\'" title="\'+(mat?mat:\'\')+\'"><input type="checkbox" data-pk="\'+pk+\'"\'+(on?\' checked\':\'\')+\' onchange="wmPatWaferToggle(this.dataset.pk,this.checked);_wmDdUpdateBtns();_wmDdSetIndeterminate()"> \'+lbl+\'</label>\';\n'
                        '    });\n'
                        '    h+=\'</div></div>\';\n'
                        '  });\n'
                        '  return h;\n'
                        '}\n'
                        'function _wmDdLotToggle(lhdr){\n'
                        '  var lc=lhdr.nextElementSibling;\n'
                        '  var exp=lhdr.querySelector(".wm-dd-exp");\n'
                        '  if(!lc)return;\n'
                        '  var open=lc.style.display!=="none";\n'
                        '  lc.style.display=open?"none":"block";\n'
                        '  if(exp)exp.innerHTML=open?"&#9654;":"&#9660;";\n'
                        '}\n'
                        'function _wmDdLotCheckAll(grp,on){\n'
                        '  var ltCb=grp.querySelector(".wm-dd-lhdr input[data-lot]");\n'
                        '  if(ltCb){ltCb.checked=on;_wmPatLotToggle(ltCb.dataset.lot,on);}\n'
                        '  grp.querySelectorAll("input[data-pk]").forEach(function(inp){\n'
                        '    inp.checked=on;\n'
                        '    wmPatWaferToggle(inp.dataset.pk,on);\n'
                        '  });\n'
                        '  _wmDdUpdateBtns();_wmDdSetIndeterminate();\n'
                        '}\n'
                        'function _wmDdSetIndeterminate(){\n'
                        '  var body=document.getElementById("wm-dd-body");if(!body)return;\n'
                        '  body.querySelectorAll(".wm-dd-lg").forEach(function(grp){\n'
                        '    var lt=grp.dataset.lot;\n'
                        '    var lotActive=!_wmPatCurLots||_wmPatCurLots.indexOf(lt)>=0;\n'
                        '    var hdr=grp.querySelector(".wm-dd-lhdr input");\n'
                        '    if(!hdr)return;\n'
                        '    if(!lotActive){hdr.checked=false;hdr.indeterminate=false;return;}\n'
                        '    var pks=[];grp.querySelectorAll("input[data-pk]").forEach(function(inp){pks.push(inp.dataset.pk);});\n'
                        '    if(!pks.length){hdr.checked=true;hdr.indeterminate=false;return;}\n'
                        '    var allOn=pks.every(function(pk){return _wmPatSelWafers===null||_wmPatSelWafers.has(pk);});\n'
                        '    var someOn=!allOn&&pks.some(function(pk){return _wmPatSelWafers===null||_wmPatSelWafers.has(pk);});\n'
                        '    hdr.checked=allOn;hdr.indeterminate=someOn;\n'
                        '  });\n'
                        '}\n'
                        'function _wmDdRefresh(){\n'
                        '  var panel=document.getElementById("wm-dd-panel");if(!panel||panel.style.display==="none")return;\n'
                        '  var body=document.getElementById("wm-dd-body");if(!body)return;\n'
                        '  var sv=panel.querySelector(".wm-dd-search");var q=sv?sv.value:"";\n'
                        '  if(_wmDdCur==="prog")body.innerHTML=_wmDdBuildProg();\n'
                        '  else if(_wmDdCur==="lotwafer"){body.innerHTML=_wmDdBuildLotWafer();_wmDdSetIndeterminate();}\n'
                        '  if(q)_wmDdSearch(q);\n'
                        '}\n'
                        'function _wmDdUpdateBtns(){\n'
                        '  var allProgs=_wmPatAllProgs(),allLots=_wmPatAllLots();\n'
                        '  var pb=document.getElementById("wm-dd-prog-btn");\n'
                        '  if(pb){if(!allProgs.length){pb.style.display="none";}else{pb.style.display="";var sp=_wmPatCurProgs?_wmPatCurProgs.length:allProgs.length;pb.textContent="Programs ("+sp+"/"+allProgs.length+") ▼";}}\n'
                        '  var lwb=document.getElementById("wm-dd-lotwafer-btn");\n'
                        '  if(lwb){lwb.style.display="";var sl=_wmPatCurLots?_wmPatCurLots.length:allLots.length;var actLots=_wmPatCurLots||allLots;var total=Object.keys(WM_PAT.wafers).filter(function(k){return actLots.indexOf(_wmPatGetLot(k))>=0;}).length;var sel=_wmPatSelWafers===null?total:_wmPatSelWafers.size;lwb.textContent="Lots/Wafers ("+sl+"/"+allLots.length+" lots \u00b7 "+sel+"/"+total+" wfrs) \u25bc";}\n'
                        '}\n'
'function _wmPatBuildCtrl(allKeys){\n'
                        '  var ctrl=document.getElementById("wm-pat-ctrl");if(!ctrl)return;\n'
                        '  ctrl.innerHTML="";\n'
                        '  _wmPatBuildRetRow(allKeys);\n'
                        '  _wmPatBuildShotRow(allKeys);\n'
                        '}\n'
                        'function _wmPatBuildBinRow(ibArr){\n'
                        '  var br=document.getElementById("wm-pat-binrow");\n'
                        '  if(!br||!ibArr.length){if(br)br.innerHTML="";_wmCloseFbSubRow();return;}\n'
                        '  var h=\'<span style="font-size:11px;font-weight:bold;color:#5d6d7e;flex-shrink:0;margin-right:4px">IB Filter:</span>\';\n'
                        '  ibArr.forEach(function(ibk){\n'
                        '    var col=_wmIbColor(ibk);\n'
                        '    var on=_wmPatBinChecked===null||_wmPatBinChecked.has(String(ibk));\n'
                        '    var _ibTag={2:"R",3:"AR",4:"CR"}[ibk]||"";\n'
                        '    var _swStyle="background:"+col+";cursor:pointer;";\n'
                        '    var _sup=_ibTag?"<sup style=\\"font-size:8px;color:#27ae60;font-weight:bold;vertical-align:super\\">"+_ibTag+"</sup>":"";\n'
                        '    var ibLabel="IB"+ibk+_sup;\n'
                        '    var hasFb=Object.keys(WM_PAT.wafers||{}).some(function(pk){var w=WM_PAT.wafers[pk];return w&&w.ibToFb&&w.ibToFb[String(ibk)]&&Object.keys(w.ibToFb[String(ibk)]).length>0;});\n'
                        '    var fbActive=_wmPatFbFilter[String(ibk)]!==undefined&&_wmPatFbFilter[String(ibk)]!==null;\n'
                        '    var arrowBtn=hasFb?\'<span title="Show FB breakdown for IB\'+ibk+\'" onclick="_wmToggleFbSubRow(\'+ibk+\');event.stopPropagation();" style="cursor:pointer;font-size:10px;color:\'+(fbActive?"#c0392b":"#2471a3")+\';margin-left:2px;padding:0 2px;border-radius:2px;background:\'+(fbActive?"#fdecea":"#eaf0fb")+\'" data-fb-arrow="\'+ibk+\'">\'+(fbActive?"&#9654;*":"&#9654;")+\'</span>\':"";\n'
                        '    h+=\'<label class="wm-pat-bincb"><span class="wm-pat-binsw" onclick="wmmIbHlClick(\'+ibk+\');event.stopPropagation();" style="\'+_swStyle+\'"></span><input type="checkbox"\'+(on?" checked":"")+\' data-ib="\'+ibk+\'" onchange="wmPatBinToggle(+this.dataset.ib,this.checked)">\'+ibLabel+\'</label>\'+arrowBtn;\n'
                        '  });\n'
                        '  h+=\'<span style="font-size:10px;color:#2471a3;cursor:pointer;text-decoration:underline;margin-left:6px" onclick="_wmPatToggleBinAll(true)">All</span>\';\n'
                        '  h+=\'<span style="font-size:10px;color:#2471a3;cursor:pointer;text-decoration:underline;margin-left:4px" onclick="_wmPatToggleBinAll(false)">None</span>\';\n'
                        '  br.innerHTML=h;\n'
                        '  _wmPatBuildBinRow.lastArr=ibArr;\n'
                        '}\n'
                        'var _wmPatFbFilter={};\n'
                        'var _wmFbSubRowIb=null;\n'
                        'function _wmToggleFbSubRow(ibk){\n'
                        '  if(_wmFbSubRowIb===ibk){_wmCloseFbSubRow();return;}\n'
                        '  _wmOpenFbSubRow(ibk);\n'
                        '}\n'
                        'function _wmCloseFbSubRow(){\n'
                        '  _wmFbSubRowIb=null;\n'
                        '  var fr=document.getElementById("wm-pat-fbrow");if(fr){fr.style.display="none";fr.innerHTML="";}\n'
                        '}\n'
                        'function _wmOpenFbSubRow(ibk){\n'
                        '  _wmFbSubRowIb=ibk;\n'
                        '  var fr=document.getElementById("wm-pat-fbrow");if(!fr)return;\n'
                        '  var fbTotals={};\n'
                        '  Object.keys(WM_PAT.wafers||{}).forEach(function(pk){\n'
                        '    var w=WM_PAT.wafers[pk];if(!w||!w.ibToFb)return;\n'
                        '    var fbMap=w.ibToFb[String(ibk)]||{};\n'
                        '    Object.keys(fbMap).forEach(function(fb){fbTotals[fb]=(fbTotals[fb]||0)+fbMap[fb];});\n'
                        '  });\n'
                        '  var fbKeys=Object.keys(fbTotals).sort(function(a,b){return fbTotals[b]-fbTotals[a];});\n'
                        '  if(!fbKeys.length){fr.style.display="none";fr.innerHTML="";_wmFbSubRowIb=null;return;}\n'
                        '  var curFilter=_wmPatFbFilter[String(ibk)];\n'
                        '  var h=\'<span style="font-size:11px;font-weight:bold;color:#2471a3;flex-shrink:0;margin-right:4px">&#9654; IB\'+ibk+\' FB Filter:</span>\';\n'
                        '  fbKeys.forEach(function(fb){\n'
                        '    var on=curFilter===undefined||curFilter===null||curFilter.has(fb);\n'
                        '    var fbDesc=(WM_PAT.fbDescriptions&&WM_PAT.fbDescriptions[fb]&&WM_PAT.fbDescriptions[fb].desc)?WM_PAT.fbDescriptions[fb].desc:"";\n'
                        '    var tip=fbDesc?"FB"+fb+" — "+fbDesc:"FB"+fb+" ("+fbTotals[fb]+" die)";\n'
                        '    h+=\'<label class="wm-pat-bincb" title="\'+tip+\'"><input type="checkbox"\'+(on?" checked":"")+\' data-ibk="\'+ibk+\'" data-fb="\'+fb+\'" onchange="_wmFbSubToggle(+this.dataset.ibk,this.dataset.fb,this.checked)">FB\'+fb+\'</label>\';\n'
                        '  });\n'
                        '  h+=\'<span style="font-size:10px;color:#2471a3;cursor:pointer;text-decoration:underline;margin-left:6px" onclick="_wmFbSubAll(\'+ibk+\',true)">All</span>\';\n'
                        '  h+=\'<span style="font-size:10px;color:#2471a3;cursor:pointer;text-decoration:underline;margin-left:4px" onclick="_wmFbSubAll(\'+ibk+\',false)">None</span>\';\n'
                        '  h+=\'<span style="font-size:10px;color:#c0392b;cursor:pointer;text-decoration:underline;margin-left:8px" onclick="_wmFbSubClear(\'+ibk+\')">&#x2715; Clear FB filter</span>\';\n'
                        '  fr.style.display="flex";\n'
                        '  fr.innerHTML=h;\n'
                        '}\n'
                        'function _wmFbSubToggle(ibk,fb,on){\n'
                        '  var fbTotals={};\n'
                        '  Object.keys(WM_PAT.wafers||{}).forEach(function(pk){var w=WM_PAT.wafers[pk];if(!w||!w.ibToFb)return;var fbMap=w.ibToFb[String(ibk)]||{};Object.keys(fbMap).forEach(function(f){fbTotals[f]=true;});});\n'
                        '  var allFbs=Object.keys(fbTotals);\n'
                        '  if(_wmPatFbFilter[String(ibk)]===undefined||_wmPatFbFilter[String(ibk)]===null){\n'
                        '    _wmPatFbFilter[String(ibk)]=new Set(allFbs);\n'
                        '  }\n'
                        '  if(on){_wmPatFbFilter[String(ibk)].add(fb);}else{_wmPatFbFilter[String(ibk)].delete(fb);}\n'
                        '  if(_wmPatFbFilter[String(ibk)].size===allFbs.length)_wmPatFbFilter[String(ibk)]=null;\n'
                        '  _wmPatRender();wmPatRenderReticle();\n'
                        '  _wmPatBuildBinRow(_wmPatBuildBinRow.lastArr||[]);\n'
                        '  _wmOpenFbSubRow(ibk);\n'
                        '}\n'
                        'function _wmFbSubAll(ibk,on){\n'
                        '  if(on){_wmPatFbFilter[String(ibk)]=null;}\n'
                        '  else{\n'
                        '    var fbTotals={};\n'
                        '    Object.keys(WM_PAT.wafers||{}).forEach(function(pk){var w=WM_PAT.wafers[pk];if(!w||!w.ibToFb)return;var fbMap=w.ibToFb[String(ibk)]||{};Object.keys(fbMap).forEach(function(f){fbTotals[f]=true;});});\n'
                        '    _wmPatFbFilter[String(ibk)]=new Set();\n'
                        '  }\n'
                        '  _wmPatRender();wmPatRenderReticle();\n'
                        '  _wmPatBuildBinRow(_wmPatBuildBinRow.lastArr||[]);\n'
                        '  _wmOpenFbSubRow(ibk);\n'
                        '}\n'
                        'function _wmFbSubClear(ibk){\n'
                        '  delete _wmPatFbFilter[String(ibk)];\n'
                        '  _wmPatRender();wmPatRenderReticle();\n'
                        '  _wmPatBuildBinRow(_wmPatBuildBinRow.lastArr||[]);\n'
                        '  _wmCloseFbSubRow();\n'
                        '}\n'
                        'function _wmPatBuildRetRow(keys){\n'
                        '  var rr=document.getElementById("wm-pat-retrow");\n'
                        '  if(!rr||!WM_PAT.hasReticle){if(rr)rr.style.display="none";return;}\n'
                        '  var sitesSeen={};\n'
                        '  keys.forEach(function(pk){\n'
                        '    var wdata=WM_PAT.wafers[pk];\n'
                        '    var dies=wdata&&wdata.dies?wdata.dies:wdata;\n'
                        '    if(!dies||!dies.length)return;\n'
                        '    var _ri=_wmRetInfoFor(pk);var rm=_ri.retMap;if(!rm)return;\n'
                        '    var rsl=_ri.retSiteLabels||WM_PAT.retSiteLabels||{};\n'
                        '    dies.forEach(function(d){\n'
                        '      var x=d[0],y=d[1];if(x===null||x===undefined)return;\n'
                        '      var info=rm[x+","+y];if(!info)return;\n'
                        '      var sk=info[0]+","+info[1];\n'
                        '      if(sitesSeen[sk]===undefined)sitesSeen[sk]=rsl[sk]!=null?rsl[sk]:null;\n'
                        '    });\n'
                        '  });\n'
                        '  var sks=Object.keys(sitesSeen);\n'
                        '  if(!sks.length){rr.style.display="none";return;}\n'
                        '  sks.sort(function(a,b){var la=sitesSeen[a],lb=sitesSeen[b];if(la!=null&&lb!=null)return la-lb;if(la!=null)return -1;if(lb!=null)return 1;var pa=a.split(","),pb=b.split(",");return(+pa[0]-+pb[0])||(+pa[1]-+pb[1]);});\n'
                        '  var _siteNum={};sks.forEach(function(sk,i){_siteNum[sk]=i+1;});\n'
                        '  WM_PAT._retSiteNum=_siteNum;\n'
                        '  var h=\'<span style="font-size:11px;font-weight:bold;color:#1f618d;flex-shrink:0;margin-right:4px">Die Loc:</span>\';\n'
                        '  sks.forEach(function(sk){\n'
                        '    var num=_siteNum[sk];\n'
                        '    var on=!(_wmPatRetUnchecked&&_wmPatRetUnchecked.has(sk));\n'
                        '    h+=\'<label class="wm-pat-bincb" title="RX\'+sk.split(",")[0]+\' RY\'+sk.split(",")[1]+\'"><input type="checkbox" data-sk="\'+sk+\'" \'+(on?"checked ":"")+ \'onchange="wmPatRetSiteToggle(this.dataset.sk,this.checked)">Loc \'+num+\'</label>\';\n'
                        '  });\n'
                        '  h+=\'<span style="font-size:10px;color:#1f618d;cursor:pointer;text-decoration:underline;margin-left:6px" onclick="_wmPatToggleRetAll(true)">All</span>\';\n'
                        '  h+=\'<span style="font-size:10px;color:#1f618d;cursor:pointer;text-decoration:underline;margin-left:4px" onclick="_wmPatToggleRetAll(false)">None</span>\';\n'
                        '  rr.style.display="";\n'
                        '  rr.innerHTML=h;\n'
                        '}\n'
                        'function _wmPatToggleRetAll(on){\n'
                        '  if(on){_wmPatRetUnchecked=null;}\n'
                        '  else{_wmPatRetUnchecked=new Set();var rr=document.getElementById("wm-pat-retrow");if(rr)rr.querySelectorAll("input[data-sk]").forEach(function(inp){_wmPatRetUnchecked.add(inp.dataset.sk);});}\n'
                        '  _wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'var _wmPatShotUnchecked=null;\n'
                        'var _wmShotAllSis=[];\n'
                        'function _wmPatBuildShotRow(keys){\n'
                        '  var sr=document.getElementById("wm-pat-shotrow");\n'
                        '  if(!sr||!WM_PAT.hasReticle){if(sr)sr.style.display="none";return;}\n'
                        '  var shotsSeen={};\n'
                        '  keys.forEach(function(pk){\n'
                        '    var wdata=WM_PAT.wafers[pk];\n'
                        '    var dies=wdata&&wdata.dies?wdata.dies:wdata;\n'
                        '    if(!dies||!dies.length)return;\n'
                        '    var _ri=_wmRetInfoFor(pk);var rm=_ri.retMap;if(!rm)return;\n'
                        '    dies.forEach(function(d){\n'
                        '      var x=d[0],y=d[1];if(x===null||x===undefined)return;\n'
                        '      var info=rm[x+","+y];if(!info)return;\n'
                        '      shotsSeen[info[2]]=true;\n'
                        '    });\n'
                        '  });\n'
                        '  var sis=Object.keys(shotsSeen).map(Number).sort(function(a,b){return a-b;});\n'
                        '  if(!sis.length){sr.style.display="none";return;}\n'
                        '  _wmShotAllSis=sis;\n'
                        '  var nSel=sis.filter(function(si){return !(_wmPatShotUnchecked&&_wmPatShotUnchecked.has(si));}).length;\n'
                        '  var btnLbl=nSel===sis.length?"All ("+sis.length+")":nSel+" / "+sis.length+" selected";\n'
                        '  var h=\'<span style="font-size:11px;font-weight:bold;color:#6c3483;flex-shrink:0;margin-right:4px">Shot #:</span>\';\n'
                        '  h+=\'<div style="position:relative;display:inline-block;vertical-align:middle">\';\n'
                        '  h+=\'<button id="wm-shot-dd-btn" onclick="_wmShotDdOpen()" style="font-size:11px;padding:1px 8px 1px 6px;border:1px solid #c39bd3;border-radius:3px;background:#f5eef8;color:#6c3483;cursor:pointer;min-width:120px;text-align:left;white-space:nowrap">\'+btnLbl+\' &#9660;</button>\';\n'
                        '  h+=\'<div id="wm-shot-dd" style="display:none;position:absolute;z-index:9999;background:#fff;border:1px solid #c39bd3;border-radius:4px;box-shadow:0 4px 12px rgba(0,0,0,.18);padding:0;min-width:160px;top:calc(100% + 2px);left:0">\';\n'
                        '  h+=\'<div style="padding:4px 6px;border-bottom:1px solid #eee"><input id="wm-shot-dd-search" type="text" placeholder="Search shots..." oninput="_wmShotDdFilter(this.value)" style="width:100%;font-size:11px;border:1px solid #ddd;border-radius:2px;padding:2px 4px;box-sizing:border-box"/></div>\';\n'
                        '  h+=\'<div style="display:flex;gap:6px;padding:2px 6px;border-bottom:1px solid #eee;font-size:10px">\';\n'
                        '  h+=\'<a href="#" onclick="_wmPatToggleShotAll(true);return false" style="color:#6c3483;font-weight:bold">All</a>\';\n'
                        '  h+=\'<a href="#" onclick="_wmPatToggleShotAll(false);return false" style="color:#c0392b;font-weight:bold">None</a></div>\';\n'
                        '  h+=\'<div id="wm-shot-dd-list" style="max-height:180px;overflow-y:auto;padding:2px 0">\';\n'
                        '  sis.forEach(function(si){\n'
                        '    var on=!(_wmPatShotUnchecked&&_wmPatShotUnchecked.has(si));\n'
                        '    h+=\'<label data-shot="\'+si+\'" style="display:flex;align-items:center;gap:5px;padding:2px 8px;font-size:11px;cursor:pointer;white-space:nowrap"><input type="checkbox" data-si="\'+si+\'" \'+(on?"checked ":"")+ \'onchange="_wmPatShotToggle(+this.dataset.si,this.checked)">Shot \'+si+\'</label>\';\n'
                        '  });\n'
                        '  h+=\'</div></div></div>\';\n'
                        '  h+=\'<span style="color:#ccc;margin:0 8px">|</span>\';\n'
                        '  h+=\'<span style="font-size:11px;font-weight:bold;color:#6d4c41;flex-shrink:0;margin-right:4px">Excl. edge rows:</span>\';\n'
                        '  var _edgeOpts=[0,1,2,3,4,5,6,7,8,9,10];\n'
                        '  h+=\'<select class="wm-edge-sel" onchange="_wmSetEdgeRows(+this.value)" style="font-size:11px;padding:1px 4px;background:#f5f5f5;color:#6d4c41;border:1px solid #bcaaa4;border-radius:3px;cursor:pointer">\';\n'
                        '  _edgeOpts.forEach(function(n){h+=\'<option value="\'+n+\'" \'+(_wmEdgeExcRows===n?"selected":"")+\'>\'+n+\'</option>\';});\n'
                        '  h+=\'</select>\';\n'
                        '  h+=\'<span style="color:#ccc;margin:0 8px">|</span>\';\n'
                        '  h+=\'<span style="font-size:11px;font-weight:bold;color:#1f618d;flex-shrink:0;margin-right:4px">&#8805;IB:</span>\';\n'
                        '  [1,2,3,4,5].forEach(function(v){h+=\'<label style="display:flex;align-items:center;gap:2px;font-size:11px;color:#1f618d;cursor:pointer;white-space:nowrap;margin-right:4px"><input type="radio" name="wm-thr-rb" value="\'+v+\'" \'+(_wmFailThr===v?"checked ":"")+\'onchange="_wmSetFailThr(+this.value)" style="cursor:pointer">\'+v+\'</label>\';});\n'
                        '  h+=\'<span style="color:#ccc;margin:0 8px">|</span>\';\n'
                        '  h+=\'<span style="font-size:11px;font-weight:bold;color:#555;margin-right:4px">Zoom:</span>\';\n'
                        '  h+=\'<button onclick="_wmPatSetZoom(-0.25)" style="font-size:13px;line-height:1;padding:0 6px;border:1px solid #bbb;border-radius:3px;background:#f5f5f5;cursor:pointer" title="Zoom out">&#8722;</button>\';\n'
                        '  h+=\'<span id="wm-zoom-pct" style="font-size:10px;color:#555;margin:0 4px;min-width:32px;text-align:center;display:inline-block">\'+Math.round(_wmPatZoom*100)+\'%</span>\';\n'
                        '  h+=\'<button onclick="_wmPatSetZoom(+0.25)" style="font-size:13px;line-height:1;padding:0 6px;border:1px solid #bbb;border-radius:3px;background:#f5f5f5;cursor:pointer" title="Zoom in">+</button>\';\n'
                        '  sr.style.display="";\n'
                        '  sr.innerHTML=h;\n'
                        '}\n'
                        'function _wmShotDdOpen(){\n'
                        '  var dd=document.getElementById("wm-shot-dd");if(!dd)return;\n'
                        '  var isOpen=dd.style.display!=="none";\n'
                        '  if(!isOpen){\n'
                        '    dd.style.display="";\n'
                        '    var inp=document.getElementById("wm-shot-dd-search");if(inp){inp.value="";_wmShotDdFilter("");inp.focus();}\n'
                        '    setTimeout(function(){document.addEventListener("click",function _cl(e){var btn=document.getElementById("wm-shot-dd-btn");if(!dd.contains(e.target)&&e.target!==btn){dd.style.display="none";document.removeEventListener("click",_cl,true);}},true);},0);\n'
                        '  } else { dd.style.display="none"; }\n'
                        '}\n'
                        'function _wmShotDdFilter(q){\n'
                        '  var list=document.getElementById("wm-shot-dd-list");if(!list)return;\n'
                        '  var s=q.trim().toLowerCase();\n'
                        '  list.querySelectorAll("label[data-shot]").forEach(function(lbl){\n'
                        '    var val="shot "+lbl.dataset.shot;\n'
                        '    lbl.style.display=(s===""||val.indexOf(s)>=0)?"":"none";\n'
                        '  });\n'
                        '}\n'
                        'function _wmShotDdRefreshBtn(){\n'
                        '  var btn=document.getElementById("wm-shot-dd-btn");if(!btn)return;\n'
                        '  var sis=_wmShotAllSis;\n'
                        '  var nSel=sis.filter(function(si){return !(_wmPatShotUnchecked&&_wmPatShotUnchecked.has(si));}).length;\n'
                        '  btn.innerHTML=(nSel===sis.length?"All ("+sis.length+")":nSel+" / "+sis.length+" selected")+" &#9660;";\n'
                        '}\n'
                        'function _wmPatShotToggle(si,on){\n'
                        '  if(_wmPatShotUnchecked===null)_wmPatShotUnchecked=new Set();\n'
                        '  if(on){_wmPatShotUnchecked.delete(si);}else{_wmPatShotUnchecked.add(si);}\n'
                        '  if(_wmPatShotUnchecked.size===0)_wmPatShotUnchecked=null;\n'
                        '  var ddInp=document.querySelector(\'#wm-shot-dd-list input[data-si="\'+si+\'"]\');if(ddInp)ddInp.checked=on;\n'
                        '  _wmShotDdRefreshBtn();\n'
                        '  _wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function _wmPatToggleShotAll(on){\n'
                        '  if(on){\n'
                        '    _wmPatShotUnchecked=null;\n'
                        '    var list=document.getElementById("wm-shot-dd-list");\n'
                        '    if(list)list.querySelectorAll("input[data-si]").forEach(function(inp){inp.checked=true;});\n'
                        '  } else {\n'
                        '    _wmPatShotUnchecked=new Set(_wmShotAllSis);\n'
                        '    var list=document.getElementById("wm-shot-dd-list");\n'
                        '    if(list)list.querySelectorAll("input[data-si]").forEach(function(inp){inp.checked=false;});\n'
                        '  }\n'
                        '  _wmShotDdRefreshBtn();\n'
                        '  _wmPatRender();wmPatRenderReticle();if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function _wmPatRender(){\n'
                        '  var maps=document.getElementById("wm-pat-maps");\n'
                        '  var tbody=document.getElementById("wm-pat-tbody");\n'
                        '  var impactBody=document.getElementById("wm-pat-impact-body");\n'
                        '  if(!maps||!tbody)return;\n'
                        '  var allKeys=Object.keys(WM_PAT.wafers).filter(function(k){return _wmPatMatchLots(k)&&_wmPatMatchProgs(k);}).sort(function(a,b){\n'
                        '    var la=_wmPatGetLot(a),lb=_wmPatGetLot(b);if(la!==lb)return la<lb?-1:1;\n'
                        '    var wa=parseInt(_wmPatGetWfr(a))||0,wb=parseInt(_wmPatGetWfr(b))||0;if(wa!==wb)return wa-wb;\n'
                        '    return(_wmPatGetProg(a)<_wmPatGetProg(b)?-1:1);\n'
                        '  });\n'
                        '  _wmPatBuildCtrl(allKeys);\n'
                        '  var keys=_wmPatSelWafers===null?allKeys:allKeys.filter(function(k){return _wmPatSelWafers.has(k);});\n'
                        '  _wmPatLastKeys=keys;\n'
                        '  if(_wmCriteriaMissOnly){var _cmk=keys.filter(function(k){return _wmGetCriteriaMissBins(k).length>0;});\n'
                        '    if(_cmk.length>0){keys=_cmk;}else{var _cm_el=maps;if(_cm_el){_cm_el.innerHTML=\'<div style="padding:20px;text-align:center;color:#e67e22;font-size:13px">&#9888; Wafers missing yield target filter is active but no wafers exceed thresholds.<br><span style="font-size:11px;color:#888">All wafers shown. Check yield spec expected values.</span></div>\';_wmPatBuildCtrl([]);if(tbody)tbody.innerHTML=\'\';if(impactBody)impactBody.innerHTML=\'\';return;}}}\n'
                        '  var FIXED_W=Math.round(190*_wmPatZoom),pad=2;\n'
                        '  var mapsHtml="",tbHtml="",ibSeen={},ibPatAcc={},sc_acc={};\n'
                        '  var _bar=function(v){var pw=Math.round(v*90);var c=v<0.35?"#27ae60":v<0.65?"#e67e22":"#c0392b";return\'<span class="wm-bar-bg"><span class="wm-bar-fg" style="width:\'+pw+\'px;background:\'+c+\'"></span></span><span style="font-size:10px;color:#555;margin-left:3px">\'+Math.round(v*100)+\'%</span>\';};\n'
                        '  keys.forEach(function(pk){\n'
                        '    var wdata=WM_PAT.wafers[pk];\n'
                        '    var dies=wdata&&wdata.dies?wdata.dies:wdata;\n'
                        '    var mLot=wdata.lot||pk.split("::")[0];\n'
                        '    var mWfr=wdata.wafer||pk.split("::")[1];\n'
                        '    var mProg=wdata.program||_wmPatGetProg(pk);\n'
                        '    var mMat=wdata.material||"";\n'
                        '    var multiProg2=_wmPatAllProgs().length>1;\n'
                        '    var mProgLbl=multiProg2&&mProg?\' <span style="font-size:9px;color:#7fb3d3">\'+mProg+"</span>":"";\n'
                        '    if(!dies||!dies.length){\n'
                        '      mapsHtml+=\'<div style="text-align:center"><div class="wm-wlbl" style="color:#aaa">\'+mLot+\' W\'+mWfr+mProgLbl+\'</div><div style="font-size:10px;color:#ccc;margin-top:4px">no data</div></div>\';\n'
                        '      tbHtml+=\'<tr><td style="font-size:10px">\'+mLot+\'</td><td style="font-weight:bold">W\'+mWfr+mProgLbl+\'</td><td colspan="9" style="color:#bbb;font-size:10px">no data</td></tr>\';\n'
                        '      return;\n'
                        '    }\n'
                        '    var _hasFbFilter=Object.keys(_wmPatFbFilter).some(function(k){return _wmPatFbFilter[k]!==null;});\n'
                        '    var _hasIbFilter=(_wmPatBinChecked!==null);\n'
                        '    if(_hasFbFilter||_hasIbFilter){\n'
                        '      var _anyVisible=dies.some(function(d){\n'
                        '        if(d[0]===null)return false;\n'
                        '        var _ibk=(d[2]!==null&&d[2]!==undefined)?d[2]:null;\n'
                        '        var _binOn=(_wmPatBinChecked===null||_wmPatBinChecked.has(String(_ibk)));\n'
                        '        if(!_binOn)return false;\n'
                        '        if(_hasFbFilter&&_ibk!==null){var _fbF=_wmPatFbFilter[String(_ibk)];if(_fbF!==undefined&&_fbF!==null){var _fbV=d[3]!==undefined&&d[3]!==null?String(d[3]):null;if(_fbV===null||!_fbF.has(_fbV))return false;}}\n'
                        '        return true;\n'
                        '      });\n'
                        '      if(!_anyVisible)return;\n'
                        '    }\n'
                        '    var xs=[],ys=[];\n'
                        '    dies.forEach(function(d){if(d[0]!==null){xs.push(d[0]);ys.push(d[1]);}});\n'
                        '    var xMin=Math.min.apply(null,xs),xMax=Math.max.apply(null,xs);\n'
                        '    var yMin=Math.min.apply(null,ys),yMax=Math.max.apply(null,ys);\n'
                        '    var cs=Math.max(2,(FIXED_W-pad*2)/(xMax-xMin+1));\n'
                        '    var xSpan=xMax-xMin,ySpan=yMax-yMin;\n'
                        '    var csy=(xSpan>0&&ySpan>0)?(cs*xSpan/ySpan):cs;\n'
                        '    var W=FIXED_W,H=Math.round((yMax-yMin+1)*csy+pad*2);\n'
                        '    var xCtr=(xMin+xMax)/2,yCtr=(yMin+yMax)/2;\n'
                        '    var xRad=(xMax-xMin)/2||1,yRad=(yMax-yMin)/2||1;\n'
                        '    var ibCoords={},failXn=[],failYn=[],failActX=[],failActY=[],totalDies=0,failDies=0;\n'
                        '    var failShotIdx=new Set();\n'
                        '    var _pkSiteFailCnt={};\n'
                        '    var _stripePatId="wmsp_"+pk.replace(/[^a-z0-9]/gi,"_");\n'
                        '    var rects=[];\n'
                        '    dies.forEach(function(d){\n'
                        '      var x=d[0],y=d[1],ib=d[2];if(x===null)return;\n'
                        '      totalDies++;\n'
                        '      var ibKey=(ib!==null&&ib!==undefined)?ib:null;\n'
                        '      ibSeen[String(ibKey)]=_wmIbColor(ibKey);\n'
                        '      var binOn=(_wmPatBinChecked===null||_wmPatBinChecked.has(String(ibKey)));\n'
                        '      if(binOn&&ibKey!==null){var _fbF=_wmPatFbFilter[String(ibKey)];if(_fbF!==undefined&&_fbF!==null){var _fbVal=d[3]!==undefined&&d[3]!==null?String(d[3]):null;if(_fbVal===null||!_fbF.has(_fbVal))binOn=false;}}\n'
                        '      if(_wmIsFail(ibKey)&&ibKey!==null&&binOn){\n'
                        '        var xn=(x-xCtr)/xRad,yn=(y-yCtr)/yRad;\n'
                        '        var _isEdge=(_wmEdgeExcRows>0&&(x<xMin+_wmEdgeExcRows||x>xMax-_wmEdgeExcRows||y<yMin+_wmEdgeExcRows||y>yMax-_wmEdgeExcRows));if(!_isEdge){failXn.push(xn);failYn.push(yn);failActX.push(x);failActY.push(y);failDies++;\n'
                        '        if(WM_PAT.hasReticle){var _wri=_wmRetInfoFor(pk);var _ri=_wri.retMap&&_wri.retMap[x+","+y];if(_ri){failShotIdx.add(_ri[2]);var _sk0=_ri[0]+\",\"+_ri[1];_pkSiteFailCnt[_sk0]=(_pkSiteFailCnt[_sk0]||0)+1;}}\n'
                        '        if(!ibCoords[ibKey])ibCoords[ibKey]={xn:[],yn:[],ax:[],ay:[]};\n'
                        '        ibCoords[ibKey].xn.push(xn);ibCoords[ibKey].yn.push(yn);\n'
                        '        ibCoords[ibKey].ax.push(x);ibCoords[ibKey].ay.push(y);}\n'
                        '      }\n'
                        '    });\n'
                        '    var _pkRetInfo=_wmRetInfoFor(pk);\n'
                        '    var _pkShots=(_pkRetInfo.retShots&&_pkRetInfo.retShots.length)?_pkRetInfo.retShots:WM_PAT.retShots;\n'
                        '    var _topLocStr="\u2014";if(WM_PAT.hasReticle&&WM_PAT._retSiteNum){var _skeys=Object.keys(_pkSiteFailCnt);if(_skeys.length){_skeys.sort(function(a,b){return _pkSiteFailCnt[b]-_pkSiteFailCnt[a];});var _topSk=_skeys[0];var _topN=WM_PAT._retSiteNum[_topSk];if(_topN!=null){var _topPct=failDies>0?Math.round(_pkSiteFailCnt[_topSk]/failDies*100):0;_topLocStr="Loc"+_topN+" ("+_topPct+"%)";}else _topLocStr=_topSk;}}\n'
                        '    var failPct=totalDies>0?(failDies/totalDies*100).toFixed(1)+"%":"0%";\n'
                        '    var driverIb="\u2014";\n'
                        '    if(failDies>0){\n'
                        '      var drKeys=Object.keys(ibCoords).sort(function(a,b){return ibCoords[b].xn.length-ibCoords[a].xn.length;});\n'
                        '      var topN=drKeys.length?ibCoords[drKeys[0]].xn.length:0;\n'
                        '      driverIb=drKeys.filter(function(k){return ibCoords[k].xn.length>=topN*0.8;}).map(function(k){return"IB"+k+"(n="+ibCoords[k].xn.length+")";}).join(", ");\n'
                        '    }\n'
                        '    var sc={center:0,edge:0,donut:0,systematic:0,reticle:0,random:0};\n'
                        '    var retSc=0;\n'
                        '    var _psc={confidence:"LOW"};\n'
                        '    if(failDies>=3){\n'
                        '      _psc=_wmScorePattern(failXn,failYn);\n'
                        '      sc.center=_psc.center;sc.edge=_psc.edge;sc.donut=_psc.donut;sc.systematic=_psc.systematic;\n'
                        '      if(WM_PAT.hasReticle&&failActX.length>0){\n'
                        '        var _wri2=_wmRetInfoFor(pk);\n'
                        '        retSc=_wmScoreReticle(failActX,failActY,_wri2.retMap||WM_PAT.retMap,_wri2.retSiteTotals||WM_PAT.retSiteTotals);\n'
                        '        sc.reticle=retSc;\n'
                        '      }\n'
                        '      var dominated=Math.max(sc.center,sc.edge,sc.donut,sc.systematic,sc.reticle);\n'
                        '      sc.random=Math.max(0,1-dominated);\n'
                        '    } else {\n'
                        '      _psc={confidence:"LOW"};\n'
                        '      sc.random=failDies>0?1:0;\n'
                        '    }\n'
                        '    sc_acc[pk]={center:sc.center,edge:sc.edge,donut:sc.donut,systematic:sc.systematic,reticle:sc.reticle,random:sc.random,topLoc:_topLocStr,siteFailCnt:_pkSiteFailCnt,failDies:failDies};\n'
                        '    var dims2=["center","edge","donut","systematic","reticle","random"];\n'
                        '    var primary="RANDOM",pCol=_pColors.RANDOM;\n'
                        '    var bestScore=sc.random;\n'
                        '    dims2.forEach(function(d){if(d!=="random"&&(sc[d]||0)>bestScore){bestScore=sc[d];primary=d.toUpperCase();pCol=_pColors[d.toUpperCase()]||"#555";}});\n'
                        '    var _confCol={HIGH:"#27ae60",MEDIUM:"#e67e22",LOW:"#e74c3c"}[_psc.confidence]||"#999";\n'
                        '    if(failDies>0){\n'
                        '      Object.keys(ibCoords).forEach(function(ibk){\n'
                        '        if(!ibPatAcc[ibk])ibPatAcc[ibk]={cnt:0,dies:0,center:0,edge:0,donut:0,systematic:0,reticle:0,random:0};\n'
                        '        var acc=ibPatAcc[ibk];acc.cnt++;acc.dies+=ibCoords[ibk].xn.length;\n'
                        '        acc.center+=sc.center;acc.edge+=sc.edge;acc.donut+=sc.donut;\n'
                        '        acc.systematic+=sc.systematic;acc.reticle+=sc.reticle;acc.random+=sc.random;\n'
                        '      });\n'
                        '    }\n'
                        '    mapsHtml+=\'<div class="wm-pat-tile-ph" data-idx="\'+_wmPatLastKeys.indexOf(pk)+\'" style="text-align:center;padding:4px 2px">\'\n'
                        '      +\'<div style="display:inline-block;border-radius:50%;transition:box-shadow 0.12s" onclick="_wmZoomWafer(\\\'\'+pk+\'\\\',this)">\'\n'
                        '      +\'<div class="wm-pat-tile-content" style="width:\'+W+\'px;height:\'+H+\'px;background:#e8ecf0;border-radius:50%;display:inline-block"></div>\'\n'
                        '      +\'</div>\'\n'
                        '      +\'<div style="font-size:10px;color:\'+pCol+\';font-weight:bold;margin-top:2px">\'+primary+\'</div>\'\n'
                        '      +(function(){\n'
                        '        var _cb=_wmGetCriteriaMissBins(pk),_ci=_wmGetCriteriaMissInfo(pk);\n'
                        '        var _icon=_cb.length?\'<span data-pk="'+pk+'" style="color:#c0392b;margin-left:5px;font-size:20px;vertical-align:middle;cursor:pointer" title="Click to see yield misses" onclick="event.stopPropagation();_wmShowIconTip(this)">&#9888;</span>\':\'\';\n'
                        '        return \'<div class="wm-wlbl" style="margin-top:3px;cursor:pointer;font-size:12px" onclick="_wmShowCriteriaTable(\\\''+pk+'\\\',event)">\'+mLot+" W"+mWfr+mProgLbl+_icon+\'</div>\'\n'
                        '          +(mMat?\'<div style="font-size:9px;color:#8e6a2a;margin-top:1px">\'+mMat+\'</div>\':\'\');\n'
                        '      })()\n'
                        '      +\'</div>\';\n'
                        '    tbHtml+=\'<tr>\'\n'
                        '      +\'<td style="font-size:10px;white-space:nowrap">\'+mLot+\'</td>\'\n'
                        '      +\'<td style="font-weight:bold;white-space:nowrap">W\'+mWfr+mProgLbl+\'</td>\'\n'
                        '      +\'<td style="font-size:10px;color:#555;white-space:nowrap">\'+mMat+\'</td>\'\n'
                        '      +\'<td style="font-weight:bold;color:\'+pCol+\'">\'+primary+\'</td>\'\n'
                        '      +\'<td style="white-space:nowrap;color:\'+_confCol+\';font-size:10px">\'+_psc.confidence+\'</td>\'\n'
                        '      +\'<td style="white-space:nowrap">\'+failPct+\'<span style="font-size:9px;color:#999;margin-left:2px">(n=\'+failDies+\')</span></td>\'\n'
                        '      +\'<td style="font-size:10px;max-width:120px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="\'+driverIb+\'">\'+driverIb+\'</td>\'\n'
                        '      +\'<td>\'+_bar(sc.center||0)+\'</td>\'\n'
                        '      +\'<td>\'+_bar(sc.edge||0)+\'</td>\'\n'
                        '      +\'<td>\'+_bar(sc.donut||0)+\'</td>\'\n'
                        '      +\'<td>\'+_bar(sc.systematic||0)+\'</td>\'\n'
                        '      +(WM_PAT.hasReticle?\'<td>\'+_bar(retSc||0)+\'</td>\':\'\')\n'
                        '      +(WM_PAT.hasReticle?\'<td style="font-size:10px;white-space:nowrap;color:#1f618d">\'+(_topLocStr||\'—\')+\'</td>\':\'\')\n'
                        '      +\'<td>\'+_bar(sc.random||0)+\'</td></tr>\';\n'
                        '  });\n'
                        '  maps.innerHTML=mapsHtml||\'<span style="color:#999;font-size:12px">No wafers selected</span>\';\n'
                        '  _wmPatRenderedKeys=new Set();\n'
                        '  if(_wmPatObserver){_wmPatObserver.disconnect();_wmPatObserver=null;}\n'
                        '  var _patWrap=document.querySelector(".wm-pat-maps-wrap");\n'
                        '  _wmPatObserver=new IntersectionObserver(function(entries){\n'
                        '    entries.forEach(function(entry){\n'
                        '      if(entry.isIntersecting){\n'
                        '        var _pidx=parseInt(entry.target.dataset.idx,10);\n'
                        '        var _ppk=_wmPatLastKeys[_pidx];\n'
                        '        var _ptc=entry.target.querySelector(".wm-pat-tile-content");\n'
                        '        if(_ppk&&_ptc)_wmPatRenderTile(_ppk,_ptc);\n'
                        '      }\n'
                        '    });\n'
                        '  },{root:_patWrap,rootMargin:"300px 0px"});\n'
                        '  maps.querySelectorAll(".wm-pat-tile-ph[data-idx]").forEach(function(el){_wmPatObserver.observe(el);});\n'
                        '  var _upmLegEl=document.getElementById("wm-upm-legend");\n'
                        '  if(_upmLegEl){\n'
                        '    if(_wmUpmOverlay&&WM_PAT.upmCols&&WM_PAT.upmCols.length){\n'
                        '      var _upmLbl=Array.isArray(WM_PAT.upmCols[0])?WM_PAT.upmCols[0][0]:WM_PAT.upmCols[0];\n'
                        '      var _upmTgt2=Array.isArray(WM_PAT.upmCols[0])?WM_PAT.upmCols[0][1]:null;\n'
                        '      _upmLegEl.style.display="flex";\n'
                        '      var _ticks=[{pct:90,label:"90%"},{pct:93,label:"93%"},{pct:97,label:"97%"},{pct:100,label:"100%+"}];\n'
                        '      var _tickHtml=_ticks.map(function(tk){var pos=Math.round((tk.pct-90)/(100-90)*140);return\'<div style="position:absolute;left:\'+pos+\'px;transform:translateX(-50%);text-align:center"><div style="width:1px;height:5px;background:rgba(255,255,255,0.6);margin:0 auto"></div><span style="font-size:9px;color:rgba(255,255,255,0.75);white-space:nowrap">\'+tk.label+\'</span></div>\';}).join("");\n'
                        '      var _colStops=_ticks.map(function(tk){return\'<span style="font-size:9px;padding:0 2px;border-radius:2px;background:\'+_wmUpmColor(tk.pct)+\';color:#fff;font-weight:bold">\'+tk.label+\'</span>\';}).join(\'<span style="font-size:10px;color:rgba(255,255,255,0.4);margin:0 2px">→</span>\');\n'
                        '      _upmLegEl.innerHTML=\n'
                        '        \'<div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap">\'\n'
                        '        +\'<span style="font-size:11px;font-weight:bold;color:#ecf0f1">UPM Heatmap</span>\'\n'
                        '        +\'<span style="font-size:10px;color:#95a5a6">\'+_upmLbl+\'</span>\'\n'
                        '        +\'<div style="display:inline-flex;flex-direction:column;gap:2px">\'\n'
                        '          +\'<div style="position:relative;width:140px;height:12px;border-radius:3px;background:linear-gradient(to right,#00dc00,#0050ff,#ff8800,#a05000);border:1px solid rgba(255,255,255,0.3)"></div>\'\n'
                        '          +\'<div style="position:relative;width:140px;height:14px">\'+_tickHtml+\'</div>\'\n'
                        '        +\'</div>\'\n'
                        '        +_colStops\n'
                        '        +\'<span style="font-size:10px;color:#ff6666;margin-left:4px">&#9632; IB-selected dies = solid red</span>\'\n'
                        '        +\'<button id="wm-upm-dist-btn" onclick="_wmToggleUpmDist()" style="font-size:10px;padding:1px 8px;border-radius:10px;border:1px solid rgba(255,255,255,0.35);background:rgba(255,255,255,0.1);color:#ecf0f1;cursor:pointer;white-space:nowrap">&#128202; Distribution</button>\'\n'
                        '        +\'</div>\'\n'
                        '        +\'<div id="wm-upm-dist-chart" style="display:none;margin-top:4px;padding:6px;background:rgba(0,0,0,0.3);border-radius:4px;border:1px solid rgba(255,255,255,0.12)"></div>\';\n'
                        '      if(_wmUpmDistOpen)_wmRenderUpmDist();\n'
                        '    } else {\n'
                        '      _upmLegEl.style.display="none";\n'
                        '    }\n'
                        '  }\n'
                        '  tbody.innerHTML=tbHtml;\n'
                        '  // Lot Trend\n'
                        '  var ltEl=document.getElementById("wm-pat-lot-trend");\n'
                        '  if(ltEl){\n'
                        '    var _lta={};\n'
                        '    keys.forEach(function(pk){\n'
                        '      var w=WM_PAT.wafers[pk];var lot=(w&&w.lot)||pk.split("::")[0];\n'
                        '      if(!_lta[lot])_lta[lot]={n:0,center:0,edge:0,donut:0,systematic:0,reticle:0,random:0,siteFailAgg:{},failDiesAgg:0};\n'
                        '      var _s=_lta[lot];_s.n++;\n'
                        '      if(sc_acc[pk]){_s.center+=sc_acc[pk].center;_s.edge+=sc_acc[pk].edge;_s.donut+=sc_acc[pk].donut;_s.systematic+=sc_acc[pk].systematic;_s.reticle+=sc_acc[pk].reticle;_s.random+=sc_acc[pk].random;_s.failDiesAgg+=sc_acc[pk].failDies||0;var _pSfc=sc_acc[pk].siteFailCnt||{};Object.keys(_pSfc).forEach(function(sk){_s.siteFailAgg[sk]=(_s.siteFailAgg[sk]||0)+_pSfc[sk];});}\n'
                        '    });\n'
                        '    var _lots=Object.keys(_lta).sort();\n'
                        '    if(_lots.length){\n'
                        '      var _lth=\'<table style="width:100%;border-collapse:collapse;font-size:11px"><thead><tr style="background:#d6eaf8"><th style="text-align:left;padding:2px 6px;color:#1a5276">Lot</th><th style="padding:2px 4px;color:#555">Wfrs</th><th style="padding:2px 6px;color:#555">Primary</th><th style="padding:2px 4px;color:#c0392b">Center</th><th style="padding:2px 4px;color:#e67e22">Edge</th><th style="padding:2px 4px;color:#8e44ad">Donut</th><th style="padding:2px 4px;color:#2471a3">Systematic</th>\'+(WM_PAT.hasReticle?\'<th style="padding:2px 4px;color:#1f618d">Reticle</th>\':\'\')+(WM_PAT.hasReticle?\'<th style="padding:2px 4px;color:#1a5276">Top Die Loc</th>\':\'\')+\'<th style="padding:2px 4px;color:#27ae60">Random</th></tr></thead><tbody>\';\n'
                        '      _lots.forEach(function(lot,li){\n'
                        '        var a=_lta[lot],n=a.n||1;\n'
                        '        var lsc={center:a.center/n,edge:a.edge/n,donut:a.donut/n,systematic:a.systematic/n,reticle:a.reticle/n,random:a.random/n};\n'
                        '        var lPrim=\'RANDOM\',lCol=_pColors.RANDOM,lV=lsc.random;\n'
                        '        [\'center\',\'edge\',\'donut\',\'systematic\',\'reticle\'].forEach(function(d){if((lsc[d]||0)>lV){lV=lsc[d];lPrim=d.toUpperCase();lCol=_pColors[d.toUpperCase()]||\'#555\';}});\n'
                        '        var bg=li%2?\'background:#f7f9fc\':\'\';\n'
                        '        _lth+=\'<tr style="\'+bg+\'"><td style="padding:2px 6px;font-weight:bold;color:#1a5276">\'+lot+\'</td>\'\n'
                        '          +\'<td style="text-align:center;padding:2px 4px;color:#555">\'+a.n+\'</td>\'\n'
                        '          +\'<td style="font-weight:bold;color:\'+lCol+\';padding:2px 6px">\'+ lPrim+\'</td>\'\n'
                        '          +\'<td style="text-align:center;padding:2px 4px">\'+Math.round(lsc.center*100)+\'%</td>\'\n'
                        '          +\'<td style="text-align:center;padding:2px 4px">\'+Math.round(lsc.edge*100)+\'%</td>\'\n'
                        '          +\'<td style="text-align:center;padding:2px 4px">\'+Math.round(lsc.donut*100)+\'%</td>\'\n'
                        '          +\'<td style="text-align:center;padding:2px 4px">\'+Math.round(lsc.systematic*100)+\'%</td>\'\n'
                        '          +(WM_PAT.hasReticle?\'<td style="text-align:center;padding:2px 4px">\'+Math.round(lsc.reticle*100)+\'%</td>\':\'\')\n'
                        '          +(WM_PAT.hasReticle?(function(){var _ltSkeys=Object.keys(a.siteFailAgg);if(!_ltSkeys.length)return\'<td style="text-align:center;padding:2px 4px;color:#1a5276;font-size:10px">\u2014</td>\';_ltSkeys.sort(function(x,y){return a.siteFailAgg[y]-a.siteFailAgg[x];});var _ltTopSk=_ltSkeys[0];var _ltTopN=WM_PAT._retSiteNum&&WM_PAT._retSiteNum[_ltTopSk];var _ltPct=a.failDiesAgg>0?Math.round(a.siteFailAgg[_ltTopSk]/a.failDiesAgg*100):0;var _ltLbl=_ltTopN!=null?"Loc"+_ltTopN+" ("+_ltPct+"%)":_ltTopSk;return\'<td style="text-align:center;padding:2px 4px;color:#1a5276;font-size:10px">\'+_ltLbl+\'</td>\';}()):\'\')\n'
                        '          +\'<td style="text-align:center;padding:2px 4px">\'+Math.round(lsc.random*100)+\'%</td></tr>\';\n'
                        '      });\n'
                        '      _lth+=\'</tbody></table>\';\n'
                        '      ltEl.innerHTML=_lth;\n'
                        '    }\n'
                        '  }\n'
                        '  var ibKeys=Object.keys(ibPatAcc).sort(function(a,b){return+a-+b;});\n'
                        '  if(impactBody&&ibKeys.length){\n'
                        '    var dims=["center","edge","donut","systematic","random"];\n'
                        '    if(WM_PAT.hasReticle)dims.splice(4,0,"reticle");\n'
                        '    var dimLbls={center:"Center",edge:"Edge",donut:"Donut",systematic:"Syst.",reticle:"Reticle",random:"Rnd"};\n'
                        '    var ibh=\'<div style="font-size:10px;color:#888;margin-bottom:6px">Avg pattern score per fail IB (across displayed wafers).</div>\';\n'
                        '    ibKeys.forEach(function(ibk){\n'
                        '      var a=ibPatAcc[ibk],cnt=a.cnt||1,nDies=a.dies||0;\n'
                        '      var col=_wmIbColor(+ibk);\n'
                        '      var bestDim="random",bestVal=a.random/cnt;\n'
                        '      dims.forEach(function(d){if(a[d]/cnt>bestVal){bestVal=a[d]/cnt;bestDim=d;}});\n'
                        '      var bdCol=_pColors[bestDim.toUpperCase()]||"#555";\n'
                        '      ibh+=\'<div class="wm-impact-row" style="margin-bottom:8px;padding-bottom:6px;border-bottom:1px solid #eee">\'\n'
                        '        +\'<div class="wm-impact-lbl" style="color:\'+col+\'">IB\'+ibk+\'<span style="font-size:9px;color:#999;margin-left:3px">(n=\'+nDies+\')</span></div>\'\n'
                        '        +\'<div style="flex:1;display:flex;flex-wrap:wrap;gap:3px 8px">\';\n'
                        '      dims.forEach(function(d){\n'
                        '        var v=a[d]/cnt,bc=v<0.35?"#27ae60":v<0.65?"#e67e22":"#c0392b";\n'
                        '        ibh+=\'<div style="display:inline-flex;align-items:center;gap:2px;font-size:10px">\'\n'
                        '          +\'<span style="width:32px;color:#666">\'+dimLbls[d]+\'</span>\'\n'
                        '          +\'<div class="wm-impact-bar" style="width:44px"><div class="wm-impact-fill" style="width:\'+Math.round(v*44)+\'px;background:\'+bc+\'"></div></div>\'\n'
                        '          +\'<span style="width:28px;font-size:10px;color:#555">\'+Math.round(v*100)+\'%</span></div>\';\n'
                        '      });\n'
                        '      ibh+=\'</div><div style="font-size:10px;font-weight:bold;color:\'+bdCol+\';white-space:nowrap;margin-left:4px">→\'+bestDim.toUpperCase()+\'</div></div>\';\n'
                        '    });\n'
                        '    impactBody.innerHTML=ibh;\n'
                        '  }else if(impactBody){impactBody.innerHTML=\'<span style="color:#aaa;font-size:11px">No fail die data</span>\';}\n'
                        '  var allIbArr=[];\n'
                        '  var _ibAllSeen={};\n'
                        '  Object.keys(WM_PAT.wafers||{}).forEach(function(pk){var w=WM_PAT.wafers[pk];if(!w||!w.dies)return;w.dies.forEach(function(d){var ib=d[2];if(ib!==null&&ib!==undefined)_ibAllSeen[String(ib)]=true;});});\n'
                        '  Object.keys(_ibAllSeen).forEach(function(k){if(k!=="null"&&k!=="undefined")allIbArr.push(+k);});\n'
                        '  allIbArr.sort(function(a,b){return a-b;});\n'
                        '  _wmPatBuildBinRow(allIbArr);\n'
                        '  _wmPatBuildRetRow(keys);\n'
                        '  _wmBuildModeMap(keys);\n'
                        '  if(typeof _wmApplyBinFilter==="function")_wmApplyBinFilter();\n'
                        '}\n'
                        'function wmPatLTab(t){\n'
                        '  document.querySelectorAll(".wm-pat-ltab").forEach(function(b){b.classList.toggle("on",b.dataset.ltab===t);});\n'
                        '  document.querySelectorAll(".wm-pat-lpane").forEach(function(p){p.classList.toggle("on",p.id==="wm-pat-lpane-"+t);});\n'
                        '}\n'
                        'function wmmIbHlClick(ibk){\n'
                        '  _wmmHlIb=(_wmmHlIb!==null&&String(_wmmHlIb)===String(ibk))?null:String(ibk);\n'
                        '  if(_wmPatBuildBinRow.lastArr)_wmPatBuildBinRow(_wmPatBuildBinRow.lastArr);\n'
                        '  if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function _wmmHeatColor(t){\n'
                        '  if(t<=0)return"#f0f0f0";\n'
                        '  t=Math.min(1,t);\n'
                        '  var hue=Math.round(30*(1-t));\n'
                        '  var sat=t<0.05?Math.round(t/0.05*100):100;\n'
                        '  var lit=Math.round(95-65*t);\n'
                        '  return"hsl("+hue+","+sat+"%,"+lit+"%)";\n'
                        '}\n'
                        'function _wmmToggleHeat(){\n'
                        '  _wmmHeatMode=!_wmmHeatMode;\n'
                        '  if(_wmPatBuildBinRow.lastMapKeys)_wmBuildModeMap(_wmPatBuildBinRow.lastMapKeys);\n'
                        '}\n'
                        'function _wmBuildModeMap(keys){\n'
                        '  var _retShots=WM_PAT.retShots||[];\n'
                        '  var _retMap=WM_PAT.retMap||{};\n'
                        '  if(keys&&keys.length){var _ri0=_wmRetInfoFor(keys[0]);if(_ri0){_retShots=_ri0.retShots||_retShots;_retMap=_ri0.retMap||_retMap;}}\n'
                        '  _wmPatBuildBinRow.lastMapKeys=keys;\n'
                        '  var el=document.getElementById("wm-pat-modemap-body");if(!el)return;\n'
                        '  var el2=document.getElementById("wm-pat-modemap-body2");\n'
                        '  if(!keys||!keys.length){el.innerHTML=`<span style="color:#999;font-size:11px">No wafers selected</span>`;if(el2)el2.innerHTML=el.innerHTML;return;}\n'
                        '  var pos={};\n'
                        '  keys.forEach(function(pk){\n'
                        '    var wdata=WM_PAT.wafers[pk];\n'
                        '    var dies=wdata&&wdata.dies?wdata.dies:wdata;\n'
                        '    if(!dies||!dies.length)return;\n'
                        '    dies.forEach(function(d){\n'
                        '      var x=d[0],y=d[1],ib=d[2];if(x===null||x===undefined)return;\n'
                        '      var key=x+","+y;\n'
                        '      if(!pos[key])pos[key]={x:x,y:y,cnt:{}};\n'
                        '      var ibk=String(ib===null||ib===undefined?"null":ib);\n'
                        '      pos[key].cnt[ibk]=(pos[key].cnt[ibk]||0)+1;\n'
                        '    });\n'
                        '  });\n'
                        '  var entries=Object.values(pos);\n'
                        '  if(!entries.length){el.innerHTML=`<span style="color:#999;font-size:11px">No die data</span>`;return;}\n'
                        '  var xs=entries.map(function(e){return e.x;}),ys=entries.map(function(e){return e.y;});\n'
                        '  var xMin=Math.min.apply(null,xs),xMax=Math.max.apply(null,xs);\n'
                        '  var yMin=Math.min.apply(null,ys),yMax=Math.max.apply(null,ys);\n'
                        '  var maxW=580,pad=8;\n'
                        '  var cs=Math.max(5,Math.floor((maxW-pad*2)/(xMax-xMin+1)));\n'
                        '  var xSpan=xMax-xMin,ySpan=yMax-yMin;\n'
                        '  var csy=(xSpan>0&&ySpan>0)?(cs*xSpan/ySpan):cs;\n'
                        '  var W=Math.round((xMax-xMin+1)*cs+pad*2),H=Math.round((ySpan+1)*csy+pad*2);\n'
                        '  var cx=pad+xSpan/2*cs+cs/2,cy=pad+ySpan/2*csy+csy/2;\n'
                        '  var rx=xSpan/2*cs+cs/2,ry=ySpan/2*csy+csy/2;\n'
                        '  var clipId="wmm-clip-"+Date.now();\n'
                        '  var _rsl=WM_PAT.retSiteLabels||{};if(keys&&keys.length){var _ri0b=_wmRetInfoFor(keys[0]);if(_ri0b&&_ri0b.retSiteLabels)_rsl=_ri0b.retSiteLabels;}\n'
                        '  var _snumM=WM_PAT._retSiteNum||{};var dieNumMap={};if(_retMap){Object.keys(_retMap).forEach(function(k){var v=_retMap[k];var sk=v[0]+","+v[1];var lbl=(_snumM[sk]!=null)?_snumM[sk]:_rsl[sk];if(lbl!=null)dieNumMap[k]=lbl;});};\n'
                        '  var fsize2=Math.max(5,Math.min(9,Math.round(cs*0.55)));\n'
                        '  var rects=[],legSeen={};\n'
                        '  var shotHlCount={},shotTotCount={};\n'
                        '  var _failCounts={},_maxFailCount=1;\n'
                        '  if(_wmmHeatMode){\n'
                        '    entries.forEach(function(e){\n'
                        '      var fc=0;Object.keys(e.cnt).forEach(function(ibk){\n'
                        '        if(ibk!=="null"&&(_wmPatBinChecked===null||_wmPatBinChecked.has(ibk)))fc+=e.cnt[ibk];\n'
                        '      });\n'
                        '      _failCounts[e.x+","+e.y]=fc;\n'
                        '      if(fc>_maxFailCount)_maxFailCount=fc;\n'
                        '    });\n'
                        '  }\n'
                        '  entries.forEach(function(e){\n'
                        '    var modeIb="null",modeCount=0;\n'
                        '    var _sortedIbks=Object.keys(e.cnt).sort(function(a,b){return e.cnt[b]-e.cnt[a];});\n'
                        '    for(var _ki=0;_ki<_sortedIbks.length;_ki++){var _ibk=_sortedIbks[_ki];if(_wmPatBinChecked===null||_wmPatBinChecked.has(_ibk)){modeIb=_ibk;modeCount=e.cnt[_ibk];break;}}\n'
                        '    var ibVal=modeIb==="null"?null:(isNaN(+modeIb)?null:+modeIb);\n'
                        '    var fill=modeCount>0?_wmIbColor(ibVal):"white";\n'
                        '    var dispFill;\n'
                        '    if(_wmmHeatMode){var _fc=_failCounts[e.x+","+e.y]||0;dispFill=_wmmHeatColor(_fc/_maxFailCount);}\n'
                        '    else if(_wmmHlIb!==null){var _hlC=e.cnt[_wmmHlIb]||0;var _hlF=_wmIbColor(+_wmmHlIb);dispFill=_hlC>0?_hlF:"white";}else{dispFill=fill;}\n'
                        '    if(_wmPatRetUnchecked&&_wmPatRetUnchecked.size>0&&_retMap){var _cri=_retMap[e.x+","+e.y];if(_cri&&_wmPatRetUnchecked.has(_cri[0]+","+_cri[1]))dispFill="white";}\n'
                        '    if(_wmPatShotUnchecked&&_wmPatShotUnchecked.size>0&&_retMap){var _csi=_retMap[e.x+","+e.y];if(_csi&&_wmPatShotUnchecked.has(_csi[2]))dispFill="white";}\n'
                        '    if(_wmEdgeExcRows>0&&(e.x<xMin+_wmEdgeExcRows||e.x>xMax-_wmEdgeExcRows||e.y<yMin+_wmEdgeExcRows||e.y>yMax-_wmEdgeExcRows))dispFill="rgba(220,220,220,0.3)";\n'
                        '    if(_retMap&&_retMap[e.x+","+e.y]){var _msi=_retMap[e.x+","+e.y][2];shotTotCount[_msi]=(shotTotCount[_msi]||0)+1;if(dispFill!=="white")shotHlCount[_msi]=(shotHlCount[_msi]||0)+1;}\n'
                        '    var px=(pad+(e.x-xMin)*cs).toFixed(1),py=(pad+(yMax-e.y)*csy).toFixed(1);\n'
                        '    var _legIb="null",_legCnt=0;Object.keys(e.cnt).forEach(function(ibk){if(e.cnt[ibk]>_legCnt){_legCnt=e.cnt[ibk];_legIb=ibk;}});\n'
                        '    var _legVal=_legIb==="null"?null:(isNaN(+_legIb)?null:+_legIb);\n'
                        '    var _ibDesc="";if(WM_PAT.ibDesc&&_legIb!="null")_ibDesc=WM_PAT.ibDesc[_legIb]||"";\n'
                        '    legSeen[_legIb]={fill:_wmIbColor(_legVal),label:_legVal===null?"N/A":"IB"+_legIb,desc:_ibDesc};\n'
                        '    rects.push(`<rect data-ib="${modeIb}" data-fill="${dispFill}" x="${px}" y="${py}" width="${cs.toFixed(1)}" height="${csy.toFixed(1)}" fill="${dispFill}" stroke="#fff" stroke-width="0.3" class="wmm-die"/>`);\n'
                        '    var _dieTag="";\n'
                        '    var _tagFs=Math.max(3,Math.min(6,Math.round(cs*0.35)));\n'
                        '    if(_dieTag&&cs>=4){rects.push(`<text x="${(parseFloat(px)+cs-0.5).toFixed(1)}" y="${(parseFloat(py)+_tagFs+0.5).toFixed(1)}" text-anchor="end" font-size="${_tagFs}" fill="#000" font-weight="bold" pointer-events="none">${_dieTag}</text>`);}\n'
                        '    var _dnum=dieNumMap[e.x+","+e.y]||"";\n'
                        '    if(_dnum){var _tc=(function(c){if(!c||c==="white")return"#444";var h=c.replace("#","");if(h.length===3)h=h[0]+h[0]+h[1]+h[1]+h[2]+h[2];var r=parseInt(h.substr(0,2),16),g=parseInt(h.substr(2,2),16),b=parseInt(h.substr(4,2),16);return(0.299*r+0.587*g+0.114*b)<128?"#fff":"#222";})(dispFill);rects.push(`<text x="${(parseFloat(px)+cs/2).toFixed(1)}" y="${(parseFloat(py)+csy*0.62).toFixed(1)}" text-anchor="middle" font-size="${fsize2}" fill="${_tc}" pointer-events="none">${_dnum}</text>`);}\n'
                        '  });\n'
                        '  var shotRects="";\n'
                        '  if(_retShots&&_retShots.length){\n'
                        '    var fsize=Math.max(7,Math.min(14,Math.round(cs*0.85)));\n'
                        '    _retShots.forEach(function(s,si){\n'
                        '      var sx=(pad+(s[0]-xMin)*cs).toFixed(1),sy=(pad+(yMax-s[3])*csy).toFixed(1);\n'
                        '      var sw=((s[2]-s[0]+1)*cs).toFixed(1),sh=((s[3]-s[1]+1)*csy).toFixed(1);\n'
                        '      shotRects+=`<rect x="${sx}" y="${sy}" width="${sw}" height="${sh}" fill="none" stroke="#1a6bb0" stroke-width="1.5" opacity="0.85"/>`;\n'
                        '      if(cs>=6){var tx=(parseFloat(sx)+parseFloat(sw)/2).toFixed(1),ty=(parseFloat(sy)+9).toFixed(1);shotRects+=`<text x="${tx}" y="${ty}" text-anchor="middle" font-size="8" fill="#1a6bb0" opacity="0.85" pointer-events="none">${si}</text>`;}\n'
                        '    });\n'
                        '  }\n'
                        '  var svgStr=`<svg id="wmm-svg" width="${W}" height="${H}" xmlns="http://www.w3.org/2000/svg" style="display:block">`\n'
                        '    +`<defs><clipPath id="${clipId}"><ellipse cx="${cx}" cy="${cy}" rx="${rx}" ry="${ry}"/></clipPath></defs>`\n'
                        '    +`<g clip-path="url(#${clipId})">${rects.join("")}</g>`\n'
                        '    +`<ellipse cx="${cx}" cy="${cy}" rx="${rx}" ry="${ry}" fill="none" stroke="#bdc3c7" stroke-width="1.5"/>`\n'
                        '    +shotRects+`</svg>`;\n'
                        '  var legKeys=Object.keys(legSeen).sort(function(a,b){var na=a==="null"?9999:+a,nb=b==="null"?9999:+b;return na-nb;});\n'
                        '  var n=keys.length;\n'
                        '  var _heatLeg=_wmmHeatMode?`<div style="display:flex;align-items:center;gap:4px;font-size:10px;color:#666;flex-shrink:0"><span>Low</span><div style="width:80px;height:8px;border-radius:3px;background:linear-gradient(to right,#f0f0f0,hsl(30,100%,62%),hsl(0,100%,30%))"></div><span>High bin density</span></div>`:\'\';\n'
                        '  var _btnStyle=_wmmHeatMode?"background:#c0392b;color:#fff;border:1px solid #c0392b":"background:#f8f9fa;color:#1a5276;border:1px solid #bdc3c7";\n'
                        '  el.innerHTML=`<div style="display:flex;flex-direction:column;align-items:center;gap:6px;width:100%">` \n'
                        '    +`<div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;justify-content:center">` \n'
                        '    +`<p style="font-size:10px;color:#666;margin:0">${_wmmHeatMode?"Bin Density":"Mode IB"} &middot; ${n} wafer${n!==1?"s":""} &mdash; ${_wmmHeatMode?"darker = more bins across wafers":"use IB Filter above to highlight"}</p>` \n'
                        '    +`<button onclick="_wmmToggleHeat()" style="font-size:9px;padding:2px 8px;border-radius:3px;cursor:pointer;flex-shrink:0;${_btnStyle}">${_wmmHeatMode?"&#127752; IB Mode":"&#128293; Bin Density"}</button>` \n'
                        '    +`</div>` \n'
                        '    +svgStr \n'
                        '    +_heatLeg \n'
                        '    +`</div>`;\n'
                        '  if(el2)el2.innerHTML=el.innerHTML;\n'
                        '  var _cIbs=legKeys.filter(function(k){return k!=="null";}).map(function(k){return+k;});\n'
                        '  if(_cIbs.length){\n'
                        '    var _bIbs=[];\n'
                        '    document.querySelectorAll("#wm-pat-binrow input[data-ib]").forEach(function(inp){_bIbs.push(+inp.dataset.ib);});\n'
                        '    var _mIbs=Array.from(new Set(_bIbs.concat(_cIbs))).sort(function(a,b){return a-b;});\n'
                        '    _wmPatBuildBinRow(_mIbs);\n'
                        '  }\n'
                        '}\n'
                        'function wmmLegClick(el){\n'
                        '  var svg=document.getElementById("wmm-svg");if(!svg)return;\n'
                        '  var ib=el.dataset.ib;\n'
                        '  var leg=document.getElementById("wmm-leg");\n'
                        '  var active=el.classList.contains("wmm-sel");\n'
                        '  var items=leg?leg.querySelectorAll("[data-ib]"):[];\n'
                        '  items.forEach(function(it){it.classList.remove("wmm-sel");it.style.borderColor="#ddd";it.style.fontWeight="";});\n'
                        '  svg.querySelectorAll(".wmm-die").forEach(function(r){r.setAttribute("fill",r.getAttribute("data-fill"));});\n'
                        '  if(!active){\n'
                        '    el.classList.add("wmm-sel");el.style.borderColor="#1a6bb0";el.style.fontWeight="bold";\n'
                        '    svg.querySelectorAll(".wmm-die").forEach(function(r){\n'
                        '      if(r.getAttribute("data-ib")!==ib)r.setAttribute("fill","white");\n'
                        '    });\n'
                        '  }\n'
                        '}\n'
                        'function _wmPatInitDrag(){\n'
                        '  var drag=document.getElementById("wm-pat-drag");\n'
                        '  var box=document.getElementById("wm-pat-box");\n'
                        '  if(!drag||!box||box._dragInit)return;\n'
                        '  box._dragInit=true;\n'
                        '  var ox=0,oy=0,bx=0,by=0;\n'
                        '  drag.addEventListener("mousedown",function(e){\n'
                        '    if(e.target.closest&&(e.target.closest("button")||e.target.closest("select")))return;\n'
                        '    ox=e.clientX;oy=e.clientY;\n'
                        '    var r=box.getBoundingClientRect();bx=r.left;by=r.top;\n'
                        '    box.style.left=bx+"px";box.style.top=by+"px";box.style.right="auto";\n'
                        '    function onMove(ev){box.style.left=(bx+ev.clientX-ox)+"px";box.style.top=(by+ev.clientY-oy)+"px";}\n'
                        '    function onUp(){document.removeEventListener("mousemove",onMove);document.removeEventListener("mouseup",onUp);}\n'
                        '    document.addEventListener("mousemove",onMove);\n'
                        '    document.addEventListener("mouseup",onUp);\n'
                        '    e.preventDefault();\n'
                        '  });\n'
                        '}\n'
                        '(function(){\n'
                        '  var resizer=document.getElementById("wm-pat-scores-resize");\n'
                        '  if(!resizer)return;\n'
                        '  resizer.addEventListener("mousedown",function(e){\n'
                        '    var panel=document.getElementById("wm-pat-scores-panel");\n'
                        '    if(!panel||panel.classList.contains("collapsed"))return;\n'
                        '    var startY=e.clientY,startH=panel.getBoundingClientRect().height;\n'
                        '    function onMove(ev){var newH=Math.max(40,startH-(ev.clientY-startY));panel.style.height=newH+"px";}\n'
                        '    function onUp(){document.removeEventListener("mousemove",onMove);document.removeEventListener("mouseup",onUp);}\n'
                        '    document.addEventListener("mousemove",onMove);\n'
                        '    document.addEventListener("mouseup",onUp);\n'
                        '    e.preventDefault();\n'
                        '  });\n'
                        '})();\n'
                        '(function(){\n'
                        '  var vs=document.getElementById("wm-pat-vsplit");\n'
                        '  if(!vs)return;\n'
                        '  vs.addEventListener("mousedown",function(e){\n'
                        '    var left=vs.previousElementSibling;\n'
                        '    var inner=vs.parentElement;\n'
                        '    if(!left||!inner)return;\n'
                        '    var startX=e.clientX,startW=left.getBoundingClientRect().width,innerW=inner.getBoundingClientRect().width;\n'
                        '    function onMove(ev){\n'
                        '      var nw=Math.max(180,Math.min(innerW-8,startW+(ev.clientX-startX)));\n'
                        '      left.style.width=nw+"px";left.style.flex="none";\n'
                        '      var fullyExpanded=(nw>=innerW-8);\n'
                        '      vs.style.display=fullyExpanded?"none":"";\n'
                        '      var right=vs.nextElementSibling;\n'
                        '      if(right){right.style.display=fullyExpanded?"none":"";}\n'
                        '    }\n'
                        '    function onUp(){document.removeEventListener("mousemove",onMove);document.removeEventListener("mouseup",onUp);}\n'
                        '    document.addEventListener("mousemove",onMove);\n'
                        '    document.addEventListener("mouseup",onUp);\n'
                        '    e.preventDefault();\n'
                        '  });\n'
                        '})();\n'
                        'function wmLoad(url,row,ev){\n'
                        '  var isCtrl=ev&&(ev.ctrlKey||ev.metaKey);\n'
                        '  if(isCtrl){\n'
                        '    if(_wmSel.has(url)){_wmSel.delete(url);if(row)row.classList.remove("wm-active");}\n'
                        '    else{_wmSel.set(url,row);if(row)row.classList.add("wm-active");}\n'
                        '  }else{\n'
                        '    _wmSel.forEach(function(r){if(r)r.classList.remove("wm-active");});\n'
                        '    _wmSel.clear();\n'
                        '    _wmSel.set(url,row);\n'
                        '    if(row)row.classList.add("wm-active");\n'
                        '  }\n'
                        '  _wmRender();\n'
                        '  /* Show/hide pattern analysis bar based on lot selection */\n'
                        '  if(!isCtrl){\n'
                        '    var _pk=row&&row.dataset&&row.dataset.patkey;\n'
                        '    var _dl=row&&row.dataset&&row.dataset.lot;\n'
                        '    var lot=_dl||(_pk&&_pk.split("::")[0])||null;\n'
                        '    var bar=document.getElementById("wm-map-bar");\n'
                        '    var barInfo=document.getElementById("wm-map-bar-info");\n'
                        '    if(lot){\n'
                        '      _wmCurLot=lot;\n'
                        '      if(bar)bar.style.display="flex";\n'
                        '    }else{\n'
                        '      _wmCurLot=null;\n'
                        '      if(bar)bar.style.display="none";\n'
                        '      wmHidePat();\n'
                        '    }\n'
                        '  }\n'
                        '}\n'
                        'function _wmRender(){\n'
                        '  var wrap=document.getElementById("wm-frames");\n'
                        '  var urls=[];\n'
                        '  _wmSel.forEach(function(r,url){urls.push(url);});\n'
                        '  var curFrames=wrap.querySelectorAll("iframe");\n'
                        '  if(urls.length===1&&curFrames.length===1){\n'
                        '    var newUrl=urls[0];\n'
                        '    var curSrc=curFrames[0].src||"";\n'
                        '    var newBase=newUrl.split("#")[0],curBase=curSrc.split("#")[0];\n'
                        '    var newFname=newBase.split("/").pop(),curFname=curBase.split("/").pop();\n'
                        '    if(newFname&&curFname&&newFname===curFname){\n'
                        '      var newHash=newUrl.indexOf("#")>=0?newUrl.slice(newUrl.indexOf("#")):"#";\n'
                        '      try{curFrames[0].contentWindow.location.hash=newHash;}catch(e){curFrames[0].src=newUrl;}\n'
                        '      return;\n'
                        '    }\n'
                        '  }\n'
                        '  wrap.innerHTML="";\n'
                        '  _wmSel.forEach(function(r,url){\n'
                        '    var f=document.createElement("iframe");\n'
                        '    f.src=url;\n'
                        '    wrap.appendChild(f);\n'
                        '  });\n'
                        '}\n'
                        'window.addEventListener("load",function(){\n'
                        '  var first=document.querySelector(".wm-lot-row");\n'
                        '  if(first)first.click();\n'
                        '});\n'
                        '(function(){\n'
                        '  var handle=document.getElementById("wm-nav-resize");\n'
                        '  var nav=document.querySelector(".wm-nav");\n'
                        '  if(!handle||!nav)return;\n'
                        '  handle.addEventListener("mousedown",function(e){\n'
                        '    if(e.button!==0)return;\n'
                        '    e.preventDefault();e.stopPropagation();\n'
                        '    handle.classList.add("dragging");\n'
                        '    document.body.style.userSelect="none";\n'
                        '    var shield=document.createElement("div");\n'
                        '    shield.style.cssText="position:fixed;top:0;left:0;right:0;bottom:0;z-index:99999;cursor:ns-resize;";\n'
                        '    document.body.appendChild(shield);\n'
                        '    var startY=e.clientY,startH=nav.getBoundingClientRect().height;\n'
                        '    function mm(ev){nav.style.height=Math.max(60,startH+(ev.clientY-startY))+"px";}\n'
                        '    function mu(){document.removeEventListener("mousemove",mm);document.removeEventListener("mouseup",mu);handle.classList.remove("dragging");document.body.style.userSelect="";shield.remove();}\n'
                        '    document.addEventListener("mousemove",mm);\n'
                        '    document.addEventListener("mouseup",mu);\n'
                        '  });\n'
                        '})();\n'
                        '(function(){\n'
                        '  var h=location.hash||"";\n'
                        '  if(h!=="#wpa")return;\n'
                        '  _wmPatIsPopup=true;\n'
                        '  document.title="Wafer Pattern Analysis";\n'
                        '  var nav=document.querySelector(".wm-nav");if(nav)nav.style.display="none";\n'
                        '  var hdr=document.querySelector(".wm-hdr");if(hdr)hdr.style.display="none";\n'
                        '  var res=document.getElementById("wm-nav-resize");if(res)res.style.display="none";\n'
                        '  var bar=document.getElementById("wm-map-bar");if(bar)bar.style.display="none";\n'
                        '  var drag=document.getElementById("wm-pat-drag");if(drag){drag.style.cursor="default";drag.style.borderRadius="0";drag.style.userSelect="auto";}\n'
                        '  var fr=document.getElementById("wm-frames");if(fr)fr.style.display="none";\n'
                        '  var ov=document.getElementById("wm-pat-overlay");\n'
                        '  if(ov){ov.style.cssText="display:block;position:static;background:none;pointer-events:auto";}\n'
                        '  var box=document.getElementById("wm-pat-box");\n'
                        '  if(box){box.style.cssText="position:static;width:100%;height:100vh;max-width:none;max-height:none;min-width:0;min-height:0;border-radius:0;resize:none;display:flex;flex-direction:column;overflow:hidden";}\n'
                        '  wmShowPatLot(_wmPatAllLots());\n'
                        '})();\n'

                        '</script>\n'
                        '</body></html>\n'
                    )
                    _wm_combined_path = out_dir / 'wafermap.html'
                    _wm_combined_path.write_text(_wm_inject(_wm_combined_html), encoding='utf-8')
                    _wm_nav = (
                        f'  <div class="sec">Wafer Map</div>\n'
                        f'  <a class="nav-link" href="wafermap.html" '
                        f'onclick="loadWpa(this);return false;">'
                        f'&#128202; Wafer Pattern Analysis</a>\n'
                        f'  <a class="nav-link" href="wafermap.html" '
                        f'onclick="load(\'wafermap.html\',this);return false;">'
                        f'&#128507; IBIN Wafer Map</a>\n'
                    )
                except Exception as _wm_ex:
                    import traceback as _tb_wm
                    print(f'[WPA ERROR] {_wm_ex}')
                    _tb_wm.print_exc()
                    _wm_entries = ''
                    for _wf in _wm_files:
                        lot_part = _wf.stem[len(f'{stem}_IBIN_WaferMap_'):]
                        _wm_entries += (
                            f'  <a class="nav-link" href="heatmap/{_wf.name}" '
                            f'onclick="load(\'heatmap/{_wf.name}\',this);return false;">'
                            f'&#128507; {lot_part}</a>\n'
                        )
                    _wm_nav = f'  <div class="sec">Wafer Map</div>\n{_wm_entries}'

            _upm_dist_sidebar = (
                f'  <div class="sec">UPM Distribution</div>\n  {_upm_dist_section}'
                if _upm_dist_section else ''
            )

            _pcm_analysis_p = out_dir / 'pcm_analysis.html'
            _pcm_sidebar = (
                '  <div class="sec">Parametric Dashboard</div>\n'
                '  <a class="nav-link" href="#" onclick="load(\'pcm_analysis.html\',this);return false;">'
                '&#128202; Parametric Dashboard</a>\n'
            ) if _pcm_analysis_p.exists() else ''

            _sidebar_inner = (
                f'<!-- SIDEBAR_START -->\n'
                f'  <h2>{_tag_display}</h2>\n'
                f'  <div class="sec">Yield</div>\n'
                f'  {yield_section}'
                f'  {_wm_nav}'
                f'<!-- AFTER_WMAP_NAV -->\n'
                f'  {_pcm_sidebar}'
                f'  {sicc_sidebar}'
                f'  {_upm_dist_sidebar}'
                f'  {plots_sidebar}\n'
                + '\n<!-- SIDEBAR_END -->'
            )
            html = f"""<!doctype html>
<html>
<head><meta charset="utf-8"><title>{_tag_display}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{display:flex;height:100vh;font-family:Arial,sans-serif;overflow:hidden}}
#sidebar{{width:270px;min-width:30px;max-height:100vh;background:#2c3e50;flex-shrink:0;position:relative;overflow:hidden;transition:width .2s}}
#sidebar.sb-collapsed{{width:30px}}
#sidebar-inner{{overflow-y:scroll;max-height:100vh;scrollbar-width:thin;scrollbar-color:#4a6a8a #2c3e50}}
#sidebar.sb-collapsed #sidebar-inner{{visibility:hidden}}
#sb-toggle{{position:absolute;top:6px;right:4px;z-index:10;width:22px;height:22px;background:#2980b9;color:#fff;border:none;border-radius:3px;cursor:pointer;font-size:15px;line-height:22px;text-align:center;padding:0}}
#sb-toggle:hover{{background:#3498db}}
#sidebar-inner::-webkit-scrollbar{{width:8px}}
#sidebar-inner::-webkit-scrollbar-track{{background:#2c3e50}}
#sidebar-inner::-webkit-scrollbar-thumb{{background:#4a6a8a;border-radius:4px}}
#sidebar-inner::-webkit-scrollbar-thumb:hover{{background:#5a8ab5}}
#sidebar h2{{color:#ecf0f1;padding:14px 30px 14px 16px;font-size:13px;background:#1a252f;letter-spacing:.5px}}
.sec{{color:#95a5a6;font-size:10px;text-transform:uppercase;padding:20px 16px 4px;letter-spacing:1px;border-top:1px solid #3d5166;margin-top:8px}}
.nav-link{{display:block;color:#bdc3c7;text-decoration:none;padding:7px 16px;font-size:12px;
           border-left:3px solid transparent;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.nav-link:hover,.nav-link.active{{background:#34495e;color:#fff;border-left-color:#3498db}}
.sub-link{{padding-left:28px;font-size:11px;color:#a9cff0}}
.sub-link:hover,.sub-link.active{{background:#34495e;color:#fff;border-left-color:#3498db}}
.subsub-link{{padding-left:44px;font-size:10px;color:#78b8e8}}
.subsub-link:hover,.subsub-link.active{{background:#34495e;color:#fff;border-left-color:#3498db}}
.xlsx-link{{color:#82e0aa;border-left-color:transparent}}
.xlsx-link:hover{{background:#34495e;color:#fff;border-left-color:#2ecc71}}
.sicc-link{{color:#f1948a;border-left-color:transparent}}
.sicc-link:hover{{background:#34495e;color:#fff;border-left-color:#e74c3c}}
#content{{flex:1;min-width:0}}
iframe{{width:100%;height:100%;border:none;display:block}}
</style>
</head>
<body>
<div id="sidebar">
  <button id="sb-toggle" title="Toggle sidebar" onclick="var sb=document.getElementById('sidebar');sb.classList.toggle('sb-collapsed');this.textContent=sb.classList.contains('sb-collapsed')?'›':'‹';">&#x2039;</button>
  <div id="sidebar-inner">
  {_sidebar_inner}
  </div>
</div>
<div id="content"><iframe id="frame" src="{first_src}"></iframe></div>
<!-- SCRIPT_START -->
<script>
var _GP_PORT={opener_port or 0};
var _GP_BASE='{str(out_dir).replace(chr(92), "/")}';
var _GP_FIRST='{first_src}';
function _gp(r){{
  if(!_GP_PORT||!r||r.indexOf('://')>=0)return r;
  var h='',s=r;var hi=s.indexOf('#');if(hi>=0){{h=s.slice(hi);s=s.slice(0,hi);}}
  if(!s)return r;
  var b=_GP_BASE;if(b.slice(-1)!='/')b+='/';
  return 'http://127.0.0.1:'+_GP_PORT+'/serve?p='+encodeURIComponent(b+s)+h;
}}
function load(url,el){{
  document.getElementById('frame').src=_gp(url);
  document.querySelectorAll('.nav-link:not(.xlsx-link)').forEach(a=>a.classList.remove('active'));
  el.classList.add('active');
}}
function loadWpa(el){{
  document.querySelectorAll('.nav-link:not(.xlsx-link)').forEach(a=>a.classList.remove('active'));
  el.classList.add('active');
  var f=document.getElementById('frame');
  var src=f.src||'';
  var isWm=src.indexOf('wafermap.html')>=0||(src.indexOf('/serve?')>=0&&src.indexOf('wafermap')>=0);
  if(isWm){{
    try{{
      var cw=f.contentWindow;
      if(cw&&cw.wmShowPatLot){{cw.wmShowPatLot(cw._wmPatAllLots?cw._wmPatAllLots():null);return;}}
    }}catch(e){{}}
    return;
  }}
  f.src=_gp('wafermap.html#wpa');
}}
var first=document.querySelector('.nav-link:not(.xlsx-link)[onclick]');
if(first){{first.classList.add('active');first.click();}}
window.addEventListener('message',function(e){{
  if(e.data&&typeof e.data.navFrame==='string'){{
    var f=document.getElementById('frame');
    if(f)f.src=_gp(e.data.navFrame)||e.data.navFrame;
  }}
}});
(function(){{if(_GP_PORT&&_GP_FIRST)document.getElementById('frame').src=_gp(_GP_FIRST);}})();
</script>
<!-- SCRIPT_END -->
</body></html>"""
            _gp_base_val = str(out_dir).replace('\\', '/')
            _script_block = (
                '<!-- SCRIPT_START -->\n'
                '<script>\n'
                f'var _GP_PORT={opener_port or 0};\n'
                f'var _GP_BASE=\'{_gp_base_val}\';\n'
                f'var _GP_FIRST=\'{first_src}\';\n'
                'function _gp(r){\n'
                '  if(!_GP_PORT||!r||r.indexOf(\'://\')>=0)return r;\n'
                '  var h=\'\',s=r;var hi=s.indexOf(\'#\');if(hi>=0){h=s.slice(hi);s=s.slice(0,hi);}\n'
                '  if(!s)return r;\n'
                '  var b=_GP_BASE;if(b.slice(-1)!=\'/\')b+=\'/\';\n'
                '  return \'http://127.0.0.1:\'+_GP_PORT+\'/serve?p=\'+encodeURIComponent(b+s)+h;\n'
                '}\n'
                'function load(url,el){\n'
                '  document.getElementById(\'frame\').src=_gp(url);\n'
                '  document.querySelectorAll(\'.nav-link:not(.xlsx-link)\').forEach(a=>a.classList.remove(\'active\'));\n'
                '  el.classList.add(\'active\');\n'
                '}\n'
                'function loadWpa(el){\n'
                '  document.querySelectorAll(\'.nav-link:not(.xlsx-link)\').forEach(a=>a.classList.remove(\'active\'));\n'
                '  el.classList.add(\'active\');\n'
                '  var f=document.getElementById(\'frame\');\n'
                '  var src=f.src||\'\';\n'
                '  var isWm=src.indexOf(\'wafermap.html\')>=0||(src.indexOf(\'/serve?\')>=0&&src.indexOf(\'wafermap\')>=0);\n'
                '  if(isWm){\n'
                '    try{var cw=f.contentWindow;if(cw&&cw.wmShowPatLot){cw.wmShowPatLot(cw._wmPatAllLots?cw._wmPatAllLots():null);return;}}catch(e){}\n'
                '    return;\n'
                '  }\n'
                '  f.src=_gp(\'wafermap.html#wpa\');\n'
                '}\n'
                'var first=document.querySelector(\'.nav-link:not(.xlsx-link)[onclick]\');\n'
                'if(first){first.classList.add(\'active\');first.click();}\n'
                'window.addEventListener(\'message\',function(e){\n'
                '  if(e.data&&typeof e.data.navFrame===\'string\'){\n'
                '    var f=document.getElementById(\'frame\');\n'
                '    if(f)f.src=_gp(e.data.navFrame)||e.data.navFrame;\n'
                '  }\n'
                '});\n'
                '(function(){if(_GP_PORT&&_GP_FIRST)document.getElementById(\'frame\').src=_gp(_GP_FIRST);}());\n'
                '</script>\n'
                '<!-- SCRIPT_END -->'
            )
            master = out_dir / 'index.html'
            if master.exists():
                import re as _reidx
                _existing = master.read_text(encoding='utf-8')
                _sidebar_re = _reidx.compile(
                    r'<!-- SIDEBAR_START -->[\s\S]*?<!-- SIDEBAR_END -->'
                )
                _script_re = _reidx.compile(
                    r'<!-- SCRIPT_START -->[\s\S]*?<!-- SCRIPT_END -->'
                )
                if _sidebar_re.search(_existing):
                    _updated = _sidebar_re.sub(lambda m: _sidebar_inner, _existing)
                    # Also update iframe src to match the current run's first_src
                    if first_src:
                        _iframe_re = _reidx.compile(r'(<div id="content"><iframe id="frame" src=")[^"]*(")')
                        _updated = _iframe_re.sub(lambda m: m.group(1) + first_src + m.group(2), _updated)
                    if _script_re.search(_updated):
                        _updated = _script_re.sub(lambda m: _script_block, _updated)
                    else:
                        # Older file with sidebar sentinel but no script sentinel
                        # — replace the bare <script>function load...block
                        _old_sc = _reidx.compile(r'<script>\s*function load\b[\s\S]*?</script>', _reidx.IGNORECASE)
                        if _old_sc.search(_updated):
                            _updated = _old_sc.sub(_script_block, _updated, count=1)
                        elif '</body>' in _updated:
                            _updated = _updated.replace('</body>', _script_block + '\n</body>', 1)
                    # Inject sidebar toggle button if missing (older files)
                    if 'id="sb-toggle"' not in _updated:
                        _sb_toggle_css = (
                            '#sidebar{min-width:30px;position:relative;overflow:hidden;transition:width .2s}\n'
                            '#sidebar.sb-collapsed{width:30px}\n'
                            '#sidebar-inner{overflow-y:scroll;max-height:100vh;scrollbar-width:thin;scrollbar-color:#4a6a8a #2c3e50}\n'
                            '#sidebar.sb-collapsed #sidebar-inner{visibility:hidden}\n'
                            '#sb-toggle{position:absolute;top:6px;right:4px;z-index:10;width:22px;height:22px;'
                            'background:#2980b9;color:#fff;border:none;border-radius:3px;cursor:pointer;'
                            'font-size:15px;line-height:22px;text-align:center;padding:0}\n'
                            '#sb-toggle:hover{background:#3498db}\n'
                        )
                        _sb_toggle_btn = '<button id="sb-toggle" title="Toggle sidebar" onclick="document.getElementById(\'sidebar\').classList.toggle(\'sb-collapsed\')">&#9776;</button>\n'
                        if '</style>' in _updated:
                            _updated = _updated.replace('</style>', _sb_toggle_css + '</style>', 1)
                        # Wrap existing sidebar content in sidebar-inner if not already done
                        import re as _re2
                        if '<div id="sidebar-inner">' not in _updated:
                            _updated = _re2.sub(
                                r'(<div id="sidebar">)\s*(<!-- SIDEBAR_START -->)',
                                r'\1\n  ' + _sb_toggle_btn.strip() + r'\n  <div id="sidebar-inner">\n  \2',
                                _updated, count=1
                            )
                            _updated = _updated.replace('<!-- SIDEBAR_END -->\n</div>', '<!-- SIDEBAR_END -->\n  </div>\n</div>', 1)
                        elif '<div id="sidebar">' in _updated:
                            _updated = _updated.replace('<div id="sidebar">', '<div id="sidebar">\n  ' + _sb_toggle_btn.strip(), 1)
                    # Update .sec CSS to current style (spacing/separator)
                    _sec_re = _reidx.compile(r'\.sec\{[^}]*\}')
                    _new_sec = '.sec{color:#95a5a6;font-size:10px;text-transform:uppercase;padding:20px 16px 4px;letter-spacing:1px;border-top:1px solid #3d5166;margin-top:8px}'
                    if _sec_re.search(_updated):
                        _updated = _sec_re.sub(_new_sec, _updated, count=1)
                    # Keep _GP_FIRST / _GP_PORT / _GP_BASE in sync for existing files
                    # (the script block replacement above handles this, but patch inline too)
                    _gp_first_re = _reidx.compile(r"var _GP_FIRST='[^']*'")
                    if _gp_first_re.search(_updated) and first_src:
                        _updated = _gp_first_re.sub(f"var _GP_FIRST='{first_src}'", _updated)
                    master.write_text(_wm_inject(_updated), encoding='utf-8')
                else:
                    # Older file without sentinels — overwrite with fresh HTML
                    master.write_text(_wm_inject(html), encoding='utf-8')
            else:
                master.write_text(_wm_inject(html), encoding='utf-8')
            return str(master)
        except Exception:
            return None

    def _update_dashboard_html(self, resolved_csv, master_html=None, dashboard_html=None, dashboard_html_dir=None, plot_html=None, plot_tag_files=None, sicc_links=None, output_dir=None):
        """Create or update Dashboard.html.
        Location priority: dashboard_html path > dashboard_html_dir > output_dir > CSV parent.
        Each run gets its own collapsible section, identified by the CSV stem.
        Existing sections for the same stem are replaced; others are kept.
        sicc_links: list of (abs_path, label, css_class) from SICC/UPM headless run."""
        try:
            from pathlib import Path as _P
            import os as _os
            from datetime import datetime as _dt

            csv_p = _P(resolved_csv)
            out_dir = csv_p.parent / 'output'
            stem = csv_p.stem
            # Display name: prefer Identifier, then TestProgram, then TestProgram_folder basename, else stem
            _tp_name = self.testprogram_id_var.get().strip() or self.testprogram_var.get().strip()
            if not _tp_name:
                _tpf = self.tp_folder_var.get().strip()
                _tp_name = os.path.basename(_tpf) if _tpf else ''
            report_name = _tp_name if _tp_name else stem
            # Use the identifier as the unique block key so re-running
            # the same identifier always overrides the existing block.
            _identifier = self.testprogram_id_var.get().strip()
            _safe_id_key = ''.join(c if c.isalnum() or c in '-_.' else '_' for c in _identifier) if _identifier else ''
            block_key = _safe_id_key if _safe_id_key else stem

            # Determine Dashboard.html location: use configured dashboard html path directly,
            # fall back to dashboard_html_dir, output_dir parent, or CSV parent.
            if dashboard_html and dashboard_html.strip():
                # Use the provided path even if the file doesn't exist yet —
                # it will be created below.
                dash_html_path = _P(dashboard_html)
                dash_dir = dash_html_path.parent
            elif dashboard_html_dir:
                dash_dir = _P(dashboard_html_dir)
                dash_html_path = dash_dir / 'Dashboard.html'
            elif output_dir:
                dash_dir = _P(output_dir).parent
                dash_html_path = dash_dir / 'Dashboard.html'
            else:
                dash_dir = csv_p.parent
                dash_html_path = dash_dir / 'Dashboard.html'
            dash_dir.mkdir(parents=True, exist_ok=True)

            # Build relative paths from dash_dir to the various output files
            def _rel(abs_path):
                try:
                    return _os.path.relpath(str(abs_path), str(dash_dir)).replace('\\', '/')
                except Exception:
                    return str(abs_path)

            # Collect links for this run (Yield Report only; SICC/UPM reserved for JSL run)
            run_links = []
            # Dashboard Yield Report (master index.html)
            _yield_html = master_html
            if not _yield_html or not _os.path.isfile(str(_yield_html)):
                # Fallback: look for an existing index.html in the output folder
                # (from a prior yield run for this same identifier)
                _scan_dirs_yield = []
                if output_dir:
                    _scan_dirs_yield.append(_P(output_dir))
                _scan_dirs_yield.extend([csv_p.parent / 'output', csv_p.parent])
                for _sd in _scan_dirs_yield:
                    _idx_cand = _sd / 'index.html'
                    if _idx_cand.exists():
                        _yield_html = str(_idx_cand)
                        break
            if not _yield_html or not _os.path.isfile(str(_yield_html)):
                # Broader fallback: output_dir may be e.g. .../ID-1 or .../ID-cdyn
                # while index.html lives in .../ID (base identifier folder).
                # Strip trailing -SUFFIX and check that sibling folder.
                if output_dir:
                    import re as _re_yield
                    _out_p = _P(output_dir)
                    _base_name = _re_yield.sub(r'-[^/\\]+$', '', _out_p.name)
                    if _base_name != _out_p.name:
                        _sibling = _out_p.parent / _base_name
                        if (_sibling / 'index.html').exists():
                            _yield_html = str(_sibling / 'index.html')
                    # Also scan all sibling dirs for index.html (most recent first)
                    if not _yield_html or not _os.path.isfile(str(_yield_html)):
                        _siblings = [d / 'index.html' for d in _out_p.parent.iterdir()
                                     if d.is_dir() and d != _out_p and (d / 'index.html').exists()]
                        if _siblings:
                            _siblings.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                            _yield_html = str(_siblings[0])
            if _yield_html and _os.path.isfile(str(_yield_html)):
                run_links.append((_rel(_yield_html), 'Dashboard Yield Report', 'report-link'))
            # SICC/UPM link removed — reserved for SICC/UPM JSL run

            _now = _dt.now()
            ts_date = _now.strftime('%Y-%m-%d')
            ts_time = _now.strftime('%H:%M')
            link_items = '\n'.join(
                f'<a class="run-link {cls}" href="{href}" target="_blank">{lbl}</a>'
                for href, lbl, cls in run_links
            )
            # Each run block is wrapped in a div with data-stem for replacement detection
            new_block = (
                f'<div class="run-block" data-stem="{block_key}">\n'
                f'<div class="run-header" onclick="toggle(this)">'
                f'<span class="arrow">&#9660;</span> {report_name}'
                f'<span class="ts"> - {ts_date} - {ts_time}</span></div>\n'
                f'<div class="run-body">\n{link_items}\n</div>\n</div>'
            )

            # CSS + JS template (only written when file is created fresh)
            page_css = """
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#1a252f;color:#ecf0f1;padding:16px}
h1{font-size:16px;margin-bottom:14px;color:#3498db;letter-spacing:.5px}
.run-block{background:#2c3e50;border-radius:6px;margin-bottom:10px;overflow:hidden}
.run-header{padding:10px 14px;cursor:pointer;display:flex;align-items:center;gap:6px;
  font-weight:bold;font-size:13px;user-select:none}
.run-header:hover{background:#34495e}
.arrow{font-size:10px;transition:transform .2s}
.run-header.collapsed .arrow{transform:rotate(-90deg)}
.ts{font-weight:normal;font-size:11px;color:#95a5a6;margin-left:auto}
.run-body{padding:8px 14px 12px;display:flex;flex-wrap:wrap;gap:6px}
.run-link{display:inline-block;padding:5px 10px;border-radius:4px;font-size:12px;
  text-decoration:none;white-space:nowrap}
.report-link{background:#2980b9;color:#fff}
.report-link:hover{background:#3498db}
.bin-link{background:#1e8449;color:#fff}
.bin-link:hover{background:#27ae60}
.heat-link{background:#7d3c98;color:#fff}
.heat-link:hover{background:#9b59b6}
.xlsx-link{background:#1f618d;color:#aed6f1}
.xlsx-link:hover{background:#2980b9;color:#fff}
.plot-link{background:#117a65;color:#a9dfbf}
.plot-link:hover{background:#1abc9c;color:#fff}
.sicc-link{background:#7b241c;color:#f1948a}
.sicc-link:hover{background:#922b21;color:#fff}
.param-link{background:#4a235a;color:#d2b4de}
.param-link:hover{background:#6c3483;color:#fff}
.param-link{background:#4a235a;color:#d2b4de}
.param-link:hover{background:#6c3483;color:#fff}
"""
            page_js = """
function toggle(hdr){
  hdr.classList.toggle('collapsed');
  hdr.nextElementSibling.style.display=hdr.classList.contains('collapsed')?'none':'';
}
"""
            # Sentinel comments for the three sections (Yield / Compare TP / Vmin)
            YIELD_START   = '<!-- YIELD_START -->'
            YIELD_END     = '<!-- YIELD_END -->'
            COMPARE_START = '<!-- COMPARE_START -->'
            COMPARE_END   = '<!-- COMPARE_END -->'
            VMIN_START    = '<!-- VMIN_START -->'
            VMIN_END      = '<!-- VMIN_END -->'
            OLD_START     = '<!-- RUNS_START -->'  # legacy
            OLD_END       = '<!-- RUNS_END -->'    # legacy
            START         = YIELD_START
            END           = YIELD_END

            def _is_valid_html(path):
                """Return True if the file looks like HTML (not a binary xlsx/zip)."""
                try:
                    with open(path, 'rb') as fh:
                        head = fh.read(256)
                    # xlsx / zip files start with PK\x03\x04
                    if head[:4] == b'PK\x03\x04':
                        return False
                    # Check for common HTML markers in the first 256 bytes
                    head_lower = head.lower()
                    if b'<!doctype' in head_lower or b'<html' in head_lower or b'<head' in head_lower:
                        return True
                    # If it's mostly printable text, treat it as HTML
                    try:
                        head.decode('utf-8')
                        return True
                    except UnicodeDecodeError:
                        return False
                except Exception:
                    return False

            if dash_html_path.exists() and _is_valid_html(dash_html_path):
                content = dash_html_path.read_text(encoding='utf-8')
                # Migrate legacy RUNS_START/END sentinels to YIELD_START/END
                content = content.replace(OLD_START, YIELD_START).replace(OLD_END, YIELD_END)
                # Inject .sicc-link CSS into existing files that pre-date this feature
                import re as _re
                if '.sicc-link' not in content and '</style>' in content:
                    _sicc_css = ('.sicc-link{background:#7b241c;color:#f1948a}\n'
                                 '.sicc-link:hover{background:#922b21;color:#fff}\n'
                                 '.param-link{background:#4a235a;color:#d2b4de}\n'
                                 '.param-link:hover{background:#6c3483;color:#fff}\n')
                    content = content.replace('</style>', _sicc_css + '</style>', 1)
                elif '.param-link' not in content and '</style>' in content:
                    _param_css = ('.param-link{background:#4a235a;color:#d2b4de}\n'
                                  '.param-link:hover{background:#6c3483;color:#fff}\n')
                    content = content.replace('</style>', _param_css + '</style>', 1)
                # Replace existing section for this block_key if present
                import re as _re

                def _make_block_re(stem_pattern):
                    """Build a regex that matches a full run-block div by counting
                    opening/closing div tags so nested content doesn't break it."""
                    return _re.compile(
                        r'<div class="run-block" data-stem="' + stem_pattern +
                        r'">\s*'
                        r'<div class="run-header"[^>]*>[\s\S]*?</div>\s*'
                        r'<div class="run-body">[\s\S]*?</div>\s*'
                        r'</div>',
                        _re.MULTILINE
                    )

                block_re = _make_block_re(_re.escape(block_key))
                # Only replace the block with the EXACT same data-stem (block_key).
                # Previous code used a broad substring match that could
                # accidentally erase unrelated blocks sharing a common prefix.
                _found = block_re.search(content)
                if _found:
                    # Preserve any param-link anchors already injected by the
                    # parametric runner so they survive a yield pipeline re-run.
                    _old_body = _found.group(0)
                    _preserved = _re.findall(
                        r'<a class="run-link param-link"[^>]*>.*?</a>', _old_body)
                    if _preserved:
                        # Append preserved links into new_block's run-body
                        _extra = '\n' + '\n'.join(_preserved)
                        new_block = new_block.replace(
                            '</div>\n</div>', _extra + '\n</div>\n</div>', 1)
                    content = block_re.sub('', content)
                    # Clean up any blank lines left behind
                    content = _re.sub(r'\n{3,}', '\n\n', content)
                    if START in content:
                        content = content.replace(START, START + '\n' + new_block)
                    else:
                        content = content.replace('</body>', new_block + '\n</body>')
                elif START in content and END in content:
                    # Prepend new block inside the runs container
                    content = content.replace(START, START + '\n' + new_block)
                else:
                    # Fallback: append before </body>
                    content = content.replace('</body>', new_block + '\n</body>')
                dash_html_path.write_text(_wm_inject(content), encoding='utf-8')
            else:
                full_html = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Yield Dashboard</title>
<style>{page_css}</style></head>
<body>
<h1>&#128202; Yield Analysis Dashboard</h1>
{START}
{new_block}
{END}
<script>{page_js}</script>
</body></html>"""
                dash_html_path.write_text(_wm_inject(full_html), encoding='utf-8')

            return str(dash_html_path)
        except Exception:
            return None

