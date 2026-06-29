#!/usr/intel/pkgs/python3/3.7.4/bin/python3
"""
Get Digital Dashboard Update: Takes a .csv input file from Crystal Ball with FBINs and
LOGTRACKER results to determine two columns for the Digital Dashboard containing the yield
summary by sub module.

Usage:
    get_dd_update.py -h --help
    get_dd_update.py -v --version
    get_dd_update.py [--data=data.csv] [--bin_defs=bin_defs.csv] [--vmax] [--wafer]
        [--log=log_file.log] [--verbose] [--debug]

Options:
    -h --help                       Show this screen.
    -v --version                    Show version.
    -d --data=<input file>          .csv Input file containing a list of die and FBINs.
    -b --bin_defs=<input file>      .csv file containing a table of DBIN and Test name.
    -g --log=<log file>             Optionally, log to a file.
    -m --vmax                       Make the Vmax summary as well.
    -w --wafer                      Print columns per wafer.
    --verbose                       Print more information than usual.
    -x --debug                      Print lots of debugging statements.
    --dashboard=<file>              Optional path to DigitalDashBoard.xlsx to append results
"""
import sys

try:
    from docopt import docopt
except ModuleNotFoundError:
    docopt = None
try:
    import pandas
except ModuleNotFoundError:
    print("Python 'pandas' module not installed. Install with pip:")
    print("  python -m pip install --user --proxy \"http://proxy-us.intel.com:911\" pandas openpyxl")
    sys.exit()
import logging
import traceback
import datetime
import pathlib
from pathlib import Path
import re
import os
import shutil
import zlib
import time
import csv
import tempfile
# Ensure user site-packages are on sys.path (needed when launched by double-click)
import site as _site, sys as _sys
_usp = _site.getusersitepackages()
if _usp not in _sys.path:
    _sys.path.insert(0, _usp)
try:
    import UsrIntel.R2  #Required for openpyxl below.
except ModuleNotFoundError:
    try:
        import UsrIntel.R1
    except ModuleNotFoundError:
        pass  # optional on non-Intel environments
try:
    import openpyxl
    HAVE_OPENPYXL = True
except Exception:
    openpyxl = None
    HAVE_OPENPYXL = False


__version__ = "get_dd_update 0.0"


#todo: If this gets out of hand, move it to a json file.
moduleMap = {
    #Good bins
    "B198_PASS": {"dd": "Bin 198 (Vmin Repair)", "vmax":"Bin 198 (Vmin Repair)"},
    r"B1((?!98)\d\d)_PASS": {"dd": "Bin 1", "vmax":"Bin 1 (No Repair)"},
    "B201_PASS": {"dd": "Bin 2 (Hard Repair)", "vmax":"Bin 2 (Hard Repair)"},
    "B202_PASS": {"dd": "Bin 202 (Vmax Repair)", "vmax":"Bin 202 (Vmax Repair)"},
    "B226_PASS": {"dd": "Bin 2 (Hard Repair)", "vmax":"Bin 2 (Hard Repair)"},
    r"B3\d\d_PASS": {"dd": "Bin 3", "vmax":"Bin 3"},
    r"B4\d\d_PASS": {"dd": "Bin 4", "vmax":"Bin 4"},
    #RESET
    r"B19\d{6}_FAIL_": {"dd": "Reset", "vmax": "Reset"},
    r"B35\d{6}_FAIL_DRV_RESET": {"dd": "Reset", "vmax": "Reset"},
    #ARR ATOM
    r"ARR_ATOM_.*_VNOM_LFM_0800": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"ARR_ATOM_.*_VATOM_.*_F1_0800": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"ARR_ATOM_.*_NOM_LFM": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"ARR_ATOM_.*_VNOM_LFM": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"ARR_ATOM_.*_VMIN_LFM_0800": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"B6326\d{4}_FAIL_HVQK": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"ARR_ATOM_.*_VMIN_.*_VATOM_.*_F6": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_TFM (B2050,B6050)"},
    r"ARR_ATOM_.*_F1_.*_MAX": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VMAX_LFM (B6050)"},
    r"ARR_ATOM_.*MAX_LFM_.": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VMAX_LFM (B6050)"},
    r"ARR_ATOM_.*_(F5|F6)_.*_MAX": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VMAX_TFM (B6050)"},
    r"ARR_ATOM_.*_MAX_TFM_": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VMAX_TFM (B6050)"},
    r"B6050\d{4}_FAIL_ARR_ATOM_.*_3200_.*": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VMAX_TFM (B6050)"},
    #ARR_CCF
    r"ARR_CCF_.*_VNOM_.*_F1_": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"B6244\d{4}_.*_VMIN_.*_F1_": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"B6226\d{4}": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"ARR_CCF_.*_VNNAON_.*_F1_": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"ARR_CCF.*_VNOM_LFM_": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"ARR_CCF_.*_(F1|LFM)_.*_800MV_": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"ARR_CCF_.*_F1_.*_MAX": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VMAX_LFM (B6242)"},
    r"ARR_CCF_.*_F6_.*_MAX": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VMAX_TFM (B2043)"},
    r"ARR_CCF_.*_VMIN_.*_(F5|F6)_.*": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_TFM (B62)"},
    r"ARR_CCF_.*_VCCR_.*_(F5|F6)_.*": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_TFM (B62)"},
    r"ARR_CCF_.*_VMIN_.*_F1_.*": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"ARR_CCF_.*_VCCR_.*_(?:F1|FMIN)_.*": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    #ARR_NONCCF
    r"ARR_MBIST_.*_ALL_NONCCF_.*_MAX": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VMAX_LFM (B61)"},
    r"ARR_MBIST_.*_ALL_NONCCF_.*_KS_\d$": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"B6126\d{4}_FAIL_HVQK": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"ARR_MBIST_.*_ALL_NONCCF_.*_RETENTION": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"ARR_MBIST_.*_(L|S)SA_NONCCF_.*_PREHVQK": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"ARR_MBIST_.*_ROM_NONCCF_.*_PREHVQK": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"ARR_UNCORE.*_VNNAON(_NOM|_X_X_X)?_LFM": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"ARR_UNCORE.*_VNNAON(_MAX|_X_X_X)?_LFM": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VMAX_LFM (B21,B60,B61)"},
    #ARR_CORE
    r"ARR_MBIST_.*_ALL_CORE_.*_VMAX_.*_F1_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VMAX_LFM (B60)"},
    r"ARR_CORE.*_MAX_.*_F1_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VMAX_LFM (B60)"},
    r"B6017\d{4}_FAIL_ARR_MBIST_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B60)"},
    r"B6019\d{4}_FAIL_ARR_MBIST_.*_F1_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B60)"},
    r"_FAIL_ARR_CORE_.*_F5_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VMAX_TFM (B60)"},
    r"^(?!.*_VMAX_).*ARR_MBIST_.*_ALL_CORE_.*_F[456]_$": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_TFM (B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_MBIST_.*_ALL_CORE_.*_F[456]_\d{4}_\d$": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_TFM (B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_MBIST_.*_CORE_.*_PREHVQK": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_)B2000\d{4}.*ARR_MBIST_.*_CORE_.*_END_.*": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_MBIST_.*_CORE_.*_END_.*_F1_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_CORE.*_CORE_.*_END_.*_LFM_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_CORE.*_CORE_.*_NOM_(?:LFM|FMIN)_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_MBIST_.*_CORE_.*_EXVF_.*_F1_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"B6026\d{4}": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"},
    r"B2026\d{4}": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"},
    r"ARR_MBIST_.*_ALL_CORE_.*_VMAX_.*_F6_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VMAX_TFM (B60)"}, #must not contain VMAX
    "FAIL_UNCORRECTABLE_ECC_ERROR": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    #FUN_ATOM
    r"B44\d{6}_FAIL_FUN_ATOM_.*_V(MIN|ATOM)_.*_F(0|1)_((?!DRAGON_SLC).)*$":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_VNOM_LFM (B445x)"},
    r"B44\d{6}_FAIL_FUN_ATOM_.*_V(MIN|ATOM)_.*_F(0|1)_.*_DRAGON_SLC":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_DRAGON_SLC_VNOM_LFM (B4463)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_F6_((?!DRAGON_SLC).)*$":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_VNOM_TFM (B446x)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_TFM_4400_ATOM_L2_DRAGON":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_VNOM_TFM (B446x)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_VATOM_.*_(?:XFM|TFM)_":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_VNOM_TFM (B446x)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_TFM_ATOM_L2":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_VNOM_TFM (B44xx)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_F6_.*_DRAGON_SLC":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_DRAGON_SLC_VNOM_TFM (B4465)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_TFM_3800_.*_DRAGON_SLC":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_DRAGON_SLC_VNOM_TFM (B4465)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_TFM_.*SLC_DRAGON":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_DRAGON_SLC_VNOM_TFM (B4465)"},
    r"B44\d{6}_FAIL_FUN_ATOM_.*_DRAGON_SLC_SPECKLE":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_DRAGON_SLC_SPECKLE (B4466)"},
    #FUN_CCF
    r"B45\d{6}_FAIL_FUN_CCF_.*_V(MIN|NOM).*_LFM_": {"dd": "FUN_CCF", "vmax": "FUN_CCF_VNOM_LFM (B4538)"},
    r"B45\d{6}_FAIL_FUN_CCF_.*_NOM.*_LFM_": {"dd": "FUN_CCF", "vmax": "FUN_CCF_VNOM_LFM (B4538)"},
    r"B45\d{6}_FAIL_FUN_CCF_.*_V(MIN|NOM)_.*_F6_": {"dd": "FUN_CCF", "vmax": "FUN_CCF_VNOM_TFM (B4538)"},
    r"B45\d{6}_FAIL_FUN_CCF_.*_V(MIN|NOM)_.*_TFM_": {"dd": "FUN_CCF", "vmax": "FUN_CCF_VNOM_TFM (B4538)"},
    r"B4526\d{4}": {"dd": "FUN_CCF", "vmax": "FUN_CCF_VNOM_LFM (B4538)"},
    #FUN_CORE
    r"B4426\d{4}": {"dd": "FUN_CORE", "vmax": "FUN_CORE_VNOM_LFM (B44)"},
    r"B44.*_SBFT_CORE_.*_VMIN_LFM_": {"dd": "FUN_CORE", "vmax": "FUN_CORE_VNOM_LFM (B44)"},
    r"B44.*_SBFT_CORE_VMIN_.*_LFM_": {"dd": "FUN_CORE", "vmax": "FUN_CORE_VNOM_LFM (B44)"},
    r"B44.*_SBFT_CORE_.*_VMAX_LFM_": {"dd": "FUN_CORE", "vmax": "FUN_CORE_VMAX_LFM (B44)"},
    r"B44.*_SBFT_CORE.*_VMIN_.*_(?:TFM|F6)_": {"dd": "FUN_CORE", "vmax": "FUN_CORE_VNOM_TFM (B44)"},
    #SCN_ATOM
    r"SCN_ATOM_.*_VNOM_LFM": {"dd": "SCN_ATOM", "vmax": "SCN_ATOM_VNOM_LFM (B41,B42,B47)"},
    r"SCN_ATOM_.*_VATOM_NOM_LFM": {"dd": "SCN_ATOM", "vmax": "SCN_ATOM_VNOM_LFM (B41,B42,B47)"},
    r"SCN_ATOM_.*_VNOM_TFM": {"dd": "SCN_ATOM", "vmax": "SCN_ATOM_VNOM_TFM (B41,B42,B47)"},
    r"SCN_ATOM_.*_VATOM_NOM_TFM": {"dd": "SCN_ATOM", "vmax": "SCN_ATOM_VNOM_TFM (B41,B42,B47)"},
    #SCN_CCF
    r"SCN_UNCORE_.*_CCF_.*_LFM": {"dd": "SCN_CCF", "vmax":"SCN_CCF_VNOM_LFM (B41,B42,B47)"},
    r"SCN_UNCORE_.*_SEC_.*_LFM": {"dd": "SCN_CCF", "vmax":"SCN_CCF_VNOM_LFM (B41,B42,B47)"},
    r"SCN_UNCORE_.*_HRY_.*_LFM": {"dd": "SCN_CCF", "vmax":"SCN_CCF_VNOM_LFM (B41,B42,B47)"},
    r"SCN_UNCORE_.*_VCCR_.*_LFM": {"dd": "SCN_CCF", "vmax":"SCN_CCF_VNOM_LFM (B41,B42,B47)"},
    r"SCN_UNCORE_.*_(?:CCF|UNCORE)_.*_TFM": {"dd": "SCN_CCF", "vmax":"SCN_CCF_VNOM_TFM (B41,B42,B47)"},
    #SCN_CORE
    r"SCN_CORE_.*_LFM": {"dd": "SCN_CORE", "vmax":"SCN_CORE_VNOM_LFM (B41,B42,B47)"},
    r"SCN_CORE_.*_TFM": {"dd": "SCN_CORE", "vmax":"SCN_CORE_VNOM_TFM (B41,B42,B47)"},
    r"B4(1|2)26\d{4}": {"dd": "SCN_CORE", "vmax":"SCN_CORE_VNOM_LFM (B41,B42,B47)"},
    #SCN_UNCORE
    r"SCN_UNCORE_.*_NONCCF": {"dd": "SCN_NONCCF", "vmax":"SCN_NONCCF_VNOM_LFM (B41,B42,B47)"},
    r"B4726\d{4}_FAIL_HVQK": {"dd": "SCN_NONCCF", "vmax":"SCN_NONCCF_VNOM_LFM (B41,B42,B47)"},
    #Analog
    r"^B24\d{6}": {"dd":"Analog", "vmax": "Analog Other (B24)"},
    r"^B27\d{6}_FAIL_PTH_BG": {"dd":"Analog", "vmax": "Analog PTH BG (B27)"},
    r"^B27\d{6}_FAIL_PTH_DLVR": {"dd":"Analog", "vmax": "Analog PTH DLVR (B27)"},
    r"^B27\d{6}_FAIL_HVQK": {"dd":"Analog", "vmax": "Analog PTH DLVR (B27)"},
    r"^B28\d{6}": {"dd":"Analog", "vmax": "Analog CLK (B28)"},
    r"^B29\d{6}": {"dd":"Analog", "vmax": "Analog PTH BGR (B29)"},
    r"^B36\d{6}": {"dd":"Analog", "vmax": "Analog MIO D2D (B36)"},
    r"^B40\d{6}": {"dd":"Analog", "vmax": "Analog PTH ODI (B40)"},
    r"^B64\d{6}": {"dd":"Analog", "vmax": "Analog PTH DTS (B64)"},
    #TPI_FOUNDRY
    "TPI_ADTL": {"dd": "TPI Foundry", "vmax": "TPI Foundry ADTL (B43)"},
    r"B4326\d{4}": {"dd": "TPI Foundry", "vmax": "TPI Foundry ADTL (B43)"},
    r"B8\d\d_": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B08\d\d_": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B89\d{6}_": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B80\d{6}_": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B10\d{6}_FAIL_TPI": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B15\d{6}_FAIL_TPI": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B40\d{6}_FAIL_PTH": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B4026\d{4}_FAIL_HVQK": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B18\d{6}_FAIL_PTH_POWER_.*SICC": {"dd": "TPI Foundry", "vmax": "TPI Foundry SICC (B18)"},
    #TPI_OTHER
    r"^B17\d{6}": {"dd":"TPI Other", "vmax":"TPI Other"},
    r"^B26\d{6}": {"dd":"HVQK (26)", "vmax":"HVQK (26)"},
    r"^B30\d{6}": {"dd":"TPI Other", "vmax": "TPI Other"},
    r"^B31\d{6}": {"dd":"TPI Other", "vmax": "TPI Other"},
    r"^B49\d{6}": {"dd":"TPI Other", "vmax": "TPI Other"},
    r"^B53\d{6}": {"dd":"TPI Other", "vmax": "TPI Other"},
    r"^B69\d{6}": {"dd":"TPI Other", "vmax": "TPI Other"},
    r"B88\d{6}_FAIL_TPI_SIU": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"B94\d{6}_": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"FAIL_DUT_TEMPERATURE": {"dd": "TPI Other", "vmax": "TPI Other"},
    "FAIL_DPS_OVERVOLTAGE_ALARM": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"FAIL_HALT_ALARM": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"B97\d{6}_": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"B98\d{6}_.*": {"dd": "TPI Other - B98", "vmax": "TPI Other"},
    r"B9099\d{4}.*_ALARM": {"dd": "TPI Other - B99", "vmax": "TPI Other"}, #IB is B99
    r"B90(?!99)\d{6}_FAIL_": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"B99\d\d.*": {"dd": "TPI Other - B99", "vmax": "TPI Other"},
    "B93": {"dd": "TPI Other - B93", "vmax": "TPI Other"},
}


def getDD(dataInFile="", binDefFile="", vmax=False, logBaseName="",
          waferLvl=False, idt=2, dashboardFile="", outDir=""):
    productInfo = {
        "ARL68-N3B": {
            "DPW": 797,
            "DEVREVSTEP": ["8PYJCVJ"],
            "TPrgx": r"(E6\w)",
            "dfBins": [3,4],
            "numCores":6,
            "numAtoms": 8,
        },
        "ARLS816": {
            "DPW": 516,
            "DEVREVSTEP": ["8PYVCVB","8PYVCVAB"],
            "TPrgx": r"(8[2,3]\w)",
            "dfBins": [3,4],
            "numCores": 8,
            "numAtoms": 16,
        },
        "NVL48": {
            "DPW": 1200,
            "DEVREVSTEP": ["8PY6CVT"],
            "TPrgx": r"(8[1,2]\w)",
            "dfBins": [3, 4],
            "numCores": 4,
            "numAtoms": 8,
        },
        "NVL816": {
            "DPW": 619,
            "DEVREVSTEP": ["8PF6CVP", "8PF6CVR", "8PF6CVER"],
            "TPrgx": r"(5[1,2]\w)",
            "dfBins": [3, 4],
            "numCores": 8,
            "numAtoms": 16,
        },
        "NVL816-BLLC": {
            "DPW": 393,
            "DEVREVSTEP": ["8PF5CVL","8PF5CVEL"],
            "TPrgx": r"(6[01]\w)",
            "dfBins": [3, 4],
            "numCores": 8,
            "numAtoms": 16,
        }
    }
    devRevSteps = {
        "8PYJCVJ": "ARL68-N3B",
        "8PYVCVB": "ARLS816",
        "8PYVCVAB": "ARLS816",
        "8PF6CVP": "NVL816",
        "8PF6CVR": "NVL816",
        "8PF5CVL": "NVL816-BLLC",
    }

    # moduleMap defined at module level (importable by external callers)

    defeatureModCnts = {} #{"FUN_CORE":18,"SCN_ATOM":37...}
    outFile = f"{pathlib.Path(dataInFile).stem}_out.xlsx"
    # If an explicit output directory was provided, use it (highest priority).
    if outDir:
        try:
            _od = pathlib.Path(outDir)
            _od.mkdir(parents=True, exist_ok=True)
            outFile = str((_od / outFile).resolve())
        except Exception:
            pass
    # If a dashboard path was provided, place the generated output workbook
    # in the same directory as the dashboard so outputs live next to it.
    elif dashboardFile:
        try:
            dbp = pathlib.Path(dashboardFile)
            if dbp.parent and str(dbp.parent) != '.':
                outFile = str((dbp.parent / outFile).resolve())
        except Exception:
            # on any error, fall back to the original outFile in cwd
            pass

    info = {}
    logging.info(f"{' ':{idt}}Checking inputs...")
    product, yieldDF = getYieldDataFrame(dataInFile, productInfo)
    info["product"] = product

    # logging.info(f"Length of df: {len(yieldDF)}")
    if not float(len(yieldDF) / productInfo[product]["DPW"]).is_integer():
        logging.warning(f"{' ':{idt}}Extra die found!  Number of die in data set ({len(yieldDF)} "
                        f"is not evenly divisible by DPW ({productInfo[product]['DPW']}).")
    numWafers = len(yieldDF) // productInfo[product]["DPW"]
    logging.info(f"{' ':{idt+2}}Found {numWafers} wafers.")
    info["numWafers"] = numWafers
    if waferLvl:
        info["numDie"] = productInfo[product]["DPW"]
    else:
        info["numDie"] = len(yieldDF)
    try:
        info["TP"] = getTPNum(df=yieldDF, rgx=productInfo[product]["TPrgx"], idt=2)
    except LookupError as e:
        logging.warning(f"{' ':{idt+2}}Can't find TP, defaulting to '??'.  Error:\n{e}")
        info["TP"] = "??"
    # binDefs = pandas.read_csv(binDefFile).set_index("B/C").to_dict("list")
    if binDefFile and os.path.isfile(str(binDefFile)):
        binDefs = dict(pandas.read_csv(binDefFile).values)
        logging.info(f"{' ':{idt}}Loaded {len(binDefs)} bin definitions from '{binDefFile}'.")
    else:
        if binDefFile:
            logging.warning(f"{' ':{idt}}bindef file '{binDefFile}' not found — building from data CSV.")
        else:
            logging.info(f"{' ':{idt}}No bindef file provided — building bin definitions from data CSV.")
        binDefs = _buildBinDefsFromDF(yieldDF)
        logging.info(f"{' ':{idt}}Built {len(binDefs)} bin definitions from data CSV.")
    lotCol  = getLotCol(df=yieldDF)
    waferCol = getWaferCol(df=yieldDF)
    if waferLvl:
        moduleYield = {}
        for lot in sorted(yieldDF[lotCol].unique()):
            for wafer in sorted(yieldDF.loc[yieldDF[lotCol] == lot][waferCol].unique()):
                lotWaferDF = yieldDF.loc[(yieldDF[lotCol] == lot) & (yieldDF[waferCol] == wafer)]
                moduleYield[f"{lot}_W{wafer}"] = getYieldByModule(yldDF=lotWaferDF,
                                                    binDefs=binDefs, modMap=moduleMap)
    else:
        moduleYield = getYieldByModule(yldDF=yieldDF, binDefs=binDefs, modMap=moduleMap)
    if not waferLvl:
        updateDefeatureModCnts(dfModCnts=defeatureModCnts, yieldDF=yieldDF,
                               waferLvl=waferLvl, productInfo=productInfo[product])
    # print(defeatureModCnts)

    if not vmax:
        if waferLvl:
            raise UnboundLocalError("Wafer level not nupported in makeOutXl!")
        makeOutXl(outFile=outFile, moduleYield=moduleYield, which="dd",
              defeatureModCnts=defeatureModCnts, info=info)
        # If a dashboard file was provided, append the output column
        if dashboardFile:
            try:
                # Only attempt xlsx append if the dashboard path looks like a workbook
                _db_suffix = pathlib.Path(dashboardFile).suffix.lower()
                if _db_suffix not in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
                    logging.debug(f"{' ':{idt}}Skipping dashboard append: '{dashboardFile}' is not an Excel workbook.")
                else:
                    # srcSheet falls back to first sheet inside append_to_dashboard if not found
                    append_to_dashboard(dashboardPath=dashboardFile, srcWorkbook=outFile,
                                         srcSheet=info.get("product", ""))
            except Exception as e:
                logging.warning(f"{' ':{idt}}Failed to append to dashboard: {e}")
    outFile = f"{pathlib.Path(dataInFile).stem}_out_vmax.xlsx"
    if outDir:
        try:
            _od = pathlib.Path(outDir)
            _od.mkdir(parents=True, exist_ok=True)
            outFile = str((_od / outFile).resolve())
        except Exception:
            pass
    elif dashboardFile:
        try:
            dbp = pathlib.Path(dashboardFile)
            if dbp.parent and str(dbp.parent) != '.':
                outFile = str((dbp.parent / outFile).resolve())
        except Exception:
            pass
    if vmax:
        makeOutXlVmax(outFile=outFile, allModuleYield=moduleYield, info=info, moduleMap=moduleMap, waferLvl=waferLvl)
    material = listOfTuplesToDict(yieldDF.value_counts([lotCol,waferCol]).keys().to_list())
    s = "; ".join([f"{key}:{','.join(str(x) for x in value)}" for key,value in material.items()])
    logging.info(f"{' ':{idt}}Material Used for this analysis:\n\n{s}\n")



    # print(len(yieldDF), len(yieldDF["FB@F24_132110"].value_counts()))

    # logging.info(f"FBin Summary:\n{yieldDF['FB@F24_132110'].value_counts().nlargest(5)}")
    # series = yieldDF["FB@F24_132110"].value_counts()
    # for k,v in series.items():
    #     print(k, v)
    # for i in range(20):
    #     for col in yieldDF:
    #         if "LOGTRACKER_AM" in col or "LOGTRACKER_CR" in col:
    #             s = yieldDF[col][i]
    #             if s:
    #                 logging.info(f"{' ':{idt+2}} Decoded:")
    #                 print(prime_error_decode(s[10:].strip("=")).split("\n"))
    #     break
    #     s = yieldDF["TPI_BIN::CTRL_UB_X_E_FINAL_X_X_X_X_LOGTRACKER_AM0_01@F24_132110"][i]
    #     if s:
    #         logging.info(f"Inflating: '{s[10:len(s)-3]}'")
    #         sDecoded = prime_error_decode(s[10:len(s)-3])
    #         logging.info(sDecoded)
    #         break
    # for i in range(10):
    #     logging.info(df["FB@F24_132110"][i])
    # for fb in df["FB@F24_132110"]:
    #     logging.info(fb)
    logging.debug("All done.")



def getProductFromProgram(program=""):
    if "ARCSDSCB0" in program:
        return "ARLS816"
    elif "ARCSDSCJ0" in program:
        return "ARL68-N3B"
    else:
        return False



def updateDefeatureModCnts(dfModCnts={}, yieldDF=pandas.DataFrame(), productInfo={}, waferLvl=False, idt=4,
                           binDefs=None, modMap=None):
    """Populate dfModCnts (module → count) for defeatured (IB3/IB4) die.

    Primary path: decode LOGTRACKER_AM/AP/CR columns to identify the defeatured module.
    Fallback path (when binDefs + modMap are supplied and LOGTRACKER yields nothing):
      look up each IB3/IB4 die's DATA_BIN in binDefs, run getModuleFromBinDesc on the
      bin description string, and use the resulting 'dd' category as the module key.
      This covers products where LOGTRACKER columns are absent or blank (e.g. NVL816-BLLC).
    """
    if waferLvl:
        raise UnboundLocalError("Wafer level not supported in updateDefeatureModCnts!")
    # print(len(yieldDF.loc[yieldDF[getIBinCol(yieldDF)].isin([3,4])]))
    dfDF = yieldDF.loc[yieldDF[getIBinCol(yieldDF)].isin(productInfo["dfBins"])]
    # print(len(dfDF))
    dfCnt = 0
    logtrackerRgx = [f"LOGTRACKER_AM[0-{productInfo['numAtoms']/4-1}]",
                     f"LOGTRACKER_AP[0-{productInfo['numAtoms']/4-1}]",
                     f"LOGTRACKER_CR[0-{productInfo['numCores']-1}]"]
    # print("*****************************************************")
    for i, row in dfDF.iterrows():
        # print(row)
        foundFlag = False
        for col,val in row.items():
            if foundFlag:
                break
            for rgx in logtrackerRgx:
                m = re.search(rgx, col)
                if m:
                    try:
                        decoded = prime_error_decode(val[10:].strip("=")).split("\n")
                    except TypeError:
                        continue
                    # print(decoded)
                    if len(decoded) > 1:
                        # print(decoded[1].split("|"))
                        modMatch = re.search(r"(\w{3}_\w{4,5}).*::.*", decoded[1].split("|")[-1])
                        if modMatch:
                            dfMod = modMatch.group(1).replace("MBIST", "CORE")
                            # print(dfMod)
                            if dfMod in dfModCnts:
                                dfModCnts[dfMod] += 1
                            else:
                                dfModCnts[dfMod] = 1
                            dfCnt += 1
                            foundFlag = True
                            break
        if not foundFlag and binDefs is None:
            # Only log per-row warnings when there is no binDefs fallback available.
            logging.warning(f"{' ':{idt}}Warning: Could not find defeatured module for row:\n")
            print(row)
    # ── Fallback: if LOGTRACKER decoding found nothing, infer module from bin description ──
    # This handles products (e.g. NVL816-BLLC) where LOGTRACKER columns are absent or blank.
    # The bin description string (e.g. "B20260000_FAIL_ARR_ATOM_LSA_...") already encodes
    # the failure domain and matches the same moduleMap regex patterns used by getYieldByModule.
    if dfCnt == 0 and len(dfDF) > 0 and binDefs is not None and modMap is not None:
        binCol = getBinCol(df=dfDF)
        if binCol:
            for dbin, n in dfDF[binCol].value_counts().items():
                try:
                    dbin_f = float(dbin)
                except (ValueError, TypeError):
                    continue
                roundDbin = round(dbin_f / 10000)
                binKey = ("FB" + str(roundDbin)) if roundDbin < 1000 else ("DB" + str(round(dbin_f)))
                binDesc = binDefs.get(binKey)
                if not binDesc:
                    logging.warning(f"{' ':{idt}}binDefs fallback: no entry for key '{binKey}' (dbin={dbin}, n={n}) — skipping.")
                    continue
                try:
                    module = getModuleFromBinDesc(modMap=modMap, binDesc=str(binDesc))
                    ddCat = module.get("dd", "")
                    if ddCat:
                        dfModCnts[ddCat] = dfModCnts.get(ddCat, 0) + n
                        dfCnt += n
                except LookupError:
                    logging.warning(f"{' ':{idt}}binDefs fallback: no moduleMap match for '{binDesc}' — skipping.")
                    continue
    if dfCnt != len(dfDF):
        logging.warning(f"{' ':{idt}}Warning: Not all Defeatured die accounted for!  There are "
                        f"{len(dfDF)} Defeatured die, but only found modules for {dfCnt} of them!")



def makeOutXlVmax(outFile="", allModuleYield={}, info={}, moduleMap={}, waferLvl=False, idt=2):
    if not HAVE_OPENPYXL:
        logging.warning(f"{' ':{idt}}openpyxl not installed — writing CSV summary instead.")
        csv_out = f"{pathlib.Path(outFile).stem}_vmax.csv"
        try:
            with open(csv_out, "w", newline="") as fh:
                writer = csv.writer(fh)
                writer.writerow(["LotWafer", "Module", "Bucket", "Count", "Percent"])
                # allModuleYield may be a dict of lotWafer -> moduleYield
                if isinstance(allModuleYield, dict):
                    for lotWafer, moduleYield in allModuleYield.items():
                        for moduleType, buckets in moduleYield.items():
                            if isinstance(buckets, dict):
                                for bucketName, cnt in buckets.items():
                                    try:
                                        pct = cnt / info.get("numDie", 1)
                                    except Exception:
                                        pct = ""
                                    writer.writerow([lotWafer, moduleType, bucketName, cnt, pct])
                else:
                    writer.writerow(["all", "vmax", str(allModuleYield), "", ""])
            logging.info(f"{' ':{idt}}CSV written to '{csv_out}'.")
        except Exception as e:
            logging.error(f"{' ':{idt}}Failed to write CSV summary: {e}")
        return
    if waferLvl:
        outFile = f"{pathlib.Path(outFile).stem}_wafer.xlsx"
    logging.info(f"{' ':{idt}}Creating or overwriting '{outFile}'.")
    outWB = openpyxl.Workbook()
    outWS = outWB[outWB.sheetnames[0]]
    outWS.title = info["product"]
    outWS.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 36
    if not waferLvl:
        allModuleYield = {"1":allModuleYield}
    for iterCnt, (lotWafer, moduleYield) in enumerate(allModuleYield.items()):
        yieldBuckets = []
        #Headers
        if iterCnt == 0:
            writeCell(ws=outWS, row=1, col=iterCnt+1, value="Sub Module", wrapText=True, bold=True)
        if waferLvl:
            s = f"{lotWafer.replace('_', ' ')} {info['TP']} Yield Loss (Fail Bins) (%)"
        else:
            s = f"{info['numWafers']}W {info['TP']} Yield Loss (Fail Bins) (%)"
        writeCell(ws=outWS, row=1, col=iterCnt+2, value=s, wrapText=True, bold=True)
        if iterCnt == 0:
            writeCell(ws=outWS, row=2, col=iterCnt+1, value="# Die", bold = True)
        writeCell(ws=outWS, row=2, col=iterCnt+2, value= info["numDie"], bold = True)

        row = 3
        sums = []
        pcntTotal = 0

        #Good Bins
        for bucket in ["Bin 1 (No Repair)", "Bin 198 (Vmin Repair)", "Bin 2 (Hard Repair)",
                       "Bin 202 (Vmax Repair)", "Bin 3", "Bin 4"]:
            if iterCnt == 0:
                writeCell(ws=outWS, row=row, col=iterCnt+1, value=bucket)
            n = getYieldFromModYield(moduleYield, "vmax", bucket)
            yieldBuckets.append(bucket)
            writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"={n}/B2", numFmt="0.0%")
            pcntTotal += (n / info["numDie"])
            row += 1
        if iterCnt == 0:
            writeCell(ws=outWS, row=row, col=iterCnt+1, value="SUM", bold=True)
        colLetter = openpyxl.utils.get_column_letter(iterCnt+2)
        writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"=SUM({colLetter}{row-6}:{colLetter}{row-1})",
                  numFmt="0.0%", bold=True)
        sums.append(row)
        row += 1

        #ARR, FUN, SCN and sub modules
        buckets = ["ARR_ATOM_VNOM_LFM", "ARR_ATOM_VNOM_TFM", "ARR_ATOM_VMAX_LFM",
                   "ARR_ATOM_VMAX_TFM", "FUN_ATOM_VNOM_LFM", "FUN_ATOM_VNOM_TFM",
                   "SCN_ATOM_VNOM_LFM", "SCN_ATOM_VNOM_TFM"]
        rslt = writeBucketsToXl(outWS=outWS, moduleYield=moduleYield, moduleMap=moduleMap,
            row=row, which="vmax", buckets=buckets, iterCnt=iterCnt, info=info)
        row = rslt["row"]
        pcntTotal += rslt["pcntTotal"]
        yieldBuckets += rslt["yieldBuckets"]
        sums.append(row)
        row += 1

        buckets = ["ARR_CCF_VNOM_LFM", "ARR_CCF_VNOM_TFM", "ARR_CCF_VMAX_LFM",
                   "ARR_CCF_VMAX_TFM", "FUN_CCF_VNOM_LFM", "FUN_CCF_VNOM_TFM",
                   "FUN_ATOM_DRAGON_SLC_VNOM_LFM", "FUN_ATOM_DRAGON_SLC_VNOM_TFM",
                   "FUN_ATOM_DRAGON_SLC_SPECKLE",
                   "SCN_CCF_VNOM_LFM", "SCN_CCF_VNOM_TFM", "ARR_NONCCF_VNOM_LFM",
                   "ARR_NONCCF_VMAX_LFM", "SCN_NONCCF_VNOM_LFM"]
        rslt = writeBucketsToXl(outWS=outWS, moduleYield=moduleYield, moduleMap=moduleMap,
            row=row, which="vmax", buckets=buckets, iterCnt=iterCnt, info=info)
        row = rslt["row"]
        pcntTotal += rslt["pcntTotal"]
        yieldBuckets += rslt["yieldBuckets"]
        sums.append(row)
        row += 1

        buckets = ["ARR_CORE_VNOM_LFM", "ARR_CORE_VNOM_TFM", "ARR_CORE_VMAX_LFM",
                   "ARR_CORE_VMAX_TFM", "FUN_CORE_VNOM_LFM", "FUN_CORE_VNOM_TFM",
                   "FUN_CORE_VMAX_LFM",
                   "SCN_CORE_VNOM_LFM", "SCN_CORE_VNOM_TFM"]
        rslt = writeBucketsToXl(outWS=outWS, moduleYield=moduleYield, moduleMap=moduleMap,
            row=row, which="vmax", buckets=buckets, iterCnt=iterCnt, info=info)
        row = rslt["row"]
        pcntTotal += rslt["pcntTotal"]
        yieldBuckets += rslt["yieldBuckets"]
        sums.append(row)
        row += 1

        #Reset
        if iterCnt == 0:
            writeCell(ws=outWS, row=row, col=iterCnt+1, value="RESET (19,35)")
        n = getYieldFromModYield(moduleYield, "vmax", "Reset")
        writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"={n}/B2", numFmt="0.0%")
        pcntTotal += (n / info["numDie"])
        yieldBuckets.append("Reset")
        sums.append(row)
        row += 1

        buckets = [ "Analog PTH BG", "Analog PTH DLVR", "Analog PTH DTS", "Analog PTH ODI",
                    "Analog CLK", "Analog MIO D2D", "Analog Other"]
        rslt = writeBucketsToXl(outWS=outWS, moduleYield=moduleYield, moduleMap=moduleMap,
                                row=row, which="vmax", buckets=buckets, iterCnt=iterCnt, info=info)
        row = rslt["row"]
        pcntTotal += rslt["pcntTotal"]
        yieldBuckets += rslt["yieldBuckets"]
        sums.append(row)
        row += 1

        #Other
        for other in ["TPI Foundry", "TPI Foundry ADTL (B43)", "TPI Foundry SICC (B18)", "TPI Other", "HVQK (B26)"]:
            if iterCnt == 0:
                writeCell(ws=outWS, row=row, col=iterCnt+1, value=f"{other}")
            n = getYieldFromModYield(moduleYield, "vmax", f"{other}")
            writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"={n}/B2", numFmt="0.0%")
            pcntTotal += (n / info["numDie"])
            yieldBuckets.append(other)
            sums.append(row)
            row += 1
        #Yield Sum
        if iterCnt == 0:
            writeCell(ws=outWS, row=row, col=iterCnt+1, value="Yield SUM (%)", bold=True)
        s = f"=SUM({colLetter}"
        s += f",{colLetter}".join(str(x) for x in sums)
        s += ")"
        writeCell(ws=outWS, row=row, col=iterCnt+2, value=s, bold=True, numFmt="0.0%")

        if round(pcntTotal*100,2) != 100.0:
            print("*************************************************")
            logging.warning(f"{' ':{idt}}WARNING!  Total yield percentage is "
                            f"{round(pcntTotal * 100,2)}%, not 100%!")
            checkBucketsUsed(moduleYield["vmax"], yieldBuckets)
        else:
            logging.info(f"{' ':{idt}}Total yield percentage is {round(pcntTotal * 100,1)}%.")

    outWB.save(filename=outFile)
    logging.info(f"{' ':{idt}}Excel written to '{outFile}'.")



def checkBucketsUsed(moduleYield={}, yieldBuckets=[], idt=6):
    logging.info(f"{' ':{idt-2}}Checking for any yield buckets that were not used...")
    # print(f"moduleYield : {moduleYield}")
    # print(f"yieldBuckets : {yieldBuckets}")
    flag = False
    for k,v in moduleYield.items():
        if k not in yieldBuckets:
            flag = True
            logging.info(f"{' ':{idt}}Bucket '{k}' with {v} die not counted!")
    if not flag:
        logging.info(f"{' ':{idt}}No unused buckets found.")


def makeOutXl(outFile="", moduleYield={}, which="dd", info={}, defeatureModCnts={}, idt=2):
    if not HAVE_OPENPYXL:
        logging.warning(f"{' ':{idt}}openpyxl not installed — writing CSV summary instead.")
        csv_out = f"{pathlib.Path(outFile).stem}.csv"
        try:
            with open(csv_out, "w", newline="") as fh:
                writer = csv.writer(fh)
                writer.writerow(["ModuleType", "Bucket", "Count", "Percent"])
                my = moduleYield.get(which, {}) if isinstance(moduleYield, dict) else {}
                for bucketName, cnt in (my.items() if isinstance(my, dict) else []):
                    try:
                        pct = cnt / info.get("numDie", 1)
                    except Exception:
                        pct = ""
                    writer.writerow([which, bucketName, cnt, pct])
            logging.info(f"{' ':{idt}}CSV written to '{csv_out}'.")
        except Exception as e:
            logging.error(f"{' ':{idt}}Failed to write CSV summary: {e}")
        return
    logging.info(f"{' ':{idt}}Creating or overwriting '{outFile}'.")
    outWB = openpyxl.Workbook()
    outWS = outWB[outWB.sheetnames[0]]
    outWS.title = info["product"]
    outWS.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 14
    #Headers
    writeCell(ws=outWS, row=1, col=1, value="Sub Module", wrapText=True, bold=True)
    s = f"{info['numWafers']}W {info['TP']} Yield Loss (Fail Bins) (%)"
    writeCell(ws=outWS, row=1, col=2, value=s, wrapText=True, bold=True)
    s = f"{info['numWafers']}W {info['TP']} Recovery Bins (3-4) (%)"
    writeCell(ws=outWS, row=1, col=3, value= s, wrapText=True, bold=True)
    writeCell(ws=outWS, row=2, col=1, value="# Die", bold = True)
    writeCell(ws=outWS, row=2, col=2, value= info["numDie"], bold = True)

    pcntTotal = 0
    row = 3
    yieldBuckets = []
    numDie = info["numDie"] or 1  # guard against div-by-zero

    # Track per-group sums (col2, col3) for cross-group SUM rows
    _grp_sums2 = []   # col2 sum per ARR/FUN/SCN group + Reset
    _grp_sums3 = []   # col3 sum per ARR/FUN/SCN group (defeature)
    _bins_sum2 = 0.0
    _other_sums2 = []

    #ARR, FUN, SCN and sub modules
    for module in ["ARR", "FUN", "SCN"]:
        grp2 = 0.0
        grp3 = 0.0
        for subMod in ["ATOM", "CCF", "CORE", "NONCCF"]:
            writeCell(ws=outWS, row=row, col=1, value=f"{module}_{subMod}")
            modKey = module + "_" + subMod
            yieldBuckets.append(modKey)
            n = getYieldFromModYield(moduleYield, "dd", modKey)
            pct2 = n / numDie
            writeCell(ws=outWS, row=row, col=2, value=pct2, numFmt="0.0%")
            grp2 += pct2
            pcntTotal += pct2
            if "ATOM" in modKey or "CORE" in modKey:
                df_n = defeatureModCnts.get(modKey, 0)
                pct3 = df_n / numDie
                writeCell(ws=outWS, row=row, col=3, value=pct3, numFmt="0.0%")
                grp3 += pct3
            row += 1
        writeCell(ws=outWS, row=row, col=1, value="SUM", bold=True)
        writeCell(ws=outWS, row=row, col=2, value=grp2, numFmt="0.0%", bold=True)
        writeCell(ws=outWS, row=row, col=3, value=grp3, numFmt="0.0%", bold=True)
        _grp_sums2.append(grp2)
        _grp_sums3.append(grp3)
        row += 1
    #Reset
    writeCell(ws=outWS, row=row, col=1, value="RESET (19,35)")
    n = getYieldFromModYield(moduleYield, "dd", "Reset")
    reset_pct = n / numDie
    writeCell(ws=outWS, row=row, col=2, value=reset_pct, numFmt="0.0%")
    yieldBuckets.append("Reset")
    pcntTotal += reset_pct
    _grp_sums2.append(reset_pct)
    row += 1
    #Good Bins
    for i in range(1,5):
        writeCell(ws=outWS, row=row, col=1, value=f"Bin {i}")
        if i == 1:
            n = getYieldFromModYield(moduleYield, "dd", f"Bin {i}") + \
                getYieldFromModYield(moduleYield, "dd", f"Bin 198 (Vmin Repair)")
            yieldBuckets.append(f"Bin {i}")
            yieldBuckets.append(f"Bin 198 (Vmin Repair)")
        elif i == 2:
            n = getYieldFromModYield(moduleYield, "dd", f"Bin 2 (Hard Repair)") + \
                getYieldFromModYield(moduleYield, "dd", f"Bin 202 (Vmax Repair)")
            yieldBuckets.append(f"Bin 2 (Hard Repair)")
            yieldBuckets.append(f"Bin 202 (Vmax Repair)")
        else:
            n = getYieldFromModYield(moduleYield, "dd", f"Bin {i}")
            yieldBuckets.append(f"Bin {i}")
        bin_pct = n / numDie
        writeCell(ws=outWS, row=row, col=2, value=bin_pct, numFmt="0.0%")
        _bins_sum2 += bin_pct
        pcntTotal += bin_pct
        row += 1
    writeCell(ws=outWS, row=row, col=1, value="SUM", bold=True)
    writeCell(ws=outWS, row=row, col=2, value=_bins_sum2, numFmt="0.0%", bold=True)
    writeCell(ws=outWS, row=row, col=3, value=sum(_grp_sums3), numFmt="0.0%", bold=True)
    _grp_sums2.append(_bins_sum2)
    row += 1

    #Repairs! #todo
    repairs_sum2 = 0.0
    for repairRow in ["Bin 198 (Vmin Repair)", "Bin 202 (Vmax Repair)", "Bin 2 (Hard Repair)"]:
        writeCell(ws=outWS, row=row, col=1, value=f"Repair {repairRow}")
        n = getYieldFromModYield(moduleYield, "dd", f"{repairRow}")
        rpr_pct = n / numDie
        writeCell(ws=outWS, row=row, col=2, value=rpr_pct, numFmt="0.0%")
        repairs_sum2 += rpr_pct
        row += 1
    writeCell(ws=outWS, row=row, col=1, value="SUM", bold=True)
    writeCell(ws=outWS, row=row, col=2, value=repairs_sum2, numFmt="0.0%", bold=True)
    row += 1
    #Other
    for other in ["Analog", "TPI Foundry", "TPI Other", "TPI Other - B93", "TPI Other - B98",
              "TPI Other - B99", "HVQK (B26)"]:
        writeCell(ws=outWS, row=row, col=1, value=f"{other}")
        n = getYieldFromModYield(moduleYield, "dd", f"{other}")
        oth_pct = n / numDie
        writeCell(ws=outWS, row=row, col=2, value=oth_pct, numFmt="0.0%")
        pcntTotal += oth_pct
        yieldBuckets.append(other)
        _other_sums2.append(oth_pct)
        row += 1
    #Yield Sum
    writeCell(ws=outWS, row=row, col=1, value="Yield SUM (%)", bold=True)
    yield_sum = sum(_grp_sums2) + sum(_other_sums2)
    writeCell(ws=outWS, row=row, col=2, value=yield_sum, bold=True, numFmt="0.0%")

    if round(pcntTotal*100,2) != 100.0:
        logging.warning(f"{' ':{idt}}WARNING!  Total yield percentage is "
                        f"{round(pcntTotal * 100,2)}%, not 100%!")
        checkBucketsUsed(moduleYield["dd"], yieldBuckets)
    else:
        logging.info(f"{' ':{idt}}Total yield percntage is {round(pcntTotal * 100,1)}%.")

    outWB.save(filename=outFile)
    logging.info(f"{' ':{idt}}Excel written to '{outFile}'.")


def writeBucketsToXl(outWS="", moduleYield={}, moduleMap={}, row=0, which="vmax",
                     buckets=[], iterCnt=0, info={}):
    pcntTotal = 0
    yieldBuckets = []
    for bucket in buckets:
        modKey = getModKey(moduleYield=moduleYield, which=which,
                           pattern=rf"{bucket} \(", moduleMap=moduleMap)
        if iterCnt == 0:
            writeCell(ws=outWS, row=row, col=1, value=modKey)
        n = getYieldFromModYield(moduleYield, which, modKey)
        writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"={n}/B2", numFmt="0.0%")
        pcntTotal += (n / info["numDie"])
        yieldBuckets.append(modKey)
        row += 1
    if iterCnt == 0:
        writeCell(ws=outWS, row=row, col=1, value="SUM", bold=True)
    colLetter = openpyxl.utils.get_column_letter(iterCnt + 2)
    writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"=SUM({colLetter}{row-len(buckets)}:{colLetter}{row-1})",
              numFmt="0.0%", bold=True)
    return {"row": row, "pcntTotal":pcntTotal, "yieldBuckets":yieldBuckets}
    #end writeBucketsToXl


def getModKey(moduleYield={}, which="dd", pattern="", moduleMap={}):
    for bucket, n in moduleYield[which].items():
        m = re.search(pattern, bucket)
        if m:
            return bucket
    #no yield data in dataset for this bucket.  Search the main one.
    for expression, maps in moduleMap.items():
        m = re.search(pattern, maps[which])
        if m:
            return maps[which]
    return pattern


def getYieldFromModYield(moduleYield={}, which="dd", modKey=""):
    try:
        return moduleYield[which][modKey]
    except KeyError:
        return 0


def getYieldByModule(yldDF=pandas.DataFrame(), binDefs=pandas.DataFrame(), modMap={}, idt=4):
    if idt < 2:
        idt = 2
    moduleYield = {}  #{"dd":{"ARR_ATOM":432,"ARR_CORE":444}}
    binTracker = {}

    logging.info(f"{' ':{idt-2}}Calculating yield by sub module.")
    binCol = getBinCol(df=yldDF)
    if not binCol:
        raise LookupError("Unable to find column with the bin number for each die, quitting!")
    else:
        logging.info(f"{' ':{idt}}Using column '{binCol}' as the Bin result.")
    # logging.info(f"DBin Summary:\n{yldDF[binCol].value_counts().nlargest(5)}")

    i = 0
    # dfDF = yieldDF.loc[yieldDF[getIBinCol(yieldDF)].isin(productInfo["dfBins"])]
    #DB do not have Bin93 for some reason.
    numB93 = len(yldDF.loc[yldDF[getIBinCol(yldDF)] == 93])
    moduleYield = updateModYield(moduleYield=moduleYield,
                                 modulesToAdd=getModuleFromBinDesc(modMap=modMap, binDesc="B93"),
                                 dieCnt=numB93)
    totalDie = numB93
    for dbin, n in yldDF[binCol].value_counts().items():
        totalDie += n
        roundDbin = round(dbin / 10000)
        if roundDbin < 1000:
            binKey = "FB" + str(roundDbin)
        else:
            binKey = "DB" + str(round(dbin))
        try:
            binDesc = binDefs[binKey]
        except KeyError:
            logging.warning(f"{'  '}Unknown bin key '{binKey}' (raw dbin={dbin}, n={n}) — not in bin_defs, skipping.")
            continue
        try:
            module = getModuleFromBinDesc(modMap=modMap, binDesc=binDesc)
        except LookupError as e:
            s = e.__str__()
            s = s.replace("',", f", ({n} die),")
            logging.warning(s + " — skipping bin.")
            continue
        # logging.info(f"{' ':{idt}}Found module '{module}' from '{binDesc}', adding {n} to these buckets.")
        try:
            if binDesc not in binTracker[module["vmax"]]:
                binTracker[module["vmax"]].append(binDesc)
        except KeyError:
            binTracker[module["vmax"]] = [binDesc]

        # if "ARR_CORE_VMAX_TFM (B60)" in module.values():
        #     logging.info(f"{' ':{idt}}Found module '{module}' from '{binDesc}', adding {n} to these buckets.")
        moduleYield = updateModYield(moduleYield=moduleYield, modulesToAdd=module, dieCnt=n)
        # print(moduleYield)
        i+=1
    if totalDie != len(yldDF):
        logging.warning(f"{' ':{idt}}WARNING! Found {totalDie} die, but there should be {len(yldDF)}!")
    # logging.info(f"{' ':{idt}}Buckets by bin:")
    # for k,v in binTracker.items():
    #     print(k)
    #     v.sort()
    #     for b in v:
    #         print(f" {b}")
    return moduleYield




def updateModYield(moduleYield={}, modulesToAdd={}, dieCnt=0):
    for k,v in modulesToAdd.items():
        if k not in moduleYield:
            moduleYield[k] = {}
        if v not in moduleYield[k]:
            moduleYield[k][v] = dieCnt
        else:
            moduleYield[k][v] = moduleYield[k][v] + dieCnt
    return moduleYield


def getModuleFromBinDesc(modMap={}, binDesc="", idt=4):
    #start with B26 because a lot of the bin names are duplicated with just B26 in front.
    m = re.search(r"B26\d", binDesc)
    if m:
        return {"dd": "HVQK (B26)", "vmax": "HVQK (B26)"}
    for rgxPattern, modDict in modMap.items():
        m = re.search(rgxPattern, binDesc)
        if m:
            # if "B60" in binDesc: # or "B20" in binDesc:
            #     print(f"{binDesc}: {rgxPattern}")
            return modDict
    raise LookupError(f"No module map for '{binDesc}', quitting.")


def _buildBinDefsFromDF(df):
    """Build a binDefs dict (matching bindef CSV format) directly from the data CSV.

    For fail bins (roundDbin >= 1000): uses the 'Bin Description_' column value,
    which already contains the full test description string that moduleMap regexes match.
    For pass bins (roundDbin < 1000): synthesizes 'B{fb}_PASS' from the FB number,
    matching the moduleMap regex patterns (e.g. 'B198_PASS', 'B201_PASS').
    """
    db_col = getBinCol(df=df)
    if not db_col:
        return {}
    bd_col = next((c for c in df.columns if c.startswith('Bin Description')), None)
    binDefs = {}
    for dbin in df[db_col].dropna().unique():
        try:
            dbin_f = float(dbin)
        except (ValueError, TypeError):
            continue
        round_dbin = round(dbin_f / 10000)
        if round_dbin < 1000:
            bin_key = "FB" + str(round_dbin)
            if bin_key not in binDefs:
                binDefs[bin_key] = f"B{round_dbin}_PASS"
        else:
            bin_key = "DB" + str(round(dbin_f))
            if bin_key not in binDefs and bd_col is not None:
                mask = df[db_col] == dbin
                desc_vals = df.loc[mask, bd_col].dropna()
                if not desc_vals.empty:
                    binDefs[bin_key] = str(desc_vals.iloc[0])
    return binDefs


def getBinCol(df=pandas.DataFrame(), productInfo={}, idt=2):
    binCols = {}
    if "DB" in df:
        return "DB"
    else:
        for col in df:
            if "DB@" in col or "DB DIEBIN" in col or "DATA_BIN" in col or "DATA_BIN_132110" in col:
                return col
    return False
    # if "FB" in df:
    #     binCols["fbin"] = "FB"
    # else:
    #     for col in df:
    #         if "FB@" in col:
    #             binCols["fbin"] = col
    # if "DB" in df:
    #     binCols["dbin"] = "DB"
    # else:
    #     for col in df:
    #         if "DB@" in col:
    #             binCols["dbin"] = col
    # if "fbin" in binCols and "dbin" in binCols:
    #     return binCols
    # else:
    #     return False


def getIBinCol(df=pandas.DataFrame(), productInfo={}, idt=2):
    binCols = {}
    if "IB" in df:
        return "DB"
    else:
        for col in df:
            if "IB@" in col or "IB DIEBIN" in col or ("INTERFACE_BIN" in col and "TOTAL" not in col):
                return col
    return False


def getLotCol(df=pandas.DataFrame(), idt=2):
   for col in df:
       if col == "LOT":
           return col
       if "Lot_132110" in col or "SORT_LOT" in col:
           return col


def getWaferCol(df=pandas.DataFrame(), idt=2):
    for col in df:
        if col == "WAFER":
            return col
        if "SORT_WAFER" in col:
            return col


def getYieldDataFrame(inFile="", productInfo={}, idt=4):
    if not inFile:
        raise FileNotFoundError(f"No input file specified: '{inFile}'.")
    df = pandas.read_csv(inFile, header=0, low_memory=False)
    product = getPart(df=df, productInfo=productInfo)
    if product:
        logging.info(f"{' ':{idt}}Opened file '{inFile}' and found product: {product}.")
    else:
        logging.warning(f"{' ':{idt}}Unable to find product, defaulting to ARLS816.")
        product = "ARLS816"
    return product, df


def getTPNum(df=pandas.DataFrame(), rgx="", idt=2):
    tpCol = ""
    for column in df:
        if "PROGRAM" in column.upper():
            tpCol = column
            break
    # Use first non-null value to avoid TypeError when first row has NaN
    series = df[tpCol].dropna()
    first_val = str(series.iloc[0]) if len(series) > 0 else ""
    m = re.search(rgx, first_val)
    if m:
        logging.info(f"{' ':{idt + 2}}Found TP line '{m.group(1)}' from {first_val}.")
        return m.group(1)
    else:
        raise LookupError(f"Unable to find a match for '{rgx}' in '{first_val}'.")


def getPart(df=pandas.DataFrame(), productInfo={}, idt=2):
    def _first_val(col):
        """Return first non-null string value from a column, or empty string."""
        series = df[col].dropna()
        return str(series.iloc[0]) if len(series) > 0 else ""
    if "Part" in df:
        return getProductFromDevRevStep(_first_val("Part"), productInfo=productInfo)
    for column in df:
        if "Part@" in column:
            return getProductFromDevRevStep(_first_val(column), productInfo=productInfo)
        if "DevRevStep" in column:
            return getProductFromDevRevStep(_first_val(column), productInfo=productInfo)
    for column in df:
        if "PROGRAM" in column:
            return getProductFromProgram(_first_val(column))

def getProductFromDevRevStep(drs="", productInfo={}):
    for product, prodInfo in productInfo.items():
        if drs in prodInfo["DEVREVSTEP"]:
            return product
    return False





###########################################
#General functions that are useful to other scripts
###########################################
IntToCharMapping = "ABCDEFGHIJKLMNOPQRSTUVWXYZ234567"
CharToIntMapping = {c: i for i, c in enumerate(IntToCharMapping)}

def prime_error_encode(text):
    bytes_str = zlib.compress(text.encode("utf-8"))
    bytes_str = bytes_str[2:-4] # discard header and tail
    bits_str = "".join([bin(b).replace("0b", "").rjust(8, "0") for b in bytes_str])
    bits_str += "0" * (5 - len(bits_str) % 5)
    encoded_str = "".join([IntToCharMapping[int(bits_str[i:i+5], 2)] for i in range(0, len(bits_str), 5)])
    return encoded_str

def prime_error_decode(encoded_str):
    if len(encoded_str) == 0:
        return ""
    bits_str = "".join([bin(CharToIntMapping[c]).replace("0b", "").rjust(5, "0") for c in encoded_str])
    bits_str += "0" * (8 - len(bits_str) % 8)
    bytes_arr = [int(bits_str[i:i+8], 2) for i in range(0, len(bits_str), 8)]
    bytes_str = b"".join([b.to_bytes(1, "big") for b in bytes_arr]) # "litte" works too
    text = zlib.decompress(bytes_str, -8).decode("utf-8")
    return text


def listOfTuplesToDict(listOfTuples=[()]):
    resultDict = {}
    #first item of each tuple is Lot #, second item is Wafer #
    for t in listOfTuples:
        try:
            resultDict[t[0]].append(t[1])
        except KeyError:
            resultDict[t[0]] = [t[1]]
    return resultDict


def writeCell(ws="", row=0, col=0, value="", numFmt="", wrapText=False, bold=False, idt=4):
    ws.cell(row=row, column=col).value = value
    if numFmt:
        ws.cell(row=row, column=col).number_format = numFmt
    if wrapText:
        ws.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True)
    if bold:
        ws.cell(row=row, column=col).font = openpyxl.styles.Font(bold=True)


def append_to_dashboard(dashboardPath="", srcWorkbook="", srcSheet="", tgtSheet=""):
    from openpyxl import load_workbook
    from pathlib import Path
    import shutil as _shutil

    dbPath = Path(dashboardPath)
    srcPath = Path(srcWorkbook)
    if not srcPath.exists():
        raise FileNotFoundError(f"Source workbook not found: {srcPath}")

    # If the dashboard file does not exist yet, bootstrap it by copying the _out.xlsx.
    # The copy becomes the new dashboard so the user can build on it over time.
    if not dbPath.exists():
        dbPath.parent.mkdir(parents=True, exist_ok=True)
        _shutil.copy2(str(srcPath), str(dbPath))
        logging.info(f"Dashboard file not found; created from source: {dbPath}")
        return

    # load source with formulas so we copy formulas and not only evaluated values
    src_wb = load_workbook(srcPath, data_only=False)
    if srcSheet not in src_wb.sheetnames:
        # fall back to the first sheet if named sheet not found
        srcSheet = src_wb.sheetnames[0]
    src_ws = src_wb[srcSheet]

    # open dashboard and append as new right-most set of columns (one per source column)
    tgt_wb = load_workbook(dbPath)
    # use the first sheet as target if tgtSheet doesn't exist
    if tgtSheet and tgtSheet in tgt_wb.sheetnames:
        tgt_ws = tgt_wb[tgtSheet]
    else:
        tgt_ws = tgt_wb[tgt_wb.sheetnames[0]]

    # append columns B and C (source columns 2 and 3) as the next right-most columns
    next_col = tgt_ws.max_column + 1
    max_row = src_ws.max_row
    # only copy columns 2 and 3 if they exist in the source
    cols_to_copy = [2, 3]
    actual_cols = [c for c in cols_to_copy if c <= src_ws.max_column]
    if not actual_cols:
        raise LookupError(f"Source sheet '{srcSheet}' does not contain columns B/C to copy.")

    # If the dashboard already contains columns with the same headers, fix formulas
    # in-place so they reference their own column (avoid stale formulas referencing B2/C2)
    from openpyxl.utils import get_column_letter
    src_headers = [src_ws.cell(row=1, column=c).value or f"{src_ws.title} {get_column_letter(c)}" for c in actual_cols]
    if src_headers:
        for tgt_c in range(1, tgt_ws.max_column + 1):
            try:
                hdr = tgt_ws.cell(row=1, column=tgt_c).value
            except Exception:
                hdr = None
            if hdr in src_headers:
                # rewrite formulas in this existing column to reference the column itself
                tgt_letter = get_column_letter(tgt_c)
                for r in range(2, tgt_ws.max_row + 1):
                    cell = tgt_ws.cell(row=r, column=tgt_c)
                    val = cell.value
                    if isinstance(val, str) and val.startswith('='):
                        # replace any occurrences of source column letters (B/C) with tgt_letter
                        for sidx in actual_cols:
                            sletter = get_column_letter(sidx)
                            val = val.replace(f"{sletter}2", f"{tgt_letter}2")
                            # replace ranges like B3:B6 -> K3:K6
                            val = val.replace(f"{sletter}", tgt_letter)
                        cell.value = val

    from openpyxl.styles import Font
    for idx, c in enumerate(actual_cols):
        tgt_col = next_col + idx
        # copy header explicitly (use source header if present, else fallback)
        src_header = src_ws.cell(row=1, column=c).value
        if not src_header:
            # fallback to sheet title + column letter
            src_header = f"{src_ws.title} {openpyxl.utils.get_column_letter(c)}"
        tgt_header = tgt_ws.cell(row=1, column=tgt_col)
        tgt_header.value = src_header
        tgt_header.font = Font(bold=True)
        try:
            tgt_header.alignment = openpyxl.styles.Alignment(wrap_text=True)
        except Exception:
            pass

        from openpyxl.utils import column_index_from_string, get_column_letter
        import re

        def rewrite_formula(formula: str):
            # replace any column-letter references that point to source columns
            def repl(m):
                col_letters = m.group(1)
                row_num = m.group(2)
                try:
                    src_idx = column_index_from_string(col_letters)
                except Exception:
                    return m.group(0)
                if src_idx in actual_cols:
                    tgt_letter = get_column_letter(next_col + actual_cols.index(src_idx))
                    return f"{tgt_letter}{row_num}"
                return m.group(0)
            return re.sub(r'([A-Z]+)(\d+)', repl, formula)

        for r in range(2, max_row + 1):
            src_cell = src_ws.cell(row=r, column=c)
            tgt_cell = tgt_ws.cell(row=r, column=tgt_col)
            val = src_cell.value
            if isinstance(val, str) and val.startswith('='):
                # rewrite formula to point to the appended dashboard columns
                try:
                    tgt_cell.value = rewrite_formula(val)
                except Exception:
                    tgt_cell.value = val
            else:
                tgt_cell.value = val

            # keep row 2 formatting from source (counts), other rows use percent
            try:
                if r == 2:
                    tgt_cell.number_format = src_cell.number_format or 'General'
                else:
                    tgt_cell.number_format = '0.0%'
            except Exception:
                pass
            try:
                tgt_cell.font = src_cell.font
                tgt_cell.alignment = src_cell.alignment
            except Exception:
                pass

    # finished copying selected columns

    tgt_wb.save(dbPath)


def findChars(s="", c=""):
    return [i for i, letter in enumerate(s) if letter == c]


def getDateTime():
    now = datetime.datetime.now().strftime("%Y%m%d-%H%M%S.%f")
    return now[:-3]


def setupLogging(logFile=False, logLevel=logging.INFO):
    handlers = [logging.StreamHandler(sys.stdout)]
    if logFile:
        handlers.append(logging.FileHandler(filename=logFile, mode="w"))
    logging.basicConfig(level=logLevel,
                        format="%(asctime)s.%(msecs)03d: %(message)s",
                        datefmt="%Y%m%d:%H:%M:%S", handlers=handlers)


def main():
    global docopt
    if docopt:
        try:
            args = docopt(__doc__, version=__version__)
        except BaseException:
            # docopt parsing failed (e.g., user passed argparse-style flags); fallback to argparse
            docopt = None
            args = None
    if not docopt:
        import argparse
        parser = argparse.ArgumentParser(description="Get Digital Dashboard Update")
        parser.add_argument("-d", "--data", dest="data", default="", help=".csv Input file containing a list of die and FBINs")
        parser.add_argument("-b", "--bin_defs", dest="bin_defs", default="", help=".csv file containing a table of DBIN and Test name")
        parser.add_argument("-g", "--log", dest="log", default="", help="Optionally, log to a file")
        parser.add_argument("-m", "--vmax", dest="vmax", action="store_true", help="Make the Vmax summary as well")
        parser.add_argument("-w", "--wafer", dest="wafer", action="store_true", help="Print columns per wafer")
        parser.add_argument("--verbose", dest="verbose", action="store_true", help="Print more information than usual")
        parser.add_argument("-x", "--debug", dest="debug", action="store_true", help="Print lots of debugging statements")
        parser.add_argument("-v", "--version", dest="version", action="store_true", help="Show version")
        parser.add_argument("--dashboard", dest="dashboard", default="", help="Path to DigitalDashBoard.xlsx to append results")
        parser.add_argument("--outdir", dest="outdir", default="", help="Output folder for generated xlsx files")
        parsed = parser.parse_args()
        args = {
            "--data": parsed.data,
            "--bin_defs": parsed.bin_defs,
            "--log": parsed.log,
            "--vmax": parsed.vmax,
            "--wafer": parsed.wafer,
            "--verbose": parsed.verbose,
            "--debug": parsed.debug,
            "--version": parsed.version,
            "--dashboard": parsed.dashboard,
            "--outdir": parsed.outdir,
        }

    # choose log level (debug overrides verbose)
    if args.get("--debug"):
        ll = logging.DEBUG
    elif args.get("--verbose"):
        ll = logging.INFO
    else:
        ll = logging.WARNING

    # determine where to place log files
    if args.get("--log"):
        user_log = args["--log"]
        p = pathlib.Path(user_log)
        suffix = p.suffix if p.suffix else ".log"
        name = p.stem if p.stem else __version__.replace(' ', '_').lower()
        logFileName = f"{name}_{getDateTime()}{suffix}"
        # if an absolute or explicit directory was provided, use it; otherwise use system temp
        if p.parent and str(p.parent) != ".":
            target_dir = p.parent
        elif p.is_absolute():
            target_dir = p.parent
        else:
            target_dir = pathlib.Path(tempfile.gettempdir())
    else:
        logFileName = f"{__version__.replace(' ', '_').lower()}_{getDateTime()}.log"
        target_dir = pathlib.Path(tempfile.gettempdir())
    try:
        pathlib.Path(target_dir).mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    logPath = str(pathlib.Path(target_dir) / logFileName)
    setupLogging(logFile=logPath, logLevel=ll)
    logBase = pathlib.Path(logFileName).stem
    try:
          logging.info(f"Welcome to {__version__}!")
          getDD(dataInFile=args.get("--data"), binDefFile=args.get("--bin_defs"), waferLvl=args.get("--wafer"),
              vmax=args.get("--vmax"), logBaseName=logBase, dashboardFile=args.get("--dashboard"),
              outDir=args.get("--outdir", ""))
    except KeyboardInterrupt:
        print("Ctl+C detected, exiting.")


if __name__ == "__main__":
    main()
