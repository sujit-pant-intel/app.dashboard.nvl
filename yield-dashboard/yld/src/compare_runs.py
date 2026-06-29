#!/usr/bin/env python3
"""
compare_runs.py
---------------
Reads Dashboard.html, resolves each run's *_out.xlsx, and generates a
standalone comparison HTML report.

Usage:
    python compare_runs.py Dashboard.html
    python compare_runs.py Dashboard.html --out my_comparison.html
    python compare_runs.py Dashboard.html --ref "NCXSDJXP0H51M202611-1"
"""

import sys
import os
import re
import argparse
import io
import base64
from pathlib import Path


def _wm_inject(html: str) -> str:
    _wm = (
        '<div id="_wm_div" style="position:fixed;top:8px;right:12px;font-size:10px;'
        'font-weight:600;pointer-events:none;z-index:99999;'
        'font-family:Arial,sans-serif;user-select:none;letter-spacing:0.04em;'
        'padding:2px 6px;border-radius:3px;background:transparent;">'
        'Pant, Sujit N \u2014 GEMS FTE</div>'
        '<script>(function(){'
        'function _wm_color(){'
        'var d=document.getElementById("_wm_div");if(!d)return;'
        'var bg=window.getComputedStyle(document.body).backgroundColor;'
        'var m=bg.match(/\\d+/g);'
        'if(m&&m.length>=3){'
        'var r=+m[0],g=+m[1],b=+m[2];'
        'var lum=0.299*r+0.587*g+0.114*b;'
        'd.style.color=lum<128?"rgba(255,255,255,0.9)":"rgba(20,20,20,0.75)";'
        '}else{d.style.color="rgba(255,255,255,0.9)";}'
        '}'
        'if(document.readyState==="loading")'
        '{document.addEventListener("DOMContentLoaded",_wm_color);}'
        'else{_wm_color();}'
        '})();</script>'
    )
    import re as _re_wm
    if '</body>' not in html:
        return html
    html = _re_wm.sub(
        r'<div[^>]*id=["\']_wm_div["\'][^>]*>[\s\S]*?</div>\s*<script[^>]*>[\s\S]*?</script>',
        '', html)
    html = _re_wm.sub(r'<div[^>]*>[^<]*GEMS FTE[^<]*</div>', '', html)
    return html.replace('</body>', _wm + '\n</body>', 1)

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np
    HAVE_MPL = True
except ImportError:
    HAVE_MPL = False

# ---------------------------------------------------------------------------
# 1. Parse Dashboard.html → run records
# ---------------------------------------------------------------------------

def parse_dashboard(dash_path: Path):
    """Return list of dicts: {stem, name, ts, index_href}.
    Only returns blocks from the Yield section (YIELD_START/END or legacy RUNS_START/END)."""
    content = dash_path.read_text(encoding='utf-8')
    # Extract only the Yield section content
    yield_html = ''
    for start_pat, end_pat in [
        (r'<!--\s*YIELD_START\s*-->', r'<!--\s*YIELD_END\s*-->'),
        (r'<!--\s*RUNS_START\s*-->',  r'<!--\s*RUNS_END\s*-->'),   # legacy
    ]:
        m = re.search(start_pat + r'(.*?)' + end_pat, content, re.S)
        if m:
            yield_html = m.group(1)
            break
    # Fall back to full content only if no sentinels found at all
    search_html = yield_html if yield_html.strip() else content
    runs = []
    # Each block: <div class="run-block" data-stem="...">...</div></div>
    block_re = re.compile(
        r'<div class="run-block" data-stem="([^"]+)">([\s\S]*?)</div>\s*</div>',
        re.MULTILINE
    )
    for m in block_re.finditer(search_html):
        stem = m.group(1)
        body = m.group(2)
        # Name and timestamp from run-header
        hdr = re.search(
            r'<span class="arrow">[^<]*</span>\s*([^<]+)<span class="ts">\s*-\s*([^<]*)</span>',
            body
        )
        name = hdr.group(1).strip() if hdr else stem
        ts   = hdr.group(2).strip() if hdr else ''
        # Yield Report href (report-link)
        link_m = re.search(r'class="run-link report-link"[^>]*href="([^"]+)"', body)
        index_href = link_m.group(1) if link_m else None
        runs.append({'stem': stem, 'name': name, 'ts': ts, 'index_href': index_href})
    return runs


# ---------------------------------------------------------------------------
# 2. Resolve *_out.xlsx from an index.html href
# ---------------------------------------------------------------------------

def find_xlsx(dash_dir: Path, index_href: str):
    """Given Dashboard.html dir and the relative/absolute href to index.html,
    return Path to *_out.xlsx in the same folder, or None."""
    if not index_href:
        return None
    # Strip file:// scheme if present
    href = re.sub(r'^file:///', '', index_href).replace('/', os.sep)
    if os.path.isabs(href):
        idx_path = Path(href)
    else:
        idx_path = dash_dir / href
    out_folder = idx_path.parent
    if not out_folder.exists():
        return None
    candidates = sorted(out_folder.glob('*_out.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


# ---------------------------------------------------------------------------
# 2b. Resolve *_BinDistribution.html and parse RDND + Bin Fail tables
# ---------------------------------------------------------------------------

def find_bin_html(output_dir: Path):
    """Return Path to *_BinDistribution.html in output_dir, or None."""
    candidates = sorted(output_dir.glob('*_BinDistribution.html'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def find_group_medians(output_dir: Path):
    """Return Path to Group_Medians.csv in output_dir, or None."""
    candidates = sorted(output_dir.glob('Group_Medians.csv'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def find_raw_csv(output_dir: Path):
    """Return Path to the raw CSV in output_dir's parent (the source CSV), or None."""
    parent = output_dir.parent
    if not parent.exists():
        return None
    # Match CSVs that look like the source data (not bindef or generated)
    candidates = [p for p in sorted(parent.glob('*.csv'))
                  if 'bindef' not in p.name.lower()
                  and '_targets_' not in p.name.lower()]
    # Prefer the one with 'reticle' in the name (most complete), else newest
    reticle = [p for p in candidates if 'reticle' in p.name.lower()]
    if reticle:
        return reticle[0]
    return candidates[0] if candidates else None


def extract_upm_from_csv(csv_path: Path, config_json: str = None):
    """Extract UPM distribution data from raw CSV using config JSON analyses.
    Returns list of dicts compatible with upm_data format, plus a detailed
    list of per-column medians."""
    import pandas as pd
    import json

    if not csv_path or not csv_path.exists():
        return None, None

    # Load config to get analyses + base
    analyses = []
    base_val = None
    upm_prefix = 'UPM_0107'
    filt_value = upm_prefix
    filt_method = 'starts_with'
    if config_json:
        try:
            cfg = json.loads(Path(config_json).read_text(encoding='utf-8'))
            analyses = cfg.get('analyses', [])
            for anl in analyses:
                if anl.get('type') == 'distribution':
                    filt = anl.get('filter', {}).get('match', {})
                    filt_method = filt.get('method', 'starts_with')
                    filt_value = filt.get('value', upm_prefix)
                    agg = anl.get('aggregation', {})
                    if agg.get('mode') == 'percentage':
                        bc = agg.get('base', {})
                        if bc.get('type') == 'fixed':
                            base_val = bc.get('value')
                    break
        except Exception:
            pass

    try:
        df = pd.read_csv(str(csv_path), dtype=object)
    except Exception:
        return None, None

    import re as _re_upm
    import fnmatch as _fnmatch_upm
    _has_wc = ('*' in filt_value or '?' in filt_value)
    if _has_wc:
        upm_cols = [c for c in df.columns if _fnmatch_upm.fnmatch(c, filt_value)]
    elif filt_method == 'regex':
        upm_cols = [c for c in df.columns if _re_upm.search(filt_value, c)]
    elif filt_method == 'contains':
        upm_cols = [c for c in df.columns if filt_value in c]
    elif filt_method == 'starts_with':
        upm_cols = [c for c in df.columns if c.startswith(filt_value)]
    else:
        upm_cols = [c for c in df.columns if c.startswith(filt_value)]
    if not upm_cols:
        return None, None

    if base_val is None:
        base_val = 9154  # default

    # Per-column detail
    col_details = []
    overall_pct = None
    for col in upm_cols:
        vals = pd.to_numeric(df[col], errors='coerce').dropna()
        if vals.empty:
            continue
        med = float(vals.median())
        pct = (med / base_val) * 100
        short = col[len(upm_prefix):].strip('_') if col.startswith(upm_prefix) else col
        col_details.append({
            'col': col, 'short': short, 'median': med,
            'mean': float(vals.mean()), 'count': len(vals), 'pct': pct
        })
        # Use 0950 column as the "main" UPM % if available
        if '0950' in col:
            overall_pct = pct

    if not col_details:
        return None, None

    # If no 0950 column, use the first
    if overall_pct is None:
        overall_pct = col_details[0]['pct']

    # Build legacy-compatible upm_data (single row with upm_pct)
    upm_data = [{'test': 'UPM', 'n_rows': col_details[0]['count'],
                 'sicc_actual': None, 'sicc_target': None,
                 'multiple': None, 'upm_pct': overall_pct}]

    return upm_data, col_details


def parse_group_medians(csv_path: Path) -> list[dict]:
    """Return list of {test, n_rows, sicc_actual, sicc_target, multiple, upm_pct}."""
    rows = []
    try:
        text = csv_path.read_text(encoding='utf-8-sig')
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            return rows
        header = [h.strip() for h in lines[0].split(',')]
        for line in lines[1:]:
            parts = [p.strip() for p in line.split(',')]
            if len(parts) < 2:
                continue
            def _f(i):
                try: return float(parts[i]) if i < len(parts) and parts[i] else None
                except: return None
            rows.append({
                'test':        parts[0] if parts else '',
                'n_rows':      _f(1),
                'sicc_actual': _f(2),
                'sicc_target': _f(3),
                'multiple':    _f(4),
                'upm_pct':     _f(5),
            })
    except Exception as e:
        print(f'Warning: could not read Group_Medians.csv: {e}')
    return rows


def find_cdyn_medians(output_dir: Path):
    """Return Path to cdyn_medians.csv in output_dir, or None."""
    candidates = sorted(output_dir.glob('cdyn_medians.csv'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def parse_cdyn_medians(csv_path: Path) -> list[dict]:
    """Return list of {test, type, actual, expected, ratio}."""
    rows = []
    try:
        text = csv_path.read_text(encoding='utf-8-sig')
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            return rows
        for line in lines[1:]:
            parts = [p.strip() for p in line.split(',')]
            if len(parts) < 2:
                continue
            def _f(i):
                try: return float(parts[i]) if i < len(parts) and parts[i] else None
                except: return None
            rows.append({
                'test':     parts[0] if parts else '',
                'type':     parts[1] if len(parts) > 1 else '',
                'actual':   _f(2),
                'expected': _f(3),
                'ratio':    _f(4),
            })
    except Exception as e:
        print(f'Warning: could not read cdyn_medians.csv: {e}')
    return rows


def _strip_tags(s: str) -> str:
    return re.sub(r'<[^>]+>', '', s).strip()


def parse_bin_html(bin_html_path: Path):
    """Return dict:
        yield_rows:    [{bin, fail_bucket, yield_pct, expected_pct}, ...]
        bin_fail_rows: [{ibin, fail_bucket, fail_pct, fail_count}, ...]
        func_bin_rows: [{ibin, fbin, fail_bucket, fail_pct, fail_count}, ...]
    """
    content = bin_html_path.read_text(encoding='utf-8')

    # --- RDND yield table ---
    yield_rows = []
    yt_m = re.search(r'<table class="yield-table">(.*?)</table>', content, re.DOTALL)
    if yt_m:
        for tr in re.findall(r'<tr>(.*?)</tr>', yt_m.group(1), re.DOTALL):
            tds = re.findall(r'<td(?:[^>]*)>(.*?)</td>', tr, re.DOTALL)
            if len(tds) < 3:
                continue
            # Column order: BIN(0), FAIL BUCKET(1), YIELD(2), EXPECTED(3)
            bin_name     = _strip_tags(tds[0])
            fail_bucket  = _strip_tags(tds[1]) if len(tds) > 1 else ''
            yield_str    = _strip_tags(tds[2]).rstrip('%') if len(tds) > 2 else ''
            expected_str = _strip_tags(tds[3]).rstrip('%') if len(tds) > 3 else ''
            try:    yield_pct    = float(yield_str)
            except: yield_pct    = None
            try:    expected_pct = float(expected_str)
            except: expected_pct = None
            if bin_name:
                yield_rows.append({'bin': bin_name, 'fail_bucket': fail_bucket,
                                   'yield_pct': yield_pct, 'expected_pct': expected_pct})

    # --- Bin Fail Summary table (pareto-tbl with Interface Bin header) ---
    bin_fail_rows = []
    func_bin_rows = []
    bin_summary_rows = []  # new 6-col format: ibin, cat, desc, total, fail_count, fail_pct
    for tbl in re.findall(r'<table class="pareto-tbl"[^>]*>(.*?)</table>', content, re.DOTALL):
        if 'Interface Bin' not in tbl:
            continue
        has_fbin    = 'Functional Bin' in tbl
        has_cat_desc = 'Category' in tbl and 'Description' in tbl
        for tr in re.findall(r'<tr[^>]*>(.*?)</tr>', tbl, re.DOTALL):
            tds = re.findall(r'<td(?:[^>]*)>(.*?)</td>', tr, re.DOTALL)
            if has_cat_desc:
                # new format: Interface Bin(0), Category(1), Description(2),
                #              Total Count(3), Fail Count(4), Yield/Fail %(5)
                if len(tds) < 6:
                    continue
                ibin         = _strip_tags(tds[0]).strip().rstrip('\u26a0').strip()
                cat          = _strip_tags(tds[1])
                desc         = _strip_tags(tds[2])
                fail_cnt_str = _strip_tags(tds[4]).replace(',', '')
                fail_pct_str = _strip_tags(tds[5]).rstrip('%')
                try:    fail_pct   = float(fail_pct_str)
                except: fail_pct   = None
                try:    fail_count = int(fail_cnt_str)
                except: fail_count = None
                if ibin and fail_pct is not None:
                    bin_summary_rows.append({'ibin': ibin, 'cat': cat, 'desc': desc,
                                             'fail_pct': fail_pct, 'fail_count': fail_count})
            elif has_fbin:
                # cols: Interface Bin(0), Functional Bin(1), Fail Bucket(2),
                #       Description(3), Total Count(4), Fail Count(5), Fail Count %(6)
                if len(tds) < 7:
                    continue
                ibin         = _strip_tags(tds[0])
                fbin         = _strip_tags(tds[1])
                fail_bucket  = _strip_tags(tds[2])
                fail_pct_str = _strip_tags(tds[6]).rstrip('%')
                fail_cnt_str = _strip_tags(tds[5]).replace(',', '')
                try:    fail_pct   = float(fail_pct_str)
                except: fail_pct   = None
                try:    fail_count = int(fail_cnt_str)
                except: fail_count = None
                if ibin and fail_pct is not None:
                    func_bin_rows.append({'ibin': ibin, 'fbin': fbin,
                                          'fail_bucket': fail_bucket,
                                          'fail_pct': fail_pct, 'fail_count': fail_count})
            else:
                # old format: Interface Bin(0), Fail Bucket(1), Total Count(2), Fail Count(3), Fail %(4)
                if len(tds) < 5:
                    continue
                ibin         = _strip_tags(tds[0])
                fail_bucket  = _strip_tags(tds[1])
                fail_pct_str = _strip_tags(tds[4]).rstrip('%')
                fail_cnt_str = _strip_tags(tds[3]).replace(',', '')
                try:    fail_pct   = float(fail_pct_str)
                except: fail_pct   = None
                try:    fail_count = int(fail_cnt_str)
                except: fail_count = None
                if ibin and fail_pct is not None:
                    bin_fail_rows.append({'ibin': ibin, 'fail_bucket': fail_bucket,
                                          'fail_pct': fail_pct, 'fail_count': fail_count})

    # Fallback for newer BinDistribution HTML where rows are rendered dynamically
    # from JS arrays/objects (DATA / BFS_DATA / FP_DATA) instead of static <tr> rows.
    if not bin_summary_rows or not func_bin_rows:
        try:
            import json as _json

            def _extract_js_array(var_name: str):
                m = re.search(
                    rf'var\s+{re.escape(var_name)}\s*=\s*(\[[\s\S]*?\]);',
                    content,
                    re.DOTALL,
                )
                if not m:
                    return []
                try:
                    return _json.loads(m.group(1))
                except Exception:
                    return []

            if not yield_rows:
                m_data = re.search(r'var\s+DATA\s*=\s*({[\s\S]*?});', content, re.DOTALL)
                if m_data:
                    try:
                        data_obj = _json.loads(m_data.group(1))
                    except Exception:
                        data_obj = {}
                    rows = data_obj.get('rows', []) if isinstance(data_obj, dict) else []
                    ydefs = data_obj.get('yieldDefs', []) if isinstance(data_obj, dict) else []
                    total_die = data_obj.get('total', 0) if isinstance(data_obj, dict) else 0
                    if not total_die:
                        total_die = sum(
                            int(r.get('total', 0) or 0)
                            for r in rows if isinstance(r, dict)
                        )
                    if total_die and ydefs:
                        for yd in ydefs:
                            if not isinstance(yd, dict):
                                continue
                            bins_list = [str(b) for b in (yd.get('bins_list') or [])]
                            cnt = 0
                            for rr in rows:
                                if not isinstance(rr, dict):
                                    continue
                                bc = rr.get('binCounts') or {}
                                if not isinstance(bc, dict):
                                    continue
                                cnt += sum(int(bc.get(bk, 0) or 0) for bk in bins_list)
                            try:
                                expected_pct = float(yd.get('expected')) if yd.get('expected') not in (None, '') else None
                            except Exception:
                                expected_pct = None
                            yield_rows.append({
                                'bin': str(yd.get('bins', '') or ''),
                                'fail_bucket': str(yd.get('bucket', '') or ''),
                                'yield_pct': (cnt / float(total_die) * 100.0) if total_die else None,
                                'expected_pct': expected_pct,
                            })

            if not bin_summary_rows:
                bfs_rows = _extract_js_array('BFS_DATA')
                for r in bfs_rows:
                    ibin_raw = str(r.get('bin', '')).strip()
                    # Stored values may contain warning symbol (e.g., "15⚠").
                    ibin = ibin_raw.rstrip('\u26a0').strip()
                    try:
                        fail_pct = float(r.get('pct'))
                    except Exception:
                        fail_pct = None
                    try:
                        fail_count = int(r.get('count'))
                    except Exception:
                        fail_count = None
                    if ibin and fail_pct is not None:
                        bin_summary_rows.append({
                            'ibin': ibin,
                            'cat': str(r.get('cat', '') or ''),
                            'desc': str(r.get('desc', '') or ''),
                            'fail_pct': fail_pct,
                            'fail_count': fail_count,
                        })

            if not func_bin_rows:
                fp_rows = _extract_js_array('FP_DATA')
                for r in fp_rows:
                    try:
                        fbin = str(int(float(r.get('fb'))))
                    except Exception:
                        fbin = str(r.get('fb', '') or '')
                    try:
                        fail_pct = float(r.get('pct'))
                    except Exception:
                        fail_pct = None
                    try:
                        fail_count = int(r.get('count'))
                    except Exception:
                        fail_count = None
                    if fbin and fail_pct is not None:
                        # Keep key name "ibin" because compare chart/table logic expects it.
                        func_bin_rows.append({
                            'ibin': fbin,
                            'fbin': fbin,
                            'fail_bucket': str(r.get('bkt', '') or ''),
                            'fail_pct': fail_pct,
                            'fail_count': fail_count,
                        })
        except Exception:
            pass

    return {'yield_rows': yield_rows, 'bin_fail_rows': bin_fail_rows,
            'func_bin_rows': func_bin_rows, 'bin_summary_rows': bin_summary_rows}


# ---------------------------------------------------------------------------
# 3. Read xlsx → structured data
# ---------------------------------------------------------------------------

def read_xlsx(xlsx_path: Path):
    """Return dict with keys:
        num_die, col_headers, groups, totals
        groups: list of {name, rows: [(label, vals...)], sum_vals: (vals...)}
        totals: {label, vals} for 'Yield SUM' row
    """
    if not HAVE_OPENPYXL:
        return None
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    ws = wb[wb.sheetnames[0]]

    # Build raw cell grid
    grid = {}
    fmt_grid = {}
    for ri, row in enumerate(ws.iter_rows(values_only=False)):
        for ci, cell in enumerate(row):
            grid[(ri, ci)] = cell.value
            fmt_grid[(ri, ci)] = cell.number_format or ''

    # Find num_die
    num_die = None
    for ri in range(ws.max_row):
        if grid.get((ri, 0)) == '# Die':
            num_die = grid.get((ri, 1))
            break

    if not num_die:
        return None

    # Memoised formula evaluator
    _ev_cache = {}

    def _ev(ri, ci):
        if (ri, ci) in _ev_cache:
            return _ev_cache[(ri, ci)]
        v = grid.get((ri, ci))
        if v is None:
            _ev_cache[(ri, ci)] = None
            return None
        if not isinstance(v, str) or not v.startswith('='):
            _ev_cache[(ri, ci)] = v
            return v
        result = None
        # =N/B2
        m1 = re.match(r'^=(-?\d+)/B\d+$', v)
        if m1:
            result = int(m1.group(1)) / num_die
        # =SUM(X3:X6)
        elif re.match(r'^=SUM\([A-Z]+\d+:[A-Z]+\d+\)$', v):
            m2 = re.match(r'^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', v)
            if m2:
                c1 = column_index_from_string(m2.group(1)) - 1
                r1 = int(m2.group(2)) - 1
                c2 = column_index_from_string(m2.group(3)) - 1
                r2 = int(m2.group(4)) - 1
                result = sum(
                    _ev(r, c1) for r in range(r1, r2 + 1)
                    if isinstance(_ev(r, c1), (int, float))
                )
        # =SUM(X3,X7,...) — comma separated refs
        elif re.match(r'^=SUM\(([^)]+)\)$', v):
            m3 = re.match(r'^=SUM\(([^)]+)\)$', v)
            if m3:
                total = 0.0
                for ref in m3.group(1).split(','):
                    mr = re.match(r'^([A-Z]+)(\d+)$', ref.strip())
                    if mr:
                        rc = column_index_from_string(mr.group(1)) - 1
                        rr = int(mr.group(2)) - 1
                        sv = _ev(rr, rc)
                        if isinstance(sv, (int, float)):
                            total += sv
                result = total
        _ev_cache[(ri, ci)] = result
        return result

    def _pct(ri, ci):
        v = _ev(ri, ci)
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return round(v * 100, 2)
        return None

    # Find column headers row
    col_hdrs = []
    hdr_ri = None
    for ri in range(ws.max_row):
        if grid.get((ri, 0)) == 'Sub Module':
            hdr_ri = ri
            col_hdrs = [grid.get((ri, ci)) for ci in range(ws.max_column)
                        if grid.get((ri, ci)) is not None]
            break

    if hdr_ri is None:
        return None

    num_cols = len(col_hdrs) - 1  # exclude 'Sub Module' column

    # Which data columns are percentages (header contains '%')
    col_is_pct = ['%' in str(h) for h in col_hdrs[1:]]

    def _read_val(ri, ci):
        """Return pct-scaled value for % columns, raw value otherwise."""
        idx = ci - 1  # 0-based into col_is_pct
        if idx < len(col_is_pct) and col_is_pct[idx]:
            return _pct(ri, ci)
        v = _ev(ri, ci)
        if isinstance(v, (int, float)):
            return v
        return None

    # Parse data rows into groups
    groups = []
    current_rows = []
    totals = None

    for ri in range(hdr_ri + 1, ws.max_row + 1):
        label = grid.get((ri, 0))
        if label is None:
            continue
        label_str = str(label).strip()
        if not label_str or label_str.startswith('#'):
            continue

        vals = [_read_val(ri, ci) for ci in range(1, 1 + num_cols)]

        if label_str.upper() == 'SUM':
            groups.append({
                'rows': current_rows,
                'sum_vals': vals,
            })
            current_rows = []
        elif re.match(r'^Yield\s+SUM', label_str, re.IGNORECASE):
            totals = {'label': label_str, 'vals': vals}
        else:
            current_rows.append({'label': label_str, 'vals': vals})

    # Flush any trailing rows without a SUM (e.g. TPI/Other group)
    if current_rows:
        groups.append({'rows': current_rows, 'sum_vals': [None] * num_cols})

    return {
        'num_die': num_die,
        'col_headers': col_hdrs[1:],
        'col_is_pct': col_is_pct,
        'groups': groups,
        'totals': totals,
        'sheet': wb.sheetnames[0],
    }


# ---------------------------------------------------------------------------
# 4. Chart helpers
# ---------------------------------------------------------------------------

def _fig_b64(fig, dpi=130):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('ascii')


def _esc(s):
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


# Color palette for identifiers (up to 10)
_ID_COLORS = [
    '#2980b9', '#27ae60', '#e74c3c', '#f39c12',
    '#8e44ad', '#16a085', '#d35400', '#2c3e50',
    '#c0392b', '#1abc9c',
]

_GROUP_COLORS = [
    '#eaf4ea', '#e3f0fc', '#fff6e6', '#f3e8fb',
    '#fdecea', '#e0f7fa', '#fdf6e3', '#f9ece8',
]


def build_sum_comparison_chart(runs_data, col_idx=0):
    """Grouped bar chart: each group's SUM value per identifier."""
    if not HAVE_MPL:
        return ''
    labels = [r['name'] for r in runs_data]
    # Collect group names from the run with most groups
    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    groups = best['data']['groups']
    group_names = []
    for g in groups:
        if g['rows']:
            # Use prefix of first row's label as group name
            first = g['rows'][0]['label']
            prefix = re.match(r'^([A-Za-z]+)', first)
            group_names.append(prefix.group(1) if prefix else first[:8])
        else:
            group_names.append('?')

    n_groups = len(group_names)
    n_runs = len(runs_data)
    x = np.arange(n_groups)
    width = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(max(10, n_groups * 1.4), 5))
    for ri, run in enumerate(runs_data):
        if not run['data']:
            continue
        vals = []
        for gi, g in enumerate(run['data']['groups']):
            sv = g['sum_vals'][col_idx] if col_idx < len(g['sum_vals']) else None
            vals.append(sv if sv is not None else 0.0)
        # Pad if fewer groups
        while len(vals) < n_groups:
            vals.append(0.0)
        offset = (ri - n_runs / 2 + 0.5) * width
        bars = ax.bar(x + offset, vals, width, label=run['name'],
                      color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)
        for bar, v in zip(bars, vals):
            if v and v > 0.3:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.1,
                        f'{v:.1f}%', ha='center', va='bottom', fontsize=6.5, rotation=90)

    col_label = runs_data[0]['data']['col_headers'][col_idx] if runs_data[0]['data'] else f'Col {col_idx+1}'
    ax.set_title(f'Group SUM Comparison — {col_label}', fontsize=13, weight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(group_names, rotation=30, ha='right', fontsize=9)
    ax.set_ylabel('Yield / Fallout (%)')
    ax.legend(fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_submodule_chart(runs_data, col_idx=0):
    """Horizontal bar chart: each sub-module row per identifier (top N)."""
    if not HAVE_MPL:
        return ''
    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    all_rows = []
    for g in best['data']['groups']:
        all_rows.extend(g['rows'])

    if not all_rows:
        return ''

    # Sort by first run's value descending (top 20)
    def _first_val(row_label):
        for r in runs_data:
            if not r['data']:
                continue
            for g in r['data']['groups']:
                for row in g['rows']:
                    if row['label'] == row_label:
                        v = row['vals'][col_idx] if col_idx < len(row['vals']) else None
                        return v or 0.0
        return 0.0

    sorted_rows = sorted(all_rows, key=lambda r: _first_val(r['label']), reverse=True)[:20]
    row_labels = [r['label'] for r in sorted_rows]

    n_runs = len(runs_data)
    n_rows = len(row_labels)
    y = np.arange(n_rows)
    height = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(9, max(5, n_rows * 0.45)))
    for ri, run in enumerate(runs_data):
        if not run['data']:
            continue
        vals = []
        for lbl in row_labels:
            v = 0.0
            for g in run['data']['groups']:
                for row in g['rows']:
                    if row['label'] == lbl:
                        v = row['vals'][col_idx] if col_idx < len(row['vals']) else 0.0
                        if v is None: v = 0.0
            vals.append(v)
        offset = (ri - n_runs / 2 + 0.5) * height
        ax.barh(y + offset, vals, height, label=run['name'],
                color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)

    col_label = runs_data[0]['data']['col_headers'][col_idx] if runs_data[0]['data'] else f'Col {col_idx+1}'
    ax.set_title(f'Sub-Module Breakdown — {col_label}', fontsize=13, weight='bold')
    ax.set_yticks(y)
    ax.set_yticklabels(row_labels, fontsize=8)
    ax.set_xlabel('Fallout / Yield (%)')
    ax.legend(fontsize=8)
    ax.grid(axis='x', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_total_yield_chart(runs_data):
    """Single bar chart: Yield SUM (%) per identifier."""
    if not HAVE_MPL:
        return ''
    labels = [r['name'] for r in runs_data]
    vals = []
    for r in runs_data:
        if r['data'] and r['data']['totals']:
            v = r['data']['totals']['vals'][0]
            vals.append(v if v is not None else 0.0)
        else:
            vals.append(0.0)

    if not any(vals):
        return ''

    fig, ax = plt.subplots(figsize=(max(6, len(labels) * 1.2), 4))
    bars = ax.bar(labels, vals, color=[_ID_COLORS[i % len(_ID_COLORS)] for i in range(len(labels))],
                  alpha=0.88, edgecolor='white')
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.2,
                f'{v:.1f}%', ha='center', va='bottom', fontsize=9, weight='bold')
    ax.set_title('Total Yield Loss SUM by Identifier', fontsize=13, weight='bold')
    ax.set_ylabel('Yield Loss (%)')
    ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=9)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_delta_heatmap(runs_data, ref_name, col_idx=0):
    """Heatmap of delta (run - ref) per sub-module."""
    if not HAVE_MPL:
        return ''
    ref_run = next((r for r in runs_data if r['name'] == ref_name), runs_data[0])
    compare_runs = [r for r in runs_data if r['name'] != ref_run['name']]
    if not compare_runs:
        return ''

    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    all_rows = []
    for g in best['data']['groups']:
        all_rows.extend(g['rows'])
    row_labels = [r['label'] for r in all_rows]

    def _get_val(run, lbl):
        if not run['data']:
            return None
        for g in run['data']['groups']:
            for row in g['rows']:
                if row['label'] == lbl:
                    return row['vals'][col_idx] if col_idx < len(row['vals']) else None
        return None

    ref_vals = [_get_val(ref_run, lbl) for lbl in row_labels]
    matrix = []
    col_labels = []
    for r in compare_runs:
        row_vals = [_get_val(r, lbl) for lbl in row_labels]
        deltas = [
            (rv - refv) if (rv is not None and refv is not None) else float('nan')
            for rv, refv in zip(row_vals, ref_vals)
        ]
        matrix.append(deltas)
        col_labels.append(r['name'])

    data_arr = np.array(matrix)  # shape: (n_compare, n_rows)
    if data_arr.size == 0:
        return ''

    fig, ax = plt.subplots(figsize=(max(8, len(row_labels) * 0.55), max(3, len(col_labels) * 0.7 + 1.5)))
    # Diverging colormap: red = higher fallout (worse), green = lower (better)
    vmax = np.nanmax(np.abs(data_arr))
    if vmax == 0:
        vmax = 1
    im = ax.imshow(data_arr, cmap='RdYlGn_r', aspect='auto',
                   vmin=-vmax, vmax=vmax)
    ax.set_xticks(range(len(row_labels)))
    ax.set_xticklabels(row_labels, rotation=60, ha='right', fontsize=7)
    ax.set_yticks(range(len(col_labels)))
    ax.set_yticklabels(col_labels, fontsize=8)
    for yi in range(len(col_labels)):
        for xi in range(len(row_labels)):
            v = data_arr[yi, xi]
            if not np.isnan(v):
                ax.text(xi, yi, f'{v:+.1f}', ha='center', va='center',
                        fontsize=6, color='black')
    plt.colorbar(im, ax=ax, label='Δ % vs reference')
    col_label = runs_data[0]['data']['col_headers'][col_idx] if runs_data[0]['data'] else ''
    ax.set_title(f'Delta vs {ref_name} — {col_label}\n(red = more fallout, green = less)',
                 fontsize=11, weight='bold')
    fig.tight_layout()
    return _fig_b64(fig)


# ---------------------------------------------------------------------------
# 4b. RDND yield-table charts and Bin Fail chart
# ---------------------------------------------------------------------------

# Palette for stacked fail segments (consistent across runs)
_FAIL_COLORS = [
    '#e74c3c', '#e67e22', '#f39c12', '#2ecc71',
    '#1abc9c', '#3498db', '#9b59b6', '#e8177d',
    '#95a5a6', '#34495e', '#c0392b', '#16a085',
]


def build_top10_pareto_chart(runs_data):
    """Horizontal bar chart: top-10 Interface Bins by max fail% — uses Functional Bin table."""
    if not HAVE_MPL:
        return ''
    # Prefer func_bin_rows (has fbin); fall back to bin_fail_rows
    use_fbin = any(r.get('bin_data') and r['bin_data'].get('func_bin_rows')
                   for r in runs_data)
    row_key  = 'func_bin_rows' if use_fbin else 'bin_fail_rows'
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data'].get(row_key)]
    if not valid:
        return ''

    # Collect all ibins; keep max fail% per ibin
    ibin_max    = {}
    ibin_fbin   = {}
    ibin_bucket = {}
    for run in valid:
        for row in run['bin_data'][row_key]:
            k = row['ibin']
            v = row['fail_pct'] or 0.0
            if v > ibin_max.get(k, 0.0):
                ibin_max[k]    = v
                ibin_bucket[k] = row.get('fail_bucket', '')
                ibin_fbin[k]   = row.get('fbin', '')

    top10 = sorted(ibin_max.keys(), key=lambda k: ibin_max[k], reverse=True)[:10]
    if not top10:
        return ''

    n_runs = len(valid)
    n_bins = len(top10)
    y      = np.arange(n_bins)
    height = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(9, max(4, n_bins * 0.65)))
    for ri, run in enumerate(valid):
        vals = [
            next((ro['fail_pct'] for ro in run['bin_data'][row_key] if ro['ibin'] == k), 0.0) or 0.0
            for k in top10
        ]
        offset = (ri - n_runs / 2 + 0.5) * height
        bars = ax.barh(y + offset, vals, height, label=run['name'],
                       color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.87)
        for bar, v in zip(bars, vals):
            if v >= 0.1:
                ax.text(v + 0.05, bar.get_y() + bar.get_height() / 2,
                        f'{v:.2f}%', va='center', ha='left', fontsize=7)

    if use_fbin:
        ylabels = [
            f'iBin {k}  FBin {ibin_fbin.get(k, "—")}  |  {ibin_bucket.get(k, "")}'
            for k in top10
        ]
    else:
        ylabels = [
            f'iBin {k}  |  {ibin_bucket.get(k, "")}'
            for k in top10
        ]
    ax.set_yticks(y)
    ax.set_yticklabels(ylabels, fontsize=7.5)
    ax.invert_yaxis()
    ax.set_xlabel('Fail (%)')
    ax.set_title('Top 10 Interface Bin Fail Pareto', fontsize=13, weight='bold')
    ax.legend(fontsize=8)
    ax.grid(axis='x', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_upm_median_chart(runs_data):
    """Grouped bar chart: SICC Si Actual Median vs Target per test, per run."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('upm_data') and r['upm_data']]
    if not valid:
        return ''

    # Collect test names from run with most rows
    best = max(valid, key=lambda r: len(r['upm_data']))
    # Only tests that have both actual and target
    tests = [row for row in best['upm_data'] if row.get('sicc_actual') is not None]
    if not tests:
        return ''
    test_labels = [row['test'] for row in tests]

    n_tests = len(test_labels)
    n_runs  = len(valid)
    x       = np.arange(n_tests)
    width   = 0.7 / max(n_runs + 1, 2)  # +1 for target bar

    fig, ax = plt.subplots(figsize=(max(7, n_tests * 0.5), 4))

    # Plot target once (from best run)
    targets = [t.get('sicc_target') for t in tests]
    has_target = any(v is not None for v in targets)

    for ri, run in enumerate(valid):
        actuals = []
        for t in tests:
            row = next((r for r in run['upm_data'] if r['test'] == t['test']), None)
            actuals.append(row['sicc_actual'] if row and row.get('sicc_actual') is not None else 0.0)
        offset = ri * width
        ax.bar(x + offset, actuals, width, label=run['name'],
               color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)

    if has_target:
        tgt_vals = [v if v is not None else 0.0 for v in targets]
        offset = n_runs * width
        ax.bar(x + offset, tgt_vals, width, label='Target',
               color='#2c3e50', alpha=0.55, hatch='//')

    # Centre ticks
    tick_offset = (n_runs * width) / 2
    ax.set_xticks(x + tick_offset)
    ax.set_xticklabels(test_labels, rotation=35, ha='right', fontsize=8)
    ax.set_ylabel('SICC Si Actual Median (A)')
    ax.set_title('SICC Median', fontsize=13, weight='bold')
    ax.yaxis.set_major_locator(plt.MaxNLocator(nbins=12, integer=False))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: '0' if v == 0 else f'{v:.3f}'))
    ax.legend(fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def build_upm_pct_chart(runs_data, upm_target_pct=None, upm_target_label='Target'):
    """Bar chart: UPM % per run (single value per run)."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('upm_data') and r['upm_data']]
    if not valid:
        return ''

    labels = []
    vals   = []
    for run in valid:
        pcts = [row['upm_pct'] for row in run['upm_data'] if row.get('upm_pct') is not None]
        if pcts:
            labels.append(run['name'])
            vals.append(pcts[0])  # UPM % is the same for all rows

    if not vals:
        return ''

    fig, ax = plt.subplots(figsize=(max(4, len(labels) * 0.9), 3.5))
    bars = ax.bar(labels, vals,
                  color=[_ID_COLORS[i % len(_ID_COLORS)] for i in range(len(labels))],
                  alpha=0.88, edgecolor='white')
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.1,
                f'{v:.2f}%', ha='center', va='bottom', fontsize=9, weight='bold')
    if upm_target_pct is not None:
        ax.axhline(upm_target_pct, color='red', linewidth=2, linestyle='--',
                   label=f'{upm_target_label}: {upm_target_pct}%', zorder=5)
        ax.legend(fontsize=9)
    ax.set_title('UPM ULVT 950mV (%)', fontsize=13, weight='bold')
    ax.set_ylabel('UPM (%)')
    _ymax = max(vals + ([upm_target_pct] if upm_target_pct is not None else [])) * 1.15 if vals else 100
    ax.set_ylim(0, _ymax)
    ax.yaxis.set_major_locator(plt.MaxNLocator(nbins=10, integer=False))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: '0' if v == 0 else f'{v:.2f}%'))
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=9)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def build_upm_detail_chart(runs_data, upm_target_pct=None, upm_target_label='Target'):
    """Grouped bar chart: UPM median % per column per run."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('upm_detail') and r['upm_detail']]
    if not valid:
        return ''

    import numpy as np

    # Collect all unique short names across runs (preserve order)
    all_shorts = []
    for run in valid:
        for d in run['upm_detail']:
            if d['short'] not in all_shorts:
                all_shorts.append(d['short'])

    if not all_shorts:
        return ''

    n_cols = len(all_shorts)
    n_runs = len(valid)
    bar_w = 0.8 / n_runs
    x = np.arange(n_cols)

    fig, ax = plt.subplots(figsize=(max(6, n_cols * 1.2 + n_runs * 0.3), 4.5))

    for ri, run in enumerate(valid):
        detail_map = {d['short']: d['pct'] for d in run['upm_detail']}
        vals = [detail_map.get(s, 0) for s in all_shorts]
        offset = x + ri * bar_w - (n_runs - 1) * bar_w / 2
        bars = ax.bar(offset, vals, width=bar_w,
                      color=_ID_COLORS[ri % len(_ID_COLORS)],
                      alpha=0.85, edgecolor='white', label=run['name'])
        for bar, v in zip(bars, vals):
            if v > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.15,
                        f'{v:.1f}%', ha='center', va='bottom', fontsize=6, weight='bold')

    if upm_target_pct is not None:
        ax.axhline(upm_target_pct, color='red', linewidth=2, linestyle='--',
                   label=f'{upm_target_label}: {upm_target_pct}%', zorder=5)

    ax.set_title('UPM Distribution Comparison (Median %)', fontsize=12, weight='bold')
    ax.set_ylabel('Median (%)', fontsize=10)
    ax.set_xticks(x)
    ax.set_xticklabels(all_shorts, rotation=45, ha='right', fontsize=7)
    ax.legend(fontsize=8, loc='best')
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def _bin_all_good(bin_str: str) -> bool:
    """Return True if every number in the bin string is <= 4."""
    nums = [int(x) for x in re.findall(r'\d+', str(bin_str))]
    return bool(nums) and all(n <= 4 for n in nums)


# Bin groups to always show in the top subplot — matched by bin string equality
_KEY_BINS = ['1/2/3/4', '1/2']
# Fallback: treat any row whose label exactly matches one of these
_KEY_BIN_TITLES = {
    '1/2/3/4': 'FF+DF  (Bin 1/2/3/4)',
    '1/2':     'FF  (Bin 1/2)',
}


def build_combined_rdnd_chart(runs_data):
    """Single mixed-format chart:
       - Stacked bars (per identifier) for Bins > 4 fail%      [left Y, 0-100%]
       - Line overlay (per key bin) for FF+DF / FF yield%       [right Y, 0-100%]
       - Dashed hline = expected yield per key bin
       - Dashed tick marker = expected fail total per identifier
    """
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]
    if not valid:
        return ''

    from matplotlib.patches import Patch
    from matplotlib.lines import Line2D

    best = max(valid, key=lambda r: len(r['bin_data']['yield_rows']))
    all_yield_rows = best['bin_data']['yield_rows']

    key_rows  = [row for row in all_yield_rows if row['bin'] in _KEY_BINS]
    if not key_rows:
        key_rows = [row for row in all_yield_rows if _bin_all_good(row['bin'])][:2]
    fail_rows = [row for row in all_yield_rows if not _bin_all_good(row['bin'])]

    n_runs    = len(valid)
    id_labels = [r['name'] for r in valid]
    x         = np.arange(n_runs)
    bar_w     = 0.5

    fig, ax_bar = plt.subplots(figsize=(max(10, n_runs * 2.2), 6))
    ax_line = ax_bar.twinx()          # shared x, independent right Y

    legend_elems = []

    # ---- Stacked bars: Bins > 4 fail% ----
    bottoms = np.zeros(n_runs)
    for si, fr in enumerate(fail_rows):
        bin_key = fr['bin']
        seg_lbl = fr.get('fail_bucket') or bin_key
        clr     = _FAIL_COLORS[si % len(_FAIL_COLORS)]
        vals    = np.array([
            (next((ro['yield_pct'] for ro in r['bin_data']['yield_rows']
                   if ro['bin'] == bin_key), None) or 0.0)
            for r in valid
        ])
        ax_bar.bar(x, vals, bar_w, bottom=bottoms, color=clr,
                   alpha=0.80, edgecolor='white', linewidth=0.5, zorder=2)
        for bi, (v, b) in enumerate(zip(vals, bottoms)):
            if v >= 0.8:
                ax_bar.text(bi, b + v / 2, f'{v:.1f}%',
                            ha='center', va='center', fontsize=6.5,
                            color='white', weight='bold')
        legend_elems.append(Patch(color=clr, alpha=0.80, label=seg_lbl))
        bottoms += vals

    # Dashed expected-total tick per identifier
    for xi, run in enumerate(valid):
        exp_total = sum(
            (next((ro['expected_pct'] for ro in run['bin_data']['yield_rows']
                   if ro['bin'] == fr['bin']), None) or 0.0)
            for fr in fail_rows
        )
        if exp_total:
            ax_bar.plot([xi - 0.28, xi + 0.28], [exp_total, exp_total],
                        color='#2c3e50', linestyle='--', linewidth=2.0, zorder=5)

    legend_elems.append(Line2D([0], [0], color='#2c3e50', linestyle='--',
                                linewidth=2.0, label='Expected Fail Total'))

    ax_bar.set_ylabel('Fail (%)',  fontsize=10)
    # Dynamic y-limit: max stacked fail total across all runs + 15% buffer
    max_fail = float(bottoms.max()) if bottoms.max() > 0 else 1.0
    also_exp = []
    for run in valid:
        exp_t = sum(
            (next((ro['expected_pct'] for ro in run['bin_data']['yield_rows']
                   if ro['bin'] == fr['bin']), None) or 0.0)
            for fr in fail_rows
        )
        if exp_t:
            also_exp.append(exp_t)
    if also_exp:
        max_fail = max(max_fail, max(also_exp))
    fail_ylim = min(100.0, max_fail * 1.20)   # 20% headroom, cap at 100
    fail_ylim = max(fail_ylim, 5.0)            # at least 5% so chart is readable
    ax_bar.set_ylim(0, fail_ylim)
    ax_bar.set_xticks(x)
    ax_bar.set_xticklabels(id_labels, rotation=15, ha='right', fontsize=9)
    ax_bar.grid(axis='y', linestyle='--', alpha=0.3, zorder=0)

    # ---- Lines: key-bin yield% (right Y axis) ----
    line_colors = ['#1a73e8', '#e53935', '#2e7d32', '#f57c00']
    for ki, kr in enumerate(key_rows):
        bin_key = kr['bin']
        lbl = _KEY_BIN_TITLES.get(bin_key,
                                  f"{bin_key} {kr.get('fail_bucket', '')}".strip())
        lclr = line_colors[ki % len(line_colors)]
        vals = np.array([
            (next((ro['yield_pct'] for ro in r['bin_data']['yield_rows']
                   if ro['bin'] == bin_key), None) or 0.0)
            for r in valid
        ])
        ax_line.plot(x, vals, marker='o', linewidth=2.4, markersize=8,
                     color=lclr, label=lbl, zorder=6)
        for xi, v in enumerate(vals):
            ax_line.text(xi, v + 1.5, f'{v:.1f}%', ha='center', va='bottom',
                         fontsize=8, color=lclr, weight='bold')
        # Dashed expected hline
        ev = kr.get('expected_pct')
        if ev is not None:
            ax_line.axhline(ev, color=lclr, linestyle=':', linewidth=1.6,
                            alpha=0.65, zorder=4)
            ax_line.text(n_runs - 0.45, ev - 2.5 - ki * 3.0, f'Exp {ev:.1f}%',
                         ha='right', va='top', fontsize=7.5,
                         color=lclr, alpha=0.85)
        legend_elems.append(Line2D([0], [0], color=lclr, marker='o',
                                   linewidth=2.4, markersize=8, label=lbl))

    ax_line.set_ylabel('Yield (%)', fontsize=10, y=0.4)
    ax_line.set_ylim(0, 100)

    ax_bar.set_title(
        'Yield (%) and Fail (%) Chart',
        fontsize=12, weight='bold'
    )
    ax_bar.legend(handles=legend_elems, fontsize=7.5,
                  loc='upper left', bbox_to_anchor=(1.18, 1.0),
                  borderaxespad=0)

    fig.tight_layout(pad=2.0)
    return _fig_b64(fig)


def build_fail_stacked_chart(runs_data):
    """Kept for backward compat — now delegates to combined chart."""
    return build_combined_rdnd_chart(runs_data)


def build_rdnd_delta_chart(runs_data):
    """Bars: (actual yield - expected) per BIN per run. Green = beating target, red = below."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]
    if not valid:
        return ''
    best = max(valid, key=lambda r: len(r['bin_data']['yield_rows']))
    yield_rows    = best['bin_data']['yield_rows']
    bin_labels    = [row['bin'] for row in yield_rows]
    expected_vals = [row['expected_pct'] for row in yield_rows]

    n_bins = len(bin_labels)
    n_runs = len(valid)
    x      = np.arange(n_bins)
    width  = 0.8 / max(n_runs, 1)

    from matplotlib.patches import Patch
    fig, ax = plt.subplots(figsize=(max(12, n_bins * 1.1), 4))
    legend_elems = []
    for ri, run in enumerate(valid):
        actuals = [
            next((r['yield_pct'] for r in run['bin_data']['yield_rows'] if r['bin'] == bl), None)
            for bl in bin_labels
        ]
        deltas = [
            (a - e) if (a is not None and e is not None) else 0.0
            for a, e in zip(actuals, expected_vals)
        ]
        offset = (ri - n_runs / 2 + 0.5) * width
        bar_colors = ['#27ae60' if d > 0 else ('#c0392b' if d < 0 else '#95a5a6')
                      for d in deltas]
        ax.bar(x + offset, deltas, width, color=bar_colors, alpha=0.85)
        legend_elems.append(Patch(color=_ID_COLORS[ri % len(_ID_COLORS)],
                                  alpha=0.85, label=run['name']))
    ax.axhline(0, color='#2c3e50', linewidth=1)
    ax.set_title('RDND Yield Delta vs Expected  (actual - expected)\n'
                 'Green = beating target  |  Red = below target',
                 fontsize=12, weight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(bin_labels, rotation=35, ha='right', fontsize=7)
    ax.set_ylabel('\u0394 Yield (%)')
    ax.legend(handles=legend_elems, fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_bin_fail_chart(runs_data):
    """Horizontal grouped bars: fail% per Interface Bin per run, sorted by max fail."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['bin_fail_rows']]
    if not valid:
        return ''

    # Union of ibins in order from the run with most rows
    best = max(valid, key=lambda r: len(r['bin_data']['bin_fail_rows']))
    ibin_meta = {}
    for row in best['bin_data']['bin_fail_rows']:
        ibin_meta.setdefault(row['ibin'], row['fail_bucket'])

    def _max_fail(ibin):
        return max(
            (next((ro['fail_pct'] for ro in r['bin_data']['bin_fail_rows'] if ro['ibin'] == ibin), 0.0) or 0.0)
            for r in valid
        )

    sorted_ibins = sorted(ibin_meta.keys(), key=_max_fail, reverse=True)
    ibin_labels  = [f"Bin {k} \u2014 {ibin_meta[k]}" for k in sorted_ibins]

    n_rows = len(sorted_ibins)
    n_runs = len(valid)
    y      = np.arange(n_rows)
    height = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(9, max(5, n_rows * 0.42)))
    for ri, run in enumerate(valid):
        vals = [
            next((ro['fail_pct'] for ro in run['bin_data']['bin_fail_rows'] if ro['ibin'] == k), 0.0) or 0.0
            for k in sorted_ibins
        ]
        offset = (ri - n_runs / 2 + 0.5) * height
        ax.barh(y + offset, vals, height, label=run['name'],
                color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)
    ax.set_title('Bin Fail Summary — Fail% per Interface Bin', fontsize=13, weight='bold')
    ax.set_yticks(y)
    ax.set_yticklabels(ibin_labels, fontsize=7)
    ax.set_xlabel('Fail (%)')
    ax.legend(fontsize=8)
    ax.grid(axis='x', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


# ---------------------------------------------------------------------------
# 4c. Run summary table (Program Name, Material Type, Wafer, # Dies)
# ---------------------------------------------------------------------------

def parse_index_meta(dash_dir: Path, index_href: str) -> dict:
    """Parse program, lot(s), wafer(s), material type(s) from a run's BinDistribution HTML
    by reading the embedded DATA JS variable written by bin_distribution_html.py."""
    result = {'program': '', 'lots': [], 'wafers': [], 'material': []}
    if not index_href:
        return result
    try:
        import json as _json_idx
        href = re.sub(r'^file:///', '', index_href).replace('/', os.sep)
        idx_path = dash_dir / href if not os.path.isabs(href) else Path(href)
        out_folder = idx_path.parent
        if not out_folder.exists():
            return result
        # Find the best BinDistribution.html: prefer *_reticle_material_BinDistribution.html
        bin_html = None
        for pat in ('*_reticle_material_BinDistribution.html',
                    '*_material_merged_*BinDistribution.html',
                    '*BinDistribution.html'):
            cands = sorted(out_folder.glob(pat), key=lambda p: p.stat().st_mtime, reverse=True)
            if cands:
                bin_html = cands[0]
                break
        if not bin_html:
            return result
        content = bin_html.read_text(encoding='utf-8', errors='replace')
        # Extract var DATA = {...}; — ends before the next var declaration
        m = re.search(r'var\s+DATA\s*=\s*(\{[\s\S]*?\});\s*(?:var\s|\Z)', content)
        if not m:
            m = re.search(r'var\s+DATA\s*=\s*(\{[\s\S]*?\});', content)
        if m:
            try:
                data = _json_idx.loads(m.group(1))
                rows = data.get('rows', [])
                programs, lots, wafers, mats = set(), set(), set(), set()
                for row in rows:
                    if row.get('program'):
                        programs.add(str(row['program']))
                    if row.get('lot'):
                        lots.add(str(row['lot']))
                    if row.get('wafer') and str(row['wafer']) not in ('', 'all'):
                        wafers.add(str(row['wafer']))
                    if row.get('material'):
                        mats.add(str(row['material']))
                def _sort_nums(s):
                    try:
                        return sorted(s, key=lambda x: int(x) if str(x).isdigit() else x)
                    except Exception:
                        return sorted(s)
                result['program']  = ', '.join(sorted(programs))
                result['lots']     = _sort_nums(lots)
                result['wafers']   = _sort_nums(wafers)
                result['material'] = sorted(mats)
            except Exception:
                pass
    except Exception:
        pass
    return result


def _find_processed_csv(output_dir: Path):
    """Find the best enriched/processed CSV for a run.
    Priority: *_reticle_material.csv > *_material_merged.csv > any CSV in output_dir > parent dir."""
    def _excl(name: str) -> bool:
        n = name.lower()
        return ('bindef' in n or '_targets_' in n or '_bindef' in n)

    def _csvs(d: Path):
        if not d or not d.exists():
            return []
        return [p for p in sorted(d.glob('*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
                if not _excl(p.name)]

    # output_dir is the identifier subfolder — enriched CSVs are written there
    for search_dir in [output_dir,
                       output_dir.parent if output_dir else None,
                       output_dir.parent.parent if output_dir and output_dir.parent else None]:
        if search_dir is None:
            continue
        cands = _csvs(search_dir)
        if not cands:
            continue
        # Strongest preference: fully enriched (reticle + material) file
        for p in cands:
            if '_reticle_material' in p.name.lower():
                return p
        # Second preference: material-merged intermediate
        for p in cands:
            if '_material_merged' in p.name.lower():
                return p
        # Third: any CSV with 'reticle' or 'material' in name
        for p in cands:
            n = p.name.lower()
            if 'reticle' in n or 'material' in n or 'enriched' in n:
                return p
        # Fallback: newest CSV in this dir
        return cands[0]
    return None


def _extract_csv_meta(output_dir: Path) -> dict:
    """Read the processed/enriched CSV to extract Program Name, Material Type, Wafer list."""
    result = {'program': '', 'material': '', 'wafers': ''}
    raw = _find_processed_csv(output_dir)
    if not raw:
        return result
    try:
        import pandas as _pd_meta
        # Read only first 10k rows for speed; use low_memory=False so mixed cols don't truncate
        df = _pd_meta.read_csv(str(raw), dtype=object, nrows=10000)
        prog_col = next((c for c in df.columns if 'program' in c.lower()), None)
        # Material Type column: look for exact match first, then substring
        mat_col = next(
            (c for c in df.columns if c.strip().lower() in
             ('material type', 'materialtype', 'material_type',
              'material type, skew, beol skew')),
            None
        ) or next((c for c in df.columns if 'material type' in c.lower()), None)
        wafer_col = (next((c for c in df.columns if 'sort_wafer' in c.lower()), None)
                     or next((c for c in df.columns if 'wafer' in c.lower()), None))
        if prog_col:
            vals = [str(v) for v in df[prog_col].dropna().unique()]
            result['program'] = ', '.join(vals[:4])
        if mat_col:
            vals = [str(v) for v in df[mat_col].dropna().unique()]
            result['material'] = ', '.join(vals[:6])
        if wafer_col:
            wvals = df[wafer_col].dropna().unique()
            try:
                wvals = sorted(wvals, key=lambda x: int(str(x)) if str(x).isdigit() else str(x))
            except Exception:
                wvals = sorted([str(w) for w in wvals])
            result['wafers'] = ', '.join([str(w) for w in wvals])
    except Exception:
        pass
    return result


def build_run_summary_table_html(runs_data, dash_dir: Path = None) -> str:
    """Compact table: Program Name / Lot(s) / Material Type / Wafer(s) / # Dies — one column per run."""
    metas = []
    for r in runs_data:
        meta = {'program': '', 'lots': '', 'material': '', 'wafers': '', 'num_die': ''}

        # # Dies from xlsx
        if r.get('data') and r['data'].get('num_die') is not None:
            try:
                meta['num_die'] = f"{int(r['data']['num_die']):,}"
            except Exception:
                meta['num_die'] = str(r['data']['num_die'])

        # Program / Lot / Wafer / Material from the run's index.html UDATA
        if dash_dir and r.get('index_href'):
            idx_meta = parse_index_meta(dash_dir, r['index_href'])
            meta['lots']     = _esc(', '.join(idx_meta['lots']))
            meta['wafers']   = _esc(', '.join(idx_meta['wafers']))
            meta['material'] = _esc(', '.join(idx_meta['material']))
            # Program comes from UDATA program field; if empty fall back to run name
            if idx_meta.get('program'):
                meta['program'] = _esc(idx_meta['program'])
            else:
                meta['program'] = _esc(r.get('name', ''))
        else:
            meta['program'] = _esc(r.get('name', ''))

        metas.append(meta)

    ROWS = [
        ('Program Name',  'program'),
        ('Lot(s)',        'lots'),
        ('Material Type', 'material'),
        ('Wafer(s)',      'wafers'),
        ('# Dies',        'num_die'),
    ]

    _th_base = 'white-space:nowrap;padding:6px 10px'
    hdr = f'<th style="background:#1a3a5c;color:#fff;{_th_base}">Metric</th>'
    for ri, r in enumerate(runs_data):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th style="background:{clr};color:#fff;font-weight:bold;'
                f'text-align:center;{_th_base}">'
                f'{_esc(r["name"])}</th>')

    _td_base = 'padding:5px 10px;font-size:20px;white-space:nowrap'
    rows_html = ''
    for label, key in ROWS:
        cells = (f'<td style="font-weight:bold;background:#e8f0fb;color:#1a3a5c;{_td_base}">{label}</td>')
        for mi, m in enumerate(metas):
            bg = '#ffffff' if mi % 2 == 0 else '#f5f8ff'
            val = m[key] or '\u2014'
            cells += (f'<td style="background:{bg};color:#222;{_td_base}">{val}</td>')
        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128204; Run Summary</h2>
  <div style="overflow-x:auto">
  <table style="border-collapse:collapse;table-layout:auto">
    <thead><tr>{hdr}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 4d. RDND and Bin Fail comparison table HTML
# ---------------------------------------------------------------------------

def build_xlsx_comparison_table(runs_data):
    """Side-by-side table of all rows from *_out.xlsx for every identifier."""
    valid = [r for r in runs_data if r.get('data')]
    if not valid:
        return ''

    # All runs (including those with no xlsx) — show '—' columns for missing ones
    all_runs = runs_data

    # Collect all unique sub-module labels in display order from the run with most groups
    best = max(valid, key=lambda r: len(r['data']['groups']))
    ordered_labels = []   # (label, is_sum, group_idx)
    for gi, group in enumerate(best['data']['groups']):
        for row in group['rows']:
            ordered_labels.append((row['label'], False, gi))
        ordered_labels.append(('SUM', True, gi))
    if best['data'] and best['data']['totals']:
        ordered_labels.append((best['data']['totals']['label'], True, -1))

    n_cols = max(len(r['data']['col_headers']) for r in valid)
    col_hdrs = best['data']['col_headers']
    col_is_pct = best['data'].get('col_is_pct', [True] * n_cols)

    def _fmt_val(v, ci):
        """Format cell value: add % for pct columns, plain number for raw columns."""
        if v is None:
            return ''
        if ci < len(col_is_pct) and col_is_pct[ci]:
            return f'{v:.1f}%'
        # Raw column (e.g. Die id count)
        return f'{int(v):,}' if isinstance(v, float) and v == int(v) else f'{v:g}'

    def _get_row_vals(run, lbl, is_sum, gi):
        """Return list of values (one per data column) for a given row."""
        if not run['data']:
            return [None] * n_cols
        if is_sum and lbl.upper() == 'SUM':
            grps = run['data']['groups']
            if gi < len(grps):
                sv = grps[gi]['sum_vals']
                return [sv[ci] if ci < len(sv) else None for ci in range(n_cols)]
            return [None] * n_cols
        if is_sum:  # Yield SUM row
            if run['data']['totals']:
                v = run['data']['totals']['vals']
                return [v[ci] if ci < len(v) else None for ci in range(n_cols)]
            return [None] * n_cols
        for g in run['data']['groups']:
            for row in g['rows']:
                if row['label'] == lbl:
                    return [row['vals'][ci] if ci < len(row['vals']) else None
                            for ci in range(n_cols)]
        return [None] * n_cols

    # Build header — use all_runs so every identifier appears
    hdr = '<th rowspan="2">Sub Module</th>'
    for ri, r in enumerate(all_runs):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        no_data = '' if r.get('data') else ' ⚠ no xlsx'
        hdr += (f'<th colspan="{n_cols}" style="background:{clr};color:#fff;'
                f'font-weight:bold;text-align:center;padding:6px 8px">'
                f'{_esc(r["name"])}{no_data}</th>')
    hdr2 = ''
    for ri, r in enumerate(all_runs):
        for ci in range(n_cols):
            if r.get('data'):
                ch = col_hdrs[ci] if ci < len(col_hdrs) else f'Col {ci+1}'
                hdr2 += f'<th style="font-size:17px;white-space:nowrap">{_esc(ch)}</th>'
            else:
                hdr2 += f'<th style="font-size:17px;white-space:nowrap;color:#aaa">—</th>'

    rows_html = ''
    for lbl, is_sum, gi in ordered_labels:
        grp_clr = _GROUP_COLORS[gi % len(_GROUP_COLORS)] if gi >= 0 else '#d5e8d4'
        is_total = (gi == -1)
        bold = 'font-weight:bold;' if is_sum else ''
        border = 'border-top:2px solid #aaa;' if is_sum and not is_total else ''
        border = 'border-top:3px solid #555;' if is_total else border
        lbl_cell = (f'<td style="{bold}{border}background:{grp_clr};'
                    f'font-size:20px;white-space:nowrap">{_esc(lbl)}</td>')
        cells = lbl_cell
        # Only include valid runs in the highlight calculation
        all_run_vals = [_get_row_vals(rx, lbl, is_sum, gi) for rx in valid]
        for ri, r in enumerate(all_runs):
            if not r.get('data'):
                # No xlsx for this run — show N/A cells
                for ci in range(n_cols):
                    base_st = f'{bold}{border}background:{grp_clr};color:#bbb;'
                    cells += f'<td class="num" style="{base_st}">—</td>'
                continue
            vals = _get_row_vals(r, lbl, is_sum, gi)
            for ci, v in enumerate(vals):
                base_st = f'{bold}{border}background:{grp_clr};'
                if col_is_pct[ci] if ci < len(col_is_pct) else False:
                    col_row_vals = [arv[ci] for arv in all_run_vals]
                    cells += _cell_hl(v, col_row_vals, extra_style=base_st) + _fmt_val(v, ci) + '</td>'
                else:
                    cells += (f'<td class="num" style="{base_st}">'
                              f'{_fmt_val(v, ci)}</td>')
        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128203; Digital Dashboard</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead>
      <tr>{hdr}</tr>
      <tr>{hdr2}</tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


def _cell_hl(v, row_vals, extra_style=''):
    """Return a <td> string. Bold-red if |v - row_mean| > 10 (percentage points)."""
    nums = [x for x in row_vals if x is not None]
    if v is None:
        return f'<td class="num" style="{extra_style}"></td>'
    alert = (len(nums) >= 2 and abs(v - (sum(nums) / len(nums))) > 10)
    st = extra_style + ('color:#c0392b;font-weight:bold;' if alert else '')
    return f'<td class="num" style="{st}">'


def build_rdnd_table_html(runs_data):
    """Table: BIN rows, Expected % + each run's actual yield% + delta vs expected."""
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]
    if not valid:
        return ''
    best = max(valid, key=lambda r: len(r['bin_data']['yield_rows']))
    yield_rows = best['bin_data']['yield_rows']

    hdr = '<th>BIN</th><th>Fail Bucket</th><th>Expected (%)</th>'
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += f'<th style="background:{clr};color:#fff;font-weight:bold;padding:5px 8px">{_esc(r["name"])}</th>'
    # add delta headers between consecutive runs
    if len(valid) >= 2:
        for ri in range(1, len(valid)):
            hdr += f'<th style="background:#34495e;color:#fff;font-weight:bold;padding:5px 8px;font-size:11px">\u0394 {_esc(valid[ri]["name"])}<br>vs {_esc(valid[0]["name"])}</th>'

    rows_html = ''
    for row in yield_rows:
        bl  = row['bin']
        exp = row.get('expected_pct')
        cells = (
            f'<td style="white-space:nowrap;font-size:20px">{_esc(bl)}</td>'
            f'<td style="font-size:20px">{_esc(row.get("fail_bucket", ""))}</td>'
            f'<td class="num" style="color:#555">{f"{exp:.1f}%" if exp is not None else ""}</td>'
        )
        run_vals = []
        for r in valid:
            v = next((ro['yield_pct'] for ro in r['bin_data']['yield_rows'] if ro['bin'] == bl), None)
            run_vals.append(v)
            row_vals = [
                next((ro['yield_pct'] for ro in rx['bin_data']['yield_rows'] if ro['bin'] == bl), None)
                for rx in valid
            ]
            cells += _cell_hl(v, row_vals) + (f'{v:.1f}%' if v is not None else '') + '</td>'
        # delta cells: each run vs first run (baseline)
        if len(valid) >= 2:
            base = run_vals[0]
            for ri in range(1, len(valid)):
                v = run_vals[ri]
                if base is not None and v is not None:
                    delta = v - base
                    sign = '+' if delta > 0 else ''
                    clr = '#27ae60' if delta > 0 else '#c0392b' if delta < 0 else '#555'
                    cells += f'<td class="num" style="color:{clr};font-weight:bold">{sign}{delta:.2f}%</td>'
                else:
                    cells += '<td class="num">\u2014</td>'
        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128203; Yield Table</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead><tr>{hdr}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


def build_bin_fail_table_html(runs_data):
    """Table: Interface Bin rows, each run's Yield/Fail% — uses new bin_summary_rows
    (Category/Description format) when available, else falls back to bin_fail_rows."""
    _CAT_PALETTE = [
        '#dbeeff','#e0f5e0','#fef3cd','#fde0d0','#ece0f8','#d0f4f4',
        '#fce4ec','#e8f5e9','#fff3e0','#e3f2fd','#f3e5f5','#e8eaf6',
    ]

    # Prefer new 6-col format
    use_summary = any(r.get('bin_data') and r['bin_data'].get('bin_summary_rows')
                      for r in runs_data)
    if use_summary:
        valid = [r for r in runs_data if r.get('bin_data') and r['bin_data'].get('bin_summary_rows')]
        if not valid:
            return ''
        best = max(valid, key=lambda r: len(r['bin_data']['bin_summary_rows']))
        all_rows = list({row['ibin']: row for row in best['bin_data']['bin_summary_rows']}.values())

        # Build category→color map in row order
        _cat_color = {}
        for row in all_rows:
            ck = row.get('cat', '').strip().lower()
            if ck and ck not in _cat_color:
                _cat_color[ck] = _CAT_PALETTE[len(_cat_color) % len(_CAT_PALETTE)]

        hdr = '<th>Bin</th><th>Category</th><th>Description</th>'
        for ri, r in enumerate(valid):
            clr = _ID_COLORS[ri % len(_ID_COLORS)]
            hdr += (f'<th style="background:{clr};color:#fff;font-weight:bold;padding:5px 8px">'
                    f'{_esc(r["name"])}<br><span style="font-size:11px;font-weight:normal">Yield/Fail%</span></th>')
        # delta headers
        if len(valid) >= 2:
            for ri in range(1, len(valid)):
                hdr += f'<th style="background:#34495e;color:#fff;font-weight:bold;padding:5px 8px;font-size:11px">\u0394 {_esc(valid[ri]["name"])}<br>vs {_esc(valid[0]["name"])}</th>'

        rows_html = ''
        for row in all_rows:
            key  = row['ibin']
            cat  = row.get('cat', '')
            desc = row.get('desc', '')
            row_bg = _cat_color.get(cat.strip().lower(), '#ffffff')
            cells = (f'<td style="background:{row_bg}">{_esc(key)}</td>'
                     f'<td style="background:{row_bg}">{_esc(cat)}</td>'
                     f'<td style="background:{row_bg}">{_esc(desc)}</td>')
            run_vals = []
            for r in valid:
                v = next((ro['fail_pct'] for ro in r['bin_data']['bin_summary_rows']
                          if ro['ibin'] == key), None)
                run_vals.append(v)
                row_all = [next((ro['fail_pct'] for ro in rx['bin_data']['bin_summary_rows'] if ro['ibin'] == key), None) for rx in valid]
                cells += _cell_hl(v, row_all, extra_style=f'background:{row_bg};') + (f'{v:.2f}%' if v is not None else '\u2014') + '</td>'
            # delta cells
            if len(valid) >= 2:
                base = run_vals[0]
                for ri in range(1, len(valid)):
                    v = run_vals[ri]
                    if base is not None and v is not None:
                        delta = v - base
                        sign = '+' if delta > 0 else ''
                        clr = '#c0392b' if delta > 0 else '#27ae60' if delta < 0 else '#555'
                        cells += f'<td class="num" style="color:{clr};font-weight:bold;background:{row_bg}">{sign}{delta:.2f}%</td>'
                    else:
                        cells += f'<td class="num" style="background:{row_bg}">\u2014</td>'
            rows_html += f'<tr>{cells}</tr>\n'
    else:
        valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['bin_fail_rows']]
        if not valid:
            return ''
        best = max(valid, key=lambda r: len(r['bin_data']['bin_fail_rows']))
        all_rows = list({row['ibin']: row for row in best['bin_data']['bin_fail_rows']}.values())

        hdr = '<th>Interface Bin</th><th>Fail Bucket</th>'
        for ri, r in enumerate(valid):
            clr = _ID_COLORS[ri % len(_ID_COLORS)]
            hdr += f'<th style="background:{clr};color:#fff;font-weight:bold;padding:5px 8px">{_esc(r["name"])}</th>'
        # delta headers
        if len(valid) >= 2:
            for ri in range(1, len(valid)):
                hdr += f'<th style="background:#34495e;color:#fff;font-weight:bold;padding:5px 8px;font-size:11px">\u0394 {_esc(valid[ri]["name"])}<br>vs {_esc(valid[0]["name"])}</th>'

        rows_html = ''
        for row in all_rows:
            key   = row['ibin']
            cells = f'<td>{_esc(key)}</td><td>{_esc(row["fail_bucket"])}</td>'
            run_vals = []
            for r in valid:
                v = next((ro['fail_pct'] for ro in r['bin_data']['bin_fail_rows']
                          if ro['ibin'] == key), None)
                run_vals.append(v)
                row_all = [next((ro['fail_pct'] for ro in rx['bin_data']['bin_fail_rows'] if ro['ibin'] == key), None) for rx in valid]
                cells += _cell_hl(v, row_all) + (f'{v:.2f}%' if v is not None else '\u2014') + '</td>'
            # delta cells
            if len(valid) >= 2:
                base = run_vals[0]
                for ri in range(1, len(valid)):
                    v = run_vals[ri]
                    if base is not None and v is not None:
                        delta = v - base
                        sign = '+' if delta > 0 else ''
                        clr = '#c0392b' if delta > 0 else '#27ae60' if delta < 0 else '#555'
                        cells += f'<td class="num" style="color:{clr};font-weight:bold">{sign}{delta:.2f}%</td>'
                    else:
                        cells += '<td class="num">\u2014</td>'
            rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128196; Bin Fail Summary</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl" style="border-collapse:collapse">
    <thead><tr>{hdr}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 4d. SICC / UPM comparison table HTML
# ---------------------------------------------------------------------------

def build_sicc_table_html(runs_data):
    """Table: one row per test, columns = SICC Actual / SICC Target / Multiple / UPM% per run."""
    valid = [r for r in runs_data if r.get('upm_data') and r['upm_data']]
    if not valid:
        return ''

    # Union of test names, ordered by the run with the most rows
    best  = max(valid, key=lambda r: len(r['upm_data']))
    tests = [row['test'] for row in best['upm_data']]

    def _get(run, test, field):
        row = next((r for r in run['upm_data'] if r['test'] == test), None)
        return row[field] if row else None

    def _fmt(v, decimals=3):
        return f'{v:.{decimals}f}' if v is not None else '—'

    # Build column headers — one group (Actual / Target / Multiple) per run
    hdr = '<th rowspan="2" style="min-width:160px">Test</th>'
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th colspan="3" style="background:{clr};color:#fff;font-weight:bold;'
                f'text-align:left;padding:6px 10px">{_esc(r["name"])}</th>')
    # UPM % — one column per run
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th style="background:{clr};color:#fff;font-weight:bold;text-align:left;'
                f'padding:6px 10px">UPM%<br><span style="font-size:11px">{_esc(r["name"])}</span></th>')

    _th_sub = 'style="font-size:11px;white-space:nowrap;text-align:left!important;padding:4px 10px"'
    hdr2 = ''
    for _ in valid:
        hdr2 += (f'<th {_th_sub}>Actual (A)</th>'
                 f'<th {_th_sub}>Target (A)</th>'
                 f'<th {_th_sub}>Multiple</th>')
    for _ in valid:
        hdr2 += f'<th {_th_sub}>UPM (%)</th>'

    rows_html = ''
    for i, test in enumerate(tests):
        bg = '#f9f9f9' if i % 2 == 0 else '#ffffff'
        cells = f'<td style="white-space:nowrap;font-size:13px;background:{bg}">{_esc(test)}</td>'

        # Collect actuals for highlight comparison
        actuals = [_get(r, test, 'sicc_actual') for r in valid]
        for r in valid:
            act  = _get(r, test, 'sicc_actual')
            tgt  = _get(r, test, 'sicc_target')
            mult = _get(r, test, 'multiple')
            # Red only when actual > target (over spec)
            if act is not None and tgt is not None and act > tgt:
                act_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                act_style = f'background:{bg}'
            # Multiple: red bold only when > 1
            if mult is not None and mult > 1:
                mult_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                mult_style = f'background:{bg}'
            _td = 'font-variant-numeric:tabular-nums;padding:4px 10px'
            cells += (f'<td style="{act_style};{_td}">{_fmt(act)}</td>'
                      f'<td style="background:{bg};{_td}">{_fmt(tgt)}</td>'
                      f'<td style="{mult_style};{_td}">{_fmt(mult, 2)}</td>')

        for r in valid:
            upm = _get(r, test, 'upm_pct')
            cells += f'<td style="background:{bg};font-variant-numeric:tabular-nums;padding:4px 10px">{_fmt(upm, 1)}</td>'

        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#9889; SICC / UPM Table</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead>
      <tr>{hdr}</tr>
      <tr>{hdr2}</tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 4e. CDYN comparison chart + table
# ---------------------------------------------------------------------------

def build_cdyn_median_chart(runs_data):
    """Grouped bar chart: CDYN Actual Median vs Expected per test, per run."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('cdyn_data') and r['cdyn_data']]
    if not valid:
        return ''

    best = max(valid, key=lambda r: len(r['cdyn_data']))
    tests = [row for row in best['cdyn_data'] if row.get('actual') is not None]
    if not tests:
        return ''
    test_labels = [row['test'] for row in tests]

    n_tests = len(test_labels)
    n_runs  = len(valid)
    x       = np.arange(n_tests)
    width   = 0.7 / max(n_runs + 1, 2)

    fig, ax = plt.subplots(figsize=(max(7, n_tests * 0.6), 4))

    expected = [t.get('expected') for t in tests]
    has_expected = any(v is not None for v in expected)

    for ri, run in enumerate(valid):
        actuals = []
        for t in tests:
            row = next((r for r in run['cdyn_data'] if r['test'] == t['test']), None)
            actuals.append(row['actual'] if row and row.get('actual') is not None else 0.0)
        offset = ri * width
        ax.bar(x + offset, actuals, width, label=run['name'],
               color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)

    if has_expected:
        exp_vals = [v if v is not None else 0.0 for v in expected]
        offset = n_runs * width
        ax.bar(x + offset, exp_vals, width, label='Expected',
               color='#2c3e50', alpha=0.55, hatch='//')

    tick_offset = (n_runs * width) / 2
    ax.set_xticks(x + tick_offset)
    ax.set_xticklabels(test_labels, rotation=35, ha='right', fontsize=8)
    ax.set_ylabel('CDYN Actual Median (nF)')
    ax.set_title('CDYN Median', fontsize=13, weight='bold')
    ax.yaxis.set_major_locator(plt.MaxNLocator(nbins=12, integer=False))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: '0' if v == 0 else f'{v:.2f}'))
    ax.legend(fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def build_cdyn_table_html(runs_data):
    """Table: one row per CDYN test, columns = Actual / Expected / Actual/Expected per run."""
    valid = [r for r in runs_data if r.get('cdyn_data') and r['cdyn_data']]
    if not valid:
        return ''

    best  = max(valid, key=lambda r: len(r['cdyn_data']))
    tests = [row['test'] for row in best['cdyn_data']]

    def _get(run, test, field):
        row = next((r for r in run['cdyn_data'] if r['test'] == test), None)
        return row[field] if row else None

    def _fmt(v, decimals=2):
        return f'{v:.{decimals}f}' if v is not None else '\u2014'

    # Header row 1 — test name + one group per run
    hdr = '<th rowspan="2" style="min-width:200px">Test</th>'
    hdr += '<th rowspan="2" style="min-width:80px">Type</th>'
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th colspan="3" style="background:{clr};color:#fff;font-weight:bold;'
                f'text-align:left;padding:6px 10px">{_esc(r["name"])}</th>')

    # Header row 2 — sub-headers
    _th_sub = 'style="font-size:11px;white-space:nowrap;text-align:left!important;padding:4px 10px"'
    hdr2 = ''
    for _ in valid:
        hdr2 += (f'<th {_th_sub}>Actual (nF)</th>'
                 f'<th {_th_sub}>Expected (nF)</th>'
                 f'<th {_th_sub}>Actual/Expected</th>')

    rows_html = ''
    for i, test in enumerate(tests):
        bg = '#f9f9f9' if i % 2 == 0 else '#ffffff'
        test_type = _get(best, test, 'type') or ''
        cells = (f'<td style="white-space:nowrap;font-size:13px;background:{bg}">{_esc(test)}</td>'
                 f'<td style="font-size:12px;background:{bg}">{_esc(test_type)}</td>')

        for r in valid:
            act  = _get(r, test, 'actual')
            exp  = _get(r, test, 'expected')
            ratio = _get(r, test, 'ratio')
            # Red when actual > expected
            if act is not None and exp is not None and act > exp:
                act_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                act_style = f'background:{bg}'
            # Ratio red when > 1
            if ratio is not None and ratio > 1:
                ratio_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                ratio_style = f'background:{bg}'
            _td = 'font-variant-numeric:tabular-nums;padding:4px 10px'
            cells += (f'<td style="{act_style};{_td}">{_fmt(act)}</td>'
                      f'<td style="background:{bg};{_td}">{_fmt(exp)}</td>'
                      f'<td style="{ratio_style};{_td}">{_fmt(ratio)}</td>')

        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#9889; CDYN Median Table</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead>
      <tr>{hdr}</tr>
      <tr>{hdr2}</tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 5. Build comparison table HTML
# ---------------------------------------------------------------------------

def build_comparison_table(runs_data, ref_name, col_idx=0):
    if not runs_data:
        return ''
    ref_run = next((r for r in runs_data if r['name'] == ref_name), runs_data[0])

    def _get_val(run, lbl):
        if not run['data']:
            return None
        for g in run['data']['groups']:
            for row in g['rows']:
                if row['label'] == lbl:
                    return row['vals'][col_idx] if col_idx < len(row['vals']) else None
            if all(row['vals'][col_idx] is None for row in g['rows']) is False:
                if g['sum_vals'] and col_idx < len(g['sum_vals']) and g['rows'] and g['rows'][0]['label'] == lbl:
                    pass
        return None

    def _get_sum(run, gi):
        if not run['data'] or gi >= len(run['data']['groups']):
            return None
        sv = run['data']['groups'][gi]['sum_vals']
        return sv[col_idx] if sv and col_idx < len(sv) else None

    def _get_total(run):
        if not run['data'] or not run['data']['totals']:
            return None
        v = run['data']['totals']['vals']
        return v[col_idx] if v and col_idx < len(v) else None

    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    groups = best['data']['groups']

    # Header row
    run_names = [r['name'] for r in runs_data]
    hdr_cells = ('<th style="background:#1a252f;color:#ecf0f1;font-size:18px">'
                 'Sub Module</th>') + ''.join(
        f'<th style="background:{_ID_COLORS[i % len(_ID_COLORS)]};color:#fff;'
        f'font-size:18px;padding:7px 10px">'
        f'{_esc(n)}</th>'
        for i, n in enumerate(run_names)
    )
    # Add delta columns (vs ref)
    compare_names = [r['name'] for r in runs_data if r['name'] != ref_run['name']]
    for cn in compare_names:
        hdr_cells += (
            f'<th style="background:#4a235a;color:#e8daef;text-align:center;'
            f'font-size:18px;padding:7px 10px">'
            f'&#916; vs <b>{_esc(ref_run["name"])}</b>'
            f'<br><span style="font-size:14px;font-weight:normal">{_esc(cn)}</span></th>'
        )

    rows_html = ''
    for gi, group in enumerate(groups):
        grp_clr = _GROUP_COLORS[gi % len(_GROUP_COLORS)]
        for row in group['rows']:
            lbl = row['label']
            vals = [_get_val(r, lbl) for r in runs_data]
            ref_v = _get_val(ref_run, lbl)
            cells = f'<td style="background:{grp_clr};font-size:20px">{_esc(lbl)}</td>'
            for v in vals:
                cells += f'<td class="num" style="background:{grp_clr}">{f"{v:.1f}%" if v is not None else ""}</td>'
            for r in runs_data:
                if r['name'] == ref_run['name']:
                    continue
                cv = _get_val(r, lbl)
                if cv is not None and ref_v is not None:
                    delta = cv - ref_v
                    clr = '#c0392b' if delta > 0 else ('#27ae60' if delta < 0 else '#555')
                    cells += f'<td class="num delta" style="color:{clr};font-weight:bold">{delta:+.1f}%</td>'
                else:
                    cells += '<td class="num delta">—</td>'
            rows_html += f'<tr>{cells}</tr>\n'

        # SUM row
        sum_vals = [_get_sum(r, gi) for r in runs_data]
        ref_sv = _get_sum(ref_run, gi)
        sum_cells = f'<td style="font-weight:bold;background:{grp_clr};border-top:2px solid #999">SUM</td>'
        for sv in sum_vals:
            sum_cells += (f'<td class="num" style="font-weight:bold;background:{grp_clr};'
                          f'border-top:2px solid #999">{f"{sv:.1f}%" if sv is not None else ""}</td>')
        for r in runs_data:
            if r['name'] == ref_run['name']:
                continue
            sv = _get_sum(r, gi)
            if sv is not None and ref_sv is not None:
                delta = sv - ref_sv
                clr = '#c0392b' if delta > 0 else ('#27ae60' if delta < 0 else '#555')
                sum_cells += (f'<td class="num delta" style="color:{clr};font-weight:bold;'
                              f'border-top:2px solid #999">{delta:+.1f}%</td>')
            else:
                sum_cells += '<td class="num delta" style="border-top:2px solid #999">—</td>'
        rows_html += f'<tr>{sum_cells}</tr>\n'

    # Total yield loss row
    if best['data'] and best['data']['totals']:
        tot_lbl = best['data']['totals']['label']
        tot_vals = [_get_total(r) for r in runs_data]
        ref_tv = _get_total(ref_run)
        tot_cells = f'<td style="font-weight:bold;background:#d5e8d4;border-top:3px solid #555">{_esc(tot_lbl)}</td>'
        for tv in tot_vals:
            tot_cells += (f'<td class="num" style="font-weight:bold;background:#d5e8d4;'
                          f'border-top:3px solid #555">{f"{tv:.1f}%" if tv is not None else ""}</td>')
        for r in runs_data:
            if r['name'] == ref_run['name']:
                continue
            tv = _get_total(r)
            if tv is not None and ref_tv is not None:
                delta = tv - ref_tv
                clr = '#c0392b' if delta > 0 else ('#27ae60' if delta < 0 else '#555')
                tot_cells += (f'<td class="num delta" style="color:{clr};font-weight:bold;'
                              f'border-top:3px solid #555">{delta:+.1f}%</td>')
            else:
                tot_cells += '<td class="num delta" style="border-top:3px solid #555">—</td>'
        rows_html += f'<tr>{tot_cells}</tr>\n'

    return f'''
<div class="section">
  <div style="background:#1a252f;border-radius:6px 6px 0 0;padding:10px 16px;
              display:flex;align-items:center;flex-wrap:wrap;gap:8px">
    <span style="font-size:20px;font-weight:bold;color:#ecf0f1">
      &#128202; Detailed Comparison Table
    </span>
    <span style="font-size:15px;color:#aab7b8;margin-left:4px">
      reference:&nbsp;<b style="color:#3498db">{_esc(ref_run["name"])}</b>
      &nbsp;&nbsp;&#124;&nbsp;&nbsp;
      <span style="color:#e74c3c">&#9650; red = more fallout</span>
      &nbsp;&nbsp;
      <span style="color:#2ecc71">&#9660; green = less</span>
    </span>
  </div>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead><tr>{hdr_cells}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 6. Generate full HTML report
# ---------------------------------------------------------------------------

def _run_sort_key(run):
    """Sort key: file mtime of xlsx (most reliable), then ts string, then name."""
    xlsx = run.get('xlsx_path')
    if xlsx:
        try:
            return Path(xlsx).stat().st_mtime
        except Exception:
            pass
    ts = run.get('ts', '')
    # Try common date patterns: YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY, etc.
    import datetime
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y',
                '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %H:%M', '%m/%d/%Y'):
        try:
            return datetime.datetime.strptime(ts.strip(), fmt).timestamp()
        except Exception:
            pass
    return 0.0


def generate_report(runs_data, output_path: Path, ref_name: str = None, config_json: str = None, dash_dir: Path = None):
    # Preserve the order as given (matches Dashboard.html order)
    # runs_data = sorted(runs_data, key=_run_sort_key)  # removed: use Dashboard order

    valid = [r for r in runs_data if r['data']]
    if not valid:
        print('ERROR: No valid xlsx data found for any run.', file=sys.stderr)
        return

    # Load UPM target from config JSON
    _upm_target_pct = None
    _upm_target_label = 'Target'
    if config_json:
        try:
            import json as _json_cr
            _cfg_cr = _json_cr.loads(open(str(config_json), encoding='utf-8').read())
            for _ut in _cfg_cr.get('upm_target', []):
                if _ut.get('target_%') is not None:
                    _upm_target_pct = float(_ut['target_%'])
                    _upm_target_label = _ut.get('test', 'Target')
                    break
        except Exception:
            pass

    # Resolve the reference run name
    if ref_name is None:
        ref_name = valid[0]['name']
    elif not any(r['name'] == ref_name for r in runs_data):
        print(f'WARNING: --ref "{ref_name}" not found; using first run.')
        ref_name = valid[0]['name']

    # Determine number of data columns (Yield Loss + optional Recovery)
    n_cols = max(len(r['data']['col_headers']) for r in valid)
    col_headers = valid[0]['data']['col_headers']

    charts_html = ''

    # --- Yield Information chart ---
    rdnd_valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]

    # --- Extract UPM from raw CSV if Group_Medians.csv not available ---
    for r in runs_data:
        if not r.get('upm_data'):
            output_dir = None
            if r.get('xlsx_path') and str(r['xlsx_path']):
                output_dir = Path(str(r['xlsx_path'])).parent
            if not output_dir:
                href = r.get('index_href', '')
                if href:
                    import re as _re_tmp
                    href = _re_tmp.sub(r'^file:///', '', href).replace('/', os.sep)
                    idx = Path(href) if os.path.isabs(href) else (output_path.parent / href)
                    output_dir = idx.parent
            if output_dir:
                raw_csv = find_raw_csv(output_dir)
                if raw_csv:
                    upm_data, upm_detail = extract_upm_from_csv(raw_csv, config_json=config_json)
                    if upm_data:
                        r['upm_data'] = upm_data
                    if upm_detail:
                        r['upm_detail'] = upm_detail

    upm_valid  = [r for r in runs_data if r.get('upm_data') and r['upm_data']]

    # --- Run Summary table (always shown) ---
    charts_html += build_run_summary_table_html(runs_data, dash_dir=dash_dir)

    if rdnd_valid:
        combined_b64 = build_combined_rdnd_chart(rdnd_valid)
        if combined_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; Yield Information</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + combined_b64 + '"/></div>')

    # --- Yield Table ---
    if rdnd_valid:
        charts_html += build_rdnd_table_html(rdnd_valid)

    # --- Bin Fail Summary ---
    bf_valid = [r for r in runs_data if r.get('bin_data') and (
        r['bin_data'].get('bin_summary_rows') or r['bin_data'].get('bin_fail_rows'))]
    if bf_valid:
        charts_html += build_bin_fail_table_html(bf_valid)

    # --- Top-10 fail pareto ---
    top10_b64 = build_top10_pareto_chart(runs_data)
    if top10_b64:
        charts_html += ('<div class="section">'
                        '<h2>&#128202; Top 10 Interface Bin Fail Pareto</h2>'
                        '<img class="chart" src="data:image/png;base64,'
                        + top10_b64 + '"/></div>')

    # --- SICC/UPM charts + table ---
    if upm_valid:
        upm_b64 = build_upm_median_chart(upm_valid)
        if upm_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; SICC Median</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + upm_b64 + '"/></div>')
        pct_b64 = build_upm_pct_chart(upm_valid, upm_target_pct=_upm_target_pct,
                                         upm_target_label=_upm_target_label)
        if pct_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; UPM ULVT 950mV (%)</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + pct_b64 + '"/></div>')
        # Detailed per-column UPM comparison
        detail_b64 = build_upm_detail_chart(runs_data, upm_target_pct=_upm_target_pct,
                                            upm_target_label=_upm_target_label)
        if detail_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128200; UPM Distribution Comparison</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + detail_b64 + '"/></div>')
        sicc_tbl = build_sicc_table_html(upm_valid)
        if sicc_tbl:
            charts_html += sicc_tbl

    # --- CDYN charts + table ---
    cdyn_valid = [r for r in runs_data if r.get('cdyn_data') and r['cdyn_data']]
    if cdyn_valid:
        cdyn_b64 = build_cdyn_median_chart(cdyn_valid)
        if cdyn_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; CDYN Median</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + cdyn_b64 + '"/></div>')
        cdyn_tbl = build_cdyn_table_html(cdyn_valid)
        if cdyn_tbl:
            charts_html += cdyn_tbl

    # --- *_out.xlsx digital dashboard (at end) ---
    charts_html += build_xlsx_comparison_table(runs_data)

    # Detailed Comparison Table removed

    # Run summary cards — show all runs, including those with no xlsx
    card_html = ''
    for ri, run in enumerate(runs_data):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        if run.get('data'):
            total = run['data']['totals']['vals'][0] if run['data']['totals'] else None
            total_str = f'{total:.1f}%' if total is not None else '—'
            num_die = run['data']['num_die']
            die_str = f'{int(num_die):,}'
            # FF+DF yield (bin 1/2/3/4) and FF yield (bin 1/2)
            yrows = run['bin_data']['yield_rows'] if run.get('bin_data') else []
            ffdf_row = next((r for r in yrows if r['bin'] == '1/2/3/4'), None)
            ff_row   = next((r for r in yrows if r['bin'] == '1/2'), None)
            ffdf_str = f"{ffdf_row['yield_pct']:.1f}%" if ffdf_row and ffdf_row.get('yield_pct') is not None else '—'
            ff_str   = f"{ff_row['yield_pct']:.1f}%"  if ff_row  and ff_row.get('yield_pct')  is not None else '—'
        else:
            total_str = 'N/A'
            die_str   = 'N/A'
            ffdf_str  = 'N/A'
            ff_str    = 'N/A'
        card_html += f'''
<div class="run-card" style="border-left:4px solid {clr}">
  <div class="run-card-name" style="color:{clr}">{_esc(run["name"])}</div>
  <div class="run-card-ts">{_esc(run["ts"])}</div>
  <div class="run-card-stat">Die: <b>{die_str}</b></div>
  <div class="run-card-stat">FF + DF Yield: <b>{ffdf_str}</b></div>
  <div class="run-card-stat">FF Yield: <b>{ff_str}</b></div>
</div>'''

    html = f'''<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>Test Program Comparison Report</title>
<style>
.dash-link{{font-size:21px;color:#2980b9;margin-bottom:4px;display:block}}
.dash-link a{{color:#2980b9;text-decoration:none}}
.dash-link a:hover{{text-decoration:underline}}
h1{{font-size:33px;color:#2c3e50;margin-bottom:6px}}
h2{{font-size:26px;color:#2c3e50;margin:20px 0 8px;padding-bottom:4px;border-bottom:2px solid #dce1e7}}
h3{{font-size:23px;color:#555;margin:10px 0 4px}}
.subtitle{{font-size:21px;color:#7f8c8d;margin-bottom:18px}}
.ref-note{{font-size:20px;font-weight:normal;color:#7f8c8d;margin-left:8px}}
.section{{background:#fff;border-radius:8px;padding:18px;margin-bottom:20px;
  box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.cards{{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:20px}}
.run-card{{background:#fff;border-radius:6px;padding:12px 16px;min-width:200px;
  box-shadow:0 1px 4px rgba(0,0,0,.08);flex:1}}
.run-card-name{{font-size:23px;font-weight:bold;margin-bottom:4px}}
.run-card-ts{{font-size:18px;color:#95a5a6;margin-bottom:6px}}
.run-card-stat{{font-size:21px;margin-bottom:2px}}
.run-card-src{{font-size:17px;color:#aaa;margin-top:6px;word-break:break-all}}
.chart{{max-width:100%;height:auto;display:block;margin:8px 0}}
.cmp-tbl{{border-collapse:collapse;font-size:20px;width:auto}}
.cmp-tbl th{{background:#34495e;color:#ecf0f1;padding:6px 10px;text-align:left;
  white-space:nowrap;font-size:18px}}
.cmp-tbl thead tr:nth-child(2) th{{text-align:right;font-size:13px;font-weight:normal;
  background:#4a5568;color:#e2e8f0;padding:3px 10px}}
.cmp-tbl td{{padding:4px 10px;border-bottom:1px solid #eee;white-space:nowrap}}
.cmp-tbl tr:hover td{{background:#f9f9f9}}
td.num{{text-align:right;font-variant-numeric:tabular-nums}}
td.delta{{text-align:right;font-variant-numeric:tabular-nums}}
th.delta-hdr{{background:#4a235a;color:#e8daef;text-align:center}}
</style>
</head>
<body>
<h1>&#128200; Test Program Comparison Report</h1>
<div class="dash-link">Source: <a href="Dashboard.html">Dashboard.html</a></div>
<div class="subtitle">
  Runs: <b>{len(runs_data)}</b> &nbsp;|&nbsp; With data: <b>{len(valid)}</b>
  &nbsp;|&nbsp; Reference: <b>{_esc(ref_name)}</b>
</div>
<div class="cards">{card_html}</div>
{charts_html}
</body>
</html>'''

    if output_path:
        output_path.write_text(_wm_inject(html), encoding='utf-8')
        print(f'Wrote comparison report: {output_path}')
    return html



# ---------------------------------------------------------------------------
# 7. Scan for compare HTML files and update links block in Dashboard.html
# ---------------------------------------------------------------------------

_CMP_PATTERNS = ('compare_*.html', 'compareTP*.html', '*_compare.html',
                 '*_comparison.html', 'compare-*.html')


def find_compare_files(dash_dir: Path) -> list[Path]:
    """Return all compare-report HTML files next to Dashboard.html, sorted by name.
    Case-insensitive: scans all .html files and matches any whose stem contains 'compare'."""
    name_re = re.compile(r'compare', re.IGNORECASE)
    results = []
    for p in sorted(dash_dir.glob('*.html')):
        if p.name.lower() == 'dashboard.html':
            continue
        if name_re.search(p.stem):
            results.append(p)
    return results


def update_dashboard_compare_links(dash_path: Path) -> None:
    """
    Scan the dashboard directory for compare HTML files and rewrite the
    <!-- COMPARISON_REPORT_START/END --> block in Dashboard.html with links.
    Safe to call after any compareTP / compare_runs run.
    """
    dash_dir = dash_path.parent
    files = find_compare_files(dash_dir)

    if not files:
        # Nothing to show — remove the block if present
        dash_html = dash_path.read_text(encoding='utf-8')
        dash_html = re.sub(
            r'\n*<!-- COMPARISON_REPORT_START -->[\s\S]*?<!-- COMPARISON_REPORT_END -->\n*',
            '', dash_html
        )
        dash_path.write_text(dash_html, encoding='utf-8')
        print('No compare files found — removed comparison block from Dashboard.html.')
        return

    link_items = ''.join(
        f'<a class="run-link report-link" href="{p.name}" target="_blank">'
        f'&#128200; {p.stem}</a>\n'
        for p in files
    )

    inject = (
        '\n\n<!-- COMPARISON_REPORT_START -->\n'
        '<div id="_cmp_embed" style="margin-top:18px;padding:10px 14px;background:#2c3e50;'
        'border-radius:6px;display:flex;flex-wrap:wrap;gap:6px;align-items:center">\n'
        '<span style="color:#ecf0f1;font-size:13px;font-weight:bold;margin-right:4px">'
        '&#128200; Comparison Reports:</span>\n'
        + link_items +
        '</div>\n'
        '<!-- COMPARISON_REPORT_END -->\n'
    )

    dash_html = dash_path.read_text(encoding='utf-8')

    # Remove any previous comparison block
    dash_html = re.sub(
        r'\n*<!-- COMPARISON_REPORT_START -->[\s\S]*?<!-- COMPARISON_REPORT_END -->\n*',
        '', dash_html
    )

    # Add ▼ Compare jump link to <h1> once
    jump = (' <a href="#_cmp_embed" style="font-size:15px;color:#2980b9;'
            'text-decoration:none;margin-left:14px;vertical-align:middle">'
            '&#9660; Compare</a>')
    if '#_cmp_embed' not in dash_html:
        dash_html = re.sub(
            r'(<h1[^>]*>)(.*?)(</h1>)',
            lambda m: m.group(1) + m.group(2) + jump + m.group(3),
            dash_html, count=1, flags=re.DOTALL
        )

    # Insert just before <!-- RUNS_END -->
    if '<!-- RUNS_END -->' in dash_html:
        dash_html = dash_html.replace('<!-- RUNS_END -->', inject + '<!-- RUNS_END -->', 1)
    elif '</body>' in dash_html:
        dash_html = dash_html.replace('</body>', inject + '</body>', 1)
    else:
        dash_html += inject

    dash_path.write_text(dash_html, encoding='utf-8')
    print(f'Updated {dash_path.name} with {len(files)} comparison link(s).')


# ---------------------------------------------------------------------------
# 8. Main
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(description='Compare yield runs from Dashboard.html')
    p.add_argument('dashboard', help='Path to Dashboard.html')
    p.add_argument('--out', default='', help='Output HTML path (default: next to Dashboard.html)')
    p.add_argument('--ref', default='', help='Reference run name for comparison (default: first run)')
    args = p.parse_args()

    if not HAVE_OPENPYXL:
        print('ERROR: openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
        sys.exit(1)
    if not HAVE_MPL:
        print('WARNING: matplotlib not installed — charts will be skipped.')

    dash_path = Path(args.dashboard).resolve()
    if not dash_path.exists():
        print(f'ERROR: Dashboard.html not found: {dash_path}', file=sys.stderr)
        sys.exit(1)

    dash_dir = dash_path.parent

    print(f'Parsing {dash_path} ...')
    run_records = parse_dashboard(dash_path)
    if not run_records:
        print('ERROR: No run blocks found in Dashboard.html', file=sys.stderr)
        sys.exit(1)

    print(f'Found {len(run_records)} run(s): {[r["name"] for r in run_records]}')

    # Load xlsx + BinDistribution HTML for each run
    runs_data = []
    for rec in run_records:
        xlsx_p = find_xlsx(dash_dir, rec['index_href'])
        data = None
        if xlsx_p:
            print(f'  [{rec["name"]}] Reading {xlsx_p.name} ...')
            data = read_xlsx(xlsx_p)
            if not data:
                print(f'    WARNING: Could not parse {xlsx_p}')
            output_dir = xlsx_p.parent
        else:
            print(f'  [{rec["name"]}] WARNING: *_out.xlsx not found')
            # Try to resolve output_dir from index_href even without xlsx
            href = re.sub(r'^file:///', '', rec['index_href'] or '').replace('/', os.sep)
            idx_path = dash_dir / href if not os.path.isabs(href) else Path(href)
            output_dir = idx_path.parent if idx_path else None

        bin_data = None
        upm_data = None
        cdyn_data = None
        if output_dir and output_dir.exists():
            bin_p = find_bin_html(output_dir)
            if bin_p:
                print(f'  [{rec["name"]}] Reading {bin_p.name} ...')
                bin_data = parse_bin_html(bin_p)
            else:
                print(f'  [{rec["name"]}] WARNING: *_BinDistribution.html not found')
            gm_p = find_group_medians(output_dir)
            if gm_p:
                print(f'  [{rec["name"]}] Reading {gm_p.name} ...')
                upm_data = parse_group_medians(gm_p)
            cdyn_p = find_cdyn_medians(output_dir)
            if cdyn_p:
                print(f'  [{rec["name"]}] Reading {cdyn_p.name} ...')
                cdyn_data = parse_cdyn_medians(cdyn_p)

        runs_data.append({**rec, 'data': data, 'xlsx_path': xlsx_p or '',
                          'bin_data': bin_data, 'upm_data': upm_data,
                          'cdyn_data': cdyn_data})

    ref_name = args.ref.strip() or None
    out_path = Path(args.out).resolve() if args.out else dash_dir / 'compare_report.html'

    print('Generating comparison report ...')
    # Find Product Config JSON in collateral/ folder
    _cfg_json = None
    _collateral = dash_dir / 'collateral'
    if _collateral.exists():
        _cfgs = sorted(_collateral.glob('Product Config*.json'),
                       key=lambda p: p.stat().st_mtime, reverse=True)
        if _cfgs:
            _cfg_json = str(_cfgs[0])
            print(f'  Config: {_cfgs[0].name}')
    generate_report(runs_data, out_path, ref_name=ref_name, config_json=_cfg_json,
                    dash_dir=dash_dir)

    # Update comparison links in Dashboard.html
    update_dashboard_compare_links(dash_path, out_path)

    # Open report in browser
    try:
        os.startfile(str(out_path))
    except Exception:
        pass


def update_dashboard_compare_links(dash_path: Path, compare_report_path: Path):
    """Inject a link to compare_report_path into the <!-- COMPARE_START/END --> section
    of Dashboard.html, using the same run-block style as the Yield section."""
    from datetime import datetime as _dt
    dash_path = Path(dash_path)
    compare_report_path = Path(compare_report_path)
    if not dash_path.exists():
        return

    content = dash_path.read_text(encoding='utf-8')

    COMPARE_START = '<!-- COMPARE_START -->'
    COMPARE_END   = '<!-- COMPARE_END -->'
    YIELD_END     = '<!-- YIELD_END -->'

    # Ensure COMPARE section exists
    if COMPARE_START not in content:
        if YIELD_END in content:
            content = content.replace(
                YIELD_END,
                YIELD_END + '\n<h2 class="section-header">&#128200; Compare TP</h2>\n'
                + COMPARE_START + '\n' + COMPARE_END)
        else:
            # No sentinels at all — append before </body>
            content = content.replace(
                '</body>',
                '<h2 class="section-header">&#128200; Compare TP</h2>\n'
                + COMPARE_START + '\n' + COMPARE_END + '\n</body>')

    # Build relative href from dash_dir to compare report
    try:
        href = os.path.relpath(str(compare_report_path), str(dash_path.parent)).replace('\\', '/')
    except Exception:
        href = compare_report_path.as_uri()

    report_stem = compare_report_path.stem
    ts = _dt.now().strftime('%Y-%m-%d %H:%M')

    # Replace existing block with same stem, or prepend new one
    block_key = report_stem
    new_block = (
        f'<div class="run-block" data-stem="{block_key}">\n'
        f'<div class="run-header" onclick="toggle(this)">'
        f'<span class="arrow">&#9660;</span> {report_stem}'
        f'<span class="ts"> - {ts}</span></div>\n'
        f'<div class="run-body">\n'
        f'<a class="run-link report-link" href="{href}" target="_blank">{report_stem}</a>\n'
        f'</div>\n</div>'
    )

    block_re = re.compile(
        r'<div class="run-block" data-stem="' + re.escape(block_key) +
        r'">\s*<div[^>]*>[\s\S]*?</div>\s*</div>', re.MULTILINE)
    if block_re.search(content):
        content = block_re.sub(new_block, content)
    else:
        content = content.replace(COMPARE_START, COMPARE_START + '\n' + new_block)

    dash_path.write_text(content, encoding='utf-8')


if __name__ == '__main__':
    main()
