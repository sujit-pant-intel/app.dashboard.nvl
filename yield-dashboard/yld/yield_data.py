"""yield_data.py — Merged data utilities and reports.

Consolidates: csv_utils, get_dd_update, add_material_type,
  apply_reticle_mapping, parse_bindef_to_crystalball,
  make_portable_dashboard, manage_dashboard, yield_report, yield_report_frame

Public API (unchanged — all original functions available at module level).
"""
from __future__ import annotations

# ════════════════════════════════════════════════════════════════
# (formerly csv_utils.py)
# ════════════════════════════════════════════════════════════════
# =============================================================================
# csv_utils.py  -  Large-file CSV helpers shared across the yield pipeline
# =============================================================================
# Provides:
#   CHUNK_SIZE          default rows per chunk (100 000)
#   detect_encoding()   try encodings in order, return first that works
#   sniff_columns()     read only the header row; return column-name list
#   read_csv_smart()    read with optional usecols (column selection)
#   iter_chunks()       generator that yields DataFrames in CHUNK_SIZE slices
#
# All functions accept an optional encoding= argument; when omitted the
# encoding is auto-detected via detect_encoding().
# =============================================================================


import io
import os
import zipfile
from pathlib import Path
from typing import Generator, Iterable

import pandas as pd

# Default number of rows loaded into RAM at a time for streaming operations.
# Callers can override per-call.  Adjust with env var CSV_CHUNK_SIZE for
# system-wide tuning without code changes.
CHUNK_SIZE: int = int(os.environ.get('CSV_CHUNK_SIZE', 100_000))

_ENCODINGS = ('utf-8-sig', 'utf-8', 'utf-16', 'latin-1')


def _resolve_csv_from_path(path: Path) -> tuple[Path | None, bytes | None]:
    """If *path* is a .zip or .7z, extract the first CSV inside and return its bytes.

    Returns ``(None, bytes)`` for archives, ``(path, None)`` for plain CSV/GZ.
    """
    if path.suffix.lower() == '.zip':
        with zipfile.ZipFile(path) as zf:
            csvs = [n for n in zf.namelist() if n.lower().endswith('.csv') and not os.path.basename(n).startswith('.')]
            if not csvs:
                # Fallback: any non-directory entry (pandas reads whatever is first)
                csvs = [n for n in zf.namelist()
                        if not n.endswith('/') and not os.path.basename(n).startswith('.')]
            if not csvs:
                raise ValueError(f'No CSV found inside zip: {path}')
            return None, zf.read(csvs[0])
    if path.suffix.lower() == '.7z':
        import py7zr
        with py7zr.SevenZipFile(path, mode='r') as sz:
            names = sz.getnames()
            csvs = [n for n in names if n.lower().endswith('.csv') and not os.path.basename(n).startswith('.')]
            if not csvs:
                csvs = [n for n in names if not os.path.basename(n).startswith('.')]
            if not csvs:
                raise ValueError(f'No CSV found inside 7z: {path}')
            extracted = sz.read(targets=[csvs[0]])
            return None, extracted[csvs[0]].read()
    return path, None


# ---------------------------------------------------------------------------
# Encoding detection
# ---------------------------------------------------------------------------

def detect_encoding(path: str | Path) -> str | None:
    """Return the first encoding that successfully reads the file header.

    Tries ``utf-8-sig``, ``utf-8``, ``utf-16``, ``latin-1`` in that order.
    Falls back to ``latin-1`` (which never raises a decode error).
    Returns ``None`` for ``.gz`` files — pandas infers compression and encoding
    automatically, so no pre-detection is needed.
    """
    path = Path(path)
    if path.suffix.lower() in ('.gz', '.7z'):
        return None   # handled via _resolve_csv_from_path or pandas auto-compression
    for enc in _ENCODINGS:
        try:
            with open(path, encoding=enc, errors='strict') as fh:
                fh.readline()   # only need to parse one line
            return enc
        except (UnicodeDecodeError, Exception):
            continue
    return 'latin-1'


# ---------------------------------------------------------------------------
# Header-only sniff
# ---------------------------------------------------------------------------

def sniff_columns(path: str | Path, encoding: str | None = None) -> list[str]:
    """Return the list of column names without loading any data rows.

    Peak RAM is proportional to the header row length only.
    Transparently handles .zip files containing a CSV.
    """
    path = Path(path)
    resolved, data = _resolve_csv_from_path(path)
    if data is not None:
        try:
            df_header = pd.read_csv(io.BytesIO(data), nrows=0, low_memory=False)
            return list(df_header.columns)
        except Exception:
            return []
    enc = encoding or detect_encoding(resolved)
    try:
        df_header = pd.read_csv(resolved, nrows=0, encoding=enc, low_memory=False)
        return list(df_header.columns)
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Smart full-load (column selection, no chunking)
# ---------------------------------------------------------------------------

def read_csv_smart(
    path: str | Path,
    usecols: list[str] | None = None,
    encoding: str | None = None,
) -> pd.DataFrame:
    """Load a CSV into a single DataFrame with optional column selection.

    Parameters
    ----------
    path:
        CSV file (or .zip containing a CSV) to read.
    usecols:
        Subset of columns to load.  Columns not present in the file are
        silently ignored so callers can pass a superset.
    encoding:
        File encoding.  Auto-detected when omitted.  Ignored for zip files
        (pandas detects encoding from the bytes stream).
    """
    path = Path(path)
    resolved, data = _resolve_csv_from_path(path)
    if data is not None:
        # zip path — read from in-memory bytes
        effective_usecols: list[str] | None = None
        if usecols is not None:
            all_cols = list(pd.read_csv(io.BytesIO(data), nrows=0, low_memory=False).columns)
            effective_usecols = [c for c in usecols if c in all_cols] or None
        return pd.read_csv(io.BytesIO(data), usecols=effective_usecols, low_memory=False)

    enc = encoding or detect_encoding(resolved)

    # Intersect requested columns with those actually in the file
    effective_usecols = None
    if usecols is not None:
        all_cols = sniff_columns(resolved, encoding=enc)
        effective_usecols = [c for c in usecols if c in all_cols] or None

    return pd.read_csv(
        resolved,
        usecols=effective_usecols,
        encoding=enc,
        low_memory=False,
    )


# ---------------------------------------------------------------------------
# Chunked iterator
# ---------------------------------------------------------------------------

def iter_chunks(
    path: str | Path,
    usecols: list[str] | None = None,
    chunksize: int = CHUNK_SIZE,
    encoding: str | None = None,
) -> Generator[pd.DataFrame, None, None]:
    """Yield successive DataFrames of at most *chunksize* rows.

    Each chunk contains only the columns listed in *usecols* (after
    intersecting with the actual column names in the file).

    Parameters
    ----------
    path:
        CSV file to read.
    usecols:
        Columns to include in every chunk.  Pass ``None`` to keep all.
    chunksize:
        Maximum rows per yielded DataFrame.
    encoding:
        File encoding.  Auto-detected when omitted.
    """
    path = Path(path)
    enc = encoding or detect_encoding(path)

    effective_usecols: list[str] | None = None
    if usecols is not None:
        all_cols = sniff_columns(path, encoding=enc)
        effective_usecols = [c for c in usecols if c in all_cols] or None

    reader = pd.read_csv(
        path,
        usecols=effective_usecols,
        encoding=enc,
        chunksize=chunksize,
        low_memory=False,
    )
    for chunk in reader:
        yield chunk


# ════════════════════════════════════════════════════════════════
# (formerly get_dd_update.py)
# ════════════════════════════════════════════════════════════════
#!/usr/intel/pkgs/python3/3.7.4/bin/python3
"""
Get Digital Dashboard Update: Takes a .csv input file from Crystal Ball with FBINs and
LOGTRACKER results to determine two columns for the Digital Dashboard containing the yield
summary by sub module.

Usage:
    get_dd_update.py -h --help
    get_dd_update.py -v --version
    get_dd_update.py [--data=data.csv] [--bin_defs=bin_defs.csv] [--vmax] [--wafer]
        [--log=log_file.log] [--verbose] [--debug]

Options:
    -h --help                       Show this screen.
    -v --version                    Show version.
    -d --data=<input file>          .csv Input file containing a list of die and FBINs.
    -b --bin_defs=<input file>      .csv file containing a table of DBIN and Test name.
    -g --log=<log file>             Optionally, log to a file.
    -m --vmax                       Make the Vmax summary as well.
    -w --wafer                      Print columns per wafer.
    --verbose                       Print more information than usual.
    -x --debug                      Print lots of debugging statements.
    --dashboard=<file>              Optional path to DigitalDashBoard.xlsx to append results
"""
import sys

try:
    from docopt import docopt
except ModuleNotFoundError:
    docopt = None
try:
    import pandas
except ModuleNotFoundError:
    print("Python 'pandas' module not installed. Install with pip:")
    print("  python -m pip install --user --proxy \"http://proxy-us.intel.com:911\" pandas openpyxl")
    sys.exit()
import logging
import traceback
import datetime
import pathlib
from pathlib import Path
import re
import os
import shutil
import zlib
import time
import csv
import tempfile
# Ensure user site-packages are on sys.path (needed when launched by double-click)
import site as _site, sys as _sys
_usp = _site.getusersitepackages()
if _usp not in _sys.path:
    _sys.path.insert(0, _usp)
try:
    import UsrIntel.R2  #Required for openpyxl below.
except ModuleNotFoundError:
    try:
        import UsrIntel.R1
    except ModuleNotFoundError:
        pass  # optional on non-Intel environments
try:
    import openpyxl
    HAVE_OPENPYXL = True
except Exception:
    openpyxl = None
    HAVE_OPENPYXL = False


__version__ = "get_dd_update 0.0"


#todo: If this gets out of hand, move it to a json file.
moduleMap = {
    #Good bins
    "B198_PASS": {"dd": "Bin 198 (Vmin Repair)", "vmax":"Bin 198 (Vmin Repair)"},
    r"B1((?!98)\d\d)_PASS": {"dd": "Bin 1", "vmax":"Bin 1 (No Repair)"},
    "B201_PASS": {"dd": "Bin 2 (Hard Repair)", "vmax":"Bin 2 (Hard Repair)"},
    "B202_PASS": {"dd": "Bin 202 (Vmax Repair)", "vmax":"Bin 202 (Vmax Repair)"},
    "B226_PASS": {"dd": "Bin 2 (Hard Repair)", "vmax":"Bin 2 (Hard Repair)"},
    r"B3\d\d_PASS": {"dd": "Bin 3", "vmax":"Bin 3"},
    r"B4\d\d_PASS": {"dd": "Bin 4", "vmax":"Bin 4"},
    #RESET
    r"B19\d{6}_FAIL_": {"dd": "Reset", "vmax": "Reset"},
    r"B35\d{6}_FAIL_DRV_RESET": {"dd": "Reset", "vmax": "Reset"},
    #ARR ATOM
    r"ARR_ATOM_.*_VNOM_LFM_0800": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"ARR_ATOM_.*_VATOM_.*_F1_0800": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"ARR_ATOM_.*_NOM_LFM": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"ARR_ATOM_.*_VNOM_LFM": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"ARR_ATOM_.*_VMIN_LFM_0800": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"B6326\d{4}_FAIL_HVQK": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_LFM (B2050,B6050)"},
    r"ARR_ATOM_.*_VMIN_.*_VATOM_.*_F6": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VNOM_TFM (B2050,B6050)"},
    r"ARR_ATOM_.*_F1_.*_MAX": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VMAX_LFM (B6050)"},
    r"ARR_ATOM_.*MAX_LFM_.": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VMAX_LFM (B6050)"},
    r"ARR_ATOM_.*_(F5|F6)_.*_MAX": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VMAX_TFM (B6050)"},
    r"ARR_ATOM_.*_MAX_TFM_": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VMAX_TFM (B6050)"},
    r"B6050\d{4}_FAIL_ARR_ATOM_.*_3200_.*": {"dd": "ARR_ATOM", "vmax": "ARR_ATOM_VMAX_TFM (B6050)"},
    #ARR_CCF
    r"ARR_CCF_.*_VNOM_.*_F1_": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"B6244\d{4}_.*_VMIN_.*_F1_": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"B6226\d{4}": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"ARR_CCF_.*_VNNAON_.*_F1_": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"ARR_CCF.*_VNOM_LFM_": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"ARR_CCF_.*_(F1|LFM)_.*_800MV_": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"ARR_CCF_.*_F1_.*_MAX": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VMAX_LFM (B6242)"},
    r"ARR_CCF_.*_F6_.*_MAX": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VMAX_TFM (B2043)"},
    r"ARR_CCF_.*_VMIN_.*_(F5|F6)_.*": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_TFM (B62)"},
    r"ARR_CCF_.*_VCCR_.*_(F5|F6)_.*": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_TFM (B62)"},
    r"ARR_CCF_.*_VMIN_.*_F1_.*": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    r"ARR_CCF_.*_VCCR_.*_(?:F1|FMIN)_.*": {"dd": "ARR_CCF", "vmax": "ARR_CCF_VNOM_LFM (B20, B33, B62)"},
    #ARR_NONCCF
    r"ARR_MBIST_.*_ALL_NONCCF_.*_MAX": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VMAX_LFM (B61)"},
    r"ARR_MBIST_.*_ALL_NONCCF_.*_KS_\d$": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"B6126\d{4}_FAIL_HVQK": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"ARR_MBIST_.*_ALL_NONCCF_.*_RETENTION": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"ARR_MBIST_.*_(L|S)SA_NONCCF_.*_PREHVQK": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"ARR_MBIST_.*_ROM_NONCCF_.*_PREHVQK": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"ARR_UNCORE.*_VNNAON(_NOM|_X_X_X)?_LFM": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    r"ARR_UNCORE.*_VNNAON(_MAX|_X_X_X)?_LFM": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VMAX_LFM (B21,B60,B61)"},
    #ARR_CORE
    r"ARR_MBIST_.*_ALL_CORE_.*_VMAX_.*_F1_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VMAX_LFM (B60)"},
    r"ARR_CORE.*_MAX_.*_F1_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VMAX_LFM (B60)"},
    r"B6017\d{4}_FAIL_ARR_MBIST_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B60)"},
    r"B6019\d{4}_FAIL_ARR_MBIST_.*_F1_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B60)"},
    r"_FAIL_ARR_CORE_.*_F5_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VMAX_TFM (B60)"},
    r"^(?!.*_VMAX_).*ARR_MBIST_.*_ALL_CORE_.*_F[456]_$": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_TFM (B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_MBIST_.*_ALL_CORE_.*_F[456]_\d{4}_\d$": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_TFM (B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_MBIST_.*_CORE_.*_PREHVQK": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_)B2000\d{4}.*ARR_MBIST_.*_CORE_.*_END_.*": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_MBIST_.*_CORE_.*_END_.*_F1_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_CORE.*_CORE_.*_END_.*_LFM_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_CORE.*_CORE_.*_NOM_(?:LFM|FMIN)_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"^(?!.*_VMAX_).*ARR_MBIST_.*_CORE_.*_EXVF_.*_F1_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"}, #must not contain VMAX
    r"B6026\d{4}": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"},
    r"B2026\d{4}": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VNOM_LFM (B20, B60)"},
    r"ARR_MBIST_.*_ALL_CORE_.*_VMAX_.*_F6_": {"dd": "ARR_CORE", "vmax":"ARR_CORE_VMAX_TFM (B60)"}, #must not contain VMAX
    "FAIL_UNCORRECTABLE_ECC_ERROR": {"dd": "ARR_NONCCF", "vmax": "ARR_NONCCF_VNOM_LFM (B21,B60,B61)"},
    #FUN_ATOM
    r"B44\d{6}_FAIL_FUN_ATOM_.*_V(MIN|ATOM)_.*_F(0|1)_((?!DRAGON_SLC).)*$":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_VNOM_LFM (B445x)"},
    r"B44\d{6}_FAIL_FUN_ATOM_.*_V(MIN|ATOM)_.*_F(0|1)_.*_DRAGON_SLC":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_DRAGON_SLC_VNOM_LFM (B4463)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_F6_((?!DRAGON_SLC).)*$":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_VNOM_TFM (B446x)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_TFM_4400_ATOM_L2_DRAGON":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_VNOM_TFM (B446x)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_VATOM_.*_(?:XFM|TFM)_":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_VNOM_TFM (B446x)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_TFM_ATOM_L2":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_VNOM_TFM (B44xx)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_F6_.*_DRAGON_SLC":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_DRAGON_SLC_VNOM_TFM (B4465)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_TFM_3800_.*_DRAGON_SLC":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_DRAGON_SLC_VNOM_TFM (B4465)"},
    r"B44\d{6}_FAIL_(FUN|SBFT)_ATOM_.*_V(MIN|ATOM)_.*_TFM_.*SLC_DRAGON":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_DRAGON_SLC_VNOM_TFM (B4465)"},
    r"B44\d{6}_FAIL_FUN_ATOM_.*_DRAGON_SLC_SPECKLE":
        {"dd": "FUN_ATOM", "vmax": "FUN_ATOM_DRAGON_SLC_SPECKLE (B4466)"},
    #FUN_CCF
    r"B45\d{6}_FAIL_FUN_CCF_.*_V(MIN|NOM).*_LFM_": {"dd": "FUN_CCF", "vmax": "FUN_CCF_VNOM_LFM (B4538)"},
    r"B45\d{6}_FAIL_FUN_CCF_.*_NOM.*_LFM_": {"dd": "FUN_CCF", "vmax": "FUN_CCF_VNOM_LFM (B4538)"},
    r"B45\d{6}_FAIL_FUN_CCF_.*_V(MIN|NOM)_.*_F6_": {"dd": "FUN_CCF", "vmax": "FUN_CCF_VNOM_TFM (B4538)"},
    r"B45\d{6}_FAIL_FUN_CCF_.*_V(MIN|NOM)_.*_TFM_": {"dd": "FUN_CCF", "vmax": "FUN_CCF_VNOM_TFM (B4538)"},
    r"B4526\d{4}": {"dd": "FUN_CCF", "vmax": "FUN_CCF_VNOM_LFM (B4538)"},
    #FUN_CORE
    r"B4426\d{4}": {"dd": "FUN_CORE", "vmax": "FUN_CORE_VNOM_LFM (B44)"},
    r"B44.*_SBFT_CORE_.*_VMIN_LFM_": {"dd": "FUN_CORE", "vmax": "FUN_CORE_VNOM_LFM (B44)"},
    r"B44.*_SBFT_CORE_VMIN_.*_LFM_": {"dd": "FUN_CORE", "vmax": "FUN_CORE_VNOM_LFM (B44)"},
    r"B44.*_SBFT_CORE_.*_VMAX_LFM_": {"dd": "FUN_CORE", "vmax": "FUN_CORE_VMAX_LFM (B44)"},
    r"B44.*_SBFT_CORE.*_VMIN_.*_(?:TFM|F6)_": {"dd": "FUN_CORE", "vmax": "FUN_CORE_VNOM_TFM (B44)"},
    #SCN_ATOM
    r"SCN_ATOM_.*_VNOM_LFM": {"dd": "SCN_ATOM", "vmax": "SCN_ATOM_VNOM_LFM (B41,B42,B47)"},
    r"SCN_ATOM_.*_VATOM_NOM_LFM": {"dd": "SCN_ATOM", "vmax": "SCN_ATOM_VNOM_LFM (B41,B42,B47)"},
    r"SCN_ATOM_.*_VNOM_TFM": {"dd": "SCN_ATOM", "vmax": "SCN_ATOM_VNOM_TFM (B41,B42,B47)"},
    r"SCN_ATOM_.*_VATOM_NOM_TFM": {"dd": "SCN_ATOM", "vmax": "SCN_ATOM_VNOM_TFM (B41,B42,B47)"},
    #SCN_CCF
    r"SCN_UNCORE_.*_CCF_.*_LFM": {"dd": "SCN_CCF", "vmax":"SCN_CCF_VNOM_LFM (B41,B42,B47)"},
    r"SCN_UNCORE_.*_SEC_.*_LFM": {"dd": "SCN_CCF", "vmax":"SCN_CCF_VNOM_LFM (B41,B42,B47)"},
    r"SCN_UNCORE_.*_HRY_.*_LFM": {"dd": "SCN_CCF", "vmax":"SCN_CCF_VNOM_LFM (B41,B42,B47)"},
    r"SCN_UNCORE_.*_VCCR_.*_LFM": {"dd": "SCN_CCF", "vmax":"SCN_CCF_VNOM_LFM (B41,B42,B47)"},
    r"SCN_UNCORE_.*_(?:CCF|UNCORE)_.*_TFM": {"dd": "SCN_CCF", "vmax":"SCN_CCF_VNOM_TFM (B41,B42,B47)"},
    #SCN_CORE
    r"SCN_CORE_.*_LFM": {"dd": "SCN_CORE", "vmax":"SCN_CORE_VNOM_LFM (B41,B42,B47)"},
    r"SCN_CORE_.*_TFM": {"dd": "SCN_CORE", "vmax":"SCN_CORE_VNOM_TFM (B41,B42,B47)"},
    r"B4(1|2)26\d{4}": {"dd": "SCN_CORE", "vmax":"SCN_CORE_VNOM_LFM (B41,B42,B47)"},
    #SCN_UNCORE
    r"SCN_UNCORE_.*_NONCCF": {"dd": "SCN_NONCCF", "vmax":"SCN_NONCCF_VNOM_LFM (B41,B42,B47)"},
    r"B4726\d{4}_FAIL_HVQK": {"dd": "SCN_NONCCF", "vmax":"SCN_NONCCF_VNOM_LFM (B41,B42,B47)"},
    #Analog
    r"^B24\d{6}": {"dd":"Analog", "vmax": "Analog Other (B24)"},
    r"^B27\d{6}_FAIL_PTH_BG": {"dd":"Analog", "vmax": "Analog PTH BG (B27)"},
    r"^B27\d{6}_FAIL_PTH_DLVR": {"dd":"Analog", "vmax": "Analog PTH DLVR (B27)"},
    r"^B27\d{6}_FAIL_HVQK": {"dd":"Analog", "vmax": "Analog PTH DLVR (B27)"},
    r"^B28\d{6}": {"dd":"Analog", "vmax": "Analog CLK (B28)"},
    r"^B29\d{6}": {"dd":"Analog", "vmax": "Analog PTH BGR (B29)"},
    r"^B36\d{6}": {"dd":"Analog", "vmax": "Analog MIO D2D (B36)"},
    r"^B40\d{6}": {"dd":"Analog", "vmax": "Analog PTH ODI (B40)"},
    r"^B64\d{6}": {"dd":"Analog", "vmax": "Analog PTH DTS (B64)"},
    #TPI_FOUNDRY
    "TPI_ADTL": {"dd": "TPI Foundry", "vmax": "TPI Foundry ADTL (B43)"},
    r"B4326\d{4}": {"dd": "TPI Foundry", "vmax": "TPI Foundry ADTL (B43)"},
    r"B8\d\d_": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B08\d\d_": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B89\d{6}_": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B80\d{6}_": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B10\d{6}_FAIL_TPI": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B15\d{6}_FAIL_TPI": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B40\d{6}_FAIL_PTH": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B4026\d{4}_FAIL_HVQK": {"dd": "TPI Foundry", "vmax": "TPI Foundry"},
    r"B18\d{6}_FAIL_PTH_POWER_.*SICC": {"dd": "TPI Foundry", "vmax": "TPI Foundry SICC (B18)"},
    #TPI_OTHER
    r"^B17\d{6}": {"dd":"TPI Other", "vmax":"TPI Other"},
    r"^B26\d{6}": {"dd":"HVQK (26)", "vmax":"HVQK (26)"},
    r"^B30\d{6}": {"dd":"TPI Other", "vmax": "TPI Other"},
    r"^B31\d{6}": {"dd":"TPI Other", "vmax": "TPI Other"},
    r"^B49\d{6}": {"dd":"TPI Other", "vmax": "TPI Other"},
    r"^B53\d{6}": {"dd":"TPI Other", "vmax": "TPI Other"},
    r"^B69\d{6}": {"dd":"TPI Other", "vmax": "TPI Other"},
    r"B88\d{6}_FAIL_TPI_SIU": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"B94\d{6}_": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"FAIL_DUT_TEMPERATURE": {"dd": "TPI Other", "vmax": "TPI Other"},
    "FAIL_DPS_OVERVOLTAGE_ALARM": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"FAIL_HALT_ALARM": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"B97\d{6}_": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"B98\d{6}_.*": {"dd": "TPI Other - B98", "vmax": "TPI Other"},
    r"B9099\d{4}.*_ALARM": {"dd": "TPI Other - B99", "vmax": "TPI Other"}, #IB is B99
    r"B90(?!99)\d{6}_FAIL_": {"dd": "TPI Other", "vmax": "TPI Other"},
    r"B99\d\d.*": {"dd": "TPI Other - B99", "vmax": "TPI Other"},
    "B93": {"dd": "TPI Other - B93", "vmax": "TPI Other"},
}


def getDD(dataInFile="", binDefFile="", vmax=False, logBaseName="",
          waferLvl=False, idt=2, dashboardFile="", outDir=""):
    productInfo = {
        "ARL68-N3B": {
            "DPW": 797,
            "DEVREVSTEP": ["8PYJCVJ"],
            "TPrgx": r"(E6\w)",
            "dfBins": [3,4],
            "numCores":6,
            "numAtoms": 8,
        },
        "ARLS816": {
            "DPW": 516,
            "DEVREVSTEP": ["8PYVCVB","8PYVCVAB"],
            "TPrgx": r"(8[2,3]\w)",
            "dfBins": [3,4],
            "numCores": 8,
            "numAtoms": 16,
        },
        "NVL48": {
            "DPW": 1200,
            "DEVREVSTEP": ["8PY6CVT"],
            "TPrgx": r"(8[1,2]\w)",
            "dfBins": [3, 4],
            "numCores": 4,
            "numAtoms": 8,
        },
        "NVL816": {
            "DPW": 619,
            "DEVREVSTEP": ["8PF6CVP", "8PF6CVR", "8PF6CVER"],
            "TPrgx": r"(5[1,2]\w)",
            "dfBins": [3, 4],
            "numCores": 8,
            "numAtoms": 16,
        },
        "NVL816-BLLC": {
            "DPW": 393,
            "DEVREVSTEP": ["8PF5CVL","8PF5CVEL"],
            "TPrgx": r"(6[01]\w)",
            "dfBins": [3, 4],
            "numCores": 8,
            "numAtoms": 16,
        },
        "NVL-GPU-512": {
            "DPW": 377,
            "DEVREVSTEP": ["8PL7CV"],
            "TPrgx": r"(NGX\w{6,8})",
            "dfBins": [3, 4],
            "numCores": 8,   # TODO: verify
            "numAtoms": 16,  # TODO: verify
        }
    }
    devRevSteps = {
        "8PYJCVJ": "ARL68-N3B",
        "8PYVCVB": "ARLS816",
        "8PYVCVAB": "ARLS816",
        "8PF6CVP": "NVL816",
        "8PF6CVR": "NVL816",
        "8PF5CVL": "NVL816-BLLC",
        "8PL7CV": "NVL-GPU-512",
    }

    # moduleMap defined at module level (importable by external callers)

    defeatureModCnts = {} #{"FUN_CORE":18,"SCN_ATOM":37...}
    outFile = f"{pathlib.Path(dataInFile).stem}_out.xlsx"
    # If an explicit output directory was provided, use it (highest priority).
    if outDir:
        try:
            _od = pathlib.Path(outDir)
            _od.mkdir(parents=True, exist_ok=True)
            outFile = str((_od / outFile).resolve())
        except Exception:
            pass
    # If a dashboard path was provided, place the generated output workbook
    # in the same directory as the dashboard so outputs live next to it.
    elif dashboardFile:
        try:
            dbp = pathlib.Path(dashboardFile)
            if dbp.parent and str(dbp.parent) != '.':
                outFile = str((dbp.parent / outFile).resolve())
        except Exception:
            # on any error, fall back to the original outFile in cwd
            pass

    info = {}
    logging.info(f"{' ':{idt}}Checking inputs...")
    product, yieldDF = getYieldDataFrame(dataInFile, productInfo)
    info["product"] = product

    # logging.info(f"Length of df: {len(yieldDF)}")
    if not float(len(yieldDF) / productInfo[product]["DPW"]).is_integer():
        logging.warning(f"{' ':{idt}}Extra die found!  Number of die in data set ({len(yieldDF)} "
                        f"is not evenly divisible by DPW ({productInfo[product]['DPW']}).")
    numWafers = len(yieldDF) // productInfo[product]["DPW"]
    logging.info(f"{' ':{idt+2}}Found {numWafers} wafers.")
    info["numWafers"] = numWafers
    if waferLvl:
        info["numDie"] = productInfo[product]["DPW"]
    else:
        info["numDie"] = len(yieldDF)
    try:
        info["TP"] = getTPNum(df=yieldDF, rgx=productInfo[product]["TPrgx"], idt=2)
    except LookupError as e:
        logging.warning(f"{' ':{idt+2}}Can't find TP, defaulting to '??'.  Error:\n{e}")
        info["TP"] = "??"
    # binDefs = pandas.read_csv(binDefFile).set_index("B/C").to_dict("list")
    if binDefFile and os.path.isfile(str(binDefFile)):
        binDefs = dict(pandas.read_csv(binDefFile).values)
        logging.info(f"{' ':{idt}}Loaded {len(binDefs)} bin definitions from '{binDefFile}'.")
    else:
        if binDefFile:
            logging.warning(f"{' ':{idt}}bindef file '{binDefFile}' not found — building from data CSV.")
        else:
            logging.info(f"{' ':{idt}}No bindef file provided — building bin definitions from data CSV.")
        binDefs = _buildBinDefsFromDF(yieldDF)
        logging.info(f"{' ':{idt}}Built {len(binDefs)} bin definitions from data CSV.")
    lotCol  = getLotCol(df=yieldDF)
    waferCol = getWaferCol(df=yieldDF)
    if waferLvl:
        moduleYield = {}
        for lot in sorted(yieldDF[lotCol].unique()):
            for wafer in sorted(yieldDF.loc[yieldDF[lotCol] == lot][waferCol].unique()):
                lotWaferDF = yieldDF.loc[(yieldDF[lotCol] == lot) & (yieldDF[waferCol] == wafer)]
                moduleYield[f"{lot}_W{wafer}"] = getYieldByModule(yldDF=lotWaferDF,
                                                    binDefs=binDefs, modMap=moduleMap)
    else:
        moduleYield = getYieldByModule(yldDF=yieldDF, binDefs=binDefs, modMap=moduleMap)
    if not waferLvl:
        updateDefeatureModCnts(dfModCnts=defeatureModCnts, yieldDF=yieldDF,
                               waferLvl=waferLvl, productInfo=productInfo[product])
    # print(defeatureModCnts)

    if not vmax:
        if waferLvl:
            raise UnboundLocalError("Wafer level not nupported in makeOutXl!")
        makeOutXl(outFile=outFile, moduleYield=moduleYield, which="dd",
              defeatureModCnts=defeatureModCnts, info=info)
        # If a dashboard file was provided, append the output column
        if dashboardFile:
            try:
                # Only attempt xlsx append if the dashboard path looks like a workbook
                _db_suffix = pathlib.Path(dashboardFile).suffix.lower()
                if _db_suffix not in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
                    logging.debug(f"{' ':{idt}}Skipping dashboard append: '{dashboardFile}' is not an Excel workbook.")
                else:
                    # srcSheet falls back to first sheet inside append_to_dashboard if not found
                    append_to_dashboard(dashboardPath=dashboardFile, srcWorkbook=outFile,
                                         srcSheet=info.get("product", ""))
            except Exception as e:
                logging.warning(f"{' ':{idt}}Failed to append to dashboard: {e}")
    outFile = f"{pathlib.Path(dataInFile).stem}_out_vmax.xlsx"
    if outDir:
        try:
            _od = pathlib.Path(outDir)
            _od.mkdir(parents=True, exist_ok=True)
            outFile = str((_od / outFile).resolve())
        except Exception:
            pass
    elif dashboardFile:
        try:
            dbp = pathlib.Path(dashboardFile)
            if dbp.parent and str(dbp.parent) != '.':
                outFile = str((dbp.parent / outFile).resolve())
        except Exception:
            pass
    if vmax:
        makeOutXlVmax(outFile=outFile, allModuleYield=moduleYield, info=info, moduleMap=moduleMap, waferLvl=waferLvl)
    material = listOfTuplesToDict(yieldDF.value_counts([lotCol,waferCol]).keys().to_list())
    s = "; ".join([f"{key}:{','.join(str(x) for x in value)}" for key,value in material.items()])
    logging.info(f"{' ':{idt}}Material Used for this analysis:\n\n{s}\n")


    # print(len(yieldDF), len(yieldDF["FB@F24_132110"].value_counts()))

    # logging.info(f"FBin Summary:\n{yieldDF['FB@F24_132110'].value_counts().nlargest(5)}")
    # series = yieldDF["FB@F24_132110"].value_counts()
    # for k,v in series.items():
    #     print(k, v)
    # for i in range(20):
    #     for col in yieldDF:
    #         if "LOGTRACKER_AM" in col or "LOGTRACKER_CR" in col:
    #             s = yieldDF[col][i]
    #             if s:
    #                 logging.info(f"{' ':{idt+2}} Decoded:")
    #                 print(prime_error_decode(s[10:].strip("=")).split("\n"))
    #     break
    #     s = yieldDF["TPI_BIN::CTRL_UB_X_E_FINAL_X_X_X_X_LOGTRACKER_AM0_01@F24_132110"][i]
    #     if s:
    #         logging.info(f"Inflating: '{s[10:len(s)-3]}'")
    #         sDecoded = prime_error_decode(s[10:len(s)-3])
    #         logging.info(sDecoded)
    #         break
    # for i in range(10):
    #     logging.info(df["FB@F24_132110"][i])
    # for fb in df["FB@F24_132110"]:
    #     logging.info(fb)
    logging.debug("All done.")


def getProductFromProgram(program=""):
    if "ARCSDSCB0" in program:
        return "ARLS816"
    elif "ARCSDSCJ0" in program:
        return "ARL68-N3B"
    elif "NGXSDSC" in program or "NGX" in program:
        return "NVL-GPU-512"
    else:
        return False


def updateDefeatureModCnts(dfModCnts={}, yieldDF=pandas.DataFrame(), productInfo={}, waferLvl=False, idt=4,
                           binDefs=None, modMap=None):
    """Populate dfModCnts (module → count) for defeatured (IB3/IB4) die.

    Primary path: decode LOGTRACKER_AM/AP/CR columns to identify the defeatured module.
    Fallback path (when binDefs + modMap are supplied and LOGTRACKER yields nothing):
      look up each IB3/IB4 die's DATA_BIN in binDefs, run getModuleFromBinDesc on the
      bin description string, and use the resulting 'dd' category as the module key.
      This covers products where LOGTRACKER columns are absent or blank (e.g. NVL816-BLLC).
    """
    if waferLvl:
        raise UnboundLocalError("Wafer level not supported in updateDefeatureModCnts!")
    # print(len(yieldDF.loc[yieldDF[getIBinCol(yieldDF)].isin([3,4])]))
    dfDF = yieldDF.loc[yieldDF[getIBinCol(yieldDF)].isin(productInfo["dfBins"])]
    # print(len(dfDF))
    dfCnt = 0
    logtrackerRgx = [f"LOGTRACKER_AM[0-{productInfo['numAtoms']/4-1}]",
                     f"LOGTRACKER_AP[0-{productInfo['numAtoms']/4-1}]",
                     f"LOGTRACKER_CR[0-{productInfo['numCores']-1}]"]
    # print("*****************************************************")
    for i, row in dfDF.iterrows():
        # print(row)
        foundFlag = False
        for col,val in row.items():
            if foundFlag:
                break
            for rgx in logtrackerRgx:
                m = re.search(rgx, col)
                if m:
                    try:
                        decoded = prime_error_decode(val[10:].strip("=")).split("\n")
                    except TypeError:
                        continue
                    # print(decoded)
                    if len(decoded) > 1:
                        # print(decoded[1].split("|"))
                        modMatch = re.search(r"(\w{3}_\w{4,5}).*::.*", decoded[1].split("|")[-1])
                        if modMatch:
                            dfMod = modMatch.group(1).replace("MBIST", "CORE")
                            # print(dfMod)
                            if dfMod in dfModCnts:
                                dfModCnts[dfMod] += 1
                            else:
                                dfModCnts[dfMod] = 1
                            dfCnt += 1
                            foundFlag = True
                            break
        if not foundFlag and binDefs is None:
            # Only log per-row warnings when there is no binDefs fallback available.
            logging.warning(f"{' ':{idt}}Warning: Could not find defeatured module for row:\n")
            print(row)
    # ── Fallback: if LOGTRACKER decoding found nothing, infer module from bin description ──
    # This handles products (e.g. NVL816-BLLC) where LOGTRACKER columns are absent or blank.
    # The bin description string (e.g. "B20260000_FAIL_ARR_ATOM_LSA_...") already encodes
    # the failure domain and matches the same moduleMap regex patterns used by getYieldByModule.
    if dfCnt == 0 and len(dfDF) > 0 and binDefs is not None and modMap is not None:
        binCol = getBinCol(df=dfDF)
        if binCol:
            for dbin, n in dfDF[binCol].value_counts().items():
                try:
                    dbin_f = float(dbin)
                except (ValueError, TypeError):
                    continue
                roundDbin = round(dbin_f / 10000)
                binKey = ("FB" + str(roundDbin)) if roundDbin < 1000 else ("DB" + str(round(dbin_f)))
                binDesc = binDefs.get(binKey)
                if not binDesc:
                    logging.warning(f"{' ':{idt}}binDefs fallback: no entry for key '{binKey}' (dbin={dbin}, n={n}) — skipping.")
                    continue
                try:
                    module = getModuleFromBinDesc(modMap=modMap, binDesc=str(binDesc))
                    ddCat = module.get("dd", "")
                    if ddCat:
                        dfModCnts[ddCat] = dfModCnts.get(ddCat, 0) + n
                        dfCnt += n
                except LookupError:
                    logging.warning(f"{' ':{idt}}binDefs fallback: no moduleMap match for '{binDesc}' — skipping.")
                    continue
    if dfCnt != len(dfDF):
        logging.warning(f"{' ':{idt}}Warning: Not all Defeatured die accounted for!  There are "
                        f"{len(dfDF)} Defeatured die, but only found modules for {dfCnt} of them!")


def makeOutXlVmax(outFile="", allModuleYield={}, info={}, moduleMap={}, waferLvl=False, idt=2):
    if not HAVE_OPENPYXL:
        logging.warning(f"{' ':{idt}}openpyxl not installed — writing CSV summary instead.")
        csv_out = f"{pathlib.Path(outFile).stem}_vmax.csv"
        try:
            with open(csv_out, "w", newline="") as fh:
                writer = csv.writer(fh)
                writer.writerow(["LotWafer", "Module", "Bucket", "Count", "Percent"])
                # allModuleYield may be a dict of lotWafer -> moduleYield
                if isinstance(allModuleYield, dict):
                    for lotWafer, moduleYield in allModuleYield.items():
                        for moduleType, buckets in moduleYield.items():
                            if isinstance(buckets, dict):
                                for bucketName, cnt in buckets.items():
                                    try:
                                        pct = cnt / info.get("numDie", 1)
                                    except Exception:
                                        pct = ""
                                    writer.writerow([lotWafer, moduleType, bucketName, cnt, pct])
                else:
                    writer.writerow(["all", "vmax", str(allModuleYield), "", ""])
            logging.info(f"{' ':{idt}}CSV written to '{csv_out}'.")
        except Exception as e:
            logging.error(f"{' ':{idt}}Failed to write CSV summary: {e}")
        return
    if waferLvl:
        outFile = f"{pathlib.Path(outFile).stem}_wafer.xlsx"
    logging.info(f"{' ':{idt}}Creating or overwriting '{outFile}'.")
    outWB = openpyxl.Workbook()
    outWS = outWB[outWB.sheetnames[0]]
    outWS.title = info["product"]
    outWS.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 36
    if not waferLvl:
        allModuleYield = {"1":allModuleYield}
    for iterCnt, (lotWafer, moduleYield) in enumerate(allModuleYield.items()):
        yieldBuckets = []
        #Headers
        if iterCnt == 0:
            writeCell(ws=outWS, row=1, col=iterCnt+1, value="Sub Module", wrapText=True, bold=True)
        if waferLvl:
            s = f"{lotWafer.replace('_', ' ')} {info['TP']} Yield Loss (Fail Bins) (%)"
        else:
            s = f"{info['numWafers']}W {info['TP']} Yield Loss (Fail Bins) (%)"
        writeCell(ws=outWS, row=1, col=iterCnt+2, value=s, wrapText=True, bold=True)
        if iterCnt == 0:
            writeCell(ws=outWS, row=2, col=iterCnt+1, value="# Die", bold = True)
        writeCell(ws=outWS, row=2, col=iterCnt+2, value= info["numDie"], bold = True)

        row = 3
        sums = []
        pcntTotal = 0

        #Good Bins
        for bucket in ["Bin 1 (No Repair)", "Bin 198 (Vmin Repair)", "Bin 2 (Hard Repair)",
                       "Bin 202 (Vmax Repair)", "Bin 3", "Bin 4"]:
            if iterCnt == 0:
                writeCell(ws=outWS, row=row, col=iterCnt+1, value=bucket)
            n = getYieldFromModYield(moduleYield, "vmax", bucket)
            yieldBuckets.append(bucket)
            writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"={n}/B2", numFmt="0.0%")
            pcntTotal += (n / info["numDie"])
            row += 1
        if iterCnt == 0:
            writeCell(ws=outWS, row=row, col=iterCnt+1, value="SUM", bold=True)
        colLetter = openpyxl.utils.get_column_letter(iterCnt+2)
        writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"=SUM({colLetter}{row-6}:{colLetter}{row-1})",
                  numFmt="0.0%", bold=True)
        sums.append(row)
        row += 1

        #ARR, FUN, SCN and sub modules
        buckets = ["ARR_ATOM_VNOM_LFM", "ARR_ATOM_VNOM_TFM", "ARR_ATOM_VMAX_LFM",
                   "ARR_ATOM_VMAX_TFM", "FUN_ATOM_VNOM_LFM", "FUN_ATOM_VNOM_TFM",
                   "SCN_ATOM_VNOM_LFM", "SCN_ATOM_VNOM_TFM"]
        rslt = writeBucketsToXl(outWS=outWS, moduleYield=moduleYield, moduleMap=moduleMap,
            row=row, which="vmax", buckets=buckets, iterCnt=iterCnt, info=info)
        row = rslt["row"]
        pcntTotal += rslt["pcntTotal"]
        yieldBuckets += rslt["yieldBuckets"]
        sums.append(row)
        row += 1

        buckets = ["ARR_CCF_VNOM_LFM", "ARR_CCF_VNOM_TFM", "ARR_CCF_VMAX_LFM",
                   "ARR_CCF_VMAX_TFM", "FUN_CCF_VNOM_LFM", "FUN_CCF_VNOM_TFM",
                   "FUN_ATOM_DRAGON_SLC_VNOM_LFM", "FUN_ATOM_DRAGON_SLC_VNOM_TFM",
                   "FUN_ATOM_DRAGON_SLC_SPECKLE",
                   "SCN_CCF_VNOM_LFM", "SCN_CCF_VNOM_TFM", "ARR_NONCCF_VNOM_LFM",
                   "ARR_NONCCF_VMAX_LFM", "SCN_NONCCF_VNOM_LFM"]
        rslt = writeBucketsToXl(outWS=outWS, moduleYield=moduleYield, moduleMap=moduleMap,
            row=row, which="vmax", buckets=buckets, iterCnt=iterCnt, info=info)
        row = rslt["row"]
        pcntTotal += rslt["pcntTotal"]
        yieldBuckets += rslt["yieldBuckets"]
        sums.append(row)
        row += 1

        buckets = ["ARR_CORE_VNOM_LFM", "ARR_CORE_VNOM_TFM", "ARR_CORE_VMAX_LFM",
                   "ARR_CORE_VMAX_TFM", "FUN_CORE_VNOM_LFM", "FUN_CORE_VNOM_TFM",
                   "FUN_CORE_VMAX_LFM",
                   "SCN_CORE_VNOM_LFM", "SCN_CORE_VNOM_TFM"]
        rslt = writeBucketsToXl(outWS=outWS, moduleYield=moduleYield, moduleMap=moduleMap,
            row=row, which="vmax", buckets=buckets, iterCnt=iterCnt, info=info)
        row = rslt["row"]
        pcntTotal += rslt["pcntTotal"]
        yieldBuckets += rslt["yieldBuckets"]
        sums.append(row)
        row += 1

        #Reset
        if iterCnt == 0:
            writeCell(ws=outWS, row=row, col=iterCnt+1, value="RESET (19,35)")
        n = getYieldFromModYield(moduleYield, "vmax", "Reset")
        writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"={n}/B2", numFmt="0.0%")
        pcntTotal += (n / info["numDie"])
        yieldBuckets.append("Reset")
        sums.append(row)
        row += 1

        buckets = [ "Analog PTH BG", "Analog PTH DLVR", "Analog PTH DTS", "Analog PTH ODI",
                    "Analog CLK", "Analog MIO D2D", "Analog Other"]
        rslt = writeBucketsToXl(outWS=outWS, moduleYield=moduleYield, moduleMap=moduleMap,
                                row=row, which="vmax", buckets=buckets, iterCnt=iterCnt, info=info)
        row = rslt["row"]
        pcntTotal += rslt["pcntTotal"]
        yieldBuckets += rslt["yieldBuckets"]
        sums.append(row)
        row += 1

        #Other
        for other in ["TPI Foundry", "TPI Foundry ADTL (B43)", "TPI Foundry SICC (B18)", "TPI Other", "HVQK (B26)"]:
            if iterCnt == 0:
                writeCell(ws=outWS, row=row, col=iterCnt+1, value=f"{other}")
            n = getYieldFromModYield(moduleYield, "vmax", f"{other}")
            writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"={n}/B2", numFmt="0.0%")
            pcntTotal += (n / info["numDie"])
            yieldBuckets.append(other)
            sums.append(row)
            row += 1
        #Yield Sum
        if iterCnt == 0:
            writeCell(ws=outWS, row=row, col=iterCnt+1, value="Yield SUM (%)", bold=True)
        s = f"=SUM({colLetter}"
        s += f",{colLetter}".join(str(x) for x in sums)
        s += ")"
        writeCell(ws=outWS, row=row, col=iterCnt+2, value=s, bold=True, numFmt="0.0%")

        if round(pcntTotal*100,2) != 100.0:
            print("*************************************************")
            logging.warning(f"{' ':{idt}}WARNING!  Total yield percentage is "
                            f"{round(pcntTotal * 100,2)}%, not 100%!")
            checkBucketsUsed(moduleYield["vmax"], yieldBuckets)
        else:
            logging.info(f"{' ':{idt}}Total yield percentage is {round(pcntTotal * 100,1)}%.")

    outWB.save(filename=outFile)
    logging.info(f"{' ':{idt}}Excel written to '{outFile}'.")


def checkBucketsUsed(moduleYield={}, yieldBuckets=[], idt=6):
    logging.info(f"{' ':{idt-2}}Checking for any yield buckets that were not used...")
    # print(f"moduleYield : {moduleYield}")
    # print(f"yieldBuckets : {yieldBuckets}")
    flag = False
    for k,v in moduleYield.items():
        if k not in yieldBuckets:
            flag = True
            logging.info(f"{' ':{idt}}Bucket '{k}' with {v} die not counted!")
    if not flag:
        logging.info(f"{' ':{idt}}No unused buckets found.")


def makeOutXl(outFile="", moduleYield={}, which="dd", info={}, defeatureModCnts={}, idt=2):
    if not HAVE_OPENPYXL:
        logging.warning(f"{' ':{idt}}openpyxl not installed — writing CSV summary instead.")
        csv_out = f"{pathlib.Path(outFile).stem}.csv"
        try:
            with open(csv_out, "w", newline="") as fh:
                writer = csv.writer(fh)
                writer.writerow(["ModuleType", "Bucket", "Count", "Percent"])
                my = moduleYield.get(which, {}) if isinstance(moduleYield, dict) else {}
                for bucketName, cnt in (my.items() if isinstance(my, dict) else []):
                    try:
                        pct = cnt / info.get("numDie", 1)
                    except Exception:
                        pct = ""
                    writer.writerow([which, bucketName, cnt, pct])
            logging.info(f"{' ':{idt}}CSV written to '{csv_out}'.")
        except Exception as e:
            logging.error(f"{' ':{idt}}Failed to write CSV summary: {e}")
        return
    logging.info(f"{' ':{idt}}Creating or overwriting '{outFile}'.")
    outWB = openpyxl.Workbook()
    outWS = outWB[outWB.sheetnames[0]]
    outWS.title = info["product"]
    outWS.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 14
    #Headers
    writeCell(ws=outWS, row=1, col=1, value="Sub Module", wrapText=True, bold=True)
    s = f"{info['numWafers']}W {info['TP']} Yield Loss (Fail Bins) (%)"
    writeCell(ws=outWS, row=1, col=2, value=s, wrapText=True, bold=True)
    s = f"{info['numWafers']}W {info['TP']} Recovery Bins (3-4) (%)"
    writeCell(ws=outWS, row=1, col=3, value= s, wrapText=True, bold=True)
    writeCell(ws=outWS, row=2, col=1, value="# Die", bold = True)
    writeCell(ws=outWS, row=2, col=2, value= info["numDie"], bold = True)

    pcntTotal = 0
    row = 3
    yieldBuckets = []
    numDie = info["numDie"] or 1  # guard against div-by-zero

    # Track per-group sums (col2, col3) for cross-group SUM rows
    _grp_sums2 = []   # col2 sum per ARR/FUN/SCN group + Reset
    _grp_sums3 = []   # col3 sum per ARR/FUN/SCN group (defeature)
    _bins_sum2 = 0.0
    _other_sums2 = []

    #ARR, FUN, SCN and sub modules
    for module in ["ARR", "FUN", "SCN"]:
        grp2 = 0.0
        grp3 = 0.0
        for subMod in ["ATOM", "CCF", "CORE", "NONCCF"]:
            writeCell(ws=outWS, row=row, col=1, value=f"{module}_{subMod}")
            modKey = module + "_" + subMod
            yieldBuckets.append(modKey)
            n = getYieldFromModYield(moduleYield, "dd", modKey)
            pct2 = n / numDie
            writeCell(ws=outWS, row=row, col=2, value=pct2, numFmt="0.0%")
            grp2 += pct2
            pcntTotal += pct2
            if "ATOM" in modKey or "CORE" in modKey:
                df_n = defeatureModCnts.get(modKey, 0)
                pct3 = df_n / numDie
                writeCell(ws=outWS, row=row, col=3, value=pct3, numFmt="0.0%")
                grp3 += pct3
            row += 1
        writeCell(ws=outWS, row=row, col=1, value="SUM", bold=True)
        writeCell(ws=outWS, row=row, col=2, value=grp2, numFmt="0.0%", bold=True)
        writeCell(ws=outWS, row=row, col=3, value=grp3, numFmt="0.0%", bold=True)
        _grp_sums2.append(grp2)
        _grp_sums3.append(grp3)
        row += 1
    #Reset
    writeCell(ws=outWS, row=row, col=1, value="RESET (19,35)")
    n = getYieldFromModYield(moduleYield, "dd", "Reset")
    reset_pct = n / numDie
    writeCell(ws=outWS, row=row, col=2, value=reset_pct, numFmt="0.0%")
    yieldBuckets.append("Reset")
    pcntTotal += reset_pct
    _grp_sums2.append(reset_pct)
    row += 1
    #Good Bins
    for i in range(1,5):
        writeCell(ws=outWS, row=row, col=1, value=f"Bin {i}")
        if i == 1:
            n = getYieldFromModYield(moduleYield, "dd", f"Bin {i}") + \
                getYieldFromModYield(moduleYield, "dd", f"Bin 198 (Vmin Repair)")
            yieldBuckets.append(f"Bin {i}")
            yieldBuckets.append(f"Bin 198 (Vmin Repair)")
        elif i == 2:
            n = getYieldFromModYield(moduleYield, "dd", f"Bin 2 (Hard Repair)") + \
                getYieldFromModYield(moduleYield, "dd", f"Bin 202 (Vmax Repair)")
            yieldBuckets.append(f"Bin 2 (Hard Repair)")
            yieldBuckets.append(f"Bin 202 (Vmax Repair)")
        else:
            n = getYieldFromModYield(moduleYield, "dd", f"Bin {i}")
            yieldBuckets.append(f"Bin {i}")
        bin_pct = n / numDie
        writeCell(ws=outWS, row=row, col=2, value=bin_pct, numFmt="0.0%")
        _bins_sum2 += bin_pct
        pcntTotal += bin_pct
        row += 1
    writeCell(ws=outWS, row=row, col=1, value="SUM", bold=True)
    writeCell(ws=outWS, row=row, col=2, value=_bins_sum2, numFmt="0.0%", bold=True)
    writeCell(ws=outWS, row=row, col=3, value=sum(_grp_sums3), numFmt="0.0%", bold=True)
    _grp_sums2.append(_bins_sum2)
    row += 1

    #Repairs! #todo
    repairs_sum2 = 0.0
    for repairRow in ["Bin 198 (Vmin Repair)", "Bin 202 (Vmax Repair)", "Bin 2 (Hard Repair)"]:
        writeCell(ws=outWS, row=row, col=1, value=f"Repair {repairRow}")
        n = getYieldFromModYield(moduleYield, "dd", f"{repairRow}")
        rpr_pct = n / numDie
        writeCell(ws=outWS, row=row, col=2, value=rpr_pct, numFmt="0.0%")
        repairs_sum2 += rpr_pct
        row += 1
    writeCell(ws=outWS, row=row, col=1, value="SUM", bold=True)
    writeCell(ws=outWS, row=row, col=2, value=repairs_sum2, numFmt="0.0%", bold=True)
    row += 1
    #Other
    for other in ["Analog", "TPI Foundry", "TPI Other", "TPI Other - B93", "TPI Other - B98",
              "TPI Other - B99", "HVQK (B26)"]:
        writeCell(ws=outWS, row=row, col=1, value=f"{other}")
        n = getYieldFromModYield(moduleYield, "dd", f"{other}")
        oth_pct = n / numDie
        writeCell(ws=outWS, row=row, col=2, value=oth_pct, numFmt="0.0%")
        pcntTotal += oth_pct
        yieldBuckets.append(other)
        _other_sums2.append(oth_pct)
        row += 1
    #Yield Sum
    writeCell(ws=outWS, row=row, col=1, value="Yield SUM (%)", bold=True)
    yield_sum = sum(_grp_sums2) + sum(_other_sums2)
    writeCell(ws=outWS, row=row, col=2, value=yield_sum, bold=True, numFmt="0.0%")

    if round(pcntTotal*100,2) != 100.0:
        logging.warning(f"{' ':{idt}}WARNING!  Total yield percentage is "
                        f"{round(pcntTotal * 100,2)}%, not 100%!")
        checkBucketsUsed(moduleYield["dd"], yieldBuckets)
    else:
        logging.info(f"{' ':{idt}}Total yield percntage is {round(pcntTotal * 100,1)}%.")

    outWB.save(filename=outFile)
    logging.info(f"{' ':{idt}}Excel written to '{outFile}'.")


def writeBucketsToXl(outWS="", moduleYield={}, moduleMap={}, row=0, which="vmax",
                     buckets=[], iterCnt=0, info={}):
    pcntTotal = 0
    yieldBuckets = []
    for bucket in buckets:
        modKey = getModKey(moduleYield=moduleYield, which=which,
                           pattern=rf"{bucket} \(", moduleMap=moduleMap)
        if iterCnt == 0:
            writeCell(ws=outWS, row=row, col=1, value=modKey)
        n = getYieldFromModYield(moduleYield, which, modKey)
        writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"={n}/B2", numFmt="0.0%")
        pcntTotal += (n / info["numDie"])
        yieldBuckets.append(modKey)
        row += 1
    if iterCnt == 0:
        writeCell(ws=outWS, row=row, col=1, value="SUM", bold=True)
    colLetter = openpyxl.utils.get_column_letter(iterCnt + 2)
    writeCell(ws=outWS, row=row, col=iterCnt+2, value=f"=SUM({colLetter}{row-len(buckets)}:{colLetter}{row-1})",
              numFmt="0.0%", bold=True)
    return {"row": row, "pcntTotal":pcntTotal, "yieldBuckets":yieldBuckets}
    #end writeBucketsToXl


def getModKey(moduleYield={}, which="dd", pattern="", moduleMap={}):
    for bucket, n in moduleYield[which].items():
        m = re.search(pattern, bucket)
        if m:
            return bucket
    #no yield data in dataset for this bucket.  Search the main one.
    for expression, maps in moduleMap.items():
        m = re.search(pattern, maps[which])
        if m:
            return maps[which]
    return pattern


def getYieldFromModYield(moduleYield={}, which="dd", modKey=""):
    try:
        return moduleYield[which][modKey]
    except KeyError:
        return 0


def getYieldByModule(yldDF=pandas.DataFrame(), binDefs=pandas.DataFrame(), modMap={}, idt=4):
    if idt < 2:
        idt = 2
    moduleYield = {}  #{"dd":{"ARR_ATOM":432,"ARR_CORE":444}}
    binTracker = {}

    logging.info(f"{' ':{idt-2}}Calculating yield by sub module.")
    binCol = getBinCol(df=yldDF)
    if not binCol:
        raise LookupError("Unable to find column with the bin number for each die, quitting!")
    else:
        logging.info(f"{' ':{idt}}Using column '{binCol}' as the Bin result.")
    # logging.info(f"DBin Summary:\n{yldDF[binCol].value_counts().nlargest(5)}")

    i = 0
    # dfDF = yieldDF.loc[yieldDF[getIBinCol(yieldDF)].isin(productInfo["dfBins"])]
    #DB do not have Bin93 for some reason.
    numB93 = len(yldDF.loc[yldDF[getIBinCol(yldDF)] == 93])
    moduleYield = updateModYield(moduleYield=moduleYield,
                                 modulesToAdd=getModuleFromBinDesc(modMap=modMap, binDesc="B93"),
                                 dieCnt=numB93)
    totalDie = numB93
    for dbin, n in yldDF[binCol].value_counts().items():
        totalDie += n
        roundDbin = round(dbin / 10000)
        if roundDbin < 1000:
            binKey = "FB" + str(roundDbin)
        else:
            binKey = "DB" + str(round(dbin))
        try:
            binDesc = binDefs[binKey]
        except KeyError:
            logging.warning(f"{'  '}Unknown bin key '{binKey}' (raw dbin={dbin}, n={n}) — not in bin_defs, skipping.")
            continue
        try:
            module = getModuleFromBinDesc(modMap=modMap, binDesc=binDesc)
        except LookupError as e:
            s = e.__str__()
            s = s.replace("',", f", ({n} die),")
            logging.warning(s + " — skipping bin.")
            continue
        # logging.info(f"{' ':{idt}}Found module '{module}' from '{binDesc}', adding {n} to these buckets.")
        try:
            if binDesc not in binTracker[module["vmax"]]:
                binTracker[module["vmax"]].append(binDesc)
        except KeyError:
            binTracker[module["vmax"]] = [binDesc]

        # if "ARR_CORE_VMAX_TFM (B60)" in module.values():
        #     logging.info(f"{' ':{idt}}Found module '{module}' from '{binDesc}', adding {n} to these buckets.")
        moduleYield = updateModYield(moduleYield=moduleYield, modulesToAdd=module, dieCnt=n)
        # print(moduleYield)
        i+=1
    if totalDie != len(yldDF):
        logging.warning(f"{' ':{idt}}WARNING! Found {totalDie} die, but there should be {len(yldDF)}!")
    # logging.info(f"{' ':{idt}}Buckets by bin:")
    # for k,v in binTracker.items():
    #     print(k)
    #     v.sort()
    #     for b in v:
    #         print(f" {b}")
    return moduleYield


def updateModYield(moduleYield={}, modulesToAdd={}, dieCnt=0):
    for k,v in modulesToAdd.items():
        if k not in moduleYield:
            moduleYield[k] = {}
        if v not in moduleYield[k]:
            moduleYield[k][v] = dieCnt
        else:
            moduleYield[k][v] = moduleYield[k][v] + dieCnt
    return moduleYield


def getModuleFromBinDesc(modMap={}, binDesc="", idt=4):
    #start with B26 because a lot of the bin names are duplicated with just B26 in front.
    m = re.search(r"B26\d", binDesc)
    if m:
        return {"dd": "HVQK (B26)", "vmax": "HVQK (B26)"}
    for rgxPattern, modDict in modMap.items():
        m = re.search(rgxPattern, binDesc)
        if m:
            # if "B60" in binDesc: # or "B20" in binDesc:
            #     print(f"{binDesc}: {rgxPattern}")
            return modDict
    raise LookupError(f"No module map for '{binDesc}', quitting.")


def _buildBinDefsFromDF(df):
    """Build a binDefs dict (matching bindef CSV format) directly from the data CSV.

    For fail bins (roundDbin >= 1000): uses the 'Bin Description_' column value,
    which already contains the full test description string that moduleMap regexes match.
    For pass bins (roundDbin < 1000): synthesizes 'B{fb}_PASS' from the FB number,
    matching the moduleMap regex patterns (e.g. 'B198_PASS', 'B201_PASS').
    """
    db_col = getBinCol(df=df)
    if not db_col:
        return {}
    bd_col = next((c for c in df.columns if c.startswith('Bin Description')), None)
    binDefs = {}
    for dbin in df[db_col].dropna().unique():
        try:
            dbin_f = float(dbin)
        except (ValueError, TypeError):
            continue
        round_dbin = round(dbin_f / 10000)
        if round_dbin < 1000:
            bin_key = "FB" + str(round_dbin)
            if bin_key not in binDefs:
                binDefs[bin_key] = f"B{round_dbin}_PASS"
        else:
            bin_key = "DB" + str(round(dbin_f))
            if bin_key not in binDefs and bd_col is not None:
                mask = df[db_col] == dbin
                desc_vals = df.loc[mask, bd_col].dropna()
                if not desc_vals.empty:
                    binDefs[bin_key] = str(desc_vals.iloc[0])
    return binDefs


def getBinCol(df=pandas.DataFrame(), productInfo={}, idt=2):
    binCols = {}
    if "DB" in df:
        return "DB"
    else:
        for col in df:
            if "DB@" in col or "DB DIEBIN" in col or "DATA_BIN" in col or "DATA_BIN_132110" in col:
                return col
    return False
    # if "FB" in df:
    #     binCols["fbin"] = "FB"
    # else:
    #     for col in df:
    #         if "FB@" in col:
    #             binCols["fbin"] = col
    # if "DB" in df:
    #     binCols["dbin"] = "DB"
    # else:
    #     for col in df:
    #         if "DB@" in col:
    #             binCols["dbin"] = col
    # if "fbin" in binCols and "dbin" in binCols:
    #     return binCols
    # else:
    #     return False


def getIBinCol(df=pandas.DataFrame(), productInfo={}, idt=2):
    binCols = {}
    if "IB" in df:
        return "DB"
    else:
        for col in df:
            if "IB@" in col or "IB DIEBIN" in col or ("INTERFACE_BIN" in col and "TOTAL" not in col):
                return col
    return False


def getLotCol(df=pandas.DataFrame(), idt=2):
   for col in df:
       if col == "LOT":
           return col
       if "Lot_132110" in col or "SORT_LOT" in col:
           return col


def getWaferCol(df=pandas.DataFrame(), idt=2):
    for col in df:
        if col == "WAFER":
            return col
        if "SORT_WAFER" in col:
            return col


def getYieldDataFrame(inFile="", productInfo={}, idt=4):
    if not inFile:
        raise FileNotFoundError(f"No input file specified: '{inFile}'.")
    df = pandas.read_csv(inFile, header=0, low_memory=False)
    product = getPart(df=df, productInfo=productInfo)
    if product:
        logging.info(f"{' ':{idt}}Opened file '{inFile}' and found product: {product}.")
    else:
        logging.warning(f"{' ':{idt}}Unable to identify product from file; defaulting to ARLS816.")
        product = "ARLS816"
    return product, df


def getTPNum(df=pandas.DataFrame(), rgx="", idt=2):
    tpCol = ""
    for column in df:
        if "PROGRAM" in column.upper():
            tpCol = column
            break
    # Use first non-null value to avoid TypeError when first row has NaN
    series = df[tpCol].dropna()
    first_val = str(series.iloc[0]) if len(series) > 0 else ""
    m = re.search(rgx, first_val)
    if m:
        logging.info(f"{' ':{idt + 2}}Found TP line '{m.group(1)}' from {first_val}.")
        return m.group(1)
    else:
        raise LookupError(f"Unable to find a match for '{rgx}' in '{first_val}'.")


def getPart(df=pandas.DataFrame(), productInfo={}, idt=2):
    def _first_val(col):
        """Return first non-null string value from a column, or empty string."""
        series = df[col].dropna()
        return str(series.iloc[0]) if len(series) > 0 else ""
    if "Part" in df:
        return getProductFromDevRevStep(_first_val("Part"), productInfo=productInfo)
    for column in df:
        if "Part@" in column:
            return getProductFromDevRevStep(_first_val(column), productInfo=productInfo)
        if "DevRevStep" in column:
            return getProductFromDevRevStep(_first_val(column), productInfo=productInfo)
    for column in df:
        if "PROGRAM" in column:
            return getProductFromProgram(_first_val(column))

def getProductFromDevRevStep(drs="", productInfo={}):
    for product, prodInfo in productInfo.items():
        if drs in prodInfo["DEVREVSTEP"]:
            return product
    # Prefix match for variants not explicitly listed (e.g. 8PL7CVA, 8PL7CVB)
    for product, prodInfo in productInfo.items():
        if any(drs.startswith(prefix) for prefix in prodInfo["DEVREVSTEP"]):
            return product
    return False


###########################################
#General functions that are useful to other scripts
###########################################
IntToCharMapping = "ABCDEFGHIJKLMNOPQRSTUVWXYZ234567"
CharToIntMapping = {c: i for i, c in enumerate(IntToCharMapping)}

def prime_error_encode(text):
    bytes_str = zlib.compress(text.encode("utf-8"))
    bytes_str = bytes_str[2:-4] # discard header and tail
    bits_str = "".join([bin(b).replace("0b", "").rjust(8, "0") for b in bytes_str])
    bits_str += "0" * (5 - len(bits_str) % 5)
    encoded_str = "".join([IntToCharMapping[int(bits_str[i:i+5], 2)] for i in range(0, len(bits_str), 5)])
    return encoded_str

def prime_error_decode(encoded_str):
    if len(encoded_str) == 0:
        return ""
    bits_str = "".join([bin(CharToIntMapping[c]).replace("0b", "").rjust(5, "0") for c in encoded_str])
    bits_str += "0" * (8 - len(bits_str) % 8)
    bytes_arr = [int(bits_str[i:i+8], 2) for i in range(0, len(bits_str), 8)]
    bytes_str = b"".join([b.to_bytes(1, "big") for b in bytes_arr]) # "litte" works too
    text = zlib.decompress(bytes_str, -8).decode("utf-8")
    return text


def listOfTuplesToDict(listOfTuples=[()]):
    resultDict = {}
    #first item of each tuple is Lot #, second item is Wafer #
    for t in listOfTuples:
        try:
            resultDict[t[0]].append(t[1])
        except KeyError:
            resultDict[t[0]] = [t[1]]
    return resultDict


def writeCell(ws="", row=0, col=0, value="", numFmt="", wrapText=False, bold=False, idt=4):
    ws.cell(row=row, column=col).value = value
    if numFmt:
        ws.cell(row=row, column=col).number_format = numFmt
    if wrapText:
        ws.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True)
    if bold:
        ws.cell(row=row, column=col).font = openpyxl.styles.Font(bold=True)


def append_to_dashboard(dashboardPath="", srcWorkbook="", srcSheet="", tgtSheet=""):
    from openpyxl import load_workbook
    from pathlib import Path
    import shutil as _shutil

    dbPath = Path(dashboardPath)
    srcPath = Path(srcWorkbook)
    if not srcPath.exists():
        raise FileNotFoundError(f"Source workbook not found: {srcPath}")

    # If the dashboard file does not exist yet, bootstrap it by copying the _out.xlsx.
    # The copy becomes the new dashboard so the user can build on it over time.
    if not dbPath.exists():
        dbPath.parent.mkdir(parents=True, exist_ok=True)
        _shutil.copy2(str(srcPath), str(dbPath))
        logging.info(f"Dashboard file not found; created from source: {dbPath}")
        return

    # load source with formulas so we copy formulas and not only evaluated values
    src_wb = load_workbook(srcPath, data_only=False)
    if srcSheet not in src_wb.sheetnames:
        # fall back to the first sheet if named sheet not found
        srcSheet = src_wb.sheetnames[0]
    src_ws = src_wb[srcSheet]

    # open dashboard and append as new right-most set of columns (one per source column)
    tgt_wb = load_workbook(dbPath)
    # use the first sheet as target if tgtSheet doesn't exist
    if tgtSheet and tgtSheet in tgt_wb.sheetnames:
        tgt_ws = tgt_wb[tgtSheet]
    else:
        tgt_ws = tgt_wb[tgt_wb.sheetnames[0]]

    # append columns B and C (source columns 2 and 3) as the next right-most columns
    next_col = tgt_ws.max_column + 1
    max_row = src_ws.max_row
    # only copy columns 2 and 3 if they exist in the source
    cols_to_copy = [2, 3]
    actual_cols = [c for c in cols_to_copy if c <= src_ws.max_column]
    if not actual_cols:
        raise LookupError(f"Source sheet '{srcSheet}' does not contain columns B/C to copy.")

    # If the dashboard already contains columns with the same headers, fix formulas
    # in-place so they reference their own column (avoid stale formulas referencing B2/C2)
    from openpyxl.utils import get_column_letter
    src_headers = [src_ws.cell(row=1, column=c).value or f"{src_ws.title} {get_column_letter(c)}" for c in actual_cols]
    if src_headers:
        for tgt_c in range(1, tgt_ws.max_column + 1):
            try:
                hdr = tgt_ws.cell(row=1, column=tgt_c).value
            except Exception:
                hdr = None
            if hdr in src_headers:
                # rewrite formulas in this existing column to reference the column itself
                tgt_letter = get_column_letter(tgt_c)
                for r in range(2, tgt_ws.max_row + 1):
                    cell = tgt_ws.cell(row=r, column=tgt_c)
                    val = cell.value
                    if isinstance(val, str) and val.startswith('='):
                        # replace any occurrences of source column letters (B/C) with tgt_letter
                        for sidx in actual_cols:
                            sletter = get_column_letter(sidx)
                            val = val.replace(f"{sletter}2", f"{tgt_letter}2")
                            # replace ranges like B3:B6 -> K3:K6
                            val = val.replace(f"{sletter}", tgt_letter)
                        cell.value = val

    from openpyxl.styles import Font
    for idx, c in enumerate(actual_cols):
        tgt_col = next_col + idx
        # copy header explicitly (use source header if present, else fallback)
        src_header = src_ws.cell(row=1, column=c).value
        if not src_header:
            # fallback to sheet title + column letter
            src_header = f"{src_ws.title} {openpyxl.utils.get_column_letter(c)}"
        tgt_header = tgt_ws.cell(row=1, column=tgt_col)
        tgt_header.value = src_header
        tgt_header.font = Font(bold=True)
        try:
            tgt_header.alignment = openpyxl.styles.Alignment(wrap_text=True)
        except Exception:
            pass

        from openpyxl.utils import column_index_from_string, get_column_letter
        import re

        def rewrite_formula(formula: str):
            # replace any column-letter references that point to source columns
            def repl(m):
                col_letters = m.group(1)
                row_num = m.group(2)
                try:
                    src_idx = column_index_from_string(col_letters)
                except Exception:
                    return m.group(0)
                if src_idx in actual_cols:
                    tgt_letter = get_column_letter(next_col + actual_cols.index(src_idx))
                    return f"{tgt_letter}{row_num}"
                return m.group(0)
            return re.sub(r'([A-Z]+)(\d+)', repl, formula)

        for r in range(2, max_row + 1):
            src_cell = src_ws.cell(row=r, column=c)
            tgt_cell = tgt_ws.cell(row=r, column=tgt_col)
            val = src_cell.value
            if isinstance(val, str) and val.startswith('='):
                # rewrite formula to point to the appended dashboard columns
                try:
                    tgt_cell.value = rewrite_formula(val)
                except Exception:
                    tgt_cell.value = val
            else:
                tgt_cell.value = val

            # keep row 2 formatting from source (counts), other rows use percent
            try:
                if r == 2:
                    tgt_cell.number_format = src_cell.number_format or 'General'
                else:
                    tgt_cell.number_format = '0.0%'
            except Exception:
                pass
            try:
                tgt_cell.font = src_cell.font
                tgt_cell.alignment = src_cell.alignment
            except Exception:
                pass

    # finished copying selected columns

    tgt_wb.save(dbPath)


def findChars(s="", c=""):
    return [i for i, letter in enumerate(s) if letter == c]


def getDateTime():
    import datetime as _dt
    now = _dt.datetime.now().strftime("%Y%m%d-%H%M%S.%f")
    return now[:-3]


def setupLogging(logFile=False, logLevel=logging.INFO):
    handlers = [logging.StreamHandler(sys.stdout)]
    if logFile:
        handlers.append(logging.FileHandler(filename=logFile, mode="w"))
    logging.basicConfig(level=logLevel,
                        format="%(asctime)s.%(msecs)03d: %(message)s",
                        datefmt="%Y%m%d:%H:%M:%S", handlers=handlers)


def main():
    global docopt
    if docopt:
        try:
            args = docopt(__doc__, version=__version__)
        except BaseException:
            # docopt parsing failed (e.g., user passed argparse-style flags); fallback to argparse
            docopt = None
            args = None
    if not docopt:
        import argparse
        parser = argparse.ArgumentParser(description="Get Digital Dashboard Update")
        parser.add_argument("-d", "--data", dest="data", default="", help=".csv Input file containing a list of die and FBINs")
        parser.add_argument("-b", "--bin_defs", dest="bin_defs", default="", help=".csv file containing a table of DBIN and Test name")
        parser.add_argument("-g", "--log", dest="log", default="", help="Optionally, log to a file")
        parser.add_argument("-m", "--vmax", dest="vmax", action="store_true", help="Make the Vmax summary as well")
        parser.add_argument("-w", "--wafer", dest="wafer", action="store_true", help="Print columns per wafer")
        parser.add_argument("--verbose", dest="verbose", action="store_true", help="Print more information than usual")
        parser.add_argument("-x", "--debug", dest="debug", action="store_true", help="Print lots of debugging statements")
        parser.add_argument("-v", "--version", dest="version", action="store_true", help="Show version")
        parser.add_argument("--dashboard", dest="dashboard", default="", help="Path to DigitalDashBoard.xlsx to append results")
        parser.add_argument("--outdir", dest="outdir", default="", help="Output folder for generated xlsx files")
        parsed = parser.parse_args()
        args = {
            "--data": parsed.data,
            "--bin_defs": parsed.bin_defs,
            "--log": parsed.log,
            "--vmax": parsed.vmax,
            "--wafer": parsed.wafer,
            "--verbose": parsed.verbose,
            "--debug": parsed.debug,
            "--version": parsed.version,
            "--dashboard": parsed.dashboard,
            "--outdir": parsed.outdir,
        }

    # choose log level (debug overrides verbose)
    if args.get("--debug"):
        ll = logging.DEBUG
    elif args.get("--verbose"):
        ll = logging.INFO
    else:
        ll = logging.WARNING

    # determine where to place log files
    if args.get("--log"):
        user_log = args["--log"]
        p = pathlib.Path(user_log)
        suffix = p.suffix if p.suffix else ".log"
        name = p.stem if p.stem else __version__.replace(' ', '_').lower()
        logFileName = f"{name}_{getDateTime()}{suffix}"
        # if an absolute or explicit directory was provided, use it; otherwise use system temp
        if p.parent and str(p.parent) != ".":
            target_dir = p.parent
        elif p.is_absolute():
            target_dir = p.parent
        else:
            target_dir = pathlib.Path(tempfile.gettempdir())
    else:
        logFileName = f"{__version__.replace(' ', '_').lower()}_{getDateTime()}.log"
        target_dir = pathlib.Path(tempfile.gettempdir())
    try:
        pathlib.Path(target_dir).mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    logPath = str(pathlib.Path(target_dir) / logFileName)
    setupLogging(logFile=logPath, logLevel=ll)
    logBase = pathlib.Path(logFileName).stem
    try:
          logging.info(f"Welcome to {__version__}!")
          getDD(dataInFile=args.get("--data"), binDefFile=args.get("--bin_defs"), waferLvl=args.get("--wafer"),
              vmax=args.get("--vmax"), logBaseName=logBase, dashboardFile=args.get("--dashboard"),
              outDir=args.get("--outdir", ""))
    except KeyboardInterrupt:
        print("Ctl+C detected, exiting.")


if __name__ == "__main__":
    main()
_get_dd_main = main  # stable alias for yield_pipeline in-process call


# ════════════════════════════════════════════════════════════════
# (formerly add_material_type.py)
# ════════════════════════════════════════════════════════════════
# =============================================================================
# add_material_type.py  -  Material type merge for yield CSV
# =============================================================================
#
# Steps (mirrors JSL add_material_type_nvl.jsl logic):
# // 1. Take the yield CSV (passed as csv_path).
# // 2. Extract DevRevStep_* column, use first 6 chars as product prefix.
# // 3. In the collateral/material folder, find CSV whose filename contains
# //    that 6-char prefix (same pattern as reticle mapping lookup).
# // 4. From the yield CSV, derive:
# //       LOT7 = first 7 chars of the lot column (Lot_119325, Lot_132322, etc.)
# //       WAFER2 = last 2 chars of the wafer column (SORT_WAFER), as integer
# // 5. Merge material data into yield CSV on LOT7 == INTEL_LOT7 and
# //    WAFER2 == WaferID.
# //    Columns added: TSMC_LOT, Material Type Skew BEOL Skew, Material Type,
# //    Production Lot.
# // 6. Return the path to the updated CSV for further analysis.
# =============================================================================


import os
from pathlib import Path

import pandas as pd


# Columns to merge from the material lookup table
MATERIAL_MERGE_COLS = [
    'TSMC_LOT',
    'Material Type, Skew, BEOL Skew',
    'Material Type',
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_material_candidates(collateral_dir: str, prefix6: str) -> list[str]:
    """Return all CSVs in collateral_dir, sorted alphabetically.

    The ``prefix6`` argument is kept for backwards compatibility but is no
    longer used for filename filtering.  Material file names use device codes
    (e.g. ``8PF6CV``) while the DevRevStep column uses product codes (e.g.
    ``NCXSDJ``) — the two share no common prefix.  Lot-number matching in
    Step 6 naturally selects the files that cover the lots present in the
    yield CSV, so it is safe to return *all* CSV files here.
    """
    if not os.path.isdir(collateral_dir):
        return []
    return [
        os.path.join(collateral_dir, fname)
        for fname in sorted(os.listdir(collateral_dir))
        if fname.lower().endswith('.csv')
    ]


def _detect_lot_wafer_columns(all_cols: list[str]) -> tuple[str | None, str | None]:
    """Detect the lot and wafer column names from the CSV header.
    Returns (lot_col, wafer_col) or (None, None) if not found."""
    # Priority order matches the JSL logic
    lot_col = None
    wafer_col = None

    # SORT_LOT_U1.U5 is preferred: values are already 7-char sort lot IDs
    if 'SORT_LOT_U1.U5' in all_cols and 'SORT_WAFER_U1.U5' in all_cols:
        return 'SORT_LOT_U1.U5', 'SORT_WAFER_U1.U5'

    # Fallback to CLASS lot column (9-char; first 7 match INTEL_LOT7)
    if 'Lot_119325_U1.U5' in all_cols and 'SORT_WAFER_U1.U5' in all_cols:
        return 'Lot_119325_U1.U5', 'SORT_WAFER_U1.U5'

    # Standard SORT columns
    if 'SORT_WAFER' in all_cols:
        wafer_col = 'SORT_WAFER'
        if 'Lot_119325' in all_cols:
            lot_col = 'Lot_119325'
        elif 'Lot_132322' in all_cols:
            lot_col = 'Lot_132322'
        elif 'Lot_1331195' in all_cols:
            lot_col = 'Lot_1331195'

    # Fallback to generic LOT/WAFER
    if not lot_col:
        for c in all_cols:
            if c.upper() == 'LOT' or c.lower().startswith('lot_'):
                lot_col = c
                break
    if not wafer_col:
        for c in all_cols:
            if c.upper() == 'WAFER' or c.upper() == 'SORT_WAFER':
                wafer_col = c
                break

    return lot_col, wafer_col


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def add_material_type(
    csv_path: str,
    collateral_dir: str,
    output_dir: str | None = None,
    log_cb=None,
) -> str:
    """Add material type columns to *csv_path* and return the path to the
    updated CSV.

    The original CSV is never modified.  The enriched copy is written to
    *output_dir* (creating it if needed).  If *output_dir* is None a temp
    directory is used and the caller is responsible for cleanup.

    Parameters
    ----------
    csv_path:
        Path to the yield CSV.
    collateral_dir:
        Folder containing lot-definition CSV files (collateral/material).
    output_dir:
        Directory to write the enriched CSV into.  Defaults to a temp dir.
    log_cb:
        Optional ``callable(msg: str)`` for progress messages.

    Returns
    -------
    str
        Path to the enriched CSV in output_dir, or original csv_path on
        error/skip.
    """

    def _log(msg: str) -> None:
        if log_cb:
            log_cb(msg)
        else:
            print(msg)

    # ── Guard: CSV must exist ─────────────────────────────────────────────────
    if not os.path.isfile(csv_path):
        _log(f'Material type: CSV not found: {csv_path}')
        return csv_path

    # ── Detect encoding ───────────────────────────────────────────────────────
    encoding = detect_encoding(csv_path)

    # ── Step 1: Sniff header ──────────────────────────────────────────────────
    all_cols = sniff_columns(csv_path, encoding=encoding)
    if not all_cols:
        _log('Material type: could not read CSV header – skipping.')
        return csv_path

    # ── Pre-check: only skip if ALL material columns are FULLY populated (zero nulls) ──
    # Do NOT skip when material is only partially filled (e.g. P0 rows have data but
    # R0 rows don't after a multi-CSV merge) — we need to fill the empty rows.
    mat_cols = [c for c in all_cols if 'material' in c.lower()]
    if mat_cols:
        try:
            df_check = pd.read_csv(csv_path, usecols=mat_cols,
                                   encoding=encoding, low_memory=False)
            if df_check[mat_cols].notna().all().all():
                _log('Material type: all material columns fully populated – skipping.')
                return csv_path
            elif df_check[mat_cols].notna().any().any():
                _log('Material type: material columns partially populated – will fill empty rows.')
            else:
                _log('Material type: Material columns present but empty – will drop and re-merge.')
        except Exception:
            pass

    # ── Step 2: Get ALL unique DevRevStep prefixes across all rows ───────────
    dev_rev_col: str | None = next(
        (c for c in all_cols if c.lower().startswith('devrevstep')), None)
    if not dev_rev_col:
        _log('Material type: no DevRevStep_* column found – skipping.')
        return csv_path

    try:
        df_drv = pd.read_csv(csv_path, usecols=[dev_rev_col],
                             encoding=encoding, low_memory=False)
    except Exception as exc:
        _log(f'Material type: failed to read DevRevStep column: {exc}')
        return csv_path

    non_null_drv = df_drv[dev_rev_col].dropna().astype(str)
    if non_null_drv.empty:
        _log('Material type: DevRevStep_* column is empty – skipping.')
        return csv_path

    # Collect all unique 6-char prefixes present in the CSV
    all_prefixes = list(dict.fromkeys(v[:6] for v in non_null_drv.unique() if len(v) >= 6))
    _log(f'Material type: DevRevStep prefixes found = {all_prefixes}')

    # ── Step 3: Find all candidate material files for every prefix ────────────
    seen_files: set[str] = set()
    candidates: list[str] = []
    for prefix6 in all_prefixes:
        for f in find_material_candidates(collateral_dir, prefix6):
            if f not in seen_files:
                seen_files.add(f)
                candidates.append(f)
    if not candidates:
        _log(f'Material type: no collateral files found for prefixes {all_prefixes} in {collateral_dir}')
        return csv_path

    _log(f'Material type: {len(candidates)} candidate file(s): {[os.path.basename(c) for c in candidates]}')

    # ── Step 4: Detect lot/wafer columns in yield CSV ─────────────────────────
    lot_col, wafer_col = _detect_lot_wafer_columns(all_cols)
    if not lot_col or not wafer_col:
        _log(
            f'Material type: could not detect lot/wafer columns '
            f'(found lot={lot_col!r}, wafer={wafer_col!r}) – skipping.'
        )
        return csv_path

    _log(f'Material type: using lot={lot_col!r}, wafer={wafer_col!r}')

    # ── Step 5: Read yield CSV, derive LOT7/WAFER2 ───────────────────────────
    try:
        df = pd.read_csv(csv_path, encoding=encoding, low_memory=False)
    except Exception as exc:
        _log(f'Material type: failed to read yield CSV: {exc}')
        return csv_path

    df = df.copy()  # defragment before column assignments to suppress PerformanceWarning

    # Derive LOT7 = first 7 characters of lot column (kept in output)
    df['LOT7'] = df[lot_col].astype(str).str[:7]

    # Derive WAFER2 = wafer number as integer, then mod 100.
    # Must convert to numeric first to avoid float-string artifacts
    # (e.g. pandas may store 202 as 202.0 → '202.0'[-2:] = '.0' → 0).
    _wafer_num = pd.to_numeric(df[wafer_col], errors='coerce')
    df['WAFER2'] = (_wafer_num.round().astype('Int64') % 100).astype(float)

    # Also add Production Lot column
    if 'Production Lot' not in df.columns:
        df['Production Lot'] = df[lot_col]

    # Save and drop ALL existing material merge columns before the merge to prevent
    # pandas from creating _x/_y suffix columns.  We restore original non-null values
    # after the merge so pre-existing data (e.g. P0 rows in a merged CSV) is preserved.
    _orig_mat_vals: dict = {}
    for _mc in MATERIAL_MERGE_COLS:
        if _mc in df.columns:
            _orig_mat_vals[_mc] = df[_mc].copy()
            df.drop(columns=[_mc], inplace=True)
    if _orig_mat_vals:
        _log(f'Material type: saving existing columns for restore after merge: {list(_orig_mat_vals)}')

    # ── Step 6: Collect lookup rows from ALL matching candidate files ────────
    # Each material CSV covers different lot numbers; we must search every file
    # so wafers from different lots all get their material type populated.
    lot7_vals = set(df['LOT7'].dropna().unique())
    lookup_frames: list[pd.DataFrame] = []
    used_files: list[str] = []
    all_merge_cols: set[str] = set()

    for material_file in candidates:
        try:
            df_material = pd.read_csv(material_file)
        except Exception as exc:
            _log(f'Material type: could not read {os.path.basename(material_file)}: {exc}')
            continue

        if 'INTEL_LOT7' not in df_material.columns or 'WaferID' not in df_material.columns:
            _log(f'Material type: {os.path.basename(material_file)} missing INTEL_LOT7/WaferID – skipping.')
            continue

        merge_cols_available = [c for c in MATERIAL_MERGE_COLS if c in df_material.columns]
        if not merge_cols_available:
            _log(f'Material type: {os.path.basename(material_file)} has no merge columns – skipping.')
            continue

        df_material['_WaferID_num'] = pd.to_numeric(df_material['WaferID'], errors='coerce')
        df_lookup = df_material[['INTEL_LOT7', '_WaferID_num'] + merge_cols_available].copy()
        # Truncate INTEL_LOT7 to first 7 chars to match yields that use only 7-char lot IDs
        df_lookup['INTEL_LOT7'] = df_lookup['INTEL_LOT7'].astype(str).str[:7]
        df_lookup = df_lookup.drop_duplicates(subset=['INTEL_LOT7', '_WaferID_num'])

        mat_lots = set(df_lookup['INTEL_LOT7'].dropna().unique())
        if not lot7_vals.intersection(mat_lots):
            _log(f'Material type: {os.path.basename(material_file)} – no matching lots, skipping.')
            continue

        lookup_frames.append(df_lookup)
        used_files.append(material_file)
        all_merge_cols.update(merge_cols_available)
        _log(f'Material type: {os.path.basename(material_file)} – matched lots, adding to lookup.')

    df_merged = None
    used_file = used_files[0] if used_files else None

    if lookup_frames:
        # Combine all matching lookup tables; later rows win on duplicate keys
        merge_cols_available = [c for c in MATERIAL_MERGE_COLS if c in all_merge_cols]
        combined_lookup = pd.concat(lookup_frames, ignore_index=True, sort=False)
        # Fill any missing merge columns with NaN so concat doesn't drop them
        for _mc in merge_cols_available:
            if _mc not in combined_lookup.columns:
                combined_lookup[_mc] = None
        combined_lookup = combined_lookup.drop_duplicates(
            subset=['INTEL_LOT7', '_WaferID_num'], keep='last'
        )
        # Split lookup into rows with WaferID and rows without (lot-level wildcard)
        _lkp_with_wafer = combined_lookup[combined_lookup['_WaferID_num'].notna()]
        _lkp_lot_only   = combined_lookup[combined_lookup['_WaferID_num'].isna()].drop_duplicates(subset=['INTEL_LOT7'], keep='last')

        # Pass 1: merge on lot + wafer (precise)
        df_merged = df.merge(
            _lkp_with_wafer[['INTEL_LOT7', '_WaferID_num'] + merge_cols_available],
            left_on=['LOT7', 'WAFER2'],
            right_on=['INTEL_LOT7', '_WaferID_num'],
            how='left',
        )
        df_merged.drop(columns=['INTEL_LOT7', '_WaferID_num'], inplace=True, errors='ignore')

        # Pass 2: for rows still missing material, fall back to lot-only wildcard rows
        if not _lkp_lot_only.empty and merge_cols_available:
            _still_missing = df_merged[merge_cols_available[0]].isna()
            if _still_missing.any():
                _lkp_lo = _lkp_lot_only[['INTEL_LOT7'] + merge_cols_available].rename(columns={'INTEL_LOT7': '_lot7_lo'})
                _fallback = df_merged.loc[_still_missing, ['LOT7']].merge(
                    _lkp_lo, left_on='LOT7', right_on='_lot7_lo', how='left'
                ).drop(columns=['LOT7', '_lot7_lo'])
                _fallback.index = df_merged.index[_still_missing]
                for _mc in merge_cols_available:
                    if _mc in _fallback.columns:
                        df_merged.loc[_still_missing, _mc] = _fallback[_mc].values
        # Restore original non-null material values: rows that already had material
        # data (e.g. P0 rows) keep their original values; empty rows (e.g. R0 rows)
        # get the newly merged values.
        df_merged = df_merged.reset_index(drop=True)
        for _mc, _orig_s in _orig_mat_vals.items():
            _orig_s = _orig_s.reset_index(drop=True)
            if _mc in df_merged.columns:
                # Keep original where non-null, else use merged result
                df_merged[_mc] = _orig_s.where(_orig_s.notna(), df_merged[_mc])
            else:
                df_merged[_mc] = _orig_s
        if len(used_files) > 1:
            _log(f'Material type: merged lookup from {len(used_files)} files: {[os.path.basename(f) for f in used_files]}')

    if df_merged is None:
        _log(f'Material type: no matching lots found in any candidate file – leaving columns empty.')
        # Still add empty columns so downstream code doesn't break
        for col in MATERIAL_MERGE_COLS:
            if col not in df.columns:
                df[col] = None
        df_merged = df

    n_matched = df_merged[MATERIAL_MERGE_COLS[0]].notna().sum() if MATERIAL_MERGE_COLS[0] in df_merged.columns else 0
    n_before = len(df_merged)
    if used_files:
        _log(f'Material type: used {len(used_files)} file(s), matched {n_matched}/{n_before} rows.')

    # ── Step 7: Write to output_dir (never modify the original) ─────────────
    try:
        if output_dir is None:
            import tempfile as _tmp
            output_dir = _tmp.mkdtemp(prefix='material_tmp_')
        os.makedirs(output_dir, exist_ok=True)
        out_name = Path(csv_path).stem + '_material_merged.csv'  # intermediate; reticle step renames to _reticle_material.csv
        out_path = os.path.join(output_dir, out_name)
        df_merged.to_csv(out_path, index=False, encoding=encoding)
        _log(f'Material type: enriched CSV saved to {out_path}')
        return out_path
    except Exception as exc:
        _log(f'Material type: failed to write CSV: {exc}')
        return csv_path


# ════════════════════════════════════════════════════════════════
# (formerly apply_reticle_mapping.py)
# ════════════════════════════════════════════════════════════════
# =============================================================================
# apply_reticle_mapping.py  -  Reticle mapping merge for yield CSV
# =============================================================================
#
# Steps:
# // 1. Take the CSV defined in "Output CSV" in GUI (passed as csv_path).
# // 2. Extract Data from DevRevStep_* column. All rows will have same value.
# //    Use the first non-null row value to identify the product prefix.
# // 3. In the Collateral folder, look for filename that contains the 1st 6
# //    characters of the DevRevStep_* value. Open that file.
# // 4. Rename DieX and DieY columns as SORT_X and SORT_Y.
# // 5. Convert to known center die using:
# //       offset_x = round((DieX.min() + DieX.max()) / 2)
# //       offset_y = round((DieY.min() + DieY.max()) / 2)
# //       SORT_X = DieX - offset_x
# //       SORT_Y = DieY - offset_y
# // 6. Merge reticle data into the output CSV based on SORT_X and SORT_Y.
# //    Only merge fields: Layout, Device, LayoutX, LayoutY, ReticleDieX,
# //    ReticleDieY, Reticle. If these fields are already present, skip merge.
# // 7. Use the merged CSV for further analysis. Copy to output folder;
# //    zip the output folder once all analysis is complete.
# // 8. Provide checkbox to save merged file. If checked, save as
# //    <Output CSV>_reticle_merged.csv in the same folder as Output CSV.
# //    Default is false.
# =============================================================================


import os
import zipfile
from pathlib import Path

import pandas as pd


# Reticle mapping columns merged into the output CSV
RETICLE_MERGE_COLS = [
    'Layout', 'Device', 'LayoutX', 'LayoutY',
    'ReticleDieX', 'ReticleDieY', 'Reticle',
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_collateral_file(collateral_dir: str, prefix6: str) -> str | None:
    """Return the first file in collateral_dir whose name contains prefix6
    (case-insensitive).  Returns None when the folder is missing or empty."""
    if not os.path.isdir(collateral_dir):
        return None
    prefix_upper = prefix6.upper()
    for fname in sorted(os.listdir(collateral_dir)):
        if prefix_upper in fname.upper():
            return os.path.join(collateral_dir, fname)
    return None


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def apply_reticle_mapping(
    csv_path: str,
    collateral_dir: str,
    save_merged: bool = False,
    output_dir: str | None = None,
    log_cb=None,
    chunksize: int | None = None,
) -> str:
    """Apply reticle mapping to *csv_path* and return the path to the
    CSV that should be used for further analysis.

    Parameters
    ----------
    csv_path:
        Path to the AQUA output CSV ("Output CSV" from the GUI).
    collateral_dir:
        Folder containing Reticle_Mapping CSV files.
    save_merged:
        If True, also save a ``<stem>_reticle_merged.csv`` next to *csv_path*.
    output_dir:
        If provided, write the merged CSV here for use by downstream
        analysis steps.  The caller should zip this folder when done.
    log_cb:
        Optional ``callable(msg: str)`` for progress messages.
    chunksize:
        Rows per chunk when streaming the AQUA CSV.  Defaults to
        ``csv_utils.CHUNK_SIZE`` (100 000).

    Returns
    -------
    str
        Path to the CSV intended for further analysis:
        - *output_dir* copy when output_dir is given
        - the original *csv_path* on any error or when merge is skipped
    """

    def _log(msg: str) -> None:
        if log_cb:
            log_cb(msg)
        else:
            print(msg)

    _chunksize = chunksize or CHUNK_SIZE

    # ── Guard: CSV must exist ─────────────────────────────────────────────────
    if not os.path.isfile(csv_path):
        _log(f'Reticle mapping: CSV not found: {csv_path}')
        return csv_path

    # ── Detect encoding once — reused for all reads of this file ─────────────
    encoding = detect_encoding(csv_path)

    # ── Step 1 & 2: Sniff header, find DevRevStep_* col, read ALL unique values ─
    # Peak RAM: header only (no data rows loaded yet).
    all_cols = sniff_columns(csv_path, encoding=encoding)
    if not all_cols:
        _log('Reticle mapping: could not read CSV header – skipping.')
        return csv_path

    dev_rev_col: str | None = next(
        (c for c in all_cols if c.lower().startswith('devrevstep')), None)
    if not dev_rev_col:
        _log('Reticle mapping: no DevRevStep_* column found – skipping.')
        return csv_path

    # ── Step 6 pre-check: skip only when ALL merge cols are present AND fully
    # populated (no null rows needing enrichment).  A previous partial-run
    # (e.g. only one product was processed) leaves null rows that must still
    # be merged – so we only bail when coverage is complete.
    if all(c in all_cols for c in RETICLE_MERGE_COLS):
        # Quick sample-check: read a modest slice and see whether any row
        # has null in the first merge col.  If fully populated, skip merge.
        try:
            _sample = pd.read_csv(csv_path, usecols=[RETICLE_MERGE_COLS[0]],
                                  encoding=encoding, low_memory=False)
            if _sample[RETICLE_MERGE_COLS[0]].notna().all():
                _log('Reticle mapping: all merge columns already fully populated – skipping merge.')
                return csv_path
            _log('Reticle mapping: merge columns present but some rows are null – re-merging.')
        except Exception:
            pass  # on any error, proceed with merge

    # Read the full DevRevStep column to collect all unique prefixes
    try:
        df_drv = pd.read_csv(csv_path, usecols=[dev_rev_col],
                             encoding=encoding, low_memory=False)
    except Exception as exc:
        _log(f'Reticle mapping: failed to read DevRevStep column: {exc}')
        return csv_path

    non_null_drv = df_drv[dev_rev_col].dropna().astype(str)
    if non_null_drv.empty:
        _log('Reticle mapping: DevRevStep_* column is empty – skipping.')
        return csv_path

    # Collect all unique 6-char prefixes (preserving order)
    all_prefixes = list(dict.fromkeys(v[:6] for v in non_null_drv.unique() if len(v) >= 6))
    _log(f'Reticle mapping: DevRevStep prefixes found = {all_prefixes}')

    # ── Step 3: Build a combined reticle lookup for all prefixes ─────────────
    # Each prefix maps to its own reticle file; offsets are computed per-file.
    # We store a dict: prefix6 → df_reticle (with SORT_X, SORT_Y already computed)
    prefix_reticle: dict[str, pd.DataFrame] = {}
    for prefix6 in all_prefixes:
        collateral_file = find_collateral_file(collateral_dir, prefix6)
        if not collateral_file:
            _log(f'Reticle mapping: no collateral file for prefix {prefix6!r} – rows with this prefix will be unmatched.')
            continue
        _log(f'Reticle mapping: prefix {prefix6!r} → {os.path.basename(collateral_file)}')
        try:
            df_ret = pd.read_csv(collateral_file)
        except Exception as exc:
            _log(f'Reticle mapping: failed to read {os.path.basename(collateral_file)}: {exc}')
            continue
        if 'DieX' not in df_ret.columns or 'DieY' not in df_ret.columns:
            _log(f'Reticle mapping: {os.path.basename(collateral_file)} missing DieX/DieY – skipping.')
            continue
        df_ret = df_ret.copy()
        die_x = df_ret['DieX'].astype(float)
        die_y = df_ret['DieY'].astype(float)
        offset_x = round((die_x.min() + die_x.max()) / 2)
        offset_y = round((die_y.min() + die_y.max()) / 2)
        _log(f'Reticle mapping: {prefix6!r} offsets  offset_x={offset_x}, offset_y={offset_y}')
        df_ret['SORT_X'] = (die_x - offset_x).astype(int)
        df_ret['SORT_Y'] = (die_y - offset_y).astype(int)
        df_ret = df_ret.drop(columns=['DieX', 'DieY'])
        avail = [c for c in RETICLE_MERGE_COLS if c in df_ret.columns]
        if not avail:
            _log(f'Reticle mapping: {os.path.basename(collateral_file)} has no merge columns – skipping.')
            continue
        keep = ['SORT_X', 'SORT_Y'] + avail
        df_ret = df_ret[keep].drop_duplicates(subset=['SORT_X', 'SORT_Y'])
        prefix_reticle[prefix6] = df_ret

    if not prefix_reticle:
        _log('Reticle mapping: no usable collateral files found – skipping.')
        return csv_path

    # If all prefixes share the same reticle (identical offsets / same file),
    # build one combined lookup (tagged with _prefix6 column internally).
    # For the streaming merge we need to know which prefix each row belongs to.
    # We add a temporary column _R_PREFIX to the chunk for joining.

    available_merge_cols = sorted(
        {c for df_r in prefix_reticle.values() for c in df_r.columns if c not in ('SORT_X', 'SORT_Y')}
    )
    available_merge_cols = [c for c in RETICLE_MERGE_COLS if c in available_merge_cols]

    # ── Check SORT_X / SORT_Y exist in output CSV ─────────────────────────────
    if 'SORT_X' not in all_cols or 'SORT_Y' not in all_cols:
        _log('Reticle mapping: output CSV missing SORT_X/SORT_Y columns – skipping merge.')
        return csv_path

    # ── Step 7: Determine output paths ───────────────────────────────────────
    merged_name = Path(csv_path).stem + '_reticle_material.csv'
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        analysis_csv = os.path.join(output_dir, merged_name)
    else:
        analysis_csv = str(Path(csv_path).with_name(merged_name))

    # ── Step 6 (streaming): merge AQUA CSV in chunks ─────────────────────────
    # When multiple prefixes are present each chunk row is matched to the
    # correct reticle lookup via its prefix6.  If only one prefix exists the
    # fast single-table path is used.

    partial_reticle_cols = [c for c in RETICLE_MERGE_COLS if c in all_cols]

    first_chunk = True
    matched_total = 0
    row_total = 0

    _log(f'Reticle mapping: streaming {_chunksize:,}-row chunks → {analysis_csv}')

    try:
        for chunk in iter_chunks(csv_path, chunksize=_chunksize, encoding=encoding):
            if partial_reticle_cols:
                chunk = chunk.drop(columns=partial_reticle_cols, errors='ignore')

            # Always split by prefix so rows from product A are never merged
            # against product B's reticle table (the old single_prefix fast-path
            # caused one product's data to eclipse the other when only one
            # collateral file was found for multiple DevRevStep prefixes).
            chunk = chunk.copy()
            chunk['_R_PREFIX'] = chunk[dev_rev_col].astype(str).str[:6]
            parts = []
            for pfx, sub in chunk.groupby('_R_PREFIX', sort=False):
                df_r = prefix_reticle.get(pfx)
                if df_r is not None:
                    # Preserve original row order through the merge by saving
                    # and restoring the chunk index (pandas merge resets it).
                    orig_idx = sub.index
                    merged_sub = sub.reset_index(drop=True).merge(
                        df_r, on=['SORT_X', 'SORT_Y'], how='left')
                    merged_sub.index = orig_idx
                else:
                    merged_sub = sub.copy()
                    for mc in available_merge_cols:
                        if mc not in merged_sub.columns:
                            merged_sub[mc] = None
                parts.append(merged_sub)
            chunk_merged = pd.concat(parts).sort_index()
            chunk_merged = chunk_merged.drop(columns=['_R_PREFIX'], errors='ignore')

            if available_merge_cols:
                matched_total += int(chunk_merged[available_merge_cols[0]].notna().sum())
            row_total += len(chunk_merged)

            chunk_merged.to_csv(
                analysis_csv,
                mode='a' if not first_chunk else 'w',
                index=False,
                header=first_chunk,
            )
            first_chunk = False

    except Exception as exc:
        _log(f'Reticle mapping: streaming merge failed: {exc}')
        return csv_path

    _log(f'Reticle mapping: {matched_total:,}/{row_total:,} rows matched.')

    # ── Step 8: Optionally zip the merged CSV inside the output folder ─────────
    if save_merged and os.path.isfile(analysis_csv):
        _zip_path = str(Path(analysis_csv).with_suffix('.zip'))
        try:
            with zipfile.ZipFile(_zip_path, 'w', zipfile.ZIP_DEFLATED) as _zf:
                _zf.write(analysis_csv, Path(analysis_csv).name)
            _log(f'Reticle mapping: zipped merged CSV → {_zip_path}')
        except Exception as exc:
            _log(f'Reticle mapping: could not zip merged CSV: {exc}')

    _log(f'Reticle mapping: analysis CSV → {analysis_csv}')
    return analysis_csv


# ---------------------------------------------------------------------------
# Zip helper  (called at the end of the pipeline)
# ---------------------------------------------------------------------------

def zip_output_folder(output_dir: str, log_cb=None) -> str | None:
    """Zip *output_dir* into ``<output_dir>.zip``.  Returns the zip path, or
    None on failure."""

    def _log(msg: str) -> None:
        if log_cb:
            log_cb(msg)
        else:
            print(msg)

    if not os.path.isdir(output_dir):
        return None

    zip_path = output_dir.rstrip('/\\') + '.zip'
    try:
        base = Path(output_dir)
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fpath in sorted(base.rglob('*')):
                if fpath.is_file():
                    zf.write(fpath, fpath.relative_to(base.parent))
        _log(f'Reticle mapping: zipped output folder → {zip_path}')
        return zip_path
    except Exception as exc:
        _log(f'Reticle mapping: could not zip output folder: {exc}')
        return None




    def __init__(
        self,
        user_id: str,
        base_url: str = AQUA_BASE_URL,
        cert: str | bool = True,
        proxies: Optional[dict] = None,
        poll_interval_s: int = DEFAULT_POLL_INTERVAL_S,
        max_wait_s: int = DEFAULT_MAX_WAIT_S,
    ):
        """
        Parameters
        ----------
        user_id       : IDSID, e.g. "snpant"
        base_url      : API root URL
        cert          : Path to IntelChain.pem, True (system), or False (skip verify — not recommended)
        proxies       : e.g. {"https": "http://proxy-us.intel.com:911"}
        poll_interval_s : seconds between status polls
        max_wait_s    : max seconds to wait for job completion
        """
        self.user_id = user_id
        self.base_url = base_url.rstrip("/")
        self.poll_interval_s = poll_interval_s
        self.max_wait_s = max_wait_s

        # Resolve cert
        if isinstance(cert, str) and not Path(cert).exists():
            print(f"[AquaRestClient] Warning: cert file not found at '{cert}', falling back to system CA")
            cert = True
        self._cert = cert

        self._session = requests.Session()
        self._session.verify = self._cert
        if proxies:
            self._session.proxies.update(proxies)

    # ── Low-level API calls ────────────────────────────────────────────────────

    def execute(
        self,
        report_txt_path: str | Path,
        output_path: str,
        email_notification: bool = True,
        format: str = "csv.gz",
    ) -> int:
        """
        POST /api/query/execute/user/{userId}
        Submit a job to run data extraction. Returns jobId.

        Parameters
        ----------
        report_txt_path   : path to exported AQUA report config .txt file
        output_path       : network/shared path where output will be written
                            (must be accessible by 'aquajobs' service account)
        email_notification: whether AQUA sends completion email
        format            : output format hint (csv.gz or parquet) — passed as outputpath suffix
        """
        report_txt_path = Path(report_txt_path)
        if not report_txt_path.exists():
            raise FileNotFoundError(f"Report config not found: {report_txt_path}")

        url = f"{self.base_url}/api/query/execute/user/{self.user_id}"
        params = {"emailNotification": "YES" if email_notification else "NO"}

        with report_txt_path.open("rb") as f:
            files = {"file": (report_txt_path.name, f, "text/plain")}
            data = {"outputpath": output_path}
            resp = self._session.post(url, params=params, files=files, data=data, timeout=60)

        resp.raise_for_status()
        job_id = int(resp.text.strip())
        print(f"[execute] Job submitted: jobId={job_id}")
        return job_id

    def generate(self, report_txt_path: str | Path) -> int:
        """
        POST /api/query/generate/user/{userId}
        Generate Midas HBase query only (no extraction). Returns jobId.
        """
        report_txt_path = Path(report_txt_path)
        if not report_txt_path.exists():
            raise FileNotFoundError(f"Report config not found: {report_txt_path}")

        url = f"{self.base_url}/api/query/generate/user/{self.user_id}"
        with report_txt_path.open("rb") as f:
            files = {"file": (report_txt_path.name, f, "text/plain")}
            resp = self._session.post(url, files=files, timeout=60)

        resp.raise_for_status()
        job_id = int(resp.text.strip())
        print(f"[generate] Job submitted: jobId={job_id}")
        return job_id

    def get_status(self, job_id: int) -> str:
        """
        GET /api/job/{jobId}/status
        Returns one of: Pending, Completed, Fail
        """
        url = f"{self.base_url}/api/job/{job_id}/status"
        resp = self._session.get(url, timeout=30)
        resp.raise_for_status()
        return resp.text.strip()

    def get_status_full(self, job_id: int) -> dict:
        """
        GET /api/job/{jobId}/status/full
        Returns dict: {jobId, status, message, sharedPath}
        """
        url = f"{self.base_url}/api/job/{job_id}/status/full"
        resp = self._session.get(url, timeout=30)
        resp.raise_for_status()
        return resp.json()

    def get_query_json(self, job_id: int) -> str:
        """
        GET /api/query/get/user/{userId}/jobId/{jobId}
        Returns the generated Midas HBase JSON query (after job completes).
        """
        url = f"{self.base_url}/api/query/get/user/{self.user_id}/jobId/{job_id}"
        resp = self._session.get(url, timeout=30)
        resp.raise_for_status()
        return resp.text

    # ── High-level helpers ─────────────────────────────────────────────────────

    def wait_for_job(self, job_id: int) -> dict:
        """
        Poll GET /api/job/{jobId}/status/full until Completed or Fail.
        Returns the full status dict.
        Raises RuntimeError on Fail or timeout.
        """
        elapsed = 0
        print(f"[wait_for_job] Polling jobId={job_id} every {self.poll_interval_s}s ...")
        while elapsed < self.max_wait_s:
            result = self.get_status_full(job_id)
            status = result.get("status", "").lower()
            print(f"  [{elapsed:>5}s] status={result.get('status')}  message={result.get('message', '')}")
            if status == "completed":
                print(f"[wait_for_job] Job {job_id} completed. sharedPath={result.get('sharedPath')}")
                return result
            if status == "fail":
                raise RuntimeError(f"Job {job_id} failed: {result.get('message')}")
            time.sleep(self.poll_interval_s)
            elapsed += self.poll_interval_s

        raise TimeoutError(f"Job {job_id} did not complete within {self.max_wait_s}s")

    def run_and_download(
        self,
        report_txt_path: str | Path,
        output_path: str,
        local_dest: str | Path | None = None,
        email_notification: bool = False,
    ) -> Path | None:
        """
        Full pipeline: submit → poll → (optionally copy file locally).

        Parameters
        ----------
        report_txt_path : AQUA report .txt config
        output_path     : network path AQUA writes result to
                          (grant aquajobs write access)
        local_dest      : if provided, copies the result file(s) here
        email_notification : send email on completion

        Returns
        -------
        Path to local copy if local_dest provided, else None.
        """
        job_id = self.execute(report_txt_path, output_path, email_notification)
        result = self.wait_for_job(job_id)
        shared_path = result.get("sharedPath")

        if local_dest and shared_path:
            local_dest = Path(local_dest)
            local_dest.mkdir(parents=True, exist_ok=True)
            shared = Path(shared_path)
            if shared.is_file():
                dest_file = local_dest / shared.name
                shutil.copy2(shared, dest_file)
                print(f"[run_and_download] Copied to {dest_file}")
                return dest_file
            elif shared.is_dir():
                # Copy all files in the output folder
                copied = []
                for f in shared.iterdir():
                    dest_file = local_dest / f.name
                    shutil.copy2(f, dest_file)
                    copied.append(dest_file)
                print(f"[run_and_download] Copied {len(copied)} file(s) to {local_dest}")
                return local_dest
            else:
                print(f"[run_and_download] sharedPath not accessible: {shared_path}")

        return None


# ── CLI ────────────────────────────────────────────────────────────────────────

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="AQUA REST API client — submit a report and download the result."
    )
    p.add_argument("--user",        required=True,  help="IDSID (e.g. snpant)")
    p.add_argument("--report",      required=True,  help="Path to exported AQUA report .txt config")
    p.add_argument("--outputpath",  required=True,  help="Network path AQUA writes output to (accessible by aquajobs)")
    p.add_argument("--outdir",      default=None,   help="Local dir to copy result into (optional)")
    p.add_argument("--url",         default=AQUA_BASE_URL, help=f"API base URL (default: {AQUA_BASE_URL})")
    p.add_argument("--cert",        default=INTEL_CERT,    help="Path to IntelChain.pem (default: env INTEL_CHAIN_PEM or IntelChain.pem)")
    p.add_argument("--no-email",    action="store_true",   help="Disable completion email notification")
    p.add_argument("--poll",        type=int, default=DEFAULT_POLL_INTERVAL_S, help="Poll interval in seconds (default: 15)")
    p.add_argument("--timeout",     type=int, default=DEFAULT_MAX_WAIT_S,      help="Max wait seconds (default: 3600)")
    p.add_argument("--proxy",       default=os.environ.get("HTTPS_PROXY"),     help="HTTPS proxy (default: HTTPS_PROXY env var)")
    p.add_argument("--status",      type=int, metavar="JOBID", help="Just check status of an existing job ID and exit")
    return p


def main():
    args = _build_parser().parse_args()

    proxies = {"https": args.proxy, "http": args.proxy} if args.proxy else None
    cert = args.cert if Path(args.cert).exists() else True

    client = AquaRestClient(
        user_id=args.user,
        base_url=args.url,
        cert=cert,
        proxies=proxies,
        poll_interval_s=args.poll,
        max_wait_s=args.timeout,
    )

    # Status-only mode
    if args.status:
        result = client.get_status_full(args.status)
        print(result)
        return

    # Full run
    client.run_and_download(
        report_txt_path=args.report,
        output_path=args.outputpath,
        local_dest=args.outdir,
        email_notification=not args.no_email,
    )


if __name__ == "__main__":
    main()


# ════════════════════════════════════════════════════════════════
# (formerly parse_bindef_to_crystalball.py)
# ════════════════════════════════════════════════════════════════
#!/usr/bin/env python3
"""
Parse a BinDefinitions.bdefs file and produce a Crystal Ball CSV.

Usage:
  python parse_bindef_to_crystalball.py --bindef "I:\\...\\BinDefinitions.bdefs" --out "C:\\tp\\workshop\\yield\\crystal_ball_input.csv"

If --out is omitted the output defaults to `crystal_ball_input.csv` in the current folder.
The script will try to preserve the same CSV structure as the attached `51K_bindef.csv`:
 - Header: `B/C,<parent-folder-name> DESCRIPTION`
 - Each mapping row: `FBxxx,DESCRIPTION`
"""

import argparse
import logging
import re
from pathlib import Path
import sys


def parse_line(line: str, current_group: str | None = None) -> tuple[str, str] | None:
    s = line.strip()
    if not s:
        return None
    # ignore comment lines
    if s.startswith("#") or s.startswith(";"):
        return None
    # Prefer explicit Bin or LeafBin definitions with numeric id and quoted label:
    # e.g. Bin b101_pass_NAME   101   : "b101_pass_NAME",... or
    #      LeafBin b10000001 10000001 : "b10000001_..."
    m = re.match(r"^(?:Bin|LeafBin)\s+([^\s]+)\s+(\d+)\s*:\s*\"([^\"]+)\"", s, re.IGNORECASE)
    if m:
        name_token = m.group(1).strip()
        num = m.group(2).strip()
        quoted = m.group(3).strip()
        # Section-specific behavior:
        group = (current_group or "").lower()
        if group == "softbins" or group == "soft_bins" or group == "passfailbins":
            key = f"FB{num}"
            val = name_token.upper()
            return key, val
        elif group == "databins" or group == "data_bins" or group == "leafbins":
            key = f"DB{num}"
            # use token name uppercased (convert leading 'b' to 'B')
            val = name_token.upper()
            return key, val
        else:
            # default to FB to preserve previous behavior
            key = f"FB{num}"
            val = quoted.upper()
            return key, val

    # Detect explicit DB keys (e.g., DB20000001) and their labels. Many bindef
    # files list DB identifiers that the yield CSV uses; map those to labels.
    # Prefer quoted labels when present.
    mdb = re.search(r"\b(DB\d{5,})\b", s, re.IGNORECASE)
    if mdb:
        dbkey = mdb.group(1).upper()
        # Try to extract a quoted label on the same line
        mq = re.search(r'"([^\"]+)"', s)
        if mq:
            return dbkey, mq.group(1).upper()
        # If comma-separated, take second field
        if "," in s:
            parts = s.split(",", 1)
            return dbkey, parts[1].strip()
        # Fallback: remove the key token and use the rest of the line
        rest = re.sub(re.escape(mdb.group(0)), "", s).strip(" ,:-")
        if rest:
            return dbkey, rest.strip()

    # If already CSV-like (but not a Bin definition line that contains commas),
    # treat as preformatted CSV. Many Bin lines include a comma after the quoted
    # label (e.g. '"label",Pass;') so avoid splitting those by checking for
    # lines that start with 'Bin '.
    if "," in s and not s.lower().startswith("bin "):
        parts = s.split(",", 1)
        key = parts[0].strip()
        val = parts[1].strip()
        if key:
            return key, val

    # Try whitespace-separated fallbacks (existing behavior)
    m2 = re.match(r"^(FB\d+|B\d+|FB\w+|B\w+)\s+(.+)$", s, re.IGNORECASE)
    if m2:
        return m2.group(1).strip(), m2.group(2).strip()

    # Try colon or equals
    for sep in [":", "=", "-"]:
        if sep in s:
            parts = s.split(sep, 1)
            key = parts[0].strip()
            val = parts[1].strip()
            if re.match(r"^(FB\d+|B\d+|FB\w+|B\w+)$", key, re.IGNORECASE):
                return key, val

    # As a last resort, look for first token as key
    toks = s.split()
    if toks and re.match(r"^(FB\d+|B\d+|FB\w+|B\w+)$", toks[0], re.IGNORECASE):
        return toks[0], " ".join(toks[1:]).strip()
    return None


def build_header(bindef_path: Path) -> str:
    parent = bindef_path.parent.name
    # follow attached file header pattern
    return f"B/C,{parent} DESCRIPTION"


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Parse BinDefinitions.bdefs into crystal_ball_input.csv")
    p.add_argument("--bindef", "-b", required=True, help="Path to BinDefinitions.bdefs")
    p.add_argument("--out", "-o", help="Output CSV path (default: crystal_ball_input.csv in current folder)")
    p.add_argument("--log", default="INFO", help="Log level")
    args = p.parse_args(argv)

    logging.basicConfig(level=getattr(logging, args.log.upper(), logging.INFO), format="%(levelname)s: %(message)s")

    bindef = Path(args.bindef)
    if not bindef.exists():
        logging.error("Bindef not found: %s", bindef)
        return 2

    out_path = Path(args.out) if args.out else Path.cwd() / "crystal_ball_input.csv"

    entries: list[tuple[str, str]] = []
    current_group = None
    with bindef.open("r", encoding="utf-8", errors="replace") as fh:
        for line in fh:
            # track BinGroup sections
            mgrp = re.match(r"^\s*BinGroup\s*,?\s*([A-Za-z0-9_]+)", line, re.IGNORECASE)
            if mgrp:
                current_group = mgrp.group(1).strip()
                continue
            parsed = parse_line(line, current_group=current_group)
            if parsed:
                k, v = parsed
                # If we're in a DataBins section, normalize keys to DB{num}
                grp = (current_group or "").lower()
                if "data" in grp:
                    # If key already starts with DB keep it
                    if not re.match(r"^DB\d+", k, re.IGNORECASE):
                        num = None
                        # Prefer numeric id inside the value if it starts with B<digits>
                        mv = re.match(r"^[bB](\d{5,})", v)
                        if mv:
                            num = mv.group(1)
                        else:
                            # fallback to numeric inside the key
                            mk = re.search(r"(\d{5,})", k)
                            if mk:
                                num = mk.group(1)
                        if num:
                            k = f"DB{num}"
                entries.append((k, v))

    if not entries:
        logging.error("No valid entries parsed from %s", bindef)
        return 3

    header = build_header(bindef)
    with out_path.open("w", encoding="utf-8", newline="") as out:
        out.write(header + "\n")
        for k, v in entries:
            out.write(f"{k},{v}\n")

    logging.info("Wrote %d entries to %s", len(entries), out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
_parse_bindef_main = main  # stable alias for yield_pipeline in-process call


# ════════════════════════════════════════════════════════════════
# (formerly make_portable_dashboard.py)
# ════════════════════════════════════════════════════════════════
#!/usr/bin/env python3
"""
make_portable_dashboard.py
--------------------------
Creates a self-contained portable copy of Dashboard.html.

Strategy per link type
  * Relative HTML hrefs  → Blob URL (JS runs properly)
  * file:// local HTML   → Blob URL (resolved to local path)
  * Relative images/CSS  → base64 data URI
  * iframe src (html)    → srcdoc  (initial load only)
  * load('x.html') JS   → pre-embedded blob map + load() override
  * xlsx / .jmp / .jmpprj / http://127.0.0.1 → disabled button

Usage:
    python src/make_portable_dashboard.py <Dashboard.html> [--out <output.html>]
"""

import argparse
import base64
import html as html_mod
import json
import mimetypes
import os
import re
import sys
from pathlib import Path


_MIME_DEFAULTS = {
    '.html': 'text/html',
    '.htm':  'text/html',
    '.png':  'image/png',
    '.jpg':  'image/jpeg',
    '.jpeg': 'image/jpeg',
    '.gif':  'image/gif',
    '.svg':  'image/svg+xml',
    '.css':  'text/css',
    '.js':   'application/javascript',
    '.ico':  'image/x-icon',
    '.webp': 'image/webp',
}

# Extensions that should be disabled (can't be embedded usefully)
_DISABLE_EXTS = {'.xlsx', '.xls', '.jmp', '.jmpprj', '.csv', '.zip', '.sas7bdat'}

# Regex that matches a full <a ...> tag including > chars inside quoted attributes
# e.g. onclick="catch(()=>...)"  — [^>]+ would break on the =>)
_ANCHOR_RE = re.compile(
    r'<a\b(?:[^>"\']*("[^"]*"|\'[^\']*\'))*[^>]*>',
    re.DOTALL)


def _mime(path: Path) -> str:
    return _MIME_DEFAULTS.get(
        path.suffix.lower(),
        mimetypes.guess_type(str(path))[0] or 'application/octet-stream',
    )


def _to_data_uri(path: Path, mime: str = None) -> str:
    mime = mime or _mime(path)
    data = base64.b64encode(path.read_bytes()).decode('ascii')
    return f'data:{mime};base64,{data}'


def _is_external(url: str) -> bool:
    return (url.startswith('data:')
            or url.startswith('http:')
            or url.startswith('https:')
            or url.startswith('//')
            or url.startswith('file://')
            or url.startswith('#')
            or url.startswith('javascript:')
            or url.startswith('mailto:'))


def _resolve_file_url(url: str) -> Path | None:
    """Convert file:// URL to a local Path if it exists.
    Handles both file:///local/path and file://server/unc/path forms."""
    try:
        from urllib.parse import unquote
        if url.startswith('file:///'):
            # Local path: file:///C:/... or file:///path/...
            from urllib.request import url2pathname
            path_str = url2pathname(url[7:])  # strip "file://"
            p = Path(path_str)
            return p if p.is_file() else None
        elif url.startswith('file://'):
            # UNC path: file://server/share/path → \\server\share\path
            rest = unquote(url[7:])   # strip "file://"
            unc  = '\\\\' + rest.replace('/', '\\')
            p = Path(unc)
            return p if p.is_file() else None
    except Exception:
        pass
    return None


def _html_to_b64(html_str: str) -> str:
    return base64.b64encode(html_str.encode('utf-8')).decode('ascii')


# ---------------------------------------------------------------------------
# Embed static resources (images, CSS) inside an HTML string.
# HTML references are NOT touched here — handled elsewhere.
# ---------------------------------------------------------------------------
def _embed_static(html: str, base: Path, depth: int = 0) -> str:
    """Embed img/script src (non-HTML) and inline CSS link tags."""
    if depth > 8:
        return html

    def _replace_src(m):
        src = m.group(1)
        if _is_external(src):
            return m.group(0)
        if src.lower().split('?')[0].endswith(('.html', '.htm')):
            return m.group(0)  # handled separately
        try:
            p = (base / src).resolve()
            if p.is_file():
                return f'src="{_to_data_uri(p)}"'
        except Exception:
            pass
        return m.group(0)

    html = re.sub(r'\bsrc="([^"]+)"', _replace_src, html)

    def _replace_link(m):
        href = m.group(1)
        if _is_external(href):
            return m.group(0)
        try:
            p = (base / href).resolve()
            if p.is_file() and p.suffix.lower() == '.css':
                css = p.read_text(encoding='utf-8', errors='replace')
                return f'<style>{css}</style>'
        except Exception:
            pass
        return m.group(0)

    html = re.sub(
        r'<link\b[^>]*\brel=["\']stylesheet["\'][^>]*\bhref="([^"]+)"[^>]*/?>',
        _replace_link, html, flags=re.IGNORECASE)

    return html


# ---------------------------------------------------------------------------
# Disable non-embeddable links (xlsx, jmp, local server)
# ---------------------------------------------------------------------------
_DISABLED_STYLE = ('style="opacity:0.4;pointer-events:none;cursor:default;'
                   'text-decoration:line-through" title="Not available in portable version"')


def _check_anchor(m):
    full = m.group(0)
    # extract href value
    hm = re.search(r'\bhref="([^"]+)"', full)
    if not hm:
        return full
    href = hm.group(1)
    # local opener server (127.0.0.1) — always disable
    if re.search(r'127\.0\.0\.1:\d+/open', href):
        return _apply_disabled(full)
    # file:// — both file:/// (local) and file://server/ (UNC)
    if href.startswith('file://'):
        ext = Path(href.split('?')[0].split('/')[-1]).suffix.lower()
        if ext in _DISABLE_EXTS:
            return _apply_disabled(full)
    if not _is_external(href):
        ext = Path(href.split('?')[0]).suffix.lower()
        if ext in _DISABLE_EXTS:
            return _apply_disabled(full)
    return full


def _apply_disabled(tag: str) -> str:
    """Inject disabled style and neutralise onclick/href on an <a> tag."""
    tag = re.sub(r'\bhref="[^"]*"', 'href="#"', tag)
    tag = re.sub(r'\bonclick="[^"]*"', '', tag)
    if 'style="' in tag:
        tag = tag.replace('style="', 'style="opacity:0.4;pointer-events:none;cursor:default;text-decoration:line-through;', 1)
    else:
        tag = tag.rstrip('>') + f' {_DISABLED_STYLE}>'
    return tag


def _disable_links(html: str) -> str:
    return _ANCHOR_RE.sub(_check_anchor, html)


# ---------------------------------------------------------------------------
# Embed a sidebar-viewer index.html:
# Pre-embed all load('x.html') targets as blobs, override window.load().
# Also embed iframe initial src as srcdoc.
# ---------------------------------------------------------------------------
def _embed_viewer_html(html: str, base: Path, depth: int) -> str:
    """Fully embed a load()-based sidebar viewer."""
    # Collect all load('...') file references
    load_targets = list(dict.fromkeys(re.findall(r"load\('([^']+\.html?)'", html)))

    # Also collect initial iframe src
    iframe_srcs = re.findall(r'<iframe\b[^>]+\bsrc="([^"]+\.html?)"', html, re.IGNORECASE)

    # Also collect direct href="*.html" anchors with target="content" (e.g. wafer map)
    href_content = re.findall(
        r'href="([^"#][^"]*\.html?)"[^>]*target=["\']content["\']',
        html, re.IGNORECASE)
    href_content += re.findall(
        r'target=["\']content["\'][^>]{0,200}href="([^"#][^"]*\.html?)"',
        html, re.IGNORECASE)

    all_targets = list(dict.fromkeys(load_targets + iframe_srcs + href_content))

    # ── Scan sub-pages for paretoNav / FP_DATA URLs and wmLoad URLs ──────
    # paretoNav targets must be in the PARENT viewer's __pl map because
    # paretoNav calls window.parent.frame.src.
    # wmLoad targets are consumed within the sub-page (local blob map).
    _wm_load_map: dict = {}   # rel → [wm_url, ...]

    for rel in list(all_targets):          # iterate a snapshot; we'll append below
        p = (base / rel)
        if not p.is_file():
            continue
        try:
            sub_raw = p.read_text(encoding='utf-8', errors='replace')
        except Exception:
            continue

        # paretoNav('url') — static onclick rows
        # Use pure relative path arithmetic (avoid resolve() to stay on mapped drive)
        for pnt in re.findall(r"paretoNav\('([^']+\.html?)'\)", sub_raw):
            pnt_rel = str(Path(rel).parent / pnt).replace('\\', '/')
            if pnt_rel not in all_targets and (base / pnt_rel).is_file():
                all_targets.append(pnt_rel)

        # "url":"..." keys inside FP_DATA / similar JS JSON blobs
        for fp_url in re.findall(r'"url"\s*:\s*"([^"]+\.html?)"', sub_raw):
            fp_rel = str(Path(rel).parent / fp_url).replace('\\', '/')
            if fp_rel not in all_targets and (base / fp_rel).is_file():
                all_targets.append(fp_rel)

        # WM_FILES={"lot":"heatmap/...html"} — wafer map targets for fbTileClick
        # Also pick up WM_URL="wafermap.html" fallback
        if 'WM_FILES' in sub_raw or 'WM_URL' in sub_raw:
            # WM_URL single fallback
            _wm_url_m = re.search(r'var\s+WM_URL\s*=\s*"([^"]+\.html?)"', sub_raw)
            if _wm_url_m:
                _wu_rel = str(Path(rel).parent / _wm_url_m.group(1)).replace('\\', '/')
                if _wu_rel not in all_targets and (base / _wu_rel).is_file():
                    all_targets.append(_wu_rel)
            # WM_FILES dict values
            _wm_block = re.search(r'var\s+WM_FILES\s*=\s*(\{[^}]*\})', sub_raw)
            if _wm_block:
                for wf_url in re.findall(r'"([^"]+\.html?)"', _wm_block.group(1)):
                    wf_rel = str(Path(rel).parent / wf_url).replace('\\', '/')
                    if wf_rel not in all_targets and (base / wf_rel).is_file():
                        all_targets.append(wf_rel)

        # wmLoad('url') — iframe targets served within the sub-page itself
        wm_urls = list(dict.fromkeys(re.findall(r"wmLoad\('([^']+)'", sub_raw)))
        if wm_urls:
            _wm_load_map[rel] = wm_urls

    # Build base64 map: relative path → b64 of (recursively embedded) HTML
    file_map = {}
    for rel in all_targets:
        p = base / rel
        if p.is_file():
            inner = p.read_text(encoding='utf-8', errors='replace')
            inner = _embed_static(inner, p.parent, depth + 1)
            inner = _disable_links(inner)

            # ── Inject paretoNav override ─────────────────────────────────
            # In portable mode, window.parent.__pl(url) must be used instead
            # of setting window.parent.frame.src (frame uses srcdoc).
            if 'paretoNav' in inner:
                _pn_ovr = (
                    '<script>(function(){'
                    'window.paretoNav=function(url){'
                    'var par=window.parent;'
                    'if(par&&par.__pl){par.__pl(url);return;}'
                    'try{var f=par.document.getElementById(\'frame\');'
                    'if(f){f.src=url;return;}}catch(e){}'
                    'window.open(url,\'_blank\');};'
                    '})();</script>'
                )
                inner = (inner.replace('</head>', _pn_ovr + '</head>', 1)
                         if '</head>' in inner else _pn_ovr + inner)

            # ── Inject fbTileClick portable nav override ──────────────────
            # fbTileClick sets window.parent.frame.src for wafer map navigation.
            # In portable mode, rewrite to use __pl instead.
            if 'fbTileClick' in inner and 'WM_URL' in inner:
                # Patch: replace frame.src assignment with __pl call
                inner = re.sub(
                    r'try\{var f=window\.parent\.document\.getElementById\([\'\"]frame[\'\"]\);if\(f\)\{f\.src=_wmTarget;\}else\{throw 0;\}\}',
                    'try{var par=window.parent;if(par&&par.__pl){par.__pl(_wmTarget);}else{var f=par.document.getElementById(\'frame\');if(f){f.src=_wmTarget;}else{throw 0;}}}',
                    inner
                )

            # ── Inject wmLoad / _wmRender override ───────────────────────
            # _wmRender creates <iframe src=url> for lot/wafer files.
            # Embed those files as a local srcdoc blob map and override
            # _wmRender to use srcdoc= so no on-disk files are needed.
            if rel in _wm_load_map:
                wm_b64: dict = {}
                for wm_url in _wm_load_map[rel]:
                    # Use base/rel-relative path arithmetic, no resolve()
                    wm_abs = base / Path(rel).parent / wm_url
                    if wm_abs.is_file():
                        try:
                            wm_html = wm_abs.read_text(encoding='utf-8', errors='replace')
                            wm_html = _embed_static(wm_html, wm_abs.parent, depth + 2)
                            wm_html = _disable_links(wm_html)
                            wm_b64[wm_url] = _html_to_b64(wm_html)
                        except Exception:
                            pass
                if wm_b64:
                    wm_map_json = json.dumps(wm_b64, separators=(',', ':'))
                    _wm_ovr = (
                        f'<script>(function(){{'
                        f'var _W={wm_map_json};'
                        f'function _dec(b64){{var b=atob(b64),a=new Uint8Array(b.length);'
                        f'for(var i=0;i<b.length;i++)a[i]=b.charCodeAt(i);'
                        f'return new TextDecoder().decode(a);}}'
                        f'function _wmFind(url){{'
                        f'var u=(url||"").replace(/\\\\/g,"/");'
                        f'var base=u.split("#")[0];'
                        f'var b64=_W[u]||_W[base];'
                        f'if(!b64){{'
                        f'var bn=base.split("/").pop();'
                        f'for(var k in _W){{if(k.split("/").pop()===bn){{b64=_W[k];break;}}}}'
                        f'}}'
                        f'return{{b64:b64,frag:(u.indexOf("#")>=0?u.split("#")[1]:"")}};'
                        f'}}'
                        f'window._wmRender=function(){{'
                        f'var wrap=document.getElementById("wm-frames");'
                        f'if(!wrap)return;wrap.innerHTML="";'
                        f'window._wmSel.forEach(function(row,url){{'
                        f'var f=document.createElement("iframe");'
                        f'var r=_wmFind(url);'
                        f'if(r.b64){{'
                        f'f.srcdoc=_dec(r.b64);'
                        f'if(r.frag){{f.addEventListener("load",function(){{'
                        f'try{{var d=f.contentDocument||f.contentWindow.document;'
                        f'var e=d.getElementById(r.frag);if(e)e.scrollIntoView();}}catch(_e){{}}'
                        f'}});}}'
                        f'}}else{{f.src=url;}}'
                        f'wrap.appendChild(f);}});}};'
                        f'}})();</script>'
                    )
                    inner = (inner.replace('</body>', _wm_ovr + '</body>', 1)
                             if '</body>' in inner else inner + _wm_ovr)

            file_map[rel] = _html_to_b64(inner)
        else:
            print(f'    [skip] not found: {base / rel}')

    if not file_map:
        return html

    # Rewrite href="x.html" target="content" anchors → onclick="__pl('x.html',this)"
    def _rewrite_content_href(m):
        tag = m.group(0)
        href_m = re.search(r'href="([^"]+)"', tag)
        if not href_m:
            return tag
        rel = href_m.group(1)
        if rel not in file_map:
            return tag
        tag = re.sub(r'\bhref="[^"]*"', 'href="#"', tag)
        tag = re.sub(r'\btarget=["\']content["\']', '', tag, flags=re.IGNORECASE)
        tag = re.sub(r'\bonclick="[^"]*"', '', tag)
        tag = tag.rstrip('>')
        tag += f" onclick=\"__pl('{rel}',this);return false;\">"
        return tag
    html = _ANCHOR_RE.sub(_rewrite_content_href, html)

    # Replace iframe src= with srcdoc= for the initial load
    def _iframe_srcdoc(m):
        src = m.group(1)
        if src in file_map:
            inner_html = base64.b64decode(file_map[src]).decode('utf-8')
            escaped = html_mod.escape(inner_html, quote=True)
            tag = m.group(0)
            tag = re.sub(r'\bsrc="[^"]*"', f'srcdoc="{escaped}"', tag)
            return tag
        return m.group(0)

    html = re.sub(r'<iframe\b([^>]*)\bsrc="([^"]+\.html?)"([^>]*)>',
                  lambda m: _iframe_srcdoc_full(m, file_map), html, flags=re.IGNORECASE)

    # Rewrite onclick="load('x.html', this)" → onclick="__pl('x.html', this)"
    # This is necessary because function declarations are hoisted and window.load
    # assignments cannot override them; rewriting the call site is the only fix.
    html = re.sub(r"""onclick="load\(""", 'onclick="__pl(', html)

    # Inject JS blob map + portable __pl loader before </body>.
    # Use frame.srcdoc (UTF-8 decoded string) instead of frame.src = blob URL,
    # because URL.createObjectURL is unreliable inside a blob-URL document context.
    map_json = json.dumps(file_map, separators=(',', ':'))
    override_js = f"""
<script id="_portable_viewer">(function(){{
  var _map={map_json};
  function _decode(b64){{
    // Decode base64 → Uint8Array → UTF-8 string (handles non-ASCII)
    var bin=atob(b64),a=new Uint8Array(bin.length);
    for(var i=0;i<bin.length;i++)a[i]=bin.charCodeAt(i);
    return new TextDecoder().decode(a);
  }}
  // __pl: portable load — replaces hoisted function load() via onclick rewrite
  window.__pl=function(url,el){{
    var frame=document.getElementById('frame');
    if(!frame)return;
    var frag='';
    var hashIdx=url.indexOf('#');
    if(hashIdx>=0){{frag=url.substring(hashIdx+1);url=url.substring(0,hashIdx);}}
    var b64=_map[url];
    if(!b64){{
      // try basename match for paths like heatmap/x.html
      var base=url.split('/').pop();
      for(var k in _map){{
        if(k.split('/').pop()===base){{b64=_map[k];break;}}
      }}
    }}
    if(b64){{
      frame.srcdoc=_decode(b64);
      if(frag){{frame.addEventListener('load',function _scrollFrag(){{
        frame.removeEventListener('load',_scrollFrag);
        try{{var d=frame.contentDocument||frame.contentWindow.document;
        var e=d.getElementById(frag);if(e)e.scrollIntoView();}}catch(_e){{}}
      }});}}
    }}
    if(el){{
      document.querySelectorAll('a.nav-link,a.sub-link,a.subsub-link').forEach(function(l){{l.classList.remove('active');}});
      el.classList.add('active');
    }}
  }};
}})();
</script>"""

    if '</body>' in html:
        html = html.replace('</body>', override_js + '\n</body>', 1)
    else:
        html += override_js

    return html


def _iframe_srcdoc_full(m, file_map):
    # group(0) = full tag, need to find src attr
    tag = m.group(0)
    src_m = re.search(r'\bsrc="([^"]+)"', tag)
    if not src_m:
        return tag
    src = src_m.group(1)
    if src in file_map:
        inner_html = base64.b64decode(file_map[src]).decode('utf-8')
        escaped = html_mod.escape(inner_html, quote=True)
        # Use lambda so escaped content is not interpreted as a regex replacement
        # (Python 3.12+ raises re.error on bare \s, \d, etc. in repl strings)
        tag = re.sub(r'\bsrc="[^"]*"', lambda _: f'srcdoc="{escaped}"', tag)
    return tag


# ---------------------------------------------------------------------------
# Top-level: detect if an HTML file is a viewer and apply appropriate embed
# ---------------------------------------------------------------------------
def _is_viewer_html(html: str) -> bool:
    """True if this HTML uses load() to swap iframe content."""
    return bool(re.search(r"window\.load\s*=|function\s+load\s*\(", html)
                or re.search(r"document\.getElementById\(['\"]frame['\"]\)\.src", html))


def _embed_sub_html_hrefs(html: str, base: Path, depth: int) -> str:
    """Embed relative and file:// HTML hrefs inside an already-embedded page
       as _openPortable blobs (new-tab opener). Only goes 3 levels deep."""
    if depth > 3:
        return html
    sub_map = {}
    counter = [0]

    def _sub(m):
        full_tag = m.group(0)
        hm = re.search(r'\bhref="([^"]+)"', full_tag)
        if not hm:
            return full_tag

        # Keep in-viewer navigation anchors intact; _embed_viewer_html handles
        # load('...') targets and rewrites them to __pl() for srcdoc rendering.
        if re.search(r'\bonclick="[^"]*\bload\(', full_tag, re.IGNORECASE):
            return full_tag

        href = hm.group(1)
        if href.startswith('#') or href.startswith('javascript:') or not href:
            return full_tag

        local_path = None
        if href.startswith('file://'):
            ext = Path(href.split('?')[0].split('/')[-1]).suffix.lower()
            if ext in _DISABLE_EXTS:
                return full_tag  # _disable_links will handle
            local_path = _resolve_file_url(href)
        elif _is_external(href):
            return full_tag
        else:
            ext = Path(href.split('?')[0]).suffix.lower()
            if ext in _DISABLE_EXTS:
                return full_tag  # _disable_links will handle
            if ext not in ('.html', '.htm'):
                return full_tag  # _embed_static handles non-HTML
            local_path = base / href.split('?')[0]

        if local_path is None or not local_path.is_file():
            return full_tag

        try:
            inner = local_path.read_text(encoding='utf-8', errors='replace')
            inner = _embed_html_file(inner, local_path.parent, depth + 1)
            key = f'_s{counter[0]}'
            counter[0] += 1
            sub_map[key] = _html_to_b64(inner)
            new_tag = re.sub(r'\bhref="[^"]*"', 'href="#"', full_tag)
            new_tag = re.sub(r'\btarget="[^"]*"', '', new_tag)
            new_tag = re.sub(r'\bonclick="(?:[^"]*|(?:"[^"]*"))*"', '', new_tag)
            new_tag = new_tag.rstrip('>')
            new_tag += f' onclick="event.preventDefault();window._openPortable(\'{key}\')">'
            return new_tag
        except Exception as e:
            print(f'  Warning: could not sub-embed {href}: {e}')
            return full_tag

    html = _ANCHOR_RE.sub(_sub, html)

    if sub_map:
        map_json = json.dumps(sub_map, separators=(',', ':'))
        inject = (f'<script>(function(){{var m={map_json};'
                  f'var p=window._openPortable;'
                  f'window._openPortable=function(k){{'
                  f'if(m[k]){{var b=atob(m[k]),a=new Uint8Array(b.length);'
                  f'for(var i=0;i<b.length;i++)a[i]=b.charCodeAt(i);'
                  f'window.open(URL.createObjectURL(new Blob([a],{{type:\'text/html\'}})),\'_blank\');}}'
                  f'else if(p)p(k);}};}})()</script>')
        html = html.replace('</body>', inject + '</body>', 1) if '</body>' in html else html + inject

    return html


def _embed_html_file(html: str, base: Path, depth: int = 0) -> str:
    """Fully embed a single HTML file: static resources + viewer logic if applicable."""
    html = _embed_static(html, base, depth)
    html = _embed_sub_html_hrefs(html, base, depth)  # embed relative HTML links
    html = _disable_links(html)
    if _is_viewer_html(html):
        print(f'    [viewer] {base.name}')
        html = _embed_viewer_html(html, base, depth)
    return html


# ---------------------------------------------------------------------------
# JS blob script injected into the top-level portable file
# ---------------------------------------------------------------------------
_TOP_BLOB_SCRIPT = """
<script id="_portable_blobs">(function(){{
  var _blobs={blob_map};
  window._openPortable=function(key){{
    var b64=_blobs[key];
    if(!b64)return;
    var b=atob(b64),a=new Uint8Array(b.length);
    for(var i=0;i<b.length;i++)a[i]=b.charCodeAt(i);
    window.open(URL.createObjectURL(new Blob([a],{{type:'text/html'}})),'_blank');
  }};
}})();
</script>
"""


def _unc_to_drive(path: Path) -> Path:
    """On Windows, try to convert a UNC path (\\\\server\\share\\...) to the
    equivalent mapped drive letter path (X:\\...).  If no mapping is found or the
    platform is not Windows, the original path is returned unchanged.

    This is needed because tkinter's file dialog can return UNC paths when the user
    navigates to a share without using its drive-letter mapping.  UNC paths work for
    simple file access but cause problems when Path arithmetic produces very long paths
    (>MAX_PATH) or when SMB latency causes is_file() checks to time out.
    """
    s = str(path)
    if not s.startswith('\\\\'):
        return path
    try:
        import subprocess as _sp
        r = _sp.run(
            ['net', 'use'], capture_output=True, text=True, timeout=5,
            creationflags=0x08000000)          # CREATE_NO_WINDOW
        for line in r.stdout.splitlines():
            # Typical net-use line:  OK    M:    \\server\share    ...
            cols = line.split()
            if len(cols) >= 3 and len(cols[1]) == 2 and cols[1][1] == ':' \
                    and cols[2].startswith('\\\\'):
                drive    = cols[1].upper()            # e.g. "M:"
                unc_root = cols[2].rstrip('\\')       # e.g. "\\server\share"
                if s.lower().startswith(unc_root.lower()):
                    remainder = s[len(unc_root):]
                    return Path(drive + '\\' + remainder.lstrip('\\'))
    except Exception:
        pass
    return path


def make_portable(dashboard_path: Path, output_path: Path = None) -> Path:
    # Normalize UNC paths (\\server\share\...) → mapped drive (X:\...) so that
    # all subsequent Path arithmetic stays on the fast, short drive-letter form.
    dashboard_path = _unc_to_drive(dashboard_path)

    if output_path is None:
        output_path = dashboard_path.parent / 'Dashboard_portable.html'

    print(f'Reading  {dashboard_path}')
    html = dashboard_path.read_text(encoding='utf-8', errors='replace')
    base = dashboard_path.parent

    print('Embedding resources …')
    blob_map = {}
    blob_counter = [0]

    def _collect_href(m):
        full_tag = m.group(0)
        href_m = re.search(r'\bhref="([^"]+)"', full_tag)
        if not href_m:
            return full_tag
        href = href_m.group(1)

        # Leave target="content" links for _embed_viewer_html to wire into __pl
        if re.search(r'\btarget=["\']content["\']', full_tag, re.IGNORECASE):
            return full_tag

        # Resolve file:// local HTML (both file:/// local and file://server/ UNC)
        local_path = None
        if href.startswith('file://'):
            # Check extension first — disable non-embeddable file:// links
            _fe = Path(href.split('?')[0].split('/')[-1]).suffix.lower()
            if _fe in _DISABLE_EXTS:
                return _apply_disabled(full_tag)
            local_path = _resolve_file_url(href)
            if local_path is None or local_path.suffix.lower() not in ('.html', '.htm'):
                return _apply_disabled(full_tag)  # unresolvable file:// link
        elif _is_external(href):
            return full_tag
        else:
            clean = href.split('?')[0].split('#')[0]
            if not clean:
                return full_tag
            ext = Path(clean).suffix.lower()
            if ext in _DISABLE_EXTS:
                return _apply_disabled(full_tag)
            local_path = base / clean
            if not local_path.is_file():
                return full_tag
            if local_path.suffix.lower() not in ('.html', '.htm'):
                return re.sub(r'\bhref="[^"]*"', f'href="{_to_data_uri(local_path)}"', full_tag)

        # Embed as blob
        try:
            inner = local_path.read_text(encoding='utf-8', errors='replace')
            print(f'  Embedding {local_path.name} …')
            inner = _embed_html_file(inner, local_path.parent, depth=1)
            encoded = _html_to_b64(inner)
            key = f'h{blob_counter[0]}'
            blob_counter[0] += 1
            blob_map[key] = encoded
            new_tag = re.sub(r'\bhref="[^"]*"', 'href="#"', full_tag)
            new_tag = re.sub(r'\bonclick="[^"]*"', '', new_tag)
            new_tag = new_tag.rstrip('>')
            new_tag += f' onclick="event.preventDefault();window._openPortable(\'{key}\')">'
            return new_tag
        except Exception as e:
            print(f'  Warning: could not embed {href}: {e}')
        return full_tag

    html = _ANCHOR_RE.sub(_collect_href, html)

    # Embed static resources in top-level HTML itself
    html = _embed_static(html, base, depth=0)

    # If the top-level file is a sidebar viewer (uses load() navigation),
    # embed all load() targets AND href="*.html" target="content" links via __pl.
    # This must run AFTER _collect_href so any remaining direct hrefs are already
    # handled, and AFTER _embed_static so images inside sub-pages are inlined.
    if _is_viewer_html(html):
        print('  [viewer] Embedding sidebar navigation targets ...')
        html = _embed_viewer_html(html, base, depth=0)

    html = _disable_links(html)

    # Inject top-level blob JS before </body>
    blob_json = json.dumps(blob_map, separators=(',', ':'))
    blob_script = _TOP_BLOB_SCRIPT.format(blob_map=blob_json)
    if '</body>' in html:
        html = html.replace('</body>', blob_script + '</body>', 1)
    else:
        html += blob_script

    # Write to a temp file first so a locked output (open in browser) doesn't
    # lose data; then atomically replace the target.
    _tmp = output_path.with_suffix('.tmp')
    _tmp.write_text(html, encoding='utf-8')
    try:
        os.replace(str(_tmp), str(output_path))
        size_mb = output_path.stat().st_size / 1_048_576
        print(f'Embedded {len(blob_map)} HTML report(s) as blob URLs')
        print(f'Wrote    {output_path}  ({size_mb:.1f} MB)')
    except PermissionError:
        size_mb = _tmp.stat().st_size / 1_048_576
        print(f'Embedded {len(blob_map)} HTML report(s) as blob URLs')
        print(f'WARNING: {output_path} is locked (close it in your browser first).')
        print(f'Wrote    {_tmp}  ({size_mb:.1f} MB)  — rename it manually.')
        output_path = _tmp
    return output_path


def main():
    ap = argparse.ArgumentParser(
        description='Make a portable self-contained Dashboard.html')
    ap.add_argument('dashboard', help='Path to Dashboard.html')
    ap.add_argument('--out', help='Output file (default: Dashboard_portable.html next to input)')
    args = ap.parse_args()

    dashboard_path = Path(args.dashboard)
    if not dashboard_path.exists():
        print(f'Error: {dashboard_path} not found', file=sys.stderr)
        sys.exit(1)

    make_portable(dashboard_path, Path(args.out) if args.out else None)


if __name__ == '__main__':
    main()


# ════════════════════════════════════════════════════════════════
# (formerly manage_dashboard.py)
# ════════════════════════════════════════════════════════════════
import os
import re
import shutil
import glob
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


# ── HTML parsing helpers ────────────────────────────────────────────────────


# Sentinel pairs for all three sections (new format) plus legacy
_SECTION_PATTERNS = [
    (r'<!--\s*YIELD_START\s*-->', r'<!--\s*YIELD_END\s*-->',   'YIELD'),
    (r'<!--\s*COMPARE_START\s*-->', r'<!--\s*COMPARE_END\s*-->', 'COMPARE'),
    (r'<!--\s*VMIN_START\s*-->',  r'<!--\s*VMIN_END\s*-->',    'VMIN'),
    (r'<!--\s*RUNS_START\s*-->',  r'<!--\s*RUNS_END\s*-->',    'YIELD'),  # legacy
]


def _all_runs_html(html: str) -> str:
    """Concatenate inner content of all known sentinel sections."""
    parts = []
    for start_re, end_re, _ in _SECTION_PATTERNS:
        m = re.search(start_re + r'(.*?)' + end_re, html, re.S)
        if m:
            parts.append(m.group(1))
    return '\n'.join(parts)


def parse_blocks(html: str) -> list[dict]:
    """Return list of dicts with keys: stem, label, ts, html, folder_hint."""
    runs_html = _all_runs_html(html)
    if not runs_html.strip():
        return []

    blocks = []
    # Each block: <div class="run-block" data-stem="...">...</div>
    for bm in re.finditer(
        r'<div class="run-block"\s+data-stem="([^"]*)">(.*?)</div>\s*</div>',
        runs_html, re.S
    ):
        stem = bm.group(1)
        inner = bm.group(2)

        # label from run-header text (after the arrow span)
        lm = re.search(r'<div class="run-header"[^>]*>.*?</span>\s*([^<]+)', inner, re.S)
        label = lm.group(1).strip() if lm else stem

        # timestamp
        tm = re.search(r'<span class="ts">\s*-\s*([^<]+)</span>', inner)
        ts = tm.group(1).strip() if tm else ''

        # first href — used to guess output folder
        hm = re.search(r'href="((?!file://)[^"]+)"', inner)
        first_href = hm.group(1) if hm else ''

        blocks.append({
            'stem': stem,
            'label': label,
            'ts': ts,
            'first_href': first_href,
            'full_div': bm.group(0),
        })
    return blocks


def resolve_output_folder(dashboard_html_path: str, first_href: str) -> str | None:
    """Derive the output folder path from the first relative href in the block."""
    if not first_href:
        return None
    # href is relative to the Dashboard.html directory
    base = os.path.dirname(dashboard_html_path)
    # strip the filename part (e.g. "52A/NCXEBJX.../index.html" → "52A/NCXEBJX.../")
    folder = os.path.normpath(os.path.join(base, os.path.dirname(first_href)))
    return folder if os.path.isdir(folder) else folder  # return even if not yet present


def remove_block(html: str, stem: str) -> str:
    """Remove the run-block div with data-stem == stem from any section."""
    escaped = re.escape(stem)
    pattern = (
        r'[ \t]*<div class="run-block"\s+data-stem="' + escaped +
        r'">' + r'.*?</div>\s*</div>\s*'
    )
    return re.sub(pattern, '', html, flags=re.S)


def section_type_of_block(html: str, stem: str) -> str | None:
    """Return the section type ('YIELD', 'COMPARE', 'VMIN') a block belongs to."""
    escaped = re.escape(stem)
    block_re = re.compile(
        r'<div class="run-block"\s+data-stem="' + escaped + r'">',
        re.S)
    for start_re, end_re, sec_type in _SECTION_PATTERNS:
        m = re.search(start_re + r'(.*?)' + end_re, html, re.S)
        if m and block_re.search(m.group(1)):
            return sec_type
    return None


def resolve_block_files(dashboard_html_path: str, block: dict) -> list[str]:
    """Return list of absolute file paths referenced by href in a block's HTML."""
    base = os.path.dirname(dashboard_html_path)
    files = []
    for hm in re.finditer(r'href="((?!file://|http)[^"]+)"', block.get('full_div', '')):
        rel = hm.group(1)
        fpath = os.path.normpath(os.path.join(base, rel))
        files.append(fpath)
    return files


def _section_of_block(html: str, stem: str) -> tuple[str, str] | tuple[None, None]:
    """Return (start_sentinel_literal, end_sentinel_literal) for the section
    that contains the given block stem, searching all known sections."""
    escaped = re.escape(stem)
    block_re = re.compile(
        r'<div class="run-block"\s+data-stem="' + escaped + r'">',
        re.S)
    for start_re, end_re, _ in _SECTION_PATTERNS:
        m = re.search(start_re + r'(.*?)' + end_re, html, re.S)
        if m and block_re.search(m.group(1)):
            # Return the actual matched sentinel strings
            full = re.search(start_re + r'.*?' + end_re, html, re.S)
            if full:
                sm = re.search(start_re, html)
                em = re.search(end_re, html)
                if sm and em:
                    return html[sm.start():sm.end()], html[em.start():em.end()]
    return None, None


def reorder_blocks(html: str, ordered_blocks: list[dict]) -> str:
    """Reorder blocks within their respective sections.
    Blocks belonging to the same section are reordered together.
    New format: YIELD/COMPARE/VMIN sentinels.  Legacy: RUNS sentinels."""
    # Group the ordered_blocks by the section they belong to
    for start_re, end_re, _ in _SECTION_PATTERNS:
        m = re.search(start_re + r'(.*?)' + end_re, html, re.S)
        if not m:
            continue
        section_html = m.group(1)
        # Find which blocks from ordered_blocks live in this section
        section_blocks = []
        for b in ordered_blocks:
            escaped = re.escape(b['stem'])
            if re.search(r'<div class="run-block"\s+data-stem="' + escaped + r'">', section_html, re.S):
                section_blocks.append(b)
        if not section_blocks:
            continue
        new_section = '\n' + ''.join(b['full_div'] + '\n' for b in section_blocks)
        html = re.sub(
            start_re + r'.*?' + end_re,
            lambda mo, ns=new_section, sr=start_re, er=end_re: (
                re.search(sr, mo.group(0)).group(0) + ns +
                re.search(er, mo.group(0)).group(0)
            ),
            html, flags=re.S, count=1
        )
    return html


# ── Main GUI ────────────────────────────────────────────────────────────────


class ManageFrame(tk.Frame):
    def __init__(self, parent=None, **kw):
        super().__init__(parent, bg='#1a252f', **kw)
        self._html_path: str = ''
        self._html: str = ''
        self._blocks: list[dict] = []
        self._build_ui()

    # ── UI construction ─────────────────────────────────────────────────────

    def _build_ui(self):
        BG  = '#1a252f'
        BG2 = '#2c3e50'
        FG  = '#ecf0f1'
        ABLU = '#3498db'

        # ── top bar ──
        top = tk.Frame(self, bg=BG)
        top.pack(fill=tk.X, padx=8, pady=6)

        self._path_var = tk.StringVar()
        tk.Label(top, text='Dashboard.html:', bg=BG, fg=FG,
                 font=('Arial', 9)).pack(side=tk.LEFT)
        tk.Entry(top, textvariable=self._path_var, width=60,
                 bg=BG2, fg='white', insertbackground='white',
                 relief='flat', font=('Consolas', 9)).pack(side=tk.LEFT, padx=(4, 4))
        tk.Button(top, text='Browse…', command=self._browse,
                  bg='#1f618d', fg='white', relief='flat',
                  font=('Arial', 9), padx=6).pack(side=tk.LEFT, padx=(0, 4))
        tk.Button(top, text='Load', command=self._load,
                  bg='#27ae60', fg='white', relief='flat',
                  font=('Arial', 9, 'bold'), padx=8).pack(side=tk.LEFT)

        # ── table ──
        cols = ('label', 'ts', 'folder')
        frm = tk.Frame(self, bg=BG)
        frm.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 4))

        style = ttk.Style()
        style.theme_use('default')
        style.configure('Dark.Treeview',
                        background=BG2, foreground=FG,
                        fieldbackground=BG2, rowheight=24,
                        font=('Consolas', 9))
        style.configure('Dark.Treeview.Heading',
                        background='#34495e', foreground=FG,
                        font=('Arial', 9, 'bold'), relief='flat')
        style.map('Dark.Treeview',
                  background=[('selected', '#2980b9')],
                  foreground=[('selected', 'white')])

        self._tree = ttk.Treeview(frm, columns=cols, show='headings',
                                  style='Dark.Treeview', selectmode='browse')
        self._tree.heading('label',  text='Identifier / Label')
        self._tree.heading('ts',     text='Timestamp')
        self._tree.heading('folder', text='Output Folder')
        self._tree.column('label',  width=280, anchor='w')
        self._tree.column('ts',     width=140, anchor='w')
        self._tree.column('folder', width=420, anchor='w')

        vsb = ttk.Scrollbar(frm, orient='vertical', command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # ── action buttons ──
        btn_frm = tk.Frame(self, bg=BG)
        btn_frm.pack(fill=tk.X, padx=8, pady=(0, 6))

        # Reorder buttons
        tk.Button(btn_frm, text='⬆ Top',
                  command=self._move_top,
                  bg='#1f618d', fg='white', relief='flat',
                  font=('Arial', 9), padx=8).pack(side=tk.LEFT, padx=(0, 2))
        tk.Button(btn_frm, text='▲ Up',
                  command=self._move_up,
                  bg='#1f618d', fg='white', relief='flat',
                  font=('Arial', 9), padx=8).pack(side=tk.LEFT, padx=(0, 2))
        tk.Button(btn_frm, text='▼ Down',
                  command=self._move_down,
                  bg='#1f618d', fg='white', relief='flat',
                  font=('Arial', 9), padx=8).pack(side=tk.LEFT, padx=(0, 2))
        tk.Button(btn_frm, text='⬇ Bottom',
                  command=self._move_bottom,
                  bg='#1f618d', fg='white', relief='flat',
                  font=('Arial', 9), padx=8).pack(side=tk.LEFT, padx=(0, 12))

        tk.Button(btn_frm, text='Delete from Dashboard.html',
                  command=self._delete_html_entry,
                  bg='#922b21', fg='white', relief='flat',
                  font=('Arial', 9, 'bold'), padx=10).pack(side=tk.LEFT, padx=(0, 6))
        tk.Button(btn_frm, text='Delete Output Folder',
                  command=self._delete_folder,
                  bg='#7d3c98', fg='white', relief='flat',
                  font=('Arial', 9, 'bold'), padx=10).pack(side=tk.LEFT, padx=(0, 6))
        tk.Button(btn_frm, text='Delete Both',
                  command=self._delete_both,
                  bg='#c0392b', fg='white', relief='flat',
                  font=('Arial', 9, 'bold'), padx=10).pack(side=tk.LEFT, padx=(0, 6))


        tk.Button(btn_frm, text='Delete Compare Files…',
                  command=self._delete_compare_files,
                  bg='#784212', fg='white', relief='flat',
                  font=('Arial', 9, 'bold'), padx=10).pack(side=tk.LEFT, padx=(0, 6))
        tk.Button(btn_frm, text='Delete Vmin Files…',
                  command=self._delete_vmin_files,
                  bg='#117a65', fg='white', relief='flat',
                  font=('Arial', 9, 'bold'), padx=10).pack(side=tk.LEFT, padx=(0, 6))

        # ── status bar ──
        self._status_var = tk.StringVar(value='Load a Dashboard.html to begin.')
        tk.Label(self, textvariable=self._status_var, bg='#151e27', fg='#95a5a6',
                 font=('Consolas', 8), anchor='w', padx=6).pack(fill=tk.X, side=tk.BOTTOM)

    # ── actions ─────────────────────────────────────────────────────────────

    def _browse(self):
        path = filedialog.askopenfilename(
            title='Select Dashboard.html',
            filetypes=[('HTML files', '*.html'), ('All files', '*.*')]
        )
        if path:
            self._path_var.set(path)
            self._load()

    def _load(self):
        path = self._path_var.get().strip()
        if not path or not os.path.isfile(path):
            messagebox.showerror('Error', f'File not found:\n{path}')
            return
        try:
            with open(path, 'r', encoding='utf-8') as f:
                self._html = f.read()
            self._html_path = path
        except Exception as e:
            messagebox.showerror('Error', f'Failed to read file:\n{e}')
            return

        self._blocks = parse_blocks(self._html)
        self._refresh_tree()
        self._status_var.set(f'Loaded {len(self._blocks)} block(s) from {path}')

    def _refresh_tree(self):
        for row in self._tree.get_children():
            self._tree.delete(row)
        for i, b in enumerate(self._blocks):
            folder = resolve_output_folder(self._html_path, b['first_href'])
            exists = '✓' if folder and os.path.isdir(folder) else '✗'
            if folder:
                # Relative path from Dashboard.html dir, strip leading 'output/'
                _html_dir = os.path.dirname(self._html_path) if self._html_path else ''
                try:
                    _rel = os.path.relpath(folder, _html_dir).replace('\\', '/') if _html_dir else ''
                except ValueError:
                    _rel = ''
                if _rel.startswith('output/'):
                    _rel = _rel[len('output/'):]
                _parts = [p for p in _rel.split('/') if p and p not in ('.', '..')]
                if len(_parts) >= 2:
                    # New 2-level structure: e.g. NVL_0H61A_20260522/NCXSDJXL0H61A002618_119325
                    _short = '/'.join(_parts[-2:])
                else:
                    # Old flat-output: use block timestamp as the "folder" prefix
                    _leaf = _parts[-1] if _parts else os.path.basename(folder.rstrip('/\\'))
                    _ts   = b.get('ts', '').strip()
                    _short = f'{_ts}/{_leaf}' if _ts else _leaf
                folder_disp = f'[{exists}] {_short}'
            else:
                folder_disp = '—'
            self._tree.insert('', tk.END, iid=str(i),
                               values=(b['label'], b['ts'], folder_disp))

    def _selected_block(self) -> dict | None:
        sel = self._tree.selection()
        if not sel:
            messagebox.showwarning('No selection', 'Select an entry first.')
            return None
        return self._blocks[int(sel[0])]

    def _delete_html_entry(self):
        block = self._selected_block()
        if not block:
            return
        if not messagebox.askyesno('Confirm',
                f'Remove this entry from Dashboard.html?\n\n{block["label"]}'):
            return
        self._html = remove_block(self._html, block['stem'])
        self._save_html()
        self._blocks = parse_blocks(self._html)
        self._refresh_tree()
        self._status_var.set(f'Removed entry: {block["label"]}')

    def _delete_folder(self):
        block = self._selected_block()
        if not block:
            return
        sec_type = section_type_of_block(self._html, block['stem'])

        # COMPARE blocks: delete only the referenced HTML file(s), not the folder
        if sec_type == 'COMPARE':
            files = resolve_block_files(self._html_path, block)
            existing = [f for f in files if os.path.isfile(f)]
            if not existing:
                messagebox.showinfo('Not found',
                    'No compare files found to delete.')
                return
            flist = '\n'.join(os.path.basename(f) for f in existing)
            if not messagebox.askyesno('Confirm',
                    f'Delete these compare file(s)?\n\n{flist}'):
                return
            errors = []
            for f in existing:
                try:
                    os.remove(f)
                except Exception as e:
                    errors.append(f'{os.path.basename(f)}: {e}')
            if errors:
                messagebox.showerror('Errors', '\n'.join(errors))
            else:
                self._status_var.set(
                    f'Deleted {len(existing)} compare file(s).')
            self._refresh_tree()
            return

        folder = resolve_output_folder(self._html_path, block['first_href'])
        if not folder or not os.path.isdir(folder):
            messagebox.showinfo('Not found',
                f'Output folder not found or already deleted:\n{folder}')
            return
        if not messagebox.askyesno('Confirm',
                f'Permanently delete this folder and ALL its contents?\n\n{folder}'):
            return
        try:
            shutil.rmtree(folder)
            self._refresh_tree()
            self._status_var.set(f'Deleted folder: {folder}')
        except Exception as e:
            messagebox.showerror('Error', f'Failed to delete folder:\n{e}')

    def _delete_both(self):
        block = self._selected_block()
        if not block:
            return
        sec_type = section_type_of_block(self._html, block['stem'])

        # COMPARE blocks: remove HTML entry + delete referenced file(s) only
        if sec_type == 'COMPARE':
            files = resolve_block_files(self._html_path, block)
            existing = [f for f in files if os.path.isfile(f)]
            flist = '\n'.join(os.path.basename(f) for f in existing) if existing else '(no files found)'
            msg = (f'Remove entry from Dashboard.html AND delete compare file(s)?\n\n'
                   f'{block["label"]}\n\nFiles:\n{flist}')
            if not messagebox.askyesno('Confirm', msg):
                return
            self._html = remove_block(self._html, block['stem'])
            self._save_html()
            errors = []
            for f in existing:
                try:
                    os.remove(f)
                except Exception as e:
                    errors.append(f'{os.path.basename(f)}: {e}')
            # Update compare links in Dashboard.html
            try:
                import pathlib
                import compare_runs as _cr
                _cr.update_dashboard_compare_links(pathlib.Path(self._html_path))
                with open(self._html_path, 'r', encoding='utf-8') as f:
                    self._html = f.read()
            except Exception:
                pass
            self._blocks = parse_blocks(self._html)
            self._refresh_tree()
            if errors:
                messagebox.showerror('Errors', '\n'.join(errors))
            else:
                self._status_var.set(
                    f'Deleted entry + {len(existing)} compare file(s): {block["label"]}')
            return

        folder = resolve_output_folder(self._html_path, block['first_href'])
        folder_exists = folder and os.path.isdir(folder)
        msg = f'Remove entry from Dashboard.html AND delete output folder?\n\n{block["label"]}'
        if folder_exists:
            msg += f'\n\nFolder to delete:\n{folder}'
        else:
            msg += '\n\n(Output folder not found — only HTML entry will be removed.)'
        if not messagebox.askyesno('Confirm', msg):
            return

        self._html = remove_block(self._html, block['stem'])
        self._save_html()

        if folder_exists:
            try:
                shutil.rmtree(folder)
            except Exception as e:
                messagebox.showerror('Error', f'Failed to delete folder:\n{e}')

        self._blocks = parse_blocks(self._html)
        self._refresh_tree()
        self._status_var.set(f'Deleted entry + folder: {block["label"]}')

    def _delete_compare_files(self):
        if not self._html_path:
            messagebox.showwarning('No file', 'Load a Dashboard.html first.')
            return
        dash_dir = os.path.dirname(self._html_path)
        import re as _re2
        name_re = _re2.compile(r'compare', _re2.IGNORECASE)
        found = [
            p for p in sorted(glob.glob(os.path.join(dash_dir, '*.html')))
            if os.path.basename(p).lower() != 'dashboard.html'
            and name_re.search(os.path.splitext(os.path.basename(p))[0])
        ]
        if not found:
            messagebox.showinfo('Nothing found', 'No compare/comparison HTML files found.')
            return

        # ── Selection dialog ────────────────────────────────────────────────
        BG  = '#1a252f'
        BG2 = '#2c3e50'
        FG  = '#ecf0f1'

        dlg = tk.Toplevel(self)
        dlg.title('Delete Compare Files')
        dlg.configure(bg=BG)
        dlg.resizable(False, False)
        dlg.grab_set()

        tk.Label(dlg, text='Select files to delete:',
                 bg=BG, fg=FG, font=('Arial', 9, 'bold'),
                 padx=12, pady=8).pack(anchor='w')

        chk_frame = tk.Frame(dlg, bg=BG2, padx=8, pady=6)
        chk_frame.pack(fill='x', padx=12, pady=(0, 8))

        vars_ = []
        for p in found:
            var = tk.BooleanVar(value=True)
            tk.Checkbutton(chk_frame, text=os.path.basename(p),
                           variable=var, bg=BG2, fg=FG,
                           selectcolor=BG, activebackground=BG2,
                           activeforeground=FG, font=('Consolas', 9),
                           relief='flat').pack(anchor='w', pady=1)
            vars_.append((p, var))

        # Select all / None helpers
        sel_row = tk.Frame(dlg, bg=BG)
        sel_row.pack(fill='x', padx=12, pady=(0, 6))
        tk.Button(sel_row, text='Select All',
                  command=lambda: [v.set(True) for _, v in vars_],
                  bg='#1f618d', fg='white', relief='flat',
                  font=('Arial', 8), padx=6).pack(side='left', padx=(0, 4))
        tk.Button(sel_row, text='Select None',
                  command=lambda: [v.set(False) for _, v in vars_],
                  bg='#1f618d', fg='white', relief='flat',
                  font=('Arial', 8), padx=6).pack(side='left')

        def _do_delete():
            to_delete = [p for p, v in vars_ if v.get()]
            if not to_delete:
                messagebox.showwarning('Nothing selected', 'Select at least one file.',
                                       parent=dlg)
                return
            dlg.destroy()
            errors = []
            for p in to_delete:
                try:
                    os.remove(p)
                except Exception as e:
                    errors.append(f'{os.path.basename(p)}: {e}')
            # Check if any compare files remain; update Dashboard.html accordingly
            try:
                import re as _re
                import sys, pathlib
                _src = str(pathlib.Path(__file__).parent)
                if _src not in sys.path:
                    sys.path.insert(0, _src)
                import compare_runs as _cr
                _cr.update_dashboard_compare_links(pathlib.Path(self._html_path))
                with open(self._html_path, 'r', encoding='utf-8') as f:
                    self._html = f.read()
            except Exception as e:
                errors.append(f'Dashboard.html update: {e}')
            if errors:
                messagebox.showerror('Errors', '\n'.join(errors))
            else:
                self._status_var.set(
                    f'Deleted {len(to_delete)} compare file(s); Dashboard.html updated.')

        btn_row = tk.Frame(dlg, bg=BG)
        btn_row.pack(fill='x', padx=12, pady=(0, 10))
        tk.Button(btn_row, text='Delete Selected',
                  command=_do_delete,
                  bg='#c0392b', fg='white', relief='flat',
                  font=('Arial', 9, 'bold'), padx=10).pack(side='left', padx=(0, 6))
        tk.Button(btn_row, text='Cancel',
                  command=dlg.destroy,
                  bg='#555', fg='white', relief='flat',
                  font=('Arial', 9), padx=10).pack(side='left')

        dlg.update_idletasks()

    # ── Delete Vmin Files dialog ─────────────────────────────────────────────

    def _delete_vmin_files(self):
        if not self._html_path:
            messagebox.showwarning('No file', 'Load a Dashboard.html first.')
            return

        # Collect only VMIN blocks (stem starts with "vmin__")
        vmin_blocks = [b for b in self._blocks if b['stem'].startswith('vmin__')]
        if not vmin_blocks:
            messagebox.showinfo('Nothing found', 'No Vmin entries found in Dashboard.html.')
            return

        BG  = '#1a252f'
        BG2 = '#2c3e50'
        FG  = '#ecf0f1'

        dlg = tk.Toplevel(self)
        dlg.title('Delete Vmin Files')
        dlg.configure(bg=BG)
        dlg.resizable(True, False)
        dlg.grab_set()

        tk.Label(dlg, text='Select Vmin runs to delete:',
                 bg=BG, fg=FG, font=('Arial', 9, 'bold'),
                 padx=12, pady=8).pack(anchor='w')

        # Column headers
        hdr = tk.Frame(dlg, bg='#34495e')
        hdr.pack(fill='x', padx=12, pady=(0, 2))
        tk.Label(hdr, text='  Delete?', bg='#34495e', fg=FG,
                 font=('Arial', 8, 'bold'), width=10, anchor='w').pack(side='left')
        tk.Label(hdr, text='Label', bg='#34495e', fg=FG,
                 font=('Arial', 8, 'bold'), width=30, anchor='w').pack(side='left')
        tk.Label(hdr, text='Output Folder', bg='#34495e', fg=FG,
                 font=('Arial', 8, 'bold'), anchor='w').pack(side='left', padx=(4, 0))

        chk_frame = tk.Frame(dlg, bg=BG2, padx=8, pady=6)
        chk_frame.pack(fill='x', padx=12, pady=(0, 4))

        entries = []  # (block, folder_path, tk.BooleanVar)
        for b in vmin_blocks:
            folder = resolve_output_folder(self._html_path, b['first_href'])
            var = tk.BooleanVar(value=True)
            row = tk.Frame(chk_frame, bg=BG2)
            row.pack(fill='x', pady=1)

            exists = folder and os.path.isdir(folder)
            folder_disp = folder if folder else '—'
            folder_color = '#a9dfbf' if exists else '#e74c3c'

            tk.Checkbutton(row, variable=var, bg=BG2, fg=FG,
                           selectcolor=BG, activebackground=BG2,
                           activeforeground=FG, relief='flat',
                           width=2).pack(side='left')
            tk.Label(row, text=b['label'], bg=BG2, fg=FG,
                     font=('Consolas', 9), width=30,
                     anchor='w').pack(side='left')
            tk.Label(row, text=folder_disp, bg=BG2, fg=folder_color,
                     font=('Consolas', 8),
                     anchor='w').pack(side='left', padx=(4, 0))

            entries.append((b, folder, var))

        # Select all / none
        sel_row = tk.Frame(dlg, bg=BG)
        sel_row.pack(fill='x', padx=12, pady=(0, 4))
        tk.Button(sel_row, text='Select All',
                  command=lambda: [v.set(True) for _, _, v in entries],
                  bg='#1f618d', fg='white', relief='flat',
                  font=('Arial', 8), padx=6).pack(side='left', padx=(0, 4))
        tk.Button(sel_row, text='Select None',
                  command=lambda: [v.set(False) for _, _, v in entries],
                  bg='#1f618d', fg='white', relief='flat',
                  font=('Arial', 8), padx=6).pack(side='left')

        # Options: what to delete
        opt_frame = tk.Frame(dlg, bg=BG, padx=12, pady=4)
        opt_frame.pack(fill='x')
        del_mode = tk.StringVar(value='both')
        for val, lbl in (('html', 'Dashboard entry only'),
                         ('folder', 'Output folder only'),
                         ('both',  'Entry + folder (recommended)')):
            tk.Radiobutton(opt_frame, text=lbl, variable=del_mode, value=val,
                           bg=BG, fg=FG, selectcolor='#2c3e50',
                           activebackground=BG, activeforeground=FG,
                           font=('Arial', 9)).pack(side='left', padx=(0, 12))

        def _do_delete():
            selected = [(b, folder) for b, folder, v in entries if v.get()]
            if not selected:
                messagebox.showwarning('Nothing selected', 'Select at least one run.',
                                       parent=dlg)
                return
            mode = del_mode.get()
            dlg.destroy()

            errors = []
            for b, folder in selected:
                # Remove HTML entry
                if mode in ('html', 'both'):
                    self._html = remove_block(self._html, b['stem'])

                # Delete output folder
                if mode in ('folder', 'both'):
                    if folder and os.path.isdir(folder):
                        try:
                            shutil.rmtree(folder)
                        except Exception as e:
                            errors.append(f'{os.path.basename(folder)}: {e}')
                    elif mode == 'folder':
                        errors.append(f'{b["label"]}: folder not found ({folder})')

            if mode in ('html', 'both'):
                self._save_html()

            self._blocks = parse_blocks(self._html)
            self._refresh_tree()

            if errors:
                messagebox.showerror('Errors', '\n'.join(errors))
            else:
                n = len(selected)
                self._status_var.set(
                    f'Deleted {n} Vmin run(s) '
                    f'({"entries + folders" if mode == "both" else mode}).')

        btn_row = tk.Frame(dlg, bg=BG)
        btn_row.pack(fill='x', padx=12, pady=(0, 10))
        tk.Button(btn_row, text='Delete Selected',
                  command=_do_delete,
                  bg='#117a65', fg='white', relief='flat',
                  font=('Arial', 9, 'bold'), padx=10).pack(side='left', padx=(0, 6))
        tk.Button(btn_row, text='Cancel',
                  command=dlg.destroy,
                  bg='#555', fg='white', relief='flat',
                  font=('Arial', 9), padx=10).pack(side='left')

        dlg.update_idletasks()
        # Centre over parent
        x = self.winfo_x() + (self.winfo_width()  - dlg.winfo_width())  // 2
        y = self.winfo_y() + (self.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f'+{x}+{y}')

    def _save_html(self):
        try:
            with open(self._html_path, 'w', encoding='utf-8') as f:
                f.write(self._html)
        except Exception as e:
            messagebox.showerror('Error', f'Failed to save Dashboard.html:\n{e}')

    # ── reorder helpers ──────────────────────────────────────────────────────

    def _selected_index(self) -> int | None:
        sel = self._tree.selection()
        if not sel:
            messagebox.showwarning('No selection', 'Select an entry first.')
            return None
        return int(sel[0])

    def _apply_reorder(self, new_idx: int):
        """Save reordered blocks, refresh tree, and re-select the moved row."""
        self._html = reorder_blocks(self._html, self._blocks)
        self._save_html()
        self._refresh_tree()
        self._tree.selection_set(str(new_idx))
        self._tree.see(str(new_idx))
        self._status_var.set(f'Reordered: {self._blocks[new_idx]["label"]}')

    def _move_up(self):
        idx = self._selected_index()
        if idx is None or idx == 0:
            return
        self._blocks[idx], self._blocks[idx - 1] = self._blocks[idx - 1], self._blocks[idx]
        self._apply_reorder(idx - 1)

    def _move_down(self):
        idx = self._selected_index()
        if idx is None or idx >= len(self._blocks) - 1:
            return
        self._blocks[idx], self._blocks[idx + 1] = self._blocks[idx + 1], self._blocks[idx]
        self._apply_reorder(idx + 1)

    def _move_top(self):
        idx = self._selected_index()
        if idx is None or idx == 0:
            return
        block = self._blocks.pop(idx)
        self._blocks.insert(0, block)
        self._apply_reorder(0)

    def _move_bottom(self):
        idx = self._selected_index()
        if idx is None or idx >= len(self._blocks) - 1:
            return
        block = self._blocks.pop(idx)
        self._blocks.append(block)
        self._apply_reorder(len(self._blocks) - 1)


# Keep standalone alias for backward compat
DashboardManager = ManageFrame


def main():
    root = tk.Tk()
    root.title('Dashboard Manager')
    root.geometry('900x560')
    frame = ManageFrame(root)
    frame.pack(fill=tk.BOTH, expand=True)
    root.mainloop()


if __name__ == '__main__':
    main()


# ════════════════════════════════════════════════════════════════
# (formerly yield_report.py)
# ════════════════════════════════════════════════════════════════
#!/usr/bin/env python3
"""
yield_report.py
---------------
Weekly pareto yield report.  Reads Dashboard.html, resolves each run's
*_BinDistribution.html, groups runs by ISO calendar week, and generates
a standalone HTML report with a bin-fail pareto chart per week.

Usage:
    python yield_report.py Dashboard.html
    python yield_report.py Dashboard.html --out my_report.html
    python yield_report.py Dashboard.html --weeks 8
"""

import sys
import os
import re
import argparse
import io
import base64
from pathlib import Path
from datetime import datetime, timedelta


def _wm_inject(html: str) -> str:
    _wm = (
        '<div id="_wm_div" style="position:fixed;top:8px;right:12px;font-size:10px;'
        'font-weight:600;pointer-events:none;z-index:99999;'
        'font-family:Arial,sans-serif;user-select:none;letter-spacing:0.04em;'
        'padding:2px 6px;border-radius:3px;background:transparent;">'
        'Pant, Sujit N \u2014 GEMS FTE</div>'
        '<script>(function(){'
        'function _wm_color(){'
        'var d=document.getElementById("_wm_div");if(!d)return;'
        'var bg=window.getComputedStyle(document.body).backgroundColor;'
        'var m=bg.match(/\\d+/g);'
        'if(m&&m.length>=3){'
        'var r=+m[0],g=+m[1],b=+m[2];'
        'var lum=0.299*r+0.587*g+0.114*b;'
        'd.style.color=lum<128?"rgba(255,255,255,0.9)":"rgba(20,20,20,0.75)";'
        '}else{d.style.color="rgba(255,255,255,0.9)";}'
        '}'
        'if(document.readyState==="loading")'
        '{document.addEventListener("DOMContentLoaded",_wm_color);}'
        'else{_wm_color();}'
        '})();</script>'
    )
    import re as _re_wm
    if '</body>' not in html:
        return html
    html = _re_wm.sub(
        r'<div[^>]*id=["\']_wm_div["\'][^>]*>[\s\S]*?</div>\s*<script[^>]*>[\s\S]*?</script>',
        '', html)
    html = _re_wm.sub(r'<div[^>]*>[^<]*GEMS FTE[^<]*</div>', '', html)
    return html.replace('</body>', _wm + '\n</body>', 1)


try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np
    HAVE_MPL = True
except ImportError:
    HAVE_MPL = False


# Re-use parse helpers from compare_runs
try:
    import compare_runs as _cr
    _parse_dashboard  = _cr.parse_dashboard
    _find_bin_html    = _cr.find_bin_html
    _parse_bin_html   = _cr.parse_bin_html
    _find_xlsx        = _cr.find_xlsx
    HAVE_CR = True
except ImportError:
    HAVE_CR = False
    _parse_dashboard = None


# ---------------------------------------------------------------------------
# Colour palette (matches compare_runs)
# ---------------------------------------------------------------------------

_WEEK_COLORS = [
    '#2980b9', '#27ae60', '#e74c3c', '#f39c12',
    '#8e44ad', '#16a085', '#d35400', '#2c3e50',
    '#c0392b', '#1abc9c', '#7f8c8d', '#8e44ad',
]


def _esc(s: str) -> str:
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def _fig_b64(fig, dpi: int = 130) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('ascii')


INTERVALS = ['daily', 'weekly', 'bi-weekly', 'monthly']

# ---------------------------------------------------------------------------
# 1.  Timestamp → period key helpers
# ---------------------------------------------------------------------------

_TS_FMTS = (
    '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
    '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y',
    '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %H:%M', '%m/%d/%Y',
    '%Y%m%d',
)


def _parse_ts(ts: str) -> datetime | None:
    ts = (ts or '').strip()
    for fmt in _TS_FMTS:
        try:
            return datetime.strptime(ts, fmt)
        except ValueError:
            pass
    m = re.search(r'(\d{8})', ts)
    if m:
        try:
            return datetime.strptime(m.group(1), '%Y%m%d')
        except ValueError:
            pass
    return None


def _ts_to_isoweek(ts: str) -> str | None:
    """Return 'YYYY-Www' or None if unparseable (kept for back-compat)."""
    dt = _parse_ts(ts)
    if dt:
        iso_yr, iso_wk, _ = dt.isocalendar()
        return f'{iso_yr}-W{iso_wk:02d}'
    return None


def _ts_to_period(ts: str, interval: str) -> str | None:
    """Return period key string based on interval."""
    dt = _parse_ts(ts)
    if dt is None:
        return None
    if interval == 'daily':
        return dt.strftime('%Y-%m-%d')
    elif interval == 'weekly':
        iso_yr, iso_wk, _ = dt.isocalendar()
        return f'{iso_yr}-W{iso_wk:02d}'
    elif interval == 'bi-weekly':
        iso_yr, iso_wk, _ = dt.isocalendar()
        biweek = ((iso_wk - 1) // 2) * 2 + 1   # odd week = start of bi-week pair
        return f'{iso_yr}-BW{biweek:02d}'
    elif interval == 'monthly':
        return dt.strftime('%Y-%m')
    return _ts_to_isoweek(ts)  # fallback


def _period_sort_key(period_str: str) -> datetime:
    """Return a datetime suitable for sorting any period key."""
    # daily: YYYY-MM-DD
    for fmt in ('%Y-%m-%d', '%Y-%m'):
        try:
            return datetime.strptime(period_str, fmt)
        except ValueError:
            pass
    # weekly: YYYY-Www
    m = re.match(r'^(\d{4})-W(\d{2})$', period_str)
    if m:
        yr, wk = int(m.group(1)), int(m.group(2))
        jan4 = datetime(yr, 1, 4)
        return jan4 - timedelta(days=jan4.weekday()) + timedelta(weeks=wk - 1)
    # bi-weekly: YYYY-BWnn
    m = re.match(r'^(\d{4})-BW(\d{2})$', period_str)
    if m:
        yr, wk = int(m.group(1)), int(m.group(2))
        jan4 = datetime(yr, 1, 4)
        return jan4 - timedelta(days=jan4.weekday()) + timedelta(weeks=wk - 1)
    return datetime.min


def _week_start(week_str: str) -> datetime:
    """Return Monday of the ISO week 'YYYY-Www' (kept for back-compat)."""
    return _period_sort_key(week_str)


# ---------------------------------------------------------------------------
# 2b.  Build runs_data from a CSV / ZIP / GZ file (via trend_chart.load_csv)
# ---------------------------------------------------------------------------

def runs_from_csv(csv_path: Path, log=None, interval: str = 'weekly') -> list[dict]:
    """
    Load an ibin CSV/ZIP/GZ file and return a list of run dicts in the same
    format produced by load_run_data(), so they can be passed to generate_report.

    Each run dict contains:
        name, label, ts, week, run_date, bin_data (with bin_summary_rows + yield_rows)
    """
    import trend_chart as tc

    runs = tc.load_csv(csv_path, log=log)
    result = []
    for r in runs:
        # Build bin_summary_rows from bin_counts
        total = r.get('total_dies', 0) or 1
        bin_summary_rows = []
        for ibin, cnt in r.get('bin_counts', {}).items():
            fail_pct = cnt / total * 100 if total else 0.0
            bin_summary_rows.append({
                'ibin': ibin,
                'fail_count': cnt,
                'fail_pct': fail_pct,
                'desc': '',
            })

        # Build yield_rows for the pass bins (1/2 = FF, 1/2/3/4 = FF+DF)
        pass_cnt_ff    = sum(r['bin_counts'].get(b, 0) for b in (1, 2))
        pass_cnt_ffdf  = sum(r['bin_counts'].get(b, 0) for b in (1, 2, 3, 4))
        yield_rows = [
            {'bin': '1/2',     'yield_pct': pass_cnt_ff   / total * 100 if total else None},
            {'bin': '1/2/3/4', 'yield_pct': pass_cnt_ffdf / total * 100 if total else None},
        ]

        run_date = r.get('date')
        ts_str   = r.get('date_str', '') or (run_date.strftime('%Y-%m-%d') if run_date else '')
        period   = _ts_to_period(ts_str, interval)
        if period is None and run_date:
            period = _ts_to_period(run_date.strftime('%Y-%m-%d'), interval)
        if period is None:
            period = 'unknown'

        result.append({
            'name':     r.get('label', r.get('lot', '')),
            'label':    r.get('label', ''),
            'ts':       ts_str,
            'week':     period,
            'run_date': run_date,
            'bin_data': {
                'bin_summary_rows': bin_summary_rows,
                'yield_rows':       yield_rows,
            },
        })
        if log:
            log(f'  [{r.get("label", "")}]  period={period}  {len(bin_summary_rows)} bins\n')
    return result


# ---------------------------------------------------------------------------
# 2.  Load run data (bin_fail) for a list of records
# ---------------------------------------------------------------------------

def load_run_data(records: list[dict], dash_dir: Path,
                  log=None, interval: str = 'weekly') -> list[dict]:
    """
    For each record return an augmented dict with:
        bin_data  — output of _parse_bin_html or None
        week      — 'YYYY-Www'  (derived from ts or xlsx mtime)
        run_date  — datetime object
    """
    import compare_runs as cr
    result = []
    for rec in records:
        href = rec.get('index_href', '')
        output_dir = None
        if href:
            href_clean = re.sub(r'^file:///', '', href).replace('/', os.sep)
            idx_path   = Path(href_clean) if os.path.isabs(href_clean) else dash_dir / href_clean
            output_dir = idx_path.parent

        bin_data = None
        run_dt   = None
        if output_dir and output_dir.exists():
            bin_p = cr.find_bin_html(output_dir)
            if bin_p:
                bin_data = cr.parse_bin_html(bin_p)
                try:
                    run_dt = datetime.fromtimestamp(bin_p.stat().st_mtime)
                except Exception:
                    pass
            if run_dt is None:
                xlsx_p = cr.find_xlsx(dash_dir, href)
                if xlsx_p:
                    try:
                        run_dt = datetime.fromtimestamp(xlsx_p.stat().st_mtime)
                    except Exception:
                        pass

        # Try ts field for date
        period = _ts_to_period(rec.get('ts', ''), interval)
        if period is None and run_dt:
            period = _ts_to_period(run_dt.strftime('%Y-%m-%d'), interval)
        if period is None:
            # Pull date from stem  e.g.  NCXSDJXP0H51M202611-1  → 202611 ≈ 2026 wk11
            m = re.search(r'(\d{6})(?![\d])', rec.get('stem', ''))
            if m:
                ds = m.group(1)
                try:
                    dt = datetime.strptime(ds, '%Y%m')
                    period = _ts_to_period(dt.strftime('%Y-%m-%d'), interval)
                except ValueError:
                    pass
        if period is None:
            period = 'unknown'

        result.append({
            **rec,
            'bin_data': bin_data,
            'week': period,      # kept as 'week' key for back-compat
            'run_date': run_dt,
        })
        if log:
            status = f'{bin_p.name}' if (output_dir and output_dir.exists() and bin_data) else 'no bin data'
            log(f'  [{rec["name"]}]  period={period}  {status}\n')
    return result


# ---------------------------------------------------------------------------
# 3.  Group runs by ISO week
# ---------------------------------------------------------------------------

def group_by_week(runs: list[dict]) -> dict[str, list[dict]]:
    """Return OrderedDict  period_str → [run, ...] sorted chronologically."""
    from collections import OrderedDict
    week_map: dict[str, list] = {}
    for r in runs:
        week_map.setdefault(r['week'], []).append(r)
    return OrderedDict(sorted(week_map.items(), key=lambda kv: _period_sort_key(kv[0])))


# ---------------------------------------------------------------------------
# 4.  Pareto charts
# ---------------------------------------------------------------------------

def _ibin_label(ibin_key: str, ibin_desc: dict, cfg: dict | None) -> str:
    """Return display label for an ibin, using product config names if available."""
    desc = ''
    if cfg:
        ibin_name_map = cfg.get('ibin_name', {})
        try:
            k_int = int(float(ibin_key))
            desc = ibin_name_map.get(k_int, ibin_name_map.get(str(k_int), ''))
        except (ValueError, TypeError):
            pass
    if not desc:
        desc = ibin_desc.get(ibin_key, '')
    return f'iBin {ibin_key}' + (f'  — {str(desc)[:40]}' if desc else '')


def build_weekly_ibin_pareto(week_str: str, week_runs: list[dict],
                              top_n: int = 15, cfg: dict | None = None) -> str:
    """Horizontal bar pareto of iBin fail% for all runs in one period."""
    if not HAVE_MPL:
        return ''

    # Prefer bin_summary_rows (new format) → func_bin_rows → bin_fail_rows
    def _get_rows(run):
        bd = run.get('bin_data') or {}
        return (bd.get('bin_summary_rows') or
                bd.get('func_bin_rows') or
                bd.get('bin_fail_rows') or [])

    valid = [r for r in week_runs if _get_rows(r)]
    if not valid:
        return ''

    # Aggregate: sum fail_count per ibin across all runs in the week
    ibin_counts: dict[str, int]   = {}
    ibin_total:  dict[str, int]   = {}   # total die across runs that have this bin
    ibin_desc:   dict[str, str]   = {}

    for run in valid:
        for row in _get_rows(run):
            k = str(row.get('ibin', '')).strip()
            if not k:
                continue
            cnt = row.get('fail_count') or 0
            ibin_counts[k]  = ibin_counts.get(k, 0) + cnt
            desc = row.get('desc') or row.get('fail_bucket') or ''
            if desc:
                ibin_desc[k] = desc

    if not ibin_counts:
        return ''

    total_fails = sum(ibin_counts.values()) or 1
    sorted_ibins = sorted(ibin_counts, key=lambda k: ibin_counts[k], reverse=True)[:top_n]

    # Pareto: horizontal bar + cumulative % line
    n     = len(sorted_ibins)
    counts = [ibin_counts[k] for k in sorted_ibins]
    cum_pcts = []
    running = 0
    for c in counts:
        running += c
        cum_pcts.append(running / total_fails * 100)

    bar_pcts  = [c / total_fails * 100 for c in counts]
    y_pos = np.arange(n)

    fig, ax_bar = plt.subplots(figsize=(10, max(4, n * 0.52)))
    ax_cum = ax_bar.twinx()

    colors = [_WEEK_COLORS[i % len(_WEEK_COLORS)] for i in range(n)]
    bars = ax_bar.barh(y_pos, bar_pcts, color=colors, alpha=0.82, edgecolor='white', linewidth=0.4)

    for bar, pct, cnt in zip(bars, bar_pcts, counts):
        if pct >= 0.05:
            ax_bar.text(bar.get_width() + 0.05, bar.get_y() + bar.get_height() / 2,
                        f'{pct:.2f}%  (n={cnt:,})',
                        va='center', ha='left', fontsize=7.5)

    ax_cum.plot(cum_pcts, y_pos, marker='o', linewidth=2, color='#2c3e50',
                markersize=4, alpha=0.9, label='Cumulative %')
    ax_cum.axvline(80, color='#e74c3c', linestyle='--', linewidth=1.2, alpha=0.8, label='80%')

    ylabels = [_ibin_label(k, ibin_desc, cfg) for k in sorted_ibins]

    ax_bar.set_yticks(y_pos)
    ax_bar.set_yticklabels(ylabels, fontsize=8)
    ax_bar.invert_yaxis()
    ax_bar.set_xlabel('Fail (%)')
    ax_bar.set_title(
        f'Interface Bin Fail Pareto  \u2014  {week_str}  '
        f'({len(valid)} run{"s" if len(valid) != 1 else ""},'
        f' {sum(counts):,} total failures)',
        fontsize=12, weight='bold'
    )

    ax_cum.set_ylabel('Cumulative Fail (%)')
    ax_cum.set_ylim(0, 110)
    ax_cum.legend(fontsize=8, loc='lower right')
    ax_bar.grid(axis='x', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_weekly_yield_trend(weeks_data: dict, metric: str = 'ff_df') -> str:
    """
    Line chart: FF+DF yield (bin 1/2/3/4) per week, one point = median of week.
    metric: 'ff_df' uses bin '1/2/3/4', 'ff' uses '1/2'.
    """
    if not HAVE_MPL:
        return ''

    bin_key = '1/2/3/4' if metric == 'ff_df' else '1/2'
    label   = 'FF+DF Yield (Bin 1/2/3/4)' if metric == 'ff_df' else 'FF Yield (Bin 1/2)'

    week_labels = []
    medians     = []
    all_vals    = []

    for wk, runs in weeks_data.items():
        vals = []
        for r in runs:
            bd = r.get('bin_data') or {}
            for row in bd.get('yield_rows', []):
                if row.get('bin') == bin_key and row.get('yield_pct') is not None:
                    vals.append(row['yield_pct'])
        if vals:
            week_labels.append(wk)
            med = sorted(vals)[len(vals) // 2]
            medians.append(med)
            all_vals.append(vals)

    if not medians:
        return ''

    n   = len(week_labels)
    x   = np.arange(n)
    fig, ax = plt.subplots(figsize=(max(7, n * 1.1), 4))

    ax.plot(x, medians, marker='o', linewidth=2.4, color='#2980b9',
            markersize=8, label='Weekly median', zorder=4)

    for xi, (med, vals) in enumerate(zip(medians, all_vals)):
        if len(vals) > 1:
            ax.vlines(xi, min(vals), max(vals), color='#aaa', linewidth=1.5,
                      zorder=2, label='Range' if xi == 0 else '')
        ax.text(xi, med + 0.3, f'{med:.1f}%', ha='center', va='bottom',
                fontsize=8, weight='bold', color='#2980b9')

    ax.set_xticks(x)
    ax.set_xticklabels(week_labels, rotation=30, ha='right', fontsize=9)
    ax.set_ylabel('Yield (%)')
    ax.set_ylim(0, 105)
    ax.set_title(f'{label} — Weekly Trend', fontsize=12, weight='bold')
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    ax.legend(fontsize=8)
    fig.tight_layout()
    return _fig_b64(fig)


def build_weekly_top_fails_trend(weeks_data: dict, top_n: int = 8,
                                  cfg: dict | None = None) -> str:
    """
    Stacked area / grouped bar: top failing iBins per period as % of total fails.
    """
    if not HAVE_MPL:
        return ''

    def _get_rows(run):
        bd = run.get('bin_data') or {}
        return (bd.get('bin_summary_rows') or
                bd.get('func_bin_rows') or
                bd.get('bin_fail_rows') or [])

    # Find global top_n iBins by total fail count
    global_counts: dict[str, int] = {}
    for runs in weeks_data.values():
        for r in runs:
            for row in _get_rows(r):
                k = str(row.get('ibin', '')).strip()
                if k:
                    global_counts[k] = global_counts.get(k, 0) + (row.get('fail_count') or 0)

    top_bins = sorted(global_counts, key=lambda k: global_counts[k], reverse=True)[:top_n]
    if not top_bins:
        return ''

    week_labels = list(weeks_data.keys())
    n_weeks = len(week_labels)
    n_bins  = len(top_bins)

    # Build matrix: weeks × bins  (fail % of total fails that week)
    matrix = np.zeros((n_weeks, n_bins))
    for wi, wk in enumerate(week_labels):
        bin_cnts: dict[str, int] = {}
        for r in weeks_data[wk]:
            for row in _get_rows(r):
                k = str(row.get('ibin', '')).strip()
                if k:
                    bin_cnts[k] = bin_cnts.get(k, 0) + (row.get('fail_count') or 0)
        total = sum(bin_cnts.values()) or 1
        for bi, bk in enumerate(top_bins):
            matrix[wi, bi] = bin_cnts.get(bk, 0) / total * 100

    x = np.arange(n_weeks)
    bar_w = 0.65

    fig, ax = plt.subplots(figsize=(max(8, n_weeks * 1.1), 5))
    bottoms = np.zeros(n_weeks)
    for bi, bk in enumerate(top_bins):
        vals = matrix[:, bi]
        lbl  = _ibin_label(bk, {}, cfg)
        ax.bar(x, vals, bar_w, bottom=bottoms,
               label=lbl, color=_WEEK_COLORS[bi % len(_WEEK_COLORS)],
               alpha=0.82, edgecolor='white', linewidth=0.3)
        for wi, (v, b) in enumerate(zip(vals, bottoms)):
            if v >= 2.0:
                ax.text(wi, b + v / 2, f'{v:.1f}%',
                        ha='center', va='center', fontsize=6.5,
                        color='white', weight='bold')
        bottoms += vals

    ax.set_xticks(x)
    ax.set_xticklabels(week_labels, rotation=30, ha='right', fontsize=9)
    ax.set_ylabel('% of Total Failures')
    ax.set_title(f'Top {n_bins} iBin Fail Mix — Week-over-Week', fontsize=12, weight='bold')
    ax.legend(fontsize=7.5, bbox_to_anchor=(1.01, 1), loc='upper left', borderaxespad=0)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


# ---------------------------------------------------------------------------
# 5.  HTML generation
# ---------------------------------------------------------------------------

def generate_report(dash_path: Path, runs_data: list[dict],
                    output_path: Path, weeks_back: int = 0,
                    interval: str = 'weekly',
                    cfg: dict | None = None) -> None:
    """
    Generate the pareto report HTML.
    weeks_back=0 → include all runs; weeks_back=N → last N periods only.
    interval: 'daily' | 'weekly' | 'bi-weekly' | 'monthly'
    cfg: product config dict (from trend_chart.load_product_config) for ibin names.
    """
    # Apply period filter
    if weeks_back > 0:
        cutoff_week = _week_start_n_back(weeks_back)
        runs_data = [r for r in runs_data
                     if _period_sort_key(r['week']) >= cutoff_week or r['week'] == 'unknown']

    weeks_data = group_by_week(runs_data)
    n_total = sum(len(v) for v in weeks_data.values())

    sections_html = ''

    # ── Yield trend ──────────────────────────────────────────────────────
    trend_b64 = build_weekly_yield_trend(weeks_data, metric='ff_df')
    if trend_b64:
        sections_html += (
            '<div class="section">'
            f'<h2>&#128200; FF+DF Yield \u2014 {interval.title()} Trend</h2>'
            f'<img class="chart" src="data:image/png;base64,{trend_b64}"/>'
            '</div>'
        )

    trend_ff_b64 = build_weekly_yield_trend(weeks_data, metric='ff')
    if trend_ff_b64:
        sections_html += (
            '<div class="section">'
            f'<h2>&#128200; FF Yield \u2014 {interval.title()} Trend</h2>'
            f'<img class="chart" src="data:image/png;base64,{trend_ff_b64}"/>'
            '</div>'
        )

    # ── Stacked fail-mix chart ────────────────────────────────────────────────
    mix_b64 = build_weekly_top_fails_trend(weeks_data, cfg=cfg)
    if mix_b64:
        sections_html += (
            '<div class="section">'
            f'<h2>&#128203; iBin Fail Mix \u2014 {interval.title()}-over-{interval.title()}</h2>'
            f'<img class="chart" src="data:image/png;base64,{mix_b64}"/>'
            '</div>'
        )

    # ── Per-period pareto charts ──────────────────────────────────────────────
    for week_str, week_runs in weeks_data.items():
        pareto_b64 = build_weekly_ibin_pareto(week_str, week_runs, cfg=cfg)
        period_dt  = _period_sort_key(week_str)
        period_label = (period_dt.strftime('%b %d, %Y') if week_str != 'unknown'
                        else 'Unknown period')
        card_rows  = ''.join(
            f'<li style="font-size:13px;color:#aaa">{_esc(r["name"])}'
            f'{(" — "+_esc(r["ts"])) if r.get("ts") else ""}</li>'
            for r in week_runs
        )
        sections_html += f'''
<div class="section">
  <h2>&#128204; {_esc(week_str)}&ensp;<span style="font-size:18px;color:#7f8c8d;font-weight:normal">
    ({period_label} — {len(week_runs)} run{"s" if len(week_runs)!=1 else ""})</span></h2>
  <ul style="margin:4px 0 10px 18px;padding:0">{card_rows}</ul>
  {'<img class="chart" src="data:image/png;base64,' + pareto_b64 + '"/>' if pareto_b64
   else '<p style="color:#888">No bin data available for this period.</p>'}
</div>'''

    ts_now = datetime.now().strftime('%Y-%m-%d %H:%M')
    html = f'''<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>Weekly Yield Pareto Report</title>
<style>
body{{font-family:Arial,sans-serif;background:#f4f6f8;margin:0;padding:16px 24px}}
h1{{font-size:28px;color:#2c3e50;margin-bottom:4px}}
h2{{font-size:22px;color:#2c3e50;margin:18px 0 8px;padding-bottom:4px;border-bottom:2px solid #dce1e7}}
.subtitle{{font-size:16px;color:#7f8c8d;margin-bottom:20px}}
.dash-link{{font-size:16px;color:#2980b9;margin-bottom:6px;display:block}}
.dash-link a{{color:#2980b9;text-decoration:none}}
.dash-link a:hover{{text-decoration:underline}}
.section{{background:#fff;border-radius:8px;padding:16px 18px;margin-bottom:18px;
  box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.chart{{max-width:100%;height:auto;display:block;margin:8px 0}}
</style>
</head>
<body>
<h1>&#128200; Yield Pareto Report</h1>
<div class="dash-link">Source: <a href="{_esc(str(dash_path.name))}">{_esc(str(dash_path.name))}</a></div>
<div class="subtitle">
  Generated: {ts_now} &nbsp;|&nbsp;
  Interval: <b>{interval.title()}</b> &nbsp;|&nbsp;
  Periods: <b>{len(weeks_data)}</b> &nbsp;|&nbsp;
  Runs: <b>{n_total}</b>
  {f'&nbsp;|&nbsp; Last <b>{weeks_back}</b> period(s)' if weeks_back > 0 else ''}
</div>
{sections_html}
</body>
</html>'''

    output_path.write_text(_wm_inject(html), encoding='utf-8')
    print(f'Wrote weekly report: {output_path}')


def _week_start_n_back(n: int) -> datetime:
    """Return the Monday of the ISO week that is n weeks before the current week."""
    today = datetime.today()
    iso_yr, iso_wk, _ = today.isocalendar()
    current_monday    = today - timedelta(days=today.weekday())
    return current_monday - timedelta(weeks=n - 1)


# ---------------------------------------------------------------------------
# 6.  Update Dashboard.html — REPORT section links
# ---------------------------------------------------------------------------

def update_dashboard_report_links(dash_path: Path, report_path: Path) -> None:
    """Inject a link to report_path into the <!-- REPORT_START/END --> section."""
    from datetime import datetime as _dt
    dash_path   = Path(dash_path)
    report_path = Path(report_path)
    if not dash_path.exists():
        return

    content = dash_path.read_text(encoding='utf-8')

    REPORT_START = '<!-- REPORT_START -->'
    REPORT_END   = '<!-- REPORT_END -->'
    YIELD_END    = '<!-- YIELD_END -->'
    COMPARE_END  = '<!-- COMPARE_END -->'

    # Ensure REPORT section exists
    if REPORT_START not in content:
        anchor = COMPARE_END if COMPARE_END in content else (
                 YIELD_END   if YIELD_END   in content else '</body>')
        insert_after = anchor if anchor != '</body>' else ''
        inject_section = (
            '\n<h2 class="section-header">&#128196; Report</h2>\n'
            + REPORT_START + '\n' + REPORT_END
        )
        if anchor == '</body>':
            content = content.replace('</body>', inject_section + '\n</body>', 1)
        else:
            content = content.replace(anchor, anchor + inject_section, 1)

    try:
        href = os.path.relpath(str(report_path), str(dash_path.parent)).replace('\\', '/')
    except Exception:
        href = report_path.as_uri()

    stem = report_path.stem
    ts   = _dt.now().strftime('%Y-%m-%d %H:%M')

    new_block = (
        f'<div class="run-block" data-stem="{stem}">\n'
        f'<div class="run-header" onclick="toggle(this)">'
        f'<span class="arrow">&#9660;</span> {stem}'
        f'<span class="ts"> - {ts}</span></div>\n'
        f'<div class="run-body">\n'
        f'<a class="run-link report-link" href="{href}" target="_blank">{stem}</a>\n'
        f'</div>\n</div>'
    )

    block_re = re.compile(
        r'<div class="run-block" data-stem="' + re.escape(stem) +
        r'">\s*<div[^>]*>[\s\S]*?</div>\s*</div>', re.MULTILINE)
    if block_re.search(content):
        content = block_re.sub(new_block, content)
    else:
        content = content.replace(REPORT_START, REPORT_START + '\n' + new_block)

    dash_path.write_text(content, encoding='utf-8')
    print(f'Updated {dash_path.name} with report link.')


# ---------------------------------------------------------------------------
# 7.  Main
# ---------------------------------------------------------------------------

def main():
    os.umask(0o002)  # ensure generated files are group-writable on NFS/Samba
    p = argparse.ArgumentParser(description='Weekly pareto yield report from Dashboard.html')
    p.add_argument('dashboard', help='Path to Dashboard.html')
    p.add_argument('--out', default='', help='Output HTML path (default: next to Dashboard.html)')
    p.add_argument('--weeks', type=int, default=0,
                   help='Limit to last N ISO weeks (0 = all)')
    args = p.parse_args()

    if not HAVE_CR:
        print('ERROR: compare_runs.py not found on sys.path', file=sys.stderr)
        sys.exit(1)
    if not HAVE_MPL:
        print('WARNING: matplotlib not installed — charts will be skipped.')

    dash_path = Path(args.dashboard).resolve()
    if not dash_path.exists():
        print(f'ERROR: Dashboard.html not found: {dash_path}', file=sys.stderr)
        sys.exit(1)

    dash_dir = dash_path.parent
    print(f'Parsing {dash_path} …')
    records = _parse_dashboard(dash_path)
    if not records:
        print('ERROR: No run blocks found.', file=sys.stderr)
        sys.exit(1)

    print(f'Found {len(records)} run(s). Loading bin data …')
    runs_data = load_run_data(records, dash_dir, log=lambda s: print(s, end=''))

    out_path = (Path(args.out).resolve() if args.out
                else dash_dir / 'yield_weekly_report.html')

    print('Generating weekly pareto report …')
    generate_report(dash_path, runs_data, out_path, weeks_back=args.weeks)
    update_dashboard_report_links(dash_path, out_path)

    try:
        os.startfile(str(out_path))
    except Exception:
        pass


if __name__ == '__main__':
    main()


# ════════════════════════════════════════════════════════════════
# (formerly yield_report_frame.py)
# ════════════════════════════════════════════════════════════════
import os
import sys
import threading
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox

sys.path.insert(0, str(Path(__file__).parent))

# ── Palette ──────────────────────────────────────────────────────────────────
BG   = '#1a252f'
BG2  = '#2c3e50'
FG   = '#ecf0f1'
FG2  = '#95a5a6'
BLUE = '#2980b9'
ABLU = '#3498db'
GRN  = '#27ae60'
AGRN = '#2ecc71'


def _btn(parent, text, cmd, color=BLUE, acolor=ABLU, width=None):
    kw = dict(text=text, command=cmd, bg=color, fg='white',
              activebackground=acolor, activeforeground='white',
              relief='flat', cursor='hand2', font=('Arial', 9),
              padx=8, pady=3)
    if width:
        kw['width'] = width
    return tk.Button(parent, **kw)


def _lf(parent, text, label_color=FG2):
    return tk.LabelFrame(parent, text=text, bg=BG, fg=label_color,
                         font=('Arial', 8, 'bold'), padx=6, pady=4,
                         relief='groove', bd=1)


# ---------------------------------------------------------------------------
# ReportFrame
# ---------------------------------------------------------------------------

class ReportFrame(tk.Frame):
    def __init__(self, parent=None, **kw):
        super().__init__(parent, bg=BG, **kw)
        self._dash_path   = tk.StringVar()
        self._cfg_var     = tk.StringVar()
        self._out_var     = tk.StringVar()
        self._weeks_var   = tk.StringVar(value='0')
        self._interval_var = tk.StringVar(value='weekly')
        self._last_report = ''
        self._build_ui()

    # ------------------------------------------------------------------ UI --

    def _build_ui(self):
        P = {'padx': 10, 'pady': 4}

        # Title
        tk.Label(self, text='Yield Pareto Report',
                 bg=BG, fg=ABLU, font=('Arial', 13, 'bold')
                 ).pack(fill='x', padx=10, pady=(8, 2))
        tk.Label(self,
                 text='Groups all runs by the chosen interval and generates a bin-fail pareto per period.',
                 bg=BG, fg=FG2, font=('Arial', 8)
                 ).pack(fill='x', padx=10, pady=(0, 4))

        # Step 1 — Input CSV / ZIP / GZ
        frm1 = _lf(self, 'Step 1 — Input CSV / ZIP / GZ', ABLU)
        frm1.pack(fill='x', **P)
        tk.Label(frm1, text='Lot, Wafer, Program Name, Interface Bin, Count, Total Dies  —  Accepts .csv, .zip, .gz',
                 bg=BG, fg=FG2, font=('Arial', 7)).pack(anchor='w', pady=(0, 2))
        entry_row = tk.Frame(frm1, bg=BG)
        entry_row.pack(fill='x')
        tk.Entry(entry_row, textvariable=self._dash_path, width=52,
                 bg=BG2, fg=FG, insertbackground=FG,
                 relief='flat', font=('Consolas', 9)
                 ).pack(side='left', padx=(0, 4), pady=2, expand=True, fill='x')
        _btn(entry_row, 'Browse…', self._browse).pack(side='left', padx=(0, 4))
        _btn(entry_row, 'Load',    self._load,  color='#1f618d').pack(side='left')

        # Run list (read-only summary)
        frm2 = _lf(self, 'Runs found in CSV', '#9b59b6')
        frm2.pack(fill='both', expand=True, **P)
        list_outer = tk.Frame(frm2, bg=BG2, relief='flat', bd=1)
        list_outer.pack(fill='both', expand=True)
        sb = tk.Scrollbar(list_outer, orient='vertical', bg=BG2, troughcolor=BG)
        self._run_listbox = tk.Listbox(
            list_outer, height=8, selectmode='extended',
            bg=BG2, fg=FG, selectbackground='#1f618d', selectforeground='white',
            activestyle='none', font=('Consolas', 9), relief='flat',
            yscrollcommand=sb.set)
        sb.config(command=self._run_listbox.yview)
        sb.pack(side='right', fill='y')
        self._run_listbox.pack(side='left', fill='both', expand=True)

        # Step 3 — Options + output
        frm3 = _lf(self, 'Step 2 — Options', FG2)
        frm3.pack(fill='x', **P)

        # Interval radio buttons
        int_row = tk.Frame(frm3, bg=BG)
        int_row.pack(fill='x', pady=(0, 4))
        tk.Label(int_row, text='Interval:', bg=BG, fg=FG,
                 font=('Arial', 9), width=22, anchor='w').pack(side='left')
        for iv in yr.INTERVALS:
            tk.Radiobutton(
                int_row, text=iv.title(), variable=self._interval_var, value=iv,
                bg=BG, fg=FG, selectcolor=BG2, activebackground=BG,
                activeforeground=FG, font=('Arial', 9), relief='flat'
            ).pack(side='left', padx=(0, 6))

        # Product config JSON (optional — for ibin names)
        cfg_row = tk.Frame(frm3, bg=BG)
        cfg_row.pack(fill='x', pady=(0, 4))
        tk.Label(cfg_row, text='Product config (optional):', bg=BG, fg=FG,
                 font=('Arial', 9), width=22, anchor='w').pack(side='left')
        tk.Entry(cfg_row, textvariable=self._cfg_var, width=38,
                 bg=BG2, fg=FG, insertbackground=FG,
                 relief='flat', font=('Consolas', 8)
                 ).pack(side='left', padx=(0, 4), expand=True, fill='x')
        _btn(cfg_row, 'Browse…', self._browse_cfg, width=8).pack(side='left', padx=(0, 4))
        _btn(cfg_row, 'Clear',   lambda: self._cfg_var.set(''),
             color='#6d3b01', width=6).pack(side='left')

        # Weeks/periods filter
        wk_row = tk.Frame(frm3, bg=BG)
        wk_row.pack(fill='x', pady=(0, 4))
        tk.Label(wk_row, text='Last N periods (0 = all):', bg=BG, fg=FG,
                 font=('Arial', 9), width=22, anchor='w').pack(side='left')
        tk.Entry(wk_row, textvariable=self._weeks_var, width=6,
                 bg=BG2, fg='white', insertbackground='white',
                 relief='flat', font=('Consolas', 9)).pack(side='left', padx=(0, 8))
        tk.Label(wk_row,
                 text='e.g. 8 = show only the 8 most recent periods',
                 bg=BG, fg=FG2, font=('Arial', 8)).pack(side='left')

        # Output file
        out_row = tk.Frame(frm3, bg=BG)
        out_row.pack(fill='x')
        tk.Label(out_row, text='Output file:', bg=BG, fg=FG,
                 font=('Arial', 9), width=12, anchor='w').pack(side='left')
        tk.Entry(out_row, textvariable=self._out_var, width=46,
                 bg=BG2, fg=FG, insertbackground=FG,
                 relief='flat', font=('Consolas', 9)
                 ).pack(side='left', padx=(0, 4), expand=True, fill='x')
        _btn(out_row, '…', self._browse_out, width=3).pack(side='left')

        # Buttons
        btn_row = tk.Frame(self, bg=BG)
        btn_row.pack(pady=(6, 2), padx=10, fill='x')
        self._run_btn = _btn(btn_row, '▶  Generate Report', self._generate,
                             color=GRN, acolor=AGRN)
        self._run_btn.config(font=('Arial', 10, 'bold'), pady=5)
        self._run_btn.pack(side='left', expand=True, fill='x', padx=(0, 4))
        self._open_btn = _btn(btn_row, '  Open Dashboard  ', self._open_dashboard,
                              color='#935116', acolor='#ca6f1e')
        self._open_btn.config(font=('Arial', 10, 'bold'), pady=5, state='disabled')
        self._open_btn.pack(side='left')

        # Log
        log_frm = _lf(self, 'Log', FG2)
        log_frm.pack(fill='both', expand=False, **P)
        self._log = tk.Text(log_frm, height=6, state='disabled',
                            font=('Consolas', 8), bg='#0d1b26', fg='#a8d8ea',
                            relief='flat', insertbackground=FG)
        self._log.pack(fill='both', expand=True)

    # ---------------------------------------------------------------- events --

    def _browse(self):
        p = filedialog.askopenfilename(
            title='Select input CSV / ZIP / GZ',
            filetypes=[
                ('Supported files', '*.csv *.zip *.gz *.gzip'),
                ('CSV files', '*.csv'),
                ('ZIP archives', '*.zip'),
                ('GZ archives', '*.gz *.gzip'),
                ('All files', '*.*'),
            ])
        if p:
            self._dash_path.set(p)
            # Auto-detect product config if not already set
            if not self._cfg_var.get().strip():
                try:
                    import trend_chart as tc
                    auto = tc._find_auto_config()
                    if auto:
                        self._cfg_var.set(str(auto))
                        self._log_write(f'Auto-detected product config: {auto.name}\n')
                except Exception:
                    pass
            self._load()

    def _browse_cfg(self):
        p = filedialog.askopenfilename(
            title='Select product config JSON',
            filetypes=[('JSON files', '*.json'), ('All files', '*.*')])
        if p:
            self._cfg_var.set(p)

    def _browse_out(self):
        p = filedialog.asksaveasfilename(
            title='Save report as',
            defaultextension='.html',
            filetypes=[('HTML files', '*.html')])
        if p:
            self._out_var.set(p)

    def _open_dashboard(self):
        rep = self._last_report
        if rep and os.path.isfile(rep):
            try:
                os.startfile(rep)
            except Exception as e:
                messagebox.showerror('Error', str(e))

    def _log_write(self, msg: str):
        def _do():
            self._log.configure(state='normal')
            self._log.insert('end', msg)
            self._log.see('end')
            self._log.configure(state='disabled')
        try:
            self.after(0, _do)
        except Exception:
            pass

    # ------------------------------------------------------------------ load --

    def _load(self):
        path_str = self._dash_path.get().strip()
        if not path_str:
            return
        csv_path = Path(path_str)
        if not csv_path.exists():
            messagebox.showerror('Not found', f'File not found:\n{csv_path}')
            return
        try:
            import trend_chart as tc
            runs = tc.load_csv(csv_path)
        except Exception as exc:
            messagebox.showerror('Parse error', str(exc))
            return
        if not runs:
            messagebox.showwarning('No runs', 'No run data found in file.')
            return

        self._run_listbox.delete(0, 'end')
        interval = self._interval_var.get()
        for r in runs:
            ts = r.get('date_str', '') or ''
            period = yr._ts_to_period(ts, interval) or '?'
            self._run_listbox.insert('end',
                f'{r.get("label", "")}   {ts}   [{period}]')

        self._out_var.set(str(csv_path.parent / (csv_path.stem + '_report.html')))
        self._open_btn.configure(state='normal')
        self._log_write(f'Loaded {len(runs)} run(s) from {csv_path.name}\n')

    # ------------------------------------------------------------- generate --

    def _generate(self):
        csv_str = self._dash_path.get().strip()
        if not csv_str:
            messagebox.showwarning('No data', 'Load a CSV / ZIP / GZ file first.')
            return
        csv_path = Path(csv_str)
        out_path = Path(self._out_var.get().strip() or
                        csv_path.parent / (csv_path.stem + '_report.html'))
        interval  = self._interval_var.get()
        try:
            weeks_back = int(self._weeks_var.get().strip() or '0')
        except ValueError:
            weeks_back = 0

        # Load product config if provided
        cfg = None
        cfg_path_str = self._cfg_var.get().strip()
        if cfg_path_str:
            try:
                import trend_chart as tc
                cfg = tc.load_product_config(cfg_path_str)
            except Exception as e:
                self._log_write(f'Warning: could not load product config: {e}\n')

        self._run_btn.configure(state='disabled', text='Working\u2026', bg=FG2)
        self._log_write('Loading run data\u2026\n')

        def _worker():
            try:
                runs_data = yr.runs_from_csv(csv_path,
                                             log=self._log_write,
                                             interval=interval)
                self._log_write('Generating report\u2026\n')
                yr.generate_report(csv_path, runs_data, out_path,
                                   weeks_back=weeks_back,
                                   interval=interval,
                                   cfg=cfg)
                self._log_write(f'Done → {out_path}\n')
                self._last_report = str(out_path)
                try:
                    os.startfile(str(out_path))
                except Exception:
                    pass
            except Exception as exc:
                import traceback
                self._log_write(f'ERROR: {exc}\n{traceback.format_exc()}\n')
            finally:
                def _re_enable():
                    self._run_btn.configure(state='normal',
                                            text='▶  Generate Report', bg=GRN)
                try:
                    self.after(0, _re_enable)
                except Exception:
                    pass

        threading.Thread(target=_worker, daemon=True).start()

