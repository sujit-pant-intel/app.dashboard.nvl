"""yield_automation.py — Merged yield automation utilities.

Consolidates (yld/automation/): compare_runs, aqua_autopull,
  auto_pull_and_run, generate_index, manage_email, patch_css, patch_css_v2,
  run_automation, serve_reports

Entry points (unchanged):
  run_automation.py  — scheduled automation orchestrator
  serve_reports.py   — local HTTP report server
"""
from __future__ import annotations

# ════════════════════════════════════════════════════════════════
# (formerly compare_runs.py)
# ════════════════════════════════════════════════════════════════
#!/usr/bin/env python3
"""
compare_runs.py
---------------
Reads Dashboard.html, resolves each run's *_out.xlsx, and generates a
standalone comparison HTML report.

Usage:
    python compare_runs.py Dashboard.html
    python compare_runs.py Dashboard.html --out my_comparison.html
    python compare_runs.py Dashboard.html --ref "NCXSDJXP0H51M202611-1"
"""

import sys
import os
import re
import argparse
import io
import base64
from pathlib import Path


def _wm_inject(html: str) -> str:
    _wm = (
        '<div id="_wm_div" style="position:fixed;top:8px;right:12px;font-size:10px;'
        'font-weight:600;pointer-events:none;z-index:99999;'
        'font-family:Arial,sans-serif;user-select:none;letter-spacing:0.04em;'
        'padding:2px 6px;border-radius:3px;background:transparent;">'
        'Pant, Sujit N \u2014 GEMS FTE</div>'
        '<script>(function(){'
        'function _wm_color(){'
        'var d=document.getElementById("_wm_div");if(!d)return;'
        'var bg=window.getComputedStyle(document.body).backgroundColor;'
        'var m=bg.match(/\\d+/g);'
        'if(m&&m.length>=3){'
        'var r=+m[0],g=+m[1],b=+m[2];'
        'var lum=0.299*r+0.587*g+0.114*b;'
        'd.style.color=lum<128?"rgba(255,255,255,0.9)":"rgba(20,20,20,0.75)";'
        '}else{d.style.color="rgba(255,255,255,0.9)";}'
        '}'
        'if(document.readyState==="loading")'
        '{document.addEventListener("DOMContentLoaded",_wm_color);}'
        'else{_wm_color();}'
        '})();</script>'
    )
    import re as _re_wm
    if '</body>' not in html:
        return html
    html = _re_wm.sub(
        r'<div[^>]*id=["\']_wm_div["\'][^>]*>[\s\S]*?</div>\s*<script[^>]*>[\s\S]*?</script>',
        '', html)
    html = _re_wm.sub(r'<div[^>]*>[^<]*GEMS FTE[^<]*</div>', '', html)
    return html.replace('</body>', _wm + '\n</body>', 1)

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np
    HAVE_MPL = True
except ImportError:
    HAVE_MPL = False

# ---------------------------------------------------------------------------
# 1. Parse Dashboard.html → run records
# ---------------------------------------------------------------------------

def parse_dashboard(dash_path: Path):
    """Return list of dicts: {stem, name, ts, index_href}.
    Only returns blocks from the Yield section (YIELD_START/END or legacy RUNS_START/END)."""
    content = dash_path.read_text(encoding='utf-8')
    # Extract only the Yield section content
    yield_html = ''
    for start_pat, end_pat in [
        (r'<!--\s*YIELD_START\s*-->', r'<!--\s*YIELD_END\s*-->'),
        (r'<!--\s*RUNS_START\s*-->',  r'<!--\s*RUNS_END\s*-->'),   # legacy
    ]:
        m = re.search(start_pat + r'(.*?)' + end_pat, content, re.S)
        if m:
            yield_html = m.group(1)
            break
    # Fall back to full content only if no sentinels found at all
    search_html = yield_html if yield_html.strip() else content
    runs = []
    # Each block: <div class="run-block" data-stem="...">...</div></div>
    block_re = re.compile(
        r'<div class="run-block" data-stem="([^"]+)">([\s\S]*?)</div>\s*</div>',
        re.MULTILINE
    )
    for m in block_re.finditer(search_html):
        stem = m.group(1)
        body = m.group(2)
        # Name and timestamp from run-header
        hdr = re.search(
            r'<span class="arrow">[^<]*</span>\s*([^<]+)<span class="ts">\s*-\s*([^<]*)</span>',
            body
        )
        name = hdr.group(1).strip() if hdr else stem
        ts   = hdr.group(2).strip() if hdr else ''
        # Yield Report href (report-link)
        link_m = re.search(r'class="run-link report-link"[^>]*href="([^"]+)"', body)
        index_href = link_m.group(1) if link_m else None
        runs.append({'stem': stem, 'name': name, 'ts': ts, 'index_href': index_href})
    return runs


# ---------------------------------------------------------------------------
# 2. Resolve *_out.xlsx from an index.html href
# ---------------------------------------------------------------------------

def find_xlsx(dash_dir: Path, index_href: str):
    """Given Dashboard.html dir and the relative/absolute href to index.html,
    return Path to *_out.xlsx in the same folder, or None."""
    if not index_href:
        return None
    # Strip file:// scheme if present
    href = re.sub(r'^file:///', '', index_href).replace('/', os.sep)
    if os.path.isabs(href):
        idx_path = Path(href)
    else:
        idx_path = dash_dir / href
    out_folder = idx_path.parent
    if not out_folder.exists():
        return None
    _bd = out_folder / 'bin_dist'
    candidates = sorted((_bd if _bd.exists() else out_folder).glob('*_out.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


# ---------------------------------------------------------------------------
# 2b. Resolve *_BinDistribution.html and parse RDND + Bin Fail tables
# ---------------------------------------------------------------------------

def find_bin_html(output_dir: Path):
    """Return Path to *_BinDistribution.html in output_dir/bin_dist, or None."""
    _bd = output_dir / 'bin_dist'
    candidates = sorted((_bd if _bd.exists() else output_dir).glob('*_BinDistribution.html'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def find_group_medians(output_dir: Path):
    """Return Path to Group_Medians.csv in output_dir or sicc/, or None."""
    candidates = sorted(
        list(output_dir.glob('Group_Medians.csv')) +
        list((output_dir / 'sicc').glob('Group_Medians.csv')),
        key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def find_raw_csv(output_dir: Path):
    """Return Path to the raw CSV in output_dir's parent (the source CSV), or None."""
    parent = output_dir.parent
    if not parent.exists():
        return None
    # Match CSVs that look like the source data (not bindef or generated)
    candidates = [p for p in sorted(parent.glob('*.csv'))
                  if 'bindef' not in p.name.lower()
                  and '_targets_' not in p.name.lower()]
    # Prefer the one with 'reticle' in the name (most complete), else newest
    reticle = [p for p in candidates if 'reticle' in p.name.lower()]
    if reticle:
        return reticle[0]
    return candidates[0] if candidates else None


def extract_upm_from_csv(csv_path: Path, config_json: str = None):
    """Extract UPM distribution data from raw CSV using config JSON analyses.
    Returns list of dicts compatible with upm_data format, plus a detailed
    list of per-column medians."""
    import pandas as pd
    import json

    if not csv_path or not csv_path.exists():
        return None, None

    # Load config to get analyses + base
    analyses = []
    base_val = None
    upm_prefix = 'UPM_0107'
    if config_json:
        try:
            cfg = json.loads(Path(config_json).read_text(encoding='utf-8'))
            analyses = cfg.get('analyses', [])
            for anl in analyses:
                if anl.get('type') == 'distribution':
                    filt = anl.get('filter', {}).get('match', {})
                    filt_method = filt.get('method', 'starts_with')
                    filt_value = filt.get('value', upm_prefix)
                    agg = anl.get('aggregation', {})
                    if agg.get('mode') == 'percentage':
                        bc = agg.get('base', {})
                        if bc.get('type') == 'fixed':
                            base_val = bc.get('value')
                    break
        except Exception:
            pass

    try:
        df = pd.read_csv(str(csv_path), dtype=object)
    except Exception:
        return None, None

    import re as _re_upm
    import fnmatch as _fnmatch_upm
    _has_wc = ('*' in filt_value or '?' in filt_value)
    if _has_wc:
        upm_cols = [c for c in df.columns if _fnmatch_upm.fnmatch(c, filt_value)]
    elif filt_method == 'regex':
        upm_cols = [c for c in df.columns if _re_upm.search(filt_value, c)]
    elif filt_method == 'contains':
        upm_cols = [c for c in df.columns if filt_value in c]
    elif filt_method == 'starts_with':
        upm_cols = [c for c in df.columns if c.startswith(filt_value)]
    else:
        upm_cols = [c for c in df.columns if c.startswith(filt_value)]
    if not upm_cols:
        return None, None

    if base_val is None:
        base_val = 9154  # default

    # Per-column detail
    col_details = []
    overall_pct = None
    for col in upm_cols:
        vals = pd.to_numeric(df[col], errors='coerce').dropna()
        if vals.empty:
            continue
        med = float(vals.median())
        pct = (med / base_val) * 100
        short = col[len(upm_prefix):].strip('_') if col.startswith(upm_prefix) else col
        col_details.append({
            'col': col, 'short': short, 'median': med,
            'mean': float(vals.mean()), 'count': len(vals), 'pct': pct
        })
        # Use 0950 column as the "main" UPM % if available
        if '0950' in col:
            overall_pct = pct

    if not col_details:
        return None, None

    # If no 0950 column, use the first
    if overall_pct is None:
        overall_pct = col_details[0]['pct']

    # Build legacy-compatible upm_data (single row with upm_pct)
    upm_data = [{'test': 'UPM', 'n_rows': col_details[0]['count'],
                 'sicc_actual': None, 'sicc_target': None,
                 'multiple': None, 'upm_pct': overall_pct}]

    return upm_data, col_details


def parse_group_medians(csv_path: Path) -> list[dict]:
    """Return list of {test, n_rows, sicc_actual, sicc_target, multiple, upm_pct}."""
    rows = []
    try:
        text = csv_path.read_text(encoding='utf-8-sig')
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            return rows
        header = [h.strip() for h in lines[0].split(',')]
        for line in lines[1:]:
            parts = [p.strip() for p in line.split(',')]
            if len(parts) < 2:
                continue
            def _f(i):
                try: return float(parts[i]) if i < len(parts) and parts[i] else None
                except: return None
            rows.append({
                'test':        parts[0] if parts else '',
                'n_rows':      _f(1),
                'sicc_actual': _f(2),
                'sicc_target': _f(3),
                'multiple':    _f(4),
                'upm_pct':     _f(5),
            })
    except Exception as e:
        print(f'Warning: could not read Group_Medians.csv: {e}')
    return rows


def find_cdyn_medians(output_dir: Path):
    """Return Path to cdyn_medians.csv in output_dir or sicc/, or None."""
    candidates = sorted(
        list(output_dir.glob('cdyn_medians.csv')) +
        list((output_dir / 'sicc').glob('cdyn_medians.csv')),
        key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def parse_cdyn_medians(csv_path: Path) -> list[dict]:
    """Return list of {test, type, actual, expected, ratio}."""
    rows = []
    try:
        text = csv_path.read_text(encoding='utf-8-sig')
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if not lines:
            return rows
        for line in lines[1:]:
            parts = [p.strip() for p in line.split(',')]
            if len(parts) < 2:
                continue
            def _f(i):
                try: return float(parts[i]) if i < len(parts) and parts[i] else None
                except: return None
            rows.append({
                'test':     parts[0] if parts else '',
                'type':     parts[1] if len(parts) > 1 else '',
                'actual':   _f(2),
                'expected': _f(3),
                'ratio':    _f(4),
            })
    except Exception as e:
        print(f'Warning: could not read cdyn_medians.csv: {e}')
    return rows


def _strip_tags(s: str) -> str:
    return re.sub(r'<[^>]+>', '', s).strip()


def parse_bin_html(bin_html_path: Path):
    """Return dict:
        yield_rows:    [{bin, fail_bucket, yield_pct, expected_pct}, ...]
        bin_fail_rows: [{ibin, fail_bucket, fail_pct, fail_count}, ...]
        func_bin_rows: [{ibin, fbin, fail_bucket, fail_pct, fail_count}, ...]
    """
    content = bin_html_path.read_text(encoding='utf-8')

    # --- RDND yield table ---
    yield_rows = []
    yt_m = re.search(r'<table class="yield-table">(.*?)</table>', content, re.DOTALL)
    if yt_m:
        for tr in re.findall(r'<tr>(.*?)</tr>', yt_m.group(1), re.DOTALL):
            tds = re.findall(r'<td(?:[^>]*)>(.*?)</td>', tr, re.DOTALL)
            if len(tds) < 3:
                continue
            # Column order: BIN(0), FAIL BUCKET(1), YIELD(2), EXPECTED(3)
            bin_name     = _strip_tags(tds[0])
            fail_bucket  = _strip_tags(tds[1]) if len(tds) > 1 else ''
            yield_str    = _strip_tags(tds[2]).rstrip('%') if len(tds) > 2 else ''
            expected_str = _strip_tags(tds[3]).rstrip('%') if len(tds) > 3 else ''
            try:    yield_pct    = float(yield_str)
            except: yield_pct    = None
            try:    expected_pct = float(expected_str)
            except: expected_pct = None
            if bin_name:
                yield_rows.append({'bin': bin_name, 'fail_bucket': fail_bucket,
                                   'yield_pct': yield_pct, 'expected_pct': expected_pct})

    # --- Bin Fail Summary table (pareto-tbl with Interface Bin header) ---
    bin_fail_rows = []
    func_bin_rows = []
    bin_summary_rows = []  # new 6-col format: ibin, cat, desc, total, fail_count, fail_pct
    for tbl in re.findall(r'<table class="pareto-tbl"[^>]*>(.*?)</table>', content, re.DOTALL):
        if 'Interface Bin' not in tbl:
            continue
        has_fbin    = 'Functional Bin' in tbl
        has_cat_desc = 'Category' in tbl and 'Description' in tbl
        for tr in re.findall(r'<tr[^>]*>(.*?)</tr>', tbl, re.DOTALL):
            tds = re.findall(r'<td(?:[^>]*)>(.*?)</td>', tr, re.DOTALL)
            if has_cat_desc:
                # new format: Interface Bin(0), Category(1), Description(2),
                #              Total Count(3), Fail Count(4), Yield/Fail %(5)
                if len(tds) < 6:
                    continue
                ibin         = _strip_tags(tds[0]).strip().rstrip('\u26a0').strip()
                cat          = _strip_tags(tds[1])
                desc         = _strip_tags(tds[2])
                fail_cnt_str = _strip_tags(tds[4]).replace(',', '')
                fail_pct_str = _strip_tags(tds[5]).rstrip('%')
                try:    fail_pct   = float(fail_pct_str)
                except: fail_pct   = None
                try:    fail_count = int(fail_cnt_str)
                except: fail_count = None
                if ibin and fail_pct is not None:
                    bin_summary_rows.append({'ibin': ibin, 'cat': cat, 'desc': desc,
                                             'fail_pct': fail_pct, 'fail_count': fail_count})
            elif has_fbin:
                # cols: Interface Bin(0), Functional Bin(1), Fail Bucket(2),
                #       Description(3), Total Count(4), Fail Count(5), Fail Count %(6)
                if len(tds) < 7:
                    continue
                ibin         = _strip_tags(tds[0])
                fbin         = _strip_tags(tds[1])
                fail_bucket  = _strip_tags(tds[2])
                fail_pct_str = _strip_tags(tds[6]).rstrip('%')
                fail_cnt_str = _strip_tags(tds[5]).replace(',', '')
                try:    fail_pct   = float(fail_pct_str)
                except: fail_pct   = None
                try:    fail_count = int(fail_cnt_str)
                except: fail_count = None
                if ibin and fail_pct is not None:
                    func_bin_rows.append({'ibin': ibin, 'fbin': fbin,
                                          'fail_bucket': fail_bucket,
                                          'fail_pct': fail_pct, 'fail_count': fail_count})
            else:
                # old format: Interface Bin(0), Fail Bucket(1), Total Count(2), Fail Count(3), Fail %(4)
                if len(tds) < 5:
                    continue
                ibin         = _strip_tags(tds[0])
                fail_bucket  = _strip_tags(tds[1])
                fail_pct_str = _strip_tags(tds[4]).rstrip('%')
                fail_cnt_str = _strip_tags(tds[3]).replace(',', '')
                try:    fail_pct   = float(fail_pct_str)
                except: fail_pct   = None
                try:    fail_count = int(fail_cnt_str)
                except: fail_count = None
                if ibin and fail_pct is not None:
                    bin_fail_rows.append({'ibin': ibin, 'fail_bucket': fail_bucket,
                                          'fail_pct': fail_pct, 'fail_count': fail_count})

    # Fallback for newer BinDistribution HTML where rows are rendered dynamically
    # from JS arrays/objects (DATA / BFS_DATA / FP_DATA) instead of static <tr> rows.
    if not bin_summary_rows or not func_bin_rows:
        try:
            import json as _json

            def _extract_js_array(var_name: str):
                m = re.search(
                    rf'var\s+{re.escape(var_name)}\s*=\s*(\[[\s\S]*?\]);',
                    content,
                    re.DOTALL,
                )
                if not m:
                    return []
                try:
                    return _json.loads(m.group(1))
                except Exception:
                    return []

            if not yield_rows:
                m_data = re.search(r'var\s+DATA\s*=\s*({[\s\S]*?});', content, re.DOTALL)
                if m_data:
                    try:
                        data_obj = _json.loads(m_data.group(1))
                    except Exception:
                        data_obj = {}
                    rows = data_obj.get('rows', []) if isinstance(data_obj, dict) else []
                    ydefs = data_obj.get('yieldDefs', []) if isinstance(data_obj, dict) else []
                    total_die = data_obj.get('total', 0) if isinstance(data_obj, dict) else 0
                    if not total_die:
                        total_die = sum(
                            int(r.get('total', 0) or 0)
                            for r in rows if isinstance(r, dict)
                        )
                    if total_die and ydefs:
                        for yd in ydefs:
                            if not isinstance(yd, dict):
                                continue
                            bins_list = [str(b) for b in (yd.get('bins_list') or [])]
                            cnt = 0
                            for rr in rows:
                                if not isinstance(rr, dict):
                                    continue
                                bc = rr.get('binCounts') or {}
                                if not isinstance(bc, dict):
                                    continue
                                cnt += sum(int(bc.get(bk, 0) or 0) for bk in bins_list)
                            try:
                                expected_pct = float(yd.get('expected')) if yd.get('expected') not in (None, '') else None
                            except Exception:
                                expected_pct = None
                            yield_rows.append({
                                'bin': str(yd.get('bins', '') or ''),
                                'fail_bucket': str(yd.get('bucket', '') or ''),
                                'yield_pct': (cnt / float(total_die) * 100.0) if total_die else None,
                                'expected_pct': expected_pct,
                            })

            if not bin_summary_rows:
                bfs_rows = _extract_js_array('BFS_DATA')
                for r in bfs_rows:
                    ibin_raw = str(r.get('bin', '')).strip()
                    # Stored values may contain warning symbol (e.g., "15⚠").
                    ibin = ibin_raw.rstrip('\u26a0').strip()
                    try:
                        fail_pct = float(r.get('pct'))
                    except Exception:
                        fail_pct = None
                    try:
                        fail_count = int(r.get('count'))
                    except Exception:
                        fail_count = None
                    if ibin and fail_pct is not None:
                        bin_summary_rows.append({
                            'ibin': ibin,
                            'cat': str(r.get('cat', '') or ''),
                            'desc': str(r.get('desc', '') or ''),
                            'fail_pct': fail_pct,
                            'fail_count': fail_count,
                        })

            if not func_bin_rows:
                fp_rows = _extract_js_array('FP_DATA')
                for r in fp_rows:
                    try:
                        fbin = str(int(float(r.get('fb'))))
                    except Exception:
                        fbin = str(r.get('fb', '') or '')
                    try:
                        fail_pct = float(r.get('pct'))
                    except Exception:
                        fail_pct = None
                    try:
                        fail_count = int(r.get('count'))
                    except Exception:
                        fail_count = None
                    if fbin and fail_pct is not None:
                        # Keep key name "ibin" because compare chart/table logic expects it.
                        func_bin_rows.append({
                            'ibin': fbin,
                            'fbin': fbin,
                            'fail_bucket': str(r.get('bkt', '') or ''),
                            'fail_pct': fail_pct,
                            'fail_count': fail_count,
                        })
        except Exception:
            pass

    return {'yield_rows': yield_rows, 'bin_fail_rows': bin_fail_rows,
            'func_bin_rows': func_bin_rows, 'bin_summary_rows': bin_summary_rows}


# ---------------------------------------------------------------------------
# 3. Read xlsx → structured data
# ---------------------------------------------------------------------------

def read_xlsx(xlsx_path: Path):
    """Return dict with keys:
        num_die, col_headers, groups, totals
        groups: list of {name, rows: [(label, vals...)], sum_vals: (vals...)}
        totals: {label, vals} for 'Yield SUM' row
    """
    if not HAVE_OPENPYXL:
        return None
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    ws = wb[wb.sheetnames[0]]

    # Build raw cell grid
    grid = {}
    fmt_grid = {}
    for ri, row in enumerate(ws.iter_rows(values_only=False)):
        for ci, cell in enumerate(row):
            grid[(ri, ci)] = cell.value
            fmt_grid[(ri, ci)] = cell.number_format or ''

    # Find num_die
    num_die = None
    for ri in range(ws.max_row):
        if grid.get((ri, 0)) == '# Die':
            num_die = grid.get((ri, 1))
            break

    if not num_die:
        return None

    # Memoised formula evaluator
    _ev_cache = {}

    def _ev(ri, ci):
        if (ri, ci) in _ev_cache:
            return _ev_cache[(ri, ci)]
        v = grid.get((ri, ci))
        if v is None:
            _ev_cache[(ri, ci)] = None
            return None
        if not isinstance(v, str) or not v.startswith('='):
            _ev_cache[(ri, ci)] = v
            return v
        result = None
        # =N/B2
        m1 = re.match(r'^=(-?\d+)/B\d+$', v)
        if m1:
            result = int(m1.group(1)) / num_die
        # =SUM(X3:X6)
        elif re.match(r'^=SUM\([A-Z]+\d+:[A-Z]+\d+\)$', v):
            m2 = re.match(r'^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', v)
            if m2:
                c1 = column_index_from_string(m2.group(1)) - 1
                r1 = int(m2.group(2)) - 1
                c2 = column_index_from_string(m2.group(3)) - 1
                r2 = int(m2.group(4)) - 1
                result = sum(
                    _ev(r, c1) for r in range(r1, r2 + 1)
                    if isinstance(_ev(r, c1), (int, float))
                )
        # =SUM(X3,X7,...) — comma separated refs
        elif re.match(r'^=SUM\(([^)]+)\)$', v):
            m3 = re.match(r'^=SUM\(([^)]+)\)$', v)
            if m3:
                total = 0.0
                for ref in m3.group(1).split(','):
                    mr = re.match(r'^([A-Z]+)(\d+)$', ref.strip())
                    if mr:
                        rc = column_index_from_string(mr.group(1)) - 1
                        rr = int(mr.group(2)) - 1
                        sv = _ev(rr, rc)
                        if isinstance(sv, (int, float)):
                            total += sv
                result = total
        _ev_cache[(ri, ci)] = result
        return result

    def _pct(ri, ci):
        v = _ev(ri, ci)
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return round(v * 100, 2)
        return None

    # Find column headers row
    col_hdrs = []
    hdr_ri = None
    for ri in range(ws.max_row):
        if grid.get((ri, 0)) == 'Sub Module':
            hdr_ri = ri
            col_hdrs = [grid.get((ri, ci)) for ci in range(ws.max_column)
                        if grid.get((ri, ci)) is not None]
            break

    if hdr_ri is None:
        return None

    num_cols = len(col_hdrs) - 1  # exclude 'Sub Module' column

    # Which data columns are percentages (header contains '%')
    col_is_pct = ['%' in str(h) for h in col_hdrs[1:]]

    def _read_val(ri, ci):
        """Return pct-scaled value for % columns, raw value otherwise."""
        idx = ci - 1  # 0-based into col_is_pct
        if idx < len(col_is_pct) and col_is_pct[idx]:
            return _pct(ri, ci)
        v = _ev(ri, ci)
        if isinstance(v, (int, float)):
            return v
        return None

    # Parse data rows into groups
    groups = []
    current_rows = []
    totals = None

    for ri in range(hdr_ri + 1, ws.max_row + 1):
        label = grid.get((ri, 0))
        if label is None:
            continue
        label_str = str(label).strip()
        if not label_str or label_str.startswith('#'):
            continue

        vals = [_read_val(ri, ci) for ci in range(1, 1 + num_cols)]

        if label_str.upper() == 'SUM':
            groups.append({
                'rows': current_rows,
                'sum_vals': vals,
            })
            current_rows = []
        elif re.match(r'^Yield\s+SUM', label_str, re.IGNORECASE):
            totals = {'label': label_str, 'vals': vals}
        else:
            current_rows.append({'label': label_str, 'vals': vals})

    # Flush any trailing rows without a SUM (e.g. TPI/Other group)
    if current_rows:
        groups.append({'rows': current_rows, 'sum_vals': [None] * num_cols})

    return {
        'num_die': num_die,
        'col_headers': col_hdrs[1:],
        'col_is_pct': col_is_pct,
        'groups': groups,
        'totals': totals,
        'sheet': wb.sheetnames[0],
    }


# ---------------------------------------------------------------------------
# 4. Chart helpers
# ---------------------------------------------------------------------------

def _fig_b64(fig, dpi=130):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('ascii')


def _esc(s):
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


# Color palette for identifiers (up to 10)
_ID_COLORS = [
    '#2980b9', '#27ae60', '#e74c3c', '#f39c12',
    '#8e44ad', '#16a085', '#d35400', '#2c3e50',
    '#c0392b', '#1abc9c',
]

_GROUP_COLORS = [
    '#eaf4ea', '#e3f0fc', '#fff6e6', '#f3e8fb',
    '#fdecea', '#e0f7fa', '#fdf6e3', '#f9ece8',
]


def build_sum_comparison_chart(runs_data, col_idx=0):
    """Grouped bar chart: each group's SUM value per identifier."""
    if not HAVE_MPL:
        return ''
    labels = [r['name'] for r in runs_data]
    # Collect group names from the run with most groups
    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    groups = best['data']['groups']
    group_names = []
    for g in groups:
        if g['rows']:
            # Use prefix of first row's label as group name
            first = g['rows'][0]['label']
            prefix = re.match(r'^([A-Za-z]+)', first)
            group_names.append(prefix.group(1) if prefix else first[:8])
        else:
            group_names.append('?')

    n_groups = len(group_names)
    n_runs = len(runs_data)
    x = np.arange(n_groups)
    width = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(max(10, n_groups * 1.4), 5))
    for ri, run in enumerate(runs_data):
        if not run['data']:
            continue
        vals = []
        for gi, g in enumerate(run['data']['groups']):
            sv = g['sum_vals'][col_idx] if col_idx < len(g['sum_vals']) else None
            vals.append(sv if sv is not None else 0.0)
        # Pad if fewer groups
        while len(vals) < n_groups:
            vals.append(0.0)
        offset = (ri - n_runs / 2 + 0.5) * width
        bars = ax.bar(x + offset, vals, width, label=run['name'],
                      color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)
        for bar, v in zip(bars, vals):
            if v and v > 0.3:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.1,
                        f'{v:.1f}%', ha='center', va='bottom', fontsize=6.5, rotation=90)

    col_label = runs_data[0]['data']['col_headers'][col_idx] if runs_data[0]['data'] else f'Col {col_idx+1}'
    ax.set_title(f'Group SUM Comparison — {col_label}', fontsize=13, weight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(group_names, rotation=30, ha='right', fontsize=9)
    ax.set_ylabel('Yield / Fallout (%)')
    ax.legend(fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_submodule_chart(runs_data, col_idx=0):
    """Horizontal bar chart: each sub-module row per identifier (top N)."""
    if not HAVE_MPL:
        return ''
    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    all_rows = []
    for g in best['data']['groups']:
        all_rows.extend(g['rows'])

    if not all_rows:
        return ''

    # Sort by first run's value descending (top 20)
    def _first_val(row_label):
        for r in runs_data:
            if not r['data']:
                continue
            for g in r['data']['groups']:
                for row in g['rows']:
                    if row['label'] == row_label:
                        v = row['vals'][col_idx] if col_idx < len(row['vals']) else None
                        return v or 0.0
        return 0.0

    sorted_rows = sorted(all_rows, key=lambda r: _first_val(r['label']), reverse=True)[:20]
    row_labels = [r['label'] for r in sorted_rows]

    n_runs = len(runs_data)
    n_rows = len(row_labels)
    y = np.arange(n_rows)
    height = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(9, max(5, n_rows * 0.45)))
    for ri, run in enumerate(runs_data):
        if not run['data']:
            continue
        vals = []
        for lbl in row_labels:
            v = 0.0
            for g in run['data']['groups']:
                for row in g['rows']:
                    if row['label'] == lbl:
                        v = row['vals'][col_idx] if col_idx < len(row['vals']) else 0.0
                        if v is None: v = 0.0
            vals.append(v)
        offset = (ri - n_runs / 2 + 0.5) * height
        ax.barh(y + offset, vals, height, label=run['name'],
                color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)

    col_label = runs_data[0]['data']['col_headers'][col_idx] if runs_data[0]['data'] else f'Col {col_idx+1}'
    ax.set_title(f'Sub-Module Breakdown — {col_label}', fontsize=13, weight='bold')
    ax.set_yticks(y)
    ax.set_yticklabels(row_labels, fontsize=8)
    ax.set_xlabel('Fallout / Yield (%)')
    ax.legend(fontsize=8)
    ax.grid(axis='x', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_total_yield_chart(runs_data):
    """Single bar chart: Yield SUM (%) per identifier."""
    if not HAVE_MPL:
        return ''
    labels = [r['name'] for r in runs_data]
    vals = []
    for r in runs_data:
        if r['data'] and r['data']['totals']:
            v = r['data']['totals']['vals'][0]
            vals.append(v if v is not None else 0.0)
        else:
            vals.append(0.0)

    if not any(vals):
        return ''

    fig, ax = plt.subplots(figsize=(max(6, len(labels) * 1.2), 4))
    bars = ax.bar(labels, vals, color=[_ID_COLORS[i % len(_ID_COLORS)] for i in range(len(labels))],
                  alpha=0.88, edgecolor='white')
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.2,
                f'{v:.1f}%', ha='center', va='bottom', fontsize=9, weight='bold')
    ax.set_title('Total Yield Loss SUM by Identifier', fontsize=13, weight='bold')
    ax.set_ylabel('Yield Loss (%)')
    ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=9)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_delta_heatmap(runs_data, ref_name, col_idx=0):
    """Heatmap of delta (run - ref) per sub-module."""
    if not HAVE_MPL:
        return ''
    ref_run = next((r for r in runs_data if r['name'] == ref_name), runs_data[0])
    compare_runs = [r for r in runs_data if r['name'] != ref_run['name']]
    if not compare_runs:
        return ''

    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    all_rows = []
    for g in best['data']['groups']:
        all_rows.extend(g['rows'])
    row_labels = [r['label'] for r in all_rows]

    def _get_val(run, lbl):
        if not run['data']:
            return None
        for g in run['data']['groups']:
            for row in g['rows']:
                if row['label'] == lbl:
                    return row['vals'][col_idx] if col_idx < len(row['vals']) else None
        return None

    ref_vals = [_get_val(ref_run, lbl) for lbl in row_labels]
    matrix = []
    col_labels = []
    for r in compare_runs:
        row_vals = [_get_val(r, lbl) for lbl in row_labels]
        deltas = [
            (rv - refv) if (rv is not None and refv is not None) else float('nan')
            for rv, refv in zip(row_vals, ref_vals)
        ]
        matrix.append(deltas)
        col_labels.append(r['name'])

    data_arr = np.array(matrix)  # shape: (n_compare, n_rows)
    if data_arr.size == 0:
        return ''

    fig, ax = plt.subplots(figsize=(max(8, len(row_labels) * 0.55), max(3, len(col_labels) * 0.7 + 1.5)))
    # Diverging colormap: red = higher fallout (worse), green = lower (better)
    vmax = np.nanmax(np.abs(data_arr))
    if vmax == 0:
        vmax = 1
    im = ax.imshow(data_arr, cmap='RdYlGn_r', aspect='auto',
                   vmin=-vmax, vmax=vmax)
    ax.set_xticks(range(len(row_labels)))
    ax.set_xticklabels(row_labels, rotation=60, ha='right', fontsize=7)
    ax.set_yticks(range(len(col_labels)))
    ax.set_yticklabels(col_labels, fontsize=8)
    for yi in range(len(col_labels)):
        for xi in range(len(row_labels)):
            v = data_arr[yi, xi]
            if not np.isnan(v):
                ax.text(xi, yi, f'{v:+.1f}', ha='center', va='center',
                        fontsize=6, color='black')
    plt.colorbar(im, ax=ax, label='Δ % vs reference')
    col_label = runs_data[0]['data']['col_headers'][col_idx] if runs_data[0]['data'] else ''
    ax.set_title(f'Delta vs {ref_name} — {col_label}\n(red = more fallout, green = less)',
                 fontsize=11, weight='bold')
    fig.tight_layout()
    return _fig_b64(fig)


# ---------------------------------------------------------------------------
# 4b. RDND yield-table charts and Bin Fail chart
# ---------------------------------------------------------------------------

# Palette for stacked fail segments (consistent across runs)
_FAIL_COLORS = [
    '#e74c3c', '#e67e22', '#f39c12', '#2ecc71',
    '#1abc9c', '#3498db', '#9b59b6', '#e8177d',
    '#95a5a6', '#34495e', '#c0392b', '#16a085',
]


def build_top10_pareto_chart(runs_data):
    """Horizontal bar chart: top-10 Interface Bins by max fail% — uses Functional Bin table."""
    if not HAVE_MPL:
        return ''
    # Prefer func_bin_rows (has fbin); fall back to bin_fail_rows
    use_fbin = any(r.get('bin_data') and r['bin_data'].get('func_bin_rows')
                   for r in runs_data)
    row_key  = 'func_bin_rows' if use_fbin else 'bin_fail_rows'
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data'].get(row_key)]
    if not valid:
        return ''

    # Collect all ibins; keep max fail% per ibin
    ibin_max    = {}
    ibin_fbin   = {}
    ibin_bucket = {}
    for run in valid:
        for row in run['bin_data'][row_key]:
            k = row['ibin']
            v = row['fail_pct'] or 0.0
            if v > ibin_max.get(k, 0.0):
                ibin_max[k]    = v
                ibin_bucket[k] = row.get('fail_bucket', '')
                ibin_fbin[k]   = row.get('fbin', '')

    top10 = sorted(ibin_max.keys(), key=lambda k: ibin_max[k], reverse=True)[:10]
    if not top10:
        return ''

    n_runs = len(valid)
    n_bins = len(top10)
    y      = np.arange(n_bins)
    height = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(9, max(4, n_bins * 0.65)))
    for ri, run in enumerate(valid):
        vals = [
            next((ro['fail_pct'] for ro in run['bin_data'][row_key] if ro['ibin'] == k), 0.0) or 0.0
            for k in top10
        ]
        offset = (ri - n_runs / 2 + 0.5) * height
        bars = ax.barh(y + offset, vals, height, label=run['name'],
                       color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.87)
        for bar, v in zip(bars, vals):
            if v >= 0.1:
                ax.text(v + 0.05, bar.get_y() + bar.get_height() / 2,
                        f'{v:.2f}%', va='center', ha='left', fontsize=7)

    if use_fbin:
        ylabels = [
            f'iBin {k}  FBin {ibin_fbin.get(k, "—")}  |  {ibin_bucket.get(k, "")}'
            for k in top10
        ]
    else:
        ylabels = [
            f'iBin {k}  |  {ibin_bucket.get(k, "")}'
            for k in top10
        ]
    ax.set_yticks(y)
    ax.set_yticklabels(ylabels, fontsize=7.5)
    ax.invert_yaxis()
    ax.set_xlabel('Fail (%)')
    ax.set_title('Top 10 Interface Bin Fail Pareto', fontsize=13, weight='bold')
    ax.legend(fontsize=8)
    ax.grid(axis='x', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_upm_median_chart(runs_data):
    """Grouped bar chart: SICC Si Actual Median vs Target per test, per run."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('upm_data') and r['upm_data']]
    if not valid:
        return ''

    # Collect test names from run with most rows
    best = max(valid, key=lambda r: len(r['upm_data']))
    # Only tests that have both actual and target
    tests = [row for row in best['upm_data'] if row.get('sicc_actual') is not None]
    if not tests:
        return ''
    test_labels = [row['test'] for row in tests]

    n_tests = len(test_labels)
    n_runs  = len(valid)
    x       = np.arange(n_tests)
    width   = 0.7 / max(n_runs + 1, 2)  # +1 for target bar

    fig, ax = plt.subplots(figsize=(max(7, n_tests * 0.5), 4))

    # Plot target once (from best run)
    targets = [t.get('sicc_target') for t in tests]
    has_target = any(v is not None for v in targets)

    for ri, run in enumerate(valid):
        actuals = []
        for t in tests:
            row = next((r for r in run['upm_data'] if r['test'] == t['test']), None)
            actuals.append(row['sicc_actual'] if row and row.get('sicc_actual') is not None else 0.0)
        offset = ri * width
        ax.bar(x + offset, actuals, width, label=run['name'],
               color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)

    if has_target:
        tgt_vals = [v if v is not None else 0.0 for v in targets]
        offset = n_runs * width
        ax.bar(x + offset, tgt_vals, width, label='Target',
               color='#2c3e50', alpha=0.55, hatch='//')

    # Centre ticks
    tick_offset = (n_runs * width) / 2
    ax.set_xticks(x + tick_offset)
    ax.set_xticklabels(test_labels, rotation=35, ha='right', fontsize=8)
    ax.set_ylabel('SICC Si Actual Median (A)')
    ax.set_title('SICC Median', fontsize=13, weight='bold')
    ax.yaxis.set_major_locator(plt.MaxNLocator(nbins=12, integer=False))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: '0' if v == 0 else f'{v:.3f}'))
    ax.legend(fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def build_upm_pct_chart(runs_data, upm_target_pct=None, upm_target_label='Target'):
    """Bar chart: UPM % per run (single value per run)."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('upm_data') and r['upm_data']]
    if not valid:
        return ''

    labels = []
    vals   = []
    for run in valid:
        pcts = [row['upm_pct'] for row in run['upm_data'] if row.get('upm_pct') is not None]
        if pcts:
            labels.append(run['name'])
            vals.append(pcts[0])  # UPM % is the same for all rows

    if not vals:
        return ''

    fig, ax = plt.subplots(figsize=(max(4, len(labels) * 0.9), 3.5))
    bars = ax.bar(labels, vals,
                  color=[_ID_COLORS[i % len(_ID_COLORS)] for i in range(len(labels))],
                  alpha=0.88, edgecolor='white')
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.1,
                f'{v:.2f}%', ha='center', va='bottom', fontsize=9, weight='bold')
    if upm_target_pct is not None:
        ax.axhline(upm_target_pct, color='red', linewidth=2, linestyle='--',
                   label=f'{upm_target_label}: {upm_target_pct}%', zorder=5)
        ax.legend(fontsize=9)
    ax.set_title('UPM ULVT 950mV (%)', fontsize=13, weight='bold')
    ax.set_ylabel('UPM (%)')
    _ymax = max(vals + ([upm_target_pct] if upm_target_pct is not None else [])) * 1.15 if vals else 100
    ax.set_ylim(0, _ymax)
    ax.yaxis.set_major_locator(plt.MaxNLocator(nbins=10, integer=False))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: '0' if v == 0 else f'{v:.2f}%'))
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=9)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def build_upm_detail_chart(runs_data, upm_target_pct=None, upm_target_label='Target'):
    """Grouped bar chart: UPM median % per column per run."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('upm_detail') and r['upm_detail']]
    if not valid:
        return ''

    import numpy as np

    # Collect all unique short names across runs (preserve order)
    all_shorts = []
    for run in valid:
        for d in run['upm_detail']:
            if d['short'] not in all_shorts:
                all_shorts.append(d['short'])

    if not all_shorts:
        return ''

    n_cols = len(all_shorts)
    n_runs = len(valid)
    bar_w = 0.8 / n_runs
    x = np.arange(n_cols)

    fig, ax = plt.subplots(figsize=(max(6, n_cols * 1.2 + n_runs * 0.3), 4.5))

    for ri, run in enumerate(valid):
        detail_map = {d['short']: d['pct'] for d in run['upm_detail']}
        vals = [detail_map.get(s, 0) for s in all_shorts]
        offset = x + ri * bar_w - (n_runs - 1) * bar_w / 2
        bars = ax.bar(offset, vals, width=bar_w,
                      color=_ID_COLORS[ri % len(_ID_COLORS)],
                      alpha=0.85, edgecolor='white', label=run['name'])
        for bar, v in zip(bars, vals):
            if v > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.15,
                        f'{v:.1f}%', ha='center', va='bottom', fontsize=6, weight='bold')

    if upm_target_pct is not None:
        ax.axhline(upm_target_pct, color='red', linewidth=2, linestyle='--',
                   label=f'{upm_target_label}: {upm_target_pct}%', zorder=5)

    ax.set_title('UPM Distribution Comparison (Median %)', fontsize=12, weight='bold')
    ax.set_ylabel('Median (%)', fontsize=10)
    ax.set_xticks(x)
    ax.set_xticklabels(all_shorts, rotation=45, ha='right', fontsize=7)
    ax.legend(fontsize=8, loc='best')
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def _bin_all_good(bin_str: str) -> bool:
    """Return True if every number in the bin string is <= 4."""
    nums = [int(x) for x in re.findall(r'\d+', str(bin_str))]
    return bool(nums) and all(n <= 4 for n in nums)


# Bin groups to always show in the top subplot — matched by bin string equality
_KEY_BINS = ['1/2/3/4', '1/2']
# Fallback: treat any row whose label exactly matches one of these
_KEY_BIN_TITLES = {
    '1/2/3/4': 'FF+DF  (Bin 1/2/3/4)',
    '1/2':     'FF  (Bin 1/2)',
}


def build_combined_rdnd_chart(runs_data):
    """Single mixed-format chart:
       - Stacked bars (per identifier) for Bins > 4 fail%      [left Y, 0-100%]
       - Line overlay (per key bin) for FF+DF / FF yield%       [right Y, 0-100%]
       - Dashed hline = expected yield per key bin
       - Dashed tick marker = expected fail total per identifier
    """
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]
    if not valid:
        return ''

    from matplotlib.patches import Patch
    from matplotlib.lines import Line2D

    best = max(valid, key=lambda r: len(r['bin_data']['yield_rows']))
    all_yield_rows = best['bin_data']['yield_rows']

    key_rows  = [row for row in all_yield_rows if row['bin'] in _KEY_BINS]
    if not key_rows:
        key_rows = [row for row in all_yield_rows if _bin_all_good(row['bin'])][:2]
    fail_rows = [row for row in all_yield_rows if not _bin_all_good(row['bin'])]

    n_runs    = len(valid)
    id_labels = [r['name'] for r in valid]
    x         = np.arange(n_runs)
    bar_w     = 0.5

    fig, ax_bar = plt.subplots(figsize=(max(10, n_runs * 2.2), 6))
    ax_line = ax_bar.twinx()          # shared x, independent right Y

    legend_elems = []

    # ---- Stacked bars: Bins > 4 fail% ----
    bottoms = np.zeros(n_runs)
    for si, fr in enumerate(fail_rows):
        bin_key = fr['bin']
        seg_lbl = fr.get('fail_bucket') or bin_key
        clr     = _FAIL_COLORS[si % len(_FAIL_COLORS)]
        vals    = np.array([
            (next((ro['yield_pct'] for ro in r['bin_data']['yield_rows']
                   if ro['bin'] == bin_key), None) or 0.0)
            for r in valid
        ])
        ax_bar.bar(x, vals, bar_w, bottom=bottoms, color=clr,
                   alpha=0.80, edgecolor='white', linewidth=0.5, zorder=2)
        for bi, (v, b) in enumerate(zip(vals, bottoms)):
            if v >= 0.8:
                ax_bar.text(bi, b + v / 2, f'{v:.1f}%',
                            ha='center', va='center', fontsize=6.5,
                            color='white', weight='bold')
        legend_elems.append(Patch(color=clr, alpha=0.80, label=seg_lbl))
        bottoms += vals

    # Dashed expected-total tick per identifier
    for xi, run in enumerate(valid):
        exp_total = sum(
            (next((ro['expected_pct'] for ro in run['bin_data']['yield_rows']
                   if ro['bin'] == fr['bin']), None) or 0.0)
            for fr in fail_rows
        )
        if exp_total:
            ax_bar.plot([xi - 0.28, xi + 0.28], [exp_total, exp_total],
                        color='#2c3e50', linestyle='--', linewidth=2.0, zorder=5)

    legend_elems.append(Line2D([0], [0], color='#2c3e50', linestyle='--',
                                linewidth=2.0, label='Expected Fail Total'))

    ax_bar.set_ylabel('Fail (%)',  fontsize=10)
    # Dynamic y-limit: max stacked fail total across all runs + 15% buffer
    max_fail = float(bottoms.max()) if bottoms.max() > 0 else 1.0
    also_exp = []
    for run in valid:
        exp_t = sum(
            (next((ro['expected_pct'] for ro in run['bin_data']['yield_rows']
                   if ro['bin'] == fr['bin']), None) or 0.0)
            for fr in fail_rows
        )
        if exp_t:
            also_exp.append(exp_t)
    if also_exp:
        max_fail = max(max_fail, max(also_exp))
    fail_ylim = min(100.0, max_fail * 1.20)   # 20% headroom, cap at 100
    fail_ylim = max(fail_ylim, 5.0)            # at least 5% so chart is readable
    ax_bar.set_ylim(0, fail_ylim)
    ax_bar.set_xticks(x)
    ax_bar.set_xticklabels(id_labels, rotation=15, ha='right', fontsize=9)
    ax_bar.grid(axis='y', linestyle='--', alpha=0.3, zorder=0)

    # ---- Lines: key-bin yield% (right Y axis) ----
    line_colors = ['#1a73e8', '#e53935', '#2e7d32', '#f57c00']
    for ki, kr in enumerate(key_rows):
        bin_key = kr['bin']
        lbl = _KEY_BIN_TITLES.get(bin_key,
                                  f"{bin_key} {kr.get('fail_bucket', '')}".strip())
        lclr = line_colors[ki % len(line_colors)]
        vals = np.array([
            (next((ro['yield_pct'] for ro in r['bin_data']['yield_rows']
                   if ro['bin'] == bin_key), None) or 0.0)
            for r in valid
        ])
        ax_line.plot(x, vals, marker='o', linewidth=2.4, markersize=8,
                     color=lclr, label=lbl, zorder=6)
        for xi, v in enumerate(vals):
            ax_line.text(xi, v + 1.5, f'{v:.1f}%', ha='center', va='bottom',
                         fontsize=8, color=lclr, weight='bold')
        # Dashed expected hline
        ev = kr.get('expected_pct')
        if ev is not None:
            ax_line.axhline(ev, color=lclr, linestyle=':', linewidth=1.6,
                            alpha=0.65, zorder=4)
            ax_line.text(n_runs - 0.45, ev - 2.5 - ki * 3.0, f'Exp {ev:.1f}%',
                         ha='right', va='top', fontsize=7.5,
                         color=lclr, alpha=0.85)
        legend_elems.append(Line2D([0], [0], color=lclr, marker='o',
                                   linewidth=2.4, markersize=8, label=lbl))

    ax_line.set_ylabel('Yield (%)', fontsize=10, y=0.4)
    ax_line.set_ylim(0, 100)

    ax_bar.set_title(
        'Yield (%) and Fail (%) Chart',
        fontsize=12, weight='bold'
    )
    ax_bar.legend(handles=legend_elems, fontsize=7.5,
                  loc='upper left', bbox_to_anchor=(1.18, 1.0),
                  borderaxespad=0)

    fig.tight_layout(pad=2.0)
    return _fig_b64(fig)


def build_fail_stacked_chart(runs_data):
    """Kept for backward compat — now delegates to combined chart."""
    return build_combined_rdnd_chart(runs_data)


def build_rdnd_delta_chart(runs_data):
    """Bars: (actual yield - expected) per BIN per run. Green = beating target, red = below."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]
    if not valid:
        return ''
    best = max(valid, key=lambda r: len(r['bin_data']['yield_rows']))
    yield_rows    = best['bin_data']['yield_rows']
    bin_labels    = [row['bin'] for row in yield_rows]
    expected_vals = [row['expected_pct'] for row in yield_rows]

    n_bins = len(bin_labels)
    n_runs = len(valid)
    x      = np.arange(n_bins)
    width  = 0.8 / max(n_runs, 1)

    from matplotlib.patches import Patch
    fig, ax = plt.subplots(figsize=(max(12, n_bins * 1.1), 4))
    legend_elems = []
    for ri, run in enumerate(valid):
        actuals = [
            next((r['yield_pct'] for r in run['bin_data']['yield_rows'] if r['bin'] == bl), None)
            for bl in bin_labels
        ]
        deltas = [
            (a - e) if (a is not None and e is not None) else 0.0
            for a, e in zip(actuals, expected_vals)
        ]
        offset = (ri - n_runs / 2 + 0.5) * width
        bar_colors = ['#27ae60' if d > 0 else ('#c0392b' if d < 0 else '#95a5a6')
                      for d in deltas]
        ax.bar(x + offset, deltas, width, color=bar_colors, alpha=0.85)
        legend_elems.append(Patch(color=_ID_COLORS[ri % len(_ID_COLORS)],
                                  alpha=0.85, label=run['name']))
    ax.axhline(0, color='#2c3e50', linewidth=1)
    ax.set_title('RDND Yield Delta vs Expected  (actual - expected)\n'
                 'Green = beating target  |  Red = below target',
                 fontsize=12, weight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(bin_labels, rotation=35, ha='right', fontsize=7)
    ax.set_ylabel('\u0394 Yield (%)')
    ax.legend(handles=legend_elems, fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


def build_bin_fail_chart(runs_data):
    """Horizontal grouped bars: fail% per Interface Bin per run, sorted by max fail."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['bin_fail_rows']]
    if not valid:
        return ''

    # Union of ibins in order from the run with most rows
    best = max(valid, key=lambda r: len(r['bin_data']['bin_fail_rows']))
    ibin_meta = {}
    for row in best['bin_data']['bin_fail_rows']:
        ibin_meta.setdefault(row['ibin'], row['fail_bucket'])

    def _max_fail(ibin):
        return max(
            (next((ro['fail_pct'] for ro in r['bin_data']['bin_fail_rows'] if ro['ibin'] == ibin), 0.0) or 0.0)
            for r in valid
        )

    sorted_ibins = sorted(ibin_meta.keys(), key=_max_fail, reverse=True)
    ibin_labels  = [f"Bin {k} \u2014 {ibin_meta[k]}" for k in sorted_ibins]

    n_rows = len(sorted_ibins)
    n_runs = len(valid)
    y      = np.arange(n_rows)
    height = 0.8 / max(n_runs, 1)

    fig, ax = plt.subplots(figsize=(9, max(5, n_rows * 0.42)))
    for ri, run in enumerate(valid):
        vals = [
            next((ro['fail_pct'] for ro in run['bin_data']['bin_fail_rows'] if ro['ibin'] == k), 0.0) or 0.0
            for k in sorted_ibins
        ]
        offset = (ri - n_runs / 2 + 0.5) * height
        ax.barh(y + offset, vals, height, label=run['name'],
                color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)
    ax.set_title('Bin Fail Summary — Fail% per Interface Bin', fontsize=13, weight='bold')
    ax.set_yticks(y)
    ax.set_yticklabels(ibin_labels, fontsize=7)
    ax.set_xlabel('Fail (%)')
    ax.legend(fontsize=8)
    ax.grid(axis='x', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig)


# ---------------------------------------------------------------------------
# 4c. Run summary table (Program Name, Material Type, Wafer, # Dies)
# ---------------------------------------------------------------------------

def parse_index_meta(dash_dir: Path, index_href: str) -> dict:
    """Parse program, lot(s), wafer(s), material type(s) from a run's BinDistribution HTML
    by reading the embedded DATA JS variable written by bin_distribution_html.py."""
    result = {'program': '', 'lots': [], 'wafers': [], 'material': []}
    if not index_href:
        return result
    try:
        import json as _json_idx
        href = re.sub(r'^file:///', '', index_href).replace('/', os.sep)
        idx_path = dash_dir / href if not os.path.isabs(href) else Path(href)
        out_folder = idx_path.parent
        if not out_folder.exists():
            return result
        # Find the best BinDistribution.html: prefer *_reticle_material_BinDistribution.html
        bin_html = None
        _bd_dir = out_folder / 'bin_dist'
        _glob_dir = _bd_dir if _bd_dir.exists() else out_folder
        for pat in ('*_reticle_material_BinDistribution.html',
                    '*_material_merged_*BinDistribution.html',
                    '*BinDistribution.html'):
            cands = sorted(_glob_dir.glob(pat), key=lambda p: p.stat().st_mtime, reverse=True)
            if cands:
                bin_html = cands[0]
                break
        if not bin_html:
            return result
        # Try data_summary.js (new format: window.DATA={...}) then inline bin_html
        _ds_js = bin_html.parent / 'data_summary.js'
        content = None
        _data_pattern = r'(?:var\s+DATA|window\.DATA)\s*=\s*'
        for _src in [_ds_js, bin_html]:
            try:
                _txt = _src.read_text(encoding='utf-8', errors='replace')
                if re.search(_data_pattern, _txt):
                    content = _txt
                    break
            except Exception:
                continue
        if content is None:
            return result
        # Extract DATA = {...}
        m = re.search(_data_pattern, content)
        if m:
            try:
                data, _ = _json_idx.JSONDecoder().raw_decode(content, m.end())
                rows = data.get('rows', [])
                programs, lots, wafers, mats = set(), set(), set(), set()
                for row in rows:
                    if row.get('program'):
                        programs.add(str(row['program']))
                    if row.get('lot'):
                        lots.add(str(row['lot']))
                    if row.get('wafer') and str(row['wafer']) not in ('', 'all'):
                        wafers.add(str(row['wafer']))
                    if row.get('material'):
                        mats.add(str(row['material']))
                def _sort_nums(s):
                    try:
                        return sorted(s, key=lambda x: int(x) if str(x).isdigit() else x)
                    except Exception:
                        return sorted(s)
                result['program']  = ', '.join(sorted(programs))
                result['lots']     = _sort_nums(lots)
                result['wafers']   = _sort_nums(wafers)
                result['material'] = sorted(mats)
            except Exception:
                pass
    except Exception:
        pass
    return result


def _find_processed_csv(output_dir: Path):
    """Find the best enriched/processed CSV for a run.
    Priority: *_reticle_material.csv > *_material_merged.csv > any CSV in output_dir > parent dir."""
    def _excl(name: str) -> bool:
        n = name.lower()
        return ('bindef' in n or '_targets_' in n or '_bindef' in n)

    def _csvs(d: Path):
        if not d or not d.exists():
            return []
        return [p for p in sorted(d.glob('*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
                if not _excl(p.name)]

    # output_dir is the identifier subfolder — enriched CSVs are written there
    for search_dir in [output_dir,
                       output_dir.parent if output_dir else None,
                       output_dir.parent.parent if output_dir and output_dir.parent else None]:
        if search_dir is None:
            continue
        cands = _csvs(search_dir)
        if not cands:
            continue
        # Strongest preference: fully enriched (reticle + material) file
        for p in cands:
            if '_reticle_material' in p.name.lower():
                return p
        # Second preference: material-merged intermediate
        for p in cands:
            if '_material_merged' in p.name.lower():
                return p
        # Third: any CSV with 'reticle' or 'material' in name
        for p in cands:
            n = p.name.lower()
            if 'reticle' in n or 'material' in n or 'enriched' in n:
                return p
        # Fallback: newest CSV in this dir
        return cands[0]
    return None


def _extract_csv_meta(output_dir: Path) -> dict:
    """Read the processed/enriched CSV to extract Program Name, Material Type, Wafer list."""
    result = {'program': '', 'material': '', 'wafers': ''}
    raw = _find_processed_csv(output_dir)
    if not raw:
        return result
    try:
        import pandas as _pd_meta
        # Read only first 10k rows for speed; use low_memory=False so mixed cols don't truncate
        df = _pd_meta.read_csv(str(raw), dtype=object, nrows=10000)
        prog_col = next((c for c in df.columns if 'program' in c.lower()), None)
        # Material Type column: look for exact match first, then substring
        mat_col = next(
            (c for c in df.columns if c.strip().lower() in
             ('material type', 'materialtype', 'material_type',
              'material type, skew, beol skew')),
            None
        ) or next((c for c in df.columns if 'material type' in c.lower()), None)
        wafer_col = (next((c for c in df.columns if 'sort_wafer' in c.lower()), None)
                     or next((c for c in df.columns if 'wafer' in c.lower()), None))
        if prog_col:
            vals = [str(v) for v in df[prog_col].dropna().unique()]
            result['program'] = ', '.join(vals[:4])
        if mat_col:
            vals = [str(v) for v in df[mat_col].dropna().unique()]
            result['material'] = ', '.join(vals[:6])
        if wafer_col:
            wvals = df[wafer_col].dropna().unique()
            try:
                wvals = sorted(wvals, key=lambda x: int(str(x)) if str(x).isdigit() else str(x))
            except Exception:
                wvals = sorted([str(w) for w in wvals])
            result['wafers'] = ', '.join([str(w) for w in wvals])
    except Exception:
        pass
    return result


def build_run_summary_table_html(runs_data, dash_dir: Path = None) -> str:
    """Compact table: Program Name / Lot(s) / Material Type / Wafer(s) / # Dies — one column per run."""
    metas = []
    for r in runs_data:
        meta = {'program': '', 'lots': '', 'material': '', 'wafers': '', 'num_die': ''}

        # # Dies from xlsx
        if r.get('data') and r['data'].get('num_die') is not None:
            try:
                meta['num_die'] = f"{int(r['data']['num_die']):,}"
            except Exception:
                meta['num_die'] = str(r['data']['num_die'])

        # Program / Lot / Wafer / Material from the run's index.html UDATA
        if dash_dir and r.get('index_href'):
            idx_meta = parse_index_meta(dash_dir, r['index_href'])
            meta['lots']     = _esc(', '.join(idx_meta['lots']))
            meta['wafers']   = _esc(', '.join(idx_meta['wafers']))
            meta['material'] = _esc(', '.join(idx_meta['material']))
            # Program comes from UDATA program field; if empty fall back to run name
            if idx_meta.get('program'):
                meta['program'] = _esc(idx_meta['program'])
            else:
                meta['program'] = _esc(r.get('name', ''))
        else:
            meta['program'] = _esc(r.get('name', ''))

        metas.append(meta)

    ROWS = [
        ('Program Name',  'program'),
        ('Lot(s)',        'lots'),
        ('Material Type', 'material'),
        ('Wafer(s)',      'wafers'),
        ('# Dies',        'num_die'),
    ]

    _th_base = 'white-space:nowrap;padding:6px 10px'
    hdr = f'<th style="background:#1a3a5c;color:#fff;{_th_base}">Metric</th>'
    for ri, r in enumerate(runs_data):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th style="background:{clr};color:#fff;font-weight:bold;'
                f'text-align:center;{_th_base}">'
                f'{_esc(r["name"])}</th>')

    _td_base = 'padding:5px 10px;font-size:20px;white-space:nowrap'
    rows_html = ''
    for label, key in ROWS:
        cells = (f'<td style="font-weight:bold;background:#e8f0fb;color:#1a3a5c;{_td_base}">{label}</td>')
        for mi, m in enumerate(metas):
            bg = '#ffffff' if mi % 2 == 0 else '#f5f8ff'
            val = m[key] or '\u2014'
            cells += (f'<td style="background:{bg};color:#222;{_td_base}">{val}</td>')
        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128204; Run Summary</h2>
  <div style="overflow-x:auto">
  <table style="border-collapse:collapse;table-layout:auto">
    <thead><tr>{hdr}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 4d. RDND and Bin Fail comparison table HTML
# ---------------------------------------------------------------------------

def build_xlsx_comparison_table(runs_data):
    """Side-by-side table of all rows from *_out.xlsx for every identifier."""
    valid = [r for r in runs_data if r.get('data')]
    if not valid:
        return ''

    # All runs (including those with no xlsx) — show '—' columns for missing ones
    all_runs = runs_data

    # Collect all unique sub-module labels in display order from the run with most groups
    best = max(valid, key=lambda r: len(r['data']['groups']))
    ordered_labels = []   # (label, is_sum, group_idx)
    for gi, group in enumerate(best['data']['groups']):
        for row in group['rows']:
            ordered_labels.append((row['label'], False, gi))
        ordered_labels.append(('SUM', True, gi))
    if best['data'] and best['data']['totals']:
        ordered_labels.append((best['data']['totals']['label'], True, -1))

    n_cols = max(len(r['data']['col_headers']) for r in valid)
    col_hdrs = best['data']['col_headers']
    col_is_pct = best['data'].get('col_is_pct', [True] * n_cols)

    def _fmt_val(v, ci):
        """Format cell value: add % for pct columns, plain number for raw columns."""
        if v is None:
            return ''
        if ci < len(col_is_pct) and col_is_pct[ci]:
            return f'{v:.1f}%'
        # Raw column (e.g. Die id count)
        return f'{int(v):,}' if isinstance(v, float) and v == int(v) else f'{v:g}'

    def _get_row_vals(run, lbl, is_sum, gi):
        """Return list of values (one per data column) for a given row."""
        if not run['data']:
            return [None] * n_cols
        if is_sum and lbl.upper() == 'SUM':
            grps = run['data']['groups']
            if gi < len(grps):
                sv = grps[gi]['sum_vals']
                return [sv[ci] if ci < len(sv) else None for ci in range(n_cols)]
            return [None] * n_cols
        if is_sum:  # Yield SUM row
            if run['data']['totals']:
                v = run['data']['totals']['vals']
                return [v[ci] if ci < len(v) else None for ci in range(n_cols)]
            return [None] * n_cols
        for g in run['data']['groups']:
            for row in g['rows']:
                if row['label'] == lbl:
                    return [row['vals'][ci] if ci < len(row['vals']) else None
                            for ci in range(n_cols)]
        return [None] * n_cols

    # Build header — use all_runs so every identifier appears
    hdr = '<th rowspan="2">Sub Module</th>'
    for ri, r in enumerate(all_runs):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        no_data = '' if r.get('data') else ' ⚠ no xlsx'
        hdr += (f'<th colspan="{n_cols}" style="background:{clr};color:#fff;'
                f'font-weight:bold;text-align:center;padding:6px 8px">'
                f'{_esc(r["name"])}{no_data}</th>')
    hdr2 = ''
    for ri, r in enumerate(all_runs):
        for ci in range(n_cols):
            if r.get('data'):
                ch = col_hdrs[ci] if ci < len(col_hdrs) else f'Col {ci+1}'
                hdr2 += f'<th style="font-size:17px;white-space:nowrap">{_esc(ch)}</th>'
            else:
                hdr2 += f'<th style="font-size:17px;white-space:nowrap;color:#aaa">—</th>'

    rows_html = ''
    for lbl, is_sum, gi in ordered_labels:
        grp_clr = _GROUP_COLORS[gi % len(_GROUP_COLORS)] if gi >= 0 else '#d5e8d4'
        is_total = (gi == -1)
        bold = 'font-weight:bold;' if is_sum else ''
        border = 'border-top:2px solid #aaa;' if is_sum and not is_total else ''
        border = 'border-top:3px solid #555;' if is_total else border
        lbl_cell = (f'<td style="{bold}{border}background:{grp_clr};'
                    f'font-size:20px;white-space:nowrap">{_esc(lbl)}</td>')
        cells = lbl_cell
        # Only include valid runs in the highlight calculation
        all_run_vals = [_get_row_vals(rx, lbl, is_sum, gi) for rx in valid]
        for ri, r in enumerate(all_runs):
            if not r.get('data'):
                # No xlsx for this run — show N/A cells
                for ci in range(n_cols):
                    base_st = f'{bold}{border}background:{grp_clr};color:#bbb;'
                    cells += f'<td class="num" style="{base_st}">—</td>'
                continue
            vals = _get_row_vals(r, lbl, is_sum, gi)
            for ci, v in enumerate(vals):
                base_st = f'{bold}{border}background:{grp_clr};'
                if col_is_pct[ci] if ci < len(col_is_pct) else False:
                    col_row_vals = [arv[ci] for arv in all_run_vals]
                    cells += _cell_hl(v, col_row_vals, extra_style=base_st) + _fmt_val(v, ci) + '</td>'
                else:
                    cells += (f'<td class="num" style="{base_st}">'
                              f'{_fmt_val(v, ci)}</td>')
        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128203; Digital Dashboard</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead>
      <tr>{hdr}</tr>
      <tr>{hdr2}</tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


def _cell_hl(v, row_vals, extra_style=''):
    """Return a <td> string. Bold-red if |v - row_mean| > 10 (percentage points)."""
    nums = [x for x in row_vals if x is not None]
    if v is None:
        return f'<td class="num" style="{extra_style}"></td>'
    alert = (len(nums) >= 2 and abs(v - (sum(nums) / len(nums))) > 10)
    st = extra_style + ('color:#c0392b;font-weight:bold;' if alert else '')
    return f'<td class="num" style="{st}">'


def build_rdnd_table_html(runs_data):
    """Table: BIN rows, Expected % + each run's actual yield% + delta vs expected."""
    valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]
    if not valid:
        return ''
    best = max(valid, key=lambda r: len(r['bin_data']['yield_rows']))
    yield_rows = best['bin_data']['yield_rows']

    hdr = '<th>BIN</th><th>Fail Bucket</th><th>Expected (%)</th>'
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += f'<th style="background:{clr};color:#fff;font-weight:bold;padding:5px 8px">{_esc(r["name"])}</th>'

    rows_html = ''
    for row in yield_rows:
        bl  = row['bin']
        exp = row.get('expected_pct')
        cells = (
            f'<td style="white-space:nowrap;font-size:20px">{_esc(bl)}</td>'
            f'<td style="font-size:20px">{_esc(row.get("fail_bucket", ""))}</td>'
            f'<td class="num" style="color:#555">{f"{exp:.1f}%" if exp is not None else ""}</td>'
        )
        # Bins where lower-than-expected is bad (yield bins)
        _yield_bins = ('1', '1/2', '1/2/3/4')
        for r in valid:
            v = next((ro['yield_pct'] for ro in r['bin_data']['yield_rows'] if ro['bin'] == bl), None)
            if v is None:
                cells += '<td class="num"></td>'
            else:
                if exp is not None:
                    if bl in _yield_bins:
                        alert = v < exp
                    else:
                        alert = v > exp
                else:
                    alert = False
                st = 'color:#c0392b;font-weight:bold;' if alert else ''
                cells += f'<td class="num" style="{st}">{v:.1f}%</td>'
        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128203; Yield Table</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead><tr>{hdr}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


def build_bin_fail_table_html(runs_data):
    """Table: Interface Bin rows, each run's Yield/Fail% — uses new bin_summary_rows
    (Category/Description format) when available, else falls back to bin_fail_rows."""
    _CAT_PALETTE = [
        '#dbeeff','#e0f5e0','#fef3cd','#fde0d0','#ece0f8','#d0f4f4',
        '#fce4ec','#e8f5e9','#fff3e0','#e3f2fd','#f3e5f5','#e8eaf6',
    ]

    # Prefer new 6-col format
    use_summary = any(r.get('bin_data') and r['bin_data'].get('bin_summary_rows')
                      for r in runs_data)
    if use_summary:
        valid = [r for r in runs_data if r.get('bin_data') and r['bin_data'].get('bin_summary_rows')]
        if not valid:
            return ''
        best = max(valid, key=lambda r: len(r['bin_data']['bin_summary_rows']))
        all_rows = list({row['ibin']: row for row in best['bin_data']['bin_summary_rows']}.values())

        # Build category→color map in row order
        _cat_color = {}
        for row in all_rows:
            ck = row.get('cat', '').strip().lower()
            if ck and ck not in _cat_color:
                _cat_color[ck] = _CAT_PALETTE[len(_cat_color) % len(_CAT_PALETTE)]

        hdr = '<th>Bin</th><th>Category</th><th>Description</th>'
        for ri, r in enumerate(valid):
            clr = _ID_COLORS[ri % len(_ID_COLORS)]
            hdr += (f'<th style="background:{clr};color:#fff;font-weight:bold;padding:5px 8px">'
                    f'{_esc(r["name"])}<br><span style="font-size:11px;font-weight:normal">Yield/Fail%</span></th>')

        rows_html = ''
        for row in all_rows:
            key  = row['ibin']
            # Skip bins 1, 2, 3, 4
            if str(key).strip() in ('1', '2', '3', '4'):
                continue
            cat  = row.get('cat', '')
            desc = row.get('desc', '')
            row_bg = _cat_color.get(cat.strip().lower(), '#ffffff')
            cells = (f'<td style="background:{row_bg}">{_esc(key)}</td>'
                     f'<td style="background:{row_bg}">{_esc(cat)}</td>'
                     f'<td style="background:{row_bg}">{_esc(desc)}</td>')
            for r in valid:
                v = next((ro['fail_pct'] for ro in r['bin_data']['bin_summary_rows']
                          if ro['ibin'] == key), None)
                row_all = [next((ro['fail_pct'] for ro in rx['bin_data']['bin_summary_rows'] if ro['ibin'] == key), None) for rx in valid]
                cells += _cell_hl(v, row_all, extra_style=f'background:{row_bg};') + (f'{v:.2f}%' if v is not None else '\u2014') + '</td>'
            rows_html += f'<tr>{cells}</tr>\n'
    else:
        valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['bin_fail_rows']]
        if not valid:
            return ''
        best = max(valid, key=lambda r: len(r['bin_data']['bin_fail_rows']))
        all_rows = list({row['ibin']: row for row in best['bin_data']['bin_fail_rows']}.values())

        hdr = '<th>Interface Bin</th><th>Fail Bucket</th>'
        for ri, r in enumerate(valid):
            clr = _ID_COLORS[ri % len(_ID_COLORS)]
            hdr += f'<th style="background:{clr};color:#fff;font-weight:bold;padding:5px 8px">{_esc(r["name"])}</th>'

        rows_html = ''
        for row in all_rows:
            key   = row['ibin']
            # Skip bins 1, 2, 3, 4
            if str(key).strip() in ('1', '2', '3', '4'):
                continue
            cells = f'<td>{_esc(key)}</td><td>{_esc(row["fail_bucket"])}</td>'
            for r in valid:
                v = next((ro['fail_pct'] for ro in r['bin_data']['bin_fail_rows']
                          if ro['ibin'] == key), None)
                row_all = [next((ro['fail_pct'] for ro in rx['bin_data']['bin_fail_rows'] if ro['ibin'] == key), None) for rx in valid]
                cells += _cell_hl(v, row_all) + (f'{v:.2f}%' if v is not None else '\u2014') + '</td>'
            rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#128196; Bin Fail Summary</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl" style="border-collapse:collapse">
    <thead><tr>{hdr}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 4d. SICC / UPM comparison table HTML
# ---------------------------------------------------------------------------

def build_sicc_table_html(runs_data):
    """Table: one row per test, columns = SICC Actual / SICC Target / Multiple / UPM% per run."""
    valid = [r for r in runs_data if r.get('upm_data') and r['upm_data']]
    if not valid:
        return ''

    # Union of test names, ordered by the run with the most rows
    best  = max(valid, key=lambda r: len(r['upm_data']))
    tests = [row['test'] for row in best['upm_data']]

    def _get(run, test, field):
        row = next((r for r in run['upm_data'] if r['test'] == test), None)
        return row[field] if row else None

    def _fmt(v, decimals=3):
        return f'{v:.{decimals}f}' if v is not None else '—'

    # Build column headers — one group (Actual / Target / Multiple) per run
    hdr = '<th rowspan="2" style="min-width:160px">Test</th>'
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th colspan="3" style="background:{clr};color:#fff;font-weight:bold;'
                f'text-align:left;padding:6px 10px">{_esc(r["name"])}</th>')
    # UPM % — one column per run
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th style="background:{clr};color:#fff;font-weight:bold;text-align:left;'
                f'padding:6px 10px">UPM%<br><span style="font-size:11px">{_esc(r["name"])}</span></th>')

    _th_sub = 'style="font-size:11px;white-space:nowrap;text-align:left!important;padding:4px 10px"'
    hdr2 = ''
    for _ in valid:
        hdr2 += (f'<th {_th_sub}>Actual (A)</th>'
                 f'<th {_th_sub}>Target (A)</th>'
                 f'<th {_th_sub}>Multiple</th>')
    for _ in valid:
        hdr2 += f'<th {_th_sub}>UPM (%)</th>'

    rows_html = ''
    for i, test in enumerate(tests):
        bg = '#f9f9f9' if i % 2 == 0 else '#ffffff'
        cells = f'<td style="white-space:nowrap;font-size:13px;background:{bg}">{_esc(test)}</td>'

        # Collect actuals for highlight comparison
        actuals = [_get(r, test, 'sicc_actual') for r in valid]
        for r in valid:
            act  = _get(r, test, 'sicc_actual')
            tgt  = _get(r, test, 'sicc_target')
            mult = _get(r, test, 'multiple')
            # Red only when actual > target (over spec)
            if act is not None and tgt is not None and act > tgt:
                act_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                act_style = f'background:{bg}'
            # Multiple: red bold only when > 1
            if mult is not None and mult > 1:
                mult_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                mult_style = f'background:{bg}'
            _td = 'font-variant-numeric:tabular-nums;padding:4px 10px'
            cells += (f'<td style="{act_style};{_td}">{_fmt(act)}</td>'
                      f'<td style="background:{bg};{_td}">{_fmt(tgt)}</td>'
                      f'<td style="{mult_style};{_td}">{_fmt(mult, 2)}</td>')

        for r in valid:
            upm = _get(r, test, 'upm_pct')
            cells += f'<td style="background:{bg};font-variant-numeric:tabular-nums;padding:4px 10px">{_fmt(upm, 1)}</td>'

        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#9889; SICC / UPM Table</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead>
      <tr>{hdr}</tr>
      <tr>{hdr2}</tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 4e. CDYN comparison chart + table
# ---------------------------------------------------------------------------

def build_cdyn_median_chart(runs_data):
    """Grouped bar chart: CDYN Actual Median vs Expected per test, per run."""
    if not HAVE_MPL:
        return ''
    valid = [r for r in runs_data if r.get('cdyn_data') and r['cdyn_data']]
    if not valid:
        return ''

    best = max(valid, key=lambda r: len(r['cdyn_data']))
    tests = [row for row in best['cdyn_data'] if row.get('actual') is not None]
    if not tests:
        return ''
    test_labels = [row['test'] for row in tests]

    n_tests = len(test_labels)
    n_runs  = len(valid)
    x       = np.arange(n_tests)
    width   = 0.7 / max(n_runs + 1, 2)

    fig, ax = plt.subplots(figsize=(max(7, n_tests * 0.6), 4))

    expected = [t.get('expected') for t in tests]
    has_expected = any(v is not None for v in expected)

    for ri, run in enumerate(valid):
        actuals = []
        for t in tests:
            row = next((r for r in run['cdyn_data'] if r['test'] == t['test']), None)
            actuals.append(row['actual'] if row and row.get('actual') is not None else 0.0)
        offset = ri * width
        ax.bar(x + offset, actuals, width, label=run['name'],
               color=_ID_COLORS[ri % len(_ID_COLORS)], alpha=0.85)

    if has_expected:
        exp_vals = [v if v is not None else 0.0 for v in expected]
        offset = n_runs * width
        ax.bar(x + offset, exp_vals, width, label='Expected',
               color='#2c3e50', alpha=0.55, hatch='//')

    tick_offset = (n_runs * width) / 2
    ax.set_xticks(x + tick_offset)
    ax.set_xticklabels(test_labels, rotation=35, ha='right', fontsize=8)
    ax.set_ylabel('CDYN Actual Median (nF)')
    ax.set_title('CDYN Median', fontsize=13, weight='bold')
    ax.yaxis.set_major_locator(plt.MaxNLocator(nbins=12, integer=False))
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: '0' if v == 0 else f'{v:.2f}'))
    ax.legend(fontsize=8)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    return _fig_b64(fig, dpi=180)


def build_cdyn_table_html(runs_data):
    """Table: one row per CDYN test, columns = Actual / Expected / Actual/Expected per run."""
    valid = [r for r in runs_data if r.get('cdyn_data') and r['cdyn_data']]
    if not valid:
        return ''

    best  = max(valid, key=lambda r: len(r['cdyn_data']))
    tests = [row['test'] for row in best['cdyn_data']]

    def _get(run, test, field):
        row = next((r for r in run['cdyn_data'] if r['test'] == test), None)
        return row[field] if row else None

    def _fmt(v, decimals=2):
        return f'{v:.{decimals}f}' if v is not None else '\u2014'

    # Header row 1 — test name + one group per run
    hdr = '<th rowspan="2" style="min-width:200px">Test</th>'
    hdr += '<th rowspan="2" style="min-width:80px">Type</th>'
    for ri, r in enumerate(valid):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        hdr += (f'<th colspan="3" style="background:{clr};color:#fff;font-weight:bold;'
                f'text-align:left;padding:6px 10px">{_esc(r["name"])}</th>')

    # Header row 2 — sub-headers
    _th_sub = 'style="font-size:11px;white-space:nowrap;text-align:left!important;padding:4px 10px"'
    hdr2 = ''
    for _ in valid:
        hdr2 += (f'<th {_th_sub}>Actual (nF)</th>'
                 f'<th {_th_sub}>Expected (nF)</th>'
                 f'<th {_th_sub}>Actual/Expected</th>')

    rows_html = ''
    for i, test in enumerate(tests):
        bg = '#f9f9f9' if i % 2 == 0 else '#ffffff'
        test_type = _get(best, test, 'type') or ''
        cells = (f'<td style="white-space:nowrap;font-size:13px;background:{bg}">{_esc(test)}</td>'
                 f'<td style="font-size:12px;background:{bg}">{_esc(test_type)}</td>')

        for r in valid:
            act  = _get(r, test, 'actual')
            exp  = _get(r, test, 'expected')
            ratio = _get(r, test, 'ratio')
            # Red when actual > expected
            if act is not None and exp is not None and act > exp:
                act_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                act_style = f'background:{bg}'
            # Ratio red when > 1
            if ratio is not None and ratio > 1:
                ratio_style = f'background:{bg};color:#c0392b;font-weight:bold'
            else:
                ratio_style = f'background:{bg}'
            _td = 'font-variant-numeric:tabular-nums;padding:4px 10px'
            cells += (f'<td style="{act_style};{_td}">{_fmt(act)}</td>'
                      f'<td style="background:{bg};{_td}">{_fmt(exp)}</td>'
                      f'<td style="{ratio_style};{_td}">{_fmt(ratio)}</td>')

        rows_html += f'<tr>{cells}</tr>\n'

    return f'''<div class="section">
  <h2>&#9889; CDYN Median Table</h2>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead>
      <tr>{hdr}</tr>
      <tr>{hdr2}</tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 5. Build comparison table HTML
# ---------------------------------------------------------------------------

def build_comparison_table(runs_data, ref_name, col_idx=0):
    if not runs_data:
        return ''
    ref_run = next((r for r in runs_data if r['name'] == ref_name), runs_data[0])

    def _get_val(run, lbl):
        if not run['data']:
            return None
        for g in run['data']['groups']:
            for row in g['rows']:
                if row['label'] == lbl:
                    return row['vals'][col_idx] if col_idx < len(row['vals']) else None
            if all(row['vals'][col_idx] is None for row in g['rows']) is False:
                if g['sum_vals'] and col_idx < len(g['sum_vals']) and g['rows'] and g['rows'][0]['label'] == lbl:
                    pass
        return None

    def _get_sum(run, gi):
        if not run['data'] or gi >= len(run['data']['groups']):
            return None
        sv = run['data']['groups'][gi]['sum_vals']
        return sv[col_idx] if sv and col_idx < len(sv) else None

    def _get_total(run):
        if not run['data'] or not run['data']['totals']:
            return None
        v = run['data']['totals']['vals']
        return v[col_idx] if v and col_idx < len(v) else None

    best = max(runs_data, key=lambda r: len(r['data']['groups']) if r['data'] else 0)
    groups = best['data']['groups']

    # Header row
    run_names = [r['name'] for r in runs_data]
    hdr_cells = ('<th style="background:#1a252f;color:#ecf0f1;font-size:18px">'
                 'Sub Module</th>') + ''.join(
        f'<th style="background:{_ID_COLORS[i % len(_ID_COLORS)]};color:#fff;'
        f'font-size:18px;padding:7px 10px">'
        f'{_esc(n)}</th>'
        for i, n in enumerate(run_names)
    )
    # Add delta columns (vs ref)
    compare_names = [r['name'] for r in runs_data if r['name'] != ref_run['name']]
    for cn in compare_names:
        hdr_cells += (
            f'<th style="background:#4a235a;color:#e8daef;text-align:center;'
            f'font-size:18px;padding:7px 10px">'
            f'&#916; vs <b>{_esc(ref_run["name"])}</b>'
            f'<br><span style="font-size:14px;font-weight:normal">{_esc(cn)}</span></th>'
        )

    rows_html = ''
    for gi, group in enumerate(groups):
        grp_clr = _GROUP_COLORS[gi % len(_GROUP_COLORS)]
        for row in group['rows']:
            lbl = row['label']
            vals = [_get_val(r, lbl) for r in runs_data]
            ref_v = _get_val(ref_run, lbl)
            cells = f'<td style="background:{grp_clr};font-size:20px">{_esc(lbl)}</td>'
            for v in vals:
                cells += f'<td class="num" style="background:{grp_clr}">{f"{v:.1f}%" if v is not None else ""}</td>'
            for r in runs_data:
                if r['name'] == ref_run['name']:
                    continue
                cv = _get_val(r, lbl)
                if cv is not None and ref_v is not None:
                    delta = cv - ref_v
                    clr = '#c0392b' if delta > 0 else ('#27ae60' if delta < 0 else '#555')
                    cells += f'<td class="num delta" style="color:{clr};font-weight:bold">{delta:+.1f}%</td>'
                else:
                    cells += '<td class="num delta">—</td>'
            rows_html += f'<tr>{cells}</tr>\n'

        # SUM row
        sum_vals = [_get_sum(r, gi) for r in runs_data]
        ref_sv = _get_sum(ref_run, gi)
        sum_cells = f'<td style="font-weight:bold;background:{grp_clr};border-top:2px solid #999">SUM</td>'
        for sv in sum_vals:
            sum_cells += (f'<td class="num" style="font-weight:bold;background:{grp_clr};'
                          f'border-top:2px solid #999">{f"{sv:.1f}%" if sv is not None else ""}</td>')
        for r in runs_data:
            if r['name'] == ref_run['name']:
                continue
            sv = _get_sum(r, gi)
            if sv is not None and ref_sv is not None:
                delta = sv - ref_sv
                clr = '#c0392b' if delta > 0 else ('#27ae60' if delta < 0 else '#555')
                sum_cells += (f'<td class="num delta" style="color:{clr};font-weight:bold;'
                              f'border-top:2px solid #999">{delta:+.1f}%</td>')
            else:
                sum_cells += '<td class="num delta" style="border-top:2px solid #999">—</td>'
        rows_html += f'<tr>{sum_cells}</tr>\n'

    # Total yield loss row
    if best['data'] and best['data']['totals']:
        tot_lbl = best['data']['totals']['label']
        tot_vals = [_get_total(r) for r in runs_data]
        ref_tv = _get_total(ref_run)
        tot_cells = f'<td style="font-weight:bold;background:#d5e8d4;border-top:3px solid #555">{_esc(tot_lbl)}</td>'
        for tv in tot_vals:
            tot_cells += (f'<td class="num" style="font-weight:bold;background:#d5e8d4;'
                          f'border-top:3px solid #555">{f"{tv:.1f}%" if tv is not None else ""}</td>')
        for r in runs_data:
            if r['name'] == ref_run['name']:
                continue
            tv = _get_total(r)
            if tv is not None and ref_tv is not None:
                delta = tv - ref_tv
                clr = '#c0392b' if delta > 0 else ('#27ae60' if delta < 0 else '#555')
                tot_cells += (f'<td class="num delta" style="color:{clr};font-weight:bold;'
                              f'border-top:3px solid #555">{delta:+.1f}%</td>')
            else:
                tot_cells += '<td class="num delta" style="border-top:3px solid #555">—</td>'
        rows_html += f'<tr>{tot_cells}</tr>\n'

    return f'''
<div class="section">
  <div style="background:#1a252f;border-radius:6px 6px 0 0;padding:10px 16px;
              display:flex;align-items:center;flex-wrap:wrap;gap:8px">
    <span style="font-size:20px;font-weight:bold;color:#ecf0f1">
      &#128202; Detailed Comparison Table
    </span>
    <span style="font-size:15px;color:#aab7b8;margin-left:4px">
      reference:&nbsp;<b style="color:#3498db">{_esc(ref_run["name"])}</b>
      &nbsp;&nbsp;&#124;&nbsp;&nbsp;
      <span style="color:#e74c3c">&#9650; red = more fallout</span>
      &nbsp;&nbsp;
      <span style="color:#2ecc71">&#9660; green = less</span>
    </span>
  </div>
  <div style="overflow-x:auto">
  <table class="cmp-tbl">
    <thead><tr>{hdr_cells}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  </div>
</div>'''


# ---------------------------------------------------------------------------
# 6. Generate full HTML report
# ---------------------------------------------------------------------------

def _run_sort_key(run):
    """Sort key: file mtime of xlsx (most reliable), then ts string, then name."""
    xlsx = run.get('xlsx_path')
    if xlsx:
        try:
            return Path(xlsx).stat().st_mtime
        except Exception:
            pass
    ts = run.get('ts', '')
    # Try common date patterns: YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY, etc.
    import datetime
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y',
                '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %H:%M', '%m/%d/%Y'):
        try:
            return datetime.datetime.strptime(ts.strip(), fmt).timestamp()
        except Exception:
            pass
    return 0.0


def generate_report(runs_data, output_path: Path, ref_name: str = None, config_json: str = None, dash_dir: Path = None):
    # Dashboard.html prepends newest runs first; reverse so latest appears on the right of all charts.
    runs_data = list(reversed(runs_data))

    valid = [r for r in runs_data if r['data']]
    if not valid:
        print('ERROR: No valid xlsx data found for any run.', file=sys.stderr)
        return

    # Load UPM target from config JSON
    _upm_target_pct = None
    _upm_target_label = 'Target'
    if config_json:
        try:
            import json as _json_cr
            _cfg_cr = _json_cr.loads(open(str(config_json), encoding='utf-8').read())
            for _ut in _cfg_cr.get('upm_target', []):
                if _ut.get('target_%') is not None:
                    _upm_target_pct = float(_ut['target_%'])
                    _upm_target_label = _ut.get('test', 'Target')
                    break
        except Exception:
            pass

    # Resolve the reference run name
    if ref_name is None:
        ref_name = valid[0]['name']
    elif not any(r['name'] == ref_name for r in runs_data):
        print(f'WARNING: --ref "{ref_name}" not found; using first run.')
        ref_name = valid[0]['name']

    # Determine number of data columns (Yield Loss + optional Recovery)
    n_cols = max(len(r['data']['col_headers']) for r in valid)
    col_headers = valid[0]['data']['col_headers']

    charts_html = ''

    # --- Yield Information chart ---
    rdnd_valid = [r for r in runs_data if r.get('bin_data') and r['bin_data']['yield_rows']]

    # --- Extract UPM from raw CSV if Group_Medians.csv not available ---
    for r in runs_data:
        if not r.get('upm_data'):
            output_dir = None
            if r.get('xlsx_path') and str(r['xlsx_path']):
                output_dir = Path(str(r['xlsx_path'])).parent
            if not output_dir:
                href = r.get('index_href', '')
                if href:
                    import re as _re_tmp
                    href = _re_tmp.sub(r'^file:///', '', href).replace('/', os.sep)
                    idx = Path(href) if os.path.isabs(href) else (output_path.parent / href)
                    output_dir = idx.parent
            if output_dir:
                raw_csv = find_raw_csv(output_dir)
                if raw_csv:
                    upm_data, upm_detail = extract_upm_from_csv(raw_csv, config_json=config_json)
                    if upm_data:
                        r['upm_data'] = upm_data
                    if upm_detail:
                        r['upm_detail'] = upm_detail

    upm_valid  = [r for r in runs_data if r.get('upm_data') and r['upm_data']]

    # --- Run Summary table (always shown) ---
    charts_html += build_run_summary_table_html(runs_data, dash_dir=dash_dir)

    if rdnd_valid:
        combined_b64 = build_combined_rdnd_chart(rdnd_valid)
        if combined_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; Yield Information</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + combined_b64 + '"/></div>')

    # --- Yield Table ---
    if rdnd_valid:
        charts_html += build_rdnd_table_html(rdnd_valid)

    # --- Top-10 fail pareto ---
    top10_b64 = build_top10_pareto_chart(runs_data)
    if top10_b64:
        charts_html += ('<div class="section">'
                        '<h2>&#128202; Top 10 Interface Bin Fail Pareto</h2>'
                        '<img class="chart" src="data:image/png;base64,'
                        + top10_b64 + '"/></div>')

    # --- Bin Fail Summary ---
    bf_valid = [r for r in runs_data if r.get('bin_data') and (
        r['bin_data'].get('bin_summary_rows') or r['bin_data'].get('bin_fail_rows'))]
    if bf_valid:
        charts_html += build_bin_fail_table_html(bf_valid)

    # --- SICC/UPM charts + table ---
    if upm_valid:
        upm_b64 = build_upm_median_chart(upm_valid)
        if upm_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; SICC Median</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + upm_b64 + '"/></div>')
        pct_b64 = build_upm_pct_chart(upm_valid, upm_target_pct=_upm_target_pct,
                                         upm_target_label=_upm_target_label)
        if pct_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; UPM ULVT 950mV (%)</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + pct_b64 + '"/></div>')
        # Detailed per-column UPM comparison
        detail_b64 = build_upm_detail_chart(runs_data, upm_target_pct=_upm_target_pct,
                                            upm_target_label=_upm_target_label)
        if detail_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128200; UPM Distribution Comparison</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + detail_b64 + '"/></div>')
        sicc_tbl = build_sicc_table_html(upm_valid)
        if sicc_tbl:
            charts_html += sicc_tbl

    # --- CDYN charts + table ---
    cdyn_valid = [r for r in runs_data if r.get('cdyn_data') and r['cdyn_data']]
    if cdyn_valid:
        cdyn_b64 = build_cdyn_median_chart(cdyn_valid)
        if cdyn_b64:
            charts_html += ('<div class="section">'
                            '<h2>&#128202; CDYN Median</h2>'
                            '<img class="chart" src="data:image/png;base64,'
                            + cdyn_b64 + '"/></div>')
        cdyn_tbl = build_cdyn_table_html(cdyn_valid)
        if cdyn_tbl:
            charts_html += cdyn_tbl

    # Run summary cards — show all runs, including those with no xlsx
    card_html = ''
    for ri, run in enumerate(runs_data):
        clr = _ID_COLORS[ri % len(_ID_COLORS)]
        if run.get('data'):
            total = run['data']['totals']['vals'][0] if run['data']['totals'] else None
            total_str = f'{total:.1f}%' if total is not None else '—'
            num_die = run['data']['num_die']
            die_str = f'{int(num_die):,}'
            # FF+DF yield (bin 1/2/3/4) and FF yield (bin 1/2)
            yrows = run['bin_data']['yield_rows'] if run.get('bin_data') else []
            ffdf_row = next((r for r in yrows if r['bin'] == '1/2/3/4'), None)
            ff_row   = next((r for r in yrows if r['bin'] == '1/2'), None)
            ffdf_str = f"{ffdf_row['yield_pct']:.1f}%" if ffdf_row and ffdf_row.get('yield_pct') is not None else '—'
            ff_str   = f"{ff_row['yield_pct']:.1f}%"  if ff_row  and ff_row.get('yield_pct')  is not None else '—'
        else:
            total_str = 'N/A'
            die_str   = 'N/A'
            ffdf_str  = 'N/A'
            ff_str    = 'N/A'
        card_html += f'''
<div class="run-card" style="border-left:4px solid {clr}">
  <div class="run-card-name" style="color:{clr}">{_esc(run["name"])}</div>
  <div class="run-card-ts">{_esc(run["ts"])}</div>
  <div class="run-card-stat">Die: <b>{die_str}</b></div>
  <div class="run-card-stat">FF + DF Yield: <b>{ffdf_str}</b></div>
  <div class="run-card-stat">FF Yield: <b>{ff_str}</b></div>
</div>'''

    mapping_table_html = _build_mapping_table_html(runs_data)

    html = f'''<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>Test Program Comparison Report</title>
<style>
.dash-link{{font-size:21px;color:#2980b9;margin-bottom:4px;display:block}}
.dash-link a{{color:#2980b9;text-decoration:none}}
.dash-link a:hover{{text-decoration:underline}}
h1{{font-size:33px;color:#2c3e50;margin-bottom:6px}}
h2{{font-size:26px;color:#2c3e50;margin:20px 0 8px;padding-bottom:4px;border-bottom:2px solid #dce1e7}}
h3{{font-size:23px;color:#555;margin:10px 0 4px}}
.subtitle{{font-size:21px;color:#7f8c8d;margin-bottom:18px}}
.ref-note{{font-size:20px;font-weight:normal;color:#7f8c8d;margin-left:8px}}
.section{{background:#fff;border-radius:8px;padding:18px;margin-bottom:20px;
  box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.cards{{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:20px}}
.run-card{{background:#fff;border-radius:6px;padding:12px 16px;min-width:200px;
  box-shadow:0 1px 4px rgba(0,0,0,.08);flex:1}}
.run-card-name{{font-size:23px;font-weight:bold;margin-bottom:4px}}
.run-card-ts{{font-size:18px;color:#95a5a6;margin-bottom:6px}}
.run-card-stat{{font-size:21px;margin-bottom:2px}}
.run-card-src{{font-size:17px;color:#aaa;margin-top:6px;word-break:break-all}}
.chart{{max-width:100%;height:auto;display:block;margin:8px 0}}
.cmp-tbl{{border-collapse:collapse;font-size:20px;width:auto}}
.cmp-tbl th{{background:#34495e;color:#ecf0f1;padding:6px 10px;text-align:left;
  white-space:nowrap;font-size:18px}}
.cmp-tbl thead tr:nth-child(2) th{{text-align:right;font-size:13px;font-weight:normal;
  background:#4a5568;color:#e2e8f0;padding:3px 10px}}
.cmp-tbl td{{padding:4px 10px;border-bottom:1px solid #eee;white-space:nowrap}}
.cmp-tbl tr:hover td{{background:#f9f9f9}}
td.num{{text-align:right;font-variant-numeric:tabular-nums}}
td.delta{{text-align:right;font-variant-numeric:tabular-nums}}
th.delta-hdr{{background:#4a235a;color:#e8daef;text-align:center}}
</style>
</head>
<body>
<h1>&#128200; Test Program Comparison Report</h1>
<div class="dash-link">Source: <a href="Dashboard.html">Dashboard.html</a></div>
<div class="subtitle">
  Runs: <b>{len(runs_data)}</b> &nbsp;|&nbsp; With data: <b>{len(valid)}</b>
  &nbsp;|&nbsp; Reference: <b>{_esc(ref_name)}</b>
</div>
<div class="cards">{card_html}</div>
{mapping_table_html}
{charts_html}
</body>
</html>'''

    if output_path:
        output_path.write_text(_wm_inject(html), encoding='utf-8')
        print(f'Wrote comparison report: {output_path}')
    return html


def _build_mapping_table_html(runs_data) -> str:
    """Build a label-to-full-name mapping table HTML section."""
    has_short = any(r.get('full_name') for r in runs_data)
    if not has_short:
        return ''
    rows = ''
    for r in runs_data:
        full = r.get('full_name', r['name'])
        if full == r['name']:
            continue  # no shortening happened
        rows += (f'<tr><td style="font-weight:bold;color:#1a3a5c;padding:4px 12px">{_esc(r["name"])}</td>'
                 f'<td style="font-family:monospace;padding:4px 12px">{_esc(full)}</td>'
                 f'<td style="color:#666;padding:4px 12px">{_esc(r.get("ts",""))}</td></tr>\n')
    if not rows:
        return ''
    return (f'<div class="section">'
            f'<h2>&#128198; Label Mapping</h2>'
            f'<table style="border-collapse:collapse;font-size:0.95em">'
            f'<thead><tr>'
            f'<th style="background:#263950;color:#4fc3f7;padding:5px 12px;text-align:left">Label</th>'
            f'<th style="background:#263950;color:#4fc3f7;padding:5px 12px;text-align:left">Full TP Key</th>'
            f'<th style="background:#263950;color:#4fc3f7;padding:5px 12px;text-align:left">Date</th>'
            f'</tr></thead>'
            f'<tbody>{rows}</tbody>'
            f'</table></div>')


# ---------------------------------------------------------------------------
# 7. Scan for compare HTML files and update links block in Dashboard.html
# ---------------------------------------------------------------------------

_CMP_PATTERNS = ('compare_*.html', 'compareTP*.html', '*_compare.html',
                 '*_comparison.html', 'compare-*.html')


def find_compare_files(dash_dir: Path) -> list[Path]:
    """Return all compare-report HTML files next to Dashboard.html, sorted by name.
    Case-insensitive: scans all .html files and matches any whose stem contains 'compare'."""
    name_re = re.compile(r'compare', re.IGNORECASE)
    results = []
    for p in sorted(dash_dir.glob('*.html')):
        if p.name.lower() == 'dashboard.html':
            continue
        if name_re.search(p.stem):
            results.append(p)
    return results


def update_dashboard_compare_links(dash_path: Path) -> None:
    """
    Scan the dashboard directory for compare HTML files and rewrite the
    <!-- COMPARISON_REPORT_START/END --> block in Dashboard.html with links.
    Safe to call after any compareTP / compare_runs run.
    """
    dash_dir = dash_path.parent
    files = find_compare_files(dash_dir)

    if not files:
        # Nothing to show — remove the block if present
        dash_html = dash_path.read_text(encoding='utf-8')
        dash_html = re.sub(
            r'\n*<!-- COMPARISON_REPORT_START -->[\s\S]*?<!-- COMPARISON_REPORT_END -->\n*',
            '', dash_html
        )
        dash_path.write_text(dash_html, encoding='utf-8')
        print('No compare files found — removed comparison block from Dashboard.html.')
        return

    link_items = ''.join(
        f'<a class="run-link report-link" href="{p.name}" target="_blank">'
        f'&#128200; {p.stem}</a>\n'
        for p in files
    )

    inject = (
        '\n\n<!-- COMPARISON_REPORT_START -->\n'
        '<div id="_cmp_embed" style="margin-top:18px;padding:10px 14px;background:#2c3e50;'
        'border-radius:6px;display:flex;flex-wrap:wrap;gap:6px;align-items:center">\n'
        '<span style="color:#ecf0f1;font-size:13px;font-weight:bold;margin-right:4px">'
        '&#128200; Comparison Reports:</span>\n'
        + link_items +
        '</div>\n'
        '<!-- COMPARISON_REPORT_END -->\n'
    )

    dash_html = dash_path.read_text(encoding='utf-8')

    # Remove any previous comparison block
    dash_html = re.sub(
        r'\n*<!-- COMPARISON_REPORT_START -->[\s\S]*?<!-- COMPARISON_REPORT_END -->\n*',
        '', dash_html
    )

    # Add ▼ Compare jump link to <h1> once
    jump = (' <a href="#_cmp_embed" style="font-size:15px;color:#2980b9;'
            'text-decoration:none;margin-left:14px;vertical-align:middle">'
            '&#9660; Compare</a>')
    if '#_cmp_embed' not in dash_html:
        dash_html = re.sub(
            r'(<h1[^>]*>)(.*?)(</h1>)',
            lambda m: m.group(1) + m.group(2) + jump + m.group(3),
            dash_html, count=1, flags=re.DOTALL
        )

    # Insert just before <!-- RUNS_END -->
    if '<!-- RUNS_END -->' in dash_html:
        dash_html = dash_html.replace('<!-- RUNS_END -->', inject + '<!-- RUNS_END -->', 1)
    elif '</body>' in dash_html:
        dash_html = dash_html.replace('</body>', inject + '</body>', 1)
    else:
        dash_html += inject

    dash_path.write_text(dash_html, encoding='utf-8')
    print(f'Updated {dash_path.name} with {len(files)} comparison link(s).')


# ---------------------------------------------------------------------------
# 8. Main
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(description='Compare yield runs from Dashboard.html')
    p.add_argument('dashboard', help='Path to Dashboard.html')
    p.add_argument('--out', default='', help='Output HTML path (default: next to Dashboard.html)')
    p.add_argument('--ref', default='', help='Reference run name for comparison (default: first run)')
    p.add_argument('--no-open', action='store_true', help='Do not open the report in the browser after building')
    args = p.parse_args()

    if not HAVE_OPENPYXL:
        print('ERROR: openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
        sys.exit(1)
    if not HAVE_MPL:
        print('WARNING: matplotlib not installed — charts will be skipped.')

    dash_path = Path(args.dashboard).resolve()
    if not dash_path.exists():
        print(f'ERROR: Dashboard.html not found: {dash_path}', file=sys.stderr)
        sys.exit(1)

    dash_dir = dash_path.parent

    print(f'Parsing {dash_path} ...')
    run_records = parse_dashboard(dash_path)
    if not run_records:
        print('ERROR: No run blocks found in Dashboard.html', file=sys.stderr)
        sys.exit(1)

    print(f'Found {len(run_records)} run(s): {[r["name"] for r in run_records]}')

    # Load xlsx + BinDistribution HTML for each run
    runs_data = []
    for rec in run_records:
        xlsx_p = find_xlsx(dash_dir, rec['index_href'])
        data = None
        if xlsx_p:
            print(f'  [{rec["name"]}] Reading {xlsx_p.name} ...')
            data = read_xlsx(xlsx_p)
            if not data:
                print(f'    WARNING: Could not parse {xlsx_p}')
            output_dir = xlsx_p.parent.parent if xlsx_p.parent.name == 'bin_dist' else xlsx_p.parent
        else:
            print(f'  [{rec["name"]}] WARNING: *_out.xlsx not found')
            # Try to resolve output_dir from index_href even without xlsx
            href = re.sub(r'^file:///', '', rec['index_href'] or '').replace('/', os.sep)
            idx_path = dash_dir / href if not os.path.isabs(href) else Path(href)
            output_dir = idx_path.parent if idx_path else None

        bin_data = None
        upm_data = None
        cdyn_data = None
        if output_dir and output_dir.exists():
            bin_p = find_bin_html(output_dir)
            if bin_p:
                print(f'  [{rec["name"]}] Reading {bin_p.name} ...')
                bin_data = parse_bin_html(bin_p)
            else:
                print(f'  [{rec["name"]}] WARNING: *_BinDistribution.html not found')
            gm_p = find_group_medians(output_dir)
            if gm_p:
                print(f'  [{rec["name"]}] Reading {gm_p.name} ...')
                upm_data = parse_group_medians(gm_p)
            cdyn_p = find_cdyn_medians(output_dir)
            if cdyn_p:
                print(f'  [{rec["name"]}] Reading {cdyn_p.name} ...')
                cdyn_data = parse_cdyn_medians(cdyn_p)

        runs_data.append({**rec, 'data': data, 'xlsx_path': xlsx_p or '',
                          'bin_data': bin_data, 'upm_data': upm_data,
                          'cdyn_data': cdyn_data})

    ref_name = args.ref.strip() or None
    out_path = Path(args.out).resolve() if args.out else dash_dir / 'compare_report.html'

    # Derive short display labels (e.g. "61C · 05-19") from full TP key + timestamp
    for r in runs_data:
        _m_l = re.search(r'0H61([A-Za-z])', r.get('name', ''))
        _letter = _m_l.group(1).upper() if _m_l else '?'
        _ts_m = re.search(r'(\d{4})-(\d{2})-(\d{2})', r.get('ts', ''))
        _date_s = f"{_ts_m.group(2)}-{_ts_m.group(3)}" if _ts_m else ''
        r['full_name'] = r['name']
        r['name'] = f"61{_letter}" + (f" · {_date_s}" if _date_s else '')

    print('Generating comparison report ...')
    # Find Product Config JSON in collateral/ folder
    _cfg_json = None
    _collateral = dash_dir / 'collateral'
    if _collateral.exists():
        _cfgs = sorted(_collateral.glob('Product Config*.json'),
                       key=lambda p: p.stat().st_mtime, reverse=True)
        if _cfgs:
            _cfg_json = str(_cfgs[0])
            print(f'  Config: {_cfgs[0].name}')
    generate_report(runs_data, out_path, ref_name=ref_name, config_json=_cfg_json,
                    dash_dir=dash_dir)

    # Update comparison links in Dashboard.html
    update_dashboard_compare_links(dash_path, out_path)

    # Open report in browser (skip when --no-open is passed, e.g. from automation)
    if not args.no_open:
        try:
            os.startfile(str(out_path))
        except Exception:
            pass


def update_dashboard_compare_links(dash_path: Path, compare_report_path: Path):
    """Inject a link to compare_report_path into the <!-- COMPARE_START/END --> section
    of Dashboard.html, using the same run-block style as the Yield section."""
    from datetime import datetime as _dt
    dash_path = Path(dash_path)
    compare_report_path = Path(compare_report_path)
    if not dash_path.exists():
        return

    content = dash_path.read_text(encoding='utf-8')

    COMPARE_START = '<!-- COMPARE_START -->'
    COMPARE_END   = '<!-- COMPARE_END -->'
    YIELD_END     = '<!-- YIELD_END -->'

    # Ensure COMPARE section exists
    if COMPARE_START not in content:
        if YIELD_END in content:
            content = content.replace(
                YIELD_END,
                YIELD_END + '\n<h2 class="section-header">&#128200; Compare TP</h2>\n'
                + COMPARE_START + '\n' + COMPARE_END)
        else:
            # No sentinels at all — append before </body>
            content = content.replace(
                '</body>',
                '<h2 class="section-header">&#128200; Compare TP</h2>\n'
                + COMPARE_START + '\n' + COMPARE_END + '\n</body>')

    # Build relative href from dash_dir to compare report
    try:
        href = os.path.relpath(str(compare_report_path), str(dash_path.parent)).replace('\\', '/')
    except Exception:
        href = compare_report_path.as_uri()

    report_stem = compare_report_path.stem
    ts = _dt.now().strftime('%Y-%m-%d %H:%M')

    # Replace existing block with same stem, or prepend new one
    block_key = report_stem
    new_block = (
        f'<div class="run-block" data-stem="{block_key}">\n'
        f'<div class="run-header" onclick="toggle(this)">'
        f'<span class="arrow">&#9660;</span> {report_stem}'
        f'<span class="ts"> - {ts}</span></div>\n'
        f'<div class="run-body">\n'
        f'<a class="run-link report-link" href="{href}" target="_blank">{report_stem}</a>\n'
        f'</div>\n</div>'
    )

    block_re = re.compile(
        r'<div class="run-block" data-stem="' + re.escape(block_key) +
        r'">\s*<div[^>]*>[\s\S]*?</div>\s*</div>', re.MULTILINE)
    if block_re.search(content):
        content = block_re.sub(new_block, content)
    else:
        content = content.replace(COMPARE_START, COMPARE_START + '\n' + new_block)

    dash_path.write_text(content, encoding='utf-8')


if __name__ == '__main__':
    main()


# ════════════════════════════════════════════════════════════════
# (formerly aqua_autopull.py)
# ════════════════════════════════════════════════════════════════

import argparse
import csv
import io
import json
import os
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────────────────
def _find_repo_root(start: Path) -> Path:
    """Walk up from `start` to the ancestor containing a shared/ dir (robust to repo reshuffles)."""
    current = Path(start).resolve()
    for _ in range(12):
        if (current / "shared").is_dir():
            return current
        parent = current.parent
        if parent == current:
            break
        current = parent
    return current


_HERE = Path(__file__).resolve().parent
_REPO_ROOT = _find_repo_root(_HERE)   # app.dashboard.nvl/

# ── CONFIG (override via CLI args) ─────────────────────────────────────────────
_AQUA_EXE_GAR   = r"\\PGSAPP3301.gar.corp.intel.com\Installer\AquaHbase\AquaCMDClient\Client\AquaCmdLine.exe"
_AQUA_EXE_AMR   = r"\\amr.corp.intel.com\ec\proj\fm\MPD\AQUA\AquaHbase\AquaCMDClient\Client\AquaCmdLine.exe"
_AQUA_EXE_GER   = r"\\HASAPP3301.ger.corp.intel.com\Installer\AquaHbase\AquaCMDClient\Client\AquaCmdLine.exe"
_DEFAULT_SERVER  = "AMR"
_DEFAULT_REPORT  = str(_REPO_ROOT / "shared" / "setup" / "aqua" / "NVL_Sort_Yield - Dashboard.txt")
_DEFAULT_OUT_DIR = r"C:\work\aqua_output"
_CACHE_FILE      = _HERE / "seen_lots.json"
_DEFAULT_OP      = "119325"
_DEFAULT_DRS     = "8PF5CV,8PF6CV"
_DEFAULT_DAYS    = 30

_EXE_MAP = {"GAR": _AQUA_EXE_GAR, "AMR": _AQUA_EXE_AMR, "GER": _AQUA_EXE_GER}


# ── Cache helpers ──────────────────────────────────────────────────────────────

def _load_cache() -> dict:
    if _CACHE_FILE.exists():
        try:
            return json.loads(_CACHE_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}   # {lot: {"pulled_at": ISO, "output": path}}


def _save_cache(cache: dict) -> None:
    _CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    _CACHE_FILE.write_text(json.dumps(cache, indent=2), encoding="utf-8")


# ── AQUA: discover lots via SUMMARY query ─────────────────────────────────────

def _discover_lots_aqua(
    aqua_exe: str,
    aqua_server: str,
    devrevsteps: list[str],
    operation: str,
    days: int,
    dry_run: bool = False,
) -> set[str]:
    """
    Run a quick AQUA FilterSet SUMMARY query to list lots for the given
    devrevsteps in the last N days.  Parses the output CSV for the 'Lot' column.
    Returns a set of lot strings.
    """
    drs_arg = ",".join(devrevsteps)
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, prefix="aqua_lot_list_") as tf:
        tmp_csv = tf.name

    cmd = [
        aqua_exe,
        "-AquaServer",       aqua_server,
        "-AnalysisType",     "SUMMARY",
        "-devrevsteps",      drs_arg,
        "-operations",       operation,
        "-lastNDaysLoadEnd", str(days),
        "-OutputFileName",   tmp_csv,
    ]

    print(f"[discover] {'DRY-RUN: ' if dry_run else ''}AQUA SUMMARY for devrevsteps={drs_arg}, op={operation}, last {days} days")
    print(f"           CMD: {' '.join(cmd)}")

    if dry_run:
        print("[discover] DRY-RUN: skipping AQUA call, returning empty set")
        try:
            Path(tmp_csv).unlink(missing_ok=True)
        except Exception:
            pass
        return set()

    lots: set[str] = set()
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
        if result.returncode != 0:
            print(f"[discover] AQUA SUMMARY error (rc={result.returncode}):\n{result.stderr.strip()}")
            return lots

        # Parse CSV — look for a 'Lot' column
        tmp_path = Path(tmp_csv)
        if not tmp_path.exists() or tmp_path.stat().st_size == 0:
            print("[discover] AQUA returned empty output")
            return lots

        with open(tmp_path, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            lot_col = next((h for h in headers if h.strip().lower() == "lot"), None)
            if not lot_col:
                print(f"[discover] No 'Lot' column found in SUMMARY output. Headers: {headers}")
                return lots
            for row in reader:
                lot = row[lot_col].strip()
                if lot:
                    lots.add(lot.upper())

        print(f"[discover] Found {len(lots)} lot(s): {sorted(lots)}")
    except subprocess.TimeoutExpired:
        print("[discover] TIMEOUT during SUMMARY query")
    except FileNotFoundError:
        print(f"[discover] ERROR: AquaCmdLine.exe not found at: {aqua_exe}")
    finally:
        try:
            Path(tmp_csv).unlink(missing_ok=True)
        except Exception:
            pass

    return lots


# ── AQUA: pull one lot ─────────────────────────────────────────────────────────

def _pull_lot(
    lot: str,
    aqua_exe: str,
    aqua_server: str,
    report_config: str,
    output_dir: str,
    dry_run: bool = False,
) -> bool:
    """
    Invoke AquaCmdLine.exe for a single lot.
    Returns True on success.
    """
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")
    out_file = str(out_dir / f"{lot}_{ts}.csv.gz")

    cmd = [
        aqua_exe,
        "-AquaServer",    aqua_server,
        "-ReportConfig",  report_config,
        "-lots",          lot,
        "-OutputFileName", out_file,
    ]

    print(f"[pull] {'DRY-RUN: ' if dry_run else ''}Pulling lot {lot} → {out_file}")
    print(f"       CMD: {' '.join(cmd)}")

    if dry_run:
        return True

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=3600)
        if result.returncode != 0:
            print(f"[pull] ERROR (rc={result.returncode}):\n{result.stderr.strip()}")
            return False
        print(f"[pull] Done: {out_file}")
        if result.stdout.strip():
            print(f"       {result.stdout.strip()}")
        return True
    except subprocess.TimeoutExpired:
        print(f"[pull] TIMEOUT for lot {lot}")
        return False
    except FileNotFoundError:
        print(f"[pull] ERROR: AquaCmdLine.exe not found at: {aqua_exe}")
        return False


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Auto-pull AQUA data for new lots detected via TRACE/XEUS.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    ap.add_argument("--devrevstep",   default=_DEFAULT_DRS,
                    help="Comma-separated devrevstep prefixes to watch (e.g. 8PF5CV,8PF6CV)")
    ap.add_argument("--operation",    default=_DEFAULT_OP,
                    help="SORT operation number")
    ap.add_argument("--days",         default=_DEFAULT_DAYS, type=int,
                    help="Look back N days in AQUA SUMMARY discovery")
    ap.add_argument("--lot",          nargs="+", default=None,
                    help="Pull specific lots now (skip AQUA discovery)")
    ap.add_argument("--aqua-server",  default=_DEFAULT_SERVER, choices=["GAR","AMR","GER"],
                    help="AQUA server domain")
    ap.add_argument("--aqua-exe",     default=None,
                    help="Path to AquaCmdLine.exe (auto-selected from --aqua-server if omitted)")
    ap.add_argument("--report-config",default=_DEFAULT_REPORT,
                    help="Path to exported AQUA report config .txt")
    ap.add_argument("--output-dir",   default=_DEFAULT_OUT_DIR,
                    help="Local folder for downloaded CSVs")
    ap.add_argument("--force",        action="store_true",
                    help="Re-pull lots even if already in cache")
    ap.add_argument("--dry-run",      action="store_true",
                    help="Show what would run without executing AQUA")
    ap.add_argument("--clear-cache",  action="store_true",
                    help="Clear the seen_lots cache and exit")
    args = ap.parse_args()

    # Handle clear-cache
    if args.clear_cache:
        if _CACHE_FILE.exists():
            _CACHE_FILE.unlink()
            print(f"[cache] Cleared: {_CACHE_FILE}")
        else:
            print(f"[cache] Nothing to clear ({_CACHE_FILE} does not exist)")
        return

    # Resolve exe
    aqua_exe = args.aqua_exe or _EXE_MAP[args.aqua_server]

    # Validate report config
    if not Path(args.report_config).exists():
        print(f"[config] ERROR: report config not found: {args.report_config}")
        sys.exit(1)

    # Determine lots to pull
    if args.lot:
        new_lots = set(l.upper() for l in args.lot)
        print(f"[main] Manual lot list: {sorted(new_lots)}")
    else:
        devrevsteps = [d.strip() for d in args.devrevstep.split(",") if d.strip()]
        discovered  = _discover_lots_aqua(
            aqua_exe=aqua_exe,
            aqua_server=args.aqua_server,
            devrevsteps=devrevsteps,
            operation=args.operation,
            days=args.days,
            dry_run=args.dry_run,
        )
        print(f"[main] AQUA discovered {len(discovered)} lot(s): {sorted(discovered)}")

        cache = _load_cache()
        if args.force:
            new_lots = discovered
        else:
            new_lots = discovered - set(cache.keys())

        print(f"[main] New lots (not yet pulled): {sorted(new_lots)}")

    if not new_lots:
        print("[main] Nothing to pull.")
        return

    # Pull each lot
    cache = _load_cache()
    for lot in sorted(new_lots):
        success = _pull_lot(
            lot=lot,
            aqua_exe=aqua_exe,
            aqua_server=args.aqua_server,
            report_config=args.report_config,
            output_dir=args.output_dir,
            dry_run=args.dry_run,
        )
        if success and not args.dry_run:
            cache[lot] = {
                "pulled_at": datetime.now(timezone.utc).isoformat(),
                "output_dir": args.output_dir,
            }
            _save_cache(cache)

    print("[main] Done.")


if __name__ == "__main__":
    main()


# ════════════════════════════════════════════════════════════════
# (formerly auto_pull_and_run.py)
# ════════════════════════════════════════════════════════════════

import argparse
import csv
import gzip
import hashlib
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from datetime import datetime, timezone
from pathlib import Path

# Ensure Unicode log output works on Windows cp1252 consoles and when
# stdout is redirected to a file (e.g. Tee-Object).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Paths ──────────────────────────────────────────────────────────────────────
_HERE      = Path(__file__).resolve().parent
_REPO_ROOT = _find_repo_root(_HERE)   # app.dashboard.nvl/

_PIPELINE_PY  = _HERE / "yield_pipeline.py"
_AQUA_EXE_GAR = r"\\PGSAPP3301.gar.corp.intel.com\Installer\AquaHbase\AquaCMDClient\Client\AquaCmdLine.exe"
_AQUA_EXE_AMR = r"\\FMSAPP3301.amr.corp.intel.com\Installer\AquaHbase\AquaCMDClient\Client\AquaCmdLine.exe"

# ── Defaults ───────────────────────────────────────────────────────────────────
_DEFAULT_SERVER      = "AMR"
_DEFAULT_DATA_DIR    = r"C:\work\auto\data"
_DEFAULT_DASH_DIR    = r"C:\work\auto\dashboard"
_DEFAULT_EMAIL       = "sujit.n.pant@intel.com"
_SNAPSHOT_FILE       = Path(r"C:\work\auto\snapshot.json")
_DEFAULT_REPORT      = str(_REPO_ROOT / "shared" / "setup" / "automation" / "yield-dashboard" / "NVL_Sort_Yield - AutoPull.txt")

# Change-detection: priority-ordered candidate column names (auto-selected from header)
_CHANGE_COL_CANDIDATES = [
    ("Lot",     ["SORT_LOT", "Lot", "Sort_Lot", "LOTFROMFS"]),
    ("Wafer",   ["SORT_WAFER", "Wafer", "Sort_Wafer_ID", "Wafer_ID", "WaferID"]),
    ("Program", ["Program Name", "Program_Name", "ProgramName"]),
    ("Date",    ["LOTS End Date Time", "End_Date_Time", "End_Date", "Start_Date_Time"]),
]

# TestProgram folder used by yield_pipeline to locate BinDefinitions.bdefs
_DEFAULT_TP_FOLDER = r"I:\program\1001\prod\hdmtprogs\nvl_ncx_sds"


# ── Helpers ────────────────────────────────────────────────────────────────────

def _ts() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")


def _log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def _load_snapshot() -> dict:
    if _SNAPSHOT_FILE.exists():
        try:
            return json.loads(_SNAPSHOT_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_snapshot(snap: dict) -> None:
    _SNAPSHOT_FILE.parent.mkdir(parents=True, exist_ok=True)
    _SNAPSHOT_FILE.write_text(json.dumps(snap, indent=2), encoding="utf-8")


# ── Step 1: Pull AQUA ─────────────────────────────────────────────────────────

def _get_aqua_report_name(config_path: str) -> str:
    """Read the '@ Report : <name>' line from an AQUA config file."""
    try:
        for line in Path(config_path).read_text(encoding="utf-8-sig", errors="replace").splitlines():
            stripped = line.strip()
            if stripped.startswith("@ Report :"):
                return stripped.split(":", 1)[1].strip()
    except Exception:
        pass
    return "NVL_Sort_Yield"  # fallback


def pull_aqua(
    aqua_exe: str,
    aqua_server: str,
    report_config: str,
    data_dir: Path,
    dry_run: bool,
) -> Path | None:
    """
    Pull via -ReportConfig — all filter settings (program, days) live in the .txt file.
    AQUA honours the OutputFileName stem but may append a different extension
    (e.g. .csv.gz instead of .zip).  We glob for any file sharing the stem.
    Falls back to scanning %TEMP% for new CSVs if nothing found in data_dir.
    Returns path to the output file, or None on failure / no data.
    """
    data_dir.mkdir(parents=True, exist_ok=True)
    ts       = _ts()
    # derive prefix from config filename (e.g. "NVLG_Sort_Yield - AutoPull.txt" -> "NVLG_Sort_Yield")
    _cfg_stem   = Path(report_config).stem  # e.g. "NVLG_Sort_Yield - AutoPull"
    _cfg_prefix = re.sub(r'\s*-\s*AutoPull.*$', '', _cfg_stem, flags=re.IGNORECASE).strip() or "AQUA"
    _cfg_prefix = re.sub(r'[^\w\-]', '_', _cfg_prefix)  # filesystem-safe
    out_base = data_dir / f"{_cfg_prefix}_{ts}"

    # AQUA ignores the extension we request and always writes .csv.gz;
    # request a .zip so we can detect the actual file by globbing out_base.*
    zip_file = out_base.with_suffix(".zip")

    # AQUA also writes to %TEMP% using internal report name — scan for fallback
    temp_dir    = Path(os.environ.get("TEMP", tempfile.gettempdir()))
    report_name = _get_aqua_report_name(report_config)
    temp_pat    = f"{report_name}*.CSV"

    cmd = [
        aqua_exe,
        "-AquaServer",    aqua_server,
        "-ReportConfig",  report_config,
        "-OutputFileName", str(zip_file),
    ]

    _log(f"{'DRY-RUN: ' if dry_run else ''}AQUA pull -> {out_base}.*")
    _log(f"  Config      : {report_config}")
    _log(f"  Report name : {report_name}")
    _log(f"  CMD         : {' '.join(cmd)}")

    if dry_run:
        return out_base.with_suffix(".csv.gz")

    # Snapshot %TEMP% before run (for fallback detection)
    before_temp = {p.resolve() for p in temp_dir.glob(temp_pat)}

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=3600)
        if result.stdout.strip():
            _log(f"AQUA stdout: {result.stdout.strip()[:500]}")
        if result.stderr.strip():
            _log(f"AQUA stderr: {result.stderr.strip()[:500]}")
        if result.returncode != 0:
            _log(f"ERROR: AQUA exited rc={result.returncode}")
            return None
    except subprocess.TimeoutExpired:
        _log("ERROR: AQUA timed out")
        return None
    except FileNotFoundError:
        _log(f"ERROR: AquaCmdLine.exe not found: {aqua_exe}")
        return None
    except Exception as e:
        _log(f"ERROR: {e}")
        return None

    # Primary: AQUA honours our OutputFileName stem but may change the extension
    # (e.g. we ask for .zip, it writes .csv.gz).  Glob for any file sharing the stem.
    written = [p for p in data_dir.glob(f"{out_base.name}*") if p.stat().st_size > 0]
    if written:
        out = max(written, key=lambda p: p.stat().st_mtime)
        _log(f"  Output: {out.name} ({out.stat().st_size:,} bytes)")
        return out

    # Fallback: check %TEMP% for new CSVs written during the run
    after_temp = {p.resolve() for p in temp_dir.glob(temp_pat)}
    new_csvs   = sorted(after_temp - before_temp, key=lambda p: p.stat().st_mtime)
    if new_csvs:
        plain = [p for p in new_csvs if p.suffix.lower() == ".csv"]
        src   = max(plain or new_csvs, key=lambda p: p.stat().st_mtime)
        dest  = data_dir / f"{_cfg_prefix}_{ts}.csv"
        shutil.copy2(src, dest)
        _log(f"  Fallback: copied from %TEMP%: {src.name} -> {dest.name} ({dest.stat().st_size:,} bytes)")
        return dest

    _log(f"ERROR: AQUA produced no output (rc={result.returncode}; nothing matching '{out_base.name}*' in {data_dir})")
    return None


# ── Step 2: Change detection ──────────────────────────────────────────────────

def _read_aqua_file(path: Path) -> list[dict]:
    """
    Read an AQUA output file (.csv.gz, .csv, or .zip) and return list of row dicts.
    AQUA typically writes gzip-compressed CSV; the inner file is tab-delimited text.
    """
    import zipfile
    try:
        raw = path.read_bytes()
        # ZIP (magic PK)
        if raw[:2] == b'PK':
            with zipfile.ZipFile(io.BytesIO(raw)) as z:
                inner = z.read(z.namelist()[0]).decode("utf-8-sig", errors="replace")
        # gzip
        elif raw[:2] == b'\x1f\x8b':
            inner = gzip.decompress(raw).decode("utf-8-sig", errors="replace")
        else:
            inner = raw.decode("utf-8-sig", errors="replace")

        # auto-detect delimiter
        first_line = inner.split("\n")[0]
        delim = "\t" if "\t" in first_line else ","
        return list(csv.DictReader(io.StringIO(inner), delimiter=delim))
    except Exception as e:
        _log(f"WARNING: could not read {path}: {e}")
        return []


def _fingerprint(rows: list[dict], cols: list[str]) -> str:
    """SHA256 of the sorted unique values of the key columns."""
    values = set()
    for row in rows:
        values.add(tuple(row.get(c, "").strip() for c in cols))
    digest = hashlib.sha256(
        "\n".join(sorted(str(v) for v in values)).encode()
    ).hexdigest()
    return digest


def has_changed(aqua_file: Path, snapshot: dict) -> bool:
    """
    Return True if the data in aqua_file differs from the last snapshot.
    Auto-detects lot/wafer/program/date columns from the file header.
    """
    rows = _read_aqua_file(aqua_file)
    if not rows:
        _log("WARNING: no rows in pull — treating as no change")
        return False

    headers = set(rows[0].keys())
    # Pick the first candidate that exists in the header for each role
    cols = []
    for _role, candidates in _CHANGE_COL_CANDIDATES:
        for c in candidates:
            if c in headers:
                cols.append(c)
                break

    if not cols:
        _log(f"WARNING: no key columns found in header {sorted(headers)[:8]}")
        _log("Treating as changed to be safe.")
        return True

    _log(f"Change-detection columns: {cols}")
    fp = _fingerprint(rows, cols)
    prev_fp = snapshot.get("fingerprint")
    _log(f"Fingerprint: {fp[:16]}...  previous: {(prev_fp or 'none')[:16]}...")

    if fp != prev_fp:
        prev_lots = set(snapshot.get("lots", []))
        lot_col   = next((c for c in cols if "lot" in c.lower()), None)
        curr_lots = {r.get(lot_col, "").strip() for r in rows if lot_col and r.get(lot_col)}
        new_lots  = curr_lots - prev_lots
        if new_lots:
            _log(f"New lots detected: {sorted(new_lots)}")
        return True

    _log("No change detected — skipping dashboard rebuild.")
    return False


# ── Step 3: Run yield dashboard ───────────────────────────────────────────────

_LOADER = _PIPELINE_PY

def _detect_test_program(aqua_csv: Path) -> str:
    """
    Return the most common TestProgram name from the AQUA CSV
    (reads 'Program Name_119325' or similar columns).
    """
    from collections import Counter
    rows = _read_aqua_file(aqua_csv)
    for col in ("Program Name_119325", "Program Name_132322", "Program Name"):
        vals = [r.get(col, "").strip() for r in rows if r.get(col, "").strip()]
        if vals:
            return Counter(vals).most_common(1)[0][0]
    return "NCXSDJXL0H61A"  # fallback

def run_dashboard(
    aqua_csv: Path,
    dash_dir: Path,
    dry_run: bool,
) -> bool:
    """
    Run the full yield pipeline headlessly against the extracted tab-delimited CSV.
    Pipes a minimal JSON config to yield_pipeline via stdin.
    Returns True on success.
    """
    dash_dir.mkdir(parents=True, exist_ok=True)
    dashboard_xlsx = dash_dir / "DigitalDashBoard.xlsx"

    test_program = _detect_test_program(aqua_csv) if not dry_run else "NCXSDJXL0H61A002618"
    tag      = datetime.now().strftime("%Y-%m-%d %H:%M")
    cfg_json = json.dumps({
        "outputFilename":   str(aqua_csv),
        "TestProgram":      test_program,
        "TestProgram_folder": _DEFAULT_TP_FOLDER,
        "output_folder":    str(dash_dir),
        "dashboard":        str(dashboard_xlsx),
        "identifier":       tag,
        "skip_aqua":        True,
    })

    cmd = [
        sys.executable,
        str(_LOADER),
        "yield_pipeline",
        "--input", "-",
        "--base",  str(aqua_csv.parent),
    ]

    _log(f"{'DRY-RUN: ' if dry_run else ''}Running yield pipeline")
    _log(f"  CSV         : {aqua_csv}")
    _log(f"  TestProgram : {test_program}")
    _log(f"  Out         : {dash_dir}")
    _log(f"  CMD         : {' '.join(cmd)}")

    if dry_run:
        return True

    try:
        result = subprocess.run(
            cmd,
            input=cfg_json,
            capture_output=True,
            text=True,
            timeout=1800,
        )
        if result.stdout.strip():
            _log(result.stdout.strip())
        if result.returncode != 0:
            _log(f"ERROR: yield_pipeline exited rc={result.returncode}\n{result.stderr.strip()}")
            return False
        _log(f"Pipeline output in: {dash_dir}")
        return True
    except subprocess.TimeoutExpired:
        _log("ERROR: yield_pipeline timed out")
        return False


# ── Step 4: Send email ────────────────────────────────────────────────────────

def _load_dashboard_html(dash_dir: Path) -> str | None:
    """Return the main Dashboard HTML from dash_dir, or None."""
    for name in ("Dashboard.html", "dashboard.html"):
        p = dash_dir / name
        if p.exists():
            return p.read_text(encoding="utf-8", errors="replace")
    htmls = [p for p in dash_dir.glob("*.html") if p.is_file()]
    if htmls:
        return max(htmls, key=lambda p: p.stat().st_mtime).read_text(encoding="utf-8", errors="replace")
    return None


def _send_via_outlook(to: str, subject: str, html_body: str) -> None:
    """Send HTML email via Outlook COM. Falls back to system python if win32com unavailable."""
    import tempfile
    try:
        import win32com.client
        ol   = win32com.client.Dispatch("Outlook.Application")
        mail = ol.CreateItem(0)
        mail.To       = to
        mail.Subject  = subject
        mail.HTMLBody = html_body
        mail.Send()
        return
    except ImportError:
        pass  # not in this env, try system python

    # Fallback: write HTML to a temp file so we avoid quoting issues in -c script.
    # Prefer the current project venv, then the Python installations available on this machine.
    candidates = [
        r"Y:\tools\scripts\.venv\Scripts\python.exe",
        r"C:\Program Files\Python314\python.exe",
        r"C:\Python313\python.exe",
        r"C:\Python312\python.exe",
    ]
    tmp_html = None
    try:
        with tempfile.NamedTemporaryFile(mode="w", suffix=".html", encoding="utf-8", delete=False) as f:
            f.write(html_body)
            tmp_html = f.name
        script = (
            "import win32com.client; "
            "ol=win32com.client.Dispatch('Outlook.Application'); "
            "m=ol.CreateItem(0); "
            f"m.To={to!r}; m.Subject={subject!r}; "
            f"m.HTMLBody=open({tmp_html!r},encoding='utf-8').read(); m.Send()"
        )
        for py in candidates:
            if os.path.exists(py):
                result = subprocess.run([py, "-c", script], capture_output=True, text=True, timeout=30)
                if result.returncode == 0:
                    return
    finally:
        if tmp_html and os.path.exists(tmp_html):
            os.unlink(tmp_html)
    raise RuntimeError("win32com not available in any known python installation")


def send_email(
    to: str,
    smtp_host: str,   # unused, kept for signature compat
    subject: str,
    html_body: str,
    dry_run: bool,
) -> None:
    _log(f"{'DRY-RUN: ' if dry_run else ''}Sending email to {to}")
    if dry_run:
        _log(f"  Subject: {subject}")
        return
    try:
        _send_via_outlook(to, subject, html_body)
        _log("Email sent.")
    except Exception as e:
        _log(f"WARNING: email failed: {e}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Auto AQUA pull + yield dashboard rerun on data change.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    ap.add_argument("--aqua-server",   default=_DEFAULT_SERVER, choices=["GAR", "AMR"])
    ap.add_argument("--aqua-exe",      default=None,
                    help="Path to AquaCmdLine.exe (auto-selected from --aqua-server)")
    ap.add_argument("--report-config", default=_DEFAULT_REPORT,
                    help="Path to AQUA .txt config file (program/days baked in)")
    ap.add_argument("--data-dir",      default=_DEFAULT_DATA_DIR,
                    help="Folder for downloaded AQUA ZIP files")
    ap.add_argument("--dashboard-dir", default=_DEFAULT_DASH_DIR,
                    help="Folder where pipeline output (xlsx, PNG) is written")
    ap.add_argument("--email",         default=_DEFAULT_EMAIL)
    ap.add_argument("--force",         action="store_true",
                    help="Skip change detection, always rerun dashboard")
    ap.add_argument("--dry-run",       action="store_true",
                    help="Show what would run without executing anything")
    args = ap.parse_args()

    aqua_exe  = args.aqua_exe or (_AQUA_EXE_GAR if args.aqua_server == "GAR" else _AQUA_EXE_AMR)
    data_dir  = Path(args.data_dir)
    dash_dir  = Path(args.dashboard_dir)
    snapshot  = _load_snapshot()

    _log("=" * 60)
    _log(f"auto_pull_and_run  [{'DRY-RUN' if args.dry_run else 'LIVE'}]")
    _log(f"Report config  : {args.report_config}")
    _log(f"Data dir       : {data_dir}")
    _log(f"Dashboard dir  : {dash_dir}")
    _log("=" * 60)

    status_lines: list[str] = []

    # ── 1. Pull ────────────────────────────────────────────────────────────────
    aqua_csv = pull_aqua(
        aqua_exe=aqua_exe,
        aqua_server=args.aqua_server,
        report_config=args.report_config,
        data_dir=data_dir,
        dry_run=args.dry_run,
    )
    if aqua_csv is None:
        _log("Pull failed — aborting.")
        send_email(
            to=args.email, smtp_host=None,
            subject="NVL816-BLLC Yield Dashboard — FAILED",
            html_body=(
                "<html><body style='font-family:Arial,sans-serif;padding:16px'>"
                "<p><b style='color:#c0392b'>AQUA pull failed.</b> Check server logs.</p>"
                f"<p style='color:#888;font-size:12px'>Run time: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>"
                "</body></html>"
            ),
            dry_run=args.dry_run,
        )
        sys.exit(1)

    status_lines.append(f"Pull  : {aqua_csv.name}")

    # ── 2. Change detection ────────────────────────────────────────────────────
    if args.force or args.dry_run:
        changed = True
        _log("Change detection skipped (--force or --dry-run).")
    else:
        changed = has_changed(aqua_csv, snapshot)

    if not changed:
        status_lines.append("Change: NONE — dashboard not rebuilt")
        send_email(
            to=args.email, smtp_host=None,
            subject="NVL816-BLLC Yield Dashboard — No new data",
            html_body=(
                "<html><body style='font-family:Arial,sans-serif;padding:16px'>"
                "<p>No new AQUA data detected &mdash; dashboard not rebuilt.</p>"
                f"<p style='color:#888;font-size:12px'>Pull: {aqua_csv.name}</p>"
                f"<p style='color:#888;font-size:12px'>Run time: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>"
                "</body></html>"
            ),
            dry_run=args.dry_run,
        )
        return

    status_lines.append("Change: YES — rebuilding dashboard")

    # ── 3. Run dashboard ─────────────────────────────────────────────────────────────────
    ok = run_dashboard(
        aqua_csv=aqua_csv,
        dash_dir=dash_dir,
        dry_run=args.dry_run,
    )
    status_lines.append(f"Dashboard: {'OK' if ok else 'FAILED'} → {dash_dir}")

    # ── 4. Update snapshot ─────────────────────────────────────────────────────
    if ok and not args.dry_run:
        rows    = _read_aqua_file(aqua_csv)
        headers = set(rows[0].keys()) if rows else set()
        cols    = []
        for _role, candidates in _CHANGE_COL_CANDIDATES:
            for c in candidates:
                if c in headers:
                    cols.append(c)
                    break
        fp       = _fingerprint(rows, cols) if rows and cols else ""
        lot_col  = next((c for c in cols if "lot" in c.lower()), None)
        lots     = sorted({r.get(lot_col, "").strip() for r in rows if lot_col and r.get(lot_col)})
        _save_snapshot({
            "fingerprint":  fp,
            "lots":         lots,
            "last_pull":    datetime.now(timezone.utc).isoformat(),
            "last_csv":     str(aqua_csv),
        })
        _log(f"Snapshot updated ({len(lots)} lots)")

    # ── 5. Email ───────────────────────────────────────────────────────────────
    subject = "NVL816-BLLC Yield Dashboard" if ok else "NVL816-BLLC Yield Dashboard — FAILED"
    _dash_html = _load_dashboard_html(dash_dir) if ok and not args.dry_run else None
    if _dash_html:
        html_body = _dash_html
    else:
        _status_html = "".join(f"<li>{ln}</li>" for ln in status_lines)
        _color = "#27ae60" if ok else "#c0392b"
        html_body = (
            "<html><body style='font-family:Arial,sans-serif;padding:16px'>"
            f"<p><b style='color:{_color}'>{'Dashboard updated.' if ok else 'Dashboard run FAILED.'}</b></p>"
            f"<ul style='font-size:13px'>{_status_html}</ul>"
            f"<p style='color:#888;font-size:12px'>Run time: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>"
            f"<p style='color:#888;font-size:12px'>Output folder: {dash_dir}</p>"
            "</body></html>"
        )

    send_email(
        to=args.email, smtp_host=None,
        subject=subject, html_body=html_body,
        dry_run=args.dry_run,
    )

    if not ok:
        sys.exit(1)
    _log("All done.")


if __name__ == "__main__":
    main()


# ════════════════════════════════════════════════════════════════
# (formerly generate_index.py)
# ════════════════════════════════════════════════════════════════
import argparse, datetime, os, re
from pathlib import Path

_UNC_REPORTS = r"\\samba.zsc10.intel.com\nfs\zsc10\disks\gsc_gwa011\users\snpant\auto\yield\reports"


def _fmt_size(n: int) -> str:
    if n < 1024:      return f"{n} B"
    if n < 1024**2:   return f"{n/1024:.0f} KB"
    return f"{n/1024**2:.1f} MB"


def build_index(base_dir: Path, product_name: str = "NVL816-BLLC") -> Path:
        """Index generation is intentionally disabled."""
        return base_dir / "reports" / "index.html"


# ════════════════════════════════════════════════════════════════
# (formerly manage_email.py)
# ════════════════════════════════════════════════════════════════

import argparse
import json
import re
import sys
import tkinter as tk
from pathlib import Path
from tkinter import font as tkfont
from tkinter import messagebox, ttk

# ── defaults (same as run_automation.py) ──────────────────────────────────────
_HERE        = Path(__file__).resolve().parent
_REPO_ROOT   = _find_repo_root(_HERE)   # app.dashboard.nvl/
_BASE_DIR    = Path(r"\\samba.zsc10.intel.com\nfs\zsc10\disks\gsc_gwa011\users\snpant\auto\yield")
_CFG_NAME    = "email_config.json"
_CFG_DIR     = _REPO_ROOT / "shared" / "setup" / "automation" / "yield-dashboard"
_EMAIL_TO    = "sujit.n.pant@intel.com"

# ── colours ───────────────────────────────────────────────────────────────────
BG          = "#1a252f"
BG2         = "#1e2e3d"
BG3         = "#263950"
FG          = "#e8f0f7"
FG_DIM      = "#90a4ae"
ACCENT      = "#4fc3f7"
GREEN       = "#66bb6a"
RED         = "#ef5350"
AMBER       = "#ffa726"
FONT_MONO   = ("Courier New", 9)
FONT_UI     = ("Segoe UI", 10)
FONT_TITLE  = ("Segoe UI", 13, "bold")
FONT_GROUP  = ("Segoe UI", 10, "bold")


def _load_config(cfg_path: Path) -> dict:
    if cfg_path.exists():
        try:
            d = json.loads(cfg_path.read_text(encoding="utf-8"))
            # migrate old single-field format
            if "email_to" in d and "email_to_report" not in d:
                d["email_to_report"] = d.pop("email_to")
            return d
        except Exception:
            pass
    return {"email_to_report": _EMAIL_TO, "email_to_alert": _EMAIL_TO, "excluded_keys": []}


def _save_config(cfg_path: Path, cfg: dict) -> None:
    cfg_path.parent.mkdir(parents=True, exist_ok=True)
    cfg_path.write_text(json.dumps(cfg, indent=2), encoding="utf-8")


def _discover_keys(base_dir: Path) -> list[str]:
    """Return all known TP keys from stored gzs."""
    prog_dir = base_dir / "data" / "programs"
    if not prog_dir.exists():
        return []
    keys = []
    for p in sorted(prog_dir.glob("*.csv.gz")):
        stem = p.name
        if stem.endswith(".csv.gz"):
            stem = stem[:-7]
        elif stem.endswith(".gz"):
            stem = stem[:-3]
        keys.append(stem)
    return sorted(keys)


def _group_keys(keys: list[str]) -> dict[str, list[str]]:
    """Group keys by full program variant (H61G, M61H …)."""
    groups: dict[str, list[str]] = {}
    for k in keys:
        m = re.search(r'([A-Za-z]\d{2}[A-Za-z])', k)
        group_key = m.group(1).upper() if m else "?"
        groups.setdefault(group_key, []).append(k)
    return dict(sorted(groups.items(), reverse=True))


# ─────────────────────────────────────────────────────────────────────────────
# Main GUI
# ─────────────────────────────────────────────────────────────────────────────

class EmailManagerApp(tk.Tk):
    def __init__(self, base_dir: Path) -> None:
        super().__init__()
        self.base_dir   = base_dir
        self.cfg_path   = _CFG_DIR / _CFG_NAME
        self.cfg        = _load_config(self.cfg_path)
        self.excluded   = set(self.cfg.get("excluded_keys", []))

        self.title("Yield Automation — Email Manager")
        self.configure(bg=BG)
        self.resizable(True, True)
        self.minsize(560, 420)

        self._build_ui()
        self._populate()

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        pad = dict(padx=14, pady=6)

        # Title
        tk.Label(self, text="Email Report Manager", font=FONT_TITLE,
                 bg=BG, fg=ACCENT).pack(anchor="w", padx=14, pady=(14, 2))
        tk.Label(self, text=f"Config: {self.cfg_path}", font=("Segoe UI", 8),
                 bg=BG, fg=FG_DIM).pack(anchor="w", padx=14, pady=(0, 8))

        # ── Email recipients ──────────────────────────────────────────────────
        frm_email = tk.LabelFrame(self, text="  Recipients  ", font=FONT_UI,
                                  bg=BG, fg=ACCENT, bd=1, relief="groove")
        frm_email.pack(fill="x", **pad)

        def _email_row(parent, row, label, var, color, note):
            tk.Label(parent, text=label, font=FONT_UI, bg=BG, fg=color
                     ).grid(row=row, column=0, sticky="w", padx=8, pady=4)
            tk.Entry(parent, textvariable=var, font=FONT_UI,
                     bg=BG2, fg=FG, insertbackground=FG, relief="flat",
                     width=50).grid(row=row, column=1, padx=8, pady=4, sticky="ew")
            tk.Label(parent, text=note, font=("Segoe UI", 7), bg=BG, fg=FG_DIM
                     ).grid(row=row, column=2, sticky="w", padx=(0, 8))

        self.report_email_var = tk.StringVar(
            value=self.cfg.get("email_to_report", _EMAIL_TO))
        self.alert_email_var  = tk.StringVar(
            value=self.cfg.get("email_to_alert",  self.cfg.get("email_to_report", _EMAIL_TO)))

        _email_row(frm_email, 0, "Report To:",  self.report_email_var,
                   GREEN,  "Final report with BinDist (semicolons OK)")
        _email_row(frm_email, 1, "Alerts To:",  self.alert_email_var,
                   AMBER,  "AQUA errors, pipeline failures")
        frm_email.columnconfigure(1, weight=1)

        # ── Program filter ────────────────────────────────────────────────────
        frm_prog = tk.LabelFrame(self, text="  Program Filter  ", font=FONT_UI,
                                 bg=BG, fg=ACCENT, bd=1, relief="groove")
        frm_prog.pack(fill="both", expand=True, **pad)

        tk.Label(frm_prog,
                 text="Unchecked programs are excluded from the email report (pipeline still runs).",
                 font=("Segoe UI", 8), bg=BG, fg=FG_DIM
                 ).pack(anchor="w", padx=8, pady=(4, 0))

        # Toolbar: select-all / deselect-all
        tb = tk.Frame(frm_prog, bg=BG)
        tb.pack(fill="x", padx=8, pady=(2, 0))
        self._btn(tb, "✔ All", self._select_all).pack(side="left", padx=(0, 6))
        self._btn(tb, "✘ None", self._deselect_all).pack(side="left", padx=(0, 6))
        self._btn(tb, "↺ Refresh", self._populate).pack(side="left")

        # Scrollable canvas for groups
        canvas_frame = tk.Frame(frm_prog, bg=BG)
        canvas_frame.pack(fill="both", expand=True, padx=8, pady=6)

        self.canvas = tk.Canvas(canvas_frame, bg=BG, highlightthickness=0)
        vsb = ttk.Scrollbar(canvas_frame, orient="vertical",
                            command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.inner = tk.Frame(self.canvas, bg=BG)
        self._canvas_win = self.canvas.create_window((0, 0), window=self.inner,
                                                      anchor="nw")
        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)

        # ── Status / buttons ──────────────────────────────────────────────────
        bot = tk.Frame(self, bg=BG)
        bot.pack(fill="x", padx=14, pady=(0, 12))

        self.status_var = tk.StringVar(value="")
        tk.Label(bot, textvariable=self.status_var, font=("Segoe UI", 9),
                 bg=BG, fg=GREEN).pack(side="left")

        self._btn(bot, "Cancel", self.destroy, fg=FG_DIM).pack(side="right", padx=(6, 0))
        self._btn(bot, "Save", self._save, bg="#1b5e20", fg="#c8e6c9").pack(side="right")

        self.check_vars: dict[str, tk.BooleanVar] = {}

    def _btn(self, parent, text, cmd, bg=BG3, fg=FG, **kw):
        return tk.Button(parent, text=text, command=cmd,
                         font=FONT_UI, bg=bg, fg=fg,
                         activebackground=ACCENT, activeforeground=BG,
                         relief="flat", padx=10, pady=3, cursor="hand2", **kw)

    # ── Populate groups ───────────────────────────────────────────────────────

    def _populate(self) -> None:
        for w in self.inner.winfo_children():
            w.destroy()
        self.check_vars.clear()

        keys   = _discover_keys(self.base_dir)
        groups = _group_keys(keys)

        if not keys:
            tk.Label(self.inner,
                     text="No TP keys found in data/programs/.\nRun automation first.",
                     font=FONT_UI, bg=BG, fg=FG_DIM
                     ).pack(padx=12, pady=20)
            return

        for letter, tp_keys in groups.items():
            prog_name = letter  # e.g. H61G, M61H

            # Group header
            hdr = tk.Frame(self.inner, bg=BG3)
            hdr.pack(fill="x", pady=(8, 0))
            tk.Label(hdr, text=f"  {prog_name}", font=FONT_GROUP,
                     bg=BG3, fg=ACCENT).pack(side="left", padx=6, pady=4)
            n_excl = sum(1 for k in tp_keys if k in self.excluded)
            if n_excl:
                tk.Label(hdr, text=f"{n_excl} excluded", font=("Segoe UI", 8),
                         bg=BG3, fg=AMBER).pack(side="right", padx=8)

            # One row per TP key
            grp_frame = tk.Frame(self.inner, bg=BG2, bd=0)
            grp_frame.pack(fill="x", pady=(0, 2))

            for tp_key in tp_keys:
                included = tp_key not in self.excluded
                var = tk.BooleanVar(value=included)
                self.check_vars[tp_key] = var

                row = tk.Frame(grp_frame, bg=BG2)
                row.pack(fill="x", padx=4, pady=1)

                cb = tk.Checkbutton(
                    row, variable=var, bg=BG2, fg=FG,
                    activebackground=BG2, activeforeground=ACCENT,
                    selectcolor=BG3, relief="flat", cursor="hand2",
                    command=lambda k=tp_key, v=var: self._on_toggle(k, v),
                )
                cb.pack(side="left")

                # Extract op number for display
                m_op  = re.search(r'_(\d{5,6})$', tp_key)
                op_lbl = f"op {m_op.group(1)}" if m_op else ""

                tk.Label(row, text=tp_key, font=FONT_MONO,
                         bg=BG2, fg=FG if included else FG_DIM).pack(side="left", padx=(2, 10))
                tk.Label(row, text=op_lbl, font=("Segoe UI", 8),
                         bg=BG2, fg=FG_DIM).pack(side="left")

                state_lbl = tk.Label(row,
                                     text="included" if included else "EXCLUDED",
                                     font=("Segoe UI", 8),
                                     bg=BG2, fg=GREEN if included else RED)
                state_lbl.pack(side="right", padx=8)

                # Keep reference to update on toggle
                var._label     = state_lbl   # type: ignore[attr-defined]
                var._key_label = row.winfo_children()[1]   # type: ignore[attr-defined]

        self.inner.update_idletasks()

    def _on_toggle(self, key: str, var: tk.BooleanVar) -> None:
        included = var.get()
        if included:
            self.excluded.discard(key)
        else:
            self.excluded.add(key)
        try:
            var._label.config(     # type: ignore[attr-defined]
                text="included" if included else "EXCLUDED",
                fg=GREEN if included else RED)
            var._key_label.config(fg=FG if included else FG_DIM)  # type: ignore[attr-defined]
        except Exception:
            pass

    def _select_all(self) -> None:
        self.excluded.clear()
        for k, v in self.check_vars.items():
            v.set(True)
            self._on_toggle(k, v)

    def _deselect_all(self) -> None:
        for k, v in self.check_vars.items():
            v.set(False)
            self._on_toggle(k, v)

    # ── Scrolling ─────────────────────────────────────────────────────────────

    def _on_inner_configure(self, _event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self._canvas_win, width=event.width)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # ── Save ──────────────────────────────────────────────────────────────────

    def _save(self) -> None:
        report_to = self.report_email_var.get().strip()
        alert_to  = self.alert_email_var.get().strip()
        if not report_to:
            messagebox.showerror("Error", "Report recipient cannot be empty.")
            return
        if not alert_to:
            alert_to = report_to  # fall back to report list

        cfg = {
            "email_to_report": report_to,
            "email_to_alert":  alert_to,
            "excluded_keys":   sorted(self.excluded),
        }
        try:
            _save_config(self.cfg_path, cfg)
            self.cfg = cfg
            n = len(self.excluded)
            self.status_var.set(
                f"Saved — {n} key(s) excluded." if n else "Saved — all keys included."
            )
            self._populate()
        except Exception as e:
            messagebox.showerror("Save failed", str(e))


# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description="Manage email report filter config.")
    ap.add_argument("--base-dir", default=str(_BASE_DIR),
                    help="Base auto directory (for TP key discovery via data/programs/)")
    args = ap.parse_args()

    base_dir = Path(args.base_dir)
    app = EmailManagerApp(base_dir)
    app.mainloop()


if __name__ == "__main__":
    main()


_SHAREPOINT_HOST   = "intel.sharepoint.com"
_SITE_PATH         = "/sites/ftesdsexecution"
_DEST_FOLDER       = "General/NVL_CDIE/NVL-N2P_trackers/NVL816-BLLC"

# ── Azure AD / MSAL settings ─────────────────────────────────────────────────
_TENANT_ID         = "46c98d88-e344-4ed4-8496-4ed7712e255d"   # Intel tenant
_CLIENT_ID         = "04b07795-8ddb-461a-bbee-02f9e1bf7b46"   # Azure CLI (public, works for delegated Graph API)
_SCOPES            = ["https://graph.microsoft.com/Files.ReadWrite",
                      "https://graph.microsoft.com/Sites.ReadWrite.All"]
_TOKEN_CACHE_FILE  = Path(__file__).parent / ".sp_token_cache.bin"

# ── Proxy settings ───────────────────────────────────────────────────────────
# Override here OR set HTTPS_PROXY / HTTP_PROXY env vars.
# Leave as empty string to use env vars (recommended).
_PROXY = os.environ.get("HTTPS_PROXY") or os.environ.get("HTTP_PROXY") or ""
# Intel corporate proxy (uncomment if env vars are not set):
# _PROXY = "http://proxy-chain.intel.com:912"

def _proxies() -> dict | None:
    """Return a requests-compatible proxies dict, or None if no proxy set."""
    p = _PROXY.strip()
    return {"http": p, "https": p} if p else None

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Auth helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get_msal_app():
    try:
        import msal
    except ImportError:
        raise RuntimeError(
            "The 'msal' package is required.\n"
            "Install it with:  pip install msal requests"
        )
    cache = msal.SerializableTokenCache()
    if _TOKEN_CACHE_FILE.exists():
        cache.deserialize(_TOKEN_CACHE_FILE.read_text(encoding="utf-8"))

    # Pass a requests Session with proxy to MSAL
    proxies = _proxies()
    http_client = None
    if proxies:
        import requests
        session = requests.Session()
        session.proxies = proxies
        http_client = session

    app = msal.PublicClientApplication(
        _CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{_TENANT_ID}",
        token_cache=cache,
        http_client=http_client,
    )
    return app, cache


def _acquire_token(progress_cb=None) -> str:
    """Return a valid access token, using cache or device-code flow."""
    app, cache = _get_msal_app()

    # Try silent first (cached token / refresh token)
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(_SCOPES, account=accounts[0])
        if result and "access_token" in result:
            _persist_cache(cache)
            return result["access_token"]

    # Fall back to device-code flow
    flow = app.initiate_device_flow(scopes=_SCOPES)
    if "user_code" not in flow:
        raise RuntimeError(f"Device flow initiation failed: {flow}")

    msg = (
        f"\nTo authenticate with SharePoint, open a browser and go to:\n"
        f"  {flow['verification_uri']}\n"
        f"Enter code: {flow['user_code']}\n"
    )
    log.info(msg)
    if progress_cb:
        progress_cb(msg)

    result = app.acquire_token_by_device_flow(flow)
    if "access_token" not in result:
        raise RuntimeError(
            f"Authentication failed: {result.get('error_description', result)}"
        )
    _persist_cache(cache)
    return result["access_token"]


def _persist_cache(cache) -> None:
    if cache.has_state_changed:
        _TOKEN_CACHE_FILE.write_text(cache.serialize(), encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# Graph API helpers
# ─────────────────────────────────────────────────────────────────────────────

def _graph_put(token: str, url: str, data: bytes, content_type: str) -> dict:
    import requests
    resp = requests.put(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": content_type,
        },
        data=data,
        timeout=60,
        proxies=_proxies(),
    )
    resp.raise_for_status()
    return resp.json()


def _upload_small(token: str, site_rel_path: str, filename: str,
                  content: bytes) -> dict:
    """Upload a file ≤4 MB via simple PUT."""
    import requests
    # Encode path components (spaces → %20, etc.) but keep slashes
    from urllib.parse import quote
    encoded_path = quote(f"{site_rel_path}/{filename}", safe="/")
    url = (
        f"https://graph.microsoft.com/v1.0"
        f"/sites/{_SHAREPOINT_HOST}:{_SITE_PATH}"
        f":/drive/root:/{encoded_path}:/content"
    )
    return _graph_put(token, url, content, "text/html; charset=utf-8")


def _upload_large(token: str, site_rel_path: str, filename: str,
                  content: bytes, progress_cb=None) -> dict:
    """Upload a file >4 MB via upload session."""
    import requests
    from urllib.parse import quote

    encoded_path = quote(f"{site_rel_path}/{filename}", safe="/")
    session_url = (
        f"https://graph.microsoft.com/v1.0"
        f"/sites/{_SHAREPOINT_HOST}:{_SITE_PATH}"
        f":/drive/root:/{encoded_path}:/createUploadSession"
    )
    sess_resp = requests.post(
        session_url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json={"item": {"@microsoft.graph.conflictBehavior": "replace"}},
        timeout=30,
        proxies=_proxies(),
    )
    sess_resp.raise_for_status()
    upload_url = sess_resp.json()["uploadUrl"]

    chunk_size = 5 * 1024 * 1024  # 5 MB chunks
    total = len(content)
    offset = 0
    result: dict = {}
    while offset < total:
        end = min(offset + chunk_size, total) - 1
        chunk = content[offset: end + 1]
        headers = {
            "Content-Length": str(len(chunk)),
            "Content-Range": f"bytes {offset}-{end}/{total}",
        }
        r = requests.put(upload_url, headers=headers, data=chunk, timeout=120,
                         proxies=_proxies())
        r.raise_for_status()
        if r.status_code in (200, 201):
            result = r.json()
        if progress_cb:
            progress_cb(f"Uploading… {min(end + 1, total) / total * 100:.0f}%")
        offset = end + 1
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def upload_report(
    file_path: Path,
    dest_folder: str = _DEST_FOLDER,
    progress_cb=None,
) -> str:
    """Upload *file_path* to SharePoint and return the web URL.

    Parameters
    ----------
    file_path   : local path to the HTML report file
    dest_folder : relative folder inside the SharePoint document library
    progress_cb : optional callable(message: str) for status updates
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Report file not found: {file_path}")

    token = _acquire_token(progress_cb)

    content = file_path.read_bytes()
    filename = file_path.name

    if progress_cb:
        progress_cb(f"Uploading {filename} ({len(content) / 1024:.0f} KB)…")

    if len(content) <= 4 * 1024 * 1024:
        result = _upload_small(token, dest_folder, filename, content)
    else:
        result = _upload_large(token, dest_folder, filename, content, progress_cb)

    web_url = result.get("webUrl", "")
    if progress_cb:
        progress_cb(f"Upload complete → {web_url}")
    return web_url


def list_reports(dest_folder: str = _DEST_FOLDER) -> list[dict]:
    """Return list of {name, webUrl, size, lastModified} dicts from SharePoint."""
    import requests
    from urllib.parse import quote
    token = _acquire_token()
    encoded_path = quote(dest_folder, safe="/")
    url = (
        f"https://graph.microsoft.com/v1.0"
        f"/sites/{_SHAREPOINT_HOST}:{_SITE_PATH}"
        f":/drive/root:/{encoded_path}:/children"
        f"?$select=name,webUrl,size,lastModifiedDateTime,file"
        f"&$orderby=lastModifiedDateTime desc"
        f"&$top=50"
    )
    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
        proxies=_proxies(),
    )
    resp.raise_for_status()
    items = resp.json().get("value", [])
    return [
        {
            "name": i["name"],
            "webUrl": i.get("webUrl", ""),
            "size": i.get("size", 0),
            "lastModified": i.get("lastModifiedDateTime", ""),
        }
        for i in items
        if "file" in i
    ]


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    if len(sys.argv) < 2:
        print("Usage: python publish_sharepoint.py <report.html>")
        sys.exit(1)
# ════════════════════════════════════════════════════════════════
# (formerly run_automation.py)
# ════════════════════════════════════════════════════════════════

import argparse
import csv
import gzip
import io
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

# ── Ensure UTF-8 output on Windows ─────────────────────────────────────────────
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Paths ──────────────────────────────────────────────────────────────────────
_HERE        = Path(__file__).resolve().parent
_REPO_ROOT   = _find_repo_root(_HERE)   # app.dashboard.nvl/
_PIPELINE    = _REPO_ROOT / "yield-dashboard" / "yld" / "yield_pipeline.py"
_COMPARE_RUNS = _HERE / "compare_runs.py"
_AQUA_CFG   = _REPO_ROOT / "shared" / "setup" / "automation" / "yield-dashboard" / "NVL_Sort_Yield - AutoPull.txt"

# ── Defaults ───────────────────────────────────────────────────────────────────
_BASE_DIR    = Path(r"\\samba.zsc10.intel.com\nfs\zsc10\disks\gsc_gwa011\users\snpant\auto\yield")
_DATA_DIR    = _BASE_DIR / "data"
_RUN_LOG     = _BASE_DIR / "run_log.html"
_EMAIL_TO    = "sujit.n.pant@intel.com"
_DEFAULT_DAYS = 7

_AQUA_EXE_GAR = r"\\gar.corp.intel.com\ec\proj\ba\aqua\AquaHbase\AquaCMDClient\Client\AquaCmdLine.exe"
_AQUA_EXE_AMR = r"\\FMSAPP3301.amr.corp.intel.com\Installer\AquaHbase\AquaCMDClient\Client\AquaCmdLine.exe"

_TP_FOLDER    = r"I:\program\1001\prod\hdmtprogs\nvl_ncx_sds"
_PROD_CFG_DIR = _REPO_ROOT / "shared" / "setup" / "config" / "yield-dashboard"
_EMAIL_CFG    = _REPO_ROOT / "shared" / "setup" / "automation" / "yield-dashboard" / "yield_setup_config.json"
_7Z_EXE       = Path(r"C:\Program Files\7-Zip\7z.exe")


# ─────────────────────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────────────────────

def _log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ─────────────────────────────────────────────────────────────────────────────
# Step 1 — AQUA pull
# ─────────────────────────────────────────────────────────────────────────────

def _aqua_report_name(config_path: Path) -> str:
    try:
        for line in config_path.read_text(encoding="utf-8-sig", errors="replace").splitlines():
            if line.strip().startswith("@ Report :"):
                return line.strip().split(":", 1)[1].strip()
    except Exception:
        pass
    return "NVL_Sort_Yield"


def _compress_aqua_to_7z(gz_path: Path) -> Path | None:
    """Re-compress a .csv.gz AQUA snapshot to .7z (better compression).
    Returns the new .7z path on success, or None on failure.
    The original .csv.gz is deleted after successful compression.
    """
    if not _7Z_EXE.exists():
        return None
    if gz_path.suffix != ".gz" or not gz_path.stem.endswith(".csv"):
        return None

    csv_path = gz_path.with_suffix("")                          # strip .gz → .csv
    z7_path  = gz_path.parent / (gz_path.stem[:-4] + ".7z")   # NAME.7z
    try:
        # 1. Decompress .csv.gz → .csv
        with gzip.open(gz_path, "rb") as fi, open(csv_path, "wb") as fo:
            shutil.copyfileobj(fi, fo)
        # 2. Compress .csv → .7z
        result = subprocess.run(
            [str(_7Z_EXE), "a", "-mx=5", "-mmt=on", str(z7_path), str(csv_path)],
            capture_output=True, text=True, timeout=300,
        )
        if result.returncode != 0:
            _log(f"  WARNING: 7z compression failed: {result.stderr.strip()[:200]}")
            return None
        # 3. Delete originals
        try: csv_path.unlink()
        except Exception: pass
        try: gz_path.unlink()
        except Exception: pass
        return z7_path
    except Exception as e:
        _log(f"  WARNING: _compress_aqua_to_7z: {e}")
        return None
    finally:
        if csv_path.exists():
            try: csv_path.unlink()
            except Exception: pass


def pull_aqua(aqua_exe: str, report_config: Path, data_dir: Path, dry_run: bool) -> Path | None:
    """Run AquaCmdLine.exe with the repo config. Returns path to the downloaded file."""
    data_dir.mkdir(parents=True, exist_ok=True)
    ts       = _ts()
    out_base = data_dir / f"aqua_pull_{ts}"
    out_req  = out_base.with_suffix(".zip")   # AQUA ignores extension; we glob after

    report_name = _aqua_report_name(report_config)
    temp_dir    = Path(os.environ.get("TEMP", tempfile.gettempdir()))
    temp_pat    = f"{report_name}*.CSV"

    # Derive server name from exe path (amr → AMR, default GAR)
    _exe_lower = str(aqua_exe).lower()
    _aqua_server = "AMR" if "amr" in _exe_lower else "GAR"

    cmd = [
        aqua_exe,
        "-AquaServer",    _aqua_server,
        "-ReportConfig",  str(report_config),
        "-OutputFileName", str(out_req),
    ]

    _log(f"{'DRY-RUN  ' if dry_run else ''}AQUA pull → {out_base}.*")
    _log(f"  Config : {report_config}")
    _log(f"  CMD    : {' '.join(cmd)}")

    if dry_run:
        _log("  DRY-RUN: skipping AQUA, returning dummy path")
        return out_base.with_suffix(".csv.gz")

    before_temp = {p.resolve() for p in temp_dir.glob(temp_pat)}

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=3600)
        if result.stdout.strip():
            _log(f"  AQUA: {result.stdout.strip()[:400]}")
        if result.returncode != 0:
            _log(f"  ERROR: AQUA rc={result.returncode}\n{result.stderr.strip()[:400]}")
            return None
    except FileNotFoundError:
        _log(f"  ERROR: AquaCmdLine.exe not found: {aqua_exe}")
        return None
    except subprocess.TimeoutExpired:
        _log("  ERROR: AQUA timed out")
        return None

    # Primary: any file written to data_dir with our stem
    written = [p for p in data_dir.glob(f"{out_base.name}*") if p.stat().st_size > 0]
    if written:
        out = max(written, key=lambda p: p.stat().st_mtime)
        _log(f"  Output: {out.name} ({out.stat().st_size:,} bytes)")
        return out

    # Fallback: new CSV in %TEMP%
    after_temp = {p.resolve() for p in temp_dir.glob(temp_pat)}
    new_csvs   = sorted(after_temp - before_temp, key=lambda p: p.stat().st_mtime)
    if new_csvs:
        src  = max(new_csvs, key=lambda p: p.stat().st_mtime)
        dest = data_dir / f"aqua_pull_{ts}.csv"
        shutil.copy2(src, dest)
        _log(f"  Fallback from %TEMP%: {src.name} → {dest.name}")
        return dest

    _log("  ERROR: AQUA produced no output file")
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — Read & split CSV by (variant, operation)
# ─────────────────────────────────────────────────────────────────────────────

def _read_aqua_file(path: Path) -> tuple[list[dict], str]:
    """
    Read an AQUA output file (.csv, .csv.gz, .zip, .7z).
    Handles nested chains: 7z→zip→csv, 7z→csv.gz, 7z→csv, zip→csv, gz→csv.
    Returns (rows, delimiter).
    """
    def _inner_from_bytes(raw: bytes) -> str:
        """Recursively unwrap zip/gz layers until we have plain CSV text."""
        if raw[:6] == b'7z\xbc\xaf\x27\x1c':
            import tempfile, subprocess as _sp
            with tempfile.TemporaryDirectory() as _tmp:
                _tmp_p = Path(_tmp)
                _sp.run([str(_7Z_EXE), "e", str(path), f"-o{_tmp}", "-y"],
                        check=True, capture_output=True)
                # Prefer .csv > .csv.gz > .zip (in case nested)
                for _pat in ("*.csv", "*.csv.gz", "*.zip"):
                    _hits = sorted(_tmp_p.glob(_pat))
                    if _hits:
                        return _inner_from_bytes(_hits[0].read_bytes())
            raise ValueError(f"No CSV/zip/gz found inside {path.name}")
        elif raw[:2] == b'PK':
            with zipfile.ZipFile(io.BytesIO(raw)) as z:
                # Pick first .csv; if none, pick first entry and recurse
                names = z.namelist()
                pick = next((n for n in names if n.lower().endswith('.csv')), names[0])
                return _inner_from_bytes(z.read(pick))
        elif raw[:2] == b'\x1f\x8b':
            return _inner_from_bytes(gzip.decompress(raw))
        else:
            return raw.decode("utf-8-sig", errors="replace")

    inner = _inner_from_bytes(path.read_bytes())
    first_line = inner.split("\n")[0]
    delim = "\t" if "\t" in first_line else ","
    rows = list(csv.DictReader(io.StringIO(inner), delimiter=delim))
    return rows, delim


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — Split by (TestProgram full name, Operation) and maintain per-TP gzs
# ─────────────────────────────────────────────────────────────────────────────

def _write_gz(rows: list[dict], fieldnames: list[str], path: Path) -> None:
    """Write rows as gzip-compressed CSV (UTF-8, comma-delimited)."""
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=fieldnames,
                       extrasaction="ignore", lineterminator="\n")
    w.writeheader()
    for row in rows:
        w.writerow({k: row.get(k, "") for k in fieldnames})
    path.write_bytes(gzip.compress(buf.getvalue().encode("utf-8"), compresslevel=6))


def _safe_filename(s: str) -> str:
    """Strip characters that are unsafe in filenames."""
    return re.sub(r'[\\/:*?"<>|]', '_', s).strip()


def split_by_tp_oper(rows: list[dict]) -> dict[str, tuple[list[dict], list[str]]]:
    """
    Split AQUA rows by (full TestProgram name, Operation code).

    Returns:
        dict  safe_key → (rows, fieldnames)
        safe_key = "{safe_tp_name}_{op_code}"
            e.g. "NCXSDJXL0H61A002618_119325"
                 "NCXSDJXL0H61B002618_119325"
                 "NCXSDJXL0H61C002618_132322"

    Wide format: columns like 'Program Name_119325', 'Lot_119325', …
        Each row spans all ops; extract per-op subset and rename columns
        (strip _{op} suffix).  Common columns (no suffix) are always included.

    Tall format: one row per die per op; has 'Program Name' and 'Operation' columns.
    """
    if not rows:
        return {}

    headers    = list(rows[0].keys())
    header_set = set(headers)

    # Detect op codes embedded in column names (5-6 digit numbers as suffix)
    op_codes: set[str] = set()
    for h in headers:
        m = re.search(r'_(\d{5,6})$', h)
        if m:
            op_codes.add(m.group(1))

    groups: dict[str, tuple[list[dict], list[str]]] = {}

    # ── Wide format ────────────────────────────────────────────────────────
    # Also handle single-op wide format (all columns have one _{op} suffix,
    # e.g. "Program Name_119325" and data rows have multiple programs in that column).
    if len(op_codes) >= 1:
        _log(f"  Wide format — ops: {sorted(op_codes)}")
        common_cols = [h for h in headers if not re.search(r'_\d{5,6}$', h)]

        for op in sorted(op_codes):
            prog_col = f"Program Name_{op}"
            if prog_col not in header_set:
                continue

            # Narrow rows for this op: common cols + op-specific cols renamed
            by_prog: dict[str, list[dict]] = {}
            for row in rows:
                prog = (row.get(prog_col) or "").strip()
                if not prog or prog.upper() in ("N/A", "NA", "NONE", "-", ""):
                    continue
                narrow: dict = {}
                for col in common_cols:
                    narrow[col] = row.get(col, "")
                for col, val in row.items():
                    if col.endswith(f"_{op}"):
                        narrow[col] = val
                by_prog.setdefault(prog, []).append(narrow)

            for prog, prog_rows in by_prog.items():
                key  = _safe_filename(f"{prog}_{op}")
                hdrs = list(prog_rows[0].keys())
                groups[key] = (prog_rows, hdrs)
                _log(f"    {key}: {len(prog_rows):,} rows")

        return groups

    # ── Tall / single-op format ─────────────────────────────────────────────
    _log("  Tall/single-op format")
    prog_col = next((h for h in headers if h.lower() in
                     ("program name", "testprogram", "test program", "program")), None)
    op_col   = next((h for h in headers if h.lower() == "operation"), None)

    for row in rows:
        prog = (row.get(prog_col) or "").strip() if prog_col else ""
        op   = (row.get(op_col)   or (next(iter(op_codes), "unknown"))).strip()
        key  = _safe_filename(f"{prog}_{op}") if prog else f"unknown_{op}"
        if key not in groups:
            groups[key] = ([], list(row.keys()))
        groups[key][0].append(row)

    for key, (rws, _) in groups.items():
        _log(f"    {key}: {len(rws):,} rows")

    return groups


def _lot_wafer_set(rows: list[dict]) -> frozenset:
    """Return a frozenset of (lot, wafer, date) strings for change-detection.
    Date is included so a re-test of the same lot/wafer with a new test date triggers a re-run.
    """
    if not rows:
        return frozenset()
    hdrs      = list(rows[0].keys())
    # Strip session suffixes like "_119325" when matching column names so that
    # AQUA columns such as "LOTS End Date Time_119325" are correctly identified.
    def _bare(h: str) -> str:
        return re.sub(r'_\d{4,}$', '', h).lower()

    lot_col   = next((h for h in hdrs if _bare(h) in
                      ("lot", "sort_lot", "lot number", "lot_number", "lot id")), None)
    wafer_col = next((h for h in hdrs if _bare(h) in
                      ("wafer", "sort_wafer", "wafer number", "wafer_number", "wafer id")), None)
    date_col  = next((h for h in hdrs if _bare(h) in
                      ("date", "test date", "test_date", "testdate",
                       "start date", "start_date", "finish date", "finish_date",
                       "insertion", "insert_date", "lots end date time",
                       "lots end date", "lots_end_date_time")), None)
    if not lot_col or not wafer_col:
        return frozenset()
    return frozenset(
        (
            str(r.get(lot_col,   "")).strip(),
            str(r.get(wafer_col, "")).strip(),
            str(r.get(date_col,  "")).strip() if date_col else "",
        )
        for r in rows
    )


def update_tp_gz(
    key: str,
    new_rows: list[dict],
    fieldnames: list[str],
    data_dir: Path,
    dry_run: bool,
) -> tuple[Path, bool]:
    """Write data_dir/programs/{letter}/{key}.csv.gz (always — no fingerprint check).

    Returns (gz_path, True) always (or (gz_path, False) in dry-run).
    """
    prog_dir = data_dir / "programs"
    _m = re.search(r'([A-Za-z]\d{2}[A-Za-z])', key)
    _letter_sub = _m.group(1).upper() if _m else "H61X"
    letter_dir  = prog_dir / _letter_sub
    gz_path     = letter_dir / f"{key}.csv.gz"
    z7_path     = letter_dir / f"{key}.7z"

    if not dry_run:
        letter_dir.mkdir(parents=True, exist_ok=True)

    _log(f"  {key}: writing {len(new_rows):,} rows")

    if not dry_run:
        _write_gz(new_rows, fieldnames, gz_path)
        _log(f"    → {gz_path.stat().st_size:,} bytes (gz)")
        _final = _compress_aqua_to_7z(gz_path)
        if _final:
            _log(f"    → compressed: {_final.name}  ({_final.stat().st_size:,} bytes)")
            return _final, True
    else:
        _log(f"    DRY-RUN: would write {gz_path}")

    return gz_path, True


# ─────────────────────────────────────────────────────────────────────────────
# Step 3 — Run pipeline for each group
# ─────────────────────────────────────────────────────────────────────────────

def _combine_gz(gz_files: list[Path], out_path: Path, dry_run: bool) -> Path:
    """
    Concatenate all per-TP gz files into a single combined.csv.gz.
    Column set = union of all files; missing values filled with empty string.
    Returns out_path.
    """
    if not gz_files:
        return out_path

    all_rows: list[dict] = []
    all_cols: list[str]  = []

    for f in gz_files:
        if not f.exists():
            _log(f"  DRY-RUN: {f.name} (not yet written)")
            continue
        rows, _ = _read_aqua_file(f)
        if not rows:
            continue
        for col in rows[0].keys():
            if col not in all_cols:
                all_cols.append(col)
        all_rows.extend(rows)

    _log(f"Combining {len(gz_files)} gz files → {len(all_rows):,} rows, {len(all_cols)} cols")

    if not dry_run:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        _write_gz(all_rows, all_cols, out_path)
        _log(f"  Combined gz: {out_path.name}  ({out_path.stat().st_size:,} bytes)")
    else:
        _log(f"  DRY-RUN: would write {out_path}")

    return out_path


_WATERMARK_CSS = """
<style id="_wm_style">
#_wm_badge {
  position: fixed;
  top: 8px;
  right: 14px;
  z-index: 99999;
  background: #6c3483;
  color: #ffffff;
  font: bold 11px/1.4 Arial, sans-serif;
  padding: 3px 10px;
  border-radius: 4px;
  letter-spacing: 0.3px;
  pointer-events: none;
  white-space: nowrap;
}
</style>
"""
_WATERMARK_HTML = '<div id="_wm_badge">Pant, Sujit N &mdash; GEMS FTE</div>'


def _inject_watermark(html_path: Path) -> None:
    """Inject the watermark badge into an HTML file in-place (idempotent)."""
    try:
        text = html_path.read_text(encoding="utf-8", errors="replace")
        if "_wm_badge" in text or "_wm_div" in text:
            return   # already watermarked
        # Insert CSS before </head> (or at top if no </head>)
        if "</head>" in text:
            text = text.replace("</head>", _WATERMARK_CSS + "</head>", 1)
        else:
            text = _WATERMARK_CSS + text
        # Insert badge div after <body …> tag
        import re as _re
        text = _re.sub(
            r'(<body[^>]*>)',
            r'\1\n' + _WATERMARK_HTML,
            text, count=1, flags=_re.IGNORECASE,
        )
        if _WATERMARK_HTML not in text:   # no <body> tag at all
            text = text + _WATERMARK_HTML
        html_path.write_text(text, encoding="utf-8")
    except Exception as e:
        _log(f"  watermark warning: {html_path.name}: {e}")


def _watermark_output_dir(output_dir: str) -> None:
    """Watermark all HTML files in the pipeline output folder."""
    d = Path(output_dir)
    if not d.exists():
        return
    html_files = list(d.rglob("*.html"))
    _log(f"  Watermarking {len(html_files)} HTML file(s) in {d.name}")
    for f in html_files:
        _inject_watermark(f)


def _rebuild_dashboard_html_for_tp(tp_key: str, base_dir: Path, out_path: Path | None = None) -> Path | None:
    """Synthesize Dashboard_{tp_key}.html at base_dir from existing historical run folders.

    Used when pipeline.py has not created/updated it (e.g. first run or fresh samba).
    Scans output/ for NVL_0H61_* run folders, finds the matching TP sub-dir in each,
    and builds a minimal Dashboard HTML with one run-block per historical run.
    Returns the written path, or None if no runs were found.
    """
    output_dir = base_dir / "output"
    if not output_dir.exists():
        return None

    # Strip the op-suffix to get the TP prefix (e.g. NCXSDJXL0H61C002620)
    tp_prefix = re.sub(r'_\d{5,6}$', '', tp_key)

    run_folders = sorted(
        [d for d in output_dir.iterdir()
         if d.is_dir() and re.search(r'_\d{8}_\d{6}$', d.name)],
        key=lambda d: d.name,   # ascending = oldest first; we'll reverse for display
    )

    # run_folders is sorted ascending (oldest first); we scan newest-last so
    # that same-date duplicates overwrite earlier same-day entries in seen_keys.
    seen_keys: dict[str, tuple] = {}   # dated_key → block tuple (newest wins)
    for rf in run_folders:
        # Find matching c_dirs (same prefix, not _R0)
        c_dirs = sorted(
            [d for d in rf.iterdir()
             if d.is_dir()
             and d.name.startswith(tp_prefix)
             and not d.name.endswith('_R0')],
            key=lambda d: d.name,
            reverse=True,
        )
        if not c_dirs:
            continue
        c_dir = c_dirs[0]
        index_html = c_dir / 'index.html'
        if not index_html.exists():
            continue
        href = os.path.relpath(str(index_html), str(base_dir)).replace('\\', '/')
        m = re.search(r'(\d{8})_(\d{6})$', rf.name)
        date_str = m.group(1) if m else '00000000'
        time_str = m.group(2) if m else '000000'
        ts = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]} {time_str[:2]}:{time_str[2:4]}"
        dated_key = f"{c_dir.name}_{date_str}"
        seen_keys[dated_key] = (  # overwrite → last (newest) run for this date wins
            dated_key,
            f'<div class="run-block" data-stem="{dated_key}">\n'
            f'<div class="run-header" onclick="toggle(this)">'
            f'<span class="arrow">&#9660;</span> {c_dir.name}'
            f'<span class="ts"> - {ts}</span></div>\n'
            f'<div class="run-body">\n'
            f'<a class="run-link report-link" href="{href}" target="_blank">Yield Report</a>\n'
            f'</div>\n</div>',
        )

    # Convert to list; seen_keys is ordered (Python 3.7+) oldest→newest (ascending scan)
    blocks = list(seen_keys.values())

    if not blocks:
        return None

    # Newest first in the HTML (reversed from our oldest-first scan order)
    blocks_html = '\n'.join(b for _, b in reversed(blocks))
    page_css = (
        '*{box-sizing:border-box;margin:0;padding:0}'
        'body{font-family:Arial,sans-serif;background:#1a252f;color:#ecf0f1;padding:16px}'
        'h1{font-size:16px;margin-bottom:14px;color:#3498db}'
        '.run-block{background:#2c3e50;border-radius:6px;margin-bottom:10px;overflow:hidden}'
        '.run-header{padding:10px 14px;cursor:pointer;display:flex;align-items:center;'
        'gap:6px;font-weight:bold;font-size:13px;user-select:none}'
        '.run-header:hover{background:#34495e}'
        '.arrow{font-size:10px;transition:transform .2s}'
        '.run-header.collapsed .arrow{transform:rotate(-90deg)}'
        '.ts{font-weight:normal;font-size:11px;color:#95a5a6;margin-left:auto}'
        '.run-body{padding:8px 14px 12px;display:flex;flex-wrap:wrap;gap:6px}'
        '.run-link{display:inline-block;padding:5px 10px;border-radius:4px;'
        'font-size:12px;text-decoration:none;white-space:nowrap}'
        '.report-link{background:#2980b9;color:#fff}'
        '.report-link:hover{background:#3498db}'
    )
    page_js = (
        "function toggle(hdr){"
        "hdr.classList.toggle('collapsed');"
        "hdr.nextElementSibling.style.display="
        "hdr.classList.contains('collapsed')?'none':'';}"
    )
    html = (
        '<!DOCTYPE html>\n<html lang="en">\n<head>'
        '<meta charset="utf-8">'
        f'<title>Dashboard {tp_key}</title>'
        f'<style>{page_css}</style>'
        f'<script>{page_js}</script>'
        '</head>\n<body>\n'
        f'<h1>Dashboard &mdash; {tp_key}</h1>\n'
        '<!-- YIELD_START -->\n'
        f'{blocks_html}\n'
        '<!-- YIELD_END -->\n'
        '</body>\n</html>\n'
    )
    dash_path = out_path or (base_dir / "output" / "misc" / f'Dashboard_{tp_key}.html')
    dash_path.parent.mkdir(parents=True, exist_ok=True)
    dash_path.write_text(html, encoding='utf-8')
    return dash_path


def _stamp_dashboard_block(dash_html_path: Path, block_key: str, date_str: str) -> None:
    """Rename the undated run-block in Dashboard HTML to a dated one.

    Each daily run writes a block with data-stem="tp_key" (always the same).
    By renaming it to data-stem="tp_key_YYYYMMDD" AFTER the pipeline writes it,
    previous days' blocks survive and compare_runs.py can show day-over-day trends.
    If a same-day dated block already exists (re-run), it is removed first.
    """
    if not dash_html_path.exists():
        return
    try:
        content = dash_html_path.read_text(encoding='utf-8')
        dated_key = f"{block_key}_{date_str}"
        # Remove any same-day dated block that may exist from an earlier re-run
        if f'data-stem="{dated_key}"' in content:
            block_re = re.compile(
                r'<div class="run-block" data-stem="' + re.escape(dated_key) + r'">'
                r'[\s\S]*?</div>\s*</div>',
                re.MULTILINE,
            )
            content = block_re.sub('', content)
        # Rename the freshly-written undated block to dated
        if f'data-stem="{block_key}"' in content:
            content = re.sub(
                r'data-stem="' + re.escape(block_key) + r'"',
                f'data-stem="{dated_key}"',
                content,
                count=1,
            )
            # Also add date in brackets to the display name in the run-header
            # e.g. "> tp_key<span" → "> tp_key (YYYY-MM-DD)<span"
            _fmt_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
            content = re.sub(
                r'(</span>\s*)(' + re.escape(block_key) + r')(<span\s+class="ts")',
                rf'\g<1>\g<2> ({_fmt_date})\g<3>',
                content,
                count=1,
            )
            dash_html_path.write_text(content, encoding='utf-8')
            _log(f"  Dashboard block stamped: {block_key} → {dated_key} ({_fmt_date})")
        else:
            _log(f"  WARNING: block data-stem=\"{block_key}\" not found in {dash_html_path.name}")
    except Exception as _e:
        _log(f"  WARNING: _stamp_dashboard_block failed: {_e}")


def _named_attachment(src: Path, display_name: str, tmp_dir: Path) -> str:
    """Copy src to tmp_dir/<display_name> so Outlook shows the desired filename."""
    dest = tmp_dir / display_name
    if Path(src).resolve() != dest.resolve():
        shutil.copy2(src, dest)
    return str(dest)


def _find_product_config(devrevstep: str) -> str | None:
    """
    Find a matching product config JSON by DevRevStep prefix (first 6 chars).
    E.g. devrevstep '8PL7CV' matches '8PL7CV - SORT - Product Config - NVL-GPU-512.json'.
    Falls back to first SORT config if no match found.
    """
    if not _PROD_CFG_DIR.exists():
        return None
    drs = devrevstep[:6].upper() if devrevstep else ""
    sort_configs = list(_PROD_CFG_DIR.glob("*.json"))
    sort_configs = [p for p in sort_configs if "SORT" in p.stem.upper()]
    if drs:
        for p in sort_configs:
            if p.stem.upper().startswith(drs):
                return str(p)
    return str(sort_configs[0]) if sort_configs else None


def run_pipeline_for_group(
    group_key: str,
    csv_path: Path,
    base_dir: Path,
    dry_run: bool,
) -> tuple[bool, str]:
    """
    Build input.json for the group and run pipeline.py --json.
    Returns (success, output_dir).
    """
    # Derive variant letter from group key '61A-119325' → 'A'
    m = re.match(r'61([A-Z])-(\d+)', group_key)
    variant = m.group(1) if m else "X"
    op_code = m.group(2) if m else "unknown"

    group_dir  = base_dir / group_key
    group_dir.mkdir(parents=True, exist_ok=True)

    identifier    = f"{group_key}_{datetime.now().strftime('%Y%m%d')}"
    output_folder = str(group_dir)
    dashboard     = str(group_dir / "Dashboard.html")
    prod_cfg      = _find_product_config(variant)
    tp_folder     = str(Path(_TP_FOLDER))

    cfg = {
        "DataCSV":             [str(csv_path)],
        "output_folder":       output_folder,
        "dashboard":           dashboard,
        "identifier":          identifier,
        "TestProgram_folder":  tp_folder,
        "run_parametric":      True,
        "keep_pcm_idw":        False,
    }
    if prod_cfg:
        cfg["product_config_json"] = prod_cfg

    json_path = group_dir / "input.json"
    json_path.write_text(json.dumps(cfg, indent=2), encoding="utf-8")
    _log(f"  [{group_key}] input.json → {json_path}")

    output_dir = str(Path(output_folder) / identifier)

    if dry_run:
        _log(f"  [{group_key}] DRY-RUN: would run pipeline.py --json {json_path}")
        return True, output_dir

    cmd = [sys.executable, str(_PIPELINE), "--json", str(json_path)]
    env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUNBUFFERED": "1"}
    _log(f"  [{group_key}] Running pipeline → {output_dir}")

    try:
        result = subprocess.run(cmd, capture_output=False, text=True,
                                timeout=3600, env=env, cwd=str(_PIPELINE.parent))
        ok = result.returncode == 0
        if not ok:
            _log(f"  [{group_key}] WARNING: pipeline exited rc={result.returncode}")
        return ok, output_dir
    except subprocess.TimeoutExpired:
        _log(f"  [{group_key}] ERROR: pipeline timed out")
        return False, output_dir


# ─────────────────────────────────────────────────────────────────────────────
# Compare-card helpers  (used by _build_run_report)
# ─────────────────────────────────────────────────────────────────────────────

def _extract_yield_summary(tp_dir: Path, row_filter: "set | None" = None) -> dict | None:
    """Parse *_BinDistribution.html (or legacy digital_dashboard.html) → {die, bins, repair_bins, sums}.
    row_filter: optional set of "lot|wafer" strings; when provided only those rows are counted.
    """
    # Try BinDistribution.html first (current pipeline output)
    _tp_bd = tp_dir / 'bin_dist'
    bd_files = sorted((_tp_bd if _tp_bd.exists() else tp_dir).glob("*_BinDistribution.html"))
    dd = bd_files[0] if bd_files else tp_dir / "digital_dashboard.html"
    if not dd.exists():
        return None
    try:
        txt = dd.read_text(encoding="utf-8", errors="replace")

        # --- New format: window.DATA={...} in data_summary.js, or var DATA inline ---
        _ds_js = dd.parent / 'data_summary.js'
        _data_pat = r'(?:var\s+DATA|window\.DATA)\s*=\s*'
        m_data = None
        txt = ''
        for _src in [_ds_js, dd]:
            try:
                txt = _src.read_text(encoding='utf-8', errors='replace')
            except Exception:
                continue
            m_data = re.search(_data_pat, txt)
            if m_data:
                break
        if m_data:
            decoder = json.JSONDecoder()
            data, _ = decoder.raw_decode(txt, m_data.end())
            all_rows = data.get("rows", [])
            # Apply lot/wafer filter if provided
            if row_filter:
                rows = [r for r in all_rows
                        if (str(r.get("lot", "")) + "|" + str(r.get("wafer", ""))) in row_filter]
            else:
                rows = all_rows
            total = sum(sum(row.get("binCounts", {}).values()) for row in rows)
            # Sum bin counts and functional bin counts across filtered wafers
            bin_totals: dict[str, int] = {}
            fb_totals:  dict[str, int] = {}
            for row in rows:
                for b, cnt in row.get("binCounts", {}).items():
                    bin_totals[b] = bin_totals.get(b, 0) + int(cnt)
                for _ib, fbmap in row.get("ibToFb", {}).items():
                    for fb, cnt in fbmap.items():
                        fb_totals[fb] = fb_totals.get(fb, 0) + int(cnt)

            def _pct(cnt: int) -> str:
                return f"{cnt / total * 100:.1f}%" if total > 0 else "–"

            named_bins  = {f"Bin {b}": _pct(c) for b, c in bin_totals.items()}
            repair_bins = {fb: _pct(c) for fb, c in fb_totals.items() if fb in ("198", "201", "202")}
            # --- HP/LP from die-level DLCP data (IB 1-4, threshold 92.5%, first UPM col) ---
            dlcp: dict = {}
            upm_med: str | None = None
            if data.get("hasUpm"):
                _ui = int(data.get("upmStart") or 5)  # first UPM column index
                _hp_tot = 0; _lp_tot = 0
                _upm_vals: list[float] = []
                for _dr in rows:
                    for _d in _dr.get("dies", []):
                        if len(_d) < 4:
                            continue
                        _ibi = int(_d[2]) if isinstance(_d[2], (int, float)) else -1
                        if _ibi not in (1, 2, 3, 4):
                            continue  # only DLCP dies
                        _up  = _d[_ui] if len(_d) > _ui else None
                        if _up is not None:
                            _upm_vals.append(float(_up))
                        if _ibi in (1, 2) and _up is not None and _up >= 92.5:
                            _hp_tot += 1
                        else:
                            _lp_tot += 1
                _dn = _hp_tot + _lp_tot
                if _dn >= 10:  # suppress if too few DLCP dies
                    dlcp = {"hp": f"{_hp_tot/_dn*100:.1f}%", "lp": f"{_lp_tot/_dn*100:.1f}%",
                            "n": _dn, "hp_n": _hp_tot, "lp_n": _lp_tot}
                if _upm_vals:
                    _upm_s = sorted(_upm_vals)
                    _nu    = len(_upm_s)
                    _umed  = _upm_s[_nu // 2] if _nu % 2 else (_upm_s[_nu//2 - 1] + _upm_s[_nu//2]) / 2
                    upm_med = f"{_umed:.1f}%"
            # ── Extract FF / FF+DF targets from yieldDefs ────────────────────
            ff_tgt   = "–"
            ffdf_tgt = "–"
            for _yd in data.get("yieldDefs", []):
                _bins_key = str(_yd.get("bins", "")).replace(" ", "")
                _exp = _yd.get("expected")
                if _exp is not None:
                    if _bins_key == "1/2":
                        ff_tgt = f"{float(_exp):.1f}%"
                    elif _bins_key == "1/2/3/4":
                        ffdf_tgt = f"{float(_exp):.1f}%"
            return {"die": f"{total:,}", "bins": named_bins, "repair_bins": repair_bins,
                    "sums": {}, "dlcp": dlcp, "upm_med": upm_med,
                    "ff_tgt": ff_tgt, "ffdf_tgt": ffdf_tgt}

        # --- Legacy format: digital_dashboard.html with DD_ROWS ---
        m_die = re.search(r"# Die: <b>([\d,]+)</b>", txt)
        total_die = m_die.group(1) if m_die else "–"
        m_rows = re.search(r"var DD_ROWS\s*=\s*(\[.*?\])\s*;", txt, re.DOTALL)
        if not m_rows:
            return {"die": total_die, "bins": {}, "repair_bins": {}, "sums": {}}
        rows = json.loads(m_rows.group(1))
        named_bins: dict[str, str] = {}
        repair_bins: dict[str, str] = {}
        section_sums: dict[str, str] = {}
        cur_section = ""
        for row in rows:
            cells = row.get("cells", [])
            if not cells:
                continue
            name = cells[0]
            val  = cells[1] if len(cells) > 1 else ""
            val  = re.sub(r'\s*\([\d,]+\)\s*$', '', val).strip()
            if   name.startswith("ARR_"):        cur_section = "ARR"
            elif name.startswith("FUN_"):        cur_section = "FUN"
            elif name.startswith("SCN_"):        cur_section = "SCN"
            elif re.match(r"^Bin \d+$", name):  cur_section = "Bins"
            if row.get("bold") and cur_section and cur_section not in section_sums:
                section_sums[cur_section] = val
            if re.match(r"^Bin \d+$", name):
                named_bins[name] = val
            m_repair = re.match(r"^Repair Bin (\d+)", name)
            if m_repair:
                repair_bins[m_repair.group(1)] = val
        return {"die": total_die, "bins": named_bins, "repair_bins": repair_bins, "sums": section_sums}
    except Exception:
        return None


def _extract_per_material_summaries(tp_dir: Path) -> list[tuple[str, dict | None]]:
    """Read BinDistribution.html in *tp_dir*, group rows by material type field,
    and return [(mat_type, summary_dict), ...] sorted by material type.
    Returns an empty list if only one material type is found (no breakdown needed).
    Falls back to empty list on any error."""
    _tp_bd2 = tp_dir / 'bin_dist'
    bd_files = sorted((_tp_bd2 if _tp_bd2.exists() else tp_dir).glob("*_BinDistribution.html"))
    dd = bd_files[0] if bd_files else tp_dir / "digital_dashboard.html"
    if not dd.exists():
        return []
    try:
        _ds_js2 = dd.parent / 'data_summary.js'
        _data_pat2 = r'(?:var\s+DATA|window\.DATA)\s*=\s*'
        m_data = None
        txt = ''
        for _src2 in [_ds_js2, dd]:
            try:
                txt = _src2.read_text(encoding='utf-8', errors='replace')
            except Exception:
                continue
            m_data = re.search(_data_pat2, txt)
            if m_data:
                break
        if not m_data:
            return []
        decoder = json.JSONDecoder()
        data, _ = decoder.raw_decode(txt, m_data.end())
        all_rows = data.get("rows", [])
        if not all_rows:
            return []

        # Try "material" field first, then fall back to "lot"
        mat_col = None
        for candidate in ("material", "material_type", "materialType", "lot"):
            if candidate in all_rows[0]:
                mat_col = candidate
                break
        if not mat_col:
            return []

        # Group rows by material type
        from collections import defaultdict as _dd
        mat_rows: dict[str, list[str]] = _dd(list)
        for row in all_rows:
            mat_id = str(row.get(mat_col, "")).strip() or "UNKNOWN"
            lot    = str(row.get("lot",   "")).strip()
            wafer  = str(row.get("wafer", "")).strip()
            mat_rows[mat_id].append(f"{lot}|{wafer}")

        results = []
        for mat_id in sorted(mat_rows):
            rf   = set(mat_rows[mat_id])
            smry = _extract_yield_summary(tp_dir, row_filter=rf)
            results.append((mat_id, smry))
        return results
    except Exception:
        return []


def _build_compare_section(sorted_groups: list, run_dir: Path, prog_series: str = "0H61",
                           prod_cfg: str | None = None, product_name: str = "NVL816-BLLC") -> str:
    """HTML for the two comparison cards shown at top of report.html."""

    def _is_stale(item) -> bool:
        return len(item) > 3 and str(item[3]).startswith("prev:")

    def _op(item) -> str:
        m = re.search(r"_(\d{5,6})$", item[0])
        return m.group(1) if m else "?"

    _auto_dir = run_dir.parent.parent  # …/auto/
    _gen = re.search(r'(\d+)$', prog_series)
    _gen = _gen.group(1) if _gen else '61'

    # ── Load FF / FF+DF targets from product config ───────────────────────────
    _ff_tgt = _ff_df_tgt = None
    try:
        cfg_path = prod_cfg or _find_product_config("")
        if cfg_path:
            _cfg = json.loads(Path(cfg_path).read_text(encoding="utf-8"))
            for entry in _cfg.get("yield_targets", []):
                bins_set = set(entry.get("bin", "").split("/"))
                y = entry.get("yield")
                if bins_set == {"1", "2"}:
                    _ff_tgt = float(y)
                elif bins_set == {"1", "2", "3", "4"}:
                    _ff_df_tgt = float(y)
    except Exception:
        pass

    def _parse_pct(s: str) -> float:
        try:
            return float(str(s).replace("%", "").strip())
        except Exception:
            return 0.0

    def _write_cross_compare_html(letter_data: list, auto_dir: Path) -> "Path | None":
        """Generate a standalone cross-program comparison HTML and save to auto_dir.
        letter_data: list of (letter, smry, tp_output_dir)"""
        if not letter_data:
            return None
        clrs = {"C": "#43a047", "B": "#1e88e5", "A": "#fb8c00"}

        def _hbar(letter: str, val: float, max_v: float, lbl: str) -> str:
            pct   = min(100.0, val / max_v * 100) if max_v > 0 else 0.0
            color = clrs.get(letter[-1], "#90a4ae")
            disp  = f"{val:.1f}%" if val > 0 else "\u2013"
            return (
                f"<div style='display:flex;align-items:center;gap:6px;margin:3px 0'>"
                f"<span style='width:36px;font-size:0.85em;text-align:right;color:#90a4ae'>{lbl}</span>"
                f"<div style='width:200px;background:#263950;height:14px;border-radius:3px'>"
                f"<div style='width:{pct:.1f}%;background:{color};height:14px;border-radius:3px'></div></div>"
                f"<span style='font-size:0.88em;min-width:46px;color:#cde'>{disp}</span>"
                f"</div>"
            )

        def _grp(title: str, pts: list, max_v: float, tgt: float | None = None) -> str:
            bars     = "".join(_hbar(l, v, max_v, l) for l, v in pts)
            tgt_note = (f"<div style='font-size:0.8em;color:#78909c;margin-left:44px'>"
                        f"Target:\u00a0{tgt:.1f}%</div>"
                        if tgt is not None else "")
            return (
                f"<div style='margin-bottom:14px'>"
                f"<div style='font-size:0.92em;font-weight:600;color:#4fc3f7;"
                f"margin-bottom:3px;margin-left:44px'>{title}</div>"
                f"{bars}{tgt_note}</div>"
            )

        ff_pts, ffdf_pts, fb198_pts, fb201_pts, fb202_pts = [], [], [], [], []
        for letter, smry, _ in letter_data:
            bins = smry.get("bins", {})        if smry else {}
            rb   = smry.get("repair_bins", {}) if smry else {}
            b1   = _parse_pct(bins.get("Bin 1", "0"))
            b2   = _parse_pct(bins.get("Bin 2", "0"))
            b3   = _parse_pct(bins.get("Bin 3", "0"))
            b4   = _parse_pct(bins.get("Bin 4", "0"))
            ff_pts.append((letter,   b1 + b2))
            ffdf_pts.append((letter, b1 + b2 + b3 + b4))
            fb198_pts.append((letter, _parse_pct(rb.get("198", "0"))))
            fb201_pts.append((letter, _parse_pct(rb.get("201") or rb.get("2") or "0")))
            fb202_pts.append((letter, _parse_pct(rb.get("202", "0"))))

        rep_max = max(
            max((v for _, v in fb198_pts), default=0),
            max((v for _, v in fb201_pts), default=0),
            max((v for _, v in fb202_pts), default=0), 1.0
        ) * 1.35

        chart_html = (
            "<div style='display:flex;gap:40px;flex-wrap:wrap;margin-bottom:28px'>"
            f"<div>{_grp('FF (1+2)', ff_pts, 100.0, _ff_tgt)}"
            f"{_grp('FF+DF (1+2+3+4)', ffdf_pts, 100.0, _ff_df_tgt)}</div>"
            "<div style='border-left:1px solid #2e4a6a;padding-left:28px'>"
            f"{_grp('FB198 Vmin Repair', fb198_pts, rep_max)}"
            f"{_grp('FB201 Vnom Repair', fb201_pts, rep_max)}"
            f"{_grp('FB202 Vmax Repair', fb202_pts, rep_max)}"
            "</div></div>"
        )

        ff_tgt_s    = f"{_ff_tgt:.1f}%"    if _ff_tgt    is not None else "\u2013"
        ff_df_tgt_s = f"{_ff_df_tgt:.1f}%" if _ff_df_tgt is not None else "\u2013"

        table_rows = ""
        for letter, smry, _ in letter_data:
            bins  = smry.get("bins", {})        if smry else {}
            rb    = smry.get("repair_bins", {}) if smry else {}
            dlcp  = smry.get("dlcp", {})        if smry else {}
            die   = smry.get("die", "\u2013")  if smry else "\u2013"
            b1    = _parse_pct(bins.get("Bin 1", "0"))
            b2    = _parse_pct(bins.get("Bin 2", "0"))
            b3    = _parse_pct(bins.get("Bin 3", "0"))
            b4    = _parse_pct(bins.get("Bin 4", "0"))
            ff    = b1 + b2
            ffdf  = b1 + b2 + b3 + b4
            ff_col    = "#66bb6a" if (_ff_tgt    is None or ff    >= _ff_tgt)    else "#ef5350"
            ffdf_col  = "#66bb6a" if (_ff_df_tgt is None or ffdf >= _ff_df_tgt) else "#ef5350"
            rv198   = rb.get("198", "\u2013")
            rv201   = rb.get("201") or rb.get("2") or "\u2013"
            rv202   = rb.get("202", "\u2013")
            rv_upm  = smry.get("upm_med", "\u2013") if smry else "\u2013"
            _d202 = dlcp  # dlcp is now a flat dict (aggregate over all DLCP IB 1-4 dies)
            if rv_upm == "\u2013":  # UPM absent → DLCP unreliable
                _d202_hp_col = "#546e7a"
                rv_hp = "\u2013"
                rv_lp = "\u2013"
            elif _d202:
                _span = "<span style='color:#546e7a;font-size:10px'>"
                _hp_ns = f"<br>{_span}({_d202['hp_n']:,})</span>" if _d202.get('hp_n') is not None else ""
                _lp_ns = f"<br>{_span}({_d202['lp_n']:,})</span>" if _d202.get('lp_n') is not None else ""
                _d202_hp_val = float(_d202['hp'].rstrip('%')) if _d202.get('hp') else 0.0
                _d202_hp_col = "#4caf50" if _d202_hp_val >= 30 else "#ef5350"
                rv_hp = f"{_d202['hp']}{_hp_ns}"
                rv_lp = f"{_d202['lp']}{_lp_ns}"
            else:
                _d202_hp_col = "#546e7a"
                rv_hp = "\u2013"
                rv_lp = "\u2013"
            prog_col = clrs.get(letter, "#80cbc4")
            table_rows += (
                f"<tr>"
                f"<td style='color:{prog_col};font-weight:bold;font-family:monospace'>{letter}</td>"
                f"<td>{die}</td>"
                f"<td style='color:{ff_col};font-weight:bold'>{ff:.1f}%</td>"
                f"<td style='color:{ffdf_col};font-weight:bold'>{ffdf:.1f}%</td>"
                f"<td>{rv_upm}</td>"
                f"<td style='color:{_d202_hp_col};font-weight:bold'>{rv_hp}</td>"
                f"<td style='color:#f0a500'>{rv_lp}</td>"
                f"<td>{rv198}</td><td>{rv201}</td><td>{rv202}</td>"
                f"</tr>\n"
            )

        bindist_cols = ""
        for letter, _, tp_dir in letter_data:
            tp_path = Path(tp_dir) if tp_dir else None
            bd_html = ""
            if tp_path and tp_path.exists():
                _bd_sub2 = tp_path / 'bin_dist'
                bdfiles = sorted((_bd_sub2 if _bd_sub2.exists() else tp_path).glob("*BinDistribution*.html"))
                if bdfiles:
                    try:
                        bd_content = bdfiles[0].read_text(encoding="utf-8", errors="replace")
                        bd_escaped = _html_mod.escape(bd_content, quote=True)
                        bd_html = (
                            f'<iframe srcdoc="{bd_escaped}" width="100%" height="920"'
                            f' style="border:none;background:#fff;border-radius:4px;display:block"></iframe>'
                        )
                    except Exception:
                        pass
            if not bd_html:
                bd_html = "<p style='color:#90a4ae;font-size:0.88em'>BinDistribution not available</p>"
            col = clrs.get(letter[-1], "#80cbc4")
            bindist_cols += (
                f"<div style='flex:1;min-width:340px'>"
                f"<div style='color:{col};font-weight:bold;font-size:1em;margin-bottom:6px'>{letter}</div>"
                f"{bd_html}</div>"
            )

        legend = "&ensp;".join(
            f"<span style='display:inline-flex;align-items:center;gap:5px'>"
            f"<span style='width:14px;height:14px;background:{clrs.get(l,'#90a4ae')};"
            f"border-radius:3px;display:inline-block'></span>"
            f"<span style='color:#cde;font-size:0.9em'>{l}</span></span>"
            for l, _, _ in letter_data
        )
        ts_now = datetime.now().strftime("%Y-%m-%d %H:%M")
        html = (
            "<!DOCTYPE html>\n<html lang='en'>\n"
            "<head><meta charset='utf-8'>"
            f"<title>{product_name} Cross-Program Comparison</title><style>\n"
            "body{font-family:Segoe UI,Arial,sans-serif;background:#1a252f;color:#e8f0f7;"
            "margin:0;padding:16px 32px 60px}\n"
            "h1{color:#4fc3f7;border-bottom:2px solid #4fc3f7;padding-bottom:8px;margin-bottom:4px}\n"
            ".ts{color:#90a4ae;font-size:0.85em;margin-top:0}\n"
            "table{border-collapse:collapse;min-width:500px;margin-bottom:6px}\n"
            "th{background:#263950;color:#4fc3f7;padding:7px 14px;text-align:center;white-space:nowrap}\n"
            "td{padding:5px 14px;border-bottom:1px solid #1e3a55;text-align:center;color:#cde}\n"
            "tr:hover td{background:#1a3050}\n"
            ".sec{color:#4fc3f7;font-size:1.05em;font-weight:bold;"
            "border-bottom:2px solid #2e4a6a;padding-bottom:6px;margin:24px 0 14px}\n"
            "</style></head>\n<body>\n"
            f"<h1>\u200b\u200b{product_name} Cross-Program Comparison</h1>\n"
            f"<p class='ts'>Generated: {ts_now}&nbsp;|&nbsp;Latest op per program</p>\n"
            f"<div style='margin-bottom:14px'>{legend}</div>\n"
            "<div class='sec'>Summary&ensp;"
            "<button onclick=\"_csvDl('cmp-all-tbl','NVL_Compare_Programs.csv')\" "
            "title='Download CSV' style='background:none;border:1px solid #4fc3f7;"
            "color:#4fc3f7;border-radius:4px;padding:1px 8px;cursor:pointer;"
            "font-size:0.82em;vertical-align:middle'>&#128190; CSV</button></div>\n"
            "<table id='cmp-all-tbl'><thead><tr>"
            "<th>Program</th><th>Die</th>"
            "<th>FF<br><span style='font-weight:normal;font-size:0.85em'>(1+2)</span></th>"
            "<th>FF+DF<br><span style='font-weight:normal;font-size:0.85em'>(1+2+3+4)</span></th>"
            "<th>UPM<br><span style='font-weight:normal;font-size:0.85em'>(Med %)</span></th>"
            "<th style='color:#5dade2'>DLCP<br><span style='font-weight:normal;font-size:0.85em'>(HP)</span></th>"
            "<th style='color:#f0a500'>DLCP<br><span style='font-weight:normal;font-size:0.85em'>(LP)</span></th>"
            "<th>FB198<br><span style='font-weight:normal;font-size:0.85em'>(Vmin&nbsp;Repair)</span></th>"
            "<th>FB201<br><span style='font-weight:normal;font-size:0.85em'>(Vnom&nbsp;Repair)</span></th>"
            "<th>FB202<br><span style='font-weight:normal;font-size:0.85em'>(Vmax&nbsp;Repair)</span></th>"
            f"</tr></thead><tbody>{table_rows}</tbody></table>\n"
            f"<p class='ts'>FF Target: {ff_tgt_s}&nbsp;|&nbsp;FF+DF Target: {ff_df_tgt_s}</p>\n"
            "<div class='sec'>Visual Comparison</div>\n"
            f"{chart_html}\n"
            "<div class='sec'>Bin Distribution</div>\n"
            f"<div style='display:flex;gap:18px;flex-wrap:wrap'>{bindist_cols}</div>\n"
            + _CSV_DL_SCRIPT +
            "\n</body></html>"
        )
        try:
            out_path = auto_dir / "output" / "compare" / "compare_report_ALL.html"
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_text(html, encoding="utf-8")
            return out_path
        except Exception:
            return None

    def _repair_cells(smry: dict | None) -> str:
        if not smry:
            return "<td>\u2013</td>" * 6  # UPM, HP, LP, FB198, FB201, FB202
        rb      = smry.get("repair_bins", {})
        dlcp    = smry.get("dlcp", {})
        upm_med = smry.get("upm_med")
        upm_cell = f"<td>{upm_med}</td>" if upm_med else "<td>\u2013</td>"
        if not upm_med:  # UPM absent → DLCP unreliable, show – too
            hp_cell = "<td>\u2013</td>"
            lp_cell = "<td>\u2013</td>"
        elif dlcp:
            _span = "<span style='color:#546e7a;font-size:10px'>"
            _hp_ns = f"<br>{_span}({dlcp['hp_n']:,})</span>" if dlcp.get('hp_n') is not None else ""
            _lp_ns = f"<br>{_span}({dlcp['lp_n']:,})</span>" if dlcp.get('lp_n') is not None else ""
            _hp_val = float(dlcp['hp'].rstrip('%')) if dlcp.get('hp') else 0.0
            _hp_col = "#4caf50" if _hp_val >= 30 else "#ef5350"
            hp_cell = f"<td style='color:{_hp_col};font-weight:bold'>{dlcp['hp']}{_hp_ns}</td>"
            lp_cell = f"<td style='color:#f0a500'>{dlcp['lp']}{_lp_ns}</td>"
        else:
            hp_cell = "<td>\u2013</td>"
            lp_cell = "<td>\u2013</td>"  # noqa
        def _rv(n: int) -> str:
            return f"<td>{rb.get(str(n), chr(8211))}</td>"
        vnom = rb.get("201") or rb.get("2") or "\u2013"
        return upm_cell + hp_cell + lp_cell + _rv(198) + f"<td>{vnom}</td>" + _rv(202)

    def _cells(smry: dict | None) -> str:
        if not smry:
            return "<td>–</td>" * 4
        bins = smry.get("bins", {})
        b1 = _parse_pct(bins.get("Bin 1", "0"))
        b2 = _parse_pct(bins.get("Bin 2", "0"))
        b3 = _parse_pct(bins.get("Bin 3", "0"))
        b4 = _parse_pct(bins.get("Bin 4", "0"))
        ff    = b1 + b2
        ff_df = b1 + b2 + b3 + b4
        ff_col    = "#66bb6a" if (_ff_tgt    is None or ff    >= _ff_tgt)    else "#ef5350"
        ff_df_col = "#66bb6a" if (_ff_df_tgt is None or ff_df >= _ff_df_tgt) else "#ef5350"
        ff_tgt_s    = f"{_ff_tgt:.1f}%"    if _ff_tgt    is not None else "–"
        ff_df_tgt_s = f"{_ff_df_tgt:.1f}%" if _ff_df_tgt is not None else "–"
        return (
            f"<td style='color:{ff_col};font-weight:bold'>{ff:.1f}%</td>"
            f"<td style='color:#546e7a'>{ff_tgt_s}</td>"
            f"<td style='color:{ff_df_col};font-weight:bold'>{ff_df:.1f}%</td>"
            f"<td style='color:#546e7a'>{ff_df_tgt_s}</td>"
        )

    # Load excluded_ops so comparison history respects the same filter as email
    _excl_ops: set[str] = set()
    try:
        _ec = json.loads(Path(_EMAIL_CFG).read_text(encoding="utf-8"))
        _excl_ops = {str(o) for o in _ec.get("excluded_ops", [])}
    except Exception:
        pass

    def _title_plot_btn(cp: Path) -> str:
        if cp and cp.exists():
            uri = cp.as_uri()
            return (
                f"<button class='cmp-plot-btn' style='font-size:2em;line-height:1;padding:0 4px' "
                f"onclick=\"window.open('{uri}','_blank','popup,width=1400,height=900')\" "
                f"title='View trend chart'>&#128202;</button>"
            )
        return ""

    _sub = "<span style='font-weight:normal;font-size:0.85em'>"
    hdr = (
        "<th>Die</th>"
        f"<th>FF<br>{_sub}(1+2)</span></th>"
        f"<th>FF&nbsp;Tgt<br>{_sub}(%)</span></th>"
        f"<th>FF+DF<br>{_sub}(1+2+3+4)</span></th>"
        f"<th>FF+DF&nbsp;Tgt<br>{_sub}(%)</span></th>"
    )
    hdr_card1 = hdr + (
        f"<th>UPM<br>{_sub}(Med %)</span></th>"
        f"<th style='color:#5dade2'>DLCP<br>{_sub}(HP)</span></th>"
        f"<th style='color:#f0a500'>DLCP<br>{_sub}(LP)</span></th>"
        f"<th>FB198<br>{_sub}(Vmin&nbsp;Repair)</span></th>"
        f"<th>FB201<br>{_sub}(Vnom&nbsp;Repair)</span></th>"
        f"<th>FB202<br>{_sub}(Vmax&nbsp;Repair)</span></th>"
    )

    _MAX_ROWS = 4
    _HDR_H    = 34   # px
    _ROW_H    = 30   # px

    # ── Card 1: latest op per TP letter (supplement with history for missing letters) ────
    letter_best: dict[str, tuple] = {}   # letter -> (item, hist_date_str|None)
    for letter, entries in sorted_groups:
        fresh = [(op, item) for op, item in entries if not _is_stale(item)]
        _, best = fresh[0] if fresh else entries[0]
        letter_best[letter] = (best, None)

    # Scan history to fill in letters absent from today's run.
    # Prefer the most-recent *tagged* run per letter; fall back to latest.
    try:
        def _folder_ts(d: Path) -> str:
            m = re.search(r'\d{8}_\d{6}', d.name)
            return m.group(0) if m else ''
        _hist_dirs = sorted(
            [d for d in run_dir.parent.iterdir()
             if d.is_dir() and re.match(rf'NVL_[A-Za-z0-9]+{_gen}[A-Za-z]_', d.name) and d != run_dir],
            key=_folder_ts, reverse=True,
        )[:30]
        _hist_latest: dict = {}  # letter -> (item, hdate)  — most recent
        _hist_tagged: dict = {}  # letter -> (item, hdate)  — most recent *tagged*
        for _rf in _hist_dirs:
            _m_rf  = re.search(r"(\d{8})_(\d{6})", _rf.name)
            _hdate = (
                f"{_m_rf.group(1)[:4]}-{_m_rf.group(1)[4:6]}-{_m_rf.group(1)[6:]}"
                if _m_rf else _rf.name
            )
            _rf_tagged = (_rf / ".tag").exists()
            try:
                _td_list = sorted(_rf.iterdir())
            except Exception:
                continue  # skip unreadable folder
            for _td in _td_list:
                if not _td.is_dir() or _td.name.endswith("_R0"):
                    continue
                _ml = re.search(rf'([A-Za-z0-9]+{_gen}[A-Za-z])', _td.name)
                if not _ml:
                    continue
                _let = _ml.group(1).upper()  # full group key e.g. H61G
                if _let in letter_best:
                    continue   # covered by current run
                _op_m = re.search(r"_(\d{5,6})$", _td.name)
                if _op_m and _op_m.group(1) in _excl_ops:
                    continue   # skip excluded op
                _hitem = ((_td.name, True, _td, f"prev: {_hdate}"), _hdate)
                if _let not in _hist_latest:
                    _hist_latest[_let] = _hitem   # newest (first hit desc)
                if _rf_tagged and _let not in _hist_tagged:
                    _hist_tagged[_let] = _hitem   # newest tagged
        # Prefer tagged over latest; latest is the fallback
        for _let, _hitem in _hist_latest.items():
            letter_best[_let] = _hist_tagged.get(_let, _hitem)
    except Exception:
        pass

    rows1 = ""
    row_count1 = 0
    letter_smry_data: list = []
    for letter, (best, hist_date) in sorted(letter_best.items(), reverse=True):
        smry      = _extract_yield_summary(Path(best[2]))
        die       = smry.get("die", "–") if smry else "–"
        # Outlook doesn't support opacity on <tr>; dim history rows by styling
        # each cell's text colour instead.
        _dim_style = " style='color:#78909c'" if hist_date else ""
        note      = f"&nbsp;<span class='ts' style='font-size:0.78em'>({hist_date})</span>" if hist_date else ""
        _idx_p    = Path(best[2]) / "index.html"
        _idx_link = (
            f"&nbsp;<a href='{_idx_p.as_uri()}' style='color:#4fc3f7;font-size:0.82em'"
            f" title='Open Dashboard'>&#128279;</a>"
            if _idx_p.exists() else ""
        )
        # For history rows use plain dimmed cells (override coloured yield cells)
        if hist_date:
            _r_cells = (
                f"<td{_dim_style}>{die}</td>"
                + re.sub(r"style='color:#[0-9a-fA-F]+;?", f"style='color:#78909c;", _cells(smry) + _repair_cells(smry))
            )
        else:
            _r_cells = f"<td>{die}</td>{_cells(smry)}{_repair_cells(smry)}"
        rows1 += (
            f"<tr>"
            f"<td class='cmp-prog'{_dim_style}>{letter}{_idx_link}{note}</td>"
            f"<td class='cmp-op'{_dim_style}>{_op(best)}</td>"
            f"{_r_cells}"
            f"</tr>\n"
        )
        letter_smry_data.append((letter, smry, Path(best[2])))
        row_count1 += 1

    cross_all_path = _write_cross_compare_html(letter_smry_data, _auto_dir) if len(letter_smry_data) > 1 else None
    cross_all_btn  = _title_plot_btn(cross_all_path) if cross_all_path else ""

    scroll1 = (
        f"max-height:{_HDR_H + _MAX_ROWS * _ROW_H}px;overflow-y:auto;"
        if row_count1 > _MAX_ROWS else ""
    )
    card1 = (
        "<div class='cmp-card'>"
        f"<div class='cmp-title'>&#128202;&nbsp;Compare by Test Program &mdash; latest op per program{cross_all_btn}"
        "<button class='cmp-plot-btn' onclick=\"_csvDl('cmp-tbl-1','NVL_Compare_Programs.csv')\" "
        "title='Download table as CSV' style='font-size:1em'>&#128190;</button></div>"
        f"<div style='overflow-x:auto;{scroll1}'><table id='cmp-tbl-1' class='cmp-tbl'><thead>"
        f"<tr><th>Program</th><th>Op</th>{hdr_card1}</tr></thead>"
        f"<tbody>{rows1}</tbody></table></div>"
        "</div>"
    )

    # ── Card 2: current-letter run history ─────────────────────────────────────
    _m_curr_ltr = re.search(rf'NVL_([A-Za-z0-9]+{_gen}[A-Za-z])_', run_dir.name)
    _curr_ltr = _m_curr_ltr.group(1).upper() if _m_curr_ltr else f'H{_gen}C'
    rows2 = ""
    row_count2 = 0
    card2_cp = None
    try:
        output_dir  = run_dir.parent
        def _folder_ts2(d: Path) -> str:
            m = re.search(r'\d{8}_\d{6}', d.name)
            return m.group(0) if m else ''
        # Sort by timestamp descending; deduplicate so old + new style folders
        # for the same program+run don't create duplicate rows.
        # Key includes the program letter so NVL_0H61A_TS and NVL_0H61C_TS
        # from the same automation run are treated as separate entries.
        _rf_all = sorted(
            [d for d in output_dir.iterdir()
             if d.is_dir() and re.match(rf'NVL_[A-Za-z0-9]+{_gen}[A-Za-z]_', d.name)],
            key=_folder_ts2, reverse=True,
        )
        _seen_ts: set[str] = set()
        run_folders: list[Path] = []
        for _rfd in _rf_all:
            _m_ltr = re.search(rf'NVL_([A-Za-z0-9]+{_gen}[A-Za-z])_', _rfd.name)
            _rk = (_m_ltr.group(1) + '_' + _folder_ts2(_rfd)
                   if _m_ltr else _folder_ts2(_rfd))
            if _rk not in _seen_ts:
                _seen_ts.add(_rk)
                run_folders.append(_rfd)
            if len(run_folders) >= 60:
                break
        for rf in run_folders:
            m = re.search(r"(\d{8})_(\d{6})", rf.name)
            date_str = (
                f"{m.group(1)[:4]}-{m.group(1)[4:6]}-{m.group(1)[6:]}"
                f" ({m.group(2)[:2]}:{m.group(2)[2:4]})"
            ) if m else rf.name
            c_dirs = sorted(
                [d for d in rf.iterdir()
                 if d.is_dir() and re.search(re.escape(_curr_ltr), d.name)
                 and not d.name.endswith("_R0")
                 and not any(d.name.endswith(f"_{op}") for op in _excl_ops)],
                key=lambda d: int(re.search(r"_(\d{5,6})$", d.name).group(1))
                              if re.search(r"_(\d{5,6})$", d.name) else 0,
                reverse=True,
            )
            if not c_dirs:
                continue
            c_dir  = c_dirs[0]
            m_op   = re.search(r"_(\d{5,6})$", c_dir.name)
            op_str = m_op.group(1) if m_op else "?"
            smry   = _extract_yield_summary(c_dir)
            die    = smry.get("die", "–") if smry else "–"
            cur    = (rf == run_dir)
            cls    = " class='cmp-current'" if cur else ""
            mark   = " &#9664;&nbsp;current" if cur else ""
            cp     = _auto_dir / f"compare_report_{c_dir.name}.html"
            if card2_cp is None:
                card2_cp = cp
            _idx2   = c_dir / "index.html"
            _idx2_lnk = (
                f"&nbsp;<a href='{_idx2.as_uri()}' style='color:#4fc3f7;font-size:0.82em' title='Open Dashboard'>&#128279;</a>"
                if _idx2.exists() else ""
            )
            rows2 += (
                f"<tr{cls}>"
                f"<td class='cmp-date'>{date_str}{mark}{_idx2_lnk}</td>"
                f"<td class='cmp-op'>{op_str}</td>"
                f"<td>{die}</td>{_cells(smry)}{_repair_cells(smry)}"
                f"</tr>\n"
            )
            row_count2 += 1
    except Exception:
        pass

    if not rows2:
        rows2 = f"<tr><td colspan='99' class='ts' style='padding:8px'>No 61{_curr_ltr} run history found.</td></tr>"

    scroll2 = (
        f"max-height:{_HDR_H + _MAX_ROWS * _ROW_H}px;overflow-y:auto;"
        if row_count2 > _MAX_ROWS else ""
    )
    card2 = (
        "<div class='cmp-card'>"
        f"<div class='cmp-title'>&#128337;&nbsp;61{_curr_ltr} Run History &mdash; newest first{_title_plot_btn(card2_cp)}"
        f"<button class='cmp-plot-btn' onclick=\"_csvDl('cmp-tbl-2','NVL_61{_curr_ltr}_History.csv')\" "
        "title='Download table as CSV' style='font-size:1em'>&#128190;</button></div>"
        f"<div style='overflow-x:auto;{scroll2}'><table id='cmp-tbl-2' class='cmp-tbl'><thead>"
        f"<tr><th>Run Date</th><th>Op</th>{hdr_card1}</tr></thead>"
        f"<tbody>{rows2}</tbody></table></div></div>"
    )

    return (
        "<div class='cmp-section'>"
        "<div class='cmp-section-hdr'>&#9660;&nbsp;Comparison Summary</div>"
        f"{card2}{card1}"
        "</div>"
    )


# Shared CSV-download JS — raw string so \r\n in the JS stays as literal backslash sequences
_CSV_DL_SCRIPT = r"""<script>
function _csvDl(tblId,fname){
  var t=document.getElementById(tblId);if(!t)return;
  var r=[],ths=t.querySelectorAll('thead th'),h=[];
  for(var i=0;i<ths.length;i++)h.push('"'+ths[i].innerText.replace(/[\r\n]+/g,' ').trim()+'"');
  r.push(h.join(','));
  var trs=t.querySelectorAll('tbody tr');
  for(var j=0;j<trs.length;j++){var cells=trs[j].querySelectorAll('td'),ro=[];
    for(var k=0;k<cells.length;k++)ro.push('"'+cells[k].innerText.replace(/[\r\n]+/g,' ').trim()+'"');
    r.push(ro.join(','));}
  var a=document.createElement('a');
  a.href='data:text/csv;charset=utf-8,'+encodeURIComponent(r.join('\r\n'));
  a.download=fname;a.click();
}
</script>"""

# ─────────────────────────────────────────────────────────────────────────────
# Step 3b — Build per-run report.html with embedded BinDistribution
# ─────────────────────────────────────────────────────────────────────────────

def _build_run_report(
    run_dir: Path,
    run_ts: str,
    aqua_file: str,
    tp_results: list[tuple[str, bool, Path]],  # (tp_key, ok, tp_output_dir[, gz_ts])
    out_path: Path | None = None,
    prog_series: str = "0H61",
    prod_cfg: str | None = None,
    product_name: str = "NVL816-BLLC",
) -> Path:
    """
    Write run_dir/report.html.
    - Grouped by program letter as collapsible top-level categories.
    - Within each category, entries sorted by op number descending (newest op first).
    - BinDistribution collapsed by default; inlined via srcdoc so the file is
      self-contained when saved from email to Downloads.
    """
    import html as _html_mod
    from collections import defaultdict

    # ── group by program letter, sort by op descending within each group ──────
    # key e.g. NCXSDJXL0H61C002620_132322  →  letter='C', op='132322'
    _ps_escaped = re.escape(prog_series)
    def _parse(tp_key: str):
        m = re.search(r'([A-Za-z]\d{2}[A-Za-z]).*?_(\d{5,6})$', tp_key)
        return (m.group(1).upper(), m.group(2)) if m else ('?', '0')

    groups: dict[str, list] = defaultdict(list)
    for item in tp_results:           # (tp_key, ok, tp_dir[, gz_ts])
        letter, op = _parse(item[0])
        if letter == '?':
            continue   # skip _R0 dirs and unparseable keys
        groups[letter].append((op, item))

    # Sort groups ascending (80A before 80B before 80C); within each group sort by op descending (newest first)
    sorted_groups = sorted(groups.items())
    for letter, entries in sorted_groups:
        entries.sort(key=lambda x: x[0], reverse=True)

    compare_html = _build_compare_section(sorted_groups, run_dir, prog_series=prog_series,
                                           prod_cfg=prod_cfg, product_name=product_name)

    # ── build HTML ────────────────────────────────────────────────────────────
    categories_html = ""
    _first_card = False  # all BinDist dropdowns start collapsed
    for letter, entries in sorted_groups:
        # display label: digits+suffix only (e.g. "80A" not "H80A")
        _m_disp = re.search(r'(\d+[A-Za-z])$', letter)
        prog_name  = _m_disp.group(1).upper() if _m_disp else letter
        ok_count   = sum(1 for _, item in entries if item[1])
        fail_count = len(entries) - ok_count
        cat_status = (f'<span class="ok">{ok_count} OK</span>'
                      if fail_count == 0 else
                      f'<span class="ok">{ok_count} OK</span>'
                      f'&ensp;<span class="fail">{fail_count} FAILED</span>')

        tp_cards = ""
        for op_num, item in entries:
            tp_key     = item[0]
            ok         = item[1]
            tp_dir     = item[2]
            gz_ts      = item[3] if len(item) > 3 else ""
            r0_label   = item[4] if len(item) > 4 else ""
            r0_dir     = Path(item[5]) if len(item) > 5 and item[5] else None
            is_stale   = gz_ts.startswith("prev:")
            status_cls = "ok" if ok else "fail"
            status_txt = "&#10004; OK" if ok else "&#10008; FAILED"
            if is_stale:
                ts_badge = f'<span class="gz-ts stale-badge">&#128337; {gz_ts}</span>'
            else:
                ts_badge = f'<span class="gz-ts">{gz_ts}</span>' if gz_ts else ""
            r0_badge = f'<span class="r0-badge">&#128204; {r0_label}</span>' if r0_label else ""

            # Links: plain run + R0 run (if available)
            index_path  = tp_dir / "index.html"
            pcm_path    = tp_dir / "pcm_analysis.html"
            compare_path = run_dir.parent / "compare" / f"compare_report_{tp_key}.html"
            links = ""
            if index_path.exists():
                links += f'<a href="{index_path.as_uri()}">Full Dashboard</a>'
            if pcm_path.exists():
                links += f' &nbsp;&nbsp; <a href="{pcm_path.as_uri()}">PCM Analysis</a>'

            if r0_dir and r0_dir.exists():
                r0_index = r0_dir / "index.html"
                r0_pcm   = r0_dir / "pcm_analysis.html"
                if r0_index.exists():
                    links += f'<br><a href="{r0_index.as_uri()}">Full Dashboard (+ R0)</a>'
                if r0_pcm.exists():
                    links += f' &nbsp;&nbsp; <a href="{r0_pcm.as_uri()}">PCM Analysis (+ R0)</a>'

            # BinDistribution: always use plain (non-R0) dir
            _bd_sub = tp_dir / 'bin_dist'
            bd_search_dir = _bd_sub if _bd_sub.exists() else tp_dir

            # Inline BinDistribution via srcdoc — open only for latest series-C, collapsed otherwise
            bd_open = not is_stale and letter[-1] == 'C'
            bindist_block = '<p class="ts">BinDistribution not available.</p>'
            if bd_search_dir.exists():
                bdfiles = sorted(bd_search_dir.glob("*BinDistribution*.html"))
                if bdfiles:
                    try:
                        bd_content = bdfiles[0].read_text(encoding="utf-8", errors="replace")
                        bd_escaped = _html_mod.escape(bd_content, quote=True)
                        bindist_block = (
                            f'<iframe srcdoc="{bd_escaped}" width="100%" height="940"'
                            f' style="border:none;display:block;margin-top:8px;'
                            f'background:#fff;border-radius:4px"></iframe>'
                        )
                    except Exception as _e:
                        bindist_block = f'<p class="ts">BinDistribution read error: {_e}</p>'

            bd_display = "block" if bd_open else "none"
            bd_arrow   = "&#x25BC;" if bd_open else "&#x25B6;"

            tp_cards += f"""
<div class="tp-card">
  <div class="tp-card-hdr">
    <span class="tp-name">{tp_key}</span>
    {ts_badge}
    {r0_badge}
    <span class="{status_cls}">{status_txt}</span>
  </div>
  <p class="links">{links if links else '<span class="ts">no output links</span>'}</p>
  <div class="bd-wrap">
    <div class="bd-hdr" onclick="(function(h){{var b=h.nextElementSibling,a=h.querySelector('.bd-arr');if(b.style.display==='none'){{b.style.display='block';a.textContent='\u25BC';}}else{{b.style.display='none';a.textContent='\u25B6';}}}})(this)">
      <span class="bd-arr">{bd_arrow}</span>&nbsp;Bin Distribution
    </div>
    <div class="bd-body" style="display:{bd_display}">{bindist_block}</div>
  </div>
</div>"""
            _first_card = False

        categories_html += f"""
<details class="prog-group" open>
  <summary class="prog-summary">
    <span class="prog-name">{prog_name}</span>
    <span class="prog-meta">{len(entries)} op(s)&ensp;&bull;&ensp;{cat_status}</span>
  </summary>
  <div class="tp-cards">
    {tp_cards}
  </div>
</details>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="utf-8"><title>{product_name} Yield Report \u2014 {run_ts}</title>{_WATERMARK_CSS}<style>
body {{ font-family: Segoe UI, Arial, sans-serif; background:#1a252f; color:#e8f0f7;
       margin:0; padding:16px 28px 60px; }}
h1   {{ color:#4fc3f7; border-bottom:2px solid #4fc3f7; padding-bottom:8px; margin-bottom:4px; }}
.ts  {{ color:#90a4ae; font-size:0.85em; margin-top:0; }}
.ok  {{ color:#66bb6a; font-size:0.85em; font-weight:bold; }}
.fail{{ color:#ef5350; font-size:0.85em; font-weight:bold; }}
.links {{ margin:2px 0 8px; }}
.links a {{ color:#4fc3f7; text-decoration:none; margin-right:16px; font-size:0.9em; }}
.links a:hover {{ text-decoration:underline; }}

/* ── Program category block ── */
.prog-group {{ border:2px solid #2e4a6a; border-radius:8px; margin-bottom:28px;
               background:#1a2e40; }}
.prog-summary {{ display:flex; align-items:baseline; gap:14px; padding:12px 18px;
                 cursor:pointer; list-style:none; user-select:none; }}
.prog-summary::-webkit-details-marker {{ display:none; }}
.prog-summary::before {{ content:"\\25B6  "; color:#4fc3f7; font-size:0.75em; }}
details[open] > .prog-summary::before {{ content:"\\25BC  "; }}
.prog-name {{ color:#4fc3f7; font-size:1.15em; font-weight:bold; }}
.prog-meta {{ color:#90a4ae; font-size:0.85em; }}

/* ── Per-op cards inside a category ── */
.tp-cards {{ padding:0 14px 14px; display:flex; flex-direction:column; gap:14px; }}
.tp-card  {{ border:1px solid #263950; border-radius:6px; padding:12px 16px 14px;
             background:#1e2e3d; }}
.tp-card-hdr {{ display:flex; align-items:center; gap:12px; margin-bottom:6px; flex-wrap:wrap; }}
.tp-name {{ color:#80cbc4; font-size:0.88em; font-weight:bold; font-family:monospace; word-break:break-all; }}
.gz-ts   {{ color:#90a4ae; font-size:0.78em; white-space:nowrap; }}
.stale-badge {{ color:#ffa726; font-size:0.78em; white-space:nowrap; font-style:italic; }}
.r0-badge    {{ color:#ce93d8; font-size:0.78em; white-space:nowrap; font-weight:bold; }}

/* ── BinDist toggle (open by default; click header to collapse) ── */
.bd-wrap {{ margin-top:6px; }}
.bd-hdr  {{ cursor:pointer; color:#4fc3f7; font-size:0.88em; padding:4px 0;
            user-select:none; display:flex; align-items:center; gap:4px; }}
.bd-hdr:hover {{ color:#80cbc4; }}
.bd-arr  {{ font-size:0.72em; }}
.bd-body {{ margin-top:4px; }}
.bd-link {{ color:#4fc3f7; font-size:0.82em; text-decoration:none; display:block;
            margin-bottom:4px; }}
.bd-link:hover {{ text-decoration:underline; }}
/* ── BinDist toggle (collapsed by default) ── */
details > summary {{ cursor:pointer; color:#4fc3f7; font-size:0.88em;
                     padding:4px 0; list-style:none; user-select:none; }}
details > summary::-webkit-details-marker {{ display:none; }}
details > summary::before {{ content:"\\25B6  "; font-size:0.75em; }}
details[open] > summary::before {{ content:"\\25BC  "; }}
/* ── Comparison cards ── */
.cmp-section {{ margin-bottom:32px; }}
.cmp-section-hdr {{ color:#4fc3f7; font-size:1.05em; font-weight:bold;
                    padding:8px 0 10px; border-bottom:2px solid #2e4a6a; margin-bottom:14px; }}
.cmp-card {{ background:#1e2e3d; border:1px solid #263950; border-radius:8px;
             padding:14px 18px; margin-bottom:14px; }}
.cmp-title {{ color:#80cbc4; font-size:0.88em; font-weight:bold; margin-bottom:8px; display:flex; align-items:center; gap:8px; }}
.cmp-tbl {{ border-collapse:collapse; font-size:0.95em; min-width:400px; }}
.cmp-tbl th {{ background:#263950; color:#4fc3f7; padding:5px 12px; text-align:center;
               white-space:nowrap; }}
.cmp-tbl td {{ padding:4px 12px; border-bottom:1px solid #1e3a55; text-align:center; color:#cde; }}
.cmp-tbl tr:hover td {{ background:#1a3050; }}
.cmp-prog {{ color:#80cbc4; font-weight:bold; font-family:monospace; text-align:left!important; }}
.cmp-op   {{ color:#90a4ae; font-family:monospace; }}
.cmp-date {{ color:#90a4ae; text-align:left!important; white-space:nowrap; }}
tr.cmp-current td {{ background:#1d3a52!important; }}
tr.cmp-current .cmp-date {{ color:#ffa726; font-weight:bold; }}
.cmp-plot-btn {{ background:none; border:none; cursor:pointer; font-size:1.1em;
                 padding:1px 4px; border-radius:4px; line-height:1;
                 transition:background 0.15s; }}
.cmp-plot-btn:hover {{ background:#1d4060; }}
</style>
</head><body>
{_WATERMARK_HTML}
<h1>{product_name} Yield Report</h1>
<p class="ts">Run: <b>{run_ts}</b>&ensp;&bull;&ensp;AQUA: {Path(aqua_file).name}&ensp;&bull;&ensp;{len(tp_results)} program(s) ({sum(1 for t in tp_results if not (len(t)>3 and str(t[3]).startswith('prev:')))} updated, {sum(1 for t in tp_results if len(t)>3 and str(t[3]).startswith('prev:'))} from previous runs)</p>
{compare_html}
{categories_html}
{_CSV_DL_SCRIPT}
</body></html>
"""
    report_path = out_path or (run_dir / "report.html")
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(html, encoding="utf-8")
    _log(f"Report: {report_path}  ({report_path.stat().st_size:,} bytes)")
    return report_path


# ─────────────────────────────────────────────────────────────────────────────
# Step 4 — Update run_log.html
# ─────────────────────────────────────────────────────────────────────────────

_RUN_LOG_CSS = """
<style>
  body { font-family: Segoe UI, Arial, sans-serif; background:#1a252f; color:#e8f0f7; margin:24px; }
  h1   { color:#4fc3f7; border-bottom:2px solid #4fc3f7; padding-bottom:8px; }
  h2   { color:#80cbc4; margin-top:32px; }
  table{ border-collapse:collapse; width:100%; margin-top:8px; }
  th   { background:#263950; color:#4fc3f7; padding:8px 12px; text-align:left; }
  td   { padding:6px 12px; border-bottom:1px solid #263950; }
  tr:hover td { background:#1e3044; }
  .ok  { color:#66bb6a; font-weight:bold; }
  .fail{ color:#ef5350; font-weight:bold; }
  .ts  { color:#90a4ae; font-size:0.85em; }
  a    { color:#4fc3f7; text-decoration:none; }
  a:hover { text-decoration:underline; }
</style>
"""

_RUN_LOG_HEADER = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Yield Dashboard — Run Log</title>
{css}
</head>
<body>
<h1>Yield Dashboard — Automation Run Log</h1>
<p class="ts">Auto-generated by run_automation.py &nbsp;|&nbsp;
Updated: <span id="ts">{ts}</span></p>
<!-- RUNS -->
""".format(css=_RUN_LOG_CSS, ts=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

_RUN_LOG_FOOTER = "\n</body>\n</html>\n"


def _make_run_section(
    run_ts: str,
    aqua_file: str,
    results: list[tuple[str, bool, str]],   # (op_code, ok, output_dir)
    report_path: Path | None = None,
) -> str:
    rows_html = ""
    for r in results:
        op_code, ok, output_dir = r[0], r[1], r[2]
        index_html = Path(output_dir) / "index.html"
        link   = f'<a href="{index_html.as_uri()}">{op_code}</a>' if index_html.exists() else op_code
        status = '<span class="ok">&#10004; OK</span>' if ok else '<span class="fail">&#10008; FAILED</span>'
        rows_html += f"<tr><td>{link}</td><td>{status}</td><td class='ts'>{output_dir}</td></tr>\n"

    report_link = ""
    if report_path and report_path.exists():
        report_link = f' &nbsp;|&nbsp; <a href="{report_path.as_uri()}">&#128196; Report</a>'

    ops_str = ", ".join(r[0] for r in results)
    return f"""
<h2>Run: {run_ts} &mdash; op(s) updated: {ops_str}</h2>
<p class="ts">AQUA: {Path(aqua_file).name}{report_link}</p>
<table>
  <tr><th>Operation</th><th>Status</th><th>Output</th></tr>
  {rows_html}
</table>
"""


def update_run_log(
    results: list[tuple[str, bool, str]],
    aqua_file: str,
    run_log: Path,
    dry_run: bool,
    report_path: Path | None = None,
) -> None:
    run_ts  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    section = _make_run_section(run_ts, aqua_file, results, report_path=report_path)

    if dry_run:
        _log(f"DRY-RUN: would prepend to {run_log}")
        return

    run_log.parent.mkdir(parents=True, exist_ok=True)

    if run_log.exists():
        existing = run_log.read_text(encoding="utf-8")
        # Prepend: insert latest section right after <!-- RUNS --> so newest is on top
        if "<!-- RUNS -->" in existing:
            updated = existing.replace("<!-- RUNS -->", "<!-- RUNS -->\n" + section, 1)
        elif "</body>" in existing:
            # Fallback for older log files without the marker
            updated = existing.replace("</body>", section + "\n</body>", 1)
        else:
            updated = existing + section
    else:
        updated = _RUN_LOG_HEADER + section + _RUN_LOG_FOOTER

    run_log.write_text(updated, encoding="utf-8")
    _log(f"Run log updated: {run_log}")


# ─────────────────────────────────────────────────────────────────────────────
# Step 5 — Send email
# ─────────────────────────────────────────────────────────────────────────────

def _email_body_html(
    run_ts: str,
    aqua_file: str,
    results: list[tuple],
    run_log: Path,
    attachments: list[str] | None = None,
) -> str:
    # Sort: 61C first (letter descending), then by op number descending within each letter
    def _email_sort_key(r):
        m = re.search(r'61([A-Za-z]).*?_(\d{5,6})$', r[0])
        letter = m.group(1).upper() if m else '?'
        op     = int(m.group(2)) if m else 0
        return (letter, op)   # both descending → negate with reverse=True
    sorted_results = sorted(results, key=_email_sort_key, reverse=True)

    rows = ""
    for r in sorted_results:
        run_tag, ok, output_dir = r[0], r[1], r[2]
        r0_dir    = r[3] if len(r) > 3 else ""
        stale_lbl = r[4] if len(r) > 4 else ""   # "prev: 2026-05-17" when not run today
        if stale_lbl:
            status = f"&#8212; {stale_lbl}"
            color  = "#90a4ae"
            row_bg = "background:#f5f5f5;"
        else:
            status = "✔ OK" if ok else "✖ FAILED"
            color  = "#66bb6a" if ok else "#ef5350"
            row_bg = ""
        index  = Path(output_dir) / "index.html"
        link   = f'<a href="{index.as_uri()}">{run_tag}</a>' if index.exists() else run_tag
        rows += (
            f"<tr style='{row_bg}'><td>{link}</td>"
            f"<td style='color:{color};font-weight:bold'>{status}</td>"
            f"<td style='color:#555'>{output_dir}</td></tr>\n"
        )
        if r0_dir:
            r0_index = Path(r0_dir) / "index.html"
            r0_link  = f'<a href="{r0_index.as_uri()}">{run_tag}_R0</a>' if r0_index.exists() else f'{run_tag}_R0'
            rows += (
                f"<tr style='background:#f0f7ff'>"
                f"<td style='padding-left:24px;color:#0071c5'>&#8627; {r0_link}</td>"
                f"<td style='color:#0071c5;font-weight:bold'>+ R0</td>"
                f"<td style='color:#0071c5'>{r0_dir}</td></tr>\n"
            )

    overall = "OK" if all(r[1] for r in results) else "FAILED"

    att_note = ""
    if attachments:
        links = "".join(
            f" &nbsp;&middot;&nbsp; <b>{Path(a).name}</b>"
            for a in attachments if Path(a).exists()
        )
        att_note = (f'<p style="background:#f0f7ff;padding:8px;border-left:4px solid #0071c5">'
                    f'<b>Attachment:</b>{links}</p>')

    return f"""
<html><body style="font-family:Segoe UI,Arial;color:#222;max-width:720px">
<h2 style="color:#0071c5;margin-bottom:4px">Yield Dashboard — {overall}</h2>
<p style="color:#555;font-size:0.9em;margin-top:0">{run_ts}</p>
{att_note}
<table border="1" cellpadding="6" cellspacing="0"
       style="border-collapse:collapse;width:100%;font-size:0.9em">
  <tr style="background:#0071c5;color:#fff">
    <th>Run</th><th>Status</th><th>Output folder</th>
  </tr>
  {rows}
</table>
<p style="color:#888;font-size:0.8em;margin-top:12px">
  AQUA: {aqua_file}<br>
  Full history: <a href="{run_log.as_uri()}">run_log.html</a> (attached)
</p>
</body></html>
"""


def _build_email_report_html(output_dir: Path, run_ts: str,
                             excluded_keys: list | None = None,
                             prog_series: str = "0H61",
                             product_name: str = "NVL816-BLLC") -> str:
    """Build self-contained HTML with sidebar tabs."""
    from collections import defaultdict

    _excluded = set(excluded_keys or [])
    _series_digits = re.search(r'(\d+)$', prog_series)
    _gen = _series_digits.group(1) if _series_digits else "61"

    run_pattern = re.compile(rf'^NVL_([A-Za-z0-9]+{_gen}[A-Za-z])_(\d{{8}}_\d{{6}})$')
    # Match both '0H61H' and 'XH61H' style TP key prefixes (some steppings differ in leading char)
    tp_pattern  = re.compile(rf'[A-Za-z0-9]H{_gen}([A-Za-z]).*?_(\d{{5,6}})$')
    history: dict[str, list[dict]] = defaultdict(list)
    _run_count: dict[str, int] = {}   # prog_key → runs already scanned
    _MAX_HISTORY = 10                 # scan at most 10 most-recent runs per program

    for rd in sorted(output_dir.iterdir(), reverse=True):   # newest first
        if not rd.is_dir():
            continue
        m = run_pattern.match(rd.name)
        if not m:
            continue
        letter, ts = m.group(1).upper(), m.group(2)
        # display label: digits+suffix only (e.g. "80A" not "0H80A")
        prog_key = re.search(r'(\d+[A-Z])$', letter).group(1) if re.search(r'(\d+[A-Z])$', letter) else letter
        if _run_count.get(prog_key, 0) >= _MAX_HISTORY:
            continue
        _run_count[prog_key] = _run_count.get(prog_key, 0) + 1
        dt_str = f"{ts[:4]}-{ts[4:6]}-{ts[6:8]} {ts[9:11]}:{ts[11:13]}"
        for tp_dir in sorted(rd.iterdir()):
            if not tp_dir.is_dir():
                continue
            if tp_dir.name in _excluded:
                continue
            tm = tp_pattern.search(tp_dir.name)
            if not tm or tm.group(1).upper() != letter[-1] or tp_dir.name.endswith("_R0"):
                continue
            history[prog_key].append({
                "ts":      ts,
                "dt_str":  dt_str,
                "op":      tm.group(2),
                "tp_key":  tp_dir.name,
                "tp_dir":  tp_dir,
                "summary": _extract_yield_summary(tp_dir),
            })

    for prog_key in history:
        # newest first so the latest run appears at the top of the table
        history[prog_key].sort(key=lambda x: x["ts"], reverse=True)
    # sort descending: newest generation/letter first (e.g. 80H before 80A before 61H)
    sorted_letters = sorted(history.keys(), key=lambda k: (int(k[:-1]) if k[:-1].isdigit() else 0, k[-1]), reverse=True)

    # ── helpers ───────────────────────────────────────────────────────────────
    def _pct_val(s):
        try:
            return float((s or "").rstrip('%'))
        except Exception:
            return 0.0

    def _sum_bins(sm, nums):
        if not sm:
            return "\u2013"
        bins = sm.get("bins", {})
        total = sum(_pct_val(bins.get(f"Bin {n}")) for n in nums)
        return f"{total:.1f}%" if total > 0 else "\u2013"

    def _yld_color(v_str):
        v = _pct_val(v_str)
        if v <= 0:
            return "#90a4ae"
        return "#66bb6a" if v >= 60 else "#ffa726" if v >= 40 else "#ef5350"

    def _get(sm, key, sub=""):
        if not sm:
            return "\u2013"
        if sub:
            return sm.get(key, {}).get(sub, "\u2013") or "\u2013"
        return sm.get(key, "\u2013") or "\u2013"

    def _idx_uri(entry):
        idx = entry["tp_dir"] / "index.html"
        return idx.as_uri() if idx.exists() else ""

    COL_HDR = (
        "<th>Material</th>"
        "<th>Run Date</th>"
        "<th>Op</th>"
        "<th>Die</th>"
        "<th>FF<br><small>(1+2)</small></th>"
        "<th>FF Tgt<br><small>(%)</small></th>"
        "<th>FF+DF<br><small>(1+2+3+4)</small></th>"
        "<th>FF+DF Tgt<br><small>(%)</small></th>"
        "<th>UPM<br><small>(Med %)</small></th>"
        "<th>DLCP<br><small>(HP)</small></th>"
        "<th>DLCP<br><small>(LP)</small></th>"
        "<th>FB198<br><small>(Vmin Repair)</small></th>"
        "<th>FB201<br><small>(Vnom Repair)</small></th>"
        "<th>FB202<br><small>(Vmax Repair)</small></th>"
    )

    def _data_row(entry, is_latest=False, prog_prefix="", material="", sub=False):
        sm    = entry["summary"]
        ff    = _sum_bins(sm, [1, 2])
        ffdf  = _sum_bins(sm, [1, 2, 3, 4])
        upm   = _get(sm, "upm_med")
        hp    = _get(sm, "dlcp", "hp")
        lp    = _get(sm, "dlcp", "lp")
        fb198 = _get(sm, "repair_bins", "198")
        fb201 = _get(sm, "repair_bins", "201")
        fb202 = _get(sm, "repair_bins", "202")
        die   = _get(sm, "die")
        link  = _idx_uri(entry)
        ff_tgt   = _get(sm, "ff_tgt")
        ffdf_tgt = _get(sm, "ffdf_tgt")
        date_cell = (f'<a href="{link}" class="tl">{entry["dt_str"]}</a>'
                     if link else entry["dt_str"])
        if is_latest and not sub:
            date_cell += ' <span class="latest-badge">latest</span>'
        if sub:
            row_cls = ' class="mat-sub-row"'
            mat_label = f'<span class="mat-sub-lbl">&#8627; {material}</span>'
        elif is_latest:
            row_cls = ' class="latest-row"'
            mat_label = material
        else:
            row_cls = ""
            mat_label = material
        mat_cell = f'<td class="c-mat mono">{mat_label}</td>'
        return (
            f'<tr{row_cls}>'
            f'{prog_prefix}'
            f'{mat_cell}'
            f'<td class="c-date">{date_cell}</td>'
            f'<td class="c-op mono">{entry["op"]}</td>'
            f'<td class="c-num">{die}</td>'
            f'<td class="c-num" style="color:{_yld_color(ff)};font-weight:bold">{ff}</td>'
            f'<td class="c-tgt">{ff_tgt}</td>'
            f'<td class="c-num" style="color:{_yld_color(ffdf)};font-weight:bold">{ffdf}</td>'
            f'<td class="c-tgt">{ffdf_tgt}</td>'
            f'<td class="c-num">{upm}</td>'
            f'<td class="c-num">{hp}</td>'
            f'<td class="c-num">{lp}</td>'
            f'<td class="c-num">{fb198}</td>'
            f'<td class="c-num">{fb201}</td>'
            f'<td class="c-num">{fb202}</td>'
            f'</tr>\n'
        )

    def _material_rows_for_entry(entry, is_latest=False, prog_prefix=""):
        """Emit one aggregate row (ALL materials combined) then one indented
        sub-row per material type (only when >1 material type is present)."""
        rows_html = _data_row(entry, is_latest=is_latest,
                              prog_prefix=prog_prefix, material="ALL")
        mat_summaries = _extract_per_material_summaries(entry["tp_dir"])
        for mat_id, mat_sm in mat_summaries:
            mat_entry = dict(entry)
            mat_entry["summary"] = mat_sm
            rows_html += _data_row(
                mat_entry, is_latest=False,
                prog_prefix=f'<td class="c-prog c-prog-sub"></td>' if prog_prefix else "",
                material=mat_id, sub=True,
            )
        return rows_html


    # ── Summary panel ─────────────────────────────────────────────────────────
    sum_rows = ""
    for letter in sorted_letters:
        if not history[letter]:
            continue
        e    = history[letter][0]   # first entry is the latest run (sorted newest-first)
        link = _idx_uri(e)
        prog_cell = (
            f'<td class="c-prog"><a href="{link}" class="tl">'
            f'<span class="prog-pill">{letter}</span></a></td>'
            if link else
            f'<td class="c-prog"><span class="prog-pill">{letter}</span></td>'
        )
        sum_rows += _material_rows_for_entry(e, is_latest=True, prog_prefix=prog_cell)

    # find most-recent group_compare.html across all run folders for this product
    _gc_files = sorted(
        (f for f in output_dir.rglob('group_compare.html') if f.is_file()),
        key=lambda f: f.stat().st_mtime, reverse=True,
    )
    _gc_link = ''
    if _gc_files:
        _gc = _gc_files[0]
        _gc_unc = str(_gc).replace('/', '\\')
        _gc_link = (
            f'<p class="panel-sub">'
            f'&#128202; <a href="{_gc.as_uri()}" style="color:#4fc3f7">Group Compare (cross-letter)</a>'
            f' &nbsp;<span style="color:#546e7a;font-size:12px">{_gc_unc}</span>'
            f'</p>\n'
        )

    summary_panel = (
        f'<div id="panel-summary" class="panel active">\n'
        f'  <h2 class="panel-hdr">&#128200; Summary \u2014 Latest Run per Program'
        f'    <button class="csv-btn" onclick="downloadCSV(this)" title="Download visible rows as CSV">&#11123; CSV</button>'
        f'  </h2>\n'
        f'  <p class="panel-sub">Generated: {run_ts}</p>\n'
        f'  {_gc_link}'
        f'  <div class="tbl-wrap">\n'
        f'  <table class="data-tbl">\n'
        f'    <thead><tr><th>Program</th>{COL_HDR}</tr></thead>\n'
        f'    <tbody>{sum_rows}</tbody>\n'
        f'  </table>\n'
        f'  </div>\n'
        f'</div>'
    )

    # ── Per-program panels ────────────────────────────────────────────────────
    prog_panels = ""
    for letter in sorted_letters:
        entries = history[letter]
        if not entries:
            continue
        def _prog_cell(e, ltr):
            link = _idx_uri(e)
            pill = f'<span class="prog-pill">{ltr}</span>'
            return (
                f'<td class="c-prog"><a href="{link}" class="tl">{pill}</a></td>'
                if link else f'<td class="c-prog">{pill}</td>'
            )
        hist_rows = "".join(
            _material_rows_for_entry(e, i == 0, prog_prefix=_prog_cell(e, letter))
            for i, e in enumerate(entries)
        )
        latest_ff = _sum_bins(entries[0]["summary"], [1, 2])
        try:
            lyf = float(latest_ff.rstrip('%'))
            badge_col = "#66bb6a" if lyf >= 60 else "#ffa726" if lyf >= 40 else "#ef5350"
        except Exception:
            badge_col = "#90a4ae"
        prog_panels += (
            f'<div id="panel-{letter}" class="panel">\n'
            f'  <h2 class="panel-hdr">\n'
            f'    <span class="prog-pill">{letter}</span>\n'
            f'    <span class="yld-badge" style="background:{badge_col}">{latest_ff} FF</span>\n'
            f'    <span class="panel-sub-inline">{len(entries)} run{"s" if len(entries)!=1 else ""}</span>\n'
            f'    <button class="csv-btn" onclick="downloadCSV(this)" title="Download visible rows as CSV">&#11123; CSV</button>\n'
            f'  </h2>\n'
            f'  <div class="tbl-wrap">\n'
            f'  <table class="data-tbl">\n'
            f'    <thead><tr><th>Program</th>{COL_HDR}</tr></thead>\n'
            f'    <tbody>{hist_rows}</tbody>\n'
            f'  </table>\n'
            f'  </div>\n'
            f'</div>\n'
        )

    # ── Sidebar ───────────────────────────────────────────────────────────────
    sb = '<li><button class="tab-btn active" data-panel="summary">&#128200;&nbsp;Summary</button></li>\n'
    for letter in sorted_letters:
        if not history[letter]:
            continue
        ff = _sum_bins(history[letter][-1]["summary"], [1, 2])
        n  = len(history[letter])
        sb += (
            f'<li><button class="tab-btn" data-panel="{letter}">'
            f'<span class="nav-prog">{letter}</span>'
            f'<span class="nav-meta">{n} run{"s" if n!=1 else ""} &bull; FF: {ff}</span>'
            f'</button></li>\n'
        )

    CSS = """
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: Segoe UI, Arial, sans-serif; font-size: 15px;
  background: #0f1923; color: #dce9f5; display: flex; min-height: 100vh;
}
#sidebar {
  width: 240px; flex-shrink: 0; background: #141f2b;
  border-right: 1px solid #1e3048; position: sticky; top: 0;
  height: 100vh; overflow-y: auto; padding: 0 0 24px;
}
#sb-hdr { background: #0f1923; border-bottom: 1px solid #1e3048; padding: 14px 16px 10px; }
#sb-hdr h3 { color: #4fc3f7; font-size: 13px; text-transform: uppercase; letter-spacing: 1px; }
#sb-hdr p  { color: #546e7a; font-size: 12px; margin-top: 3px; }
#sidebar ul { list-style: none; padding: 6px 0; }
#sidebar li { margin: 0; }
.tab-btn {
  width: 100%; background: none; border: none; border-left: 3px solid transparent;
  cursor: pointer; text-align: left; padding: 10px 14px; color: #78909c;
  display: flex; flex-direction: column; gap: 3px; font-size: 15px;
  transition: background .15s, color .15s;
}
.tab-btn:hover { background: #1a2f45; color: #dce9f5; border-left-color: #546e7a; }
.tab-btn.active { background: #1a3a55; color: #4fc3f7; border-left-color: #4fc3f7; font-weight: bold; }
.nav-prog { font-size: 15px; }
.nav-meta { font-size: 12px; color: #546e7a; }
.tab-btn.active .nav-meta { color: #607d8b; }
#main { flex: 1; padding: 22px 28px 60px; overflow-x: auto; min-width: 0; }
.panel { display: none; }
.panel.active { display: block; }
.panel-hdr {
  font-size: 18px; color: #4fc3f7;
  display: flex; align-items: center; gap: 10px; flex-wrap: wrap;
  padding-bottom: 8px; border-bottom: 2px solid #1e3048; margin-bottom: 6px;
}
.panel-sub { color: #546e7a; font-size: 13px; margin-bottom: 14px; margin-top: 4px; }
.panel-sub-inline { color: #546e7a; font-size: 13px; font-weight: normal; }
.prog-pill { background: #1a3a55; color: #80cbc4; border-radius: 4px; padding: 2px 10px; font-family: monospace; font-size: 16px; }
.yld-badge { color: #fff; border-radius: 12px; padding: 2px 10px; font-size: 14px; font-weight: bold; }
.latest-badge { background: #4fc3f7; color: #0f1923; border-radius: 8px; padding: 1px 6px; font-size: 11px; font-weight: bold; margin-left: 4px; vertical-align: middle; }
.tbl-wrap { overflow-x: auto; margin-top: 6px; }
.data-tbl { border-collapse: collapse; width: 100%; font-size: 14px; min-width: 900px; }
.data-tbl th {
  background: #1a3a55; color: #4fc3f7;
  padding: 8px 12px; text-align: left;
  border-bottom: 2px solid #0f1923; font-size: 13px; line-height: 1.4; white-space: nowrap;
}
.data-tbl th small { color: #607d8b; font-weight: normal; display: block; font-size: 12px; }
.data-tbl td { padding: 6px 12px; border-bottom: 1px solid #1a2f45; vertical-align: middle; text-align: left; }
.data-tbl tr:hover td { background: #14253a; }
.latest-row td { background: #0f2233 !important; }
.c-date { white-space: nowrap; color: #90a4ae; }
.c-prog { }
.c-prog-sub { }
.c-mat  { white-space: nowrap; color: #ce93d8; font-size: 13px; }
.mat-sub-row td { background: #111e2a !important; font-size: 13px; color: #a5d6e8; }
.mat-sub-row .c-num { color: #80cbc4; }
.mat-sub-row .c-tgt { color: #7986cb; }
.mat-sub-row .c-date { color: #607d8b; }
.mat-sub-row .c-op   { color: #4db6ac; }
.mat-sub-row .c-mat { padding-left: 6px; }
.mat-sub-lbl { color: #80deea; font-size: 13px; font-style: italic; font-weight: 500; }
.c-op   { white-space: nowrap; color: #80cbc4; }
.c-num  { white-space: nowrap; }
.c-tgt  { color: #78909c; font-size: 13px; }
.mono   { font-family: monospace; font-size: 13px; }
.tl     { color: #4fc3f7; text-decoration: none; }
.tl:hover { text-decoration: underline; }
.sort-arrow { font-size: 11px; color: #4fc3f7; margin-left: 3px; }
/* ── Column filter dropdown ──────────────────────────────────────────────── */
.flt-btn {
  display: inline-block; margin-left: 5px; cursor: pointer;
  font-size: 10px; color: #607d8b; vertical-align: middle;
  padding: 0 3px; border-radius: 3px; line-height: 1;
  transition: color .15s;
}
.flt-btn:hover { color: #4fc3f7; }
.flt-btn.flt-active { color: #ffa726; }
.flt-drop {
  position: absolute; z-index: 9999;
  background: #1a2f45; border: 1px solid #263950;
  border-radius: 6px; padding: 8px 0 6px;
  box-shadow: 0 4px 18px rgba(0,0,0,.6);
  min-width: 220px; max-width: 320px;
}
.flt-search-row { padding: 0 10px 6px; }
.flt-text {
  width: 100%; box-sizing: border-box;
  background: #0f1923; border: 1px solid #263950;
  color: #dce9f5; border-radius: 4px; padding: 5px 8px;
  font-size: 13px; outline: none;
}
.flt-text:focus { border-color: #4fc3f7; }
.flt-cb-list {
  max-height: 180px; overflow-y: auto; padding: 0 10px;
  border-top: 1px solid #263950; border-bottom: 1px solid #263950;
  margin-bottom: 4px;
}
.flt-cb-lbl {
  display: flex; align-items: center; gap: 6px;
  padding: 4px 2px; font-size: 13px; color: #c8d8e8;
  cursor: pointer; white-space: nowrap;
}
.flt-cb-lbl:hover { color: #fff; }
.flt-cb-lbl input[type=checkbox] { accent-color: #4fc3f7; cursor: pointer; }
.flt-footer { display: flex; gap: 6px; padding: 4px 10px 0; }
.flt-footer button {
  flex: 1; background: #263950; border: none; color: #90a4ae;
  border-radius: 4px; padding: 4px 0; font-size: 12px; cursor: pointer;
}
.flt-footer button:hover { background: #2e4a6a; color: #dce9f5; }
.flt-footer .flt-apply { background: #1b5e20; color: #a5d6a7; }
.flt-footer .flt-apply:hover { background: #2e7d32; }
.flt-footer .flt-clear { color: #ef9a9a; }
/* ── CSV download button ─────────────────────────────────────────────────── */
.csv-btn {
  margin-left: auto; background: #1a3a2c; border: 1px solid #2e6b4a;
  color: #80deea; border-radius: 5px; padding: 4px 12px;
  font-size: 13px; cursor: pointer; white-space: nowrap;
  transition: background .15s, color .15s;
}
.csv-btn:hover { background: #235c40; color: #b2ebf2; }
"""

    JS = """
(function(){
  /* ── Tab navigation ───────────────────────────────────────────────────── */
  var btns   = document.querySelectorAll('.tab-btn');
  var panels = document.querySelectorAll('.panel');
  btns.forEach(function(b){
    b.addEventListener('click', function(){
      btns.forEach(function(x){x.classList.remove('active');});
      panels.forEach(function(p){p.classList.remove('active');});
      b.classList.add('active');
      var p = document.getElementById('panel-'+b.dataset.panel);
      if(p) p.classList.add('active');
    });
  });

  /* ── Column sort + filter for .data-tbl ──────────────────────────────── */
  var _sortState = {};   /* key: tableId+colIdx -> {asc:bool} */
  var _filterState = {}; /* key: tableId+colIdx -> {text:'', checked:Set} */
  var _activeDropdown = null;

  function tableId(tbl){ return tbl.id || (tbl.id='tbl'+(Math.random()*1e9|0)); }

  /* Build a unique key per table+column */
  function fkey(tbl,ci){ return tableId(tbl)+':'+ci; }

  /* Collect unique values for a column (ignores hidden rows) */
  function colValues(tbl,ci){
    var vals=new Set();
    Array.from(tbl.querySelectorAll('tbody tr')).forEach(function(r){
      var c=r.cells[ci]; if(c) vals.add(c.innerText.trim());
    });
    return Array.from(vals).sort();
  }

  /* Apply all active filters to a table */
  function applyFilters(tbl){
    var tid=tableId(tbl);
    Array.from(tbl.querySelectorAll('tbody tr')).forEach(function(row){
      var show=true;
      Object.keys(_filterState).forEach(function(k){
        if(k.split(':')[0]!==tid) return;
        var ci=parseInt(k.split(':')[1]);
        var st=_filterState[k];
        var cell=row.cells[ci];
        var val=cell ? cell.innerText.trim() : '';
        if(st.text && val.toLowerCase().indexOf(st.text.toLowerCase())===-1){ show=false; }
        if(st.checked && st.checked.size>0 && !st.checked.has(val)){ show=false; }
      });
      row.style.display = show ? '' : 'none';
    });
  }

  /* Close any open dropdown */
  function closeDropdown(){
    if(_activeDropdown){ _activeDropdown.remove(); _activeDropdown=null; }
  }
  document.addEventListener('click', function(e){
    if(_activeDropdown && !_activeDropdown.contains(e.target) && !e.target.classList.contains('flt-btn')){
      closeDropdown();
    }
  });

  /* Build and show the filter dropdown for a th */
  function showDropdown(th, tbl, ci){
    closeDropdown();
    var k=fkey(tbl,ci);
    if(!_filterState[k]) _filterState[k]={text:'',checked:new Set()};
    var st=_filterState[k];

    var dd=document.createElement('div');
    dd.className='flt-drop';
    dd.innerHTML=
      '<div class="flt-search-row">'+
        '<input class="flt-text" type="text" placeholder="Search..." value="'+st.text+'">'+
      '</div>'+
      '<div class="flt-cb-list"></div>'+
      '<div class="flt-footer">'+
        '<button class="flt-all">All</button>'+
        '<button class="flt-none">None</button>'+
        '<button class="flt-apply">Apply</button>'+
        '<button class="flt-clear">Clear</button>'+
      '</div>';

    /* Position below th */
    var rect=th.getBoundingClientRect();
    dd.style.top=(rect.bottom+window.scrollY)+'px';
    dd.style.left=(rect.left+window.scrollX)+'px';
    document.body.appendChild(dd);
    _activeDropdown=dd;

    /* Populate checkboxes */
    var cbList=dd.querySelector('.flt-cb-list');
    var vals=colValues(tbl,ci);
    vals.forEach(function(v){
      var lbl=document.createElement('label');
      lbl.className='flt-cb-lbl';
      var chk=document.createElement('input');
      chk.type='checkbox';
      chk.value=v;
      chk.checked = st.checked.size===0 || st.checked.has(v);
      lbl.appendChild(chk);
      lbl.appendChild(document.createTextNode(' '+v));
      cbList.appendChild(lbl);
    });

    /* Text filter live preview */
    dd.querySelector('.flt-text').addEventListener('input',function(){
      var q=this.value.toLowerCase();
      cbList.querySelectorAll('label').forEach(function(l){
        l.style.display=l.textContent.toLowerCase().indexOf(q)>=0?'':'none';
      });
    });

    dd.querySelector('.flt-all').addEventListener('click',function(e){
      e.stopPropagation();
      cbList.querySelectorAll('input').forEach(function(c){c.checked=true;});
    });
    dd.querySelector('.flt-none').addEventListener('click',function(e){
      e.stopPropagation();
      cbList.querySelectorAll('input').forEach(function(c){c.checked=false;});
    });
    dd.querySelector('.flt-apply').addEventListener('click',function(e){
      e.stopPropagation();
      st.text=dd.querySelector('.flt-text').value.trim();
      st.checked=new Set();
      var all=cbList.querySelectorAll('input');
      var anyUnchecked=false;
      all.forEach(function(c){ if(!c.checked) anyUnchecked=true; });
      if(anyUnchecked) all.forEach(function(c){ if(c.checked) st.checked.add(c.value); });
      applyFilters(tbl);
      /* Update filter-active indicator */
      var active=(st.text||st.checked.size>0);
      th.querySelector('.flt-btn').classList.toggle('flt-active',active);
      closeDropdown();
    });
    dd.querySelector('.flt-clear').addEventListener('click',function(e){
      e.stopPropagation();
      st.text=''; st.checked=new Set();
      applyFilters(tbl);
      th.querySelector('.flt-btn').classList.remove('flt-active');
      closeDropdown();
    });
  }

  /* Attach sort + filter controls to each filterable column header */
  function initTable(tbl){
    /* cols 0=Program, 1=Material (0-indexed) */
    var filterCols=[0,1];
    var ths=Array.from(tbl.querySelectorAll('thead th'));
    ths.forEach(function(th,ci){
      /* Sort on th text click */
      th.style.cursor='pointer';
      th.title='Click to sort';
      th.addEventListener('click',function(e){
        if(e.target.classList.contains('flt-btn')) return;
        var tid=tableId(tbl);
        var k=tid+':sort:'+ci;
        _sortState[k]=!_sortState[k];
        var asc=!_sortState[k];
        var rows=Array.from(tbl.querySelectorAll('tbody tr'));
        rows.sort(function(a,b){
          var av=(a.cells[ci]||{}).innerText||'';
          var bv=(b.cells[ci]||{}).innerText||'';
          return asc ? av.localeCompare(bv) : bv.localeCompare(av);
        });
        var tbody=tbl.querySelector('tbody');
        rows.forEach(function(r){ tbody.appendChild(r); });
        /* Update sort arrows */
        ths.forEach(function(h){
          var arrow=h.querySelector('.sort-arrow');
          if(arrow) arrow.textContent='';
        });
        var arrow=th.querySelector('.sort-arrow');
        if(arrow) arrow.textContent=asc?' ↑':' ↓';
      });

      /* Sort-arrow span */
      var arrowSpan=document.createElement('span');
      arrowSpan.className='sort-arrow';
      th.appendChild(arrowSpan);

      /* Filter button for Program and Material only */
      if(filterCols.indexOf(ci)>=0){
        var btn=document.createElement('span');
        btn.className='flt-btn';
        btn.title='Filter';
        btn.innerHTML='&#9660;';
        btn.addEventListener('click',function(e){
          e.stopPropagation();
          showDropdown(th,tbl,ci);
        });
        th.appendChild(btn);
      }
    });
  }

  /* Init all current tables */
  document.querySelectorAll('.data-tbl').forEach(initTable);

  /* Re-init when tab switches (panels reuse same DOM so only once needed) */

  /* ── CSV download ─────────────────────────────────────────────────────── */
  window.downloadCSV = function(btn){
    var panel = btn.closest('.panel');
    var tbl   = panel && panel.querySelector('.data-tbl');
    if(!tbl) return;

    /* Headers — strip HTML, sort-arrow spans and filter buttons */
    var headers = Array.from(tbl.querySelectorAll('thead th')).map(function(th){
      return th.cloneNode(true).innerText.replace(/[\u2191\u2193\u25bc\u25be]/g,'').trim();
    });

    /* Visible rows only */
    var rows = Array.from(tbl.querySelectorAll('tbody tr')).filter(function(r){
      return r.style.display !== 'none';
    });

    var LF = String.fromCharCode(10), CR = String.fromCharCode(13);
    function escCSV(v){
      v = v.split(LF).join(' ').split(CR).join('').trim();
      if(v.indexOf(',')>=0 || v.indexOf('"')>=0)
        return '"'+v.split('"').join('""')+'"';
      return v;
    }

    var NL = CR+LF;
    var lines = [headers.map(escCSV).join(',')];
    rows.forEach(function(r){
      lines.push(Array.from(r.cells).map(function(c){
        return escCSV(c.innerText.trim());
      }).join(','));
    });

    /* Build panel name for filename */
    var panelId = panel.id || 'report';
    var ts = new Date().toISOString().replace(/[T:]/g,'-').slice(0,19);
    var filename = 'NVL_Yield_'+panelId+'_'+ts+'.csv';

    var blob = new Blob([lines.join(NL)], {type:'text/csv;charset=utf-8;'});
    var url  = URL.createObjectURL(blob);
    var a    = document.createElement('a');
    a.href     = url;
    a.download = filename;
    document.body.appendChild(a);
    a.click();
    setTimeout(function(){ document.body.removeChild(a); URL.revokeObjectURL(url); }, 500);
  };
})();
"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{product_name} Yield Report — {run_ts}</title><style>{CSS}</style>
</head>
<body>
<nav id="sidebar">
  <div id="sb-hdr">
    <h3>{product_name} Yield</h3>
    <p>{run_ts}</p>
  </div>
  <ul>{sb}</ul>
</nav>
<div id="main">
  {summary_panel}
  {prog_panels}
</div>
<script>{JS}</script>
</body>
</html>"""


def _collapse_report_html(html: str) -> str:
    """Return a copy of report HTML suitable for Outlook email body.
    - Converts <details>/<summary> to plain <div> elements (Outlook strips them)
    - Removes the BinDist section entirely (iframes not supported in Outlook)
    """
    # Strip entire bd-wrap blocks (iframe content can't render in Outlook)
    html = re.sub(r'<div class="bd-wrap">.*?</div>\s*</div>', '', html, flags=re.DOTALL)
    # Convert <details>/<summary> to plain divs — Outlook (Word engine) strips
    # these tags and may discard their entire content, causing visible truncation.
    html = re.sub(r'<details\b[^>]*>', '<div class="prog-group">', html)
    html = re.sub(r'</details>', '</div>', html)
    html = re.sub(r'<summary\b[^>]*>', '<div class="prog-summary">', html)
    html = re.sub(r'</summary>', '</div>', html)
    return html


_SMTP_SERVER  = "smtpauth.intel.com"
_SMTP_PORT    = 587
_SMTP_FROM    = "sujit.n.pant@intel.com"


def _send_via_outlook(to: str, subject: str, body_html: str,
                      attachments: list[str]) -> None:
    """Send via Outlook COM (requires Outlook running in user session)."""
    import win32com.client as _w
    _ol = _w.Dispatch("Outlook.Application")
    _m  = _ol.CreateItem(0)
    _m.To       = to
    _m.Subject  = subject
    _m.HTMLBody = body_html
    for att in attachments:
        if Path(att).exists():
            _m.Attachments.Add(att)
            _log(f"  Attaching : {Path(att).name}")
    try:
        _m.Send()
    except Exception as _send_err:
        # Outlook sometimes raises a COM error after the item is already
        # dispatched (e.g. "The operation failed").  Treat as sent.
        _log(f"  Outlook COM: Send() raised {_send_err!r} — email likely dispatched.")
    _log("  Email sent via Outlook COM.")


def _send_via_smtp(to: str, subject: str, body_html: str,
                   attachments: list[str]) -> None:
    """Send via Intel SMTP relay — works without Outlook (scheduled tasks)."""
    import smtplib
    import time
    import os
    import socket
    from email.mime.multipart import MIMEMultipart
    from email.mime.text      import MIMEText
    from email.mime.base      import MIMEBase
    from email                import encoders

    msg = MIMEMultipart("mixed")
    msg["From"]    = _SMTP_FROM
    msg["To"]      = to
    msg["Subject"] = subject
    msg.attach(MIMEText(body_html, "html", "utf-8"))

    for att in attachments:
        p = Path(att)
        if p.exists():
            part = MIMEBase("application", "octet-stream")
            part.set_payload(p.read_bytes())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment",
                            filename=p.name)
            msg.attach(part)
            _log(f"  Attaching : {p.name}")

    recipients = [a.strip() for a in to.split(";")]
    msg_str = msg.as_string()

    # Get proxy from environment or use Intel DMZ proxy
    proxy = os.getenv("HTTP_PROXY") or os.getenv("HTTPS_PROXY") or "http://proxy-dmz.intel.com:912"
    
    # Retry with exponential backoff (account for network issues / proxy)
    max_retries = 3
    base_delay = 2  # seconds
    for attempt in range(1, max_retries + 1):
        try:
            _log(f"  SMTP attempt {attempt}/{max_retries} via proxy {proxy}...")
            
            # Try direct connection first
            try:
                with smtplib.SMTP(_SMTP_SERVER, _SMTP_PORT, timeout=60) as s:
                    s.starttls()
                    s.sendmail(_SMTP_FROM, recipients, msg_str)
                _log(f"  Email sent via SMTP ({_SMTP_SERVER}) — direct.")
                return
            except (smtplib.SMTPException, OSError, TimeoutError) as direct_err:
                # If direct fails, try via proxy
                _log(f"  Direct connection failed ({direct_err}), trying via proxy...")
                
                try:
                    import socks
                    # Parse proxy URL
                    if proxy.startswith("http://"):
                        proxy_addr = proxy[7:]
                    else:
                        proxy_addr = proxy
                    proxy_host, proxy_port = proxy_addr.rsplit(":", 1)
                    proxy_port = int(proxy_port)
                    
                    # Create SOCKS5 proxy socket tunnel for HTTP CONNECT
                    sock = socks.socksocket(socket.AF_INET, socket.SOCK_STREAM)
                    sock.set_proxy(socks.HTTP, proxy_host, proxy_port)
                    sock.connect((_SMTP_SERVER, _SMTP_PORT))
                    
                    with smtplib.SMTP(sock=sock, timeout=60) as s:
                        s.starttls()
                        s.sendmail(_SMTP_FROM, recipients, msg_str)
                    _log(f"  Email sent via SMTP ({_SMTP_SERVER}) — via proxy.")
                    return
                except ImportError:
                    _log("  PySocks not available, retrying direct connection...")
                    raise direct_err
        except (smtplib.SMTPException, OSError, TimeoutError) as e:
            if attempt < max_retries:
                delay = base_delay * (2 ** (attempt - 1))
                _log(f"  SMTP attempt {attempt} failed: {e}")
                _log(f"  Retrying in {delay}s...")
                time.sleep(delay)
            else:
                _log(f"  SMTP all {max_retries} attempts failed: {e}")
                raise


def send_email(
    to: str,
    subject: str,
    body_html: str,
    dry_run: bool,
    attachments: list[str] | None = None,
) -> None:
    _log(f"{'DRY-RUN: ' if dry_run else ''}Sending email → {to}")
    if dry_run:
        _log(f"  Subject   : {subject}")
        for a in (attachments or []):
            _log(f"  Attach    : {a}")
        return

    atts = attachments or []
    try:
        _send_via_outlook(to, subject, body_html, atts)
        return
    except ImportError:
        _log("  win32com not available — falling back to SMTP.")
    except Exception as e:
        _log(f"  Outlook COM failed ({e}) — falling back to SMTP.")

    try:
        _send_via_smtp(to, subject, body_html, atts)
    except Exception as e:
        _log(f"  ERROR sending email via SMTP: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# No-new-data helper
# ─────────────────────────────────────────────────────────────────────────────

def _send_no_new_data_email(base_dir: Path, args, product_name: str = "NVL816-BLLC") -> None:
    """Send a brief daily 'no new data' email so the user always hears back."""
    ecfg_path = _EMAIL_CFG
    ecfg: dict = {}
    if ecfg_path.exists():
        try:
            ecfg = json.loads(ecfg_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    to = (ecfg.get("email_to_report")
          or ecfg.get("email_to")
          or getattr(args, "email", _EMAIL_TO)
          or _EMAIL_TO)

    # Find the most recent report from a previous run
    last_report_link = ""
    out_dir = base_dir / "output"
    if out_dir.exists():
        runs = sorted(out_dir.iterdir(), reverse=True)
        for r in runs:
            rpt = r / "report.html"
            if rpt.exists():
                last_report_link = (
                    f'<p>Last report: <a href="{rpt}">{rpt.name}</a> '
                    f'(from run <code>{r.name}</code>)</p>'
                )
                break

    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    body = f"""<!DOCTYPE html><html><body style="font-family:sans-serif">
<h2 style="color:#6c3483">NVL Yield Automation — No New Data</h2>
<p>Run at <strong>{run_ts}</strong>: AQUA pull completed but no new lot/wafer data
was detected since the last run. Pipelines were not re-executed.</p>
{last_report_link}
<hr/><p style="font-size:0.85em;color:#888">Pant, Sujit N — GEMS FTE</p>
</body></html>"""

    subject = f"{product_name} Yield Dashboard"
    send_email(to=to, subject=subject, body_html=body,
               dry_run=getattr(args, "dry_run", False))


# ─────────────────────────────────────────────────────────────────────────────
# Output cleanup
# ─────────────────────────────────────────────────────────────────────────────

def cleanup_old_runs(output_dir: Path, keep_runs: int, dry_run: bool = False, prog_series: str = "0H61") -> int:
    """Delete old output run folders, keeping the *keep_runs* most-recent per
    program variant (H61G, M61H, …).

    Rules:
      - Folders are grouped by program variant (NVL_H61G_* → group H61G).
      - Within each group, folders are sorted newest-first by name (date-encoded).
      - The *keep_runs* most-recent folders are kept.
      - Folders that contain a ``.tag`` file are **always** preserved (not counted
        against keep_runs).
      - Returns the number of folders actually deleted (0 in dry-run mode).
    """
    if keep_runs <= 0 or not output_dir.exists():
        return 0

    _gd = re.search(r'(\d+)$', prog_series)
    _gen = _gd.group(1) if _gd else '61'
    pattern = re.compile(rf'^NVL_([A-Za-z0-9]+{re.escape(_gen)}[A-Za-z])_', re.IGNORECASE)
    letter_groups: dict[str, list[Path]] = {}
    for d in output_dir.iterdir():
        if not d.is_dir():
            continue
        m = pattern.match(d.name)
        if m:
            letter = m.group(1).upper()
            letter_groups.setdefault(letter, []).append(d)

    deleted = 0
    for letter in sorted(letter_groups):
        # Newest run first (folder names are NVL_H61G_YYYYMMDD_HHMMSS)
        folders = sorted(letter_groups[letter], key=lambda d: d.name, reverse=True)
        kept = 0
        for d in folders:
            is_tagged = (d / ".tag").exists()
            if is_tagged:
                continue          # tagged → always preserved, don't count
            if kept < keep_runs:
                kept += 1
                continue          # within retention quota → keep
            # Beyond quota → schedule for deletion
            if dry_run:
                _log(f"  CLEANUP DRY-RUN: would delete {d.name}")
            else:
                try:
                    shutil.rmtree(d)
                    _log(f"  Cleanup: deleted old run {d.name}")
                    deleted += 1
                except Exception as e:
                    _log(f"  WARNING: cleanup could not delete {d.name}: {e}")

    return deleted


def _preview_cleanup(output_dir: Path, keep_runs: int, prog_series: str = "0H61") -> list[Path]:
    """Return the list of run folders that *would* be deleted by cleanup_old_runs.
    Tagged folders are excluded from the result (they are never deleted).
    """
    if keep_runs <= 0 or not output_dir.exists():
        return []

    _gd2 = re.search(r'(\d+)$', prog_series)
    _gen2 = _gd2.group(1) if _gd2 else '61'
    pattern = re.compile(rf'^NVL_([A-Za-z0-9]+{re.escape(_gen2)}[A-Za-z])_', re.IGNORECASE)
    letter_groups: dict[str, list[Path]] = {}
    for d in output_dir.iterdir():
        if not d.is_dir():
            continue
        m = pattern.match(d.name)
        if m:
            letter = m.group(1).upper()
            letter_groups.setdefault(letter, []).append(d)

    to_delete: list[Path] = []
    for letter in sorted(letter_groups):
        folders = sorted(letter_groups[letter], key=lambda d: d.name, reverse=True)
        kept = 0
        for d in folders:
            if (d / ".tag").exists():
                continue
            if kept < keep_runs:
                kept += 1
                continue
            to_delete.append(d)

    return to_delete


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Auto-pull AQUA + split by program/op + run yield dashboards + email.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    ap.add_argument("--aqua-exe",      default=_AQUA_EXE_AMR,
                    help="Path to AquaCmdLine.exe")
    ap.add_argument("--report-config", default=str(_AQUA_CFG),
                    help="AQUA .txt config (program + days baked in)")
    ap.add_argument("--base-dir",      default=None,
                    help="Root output directory (samba share); defaults to per-product base_dir in yield_setup_config.json")
    ap.add_argument("--days",          type=int, default=_DEFAULT_DAYS,
                    help="Look-back days")
    ap.add_argument("--local-csv",     default=None,
                    help="Skip AQUA pull; use this existing CSV/gz/zip (glob ok)")
    ap.add_argument("--use-cached",    action="store_true",
                    help="Skip AQUA pull; use most-recent raw_*.7z from data/programs/*/")
    ap.add_argument("--keys",          default=None,
                    help="Comma-separated substrings to filter TP keys (e.g. '0H61C,119325'). Only matching keys run.")
    ap.add_argument("--force",         action="store_true",
                    help="Rerun all ops even if no new data")
    ap.add_argument("--dry-run",       action="store_true",
                    help="Show plan without executing")
    ap.add_argument("--email",         default=_EMAIL_TO)
    ap.add_argument("--product-name",   default="NVL816-BLLC",
                    help="Product label used in email subjects and report filenames")
    ap.add_argument("--program-series",  default="H61",
                    help="Program series code in TP keys/folder names (e.g. H61 or H80)")
    ap.add_argument("--keep-runs",     type=int, default=None, metavar="N",
                    help="Keep the N most-recent output run folders per program letter "
                         "after this run; older folders are deleted automatically. "
                         "0 = disabled. Reads from email_config.json (keep_runs) "
                         "when not set; default in config is 5.")
    ap.add_argument("--serve",         action="store_true",
                    help="Start local resend server (localhost:17450) and block. "
                         "Enables the 'Resend Email' button in BinDistribution.html.")
    ap.add_argument("--port",          type=int, default=17450,
                    help="Port for --serve mode (default: 17450)")
    args = ap.parse_args()

    # ── --serve mode: just start the server and block ─────────────────────────
    if args.serve:
        base_dir = Path(args.base_dir)
        _run_resend_server(base_dir, port=args.port, email_to=args.email, prog_series=_prog_series,
                           product_name=args.product_name)
        return

    _product_name  = args.product_name
    _prog_series   = args.program_series  # e.g. '0H61' or '0H80'
    _series_digits = re.search(r'(\d+)$', _prog_series)
    _series_code   = f"H{_series_digits.group(1)}" if _series_digits else _prog_series

    # Resolve base_dir: CLI arg → product config → module default
    _setup_cfg: dict = {}
    _prod_sect: dict = {}
    try:
        _setup_cfg = json.loads(_EMAIL_CFG.read_text(encoding="utf-8"))
        _prod_sect = _setup_cfg.get("products", {}).get(_product_name, {})
    except Exception:
        pass

    if args.base_dir:
        base_dir = Path(args.base_dir)
    elif _prod_sect.get("base_dir"):
        base_dir = Path(_prod_sect["base_dir"])
    else:
        base_dir = _BASE_DIR

    # Config always wins over the argparse default; CLI explicit value wins over config
    if _prod_sect.get("program_series"):
        _prog_series = _prod_sect["program_series"]
    if args.program_series != "H61":   # user explicitly passed --program-series
        _prog_series = args.program_series

    # Load per-product tp_folder from yield_setup_config.json early
    _tp_folder = _prod_sect.get("tp_folder") or _TP_FOLDER

    data_dir = base_dir / "data"
    run_log  = base_dir / "run_log.html"
    ts       = datetime.now().strftime('%Y%m%d_%H%M%S')

    _log("=" * 65)
    _log(f"run_automation  [{'DRY-RUN' if args.dry_run else 'LIVE'}]")
    _log(f"Base dir     : {base_dir}")
    _log(f"Pipeline     : {_PIPELINE}")
    _log("=" * 65)

    # ── 1. Get AQUA data ──────────────────────────────────────────────────────
    _local_7z_tmpdir = None   # kept alive so extracted CSV path stays valid
    if args.local_csv:
        import glob as _glob
        _matches = sorted(_glob.glob(args.local_csv), key=os.path.getmtime)
        if _matches:
            aqua_file = Path(_matches[-1])
            _log(f"Local CSV: {aqua_file}  ({len(_matches)} match(es) for {args.local_csv!r})")
        elif '*' in args.local_csv or '?' in args.local_csv:
            _log(f"ERROR: no files matched glob: {args.local_csv!r}")
            sys.exit(1)
        else:
            aqua_file = Path(args.local_csv)
            _log(f"Local CSV: {aqua_file}")
        # If the local file is a .7z archive, extract it to a temp dir first so
        # aqua_file points to the real CSV/gz for timestamp parsing and downstream use.
        if aqua_file.suffix.lower() == '.7z':
            import tempfile as _tempfile2
            import subprocess as _sp_7z
            _local_7z_tmpdir = _tempfile2.TemporaryDirectory(prefix='yield_auto_7z_')
            _7z_out = Path(_local_7z_tmpdir.name)
            _log(f"  Extracting {aqua_file.name} → {_7z_out}")
            try:
                _sp_7z.run([str(_7Z_EXE), 'e', str(aqua_file), f'-o{_7z_out}', '-y'],
                           check=True, capture_output=True)
            except Exception as _e7z:
                _log(f"  ERROR extracting {aqua_file.name}: {_e7z}")
                sys.exit(1)
            _extracted = None
            for _pat7 in ('*.csv.gz', '*.csv'):
                _hits7 = sorted(_7z_out.glob(_pat7), key=lambda p: p.stat().st_size, reverse=True)
                if _hits7:
                    _extracted = _hits7[0]
                    break
            if _extracted is None:
                _log(f"  ERROR: no CSV/gz found inside {aqua_file.name}")
                sys.exit(1)
            _log(f"  Extracted: {_extracted.name}  ({_extracted.stat().st_size:,} bytes)")
            aqua_file = _extracted
    else:
        # ── Use most-recent cached raw_*.7z from data/programs/*/ ──────────
        if args.use_cached or not args.report_config:
            _cached_7z: Path | None = None
            _prog_dir = data_dir / "programs"
            if _prog_dir.exists():
                _candidates = sorted(
                    _prog_dir.glob("*/raw_*.7z"),
                    key=lambda p: p.stat().st_mtime, reverse=True,
                )
                if not _candidates:
                    _candidates = sorted(
                        _prog_dir.glob("*/raw_*.csv.gz"),
                        key=lambda p: p.stat().st_mtime, reverse=True,
                    )
                if _candidates:
                    _cached_7z = _candidates[0]
            if _cached_7z:
                _log(f"Using cached file: {_cached_7z}")
                args.local_csv = str(_cached_7z)
                # Re-run the local-csv branch by recursing into it inline
                import glob as _glob2
                _matches2 = sorted(_glob2.glob(args.local_csv), key=os.path.getmtime)
                aqua_file = Path(_matches2[-1]) if _matches2 else Path(args.local_csv)
                if aqua_file.suffix.lower() == '.7z':
                    import tempfile as _tempfile3
                    _local_7z_tmpdir2 = _tempfile3.TemporaryDirectory(prefix='yield_auto_7z2_')
                    _7z_out2 = Path(_local_7z_tmpdir2.name)
                    _sp_7z2 = __import__('subprocess')
                    _sp_7z2.run([str(_7Z_EXE), 'e', str(aqua_file), f'-o{_7z_out2}', '-y'],
                                check=True, capture_output=True)
                    for _pat8 in ('*.csv.gz', '*.csv'):
                        _hits8 = sorted(_7z_out2.glob(_pat8), key=lambda p: p.stat().st_size, reverse=True)
                        if _hits8:
                            aqua_file = _hits8[0]
                            break
            else:
                _log("ERROR: --use-cached specified but no raw_*.7z found in data/programs/*/")
                sys.exit(1)
        else:
            aqua_file = pull_aqua(
                aqua_exe=args.aqua_exe,
                report_config=Path(args.report_config),
                data_dir=data_dir,
                dry_run=args.dry_run,
            )
            if aqua_file is None:
                _log("AQUA pull failed — aborting.")
                # Read email_config so failure alert goes to the right recipient
                _ecfg = {}
                _ecfg_path = _EMAIL_CFG
                if _ecfg_path.exists():
                    try:
                        _ecfg = json.loads(_ecfg_path.read_text(encoding="utf-8"))
                    except Exception:
                        pass
                _err_to = _ecfg.get("email_to_alert",
                            _ecfg.get("email_to_report",
                              _ecfg.get("email_to", args.email))) or args.email
                send_email(
                    to=_err_to,
                    subject=f"{_product_name} Yield Dashboard",
                    body_html="<p>AQUA pull failed. Check automation logs.</p>",
                    dry_run=args.dry_run,
                )
                sys.exit(1)

    # ── 2. Split by (TestProgram, Operation) and update per-TP gzs ──────────
    _log(f"\nReading: {aqua_file}")
    new_rows, _ = _read_aqua_file(aqua_file)
    _log(f"  {len(new_rows):,} rows")

    groups = split_by_tp_oper(new_rows)
    if not groups and not args.dry_run:
        _log("No groups found — nothing to run.")
        sys.exit(0)

    # ── 2a. Distribute raw AQUA rows into per-program-letter folders ──────────
    # Write a dated per-letter raw snapshot BEFORE any processing so that each
    # program's input is independently recoverable regardless of how many programs
    # (1, 2, 3, 4 …) are present in the AQUA pull.
    #   data/programs/H61A/raw_YYYYMMDD_HHMMSS.csv.gz
    #   data/programs/H61B/raw_YYYYMMDD_HHMMSS.csv.gz   …etc.
    _ts_match = re.search(r'(\d{8}_\d{6})', Path(aqua_file).stem)
    _raw_ts   = _ts_match.group(1) if _ts_match else datetime.now().strftime('%Y%m%d_%H%M%S')

    # Collect rows + union-headers per program letter
    _letter_rows: dict[str, tuple[list[dict], list[str]]] = {}
    for _key, (_krows, _khdrs) in groups.items():
        # Canonicalize per-program folder names to HxxA (e.g. H80A),
        # even when configured series is 0H80.
        _m = re.search(rf'[A-Za-z0-9]?H{re.escape(_series_code[1:])}([A-Za-z])', _key)
        _letter = f"{_series_code}{_m.group(1).upper()}" if _m else f"{_series_code}X"
        if _letter not in _letter_rows:
            _letter_rows[_letter] = ([], list(_khdrs))
        _lrows, _lhdrs = _letter_rows[_letter]
        _lrows.extend(_krows)
        for _h in _khdrs:
            if _h not in _lhdrs:
                _lhdrs.append(_h)

    _log("\nDistributing raw AQUA data to per-program folders…")
    for _letter, (_lrows, _lhdrs) in sorted(_letter_rows.items()):
        _letter_dir = data_dir / "programs" / _letter
        _raw_dest   = _letter_dir / f"raw_{_raw_ts}.csv.gz"
        _raw_z7 = _letter_dir / f"raw_{_raw_ts}.7z"
        if not args.dry_run:
            _letter_dir.mkdir(parents=True, exist_ok=True)
            if _raw_dest.exists() or _raw_z7.exists():
                _log(f"  {_letter}/raw_{_raw_ts}.*  already exists — skipping")
            else:
                _write_gz(_lrows, _lhdrs, _raw_dest)
                _log(f"  {_letter}/raw_{_raw_ts}.csv.gz  ({len(_lrows):,} rows, {_raw_dest.stat().st_size:,} bytes)")
                _z7 = _compress_aqua_to_7z(_raw_dest)
                if _z7:
                    _log(f"    → compressed: {_z7.name}  ({_z7.stat().st_size:,} bytes)")
        else:
            _log(f"  DRY-RUN: would write {_letter}/raw_{_raw_ts}.7z ({len(_lrows):,} rows)")

    # Remove the combined raw file from data/ now that per-program slices are in place.
    # Only removes files that live directly in data_dir (auto-pull location).
    if not args.dry_run:
        try:
            _af = Path(aqua_file)
            if _af.exists() and _af.parent.resolve() == data_dir.resolve():
                _af.unlink()
                _log(f"  Removed combined raw file: {_af.name}")
        except Exception as _de:
            _log(f"\nWARNING: could not remove combined raw file: {_de}")

    # ── 3. Build list of TP keys to run ──────────────────────────────────────
    prog_dir = data_dir / "programs"
    # No persistent per-TP gz files; all_stored_keys only used for stale-TP detection (section 4b).
    all_stored_keys: list[str] = []
    # Always run all programs from the current AQUA pull
    keys_to_run = sorted(groups.keys())

    # ── Excluded ops (email_config.json → excluded_ops): skip execution entirely ──
    _excl_ops: set[str] = set()
    try:
        _ec = json.loads(_EMAIL_CFG.read_text(encoding="utf-8"))
        _excl_ops = {str(o) for o in _ec.get("excluded_ops", [])}
    except Exception:
        pass
    if _excl_ops:
        _before = len(keys_to_run)
        keys_to_run = [
            k for k in keys_to_run
            if not any(k.endswith(f"_{op}") for op in _excl_ops)
        ]
        _skipped = _before - len(keys_to_run)
        if _skipped:
            _log(f"  excluded_ops {sorted(_excl_ops)} → skipped {_skipped} key(s)")

    # ── Optional key filter (--keys) ─────────────────────────────────────────
    if args.keys:
        _kf = [s.strip() for s in args.keys.split(',') if s.strip()]
        keys_to_run = [k for k in keys_to_run if any(f in k for f in _kf)]
        _log(f"  --keys filter '{args.keys}' → {len(keys_to_run)} key(s)")

    _log(f"\nTP programs to run ({len(keys_to_run)}): {keys_to_run or '(none)'}")

    if not keys_to_run:
        _log("Nothing to run — sending no-new-data email and exiting.")
        _send_no_new_data_email(base_dir, args, product_name=_product_name)
        sys.exit(0)

    # ── 4. Per-program-letter run folders ────────────────────────────────────
    #  Group tp_keys by 0H61X letter so each letter gets its own run folder,
    #  its own report.html and its own run-log entry:
    #    base_dir/output/NVL_0H61{letter}_{ts}/    ← one folder per letter
    #      NCXSDJXL0H61{letter}XXXXXX_NNNNNN/      ← one subfolder per TP-op
    #      report.html
    #      input_{tp_key}.json
    #    base_dir/Dashboard_{tp_key}.html           ← top-level latest pointer
    _letter_groups: dict[str, list[str]] = {}
    for _k in sorted(keys_to_run):
        _m = re.search(r'([A-Za-z]\d{2}[A-Za-z])', _k)
        _letter_groups.setdefault(_m.group(1).upper() if _m else '?', []).append(_k)
    _log(f"\nProgram groups: {list(_letter_groups.keys())} ({len(_letter_groups)} run folder(s))")

    all_results:       list[tuple[str, bool, str]] = []
    all_tp_outputs:    list[tuple] = []
    letter_report_paths: list[Path] = []

    env      = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUNBUFFERED": "1"}
    # Extract DevRevStep from first available CSV row (column has op-suffix like DevRevStep_119325)
    _drs = ""
    for _grp_rows, _ in groups.values():
        if _grp_rows:
            _first = _grp_rows[0]
            _drs_col = next((h for h in _first if re.match(r'DevRevStep', h, re.IGNORECASE)), None)
            if _drs_col:
                _drs = str(_first.get(_drs_col, "")).strip()
            break
    prod_cfg = _find_product_config(_drs)
    if _drs:
        _log(f"  DevRevStep: {_drs}  →  {Path(prod_cfg).name if prod_cfg else '(none)'}")
    # Prefer testprogram_folder from product config JSON over yield_setup_config.json
    if prod_cfg:
        try:
            _pcfg_data = json.loads(Path(prod_cfg).read_text(encoding="utf-8"))
            _tp_from_pcfg = str(_pcfg_data.get("testprogram_folder", "")).strip()
            if _tp_from_pcfg:
                _tp_folder = _tp_from_pcfg
        except Exception:
            pass
    run_dir  = base_dir / "output" / f"NVL_unknown_{ts}"   # fallback; overwritten per-group

    for _letter, _letter_keys in sorted(_letter_groups.items(), reverse=True):
        run_dir = base_dir / "output" / f"NVL_{_letter}_{ts}"
        _log(f"\n{'='*65}")
        _log(f"=== Program {_letter}  ({len(_letter_keys)} TP(s))  →  {run_dir.name} ===")

        results:    list[tuple[str, bool, str]] = []
        tp_outputs: list[tuple] = []

        # Merge all TPs for this letter into one combined dataset so the pipeline
        # runs once with a single combined CSV (e.g. L0 + L5 both containing H61E).
        # The L0 key (matching '0H61') is used as the primary identifier.
        _primary_key = next(
            (k for k in sorted(_letter_keys) if re.search(re.escape(_prog_series), k)),
            sorted(_letter_keys)[0]
        )
        if len(_letter_keys) > 1:
            _comb_rows: list[dict] = []
            _comb_hdrs: list[str] = []
            for _k in sorted(_letter_keys):
                if _k in groups:
                    _k_rows, _k_hdrs = groups[_k]
                    _comb_rows.extend(_k_rows)
                    for _h in _k_hdrs:
                        if _h not in _comb_hdrs:
                            _comb_hdrs.append(_h)
            _log(f"  Merging {len(_letter_keys)} TPs → combined: {len(_comb_rows):,} rows  (primary: {_primary_key})")
            groups[_primary_key] = (_comb_rows, _comb_hdrs)
        _exec_keys = [_primary_key]

        for tp_key in _exec_keys:
            _m_key  = re.search(r'([A-Za-z]\d{2}[A-Za-z])', tp_key)
            _sub    = _m_key.group(1).upper() if _m_key else "0H61X"
            # Write temp gz from in-memory data (deleted after pipeline; raw_<ts>.7z is the archival copy)
            _tp_letter_dir = prog_dir / _sub
            if not args.dry_run:
                _tp_letter_dir.mkdir(parents=True, exist_ok=True)
            gz_path = _tp_letter_dir / f"tmp_{tp_key}.csv.gz"
            if tp_key in groups and not args.dry_run:
                _tp_rows, _tp_hdrs = groups[tp_key]
                _write_gz(_tp_rows, _tp_hdrs, gz_path)
            _log(f"\n{'='*55}")
            _log(f"TP: {tp_key}")

            tp_output_dir = run_dir / tp_key
            _misc_dir = base_dir / "output" / "misc"
            _misc_dir.mkdir(parents=True, exist_ok=True)
            dashboard     = str(_misc_dir / f"Dashboard_{tp_key}.html")  # top-level latest pointer

            # extract revision letter (last char of group key) for R0 merge logic
            _m_let    = re.search(r'[A-Za-z]\d+([A-Za-z])', tp_key)
            _tp_letter = _m_let.group(1).upper() if _m_let else ''
            _r0_gz    = base_dir / "data" / "NVL816-R0-Data.csv.gz"
            _use_r0   = _tp_letter in ('C', 'D') and _r0_gz.exists() and "119325" in tp_key
            _r0_label = f"H61{_tp_letter} + NVL816-R0" if _use_r0 else ""
            if _use_r0:
                _log(f"  R0 merge   : {_r0_gz.name}  (label: {_r0_label})")

            cfg = {
                "DataCSV":            [str(gz_path)],
                "output_folder":      str(run_dir),   # pipeline writes to run_dir/tp_key/
                "dashboard":          dashboard,
                "identifier":         tp_key,         # subfolder name inside run_dir
                "TestProgram_folder": _tp_folder,
                "run_parametric":     True,
                "keep_pcm_idw":       False,
            }
            if prod_cfg:
                cfg["product_config_json"] = prod_cfg

            json_path = run_dir / f"input_{tp_key}.json"
            if not args.dry_run:
                run_dir.mkdir(parents=True, exist_ok=True)
                json_path.write_text(json.dumps(cfg, indent=2), encoding="utf-8")
            _log(f"  Config → {json_path}")
            _log(f"  Output → {tp_output_dir}")

            if args.dry_run:
                _log(f"  DRY-RUN: would run pipeline.py --json {json_path}")
                ok = True
            else:
                if not gz_path.exists():
                    _log(f"  WARNING: gz not found: {gz_path} — skipping")
                    continue
                cmd = [sys.executable, str(_PIPELINE), "--json", str(json_path)]
                _log("  Running pipeline…")
                try:
                    result = subprocess.run(cmd, capture_output=False, text=True,
                                            timeout=3600, env=env, cwd=str(_PIPELINE.parent))
                    ok = result.returncode == 0
                    if not ok:
                        _log(f"  WARNING: pipeline rc={result.returncode}")
                except subprocess.TimeoutExpired:
                    _log("  ERROR: pipeline timed out")
                    ok = False

            # gz mtime → data freshness timestamp shown in report
            try:
                gz_ts = datetime.fromtimestamp(gz_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M") if gz_path.exists() else ""
            except Exception:
                gz_ts = ""

            # ── R0 second run (61C/61D only) ─────────────────────────────────
            tp_output_dir_r0: Path | None = None
            if _use_r0 and not args.dry_run and ok and gz_path.exists():
                tp_output_dir_r0 = run_dir / (tp_key + "_R0")
                cfg_r0 = {
                    "DataCSV":            [str(gz_path), str(_r0_gz)],
                    "output_folder":      str(run_dir),
                    "dashboard":          str(base_dir / "output" / "misc" / f"Dashboard_{tp_key}_R0.html"),
                    "identifier":         tp_key + "_R0",
                    "TestProgram_folder": _tp_folder,
                    "run_parametric":     True,
                    "keep_pcm_idw":       False,
                }
                if prod_cfg:
                    cfg_r0["product_config_json"] = prod_cfg
                json_r0 = run_dir / f"input_{tp_key}_R0.json"
                json_r0.write_text(json.dumps(cfg_r0, indent=2), encoding="utf-8")
                _log(f"  R0 run → {tp_output_dir_r0}")
                cmd_r0 = [sys.executable, str(_PIPELINE), "--json", str(json_r0)]
                try:
                    res_r0 = subprocess.run(cmd_r0, capture_output=False, text=True,
                                            timeout=3600, env=env, cwd=str(_PIPELINE.parent))
                    if res_r0.returncode != 0:
                        _log(f"  WARNING: R0 pipeline rc={res_r0.returncode}")
                        tp_output_dir_r0 = None
                    else:
                        _watermark_output_dir(str(tp_output_dir_r0))
                except subprocess.TimeoutExpired:
                    _log("  ERROR: R0 pipeline timed out")
                    tp_output_dir_r0 = None

            tp_outputs.append((tp_key, ok, tp_output_dir, gz_ts, _r0_label, tp_output_dir_r0))
            results.append((tp_key, ok, str(tp_output_dir), str(tp_output_dir_r0) if tp_output_dir_r0 else ""))

            if not args.dry_run and ok:
                _watermark_output_dir(str(tp_output_dir))
                _stamp_dashboard_block(Path(dashboard), tp_key, ts[:8])
            # Delete temp gz (raw_<ts>.7z is the archival copy; no persistent per-TP files)
            if not args.dry_run:
                try:
                    if gz_path.exists() and gz_path.name.startswith("tmp_"):
                        gz_path.unlink()
                except Exception:
                    pass

        # ── 4b. Add previous-run data for TPs not updated this cycle ─────────
        if not args.dry_run:
            run_keys_set  = set(_letter_keys)
            out_root      = base_dir / "output"
            prev_run_dirs = sorted(
                (d for d in out_root.iterdir() if d.is_dir() and d != run_dir),
                reverse=True,
            ) if out_root.exists() else []

            # Stale keys for this letter only (from stored gz + previous output dirs)
            hist_keys: set[str] = {
                k for k in all_stored_keys
                if re.search(rf'{re.escape(_prog_series)}{_letter}', k, re.IGNORECASE)
            }
            for _prev in prev_run_dirs:
                for _sub in _prev.iterdir():
                    if _sub.is_dir() and not _sub.name.endswith("_R0"):
                        _sub_m = re.search(rf'{re.escape(_prog_series)}([A-Za-z])', _sub.name)
                        if (_sub_m.group(1).upper() if _sub_m else 'X') == _letter:
                            hist_keys.add(_sub.name)

            stale_cands = sorted(hist_keys - run_keys_set)
            if stale_cands:
                for stale_key in stale_cands:
                    for prev_run_dir in prev_run_dirs:
                        prev_tp_dir = prev_run_dir / stale_key
                        if prev_tp_dir.is_dir():
                            m = re.search(r'_(\d{8})_', prev_run_dir.name)
                            if m:
                                d = m.group(1)
                                prev_label = f"prev: {d[:4]}-{d[4:6]}-{d[6:]}"
                            else:
                                prev_label = f"prev: {prev_run_dir.name}"
                            tp_outputs.append((stale_key, True, prev_tp_dir, prev_label, "", None))
                            _log(f"  Stale TP in report : {stale_key} ← {prev_run_dir.name}")
                            break

        # ── 4c. Generate per-TP compare_report (trend over daily runs) ───────
        if not args.dry_run and _COMPARE_RUNS.exists():
            _log(f"\nGenerating compare reports for {_prog_series}{_letter}…")
            for tp_key, ok, tp_output_dir, gz_ts, r0_label, tp_output_dir_r0 in tp_outputs:
                compare_out = base_dir / "output" / "compare" / f"compare_report_{tp_key}.html"
                compare_out.parent.mkdir(parents=True, exist_ok=True)
                if gz_ts.startswith("prev:") and compare_out.exists():
                    continue   # stale TP — compare report already generated; skip
                dash_html = base_dir / "output" / "misc" / f"Dashboard_{tp_key}.html"
                # Build a temporary enriched Dashboard for compare_runs that includes ALL
                # historical output dirs (including prior op numbers with same prefix,
                # e.g. 132222 runs when the current key is 132322).  Write to a sidecar
                # file so the real pipeline.py Dashboard.html is never overwritten.
                _cmp_dash = base_dir / "output" / "misc" / f"_cmp_dash_{tp_key}.html"
                rebuilt = _rebuild_dashboard_html_for_tp(tp_key, base_dir, out_path=_cmp_dash)
                if rebuilt and rebuilt.stat().st_size > 0:
                    _n_blocks = rebuilt.read_text(encoding='utf-8').count('class="run-block"')
                    _log(f"  Compare dashboard: {rebuilt.name} ({_n_blocks} run blocks)")
                    dash_html = rebuilt
                elif not dash_html.exists():
                    _log(f"  SKIP {tp_key}: Dashboard.html not found and could not rebuild (no historical runs?)")
                    continue
                cmd_cmp = [sys.executable, str(_COMPARE_RUNS),
                           str(dash_html), "--out", str(compare_out), "--no-open"]
                _log(f"  Compare → {compare_out.name}")
                try:
                    res_cmp = subprocess.run(cmd_cmp, capture_output=True, text=True,
                                             timeout=300, env=env,
                                             cwd=str(_COMPARE_RUNS.parent))
                    if res_cmp.returncode != 0:
                        _log(f"  WARNING: compare_runs rc={res_cmp.returncode}: {res_cmp.stderr[:200]}")
                    else:
                        _log(f"  OK  ({compare_out.stat().st_size:,} bytes)")
                except subprocess.TimeoutExpired:
                    _log(f"  WARNING: compare_runs timed out for {tp_key}")
                except Exception as _ce:
                    _log(f"  WARNING: compare_runs error: {_ce}")

        # ── Per-letter report.html ────────────────────────────────────────────
        letter_report_path: Path | None = None
        if not args.dry_run and results:
            letter_report_path = _build_run_report(
                run_dir,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                str(aqua_file),
                tp_outputs,
                prog_series=_prog_series,
                prod_cfg=prod_cfg,
                product_name=_product_name,
            )
            if letter_report_path:
                letter_report_paths.append(letter_report_path)

        # ── Per-letter run log entry ──────────────────────────────────────────
        _log(f"\nUpdating run log for {_prog_series}{_letter}…")
        update_run_log(
            results=results,
            aqua_file=str(aqua_file),
            run_log=run_log,
            dry_run=args.dry_run,
            report_path=letter_report_path,
        )

        all_results.extend(results)
        all_tp_outputs.extend(tp_outputs)

    # ── 4d. Group compare across fresh program letters ────────────────────────
    if not args.dry_run and len(_letter_groups) >= 2:
        _log('\nRunning cross-letter group compare…')
        try:
            import sys as _sys
            _sys.path.insert(0, str(_HERE))
            from yield_group_compare import run_group_compare_headless as _gc_headless
            _gc_pairs: list[tuple[str, Path]] = []
            for _tp_key, _ok, _tp_output_dir, _gz_ts, _r0_label, _tp_output_dir_r0 in all_tp_outputs:
                if not _ok or str(_gz_ts).startswith('prev:'):
                    continue
                _idx = Path(str(_tp_output_dir)) / 'index.html'
                _m_let = re.search(r'[A-Za-z](\d{2,})([A-Za-z])', _tp_key)
                # group name e.g. "80C" (digits + letter, no leading series prefix)
                _grp = (_m_let.group(1) + _m_let.group(2).upper()) if _m_let else _prog_series + '?'
                _gc_pairs.append((_grp, _idx))
            # sort ascending by letter so columns appear 80A, 80B, 80C …
            _gc_pairs.sort(key=lambda t: t[0])
            if len(_gc_pairs) >= 2:
                # place alongside the last letter's report.html
                _gc_out = letter_report_paths[-1].parent / 'group_compare.html' if letter_report_paths else base_dir / 'output' / 'group_compare.html'
                _gc_result = _gc_headless(_gc_pairs, _gc_out, log=_log)
                if _gc_result:
                    _log(f'  Group compare OK ({len(_gc_pairs)} groups) \u2192 {_gc_result}')
            else:
                _log(f'  Group compare skipped — only {len(_gc_pairs)} fresh TP(s)')
        except Exception as _gce:
            import traceback as _tb
            _log(f'  WARNING: group compare failed: {_gce}')
            _log(_tb.format_exc())

    # ── Flatten results from all letter groups for email ─────────────────────
    results    = all_results
    tp_outputs = all_tp_outputs
    report_path = letter_report_paths[-1] if letter_report_paths else None

    # ── 5. Load email_config.json (set by manage_email.py GUI) ───────────────
    email_cfg_path = _EMAIL_CFG
    email_cfg: dict = {}
    if email_cfg_path.exists():
        try:
            email_cfg = json.loads(email_cfg_path.read_text(encoding="utf-8"))
            # migrate old single-field format
            if "email_to" in email_cfg and "email_to_report" not in email_cfg:
                email_cfg["email_to_report"] = email_cfg.pop("email_to")
            _log(f"\nEmail config: {email_cfg_path}")
        except Exception as _e:
            _log(f"\nWARNING: could not read email_config.json: {_e}")

    excluded_keys: set[str] = set(email_cfg.get("excluded_keys", []))
    excluded_ops:  set[str] = set(str(o) for o in email_cfg.get("excluded_ops", []))
    email_to_report = email_cfg.get("email_to_report", args.email) or args.email
    email_to_alert  = email_cfg.get("email_to_alert",  email_to_report)
    if excluded_keys or excluded_ops:
        _log(f"  Excluded from report: keys={sorted(excluded_keys)} ops={sorted(excluded_ops)}")

    def _is_excluded(tp_key: str) -> bool:
        if tp_key in excluded_keys:
            return True
        for op in excluded_ops:
            if tp_key.endswith(f"_{op}") or f"_{op}_" in tp_key:
                return True
        return False

    # Filter tp_outputs to only keys allowed in the email report
    tp_outputs_email = [t for t in tp_outputs if not _is_excluded(t[0])]

    # ── 5b. Cleanup old run folders (before sending email) ───────────────────
    _keep_runs = args.keep_runs
    if _keep_runs is None:
        _keep_runs = int(email_cfg.get("keep_runs", 5))
    if _keep_runs > 0:
        _log(f"\nCleaning up old runs (keep last {_keep_runs} per letter) …")
        _n_deleted = cleanup_old_runs(
            base_dir / "output", _keep_runs, dry_run=args.dry_run, prog_series=_prog_series
        )
        if not args.dry_run:
            _log(f"  Deleted {_n_deleted} old run folder(s).")

    # ── 6. Send email ─────────────────────────────────────────────────────────
    run_ts  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    all_ok  = all(r[1] for r in results)
    subject = f"{_product_name} Yield Dashboard"

    tmp_att_dir = Path(tempfile.mkdtemp(prefix="nvl_att_"))
    email_attachments: list[str] = []
    try:
        # Build ONE combined report from all allowed tp_outputs.
        # This single file is used as BOTH the email body and the attachment
        # so they are guaranteed to be identical.
        _ts_label   = datetime.now().strftime("%Y%m%d_%H%M%S")
        _att_name   = f"{_product_name} Yield Report {_ts_label}.html"
        _comb_path  = tmp_att_dir / _att_name

        _email_body_rpt: str | None = None
        _email_body_collapsed: str | None = None
        try:
            _excl_keys = list(email_cfg.get("excluded_keys", []))
            _email_body_rpt = _build_email_report_html(
                base_dir / "output", run_ts,
                excluded_keys=_excl_keys,
                prog_series=_prog_series,
                product_name=_product_name,
            )
            _comb_path.write_text(_email_body_rpt, encoding="utf-8")
            email_attachments.append(str(_comb_path))
            # Also save a persistent copy to reports/
            _reports_dir = base_dir / "reports"
            _reports_dir.mkdir(parents=True, exist_ok=True)
            _report_save = _reports_dir / f"Yield_Report_{_ts_label}.html"
            _report_save.write_text(_email_body_rpt, encoding="utf-8")
            _log(f"Report saved: {_report_save}")
            # Collapse <details>/<summary> and strip iframes for Outlook compatibility
            _email_body_collapsed = _collapse_report_html(_email_body_rpt)
        except Exception as _be:
            _log(f"  WARNING: combined email report build failed: {_be}")

        if not email_attachments and letter_report_paths:
            # Fallback: attach the per-letter reports if combined build failed
            for _lp in letter_report_paths:
                if _lp and _lp.exists():
                    email_attachments.append(str(_lp))

        _fallback_body = _email_body_html(
            run_ts, str(aqua_file),
            [r for r in results if not _is_excluded(r[0])]
            + [
                (tp_key, True, str(tp_dir), "", gz_ts)
                for tp_key, ok, tp_dir, gz_ts, _r0lbl, _r0dir in tp_outputs_email
                if gz_ts.startswith("prev:") and tp_key not in {r[0] for r in results}
            ],
            run_log,
            attachments=email_attachments,
        )

        send_email(
            to=email_to_report,
            subject=f"{_product_name} Yield Report",
            body_html=_email_body_collapsed or _email_body_rpt or _fallback_body,
            dry_run=args.dry_run,
            attachments=email_attachments,
        )
    finally:
        shutil.rmtree(tmp_att_dir, ignore_errors=True)

    _log("\n" + "=" * 65)
    for _lp in letter_report_paths:
        _log(f"Run folder : {_lp.parent}")
    for r in results:
        _log(f"  {'OK' if r[1] else 'FAILED'}  op={r[0]}  → {r[2]}")
    _log(f"Run log    : {run_log}")
    _log("=" * 65)

    # Per-TP gz files are kept in data/programs/0H61{letter}/ for selective deletion.

    # Compress the AQUA pull snapshot to .7z for better long-term storage
    if not args.dry_run and not args.local_csv and isinstance(aqua_file, Path):
        if aqua_file.suffix == ".gz" and aqua_file.stem.endswith(".csv"):
            _log(f"\nCompressing {aqua_file.name} → .7z …")
            z7 = _compress_aqua_to_7z(aqua_file)
            if z7:
                _log(f"  {aqua_file.name} → {z7.name}  ({z7.stat().st_size / 1024:.0f} KB)")
            else:
                _log(f"  WARNING: 7z compression failed; keeping {aqua_file.name}")

    if not all_ok:
        sys.exit(1)

    # ── Hint: start resend server if not dry-run ──────────────────────────
    if not args.dry_run:
        _log(f"\nTip: run  python run_automation.py --serve  to enable the")
        _log(f"     'Resend Email' button in BinDistribution.html.")


# ─────────────────────────────────────────────────────────────────────────────
# Resend server  (python run_automation.py --serve [--port N] [--base-dir D])
# ─────────────────────────────────────────────────────────────────────────────

def _run_resend_server(base_dir: Path, port: int = 17450, email_to: str = "", prog_series: str = "0H61", product_name: str = "NVL816-BLLC") -> None:
    """Start a local HTTP server on localhost:port that handles POST /resend.
    Blocks until Ctrl+C.  Looks at base_dir/output/ to find the latest run.
    """
    import http.server
    import threading

    def _find_latest_tp_dirs(base_dir: Path) -> list[Path]:
        """Return tp_output dirs from the most-recent run for each program group."""
        out_root = base_dir / "output"
        if not out_root.exists():
            return []
        _gd3 = re.search(r'(\d+)$', prog_series)
        _gen3 = _gd3.group(1) if _gd3 else '61'
        run_dirs = sorted(
            (d for d in out_root.iterdir()
             if d.is_dir() and re.match(rf'NVL_[A-Za-z]{_gen3}[A-Z]_', d.name)),
            key=lambda d: d.name, reverse=True,
        )
        # One latest run per group key
        letter_run: dict[str, Path] = {}
        for rd in run_dirs:
            m = re.search(rf'NVL_([A-Za-z]{_gen3}[A-Z])_', rd.name)
            if m and m.group(1) not in letter_run:
                letter_run[m.group(1)] = rd
        tp_dirs: list[Path] = []
        for letter in sorted(letter_run):
            rd = letter_run[letter]
            for sub in sorted(rd.iterdir()):
                if sub.is_dir() and not sub.name.endswith("_R0"):
                    tp_dirs.append(sub)
        return tp_dirs

    def _handle_resend(body: dict) -> dict:
        lw_raw = body.get("lots_wafers", "all")
        row_filter: "set | None" = None
        if lw_raw != "all" and isinstance(lw_raw, list):
            row_filter = set(str(x) for x in lw_raw)
        tp_dirs = _find_latest_tp_dirs(base_dir)
        if not tp_dirs:
            return {"status": "error", "message": "No run output found in " + str(base_dir)}

        # Re-read summaries with filter
        summaries: list[tuple[str, dict]] = []
        for tp_dir in tp_dirs:
            smry = _extract_yield_summary(tp_dir, row_filter=row_filter)
            if smry:
                summaries.append((tp_dir.name, smry))

        if not summaries:
            return {"status": "error", "message": "No yield data found (filter too narrow?)"}

        # Build a simple email table
        def _parse_pct(s: str) -> float:
            try:
                return float(str(s).replace("%", "").strip())
            except Exception:
                return 0.0

        filter_desc = ""
        if row_filter:
            lots  = sorted(set(lw.split("|")[0] for lw in row_filter if "|" in lw))
            wafs  = sorted(set(lw.split("|")[1] for lw in row_filter if "|" in lw))
            filter_desc = f"Lots: {', '.join(lots)} &nbsp;|&nbsp; Wafers: {', '.join(wafs)}"
        else:
            filter_desc = "All wafers"

        rows_html = ""
        for tp_key, smry in sorted(summaries, reverse=True):
            bins = smry.get("bins", {})
            b1 = _parse_pct(bins.get("Bin 1", "0"))
            b2 = _parse_pct(bins.get("Bin 2", "0"))
            b3 = _parse_pct(bins.get("Bin 3", "0"))
            b4 = _parse_pct(bins.get("Bin 4", "0"))
            ff   = b1 + b2
            ffdf = b1 + b2 + b3 + b4
            rb   = smry.get("repair_bins", {})
            dlcp = smry.get("dlcp", {})
            rv_hp = f"{dlcp['hp']} ({dlcp['hp_n']:,})" if dlcp else "–"
            rv_lp = f"{dlcp['lp']} ({dlcp['lp_n']:,})" if dlcp else "–"
            rv198 = rb.get("198", "–")
            rv201 = rb.get("201") or rb.get("2") or "–"
            rv202 = rb.get("202", "–")
            rows_html += (
                f"<tr><td style='font-family:monospace;color:#0071c5'>{tp_key}</td>"
                f"<td style='text-align:center'>{smry.get('die','–')}</td>"
                f"<td style='text-align:center'>{ff:.1f}%</td>"
                f"<td style='text-align:center'>{ffdf:.1f}%</td>"
                f"<td style='text-align:center;color:#1565c0'>{rv_hp}</td>"
                f"<td style='text-align:center;color:#e65100'>{rv_lp}</td>"
                f"<td style='text-align:center'>{rv198}</td>"
                f"<td style='text-align:center'>{rv201}</td>"
                f"<td style='text-align:center'>{rv202}</td>"
                f"</tr>\n"
            )

        run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        body_html = f"""<!DOCTYPE html><html><body style="font-family:Segoe UI,Arial;color:#222;max-width:860px">
<h2 style="color:#0071c5;margin-bottom:4px">Yield Dashboard — Filtered Resend</h2>
<p style="color:#555;font-size:0.9em;margin-top:0">{run_ts}</p>
<p style="background:#fff8e1;padding:8px 12px;border-left:4px solid #f9a825;font-size:0.9em">
  <b>Filter applied:</b>&nbsp;{filter_desc}</p>
<table border="1" cellpadding="6" cellspacing="0"
       style="border-collapse:collapse;width:100%;font-size:0.9em">
  <tr style="background:#0071c5;color:#fff">
    <th>TP Key</th><th>Die</th><th>FF (1+2)</th><th>FF+DF</th>
    <th style="color:#90caf9">DLCP HP</th><th style="color:#ffcc80">DLCP LP</th>
    <th>FB198</th><th>FB201</th><th>FB202</th>
  </tr>
  {rows_html}
</table>
<hr/><p style="font-size:0.8em;color:#888">Sent via resend server — base dir: {base_dir}</p>
</body></html>"""

        ecfg: dict = {}
        if _EMAIL_CFG.exists():
            try:
                ecfg = json.loads(_EMAIL_CFG.read_text(encoding="utf-8"))
            except Exception:
                pass
        to = email_to or ecfg.get("email_to_report") or ecfg.get("email_to") or _EMAIL_TO
        subject = f"{product_name} Yield Dashboard (filtered resend)"

        class _FakeArgs:
            dry_run = False

        try:
            send_email(to=to, subject=subject, body_html=body_html, dry_run=False)
            n_tp = len(summaries)
            return {"status": "ok",
                    "message": f"Email sent to {to}  ({n_tp} TP(s), {filter_desc})"}
        except Exception as exc:
            return {"status": "error", "message": f"Email failed: {exc}"}

    # ── HTTP server ──────────────────────────────────────────────────────────
    class _Handler(http.server.BaseHTTPRequestHandler):
        def log_message(self, *_):
            pass  # suppress default access log

        def _cors(self):
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
            self.send_header("Access-Control-Allow-Headers", "Content-Type")

        def do_OPTIONS(self):
            self.send_response(200)
            self._cors()
            self.end_headers()

        def do_GET(self):
            if self.path == "/status":
                resp = json.dumps({"status": "ok", "base_dir": str(base_dir)}).encode()
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self._cors()
                self.send_header("Content-Length", len(resp))
                self.end_headers()
                self.wfile.write(resp)
            else:
                self.send_response(404)
                self.end_headers()

        def do_POST(self):
            if self.path == "/resend":
                try:
                    length = int(self.headers.get("Content-Length", 0))
                    body = json.loads(self.rfile.read(length))
                    result = _handle_resend(body)
                except Exception as exc:
                    result = {"status": "error", "message": str(exc)}
                resp = json.dumps(result).encode()
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self._cors()
                self.send_header("Content-Length", len(resp))
                self.end_headers()
                self.wfile.write(resp)
            else:
                self.send_response(404)
                self.end_headers()

    server = http.server.HTTPServer(("127.0.0.1", port), _Handler)
    print(f"\nResend server listening on http://localhost:{port}/resend")
    print(f"Base dir : {base_dir}")
    print("Press Ctrl+C to stop.\n")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped.")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()


# ════════════════════════════════════════════════════════════════
# (formerly serve_reports.py)
# ════════════════════════════════════════════════════════════════

import argparse
from datetime import datetime
import http.server
import json
import logging
import os
import re
import threading
from pathlib import Path

log = logging.getLogger(__name__)

_HERE      = Path(__file__).resolve().parent.parent   # yield-dashboard/
_BASE_DIR  = Path(r"\\samba.zsc10.intel.com\nfs\zsc10\disks\gsc_gwa011\users\snpant\auto\yield")
_DEFAULT_PORT = 8765


def _fmt_size(n: int) -> str:
    if n < 1024:       return f"{n} B"
    if n < 1024 ** 2:  return f"{n / 1024:.0f} KB"
    return f"{n / 1024 ** 2:.1f} MB"


class ReportHandler(http.server.BaseHTTPRequestHandler):
    base_dir: Path = _BASE_DIR   # set by factory

    def _cors(self) -> None:
        self.send_header("Access-Control-Allow-Origin",  "*")
        self.send_header("Access-Control-Allow-Methods", "GET, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def do_OPTIONS(self) -> None:   # preflight
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_GET(self) -> None:
        path = self.path.split("?")[0].rstrip("/")

        if path == "/api/reports":
            self._serve_manifest()
        elif path.startswith("/reports/"):
            fname = Path(path[len("/reports/"):]).name
            self._serve_file(fname)
        elif path in ("", "/"):
            self._serve_file("index.html")
        else:
            self.send_response(404)
            self.end_headers()

    def _serve_manifest(self) -> None:
        reports_dir = self.base_dir / "reports"
        files = sorted(
            [f for f in reports_dir.glob("Yield_Report_*.html")
             if f.name != "index.html"],
            key=lambda f: f.stat().st_mtime,
            reverse=True,
        )
        data = [
            {
                "name":     f.name,
                "size":     _fmt_size(f.stat().st_size),
                "mtime":    datetime.fromtimestamp(
                                f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                "url":      f"/reports/{f.name}",
            }
            for f in files
        ]
        body = json.dumps(data, indent=2).encode()
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self._cors()
        self.end_headers()
        self.wfile.write(body)

    def _serve_file(self, fname: str) -> None:
        # Safety: strip any path traversal
        fname = Path(fname).name
        fpath = self.base_dir / "reports" / fname
        if not fpath.exists():
            self.send_response(404)
            self.end_headers()
            return
        body = fpath.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self._cors()
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, fmt, *args) -> None:  # suppress default stdout noise
        log.debug(fmt % args)


def _ensure_cert(cert_dir: Path) -> tuple[Path, Path]:
    """Return (certfile, keyfile), generating a self-signed cert if needed.
    Uses the built-in `ssl` + `cryptography` package if available,
    falls back to `openssl` CLI, or skips HTTPS if neither is available.
    """
    cert_dir.mkdir(parents=True, exist_ok=True)
    cert_file = cert_dir / "server.crt"
    key_file  = cert_dir / "server.key"
    if cert_file.exists() and key_file.exists():
        return cert_file, key_file

    # Try cryptography package first (pure Python, no external tools)
    try:
        import datetime as _dt
        from cryptography import x509
        from cryptography.x509.oid import NameOID
        from cryptography.hazmat.primitives import hashes, serialization
        from cryptography.hazmat.primitives.asymmetric import rsa
        import ipaddress

        key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
        hostname = __import__("socket").gethostname()
        subject = issuer = x509.Name([
            x509.NameAttribute(NameOID.COMMON_NAME, hostname),
        ])
        cert = (
            x509.CertificateBuilder()
            .subject_name(subject)
            .issuer_name(issuer)
            .public_key(key.public_key())
            .serial_number(x509.random_serial_number())
            .not_valid_before(_dt.datetime.utcnow())
            .not_valid_after(_dt.datetime.utcnow() + _dt.timedelta(days=3650))
            .add_extension(x509.SubjectAlternativeName([
                x509.DNSName(hostname),
                x509.DNSName("localhost"),
                x509.IPAddress(ipaddress.IPv4Address("127.0.0.1")),
            ]), critical=False)
            .sign(key, hashes.SHA256())
        )
        key_file.write_bytes(
            key.private_bytes(serialization.Encoding.PEM,
                              serialization.PrivateFormat.TraditionalOpenSSL,
                              serialization.NoEncryption()))
        cert_file.write_bytes(cert.public_bytes(serialization.Encoding.PEM))
        log.info(f"Self-signed cert generated → {cert_file}")
        return cert_file, key_file
    except ImportError:
        pass

    # Fallback: openssl CLI
    import subprocess as _sp
    hostname = __import__("socket").gethostname()
    r = _sp.run([
        "openssl", "req", "-x509", "-newkey", "rsa:2048",
        "-keyout", str(key_file), "-out", str(cert_file),
        "-days", "3650", "-nodes",
        "-subj", f"/CN={hostname}",
    ], capture_output=True)
    if r.returncode == 0:
        log.info(f"Self-signed cert generated via openssl → {cert_file}")
        return cert_file, key_file

    raise RuntimeError(
        "Cannot generate TLS cert. Install 'cryptography':\n"
        "  pip install cryptography"
    )


def make_server(base_dir: Path, port: int,
                https: bool = True) -> http.server.HTTPServer:
    handler = type("Handler", (ReportHandler,), {"base_dir": base_dir})
    server  = http.server.HTTPServer(("", port), handler)
    if https:
        try:
            import ssl
            cert_dir  = Path(__file__).parent / ".tls"
            cert_file, key_file = _ensure_cert(cert_dir)
            ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
            ctx.load_cert_chain(str(cert_file), str(key_file))
            server.socket = ctx.wrap_socket(server.socket, server_side=True)
            server._is_https = True
        except Exception as e:
            log.warning(f"HTTPS setup failed ({e}), falling back to HTTP.")
            server._is_https = False
    else:
        server._is_https = False
    return server


def run(base_dir: Path = _BASE_DIR, port: int = _DEFAULT_PORT,
        daemon: bool = False) -> http.server.HTTPServer:
    """Start the server (optionally in a daemon thread) and return it."""
    server = make_server(base_dir, port)
    proto  = "https" if getattr(server, "_is_https", False) else "http"
    t = threading.Thread(target=server.serve_forever, daemon=daemon)
    t.start()
    log.info(f"Reports server listening on {proto}://localhost:{port}")
    return server


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    ap = argparse.ArgumentParser()
    ap.add_argument("--port",       type=int, default=_DEFAULT_PORT)
    ap.add_argument("--base-dir",   default=None)
    ap.add_argument("--no-https",   action="store_true",
                    help="Disable HTTPS (use plain HTTP)")
    args = ap.parse_args()

    base  = Path(args.base_dir) if args.base_dir else _BASE_DIR
    https = not args.no_https
    srv   = make_server(base, args.port, https=https)
    proto = "https" if getattr(srv, "_is_https", False) else "http"
    print(f"Starting reports server on {proto}://0.0.0.0:{args.port}")
    print(f"  Reports dir : {base / 'reports'}")
    print(f"  API         : {proto}://localhost:{args.port}/api/reports")
    if getattr(srv, "_is_https", False):
        print("  NOTE: self-signed cert — browser will show a security warning.")
        print("  Open the URL once and click 'Advanced -> Proceed' to trust it.")
    print("Press Ctrl+C to stop.")
    srv.serve_forever()

